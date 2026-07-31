#!/usr/bin/env python3

import datetime
import json
import pathlib
import shutil
import tempfile
import zipfile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path.cwd()
PACK = ROOT / "content-reference/apocalyptic-shadow-v1"
DATA = ROOT / "config/apocalyptic-shadow/data"
TEMPLATES = ROOT / "config/apocalyptic-shadow-generated/templates"
GROUPS = {
    "ApocalypticShadow.xlsx": [
        "profiles", "periods", "stages", "nodes", "participant-policies",
        "team-slots", "loadout-records", "attempts", "transitions", "clocks",
        "boss-progress", "scores", "objectives", "stars", "safeguards",
        "axioms", "embers", "buffs", "mechanic-contributions",
    ],
    "ApocalypticShadowBindings.xlsx": [
        "pool-audits", "encounters", "encounter-waves", "enemy-slots", "enemies",
        "enemy-skills", "enemy-statuses", "ability-bindings",
    ],
    "ApocalypticShadowReview.xlsx": [
        "mechanic-rules", "sources", "reconciliation", "research-gaps", "coverage",
        "review-fixtures", "manifest", "pack-index",
    ],
}
HEADERS = [
    "id", "stable_key", "row_order", "name_en", "name_zh_cn", "summary_en",
    "summary_zh_cn", "ownership", "coverage_state", "evidence_quality",
    "mechanism_quality", "manifest_record_ids", "source_ref_ids", "payload_json",
    "runtime_executable",
]


def pascal(value: str) -> str:
    return "".join(part[:1].upper() + part[1:] for part in value.split("-"))


def normalize_xlsx(path: pathlib.Path) -> None:
    with tempfile.TemporaryDirectory(prefix="starclock-g18-xlsx-") as tmp:
        temporary = pathlib.Path(tmp) / path.name
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
            temporary, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as target:
            for name in sorted(source.namelist()):
                info = zipfile.ZipInfo(name, date_time=(2000, 1, 1, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                info.external_attr = source.getinfo(name).external_attr
                target.writestr(info, source.read(name))
        shutil.copyfile(temporary, path)


def create_workbook(filename: str, files: list[str]) -> None:
    workbook = load_workbook(TEMPLATES / filename)
    workbook.properties.created = datetime.datetime(2000, 1, 1)
    workbook.properties.modified = datetime.datetime(2000, 1, 1)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for file_order, file in enumerate(files, start=1):
        document = json.loads((PACK / f"{file}.json").read_text())
        sheet = workbook[pascal(file)]
        sheet.freeze_panes = "B8"
        sheet.auto_filter.ref = f"B2:{get_column_letter(len(HEADERS) + 1)}2"
        for cell in sheet[2][1:]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row_order, record in enumerate(document["records"], start=1):
            payload = dict(record)
            source_ids = [ref["source_id"] for ref in record.get("source_refs", [])]
            values = [
                file_order * 1_000_000 + row_order,
                record["id"], row_order, record["name_en"], record["name_zh_cn"],
                record["summary_en"], record["summary_zh_cn"], record["ownership"],
                record["coverage_state"], record["evidence_quality"],
                record["mechanism_quality"], "|".join(record["manifest_record_ids"]),
                "|".join(source_ids), json.dumps(payload, ensure_ascii=False,
                                              separators=(",", ":"), sort_keys=True),
                record["runtime_executable"],
            ]
            sheet.append([None, *values])
        widths = [12, 42, 12, 28, 24, 48, 48, 20, 16, 24, 24, 48, 48, 80, 18]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index + 1)].width = width
        sheet.column_dimensions["A"].width = 3
        for row in sheet.iter_rows(min_row=8):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    DATA.mkdir(parents=True, exist_ok=True)
    output = DATA / filename
    workbook.save(output)
    normalize_xlsx(output)
    loaded = load_workbook(output, read_only=True, data_only=False)
    assert loaded.sheetnames == [pascal(file) for file in files]
    loaded.close()


for workbook_name, file_names in GROUPS.items():
    create_workbook(workbook_name, file_names)
print("Apocalyptic Shadow workbooks: 3 files, 35 sheets.")
