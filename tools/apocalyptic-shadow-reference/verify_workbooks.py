#!/usr/bin/env python3

import json
import pathlib
import sys

from openpyxl import load_workbook

ROOT = pathlib.Path.cwd()
sys.path.insert(0, str(ROOT / "tools/apocalyptic-shadow-reference"))
from author_workbooks import GROUPS, HEADERS, pascal  # noqa: E402

data_root = ROOT / "config/apocalyptic-shadow/data"
sheet_count = 0
row_count = 0
for workbook_name, files in GROUPS.items():
    workbook = load_workbook(data_root / workbook_name, read_only=False,
                             data_only=False)
    expected_sheets = [pascal(file) for file in files]
    assert workbook.sheetnames == expected_sheets
    for file, sheet_name in zip(files, expected_sheets, strict=True):
        sheet = workbook[sheet_name]
        document = json.loads((ROOT / "content-reference/apocalyptic-shadow-v1" /
                               f"{file}.json").read_text())
        assert sheet["A1"].value == "@table"
        assert sheet["C1"].value == "@mode"
        assert sheet["D1"].value == "map"
        assert [cell.value for cell in sheet[3][1:16]] == HEADERS
        authored = max(sheet.max_row - 7, 0)
        assert authored == len(document["records"]), (file, authored,
                                                       len(document["records"]))
        stable_keys = [sheet.cell(row=row, column=3).value
                       for row in range(8, sheet.max_row + 1)]
        assert stable_keys == [record["id"] for record in document["records"]]
        assert len(stable_keys) == len(set(stable_keys))
        for row in sheet.iter_rows(min_row=8, min_col=2, max_col=16):
            assert not any(isinstance(cell.value, str) and
                           cell.value.startswith("=") for cell in row)
            assert row[-1].value is False
        sheet_count += 1
        row_count += authored
    workbook.close()
assert sheet_count == 35
assert row_count == 1246
print(f"Apocalyptic Shadow workbooks verified: {sheet_count} sheets, "
      f"{row_count} rows, zero formulas/runtime rows.")
