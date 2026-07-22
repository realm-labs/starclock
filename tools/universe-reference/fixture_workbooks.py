"""Build disposable empty or representative Universe workbooks with openpyxl."""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
from pathlib import Path

from openpyxl import load_workbook


def write_row(sheet, values: dict[str, object]) -> None:
    fields = {cell.value: cell.column for cell in sheet[3] if cell.value and cell.value != "#field"}
    unknown = sorted(set(values) - set(fields))
    if unknown:
        raise ValueError(f"{sheet.title}: unknown fields {unknown}")
    row = sheet.max_row + 1
    for field, value in values.items():
        sheet.cell(row=row, column=fields[field], value=value)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def populate(root: Path, output: Path, mode: str) -> None:
    if output.exists():
        raise FileExistsError(f"refusing to overwrite {output}")
    shutil.copytree(root / "config" / "data", output)
    templates = root / "config" / "generated" / "templates"
    for name in ("Universe.xlsx", "UniverseBindings.xlsx", "UniverseEvidence.xlsx"):
        shutil.copy2(templates / name, output / name)
    if mode == "empty":
        return

    pack = json.loads((root / "content-reference" / "standard-universe-v1" / "pack-index.json").read_text(encoding="utf-8"))
    universe = load_workbook(output / "Universe.xlsx")
    write_row(universe["UniverseProfile"], {
        "id": 1, "stable_key": "universe.profile.standard-main-world", "game_version": "4.4",
        "snapshot_date": "2026-07-22", "content_manifest_sha256": sha256(root / "content-manifests" / "standard-universe-v1" / "content-manifest.json"),
        "pack_sha256": pack["pack_sha256"], "world_count": 9, "path_count": 9,
        "runtime_loading": "ForbiddenStagingOnly",
    })
    write_row(universe["UniverseWorld"], {
        "id": 1, "profile_id": 1, "stable_key": "universe.world.01", "world_number": 1,
        "name_en": "World 1", "name_zh_cn": "第一世界", "summary_en": "Representative Standard world schema row.",
        "summary_zh_cn": "标准宇宙世界结构的代表性记录。", "entry_rule_stable_key": "universe.rule.run-entry.standard",
        "terminal_rule_stable_key": "universe.rule.run-terminal.standard",
    })
    write_row(universe["UniverseDomain"], {
        "id": 1, "stable_key": "universe.domain.combat-primary", "source_type": 1,
        "kind": "CombatPrimary", "decision_policy": "BattleHandoff", "terminal": False,
        "name_en": "Domain — Combat", "name_zh_cn": "区域—战斗", "summary_en": "Representative battle-handoff domain.",
        "summary_zh_cn": "代表性的战斗交接区域。",
    })
    write_row(universe["UniversePath"], {
        "id": 1, "stable_key": "universe.path.preservation", "buff_type": "Preservation",
        "name_en": "Preservation", "name_zh_cn": "存护", "summary_en": "Representative selectable Path row.",
        "summary_zh_cn": "代表性的可选命途记录。", "unlock_policy_stable_key": "universe.policy.path.preservation",
    })
    universe.save(output / "Universe.xlsx")

    bindings = load_workbook(output / "UniverseBindings.xlsx")
    write_row(bindings["UniverseContentPool"], {
        "id": 1, "stable_key": "universe.pool.blessings.representative", "kind": "Blessing",
        "ordering": "StableId", "replacement": False,
    })
    write_row(bindings["UniverseActivityBinding"], {
        "id": 1, "stable_key": "universe.activity-binding.standard", "profile_id": 1,
        "activity_stable_key": "activity.standard-universe", "participant_digest_locked": True,
        "scoped_slots_supported": True, "fork_join_reserved": True,
        "battle_handoff_contract": "BattleSpecDecisionV1", "external_outcome_contract": "ExternalOutcomeCommandV1",
    })
    write_row(bindings["UniverseActivityDomainBinding"], {
        "activity_binding_id": 1, "sequence": 1, "domain_id": 1, "decision_kind": "BattleCommand",
    })
    bindings.save(output / "UniverseBindings.xlsx")

    evidence = load_workbook(output / "UniverseEvidence.xlsx")
    write_row(evidence["UniverseSourceRecord"], {
        "id": 1, "stable_key": "source.standard-su.fixture", "source_kind": "ProjectPolicy",
        "repository_or_url": "https://github.com/realm-labs/sora", "revision_or_access_date": "2026-07-22",
        "game_version": "4.4", "relative_path_or_page": "representative fixture", "row_locator": "1",
        "evidence_sha256": "0" * 64, "quality": "ProjectPolicy",
        "license_note": "Synthetic schema proof only.", "note": "Not production content.",
    })
    write_row(evidence["UniverseContentAudit"], {
        "id": 1, "content_stable_key": "universe.profile.standard-main-world", "source_file": "manifest.json",
        "enabled": True, "mode_owner": "Standard", "quality": "ProjectPolicy", "mechanism_quality": "ProjectPolicy",
        "coverage_state": "DataReady", "provenance_ids": "1", "source_ids": "fixture", "note": "Schema proof only.",
    })
    write_row(evidence["UniverseCoverage"], {
        "id": 1, "category": "representative", "source_file": "fixture", "required": 1,
        "accounted": 1, "data_ready": 1, "coverage_percent_decimal": "100",
    })
    write_row(evidence["UniverseReviewFixture"], {
        "id": 1, "stable_key": "universe.fixture.representative", "mechanic_family": "schema-proof",
        "input_stable_keys": "universe.profile.standard-main-world", "initial_state_json": "{}", "commands_json": "[]",
        "expected_facts_json": "[]", "quality_floor": "ProjectPolicy", "provenance_ids": "1",
    })
    write_row(evidence["UniversePackFile"], {
        "id": 1, "relative_path": "manifest.json", "bytes": 1, "rows": 1, "sha256": "0" * 64,
    })
    evidence.save(output / "UniverseEvidence.xlsx")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--mode", choices=("empty", "representative"), required=True)
    args = parser.parse_args()
    populate(args.root.resolve(), args.output.resolve(), args.mode)


if __name__ == "__main__":
    main()
