"""Deterministic openpyxl writer and semantic workbook QA."""

from __future__ import annotations

import hashlib
import json
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

WORKBOOKS = ("Universe.xlsx", "UniverseBindings.xlsx", "UniverseEvidence.xlsx")
FIXED_TIME = datetime(2000, 1, 1, tzinfo=timezone.utc)
HEADER_FILL = PatternFill("solid", fgColor="17365D")
EVEN_FILL = PatternFill("solid", fgColor="EAF2F8")
ODD_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="B7C9E2"))


def schema_tables(root: Path) -> dict[str, dict]:
    lock = json.loads((root / "config" / "generated" / "schema.lock").read_text(encoding="utf-8"))
    return {
        table["name"]: table
        for table in lock["schema"]["tables"]
        if table["name"].startswith("Universe")
    }


def field_columns(sheet) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }


def write_rows(sheet, rows: Iterable[dict[str, object]]) -> int:
    columns = field_columns(sheet)
    row_list = list(rows)
    for offset, values in enumerate(row_list, start=8):
        unknown = sorted(set(values) - set(columns))
        if unknown:
            raise ValueError(f"{sheet.title}: unknown fields {unknown}")
        for field, value in values.items():
            sheet.cell(row=offset, column=columns[field], value=value)
    return len(row_list)


def style_sheet(sheet, row_count: int) -> None:
    maximum_row = max(7 + row_count, 7)
    maximum_column = sheet.max_column
    for cell in sheet[3]:
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=8, max_row=maximum_row, max_col=maximum_column):
        fill = EVEN_FILL if row[0].row % 2 == 0 else ODD_FILL
        for cell in row:
            cell.fill = copy(fill)
            cell.border = copy(THIN_BORDER)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A8"
    sheet.auto_filter.ref = f"A3:{sheet.cell(row=3, column=maximum_column).column_letter}{maximum_row}"
    sheet.sheet_view.showGridLines = False
    for column in range(1, maximum_column + 1):
        values = [str(sheet.cell(row=row, column=column).value or "") for row in range(2, min(maximum_row, 200) + 1)]
        width = min(60, max(10, max((len(value) for value in values), default=10) + 2))
        sheet.column_dimensions[sheet.cell(row=3, column=column).column_letter].width = width


def prepare_workbook(template: Path, target: Path, tables: dict[str, dict], rows: dict[str, list[dict]]) -> dict[str, int]:
    workbook = load_workbook(template)
    expected_sheets = [name for name, table in tables.items() if table["source"]["file"] == template.name]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"{template.name}: sheet order differs from schema")
    counts: dict[str, int] = {}
    for sheet_name in workbook.sheetnames:
        count = write_rows(workbook[sheet_name], rows.get(sheet_name, []))
        style_sheet(workbook[sheet_name], count)
        counts[sheet_name] = count
    workbook.properties.creator = "Starclock Goal 03 openpyxl bootstrap"
    workbook.properties.lastModifiedBy = "Starclock Goal 03 openpyxl bootstrap"
    workbook.properties.created = FIXED_TIME
    workbook.properties.modified = FIXED_TIME
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.save(target)
    return counts


def author(root: Path, output: Path, rows: dict[str, list[dict]]) -> dict[str, int]:
    output.mkdir(parents=True, exist_ok=True)
    targets = [output / name for name in WORKBOOKS]
    existing = [target for target in targets if target.exists()]
    if existing:
        raise FileExistsError(f"refusing to overwrite authored workbook(s): {', '.join(map(str, existing))}")
    tables = schema_tables(root)
    counts: dict[str, int] = {}
    for target in targets:
        counts.update(prepare_workbook(root / "config" / "generated" / "templates" / target.name, target, tables, rows))
    verify(root, output, counts)
    return counts


def verify(root: Path, directory: Path, expected_counts: dict[str, int] | None = None) -> dict[str, int]:
    tables = schema_tables(root)
    counts: dict[str, int] = {}
    for name in WORKBOOKS:
        workbook = load_workbook(directory / name, data_only=False)
        expected_sheets = [table_name for table_name, table in tables.items() if table["source"]["file"] == name]
        if workbook.sheetnames != expected_sheets:
            raise ValueError(f"{name}: missing or reordered sheet")
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if sheet["A1"].value != "@table" or sheet["B1"].value != sheet_name or sheet["A3"].value != "#field":
                raise ValueError(f"{name}/{sheet_name}: Sora metadata drifted")
            count = max(0, sheet.max_row - 7)
            counts[sheet_name] = count
            if sheet.freeze_panes != "A8" or not sheet.auto_filter.ref:
                raise ValueError(f"{name}/{sheet_name}: authoring affordances missing")
            for row in sheet.iter_rows(min_row=8):
                for cell in row:
                    if cell.data_type in (TYPE_FORMULA, TYPE_ERROR):
                        raise ValueError(f"{name}/{sheet_name}/{cell.coordinate}: formula or Excel error forbidden")
    if expected_counts is not None and counts != expected_counts:
        raise ValueError("workbook row counts changed after save/reload")
    return counts


def semantic_digest(directory: Path) -> str:
    payload: list[object] = []
    for name in WORKBOOKS:
        workbook = load_workbook(directory / name, data_only=False)
        for sheet in workbook.worksheets:
            cells = [[cell.value for cell in row] for row in sheet.iter_rows()]
            widths = {key: value.width for key, value in sheet.column_dimensions.items() if value.width is not None}
            payload.append([name, sheet.title, cells, sheet.freeze_panes, sheet.auto_filter.ref, widths])
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()
