#!/usr/bin/env python3
"""Create disposable first-row workbook copies for Goal 13 visual review."""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_ROOT))

from workbook.common import WORKBOOKS  # noqa: E402


def prepare(source_root: Path, output: Path) -> None:
    if output.exists():
        raise FileExistsError(
            f"refusing to overwrite visual-review directory {output}"
        )
    output.mkdir(parents=True)
    for name in WORKBOOKS:
        target = output / name
        shutil.copy2(source_root / name, target)
        workbook = load_workbook(target)
        for sheet in workbook.worksheets:
            last_column = sheet.cell(
                row=3, column=sheet.max_column
            ).column_letter
            last_row = min(sheet.max_row, 12)
            sheet.print_area = f"A1:{last_column}{last_row}"
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_setup.orientation = "landscape"
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 1
            sheet.sheet_view.zoomScale = 80
            for row in range(8, last_row + 1):
                sheet.row_dimensions[row].height = 42
        workbook.save(target)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    prepare(args.source.resolve(), args.output.resolve())


if __name__ == "__main__":
    main()
