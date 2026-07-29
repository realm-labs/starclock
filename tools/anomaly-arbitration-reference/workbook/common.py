"""Deterministic openpyxl authoring and semantic QA for Goal 13."""

from __future__ import annotations

import hashlib
import json
import re
import zipfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

WORKBOOKS = (
    "AnomalyArbitration.xlsx",
    "AnomalyArbitrationBindings.xlsx",
    "AnomalyArbitrationReview.xlsx",
)
FIXED_TIME = datetime(2000, 1, 1, tzinfo=timezone.utc)
HEADER_FILL = PatternFill("solid", fgColor="17365D")
EVEN_FILL = PatternFill("solid", fgColor="EAF2F8")
ODD_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="B7C9E2"))


def schema_tables(root: Path) -> list[dict]:
    lock_path = root / "config" / "anomaly-arbitration-generated" / "schema.lock"
    lock = json.loads(lock_path.read_text(encoding="utf-8"))
    return [table for table in lock["schema"]["tables"] if table["name"].startswith("Arb")]


def workbook_tables(root: Path, workbook_name: str) -> list[dict]:
    return [
        table
        for table in schema_tables(root)
        if table["source"]["file"] == workbook_name
    ]


def field_columns(sheet) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }


def write_rows(sheet, rows: Iterable[dict[str, object]]) -> int:
    columns = field_columns(sheet)
    row_list = list(rows)
    for offset, values in enumerate(row_list, start=8):
        unknown = sorted(set(values) - set(columns))
        missing = sorted(set(columns) - set(values))
        if unknown or missing:
            raise ValueError(
                f"{sheet.title}: field mismatch; unknown={unknown}, missing={missing}"
            )
        for field, value in values.items():
            sheet.cell(row=offset, column=columns[field], value=value)
    return len(row_list)


def style_sheet(sheet, row_count: int) -> None:
    maximum_row = max(7 + row_count, 7)
    maximum_column = sheet.max_column
    for cell in sheet[3]:
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    for row in sheet.iter_rows(
        min_row=8, max_row=maximum_row, max_col=maximum_column
    ):
        fill = EVEN_FILL if row[0].row % 2 == 0 else ODD_FILL
        for cell in row:
            cell.fill = copy(fill)
            cell.border = copy(THIN_BORDER)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A8"
    last_column = sheet.cell(row=3, column=maximum_column).column_letter
    sheet.auto_filter.ref = f"A3:{last_column}{maximum_row}"
    sheet.sheet_view.showGridLines = False
    for column in range(1, maximum_column + 1):
        values = [
            str(sheet.cell(row=row, column=column).value or "")
            for row in range(2, min(maximum_row, 200) + 1)
        ]
        width = min(
            60,
            max(10, max((len(value) for value in values), default=10) + 2),
        )
        letter = sheet.cell(row=3, column=column).column_letter
        sheet.column_dimensions[letter].width = width


def prepare_workbook(
    root: Path,
    template: Path,
    target: Path,
    rows: dict[str, list[dict[str, object]]],
) -> dict[str, int]:
    workbook = load_workbook(template)
    tables = workbook_tables(root, template.name)
    expected_sheets = [table["source"]["sheet"] for table in tables]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"{template.name}: sheet order differs from schema")
    counts: dict[str, int] = {}
    for table in tables:
        sheet_name = table["source"]["sheet"]
        sheet = workbook[sheet_name]
        if sheet["A1"].value != "@table" or sheet["B1"].value != table["name"]:
            raise ValueError(f"{template.name}/{sheet_name}: Sora metadata drifted")
        count = write_rows(sheet, rows[sheet_name])
        style_sheet(sheet, count)
        counts[sheet_name] = count
    workbook.properties.creator = "Starclock Goal 13 openpyxl author"
    workbook.properties.lastModifiedBy = "Starclock Goal 13 openpyxl author"
    workbook.properties.created = FIXED_TIME
    workbook.properties.modified = FIXED_TIME
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.save(target)
    normalize_archive(target)
    return counts


def normalize_archive(path: Path) -> None:
    """Canonicalize workbook ZIP metadata without altering workbook semantics."""
    temporary = path.with_suffix(f"{path.suffix}.canonical")
    with zipfile.ZipFile(path, "r") as source:
        members = [(name, source.read(name)) for name in sorted(source.namelist())]
    with zipfile.ZipFile(
        temporary, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as target:
        for name, payload in members:
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
                    rb"\g<1>2000-01-01T00:00:00Z\g<2>",
                    payload,
                )
            info = zipfile.ZipInfo(name, date_time=(2000, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0
            target.writestr(
                info,
                payload,
                compress_type=zipfile.ZIP_DEFLATED,
                compresslevel=9,
            )
    temporary.replace(path)


def author(
    root: Path,
    output: Path,
    rows: dict[str, list[dict[str, object]]],
) -> dict[str, int]:
    output.mkdir(parents=True, exist_ok=True)
    targets = [output / name for name in WORKBOOKS]
    existing = [target for target in targets if target.exists()]
    if existing:
        names = ", ".join(str(target) for target in existing)
        raise FileExistsError(f"refusing to overwrite authored workbook(s): {names}")
    templates = root / "config" / "anomaly-arbitration-generated" / "templates"
    counts: dict[str, int] = {}
    for target in targets:
        counts.update(
            prepare_workbook(root, templates / target.name, target, rows)
        )
    verify(root, output, counts)
    return counts


def verify(
    root: Path,
    directory: Path,
    expected_counts: dict[str, int] | None = None,
) -> dict[str, int]:
    counts: dict[str, int] = {}
    for name in WORKBOOKS:
        workbook = load_workbook(directory / name, data_only=False)
        tables = workbook_tables(root, name)
        expected_sheets = [table["source"]["sheet"] for table in tables]
        if workbook.sheetnames != expected_sheets:
            raise ValueError(f"{name}: missing or reordered sheet")
        for table in tables:
            sheet_name = table["source"]["sheet"]
            sheet = workbook[sheet_name]
            if (
                sheet["A1"].value != "@table"
                or sheet["B1"].value != table["name"]
                or sheet["A3"].value != "#field"
            ):
                raise ValueError(f"{name}/{sheet_name}: Sora metadata drifted")
            count = max(0, sheet.max_row - 7)
            counts[sheet_name] = count
            if sheet.freeze_panes != "A8" or not sheet.auto_filter.ref:
                raise ValueError(
                    f"{name}/{sheet_name}: authoring affordances missing"
                )
            for cell in sheet[3]:
                if cell.value is not None and not cell.alignment.wrap_text:
                    raise ValueError(
                        f"{name}/{sheet_name}/{cell.coordinate}: header wrapping missing"
                    )
            for column in range(1, sheet.max_column + 1):
                letter = sheet.cell(row=3, column=column).column_letter
                width = sheet.column_dimensions[letter].width
                if width is None or not 10 <= width <= 60:
                    raise ValueError(
                        f"{name}/{sheet_name}/{letter}: invalid width {width}"
                    )
            for row in sheet.iter_rows(min_row=8):
                for cell in row:
                    if cell.data_type in (TYPE_FORMULA, TYPE_ERROR):
                        raise ValueError(
                            f"{name}/{sheet_name}/{cell.coordinate}: "
                            "formula or Excel error forbidden"
                        )
                    if cell.value is not None and not cell.alignment.wrap_text:
                        raise ValueError(
                            f"{name}/{sheet_name}/{cell.coordinate}: "
                            "data wrapping missing"
                        )
            columns = field_columns(sheet)
            for row_number in range(8, 8 + count):
                authored_id = sheet.cell(row=row_number, column=columns["id"]).value
                if authored_id != row_number - 7:
                    raise ValueError(
                        f"{name}/{sheet_name}/{row_number}: non-sequential id"
                    )
                stable_key = sheet.cell(
                    row=row_number, column=columns["stable_key"]
                ).value
                payload_text = sheet.cell(
                    row=row_number, column=columns["payload_json"]
                ).value
                payload = json.loads(payload_text)
                if stable_key != payload.get("id"):
                    raise ValueError(
                        f"{name}/{sheet_name}/{row_number}: payload identity drift"
                    )
                expected_sources = "|".join(
                    reference["source_id"]
                    for reference in payload["source_refs"]
                )
                authored_sources = sheet.cell(
                    row=row_number, column=columns["source_ref_ids"]
                ).value
                if authored_sources != expected_sources:
                    raise ValueError(
                        f"{name}/{sheet_name}/{row_number}: source projection drift"
                    )
                expected_manifest = "|".join(payload["manifest_record_ids"])
                authored_manifest = sheet.cell(
                    row=row_number, column=columns["manifest_record_ids"]
                ).value or ""
                if authored_manifest != expected_manifest:
                    raise ValueError(
                        f"{name}/{sheet_name}/{row_number}: manifest projection drift"
                    )
    if expected_counts is not None and counts != expected_counts:
        raise ValueError("workbook row counts changed after save/reload")
    return counts


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def semantic_digest(directory: Path) -> str:
    payload: list[object] = []
    for name in WORKBOOKS:
        workbook = load_workbook(directory / name, data_only=False)
        for sheet in workbook.worksheets:
            cells = [[cell.value for cell in row] for row in sheet.iter_rows()]
            widths = {
                key: value.width
                for key, value in sheet.column_dimensions.items()
                if value.width is not None
            }
            payload.append(
                [
                    name,
                    sheet.title,
                    cells,
                    sheet.freeze_panes,
                    sheet.auto_filter.ref,
                    widths,
                ]
            )
    encoded = json.dumps(
        payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()
