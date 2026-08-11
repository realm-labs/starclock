#!/usr/bin/env python3
"""Author the production Anomaly Arbitration workbook from reviewed Candidate rows."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def records(root: Path, name: str) -> list[dict]:
    document = json.loads(
        (root / "content-reference/anomaly-arbitration-v1" / f"{name}.json")
        .read_text(encoding="utf-8")
    )
    return document["records"]


def source_ref(row: dict) -> str:
    return json.dumps(
        row["source_refs"][0], ensure_ascii=False, sort_keys=True, separators=(",", ":")
    )


def rows(root: Path) -> dict[str, list[dict]]:
    profiles = records(root, "profiles")
    stages = records(root, "stages")
    targets = {row["source_numeric_id"]: row for row in records(root, "targets")}
    quadrants = records(root, "quadrant-options")
    clock_rows = {row["id"]: row for row in records(root, "clocks")}
    profile = profiles[0]
    authored_profiles = [{
        "id": 1,
        "stable_key": profile["id"],
        "game_version": "4.4",
        # Released text proves only that the first cycle is longer. These
        # deterministic values are isolated by the policy row below.
        "first_window_scaled": 150_000_000,
        "later_window_scaled": 100_000_000,
        "source_ref": source_ref(profile),
    }]
    cycle_limits = {"Knight": 6, "KingNormal": 6, "KingPlight": 2}
    stage_kinds = {
        ("Knight", "Normal"): "Knight",
        ("King", "Normal"): "KingNormal",
        ("King", "Plight"): "KingPlight",
    }
    stage_ids = {row["id"]: index for index, row in enumerate(stages, start=1)}
    authored_stages = []
    authored_targets = []
    for row in stages:
        kind = stage_kinds[(row["stage_kind"], row["difficulty"])]
        stage_id = stage_ids[row["id"]]
        authored_stages.append({
            "id": stage_id,
            "profile_id": 1,
            "stable_key": row["id"],
            "source_stage_id": row["source_stage_id"],
            "kind": kind,
            "display_order": row["display_order"],
            "team_index": min(row["display_order"] - 1, 3),
            "cycle_limit": cycle_limits[kind],
            "encounter_id": row["source_stage_id"],
            "source_ref": source_ref(row),
        })
        for target_id in row["battle_target_ids"]:
            target = targets[int(target_id)]
            is_death = target["source_ability_name"] == "BattleTarget_DeathCount"
            authored_targets.append({
                "id": len(authored_targets) + 1,
                "stage_id": stage_id,
                "stable_key": f"{row['id']}.{target['id']}",
                "kind": "NoDefeatedParticipants" if is_death else "ConsumedCyclesAtMost",
                "threshold": target["comparison_parameter"],
                "source_ref": source_ref(target),
            })
    authored_quadrants = [{
        "id": index,
        "stable_key": row["id"],
        "upstream_buff_id": row["source_numeric_id"],
        "rule_bundle_id": row["source_numeric_id"],
        "behavior_exact": row["binding_program_state"] == "ResolvedInExtractedAbilityList",
        "source_ref": source_ref(row),
    } for index, row in enumerate(quadrants, start=1)]
    first_cycle = clock_rows["clock.first-cycle-action-value"]
    policies = [
        {
            "id": 1,
            "profile_id": 1,
            "stable_key": "policy.anomaly.first-cycle-action-value",
            "known_facts": "Released official text states that the first cycle has greater total action value than later cycles, without a numeric constant.",
            "selected_behavior": "Use 150 action value for the first cycle and 100 for later cycles, while preserving one stage-local clock across waves.",
            "rejected_alternatives": "omit the stage clock|use equal cycle windows",
            "rationale": "The standard challenge clock requires exact integer windows; 150/100 preserves the only released ordering and matches the shared cycle convention.",
            "affected_tests": "challenge_anomaly::tests::production_profile_preserves_released_topology|anomaly_runtime::tests::normal_king_requires_three_knight_clears",
            "confidence": "Low",
            "replacement_condition": "Replace with a released configuration constant or reproducible action-value trace.",
            "source_ref": source_ref(first_cycle),
        },
        {
            "id": 2,
            "profile_id": 1,
            "stable_key": "policy.anomaly.king-protection-numeric-effects",
            "known_facts": "Released official text proves that uncleared Knights protect the King and that Plight starts with active protection; no complete numeric effect program is public.",
            "selected_behavior": "Track one active protection contribution per uncleared Knight and expose the count to battle assembly without applying invented combat numbers.",
            "rejected_alternatives": "invent three stacking buffs|treat protection as absent",
            "rationale": "The lifecycle and availability remain executable while combat parity stays explicitly bounded.",
            "affected_tests": "anomaly_runtime::tests::protection_count_tracks_stage_kind|anomaly_runtime::tests::plight_is_directly_available",
            "confidence": "Low",
            "replacement_condition": "Replace when released ability programs identify the protection contributions and parameters.",
            "source_ref": source_ref(records(root, "king-protection")[0]),
        },
        {
            "id": 3,
            "profile_id": 1,
            "stable_key": "policy.anomaly.fixed-runtime-roster",
            "known_facts": "Successful Knight clears record participants and account equipment instances; retries may use current progression and have explicit replacement/reset behavior.",
            "selected_behavior": "One runtime instance locks three disjoint Knight teams and one independent King team. A changed build starts a new instance; no account equipment identity is inferred from BuildDigest.",
            "rejected_alternatives": "treat opaque build digests as equipment instance IDs|mutate a sealed participant lock",
            "rationale": "The current shared participant seam intentionally owns immutable resolved builds and does not expose account inventory identities.",
            "affected_tests": "anomaly_runtime::tests::definition_rejects_duplicate_knight_characters|anomaly_runtime::tests::failed_retry_preserves_best_record",
            "confidence": "High",
            "replacement_condition": "Replace when activity preparation accepts a typed account-equipment snapshot and deterministic lock replacement command.",
            "source_ref": source_ref(records(root, "loadout-records")[0]),
        },
    ]
    return {
        "Profiles": authored_profiles,
        "Stages": authored_stages,
        "Targets": authored_targets,
        "Quadrants": authored_quadrants,
        "Policies": policies,
    }


def columns(sheet) -> list[str]:
    return [str(cell.value) for cell in sheet[3] if cell.value not in (None, "#field")]


def style_sheet(sheet) -> None:
    for column in range(1, sheet.max_column + 1):
        values = (
            str(sheet.cell(row, column).value or "")
            for row in range(1, sheet.max_row + 1)
        )
        width = min(max(max(map(len, values), default=0) + 2, 10), 40)
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows(min_row=8):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:7"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--replace",
        action="store_true",
        help="atomically replace the known generated production workbook",
    )
    args = parser.parse_args()
    if args.output.exists() and not args.replace:
        raise FileExistsError(f"refusing to overwrite {args.output}")
    workbook = openpyxl.load_workbook(args.template)
    authored = rows(args.root.resolve())
    if workbook.sheetnames != list(authored):
        raise ValueError(f"template sheets differ: {workbook.sheetnames}")
    for name, values in authored.items():
        sheet = workbook[name]
        if sheet.max_row >= 8:
            sheet.delete_rows(8, sheet.max_row - 7)
        headers = columns(sheet)
        for value in values:
            missing = set(headers) - set(value)
            if missing:
                raise ValueError(f"{name}: missing fields {sorted(missing)}")
            sheet.append([None, *[value[column] for column in headers]])
        sheet.freeze_panes = "A8"
        sheet.sheet_view.showGridLines = False
        sheet.auto_filter.ref = f"A3:{sheet.cell(3, sheet.max_column).column_letter}{sheet.max_row}"
        style_sheet(sheet)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    if args.replace:
        temporary = args.output.with_suffix(".authoring.tmp.xlsx")
        if temporary.exists():
            raise FileExistsError(f"refusing to overwrite stale temporary file {temporary}")
        workbook.save(temporary)
        temporary.replace(args.output)
    else:
        workbook.save(args.output)
    print(f"Authored {args.output} with {sum(map(len, authored.values()))} typed rows")


if __name__ == "__main__":
    main()
