#!/usr/bin/env python3
"""Author a clean typed Memory of Chaos runtime workbook from Candidate evidence."""

from __future__ import annotations

import argparse
import hashlib
import json
from decimal import Decimal, ROUND_HALF_EVEN
from pathlib import Path
from urllib.request import urlopen

import openpyxl


SOURCE_URL = "https://gitlab.com/Dimbreath/turnbasedgamedata.git"
SOURCE_REVISION = "fd978d6ef09f941fba644c731ab54abd6f7c3568"
SOURCE_RAW = f"{SOURCE_URL}/-/raw/{SOURCE_REVISION}"

# Candidate rows prove these exact identities and encounter placements, but the
# shared production catalog does not yet contain reviewed executable behavior
# for every Version 4.4 variant. Missing rows borrow a deterministic same-rank
# behavior until their exact ability programs are lowered. The workbook and its
# ProjectPolicy row make every such substitution inspectable.
BEHAVIOR_DONORS = {
    "enemy.silvermane-lieutenant.elite.variant.01": "enemy.silvermane-lieutenant-bug.elite.variant.01",
    "enemy.gepard.littleboss.variant.11": "enemy.gepard-complete.littleboss.variant.01",
    "enemy.svarog.littleboss.variant.13": "enemy.svarog-complete.littleboss.variant.01",
    "enemy.howling-casket.elite.variant.01": "enemy.the-ascended.elite.variant.01",
    "enemy.borisin-warhead-hoolay.littleboss.variant.01": "enemy.svarog-complete.littleboss.variant.01",
    "enemy.present-inebriated-in-revelry.elite.variant.01": "enemy.the-ascended.elite.variant.01",
    "enemy.wonder-forests-banacademic-office-staff.littleboss.variant.01": "enemy.harmonious-choir-the-great-septimus.bigboss.variant.01",
    "enemy.argenti.littleboss.variant.13": "enemy.cocolia-mother-of-deception.bigboss.variant.01",
    "enemy.tide-eroded-blade.minionlv2.variant.01": "enemy.everwinter-shadewalker.minionlv2.variant.01",
    "enemy.black-tides-corroded-axe.minionlv2.variant.01": "enemy.incineration-shadewalker.minionlv2.variant.01",
    "enemy.black-tides-corroded-daemon.elite.variant.01": "enemy.frigid-prowler.elite.variant.01",
    "enemy.black-tides-champion.elite.variant.01": "enemy.stormbringer.elite.variant.01",
    "enemy.dark-sun-gryphon.elite.variant.01": "enemy.stormbringer.elite.variant.01",
    "enemy.flame-reaver-of-the-deepest-dark.littleboss.variant.01": "enemy.cocolia-mother-of-deception.bigboss.variant.01",
    "enemy.aggressive-reading-material.minionlv2.variant.01": "enemy.dreamjolt-troupes-bubble-hound.minionlv2.variant.01",
    "enemy.rocking-rebel.elite.variant.01": "enemy.dreamjolt-troupes-beyond-overcooked.elite.variant.01",
    "enemy.canvas-peacock.elite.variant.01": "enemy.dreamjolt-troupes-sweet-gorilla.elite.variant.01",
    "enemy.daybreak-squadron-azurewing.elite.variant.01": "enemy.stormbringer.elite.variant.01",
    "enemy.daybreak-squadron-dawnlance.elite.variant.01": "enemy.automaton-direwolf.elite.variant.01",
    "enemy.sparxiconofficial.littleboss.variant.01": "enemy.harmonious-choir-the-great-septimus.bigboss.variant.01",
    "enemy.murata-graphia-founding-artist.littleboss.variant.01": "enemy.cocolia-mother-of-deception.bigboss.variant.01",
    "enemy.god-devourer-offspring.elite.variant.01": "enemy.the-ascended.elite.variant.01",
}


def payloads(root: Path, table: str) -> list[dict]:
    source = root / "config/memory-of-chaos-generated/debug-json" / f"{table}.json"
    document = json.loads(source.read_text(encoding="utf-8"))
    return [json.loads(row["values"]["payload_json"]["String"]) for row in document["table"]["rows"]]


def evidence(row: dict) -> dict:
    reference = row["evidence_refs"][0]
    return {
        "source_url": reference["repository_or_url"],
        "source_revision": reference["revision_or_access_date"],
        "source_path": reference["path_or_page"],
        "source_locator": reference["row_locator"],
        "evidence_digest": reference["evidence_sha256"],
        "evidence_quality": row.get("evidence_quality", reference["quality"]),
        "mechanism_quality": row.get("mechanism_quality", reference["mechanism_quality"]),
    }


def exact_source(root: Path, name: str) -> list[dict]:
    cache = root / ".cache/challenge-runtime-source" / f"{name}.json"
    if not cache.exists():
        cache.parent.mkdir(parents=True, exist_ok=True)
        with urlopen(f"{SOURCE_RAW}/ExcelOutput/{name}.json") as response:
            cache.write_bytes(response.read())
    return json.loads(cache.read_text(encoding="utf-8"), parse_float=Decimal)


def decimal(value) -> Decimal:
    return value if isinstance(value, Decimal) else Decimal(str(value))


def scaled(value: Decimal) -> int:
    return int((value * 1_000_000).to_integral_value(rounding=ROUND_HALF_EVEN))


def value(row: dict, field: str, default: str = "1") -> Decimal:
    return decimal(row.get(field, {"Value": default})["Value"])


def stat_evidence(
    slot: dict,
    encounter: dict,
    variant: dict,
    template: dict,
    hard: dict,
    elite: dict,
) -> dict:
    selected = [slot, encounter, variant["definition"], template["definition"], hard, elite]
    encoded = json.dumps(
        selected,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")
    return {
        "source_url": SOURCE_URL,
        "source_revision": SOURCE_REVISION,
        "source_path": (
            "ExcelOutput/StageConfig.json|ExcelOutput/MonsterConfig.json|"
            "ExcelOutput/MonsterTemplateConfig.json|ExcelOutput/HardLevelGroup.json|"
            "ExcelOutput/EliteGroup.json"
        ),
        "source_locator": (
            f"StageID={encounter['upstream_stage_config_id']};"
            f"MonsterID={variant['upstream_monster_id']};"
            f"MonsterTemplateID={template['source_template_id']};"
            f"HardLevelGroup={encounter['hard_level_group']},Level={encounter['level']};"
            f"EliteGroup={encounter['elite_group']}"
        ),
        "evidence_digest": hashlib.sha256(encoded).hexdigest(),
        "evidence_quality": "ExactStructured",
        "mechanism_quality": "ExactFormulaProjection",
    }


def rows(root: Path) -> dict[str, list[dict]]:
    stages = payloads(root, "MocStages")
    nodes = payloads(root, "MocNodes")
    tierce = payloads(root, "MocTierce")[0]
    objectives = payloads(root, "MocObjectives")
    clocks = {row["id"]: row for row in payloads(root, "MocClockRules")}
    encounters = payloads(root, "MocEncounters")
    waves = payloads(root, "MocWaves")
    enemy_slots = payloads(root, "MocEnemySlots")
    enemy_variants = payloads(root, "MocEnemyVariants")
    enemy_templates = payloads(root, "MocEnemyTemplates")
    hard_levels = exact_source(root, "HardLevelGroup")
    elite_groups = exact_source(root, "EliteGroup")
    profile_clock = clocks["clock.first-cycle-av-window"]
    profile = [{
        "id": 1,
        "stable_key": "profile.memory-of-chaos.v4.4",
        "game_version": "4.4",
        "initial_cycles": 30,
        "first_window_scaled": 150_000_000,
        "later_window_scaled": 100_000_000,
        "reset_window_on_wave": True,
        "expiry": "Lose",
        **evidence(profile_clock),
    }]
    stage_ids = {stage["id"]: index for index, stage in enumerate(stages, start=1)}
    authored_stages = [{
        "id": stage_ids[stage["id"]],
        "profile_id": 1,
        "stable_key": stage["id"],
        "upstream_stage_id": stage["upstream_stage_id"],
        "floor": stage["floor"],
        "initial_cycles": stage["challenge_countdown"],
        "turbulence_upstream_id": int(stage["turbulence_id"].rsplit(".", 1)[1]),
        **evidence(stage),
    } for stage in stages]
    authored_stages.append({
        "id": 13,
        "profile_id": 1,
        "stable_key": "stage.5212.starward",
        "upstream_stage_id": tierce["upstream_tierce_id"],
        "floor": 12,
        "initial_cycles": tierce["challenge_countdown"],
        "turbulence_upstream_id": int(stages[-1]["turbulence_id"].rsplit(".", 1)[1]),
        **evidence(tierce),
    })
    authored_nodes = [{
        "id": index,
        "stage_id": stage_ids[node["stage_id"]],
        "stable_key": node["id"],
        "node_index": node["node_index"],
        "team_index": node["node_index"] - 1,
        "upstream_encounter_id": node["upstream_stage_config_id"],
        **evidence(node),
    } for index, node in enumerate(nodes, start=1)]
    floor_twelve_nodes = sorted(
        (node for node in nodes if node["stage_id"] == stages[-1]["id"]),
        key=lambda node: node["node_index"],
    )
    for node_index, encounter_id in enumerate([
        *(node["upstream_stage_config_id"] for node in floor_twelve_nodes),
        int(tierce["stage_config_ids"][0].rsplit(".", 1)[1]),
    ], start=1):
        authored_nodes.append({
            "id": len(authored_nodes) + 1,
            "stage_id": 13,
            "stable_key": f"node.5212.starward.{node_index}",
            "node_index": node_index,
            "team_index": node_index - 1,
            "upstream_encounter_id": encounter_id,
            **evidence(tierce),
        })
    objective_kinds = {
        "objective.251": "RemainingCyclesAtLeast",
        "objective.252": "RemainingCyclesAtLeast",
        "objective.253": "NoDefeatedParticipants",
        "objective.601": "RemainingCyclesAtLeast",
        "objective.602": "RemainingCyclesAtLeast",
        "objective.603": "NoDefeatedParticipants",
    }
    authored_objectives = [{
        "id": index,
        "profile_id": 1,
        "stable_key": objective["id"],
        "kind": objective_kinds[objective["id"]],
        "threshold": objective["threshold"],
        **evidence(objective),
    } for index, objective in enumerate(objectives, start=1)]
    policy_rows = []
    for clock_id in (
        "clock.first-cycle-av-window",
        "clock.cycle-tick-boundary",
        "clock.node-carry",
        "clock.wave-carry",
        "clock.expiry-failure",
    ):
        clock = clocks[clock_id]
        for approximation_index, approximation in enumerate(clock.get("approximations", []), start=1):
            policy_rows.append({
                "id": len(policy_rows) + 1,
                "profile_id": 1,
                "stable_key": f"policy.{clock_id}.{approximation_index}",
                "known_facts": approximation["known_facts"],
                "selected_behavior": approximation["selected_behavior"],
                "rejected_alternatives": "|".join(approximation["rejected_alternatives"]),
                "rationale": approximation["rationale"],
                "affected_tests": "|".join(approximation["affected_fixture_ids"]),
                "confidence": approximation["confidence"],
                "replacement_condition": approximation["replacement_condition"],
                **evidence(clock),
            })
    production_identities = payload_table(
        root / "config/generated/debug-json/ContentIdentity.json"
    )
    production_enemy_keys = {
        typed_value(row, "stable_key")
        for row in production_identities
        if typed_value(row, "content_kind") == "EnemyVariant"
    }
    shared_variants = {
        row["id"]: row
        for row in json.loads(
            (root / "content-reference/v4.4/enemy-variants.json").read_text(encoding="utf-8")
        )
    }
    shared_templates = {
        row["id"]: row
        for row in json.loads(
            (root / "content-reference/v4.4/enemy-templates.json").read_text(encoding="utf-8")
        )
    }
    variant_by_upstream = {
        int(row["upstream_monster_id"]): row for row in enemy_variants
    }
    template_by_id = {row["id"]: row for row in enemy_templates}
    hard_level_by_key = {
        (row["HardLevelGroup"], row["Level"]): row for row in hard_levels
    }
    elite_group_by_id = {row["EliteGroup"]: row for row in elite_groups}
    selected_encounter_ids = {
        node["upstream_encounter_id"] for node in authored_nodes
    }
    selected_encounters = [
        row for row in encounters
        if row["upstream_stage_config_id"] in selected_encounter_ids
    ]
    selected_waves = [
        row for row in waves
        if row["upstream_stage_config_id"] in selected_encounter_ids
    ]
    selected_slots = [
        row for row in enemy_slots
        if row["upstream_stage_config_id"] in selected_encounter_ids
    ]
    referenced_variants = sorted({
        row["upstream_enemy_variant_id"] for row in selected_slots
    })
    binding_id = {
        upstream: index for index, upstream in enumerate(referenced_variants, start=1)
    }
    authored_bindings = []
    for upstream in referenced_variants:
        variant = variant_by_upstream[upstream]
        stable_key = variant["shared_variant_id"]
        exact = stable_key in production_enemy_keys
        donor = stable_key if exact else BEHAVIOR_DONORS.get(stable_key)
        if donor is None or donor not in production_enemy_keys:
            raise ValueError(f"missing executable behavior donor for {stable_key}")
        authored_bindings.append({
            "id": binding_id[upstream],
            "upstream_variant_id": upstream,
            "stable_key": stable_key,
            "behavior_source_key": donor,
            "behavior_exact": exact,
            **evidence(variant),
        })
    node_id_by_encounter = {
        node["upstream_encounter_id"]: node["id"] for node in authored_nodes
    }
    authored_encounters = [{
        "id": row["upstream_stage_config_id"],
        "node_id": node_id_by_encounter[row["upstream_stage_config_id"]],
        "level": row["level"],
        "hard_level_group": row["hard_level_group"],
        **evidence(row),
    } for row in selected_encounters]
    wave_id = {
        row["id"]: index for index, row in enumerate(selected_waves, start=1)
    }
    authored_waves = [{
        "id": wave_id[row["id"]],
        "encounter_id": row["upstream_stage_config_id"],
        "sequence": row["wave_index"],
        **evidence(row),
    } for row in selected_waves]
    selected_slots.sort(key=lambda row: row["order_key"])
    encounter_by_id = {
        row["upstream_stage_config_id"]: row for row in selected_encounters
    }
    authored_slots = []
    for index, row in enumerate(selected_slots, start=1):
        encounter = encounter_by_id[row["upstream_stage_config_id"]]
        variant = variant_by_upstream[row["upstream_enemy_variant_id"]]
        template = template_by_id[variant["enemy_template_id"]]
        hard = hard_level_by_key[(encounter["hard_level_group"], encounter["level"])]
        elite = elite_group_by_id[encounter["elite_group"]]
        base = template["definition"]["base_stats"]
        multiplier = variant["definition"]["stat_multipliers"]

        def base_stat(stat: str) -> Decimal:
            authored = base[stat]
            if authored is not None:
                return decimal(authored)
            donor_key = BEHAVIOR_DONORS[variant["shared_variant_id"]]
            donor_variant = shared_variants[donor_key]
            return decimal(shared_templates[donor_variant["enemy_id"]]["base_stats"][stat])

        def resolved(stat: str, hard_field: str, elite_field: str) -> Decimal:
            return (
                base_stat(stat)
                * decimal(multiplier[stat])
                * value(elite, elite_field)
                * value(hard, hard_field)
            )

        rank = template["definition"]["rank"]
        authored_slots.append({
            "id": index,
            "wave_id": wave_id[row["wave_id"]],
            "enemy_binding_id": binding_id[row["upstream_enemy_variant_id"]],
            "spawn_sequence": row["slot_index"] + 1,
            "formation_index": row["slot_index"],
            "rank": (
                "Boss"
                if rank in {"LittleBoss", "BigBoss", "Boss"}
                else "Elite"
                if rank == "Elite"
                else "Normal"
            ),
            "weaknesses": "|".join(sorted(variant["definition"]["weaknesses"])),
            "maximum_hp_scaled": scaled(resolved("hp", "HPRatio", "HPRatio")),
            "attack_scaled": scaled(resolved("atk", "AttackRatio", "AttackRatio")),
            "defense_scaled": scaled(resolved("def", "DefenceRatio", "DefenceRatio")),
            "speed_scaled": scaled(resolved("spd", "SpeedRatio", "SpeedRatio")),
            "effect_hit_rate_scaled": scaled(value(hard, "StatusProbability", "0")),
            "effect_resistance_scaled": scaled(
                decimal(base["effect_res"]) + value(hard, "StatusResistance", "0")
            ),
            "toughness_scaled": scaled(
                resolved("toughness", "StanceRatio", "StanceRatio")
            ),
            **stat_evidence(row, encounter, variant, template, hard, elite),
        })
    approximate = [row for row in authored_bindings if not row["behavior_exact"]]
    if len(authored_encounters) != 25 or len(authored_waves) != 50 or len(authored_slots) != 99:
        raise ValueError("Memory ordinary/Starward encounter closure denominator drift")
    if len(authored_bindings) != 41:
        raise ValueError("Memory enemy behavior closure denominator drift")
    first_approximation = variant_by_upstream[approximate[0]["upstream_variant_id"]]
    policy_rows.append({
        "id": len(policy_rows) + 1,
        "profile_id": 1,
        "stable_key": "policy.memory.enemy-behavior-donors",
        "known_facts": (
            f"The released Version 4.4 rows prove {len(authored_bindings)} enemy identities "
            f"reachable from {len(authored_encounters)} playable encounters, "
            f"{len(authored_waves)} waves and {len(authored_slots)} slots. The shared executable "
            f"catalog currently contains reviewed behavior for "
            f"{len(authored_bindings) - len(approximate)} identities; {len(approximate)} "
            "normalized ability programs are reference-only and are not yet lowered."
        ),
        "selected_behavior": (
            f"Keep every exact encounter identity and placement. For each of the {len(approximate)} missing "
            "executable variants, clone the explicitly authored same-rank behavior_source_key "
            "while retaining a distinct mode-owned EnemyDefinitionId."
        ),
        "rejected_alternatives": (
            "parse Candidate payload_json at runtime|silently drop missing enemies|"
            "claim normalized reference programs are executable parity"
        ),
        "rationale": (
            "This keeps battles deterministic and runnable without crossing the Sora runtime "
            "boundary, and makes every temporary behavior substitution visible by ID."
        ),
        "affected_tests": (
            "challenge::tests::production_memory_combat_definitions_lower_exact_closure|"
            "challenge_combat::tests::memory_combat_catalog_composes_all_playable_encounters"
        ),
        "confidence": "Low",
        "replacement_condition": (
            "Replace each donor binding when its exact Version 4.4 enemy ability, AI, summon "
            "and phase closure is reviewed and lowered into the shared executable catalog."
        ),
        **evidence(first_approximation),
    })
    policy_rows.append({
        "id": len(policy_rows) + 1,
        "profile_id": 1,
        "stable_key": "policy.memory.starward-composition",
        "known_facts": (
            "Released Tierce 5213 selects encounter 30123123 after floor 12, publishes a "
            "45-cycle clock and objectives at 15/30 remaining cycles plus no defeated "
            "participants. Released public guidance confirms three teams share 45 cycles."
        ),
        "selected_behavior": (
            "Represent Starward as a thirteenth three-node stage containing both floor-12 "
            "nodes followed by Tierce. Carry one 45-cycle Activity clock through all three "
            "battles and apply the active Memory Turbulence to Tierce."
        ),
        "rejected_alternatives": (
            "run Tierce as an unrelated one-team activity|give each node 45 cycles|"
            "silently omit Memory Turbulence from the third node"
        ),
        "rationale": (
            "This matches the released predecessor, encounter, clock, objective and public "
            "three-team behavior; Turbulence dispatch to Tierce remains the one inferred edge."
        ),
        "affected_tests": (
            "challenge::tests::production_memory_profile_lowers_all_active_stages_and_policies|"
            "memory_runtime::tests::starward_carries_cycles_through_three_nodes"
        ),
        "confidence": "Medium",
        "replacement_condition": (
            "Replace the Turbulence inference when a released Tierce battle-creation trace or "
            "decoded MazeBuff dispatch field proves its exact attachment."
        ),
        **evidence(tierce),
    })
    office_staff = next(
        row for row in enemy_variants
        if row["shared_variant_id"]
        == "enemy.wonder-forests-banacademic-office-staff.littleboss.variant.01"
    )
    policy_rows.append({
        "id": len(policy_rows) + 1,
        "profile_id": 1,
        "stable_key": "policy.memory.office-staff-toughness-donor",
        "known_facts": (
            "The released office-staff template has no numeric base Toughness in the normalized "
            "Version 4.4 row. Its encounter identity, level, formation and other stats are exact."
        ),
        "selected_behavior": (
            "Use the 180 base Toughness of its explicit Great Septimus behavior donor, then apply "
            "the exact encounter EliteGroup and HardLevelGroup Stance ratios."
        ),
        "rejected_alternatives": (
            "invent a new boss Toughness value|create no Toughness layer|"
            "parse an undocumented runtime object graph"
        ),
        "rationale": (
            "The donor already supplies this temporary boss behavior, and reusing its reviewed "
            "Toughness is deterministic and easier to replace than an unrelated literal."
        ),
        "affected_tests": (
            "challenge::tests::production_memory_combat_definitions_lower_exact_closure|"
            "challenge_combat::tests::memory_combat_catalog_composes_all_playable_encounters"
        ),
        "confidence": "Low",
        "replacement_condition": (
            "Replace the donor value when a released numeric base Toughness or decoded shared-"
            "Toughness phase definition for the office-staff boss is available."
        ),
        **evidence(office_staff),
    })
    return {
        "Profiles": profile,
        "Stages": authored_stages,
        "Nodes": authored_nodes,
        "Objectives": authored_objectives,
        "Policies": policy_rows,
        "EnemyBindings": authored_bindings,
        "Encounters": authored_encounters,
        "Waves": authored_waves,
        "EnemySlots": authored_slots,
    }


def payload_table(source: Path) -> list[dict]:
    return json.loads(source.read_text(encoding="utf-8"))["table"]["rows"]


def typed_value(row: dict, field: str):
    encoded = row["values"][field]
    return next(iter(encoded.values()))


def columns(sheet) -> list[str]:
    return [str(cell.value) for cell in sheet[3] if cell.value not in (None, "#field")]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    if args.output.exists():
        raise FileExistsError(f"refusing to overwrite {args.output}")
    workbook = openpyxl.load_workbook(args.template)
    authored = rows(args.root.resolve())
    if workbook.sheetnames != list(authored):
        raise ValueError(f"template sheets differ: {workbook.sheetnames}")
    for name, values in authored.items():
        sheet = workbook[name]
        if sheet.max_row >= 8:
            sheet.delete_rows(8, sheet.max_row - 7)
        headers = columns(sheet)
        for value in values:
            missing = set(headers) - set(value)
            if missing:
                raise ValueError(f"{name}: missing fields {sorted(missing)}")
            sheet.append([None, *[value[column] for column in headers]])
        sheet.freeze_panes = "A8"
        sheet.auto_filter.ref = f"A3:{sheet.cell(3, sheet.max_column).column_letter}{sheet.max_row}"
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Authored {args.output} with {sum(map(len, authored.values()))} typed rows")


if __name__ == "__main__":
    main()
