#!/usr/bin/env python3
"""Author a clean typed Pure Fiction runtime workbook from Candidate evidence."""

from __future__ import annotations

import argparse
import json
from decimal import Decimal, ROUND_HALF_EVEN
from pathlib import Path

import openpyxl


NORMAL_DONOR = "enemy.everwinter-shadewalker.minionlv2.variant.01"
ELITE_DONOR = "enemy.the-ascended.elite.variant.01"
BOSS_DONOR = "enemy.harmonious-choir-the-great-septimus.bigboss.variant.01"


def payloads(root: Path, table: str) -> list[dict]:
    source = root / "config/pure-fiction-generated/debug-json" / f"{table}.json"
    document = json.loads(source.read_text(encoding="utf-8"))
    return [json.loads(row["values"]["payload_json"]["String"])
            for row in document["table"]["rows"]]


def source_map(root: Path) -> dict[str, dict]:
    return {row["id"]: row for row in payloads(root, "PfSourceRecord")}


def source_ref(row: dict, sources: dict[str, dict]) -> str:
    records = [sources[value] for value in row["source_record_ids"]]
    return json.dumps(records, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def stable_variants(root: Path) -> dict[int, str]:
    rows = json.loads((root / "content-reference/v4.4/enemy-variants.json").read_text(
        encoding="utf-8"
    ))
    return {int(row["source_monster_id"]): row["id"] for row in rows}


def typed_value(row: dict, field: str):
    encoded = row["values"][field]
    return next(iter(encoded.values()))


def production_enemy_keys(root: Path) -> set[str]:
    document = json.loads((root / "config/generated/debug-json/ContentIdentity.json").read_text(
        encoding="utf-8"
    ))
    return {
        typed_value(row, "stable_key")
        for row in document["table"]["rows"]
        if typed_value(row, "content_kind") == "EnemyVariant"
    }


def decimal(value) -> Decimal:
    return value if isinstance(value, Decimal) else Decimal(str(value))


def scaled(value: Decimal) -> int:
    return int((value * 1_000_000).to_integral_value(rounding=ROUND_HALF_EVEN))


def hard_levels(root: Path) -> dict[tuple[int, int], dict]:
    path = root / ".cache/challenge-runtime-source/HardLevelGroup.json"
    rows = json.loads(path.read_text(encoding="utf-8"), parse_float=Decimal)
    return {(row["HardLevelGroup"], row["Level"]): row for row in rows}


def hard_value(row: dict, field: str, default: str = "1") -> Decimal:
    return decimal(row.get(field, {"Value": default})["Value"])


def rank_name(value: str) -> str:
    if value in {"LittleBoss", "BigBoss", "Boss"}:
        return "Boss"
    if value == "Elite":
        return "Elite"
    return "Normal"


def rows(root: Path) -> dict[str, list[dict]]:
    sources = source_map(root)
    starward_public_source = json.dumps({
        "repository_or_url": "https://honkai-star-rail.fandom.com/wiki/Pure_Fiction",
        "revision_or_access_date": "2026-08-11",
        "path_or_page": "Starward Mode",
        "quality": "ReleasedPublicCrossCheck",
    }, sort_keys=True, separators=(",", ":"))
    stages = payloads(root, "PfStage")[:4]
    tierce = payloads(root, "PfTierceStarward")[0]
    nodes = [row for row in payloads(root, "PfNode") if row["stage_id"] in {
        stage["stage_id"] for stage in stages
    }]
    objectives = [row for row in payloads(root, "PfObjective") if row["family"] == "ordinary"]
    cacophonies = payloads(root, "PfCacophony")
    encounters = payloads(root, "PfEncounter")
    waves = payloads(root, "PfWave")
    slots = payloads(root, "PfEnemySlot")
    variants = {row["monster_id"]: row for row in payloads(root, "PfEnemyVariant")}
    templates = {row["template_id"]: row for row in payloads(root, "PfEnemyTemplate")}
    clocks = {row["stage_id"]: row for row in payloads(root, "PfClock")}
    stable = stable_variants(root)
    executable = production_enemy_keys(root)
    hard = hard_levels(root)

    profile = [{
        "id": 1,
        "stable_key": "profile.pure-fiction.v4.4",
        "game_version": "4.4",
        "first_window_scaled": 150_000_000,
        "later_window_scaled": 100_000_000,
        "expiry": "Finalize",
        "node_score_maximum": 40_000,
        "stage_score_maximum": 120_000,
        "source_ref": source_ref(clocks[stages[0]["stage_id"]], sources),
    }]
    stage_ids = {row["stage_id"]: index for index, row in enumerate(stages, start=1)}
    authored_stages = [{
        "id": stage_ids[row["stage_id"]],
        "profile_id": 1,
        "stable_key": row["id"],
        "upstream_stage_id": row["stage_id"],
        "floor": row["floor"],
        "initial_cycles": clocks[row["stage_id"]]["turn_limit"],
        "clear_score": row["clear_score"],
        "source_ref": source_ref(row, sources),
    } for row in stages]
    authored_stages.append({
        "id": 5,
        "profile_id": 1,
        "stable_key": "pf.stage.20244.starward",
        "upstream_stage_id": tierce["tierce_id"],
        "floor": 4,
        "initial_cycles": clocks[stages[-1]["stage_id"]]["turn_limit"],
        "clear_score": tierce["clear_score"],
        "source_ref": source_ref(stages[-1], sources) + "|" + source_ref(tierce, sources),
    })
    authored_nodes = [{
        "id": index,
        "stage_id": stage_ids[row["stage_id"]],
        "stable_key": row["id"],
        "node_index": row["side"],
        "team_index": row["side"] - 1,
        "encounter_id": row["stage_config_id"],
        "whimsicality_id": row["maze_buff_id"],
        "source_ref": source_ref(row, sources),
    } for index, row in enumerate(nodes, start=1)]
    stage_four_nodes = sorted(
        (row for row in nodes if row["stage_id"] == stages[-1]["stage_id"]),
        key=lambda row: row["side"],
    )
    tierce_node = next(row for row in payloads(root, "PfNode")
                       if row["stage_id"] == tierce["tierce_id"])
    for node_index, row in enumerate([*stage_four_nodes, tierce_node], start=1):
        authored_nodes.append({
            "id": len(authored_nodes) + 1,
            "stage_id": 5,
            "stable_key": f"pf.node.20244.starward.{node_index}",
            "node_index": node_index,
            "team_index": node_index - 1,
            "encounter_id": row["stage_config_id"],
            "whimsicality_id": row.get("maze_buff_id", stage_four_nodes[-1]["maze_buff_id"]),
            "source_ref": source_ref(row, sources) + "|" + source_ref(tierce, sources),
        })
    authored_objectives = [{
        "id": row["target_id"],
        "profile_id": 1,
        "stable_key": row["id"],
        "kind": "ScoreAtLeast",
        "threshold": row["threshold"],
        "source_ref": source_ref(row, sources),
    } for row in objectives]
    for target_id, threshold in zip(tierce["target_ids"], [60_000, 75_000, 90_000], strict=True):
        authored_objectives.append({
            "id": target_id,
            "profile_id": 1,
            "stable_key": f"pf.objective.starward.{target_id}",
            "kind": "ScoreAtLeast",
            "threshold": threshold,
            "source_ref": source_ref(tierce, sources) + "|" + starward_public_source,
        })
    authored_cacophonies = [{
        "id": index,
        "profile_id": 1,
        "stable_key": row["id"],
        "upstream_buff_id": row["buff_id"],
        "binding_key": row["binding_key"],
        "parameters": "|".join(row["parameters"]),
        "source_ref": source_ref(row, sources),
    } for index, row in enumerate(cacophonies, start=1)]

    enemy_ids = {monster: index for index, monster in enumerate(sorted(variants), start=1)}
    authored_enemies = []
    for monster in sorted(variants):
        variant = variants[monster]
        template = templates[variant["template_id"]]
        rank = rank_name(template["rank"])
        stable_key = stable[monster]
        exact = stable_key in executable
        donor = stable_key if exact else {
            "Normal": NORMAL_DONOR,
            "Elite": ELITE_DONOR,
            "Boss": BOSS_DONOR,
        }[rank]
        authored_enemies.append({
            "id": enemy_ids[monster],
            "upstream_monster_id": monster,
            "stable_key": stable_key,
            "behavior_source_key": donor,
            "behavior_exact": exact,
            "rank": rank,
            "weaknesses": "|".join(sorted(
                "Lightning" if value == "Thunder" else value
                for value in variant["weaknesses"]
            )),
            "source_ref": source_ref(variant, sources),
        })

    node_by_encounter = {row["encounter_id"]: row["id"] for row in authored_nodes}
    authored_encounters = [{
        "id": row["stage_config_id"],
        "node_id": node_by_encounter.get(row["stage_config_id"], ""),
        "level": row["level"],
        "battle_event_id": row["battle_event_id"],
        "source_ref": source_ref(row, sources),
    } for row in encounters]
    encounter_ids = {row["id"]: row["stage_config_id"] for row in encounters}
    wave_ids = {row["id"]: index for index, row in enumerate(waves, start=1)}
    authored_waves = []
    for row in waves:
        sequence = row["order"]
        authored_waves.append({
            "id": wave_ids[row["id"]],
            "encounter_id": encounter_ids[row["encounter_id"]],
            "sequence": sequence,
            "spawn_end": "DefeatQuota" if sequence == 1 else "RequiredSlots",
            "defeat_quota": 20 if sequence == 1 else "",
            "refill_source_wave": "" if sequence == 1 else 1,
            "maximum_simultaneous": 5,
            "score_cap": 8_000 if sequence == 1 else 16_000,
            "normal_defeat_true_damage_scaled": 0 if sequence == 1 else 30_000,
            "source_ref": source_ref(row, sources),
        })

    encounter_by_key = {row["id"]: row for row in encounters}
    authored_slots = []
    for index, row in enumerate(slots, start=1):
        encounter = encounter_by_key[row["encounter_id"]]
        variant = variants[row["monster_id"]]
        template = templates[variant["template_id"]]
        hard_row = hard[(variant["hard_level_group"], encounter["level"])]
        base = template["base_stats"]
        ratios = variant["stat_ratios"]

        def resolved(base_field: str, ratio_field: str, hard_field: str) -> Decimal:
            authored = base[base_field]
            if authored is None:
                authored = "180" if base_field == "stance" else "1"
            return decimal(authored) * decimal(ratios[ratio_field]) * hard_value(hard_row, hard_field)

        hp = resolved("hp", "hp", "HPRatio")
        formation = row["slot_order"] - 1 if row["wave_order"] == 1 else 4
        authored_slots.append({
            "id": index,
            "wave_id": wave_ids[f"pf.wave.{encounter['stage_config_id']}.{row['wave_order']}"],
            "enemy_id": enemy_ids[row["monster_id"]],
            "spawn_sequence": row["slot_order"],
            "formation_index": formation,
            "maximum_hp": max(1, int(hp.to_integral_value(rounding=ROUND_HALF_EVEN))),
            "attack_scaled": scaled(resolved("attack", "attack", "AttackRatio")),
            "defense_scaled": scaled(resolved("defence", "defence", "DefenceRatio")),
            "speed_scaled": scaled(resolved("speed", "speed", "SpeedRatio")),
            "effect_hit_rate_scaled": scaled(hard_value(hard_row, "StatusProbability", "0")),
            "effect_resistance_scaled": scaled(hard_value(hard_row, "StatusResistance", "0")),
            "toughness_scaled": scaled(resolved("stance", "stance", "StanceRatio")),
            "source_ref": source_ref(row, sources),
        })

    public_source = json.dumps({
        "repository_or_url": "https://honkai-star-rail.fandom.com/wiki/Pure_Fiction",
        "revision_or_access_date": "2026-08-11",
        "path_or_page": "Node Format, scoring, clocks and Grit Effect",
        "quality": "ReleasedPublicCrossCheck",
    }, sort_keys=True, separators=(",", ":"))
    candidate_spawn = payloads(root, "PfSpawnProgram")[0]
    score = payloads(root, "PfScoreProgram")[0]
    seasonal = payloads(root, "PfSeasonalMechanic")[0]
    policies = [
        {
            "id": 1, "profile_id": 1,
            "stable_key": "policy.pure-fiction.spawn-runtime-identity",
            "known_facts": "Released data proves an infinite group, five initial wave-1 slots, one wave-2/3 scoring target and after-defeat refill. Public released guidance confirms continuous replenishment, 20 wave-1 defeats and required targets in waves 2/3. Dynamic occurrence identity and hidden pool cursor are not exposed.",
            "selected_behavior": "Refill in stable formation order after each action. Reuse the defeated runtime UnitId, allocate a fresh SpawnSequence, fully reset occurrence state, and derive four wave-2/3 normal refill slots from wave 1 while reserving formation 4 for the scoring target.",
            "rejected_alternatives": "ordinary next-wave victory without refill|unbounded dynamic catalog lookup|preallocate an arbitrary number of hidden units",
            "rationale": "The policy is bounded, replayable, inspectable and isolates the only unavailable identity detail without content-ID branches in Combat.",
            "affected_tests": "combat_damage_lifecycle::continuous_spawn_refills_in_slot_order_until_the_authored_quota|pure_fiction_runtime::tests::aggregates_two_independent_nodes",
            "confidence": "Medium",
            "replacement_condition": "Replace runtime-slot reuse and the derived refill pool when released engine traces expose occurrence allocation and pool cursor semantics.",
            "source_ref": source_ref(candidate_spawn, sources) + "|" + public_source,
        },
        {
            "id": 2, "profile_id": 1,
            "stable_key": "policy.pure-fiction.score-lowering",
            "known_facts": "Public released guidance gives 400 points per wave-1 defeat up to 8000, and 40 points per 0.25% scoring-target HP lost in waves 2/3 up to 16000 each. Candidate data proves 30000 ordinary clear score and 40000/50000/60000 objectives.",
            "selected_behavior": "Read the authoritative per-wave continuous-spawn defeat counter, which is incremented by committed Defeated events, and floor each scoring target's depleted-HP ratio at a 16000-point cap. Sum three wave scores, then sum two node scores.",
            "rejected_alternatives": "score displayed rounded damage|award full wave score on timeout|round partial target progress to nearest",
            "rationale": "This directly follows released score units and keeps applied state/events authoritative.",
            "affected_tests": "pure_fiction::tests::scores_three_wave_progress|pure_fiction_runtime::tests::aggregates_two_independent_nodes",
            "confidence": "High",
            "replacement_condition": "Replace only if a pinned released scoring program demonstrates different integer boundary ordering.",
            "source_ref": source_ref(score, sources) + "|" + public_source,
        },
        {
            "id": 3, "profile_id": 1,
            "stable_key": "policy.pure-fiction.enemy-behavior-donors",
            "known_facts": "Candidate data proves 42 variants, 41 templates, exact levels, base stats, ratios, weaknesses and 63 authored slots. Not every normalized AI/ability program is executable in the shared production catalog.",
            "selected_behavior": "Project exact occurrence stats through HardLevelGroup and use explicit same-rank reviewed behavior donors only for missing executable variants.",
            "rejected_alternatives": "parse Candidate JSON at runtime|drop unavailable enemies|claim donor behavior is observed parity",
            "rationale": "Every identity and stat stays exact while temporary behavior substitutions remain visible in typed data.",
            "affected_tests": "challenge::tests::production_pure_fiction_combat_definitions_lower|pure_fiction_combat::tests::production_pure_fiction_catalog_composes_all_playable_encounters",
            "confidence": "Low",
            "replacement_condition": "Replace each donor when that variant's reviewed ability, AI, summon and phase closure is executable.",
            "source_ref": source_ref(next(iter(variants.values())), sources),
        },
        {
            "id": 4, "profile_id": 1,
            "stable_key": "policy.pure-fiction.grit-fever-and-cacophony",
            "known_facts": "Released public text and exact MazeBuff parameters prove Grit and Resurging Tide caps, the speed-100 Surging Grit countdown, 80% entry-damage parameter, 50% vulnerability, 1% two-turn Dejection with 20 stacks, the per-enemy 10-trigger debuff cap, and all three Cacophonies. Released material does not expose the fixed-damage level curve, equal-boundary ordering, a generic per-applier Indulgence key, or the shared Punchline cap.",
            "selected_behavior": "Apply negative-effect triggers in committed event order and track their 10-use cap on each enemy. Aggregate Surging entry damage into one true-damage event per enemy at 80% of the mode-rule host's base ATK for each current Debuff, Control or DoT stack. Use a speed-100 countdown, apply Dejection at one stack or Toccata's two stacks, and convert up to 30 Tide on exit. Model Indulgence as an independent two-turn instance per Elation action and create a missing shared Punchline resource with cap 999 before Mirthful Cadence grants 40.",
            "rejected_alternatives": "leave the seasonal mechanic silently inactive|guess from parameter vectors without public text|put mode IDs in the shared resolver",
            "rationale": "The released descriptions and exact parameters support typed rules, while the four unavailable engine joins remain deterministic, inspectable and replaceable without adding mode IDs to Combat.",
            "affected_tests": "pure_fiction_combat::tests::production_pure_fiction_catalog_composes_all_playable_encounters|challenge::tests::production_pure_fiction_profile_lowers_all_playable_stages",
            "confidence": "Medium",
            "replacement_condition": "Replace host-ATK entry damage, event aggregation/order, per-action Indulgence identity or Punchline cap when a released ability graph or reproducible public trace proves the corresponding engine behavior.",
            "source_ref": source_ref(seasonal, sources) + "|" + public_source,
        },
        {
            "id": 5, "profile_id": 1,
            "stable_key": "policy.pure-fiction.tierce-starward-exclusion",
            "known_facts": "Candidate data proves Tierce encounter 30322043, predecessor stage 20244, clear score 45000 and target IDs 4001/4002/4003. Released public guidance proves Starward Stage 4 uses three independent teams/nodes with 60000/75000/90000 star targets and retains the best score per node.",
            "selected_behavior": "Expose Starward as a fifth selectable runtime stage composed from ordinary Stage 4 nodes plus Tierce, with three Section-locked teams, independent four-cycle battles, 45000 clear score and 60000/75000/90000 objectives. Reuse the active seasonal Whimsicality for Tierce because its exact row omits a separate selector.",
            "rejected_alternatives": "silently exclude released Starward|treat Tierce as a standalone one-team stage|share one battle clock across all three teams",
            "rationale": "The three-node topology and thresholds are public and the encounter identity is exact; only the omitted Tierce Whimsicality join remains replaceable policy.",
            "affected_tests": "challenge::tests::production_pure_fiction_profile_lowers_all_playable_stages|pure_fiction_runtime_tests::tests::starward_runs_three_independent_nodes",
            "confidence": "Medium",
            "replacement_condition": "Replace the inherited Tierce Whimsicality or target-ID mapping when a released selector row or decoded schema provides exact joins.",
            "source_ref": source_ref(encounters[-1], sources) + "|" + starward_public_source,
        },
    ]
    if len(authored_stages) != 5 or len(authored_nodes) != 11:
        raise ValueError("playable Pure Fiction stage/node denominator drift")
    if len(authored_encounters) != 9 or len(authored_waves) != 27 or len(authored_slots) != 63:
        raise ValueError("Pure Fiction encounter closure denominator drift")
    if len(authored_enemies) != 42 or len(authored_cacophonies) != 3:
        raise ValueError("Pure Fiction enemy/Cacophony denominator drift")
    return {
        "Profiles": profile,
        "Stages": authored_stages,
        "Nodes": authored_nodes,
        "Objectives": authored_objectives,
        "Cacophonies": authored_cacophonies,
        "Policies": policies,
        "Enemies": authored_enemies,
        "Encounters": authored_encounters,
        "Waves": authored_waves,
        "EnemySlots": authored_slots,
    }


def columns(sheet) -> list[str]:
    return [str(cell.value) for cell in sheet[3] if cell.value not in (None, "#field")]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    if args.output.exists():
        raise FileExistsError(f"refusing to overwrite {args.output}")
    workbook = openpyxl.load_workbook(args.template)
    authored = rows(args.root.resolve())
    if workbook.sheetnames != list(authored):
        raise ValueError(f"template sheets differ: {workbook.sheetnames}")
    for name, values in authored.items():
        sheet = workbook[name]
        if sheet.max_row >= 8:
            sheet.delete_rows(8, sheet.max_row - 7)
        headers = columns(sheet)
        for value in values:
            missing = set(headers) - set(value)
            if missing:
                raise ValueError(f"{name}: missing fields {sorted(missing)}")
            sheet.append([None, *[value[column] for column in headers]])
        sheet.freeze_panes = "A8"
        sheet.auto_filter.ref = f"A3:{sheet.cell(3, sheet.max_column).column_letter}{sheet.max_row}"
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Authored {args.output} with {sum(map(len, authored.values()))} typed rows")


if __name__ == "__main__":
    main()
