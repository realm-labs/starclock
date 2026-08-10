#!/usr/bin/env python3
"""Author a clean typed Apocalyptic Shadow runtime workbook from Candidate data."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl


LEVELS = {30191: 65, 30192: 75, 30193: 85, 30194: 95}
PRIMARY_HP = {30191: 500_000, 30192: 1_000_000, 30193: 2_000_000, 30194: 4_000_000}
DONORS = {
    "Normal": "enemy.everwinter-shadewalker.minionlv2.variant.01",
    "Elite": "enemy.the-ascended.elite.variant.01",
    "Boss": "enemy.harmonious-choir-the-great-septimus.bigboss.variant.01",
}


def payloads(root: Path, table: str) -> list[dict]:
    source = root / "config/apocalyptic-shadow-generated/debug-json" / f"{table}.json"
    document = json.loads(source.read_text(encoding="utf-8"))
    return [json.loads(row["values"]["payload_json"]["String"])
            for row in document["table"]["rows"]]


def source_ref(row: dict) -> str:
    reference = row["source_refs"][0]
    return json.dumps(reference, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def stable_variants(root: Path) -> dict[int, str]:
    rows = json.loads((root / "content-reference/v4.4/enemy-variants.json").read_text(
        encoding="utf-8"
    ))
    return {int(row["source_monster_id"]): row["id"] for row in rows}


def rows(root: Path) -> dict[str, list[dict]]:
    public_axiom_source = json.dumps({
        "repository_or_url": "https://www.gamsgo.com/blog/hsr-apocalyptic-shadow-teams",
        "revision_or_access_date": "2026-08-11",
        "path_or_page": "Finality's Axiom and Ruinous Embers 4.4 tables",
        "quality": "ReleasedPublicCrossCheck",
    }, sort_keys=True, separators=(",", ":"))
    public_starward_source = json.dumps({
        "repository_or_url": "https://honkai-star-rail.fandom.com/wiki/Apocalyptic_Shadow",
        "revision_or_access_date": "2026-08-11",
        "path_or_page": "Starward Mode released tutorial text",
        "quality": "ReleasedPublicCrossCheck",
    }, sort_keys=True, separators=(",", ":"))
    all_stages = payloads(root, "ApsStages")
    stages = [row for row in all_stages if not row["tierce"]]
    tierce = next(row for row in all_stages if row["tierce"])
    all_nodes = payloads(root, "ApsNodes")
    nodes = [row for row in all_nodes if row["stage_id"] in LEVELS]
    objectives = payloads(root, "ApsObjectives")
    encounters = payloads(root, "ApsEncounters")
    slots = payloads(root, "ApsEnemySlots")
    axioms = payloads(root, "ApsAxioms")
    enemies = {row["monster_id"]: row for row in payloads(root, "ApsEnemies")}
    scores = payloads(root, "ApsScores")
    stable = stable_variants(root)

    profile = [{
        "id": 1,
        "stable_key": "profile.apocalyptic-shadow.v4.4",
        "game_version": "4.4",
        "initial_action_value_scaled": 2_000_000_000,
        "expiry": "Finalize",
        "boss_progress_maximum": 2_000,
        "action_value_score_maximum": 2_000,
        "source_ref": source_ref(scores[0]),
    }]
    stage_ids = {row["stage_id"]: index for index, row in enumerate(stages, start=1)}
    authored_stages = [{
        "id": stage_ids[row["stage_id"]],
        "profile_id": 1,
        "stable_key": row["id"],
        "upstream_stage_id": row["stage_id"],
        "floor": row["floor"],
        "source_ref": source_ref(row),
    } for row in stages]
    authored_stages.append({
        "id": 5,
        "profile_id": 1,
        "stable_key": "stage.30194.starward",
        "upstream_stage_id": tierce["stage_id"],
        "floor": 4,
        "source_ref": source_ref(stages[-1]) + "|" + source_ref(tierce),
    })
    axiom_ids = {
        group: "|".join(str(row["buff_id"]) for row in sorted(
            (row for row in axioms if row["option_group"] == group),
            key=lambda row: row["buff_id"],
        ))
        for group in (1, 2, 3)
    }
    authored_nodes = [{
        "id": index,
        "stage_id": stage_ids[row["stage_id"]],
        "stable_key": row["id"],
        "node_index": row["order"],
        "team_index": row["order"] - 1,
        "encounter_id": row["event_ids"][0],
        "maze_buff_id": row["maze_buff_id"],
        "axiom_bundle_ids": axiom_ids[row["order"]],
        "source_ref": source_ref(row),
    } for index, row in enumerate(nodes, start=1)]
    stage_four_nodes = sorted(
        (row for row in nodes if row["stage_id"] == stages[-1]["stage_id"]),
        key=lambda row: row["order"],
    )
    tierce_node = next(row for row in all_nodes if row["stage_id"] == tierce["stage_id"])
    for node_index, row in enumerate([*stage_four_nodes, tierce_node], start=1):
        authored_nodes.append({
            "id": len(authored_nodes) + 1,
            "stage_id": 5,
            "stable_key": f"node.30194.starward.{node_index}",
            "node_index": node_index,
            "team_index": node_index - 1,
            "encounter_id": row["event_ids"][0],
            "maze_buff_id": row["maze_buff_id"],
            "axiom_bundle_ids": axiom_ids[node_index],
            "source_ref": source_ref(row) + "|" + source_ref(tierce),
        })
    authored_objectives = [{
        "id": row["target_id"],
        "profile_id": 1,
        "stable_key": row["id"],
        "kind": "ScoreAtLeast",
        "threshold": row["threshold"],
        "source_ref": source_ref(row),
    } for row in objectives]

    encountered_monsters = sorted({row["monster_id"] for row in slots})
    enemy_ids = {monster: index for index, monster in enumerate(encountered_monsters, start=1)}
    authored_enemies = []
    for monster in encountered_monsters:
        row = enemies[monster]
        rank = "Elite" if monster == 3003015 else "Boss"
        authored_enemies.append({
            "id": enemy_ids[monster],
            "upstream_monster_id": monster,
            "stable_key": stable[monster],
            "behavior_source_key": DONORS[rank],
            "behavior_exact": False,
            "rank": rank,
            "weaknesses": "|".join(sorted(
                "Lightning" if value == "Thunder" else value
                for value in row["weaknesses"]
            )),
            "source_ref": source_ref(row),
        })

    ordinary_node_by_key = {
        row["stable_key"]: row for row in authored_nodes if not ".starward." in row["stable_key"]
    }
    ordinary_node_by_key[tierce_node["id"]] = authored_nodes[-1]
    authored_encounters = []
    encounter_ids = {}
    for row in encounters:
        node = ordinary_node_by_key[row["node_id"]]
        stage_id = int(row["node_id"].split(".")[1])
        encounter_ids[row["id"]] = node["encounter_id"]
        authored_encounters.append({
            "id": node["encounter_id"],
            "node_id": node["id"],
            "event_id": row["event_id"],
            "level": LEVELS.get(stage_id, LEVELS[30194]),
            "source_ref": source_ref(row),
        })

    authored_slots = []
    for index, row in enumerate(slots, start=1):
        stage_id = int(row["encounter_id"].split(".")[1])
        auxiliary = row["role"] == "auxiliary-scoring-boss"
        authored_slots.append({
            "id": index,
            "encounter_id": encounter_ids[row["encounter_id"]],
            "enemy_id": enemy_ids[row["monster_id"]],
            "formation_index": row["slot_order"] - 1,
            "score_included": True,
            "maximum_hp": 1_000_000 if auxiliary else PRIMARY_HP.get(stage_id, 6_000_000),
            "attack_scaled": (1_000 + (min(stage_id, 30194) - 30191) * 500) * 1_000_000,
            "defense_scaled": (800 + (min(stage_id, 30194) - 30191) * 100) * 1_000_000,
            "speed_scaled": 100_000_000,
            "toughness_scaled": (180 if auxiliary else 300) * 1_000_000,
            "source_ref": source_ref(row),
        })

    policy_ref = source_ref(scores[0])
    policies = [
        {
            "id": 1,
            "profile_id": 1,
            "stable_key": "policy.apocalyptic.score-postfix-lowering",
            "known_facts": (
                "Released structured data retains scoring items 90004/90005, the 2000-point "
                "constants, all-wave boss HP inputs and global Action Value updates. Public "
                "released guidance confirms up to 2000 boss-progress points plus remaining AV "
                "only after defeat. The opcode evaluator itself is not documented."
            ),
            "selected_behavior": (
                "Floor 2000 * depleted included-boss HP / included-boss maximum HP. On victory "
                "add floor remaining Action Value, clamped to 0..2000; otherwise add zero."
            ),
            "rejected_alternatives": (
                "claim undocumented postfix opcodes are exact|award remaining AV on timeout|"
                "round boss progress to nearest"
            ),
            "rationale": "This is the narrowest deterministic formula matching released constants and public behavior.",
            "affected_tests": "apocalyptic_shadow::tests::scores_terminal_boss_progress|apocalyptic_runtime::tests::aggregates_two_node_scores",
            "confidence": "High",
            "replacement_condition": "Replace when the released postfix evaluator and dynamic-hash mapping are independently decoded and fixture-tested.",
            "source_ref": policy_ref,
        },
        {
            "id": 2,
            "profile_id": 1,
            "stable_key": "policy.apocalyptic.boss-score-closure",
            "known_facts": "Candidate data identifies each primary boss and one Stage 30194 auxiliary-scoring-boss; the source program also contains boss-specific summon adjustments.",
            "selected_behavior": "Include every retained enemy unit authored as Boss rank in the score HP closure; the explicit auxiliary slot is authored as Boss.",
            "rejected_alternatives": "ignore auxiliary scoring bosses|infer undocumented summon IDs at runtime|include ordinary summons",
            "rationale": "Rank closure is stable, inspectable and includes the one exact active auxiliary binding without content-ID branches in Combat.",
            "affected_tests": "apocalyptic_shadow::tests::scores_terminal_boss_progress|challenge::tests::production_apocalyptic_profile_lowers",
            "confidence": "Medium",
            "replacement_condition": "Replace with typed per-encounter include/exclude selectors when every active summon adjustment is decoded.",
            "source_ref": source_ref(next(row for row in slots if row["role"] == "auxiliary-scoring-boss")),
        },
        {
            "id": 3,
            "profile_id": 1,
            "stable_key": "policy.apocalyptic.enemy-behavior-and-stats",
            "known_facts": "The Candidate package proves nine ordinary boss slots, exact variant identities, roles and weaknesses, but does not publish executable AI/phase programs or a node-local level/stat projection.",
            "selected_behavior": "Use explicit same-rank production behavior donors and deterministic floor-scaled placeholder stats: levels 65/75/85/95 and primary HP 0.5m/1m/2m/4m; the auxiliary has 1m HP.",
            "rejected_alternatives": "parse Candidate JSON at runtime|silently omit encounters|present placeholder AI or stats as observed parity",
            "rationale": "The mode remains runnable and every guessed value is centralized in typed authored data instead of hidden in Combat branches.",
            "affected_tests": "challenge::tests::production_apocalyptic_combat_definitions_lower|challenge_combat::tests::apocalyptic_catalog_composes_all_playable_encounters",
            "confidence": "Low",
            "replacement_condition": "Replace each row when released event-owned phase expansion, exact occurrence stats and reviewed ability programs are lowered.",
            "source_ref": source_ref(enemies[encountered_monsters[0]]),
        },
        {
            "id": 4,
            "profile_id": 1,
            "stable_key": "policy.apocalyptic.axioms-and-embers",
            "known_facts": "Released rows prove Ruinous Embers parameters 0.25/0.15 and nine selectable Axiom IDs/parameters. Released 4.4 public descriptions independently establish every trigger, target and cap.",
            "selected_behavior": "Execute all nine selectable Axioms and Ruinous Embers in mode-owned Rule IR. Unconditionally apply Knowledge and Decorum after selection because ResolvedCombatantSpec does not retain Path; approximate Oppose With Tenderness armor break as a boss Weakness Break; activate Energy-based Ultimates by setting Energy to maximum and do not invent non-Energy readiness resources.",
            "rejected_alternatives": "leave selected buffs inactive|add content IDs to Combat resolver branches|add build-catalog queries to Combat|invent boss-specific armor effects absent from placeholder AI",
            "rationale": "The exact released effects now execute through generic operations; three unsupported predicates are isolated, inspectable ProjectPolicy choices.",
            "affected_tests": "apocalyptic_mechanics::tests::active_definitions_cover_every_released_bundle|challenge_combat::tests::apocalyptic_catalog_composes_all_encounters|apocalyptic_runtime::tests::starward_runs_three_independent_nodes",
            "confidence": "Medium",
            "replacement_condition": "Replace the three approximations when combat specs retain Path, reviewed boss armor effects are executable, and non-Energy Ultimate readiness resources are authored.",
            "source_ref": source_ref(payloads(root, "ApsEmbers")[0]) + "|" + public_axiom_source,
        },
        {
            "id": 5,
            "profile_id": 1,
            "stable_key": "policy.apocalyptic.starward-composition",
            "known_facts": "Released Tierce data selects Stage 30195 after 30194, exposes one third node, and publishes aggregate thresholds 6000/7800/9900. Released public guidance confirms Starward Difficulty 4 has three nodes.",
            "selected_behavior": "Represent Starward as a fifth three-node stage containing both Difficulty-4 nodes followed by Tierce; use one independent team, battle, clock and 4000-point score closure per node.",
            "rejected_alternatives": "replace ordinary Difficulty 4|score Tierce alone against aggregate thresholds|share one AV clock between nodes",
            "rationale": "This preserves the exact predecessor relationship, node count, thresholds and ordinary node semantics.",
            "affected_tests": "challenge::tests::production_apocalyptic_profile_lowers|apocalyptic_runtime::tests::starward_runs_three_independent_nodes",
            "confidence": "High",
            "replacement_condition": "Replace only if released runtime evidence proves a different carry or aggregate rule.",
            "source_ref": source_ref(tierce) + "|" + public_starward_source,
        },
    ]
    return {
        "Profiles": profile,
        "Stages": authored_stages,
        "Nodes": authored_nodes,
        "Objectives": authored_objectives,
        "Policies": policies,
        "Enemies": authored_enemies,
        "Encounters": authored_encounters,
        "EnemySlots": authored_slots,
    }


def columns(sheet) -> list[str]:
    return [str(cell.value) for cell in sheet[3] if cell.value not in (None, "#field")]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    if args.output.exists():
        raise FileExistsError(f"refusing to overwrite {args.output}")
    workbook = openpyxl.load_workbook(args.template)
    authored = rows(args.root.resolve())
    if workbook.sheetnames != list(authored):
        raise ValueError(f"template sheets differ: {workbook.sheetnames}")
    for name, values in authored.items():
        sheet = workbook[name]
        if sheet.max_row >= 8:
            sheet.delete_rows(8, sheet.max_row - 7)
        headers = columns(sheet)
        for value in values:
            missing = set(headers) - set(value)
            if missing:
                raise ValueError(f"{name}: missing fields {sorted(missing)}")
            sheet.append([None, *[value[column] for column in headers]])
        sheet.freeze_panes = "A8"
        sheet.auto_filter.ref = f"A3:{sheet.cell(3, sheet.max_column).column_letter}{sheet.max_row}"
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Authored {args.output} with {sum(map(len, authored.values()))} typed rows")


if __name__ == "__main__":
    main()
