"""Deterministic openpyxl authoring and verification for Goal 17 workbooks."""

from __future__ import annotations

import hashlib
import json
import re
import zipfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

if openpyxl.__version__ != "3.1.5":
    raise RuntimeError(f"openpyxl 3.1.5 required, got {openpyxl.__version__}")

FIXED_TIME = datetime(2000, 1, 1, tzinfo=timezone.utc)
TITLE_FILL = PatternFill("solid", fgColor="172554")
HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
META_FILL = PatternFill("solid", fgColor="DBEAFE")
EVEN_FILL = PatternFill("solid", fgColor="F8FAFC")
ODD_FILL = PatternFill("solid", fgColor="FFFFFF")
POLICY_FILL = PatternFill("solid", fgColor="FEF3C7")
TITLE_FONT = Font(name="Aptos Display", color="FFFFFF", bold=True)
HEADER_FONT = Font(name="Aptos", color="FFFFFF", bold=True)
META_FONT = Font(name="Aptos", color="1E3A8A", italic=True, size=9)
BODY_FONT = Font(name="Aptos", size=10)
THIN_BORDER = Border(bottom=Side(style="thin", color="CBD5E1"))
WORKBOOKS = (
    "MemoryOfChaos.xlsx",
    "MemoryOfChaosBindings.xlsx",
    "MemoryOfChaosReview.xlsx",
)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def compact(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def sheet_name(file_name: str) -> str:
    parts = re.sub(r"\.json$", "", file_name).split("-")
    return "".join(part[:1].upper() + part[1:] for part in parts)[:31]


def layout(root: Path) -> dict[str, list[str]]:
    contract = read_json(
        root / "content-manifests/memory-of-chaos-v1/authoring-contract.json"
    )
    result = {name: [] for name in WORKBOOKS}
    bindings = sorted(
        contract["normalized_family_bindings"].items(),
        key=lambda item: item[1]["order"],
    )
    for file_name, binding in bindings:
        result[binding["workbook"]].append(file_name)
    if list(result) != list(WORKBOOKS) or sum(map(len, result.values())) != 27:
        raise ValueError("workbook partition drift")
    return result


def normalized_rows(root: Path, file_name: str) -> list[dict[str, Any]]:
    value = read_json(root / "content-reference/memory-of-chaos-v1" / file_name)
    if value["file"] != file_name or not isinstance(value["records"], list):
        raise ValueError(f"{file_name}: normalized envelope drift")
    return value["records"]


def id_maps(root: Path) -> dict[str, dict[str, int]]:
    return {
        file_name: {
            row["id"]: index
            for index, row in enumerate(normalized_rows(root, file_name), start=1)
        }
        for files in layout(root).values()
        for file_name in files
    }


def strongest_quality(row: dict[str, Any]) -> str:
    if isinstance(row.get("evidence_quality"), str):
        return row["evidence_quality"]
    if isinstance(row.get("quality"), str):
        return row["quality"]
    refs = row.get("evidence_refs", [])
    return refs[0]["quality"] if refs else "ExactStructured"


def mechanism_quality(row: dict[str, Any]) -> str:
    if isinstance(row.get("mechanism_quality"), str):
        return row["mechanism_quality"]
    refs = row.get("evidence_refs", [])
    return refs[0]["mechanism_quality"] if refs else "ExactRelationship"


def relation_values(
    file_name: str,
    row: dict[str, Any],
    maps: dict[str, dict[str, int]],
) -> dict[str, int]:
    values: dict[str, int] = {}
    if file_name != "profile.json":
        values["profile_id"] = maps["profile.json"]["profile.memory-of-chaos"]
    if file_name == "nodes.json":
        values["stage_id"] = maps["stages.json"][row["stage_id"]]
    elif file_name == "waves.json":
        values["encounter_id"] = maps["encounters.json"][row["encounter_id"]]
    elif file_name == "enemy-slots.json":
        values["wave_id"] = maps["waves.json"][row["wave_id"]]
        values["enemy_variant_id"] = maps["enemy-variants.json"][row["enemy_variant_id"]]
    elif file_name == "enemy-variants.json":
        values["enemy_template_id"] = maps["enemy-templates.json"][row["enemy_template_id"]]
    elif file_name == "enemy-abilities.json":
        values["enemy_template_id"] = maps["enemy-templates.json"][row["enemy_template_id"]]
    return values


def authored_rows(
    root: Path,
    file_name: str,
    maps: dict[str, dict[str, int]],
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for private_id, row in enumerate(normalized_rows(root, file_name), start=1):
        values: dict[str, Any] = {
            "id": private_id,
            "stable_key": row["id"],
            "row_order": private_id,
            "name_en": row["name_en"],
            "name_zh_cn": row["name_zh_cn"],
            "summary_en": row["summary_en"],
            "summary_zh_cn": row["summary_zh_cn"],
            "ownership": row["ownership"],
            "coverage_state": row["coverage_state"],
            "evidence_quality": strongest_quality(row),
            "mechanism_quality": mechanism_quality(row),
            "manifest_record_ids": "|".join(row["source_record_ids"]) or None,
            "source_ref_ids": "|".join(
                reference["id"] for reference in row["evidence_refs"]
            ) or None,
            "tags": "|".join(row["tags"]) or None,
            "payload_json": compact(row),
            "runtime_executable": False,
            **relation_values(file_name, row, maps),
        }
        for field, value in values.items():
            if isinstance(value, str) and len(value) > 32767:
                raise ValueError(
                    f"{file_name}/{row['id']}/{field}: Excel cell limit exceeded"
                )
        result.append(values)
    return result


def columns(sheet) -> list[str]:
    return [
        str(cell.value)
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    ]


def add_validation(sheet, field: str, values: list[str], maximum_row: int) -> None:
    by_name = {str(cell.value): cell.column for cell in sheet[3] if cell.value}
    column = by_name.get(field)
    if column is None:
        return
    validation = DataValidation(
        type="list", formula1=f'"{",".join(values)}"', allow_blank=False
    )
    validation.error = f"Choose one of: {', '.join(values)}"
    validation.errorTitle = f"Invalid {field}"
    validation.showErrorMessage = True
    sheet.add_data_validation(validation)
    validation.add(
        f"{sheet.cell(8, column).coordinate}:"
        f"{sheet.cell(max(8, maximum_row), column).coordinate}"
    )


def style_sheet(sheet, row_count: int) -> None:
    maximum_row = max(7 + row_count, 7)
    maximum_column = sheet.max_column
    for row_number in range(1, 8):
        sheet.row_dimensions[row_number].height = 22
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[3].height = 34
    for cell in sheet[1]:
        cell.fill = copy(TITLE_FILL)
        cell.font = copy(TITLE_FONT)
        cell.alignment = Alignment(vertical="center")
    for row_number in (2, 4, 5, 6, 7):
        for cell in sheet[row_number]:
            cell.fill = copy(META_FILL)
            cell.font = copy(META_FONT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in sheet[3]:
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=8, max_row=maximum_row, max_col=maximum_column):
        fill = EVEN_FILL if row[0].row % 2 == 0 else ODD_FILL
        sheet.row_dimensions[row[0].row].height = 38
        for cell in row:
            cell.fill = copy(fill)
            cell.border = copy(THIN_BORDER)
            cell.font = copy(BODY_FONT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A8"
    last_column = sheet.cell(3, maximum_column).column_letter
    sheet.auto_filter.ref = f"A3:{last_column}{maximum_row}"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    for column in range(1, maximum_column + 1):
        header = str(sheet.cell(3, column).value or "")
        samples = [
            str(sheet.cell(row, column).value or "")
            for row in range(1, min(maximum_row, 107) + 1)
        ]
        cap = 64 if header in {"payload_json", "summary_en", "summary_zh_cn"} else 44
        width = min(cap, max(11, max((min(len(value), cap - 2) for value in samples), default=11) + 2))
        sheet.column_dimensions[sheet.cell(3, column).column_letter].width = width
    add_validation(sheet, "ownership", ["MemoryOfChaos", "Shared"], maximum_row)
    add_validation(sheet, "coverage_state", ["DataReady"], maximum_row)
    by_name = {str(cell.value): cell.column for cell in sheet[3] if cell.value}
    quality_column = by_name.get("mechanism_quality")
    if quality_column and row_count:
        letter = sheet.cell(8, quality_column).column_letter
        sheet.conditional_formatting.add(
            f"{letter}8:{letter}{maximum_row}",
            FormulaRule(formula=[f'ISNUMBER(SEARCH("Policy",{letter}8))'], fill=copy(POLICY_FILL)),
        )


def normalize_archive(file: Path) -> None:
    temporary = file.with_suffix(f"{file.suffix}.canonical")
    with zipfile.ZipFile(file, "r") as source:
        members = [(name, source.read(name)) for name in sorted(source.namelist())]
    with zipfile.ZipFile(temporary, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as target:
        for name, payload in members:
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
                    rb"\g<1>2000-01-01T00:00:00Z\g<2>",
                    payload,
                )
            info = zipfile.ZipInfo(name, date_time=(2000, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0
            target.writestr(info, payload, compress_type=zipfile.ZIP_DEFLATED, compresslevel=9)
    temporary.replace(file)


def author(root: Path, output: Path, templates: Path) -> dict[str, int]:
    output.mkdir(parents=True, exist_ok=True)
    expected_layout = layout(root)
    existing = [output / name for name in WORKBOOKS if (output / name).exists()]
    if existing:
        raise FileExistsError("refusing to overwrite authored workbook(s): " + ", ".join(map(str, existing)))
    maps = id_maps(root)
    counts: dict[str, int] = {}
    for workbook_name, files in expected_layout.items():
        template = templates / workbook_name
        workbook = load_workbook(template, data_only=False)
        expected_sheets = [sheet_name(file_name) for file_name in files]
        if workbook.sheetnames != expected_sheets:
            raise ValueError(f"{workbook_name}: template sheet partition drift")
        for file_name in files:
            sheet = workbook[sheet_name(file_name)]
            rows = authored_rows(root, file_name, maps)
            template_columns = columns(sheet)
            if sheet.max_row != 7:
                raise ValueError(f"{workbook_name}/{sheet.title}: template contains data")
            for values in rows:
                missing = set(template_columns) - set(values)
                if missing:
                    raise ValueError(f"{file_name}: missing authored fields {sorted(missing)}")
                sheet.append([None, *[values[column] for column in template_columns]])
            style_sheet(sheet, len(rows))
            counts[f"{workbook_name}/{sheet.title}"] = len(rows)
        workbook.properties.creator = "Starclock Goal 17 openpyxl authoring"
        workbook.properties.lastModifiedBy = "Starclock Goal 17 openpyxl authoring"
        workbook.properties.created = FIXED_TIME
        workbook.properties.modified = FIXED_TIME
        workbook.calculation.fullCalcOnLoad = False
        workbook.calculation.forceFullCalc = False
        target = output / workbook_name
        workbook.save(target)
        workbook.close()
        normalize_archive(target)
    verify(root, output, templates, counts)
    return counts


def verify(
    root: Path,
    directory: Path,
    templates: Path,
    authored_counts: dict[str, int] | None = None,
) -> dict[str, int]:
    expected_layout = layout(root)
    maps = id_maps(root)
    observed: dict[str, int] = {}
    for workbook_name, files in expected_layout.items():
        template = load_workbook(templates / workbook_name, read_only=True, data_only=False)
        workbook = load_workbook(directory / workbook_name, data_only=False)
        expected_sheets = [sheet_name(file_name) for file_name in files]
        if workbook.sheetnames != expected_sheets or template.sheetnames != expected_sheets:
            raise ValueError(f"{workbook_name}: sheet order drift")
        for file_name, sheet_title in zip(files, expected_sheets):
            sheet = workbook[sheet_title]
            template_sheet = template[sheet_title]
            template_columns = columns(template_sheet)
            expected_rows = authored_rows(root, file_name, maps)
            for row_number in range(1, 8):
                actual = [sheet.cell(row_number, column).value for column in range(1, sheet.max_column + 1)]
                expected = [template_sheet.cell(row_number, column).value for column in range(1, template_sheet.max_column + 1)]
                if actual != expected:
                    raise ValueError(f"{workbook_name}/{sheet_title}: metadata row {row_number} drift")
            key = f"{workbook_name}/{sheet_title}"
            observed[key] = len(expected_rows)
            if sheet.max_row != 7 + len(expected_rows):
                raise ValueError(f"{key}: row count drift")
            for row_number, expected in enumerate(expected_rows, start=8):
                actual = [sheet.cell(row_number, column).value for column in range(2, len(template_columns) + 2)]
                expected_values = [expected[column] for column in template_columns]
                if actual != expected_values:
                    raise ValueError(f"{key}/row {row_number}: semantic drift")
            if sheet.freeze_panes != "A8" or not sheet.auto_filter.ref:
                raise ValueError(f"{key}: authoring affordances missing")
            for row in sheet.iter_rows(min_row=8):
                for cell in row:
                    if cell.data_type in (TYPE_FORMULA, TYPE_ERROR):
                        raise ValueError(f"{key}/{cell.coordinate}: formula/error forbidden")
                    if isinstance(cell.value, str) and len(cell.value) > 32767:
                        raise ValueError(f"{key}/{cell.coordinate}: Excel text overflow")
        workbook.close()
        template.close()
    if authored_counts is not None and authored_counts != observed:
        raise ValueError("workbook counts changed after reload")
    return observed


def semantic_digest(directory: Path) -> str:
    result = hashlib.sha256()
    for workbook_name in WORKBOOKS:
        workbook = load_workbook(directory / workbook_name, read_only=True, data_only=False)
        result.update(f"{workbook_name}\n".encode())
        for sheet in workbook.worksheets:
            result.update(f"{sheet.title}\n".encode())
            for row in sheet.iter_rows(values_only=True):
                result.update((compact(list(row)) + "\n").encode("utf-8"))
        workbook.close()
    return result.hexdigest()


def byte_digests(directory: Path) -> dict[str, str]:
    return {
        name: hashlib.sha256((directory / name).read_bytes()).hexdigest()
        for name in WORKBOOKS
    }
