"""Deterministic openpyxl authoring and QA for Goal 12 workbooks."""

from __future__ import annotations

import hashlib
import json
import re
import zipfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKBOOKS = (
    "CurrencyWars.xlsx",
    "CurrencyWarsBindings.xlsx",
    "CurrencyWarsReview.xlsx",
)
FIXED_TIME = datetime(2000, 1, 1, tzinfo=timezone.utc)
HEADER_FILL = PatternFill("solid", fgColor="17365D")
EVEN_FILL = PatternFill("solid", fgColor="EAF2F8")
ODD_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(name="Aptos", color="FFFFFF", bold=True)
BODY_FONT = Font(name="Aptos", size=10)
THIN_BORDER = Border(bottom=Side(style="thin", color="B7C9E2"))
QUALITY_FILL = PatternFill("solid", fgColor="FFF2CC")


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def compact(value: Any) -> str:
    if isinstance(value, str):
        return value
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def source_key(value: dict[str, Any]) -> str:
    return compact([
        value["repository"],
        value["revision"],
        value["path"],
        value["locator"],
        value["sha256"],
        value["evidence_quality"],
    ])


def sheet_name(file_name: str) -> str:
    parts = re.sub(r"\.json$", "", file_name).split("-")
    return "".join(part[:1].upper() + part[1:] for part in parts)[:31]


def contracts(root: Path) -> list[dict[str, Any]]:
    schema = read_json(
        root / "content-manifests/currency-wars-v1/normalized-schema.json"
    )
    return schema["files"]


def tables(root: Path) -> list[dict[str, Any]]:
    lock = read_json(root / "config/currency-wars-generated/schema.lock")
    return lock["schema"]["tables"]


def source_id_map(root: Path) -> dict[str, str]:
    result: dict[str, str] = {}
    for row in read_json(root / "content-reference/currency-wars-v1/sources.json"):
        key = source_key(row)
        if key in result:
            raise ValueError(f"duplicate source key for {row['id']}")
        result[key] = row["id"]
    return result


def normalized_rows(
    root: Path,
    contract: dict[str, Any],
    source_ids: dict[str, str],
) -> list[dict[str, Any]]:
    source = root / "content-reference/currency-wars-v1" / contract["file"]
    rows = read_json(source) if source.exists() else []
    authored: list[dict[str, Any]] = []
    for private_id, row in enumerate(rows, start=1):
        refs = []
        for reference in row["source_refs"]:
            key = source_key(reference)
            if key not in source_ids:
                raise ValueError(f"{contract['file']}/{row['id']}: missing source")
            refs.append(source_ids[key])
        values: dict[str, Any] = {
            "id": private_id,
            "stable_key": row["id"],
            "schema_revision": row["schema_revision"],
            "kind": row["kind"],
            "name_en": row["name_en"],
            "name_zh_cn": row["name_zh_cn"],
            "summary_en": row["summary_en"],
            "summary_zh_cn": row["summary_zh_cn"],
            "ownership": row["ownership"],
            "coverage_state": row["coverage_state"],
            "evidence_quality": row["evidence_quality"],
            "source_refs_json": compact(refs),
            "tags_json": compact(row["tags"]),
        }
        for field in contract["required_domain_fields"]:
            values[field] = compact(row[field])
        for field, value in values.items():
            if isinstance(value, str) and len(value) > 32767:
                raise ValueError(
                    f"{contract['file']}/{row['id']}/{field}: "
                    f"{len(value)} exceeds Excel cell limit"
                )
        authored.append(values)
    return authored


def field_columns(sheet) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }


def metadata_values(sheet) -> list[list[Any]]:
    return [
        [cell.value for cell in row]
        for row in sheet.iter_rows(
            min_row=1,
            max_row=7,
            max_col=sheet.max_column,
        )
    ]


def write_rows(sheet, rows: list[dict[str, Any]]) -> None:
    columns = field_columns(sheet)
    for offset, values in enumerate(rows, start=8):
        unknown = sorted(set(values) - set(columns))
        if unknown:
            raise ValueError(f"{sheet.title}: unknown fields {unknown}")
        for field, value in values.items():
            sheet.cell(row=offset, column=columns[field], value=value)


def add_validation(sheet, field: str, values: list[str], maximum_row: int) -> None:
    column = field_columns(sheet).get(field)
    if column is None:
        return
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(values)}"',
        allow_blank=False,
    )
    validation.error = f"Choose one of: {', '.join(values)}"
    validation.errorTitle = f"Invalid {field}"
    sheet.add_data_validation(validation)
    validation.add(
        f"{sheet.cell(row=8, column=column).coordinate}:"
        f"{sheet.cell(row=max(8, maximum_row), column=column).coordinate}"
    )


def style_sheet(sheet, row_count: int) -> None:
    maximum_row = max(7 + row_count, 7)
    maximum_column = sheet.max_column
    for row_number in range(1, 8):
        sheet.row_dimensions[row_number].height = 22
    sheet.row_dimensions[3].height = 30
    for cell in sheet[3]:
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    if row_count:
        for row in sheet.iter_rows(
            min_row=8,
            max_row=maximum_row,
            max_col=maximum_column,
        ):
            fill = EVEN_FILL if row[0].row % 2 == 0 else ODD_FILL
            sheet.row_dimensions[row[0].row].height = 30
            for cell in row:
                cell.fill = copy(fill)
                cell.border = copy(THIN_BORDER)
                cell.font = copy(BODY_FONT)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A8"
    last_column = sheet.cell(row=3, column=maximum_column).column_letter
    sheet.auto_filter.ref = f"A3:{last_column}{maximum_row}"
    sheet.sheet_view.showGridLines = False
    for column in range(1, maximum_column + 1):
        samples = [
            str(sheet.cell(row=row, column=column).value or "")
            for row in range(2, min(maximum_row, 107) + 1)
        ]
        width = min(
            50,
            max(10, max((min(len(value), 48) for value in samples), default=10) + 2),
        )
        letter = sheet.cell(row=3, column=column).column_letter
        sheet.column_dimensions[letter].width = width
    add_validation(
        sheet,
        "ownership",
        ["CurrencyWars", "Shared"],
        maximum_row,
    )
    add_validation(
        sheet,
        "coverage_state",
        ["Cataloged", "Researched", "DataReady", "Blocked"],
        maximum_row,
    )
    add_validation(
        sheet,
        "evidence_quality",
        [
            "ExactStructured",
            "ExactPublicText",
            "Observed",
            "ApproximateFromReleasedText",
            "ProjectPolicy",
        ],
        maximum_row,
    )
    quality_column = field_columns(sheet).get("evidence_quality")
    if quality_column is not None and row_count:
        letter = sheet.cell(row=8, column=quality_column).column_letter
        target = f"{letter}8:{letter}{maximum_row}"
        sheet.conditional_formatting.add(
            target,
            FormulaRule(
                formula=[
                    f'OR({letter}8="ProjectPolicy",'
                    f'{letter}8="ApproximateFromReleasedText")'
                ],
                fill=copy(QUALITY_FILL),
            ),
        )


def normalize_archive(file: Path) -> None:
    temporary = file.with_suffix(f"{file.suffix}.canonical")
    with zipfile.ZipFile(file, "r") as source:
        members = [(name, source.read(name)) for name in sorted(source.namelist())]
    with zipfile.ZipFile(
        temporary,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as target:
        for name, payload in members:
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2000-01-01T00:00:00Z\g<2>",
                    payload,
                )
            info = zipfile.ZipInfo(name, date_time=(2000, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0
            target.writestr(
                info,
                payload,
                compress_type=zipfile.ZIP_DEFLATED,
                compresslevel=9,
            )
    temporary.replace(file)


def author(root: Path, output: Path) -> dict[str, int]:
    output.mkdir(parents=True, exist_ok=True)
    targets = [output / name for name in WORKBOOKS]
    existing = [target for target in targets if target.exists()]
    if existing:
        raise FileExistsError(
            "refusing to overwrite authored workbook(s): "
            + ", ".join(map(str, existing))
        )
    table_by_sheet = {
        (table["source"]["file"], table["source"]["sheet"]): table
        for table in tables(root)
    }
    contract_by_sheet = {
        sheet_name(contract["file"]): contract
        for contract in contracts(root)
    }
    source_ids = source_id_map(root)
    counts: dict[str, int] = {}
    template_root = root / "config/currency-wars-generated/templates"
    for workbook_name in WORKBOOKS:
        template = template_root / workbook_name
        workbook = load_workbook(template)
        template_metadata = {
            sheet.title: metadata_values(sheet)
            for sheet in workbook.worksheets
        }
        for sheet in workbook.worksheets:
            if (workbook_name, sheet.title) not in table_by_sheet:
                raise ValueError(f"{workbook_name}/{sheet.title}: schema missing")
            contract = contract_by_sheet.get(sheet.title)
            if contract is None:
                raise ValueError(f"{workbook_name}/{sheet.title}: contract missing")
            rows = normalized_rows(root, contract, source_ids)
            write_rows(sheet, rows)
            style_sheet(sheet, len(rows))
            if metadata_values(sheet) != template_metadata[sheet.title]:
                raise ValueError(f"{workbook_name}/{sheet.title}: metadata mutated")
            counts[f"{workbook_name}/{sheet.title}"] = len(rows)
        workbook.properties.creator = "Starclock Goal 12 openpyxl authoring"
        workbook.properties.lastModifiedBy = "Starclock Goal 12 openpyxl authoring"
        workbook.properties.created = FIXED_TIME
        workbook.properties.modified = FIXED_TIME
        workbook.calculation.fullCalcOnLoad = False
        workbook.calculation.forceFullCalc = False
        target = output / workbook_name
        workbook.save(target)
        workbook.close()
        normalize_archive(target)
    verify(root, output, counts)
    return counts


def expected_counts(root: Path) -> dict[str, int]:
    authoring = read_json(
        root / "content-manifests/currency-wars-v1/authoring-contract.json"
    )
    result: dict[str, int] = {}
    for workbook in authoring["workbooks"]:
        for file_name in workbook["normalized_files"]:
            source = root / "content-reference/currency-wars-v1" / file_name
            count = len(read_json(source)) if source.exists() else 0
            result[
                f"{workbook['file']}/{sheet_name(file_name)}"
            ] = count
    return result


def verify(
    root: Path,
    directory: Path,
    authored_counts: dict[str, int] | None = None,
) -> dict[str, int]:
    expected = expected_counts(root)
    template_root = root / "config/currency-wars-generated/templates"
    observed: dict[str, int] = {}
    for workbook_name in WORKBOOKS:
        template = load_workbook(template_root / workbook_name, read_only=True)
        workbook = load_workbook(directory / workbook_name, data_only=False)
        if workbook.sheetnames != template.sheetnames:
            raise ValueError(f"{workbook_name}: missing or reordered sheet")
        for sheet_name_value in workbook.sheetnames:
            sheet = workbook[sheet_name_value]
            template_sheet = template[sheet_name_value]
            if metadata_values(sheet) != metadata_values(template_sheet):
                raise ValueError(
                    f"{workbook_name}/{sheet_name_value}: Sora metadata drift"
                )
            key = f"{workbook_name}/{sheet_name_value}"
            count = expected[key]
            observed[key] = count
            actual_rows = max(0, sheet.max_row - 7)
            if actual_rows != count:
                raise ValueError(
                    f"{key}: expected {count} rows, got {actual_rows}"
                )
            if sheet.freeze_panes != "A8" or not sheet.auto_filter.ref:
                raise ValueError(f"{key}: authoring affordances missing")
            columns = field_columns(sheet)
            if count:
                stable_column = columns["stable_key"]
                id_column = columns["id"]
                ids = [
                    sheet.cell(row=row, column=id_column).value
                    for row in range(8, 8 + count)
                ]
                if ids != list(range(1, count + 1)):
                    raise ValueError(f"{key}: private ID sequence drift")
                contract_file = next(
                    contract["file"]
                    for contract in contracts(root)
                    if sheet_name(contract["file"]) == sheet_name_value
                )
                normalized = read_json(
                    root / "content-reference/currency-wars-v1" / contract_file
                )
                stable_keys = [
                    sheet.cell(row=row, column=stable_column).value
                    for row in range(8, 8 + count)
                ]
                if stable_keys != [row["id"] for row in normalized]:
                    raise ValueError(f"{key}: stable-key order drift")
            for cell in sheet[3]:
                if cell.value is not None and not cell.alignment.wrap_text:
                    raise ValueError(f"{key}/{cell.coordinate}: header wrap missing")
            for column in range(1, sheet.max_column + 1):
                letter = sheet.cell(row=3, column=column).column_letter
                width = sheet.column_dimensions[letter].width
                if width is None or not 10 <= width <= 50:
                    raise ValueError(f"{key}/{letter}: width {width} invalid")
            for row in sheet.iter_rows(min_row=8):
                for cell in row:
                    if cell.data_type in (TYPE_FORMULA, TYPE_ERROR):
                        raise ValueError(
                            f"{key}/{cell.coordinate}: formula/error forbidden"
                        )
                    if isinstance(cell.value, str) and len(cell.value) > 32767:
                        raise ValueError(
                            f"{key}/{cell.coordinate}: Excel text overflow"
                        )
        template.close()
        workbook.close()
    if observed != expected:
        raise ValueError("workbook count closure drift")
    if authored_counts is not None and observed != authored_counts:
        raise ValueError("workbook counts changed after save/reload")
    return observed


def semantic_digest(directory: Path) -> str:
    digest = hashlib.sha256()
    for workbook_name in WORKBOOKS:
        workbook = load_workbook(
            directory / workbook_name,
            read_only=True,
            data_only=False,
        )
        digest.update(f"{workbook_name}\n".encode())
        for sheet in workbook.worksheets:
            digest.update(f"{sheet.title}\n".encode())
            for row in sheet.iter_rows(values_only=True):
                digest.update(
                    (compact(list(row)) + "\n").encode("utf-8")
                )
        workbook.close()
    return digest.hexdigest()
