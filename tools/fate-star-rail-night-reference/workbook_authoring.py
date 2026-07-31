"""Deterministic openpyxl authoring and verification for Goal 19."""

from __future__ import annotations

import json
import re
import subprocess
import tempfile
import zipfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

if openpyxl.__version__ != "3.1.5":
    raise RuntimeError(f"openpyxl 3.1.5 required, got {openpyxl.__version__}")

FIXED_TIME = datetime(2000, 1, 1, tzinfo=timezone.utc)
TITLE_FILL = PatternFill("solid", fgColor="243447")
META_FILL = PatternFill("solid", fgColor="E8EEF4")
HEADER_FILL = PatternFill("solid", fgColor="2F6F8F")
EVEN_FILL = PatternFill("solid", fgColor="F7FAFC")
ODD_FILL = PatternFill("solid", fgColor="FFFFFF")
POLICY_FILL = PatternFill("solid", fgColor="FFF2CC")
EVIDENCE_FILL = PatternFill("solid", fgColor="E2F0D9")
WHITE_FONT = Font(name="Aptos", size=10, color="FFFFFF", bold=True)
META_FONT = Font(name="Aptos", size=9, color="354052")
BODY_FONT = Font(name="Aptos", size=9, color="20262E")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9E1E8"))


def normalize_archive(file: Path) -> None:
    temporary = file.with_suffix(f"{file.suffix}.canonical")
    with zipfile.ZipFile(file, "r") as source:
        members = [(name, source.read(name)) for name in sorted(source.namelist())]
    with zipfile.ZipFile(temporary, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as target:
        for name, payload in members:
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
                    rb"\g<1>2000-01-01T00:00:00Z\g<2>",
                    payload,
                )
            info = zipfile.ZipInfo(name, date_time=(2000, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0
            target.writestr(info, payload, compress_type=zipfile.ZIP_DEFLATED, compresslevel=9)
    temporary.replace(file)


def load_data(root: Path) -> dict:
    with tempfile.TemporaryDirectory(prefix="g19-workbook-data-") as temporary:
        target = Path(temporary) / "workbooks.json"
        subprocess.run(
            [
                "node",
                str(root / "tools/fate-star-rail-night-reference/workbook-data.mjs"),
                "--root",
                str(root),
                "--output",
                str(target),
            ],
            cwd=root,
            check=True,
            stdout=subprocess.DEVNULL,
        )
        return json.loads(target.read_text(encoding="utf-8"))


def add_validation(sheet, field: str, values: list[str], maximum_row: int) -> None:
    columns = {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }
    column = columns.get(field)
    if column is None:
        return
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(values)}"',
        allow_blank=False,
    )
    validation.error = f"Choose one of: {', '.join(values)}"
    validation.errorTitle = f"Invalid {field}"
    validation.prompt = f"Goal 19 {field}"
    validation.promptTitle = "Starclock authoring contract"
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(
        f"{sheet.cell(row=8, column=column).coordinate}:"
        f"{sheet.cell(row=max(8, maximum_row), column=column).coordinate}"
    )


def style_sheet(sheet, row_count: int) -> None:
    maximum_row = max(7 + row_count, 7)
    maximum_column = sheet.max_column
    for row_number in range(1, 8):
        sheet.row_dimensions[row_number].height = 22
    sheet.row_dimensions[1].height = 27
    sheet.row_dimensions[3].height = 32
    for cell in sheet[1]:
        cell.fill = copy(TITLE_FILL)
        cell.font = copy(WHITE_FONT)
        cell.alignment = Alignment(vertical="center")
    for row_number in (2, 4, 5, 6, 7):
        for cell in sheet[row_number]:
            cell.fill = copy(META_FILL)
            cell.font = copy(META_FONT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in sheet[3]:
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(WHITE_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=8, max_row=maximum_row, max_col=maximum_column):
        fill = EVEN_FILL if row[0].row % 2 == 0 else ODD_FILL
        sheet.row_dimensions[row[0].row].height = 30
        for cell in row:
            cell.fill = copy(fill)
            cell.border = copy(THIN_BORDER)
            cell.font = copy(BODY_FONT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "B8"
    last_column = sheet.cell(row=3, column=maximum_column).column_letter
    sheet.auto_filter.ref = f"A3:{last_column}{maximum_row}"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    for column in range(1, maximum_column + 1):
        samples = [
            str(sheet.cell(row=row, column=column).value or "")
            for row in range(1, min(maximum_row, 107) + 1)
        ]
        width = min(48, max(11, max((min(len(value), 46) for value in samples), default=11) + 2))
        sheet.column_dimensions[sheet.cell(row=3, column=column).column_letter].width = width
    add_validation(sheet, "ownership", ["FateStarRailNight", "Shared", "EvidenceOnly"], maximum_row)
    add_validation(sheet, "disposition", ["DataReady", "EvidenceOnly", "ResearchRequired"], maximum_row)
    add_validation(sheet, "enabled", ["true", "false"], maximum_row)
    columns = {str(cell.value): cell.column for cell in sheet[3] if cell.value}
    quality_column = columns.get("evidence_quality")
    disposition_column = columns.get("disposition")
    if quality_column and row_count:
        letter = sheet.cell(row=8, column=quality_column).column_letter
        sheet.conditional_formatting.add(
            f"{letter}8:{letter}{maximum_row}",
            FormulaRule(formula=[f'{letter}8="ProjectPolicy"'], fill=copy(POLICY_FILL)),
        )
    if disposition_column and row_count:
        letter = sheet.cell(row=8, column=disposition_column).column_letter
        sheet.conditional_formatting.add(
            f"{letter}8:{letter}{maximum_row}",
            FormulaRule(formula=[f'{letter}8="EvidenceOnly"'], fill=copy(EVIDENCE_FILL)),
        )


def author(root: Path, output: Path, templates: Path) -> dict[str, int]:
    data = load_data(root)
    targets = [output / name for name in data["workbooks"]]
    existing = [target for target in targets if target.exists()]
    if existing:
        raise FileExistsError("refusing to overwrite authored workbook(s): " + ", ".join(map(str, existing)))
    output.mkdir(parents=True, exist_ok=True)
    counts: dict[str, int] = {}
    columns = data["columns"]
    for workbook_name, tables in data["workbooks"].items():
        template = templates / workbook_name
        workbook = load_workbook(template, data_only=False)
        expected_sheets = [table["sheet"] for table in tables]
        if workbook.sheetnames != expected_sheets:
            raise ValueError(f"{workbook_name}: Sora template sheet partition drift")
        for table in tables:
            sheet = workbook[table["sheet"]]
            template_fields = [cell.value for cell in sheet[3] if cell.value is not None]
            if template_fields != ["#field", *columns]:
                raise ValueError(f"{workbook_name}/{sheet.title}: Sora template field drift")
            if sheet.max_row != 7:
                raise ValueError(f"{workbook_name}/{sheet.title}: Sora template contains authored rows")
            for row in table["rows"]:
                values = [row[column] for column in columns]
                for field, value in zip(columns, values):
                    if isinstance(value, str) and len(value) > 32767:
                        raise ValueError(f"{sheet.title}/{row['stable_key']}/{field}: Excel cell limit")
                sheet.append([None, *values])
            style_sheet(sheet, len(table["rows"]))
            counts[f"{workbook_name}/{sheet.title}"] = len(table["rows"])
        workbook.properties.creator = "Starclock Goal 19 openpyxl authoring"
        workbook.properties.lastModifiedBy = "Starclock Goal 19 openpyxl authoring"
        workbook.properties.created = FIXED_TIME
        workbook.properties.modified = FIXED_TIME
        workbook.calculation.fullCalcOnLoad = False
        workbook.calculation.forceFullCalc = False
        target = output / workbook_name
        workbook.save(target)
        workbook.close()
        normalize_archive(target)
    verify(root, output, templates)
    return counts


def verify(root: Path, directory: Path, templates: Path) -> dict[str, int]:
    data = load_data(root)
    columns = data["columns"]
    counts: dict[str, int] = {}
    for workbook_name, tables in data["workbooks"].items():
        workbook = load_workbook(directory / workbook_name, data_only=False)
        template = load_workbook(templates / workbook_name, read_only=True, data_only=False)
        expected_sheets = [table["sheet"] for table in tables]
        if workbook.sheetnames != expected_sheets or template.sheetnames != expected_sheets:
            raise ValueError(f"{workbook_name}: missing or reordered sheet")
        for table in tables:
            sheet = workbook[table["sheet"]]
            template_sheet = template[table["sheet"]]
            for row_number in range(1, 8):
                observed = [sheet.cell(row_number, column).value for column in range(1, len(columns) + 2)]
                expected = [template_sheet.cell(row_number, column).value for column in range(1, len(columns) + 2)]
                if observed != expected:
                    raise ValueError(f"{workbook_name}/{sheet.title}: metadata row {row_number} drift")
            if sheet.max_row != 7 + len(table["rows"]):
                raise ValueError(f"{workbook_name}/{sheet.title}: row denominator drift")
            for offset, expected_row in enumerate(table["rows"], start=8):
                observed = [sheet.cell(offset, column).value for column in range(2, len(columns) + 2)]
                expected = [expected_row[column] if expected_row[column] != "" else None for column in columns]
                if observed != expected:
                    raise ValueError(f"{workbook_name}/{sheet.title}/row {offset}: semantic drift")
                if any(cell.data_type == "f" for cell in sheet[offset]):
                    raise ValueError(f"{workbook_name}/{sheet.title}/row {offset}: formula forbidden")
            counts[f"{workbook_name}/{sheet.title}"] = len(table["rows"])
        workbook.close()
        template.close()
    return counts
