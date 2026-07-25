"""Author the Goal 07 battle-start state-slot reset through production Excel.

The reset is behavior-neutral for a newly constructed battle because the slot
already holds its authored initial value. It proves that lifecycle reset rows
survive the required openpyxl -> Sora -> domain-definition path.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "config" / "data" / "StateSlotReset.xlsx"
EXPECTED = (24_003, 1, "BattleStart")


def records(sheet) -> list[tuple[int, int, str]]:
    result = []
    for row in sheet.iter_rows(min_row=8, values_only=True):
        if row[1] is None:
            continue
        result.append((int(row[1]), int(row[2]), str(row[3])))
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--write", action="store_true")
    mode.add_argument("--check", action="store_true")
    args = parser.parse_args()

    workbook = load_workbook(WORKBOOK, read_only=False, data_only=False)
    sheet = workbook.active
    matching = [row for row in records(sheet) if row[:2] == EXPECTED[:2]]
    if args.check:
        if matching != [EXPECTED]:
            raise ValueError("Goal 07 effect/state workbook probe is missing or changed")
        print("Goal 07 effect/state workbook probe matches.")
        return
    if matching and matching != [EXPECTED]:
        raise ValueError("Goal 07 effect/state probe key is already owned")
    if not matching:
        row = max(sheet.max_row + 1, 8)
        sheet.cell(row, 1, None)
        for column, value in enumerate(EXPECTED, start=2):
            sheet.cell(row, column, value)
    workbook.save(WORKBOOK)
    print("Authored Goal 07 effect/state workbook probe.")


if __name__ == "__main__":
    main()
