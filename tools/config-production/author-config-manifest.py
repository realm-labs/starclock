"""Author or verify the current production ConfigManifest workbook.

Run with the repository-approved adapter:
  uv run --with openpyxl==3.1.5 python tools/config-production/author-config-manifest.py --write
  uv run --with openpyxl==3.1.5 python tools/config-production/author-config-manifest.py --check
"""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "config" / "data" / "ConfigManifest.xlsx"
TEMPLATE = ROOT / "config" / "generated" / "templates" / "ConfigManifest.xlsx"
EXPECTED = {
    "game_version": "4.4",
    "snapshot_date": "2026-07-17",
    "sora_cli_version": "0.3.0",
}


def fields(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    return [cell.value for cell in workbook.active[3][1:] if cell.value]


def rows(path: Path) -> list[dict[str, Any]]:
    names = fields(path)
    workbook = load_workbook(path, read_only=True, data_only=False)
    records = []
    for values in workbook.active.iter_rows(min_row=8, values_only=True):
        record = {name: values[index + 1] for index, name in enumerate(names)}
        if any(value is not None for value in record.values()):
            records.append(record)
    return records


def normalized(record: dict[str, Any]) -> dict[str, str | None]:
    return {
        name: None if record.get(name) is None else str(record[name])
        for name in fields(DATA)
    }


def check() -> None:
    if fields(DATA) != list(EXPECTED):
        raise ValueError("ConfigManifest.xlsx fields differ from the current schema template")
    actual = rows(DATA)
    if len(actual) != 1 or normalized(actual[0]) != normalized(EXPECTED):
        raise ValueError("ConfigManifest.xlsx differs from current deterministic values")


def write() -> None:
    if fields(DATA) != fields(TEMPLATE):
        shutil.copyfile(TEMPLATE, DATA)
    if rows(DATA) == [EXPECTED]:
        return
    workbook = load_workbook(DATA)
    sheet = workbook.active
    names = fields(DATA)
    if sheet.max_row >= 8:
        sheet.delete_rows(8, sheet.max_row - 7)
    for column, name in enumerate(names, start=2):
        sheet.cell(row=8, column=column, value=EXPECTED[name])
    workbook.save(DATA)


def main() -> None:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--write", action="store_true")
    mode.add_argument("--check", action="store_true")
    arguments = parser.parse_args()
    if arguments.write:
        write()
    check()
    print("Current production config manifest matches schema and authored values.")


if __name__ == "__main__":
    main()
