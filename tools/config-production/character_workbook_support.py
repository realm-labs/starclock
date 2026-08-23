"""Shared deterministic workbook helpers for released character authoring."""

from __future__ import annotations

import json
from decimal import Decimal, ROUND_HALF_EVEN
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "config" / "data"


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def canonical_decimal(value: Decimal | str | int) -> str:
    decimal = Decimal(str(value))
    if decimal.as_tuple().exponent < -6:
        decimal = decimal.quantize(Decimal("0.000001"), rounding=ROUND_HALF_EVEN)
    text = format(decimal, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def workbook_rows(name: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(DATA / f"{name}.xlsx", read_only=True, data_only=False)
    sheet = workbook.active
    fields = [cell.value for cell in sheet[3][1:] if cell.value]
    rows = []
    for values in sheet.iter_rows(min_row=8, values_only=True):
        record = {field: values[index + 1] for index, field in enumerate(fields)}
        if any(value is not None for value in record.values()):
            rows.append(record)
    return fields, rows


def write_rows(name: str, records: list[dict[str, Any]]) -> None:
    path = DATA / f"{name}.xlsx"
    workbook = load_workbook(path)
    sheet = workbook.active
    fields = [cell.value for cell in sheet[3][1:] if cell.value]
    for record in records:
        unknown = set(record) - set(fields)
        if unknown:
            raise ValueError(f"{name} has unknown fields {sorted(unknown)}")
    if sheet.max_row >= 8:
        sheet.delete_rows(8, sheet.max_row - 7)
    for row_index, record in enumerate(records, start=8):
        for field_index, field in enumerate(fields, start=2):
            value = record.get(field)
            if value is not None:
                sheet.cell(row=row_index, column=field_index, value=value)
    workbook.save(path)


def normalized(value: Any) -> Any:
    if value == "":
        return None
    if value is None or isinstance(value, bool):
        return value
    return str(value)


def identity(
    id_: int, stable_key: str, kind: str, name_en: str, name_zh_cn: str, summary: str
) -> dict[str, Any]:
    return {
        "id": id_,
        "stable_key": stable_key,
        "content_kind": kind,
        "name_en": name_en,
        "name_zh_cn": name_zh_cn,
        "summary_en": summary,
        "summary_zh_cn": "版本4.4准备数据中完整转录并绑定来源的战斗内容。",
        "game_version_introduced": "unresolved",
        "game_version_snapshot": "4.4",
        "release_state": "Released",
        "enabled": True,
        "coverage_state": "DataReady",
        "source_record_ids": "1",
    }


def ability_kind(row: dict[str, Any]) -> str:
    kind = row["kind"]
    stable_key = row["id"]
    if kind == "Normal":
        if any(token in stable_key for token in ("pyrogenic-decimation", "slash-by-a-thousandfold", "bonus-stage", "big-flipping", "funky-munch", "kaboom")):
            return "EnhancedBasic"
        return "Basic"
    if kind == "BPSkill":
        return "EnhancedSkill" if "deathstar-overload" in stable_key else "Skill"
    return {
        "Ultra": "Ultimate",
        "Talent": "Talent",
        "Passive": "Talent",
        "Maze": "Technique",
        "MazeNormal": "Entry",
        "ElationDamage": "Passive",
    }.get(kind, "Passive")


def ability_slot(kind: str) -> str:
    return {
        "Basic": "Basic",
        "Skill": "Skill",
        "Ultimate": "Ultimate",
        "Talent": "Talent",
        "Technique": "Technique",
        "EnhancedBasic": "Enhanced",
        "EnhancedSkill": "Enhanced",
        "FollowUp": "Passive",
        "Counter": "Passive",
        "Summon": "Summon",
        "Memosprite": "Memosprite",
        "Entry": "Passive",
        "Passive": "Passive",
    }[kind]


def invested_level_cap(source_kind: str, effective_cap: int) -> int:
    """Separate legal player investment from the prepared effective-level table."""
    if source_kind in ("Maze", "MazeNormal"):
        return 1
    if source_kind == "Normal":
        return min(6, effective_cap)
    return min(10, effective_cap)


def target_pattern(row: dict[str, Any]) -> str:
    return {
        "SingleEnemy": "SingleTarget",
        "Blast": "Blast",
        "AllEnemies": "Aoe",
        "RandomEnemy": "Bounce",
        "AllAllies": "Support",
        "SingleAlly": "Support",
        "Self": "Enhance",
        "Battlefield": "ContentDefined",
        "": "None",
    }.get(row["mechanic_hints"]["target_hint"], "ContentDefined")


def semantic_mask(kind: str, row: dict[str, Any]) -> int:
    tags = set(row["mechanic_hints"]["operation_tags"])
    mask = 1 if "damage" in tags else 0
    family = {
        "Basic": 1,
        "EnhancedBasic": 1,
        "Skill": 2,
        "EnhancedSkill": 2,
        "Ultimate": 3,
        "FollowUp": 4,
        "Counter": 5,
        "Summon": 6,
        "Memosprite": 7,
        "Technique": 0,
    }.get(kind)
    if family is not None:
        mask |= 1 << family
    if row["character_id"] == "character.aglaea" and ("summon" in tags or kind == "Talent"):
        mask |= 1 << 7
    if row["character_id"] == "character.silver-wolf-lv-999" and row["kind"] == "ElationDamage":
        mask |= 1 << 10
    if row["kind"] == "Assist":
        mask |= 1 << 11
    return mask


def check_exact(name: str, expected: list[dict[str, Any]]) -> None:
    fields, actual = workbook_rows(name)
    project = lambda rows: [
        {field: normalized(row.get(field)) for field in fields}
        for row in rows
    ]
    if project(actual) != project(expected):
        raise ValueError(f"{name}.xlsx differs from deterministic character authoring output")
