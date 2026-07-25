"""Author the Goal 07 combat/replay compatibility revisions in production Excel."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "config" / "data" / "ConfigManifest.xlsx"
STATE_HASH_REVISION = "sha256-v5"


def main() -> None:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--write", action="store_true")
    mode.add_argument("--check", action="store_true")
    args = parser.parse_args()

    workbook = load_workbook(WORKBOOK, read_only=False, data_only=False)
    sheet = workbook["ConfigManifest"]
    header = [cell.value for cell in sheet[3]]
    column = header.index("state_hash_revision") + 1
    row = 8
    current = sheet.cell(row, column).value
    if args.check:
        if current != STATE_HASH_REVISION:
            raise ValueError(
                f"state hash revision is {current!r}, expected {STATE_HASH_REVISION!r}"
            )
        print("Goal 07 runtime revision workbook value matches.")
        return

    sheet.cell(row, column, STATE_HASH_REVISION)
    workbook.save(WORKBOOK)
    print("Authored Goal 07 runtime revision workbook value.")


if __name__ == "__main__":
    main()
