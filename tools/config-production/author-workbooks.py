"""Author or verify the small current production workbook surfaces.

Run with the repository-approved adapter after generating current templates:
  uv run --with openpyxl==3.1.5 python tools/config-production/author-workbooks.py --write
  uv run --with openpyxl==3.1.5 python tools/config-production/author-workbooks.py --check
"""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "config" / "data"
TEMPLATES = ROOT / "config" / "generated" / "templates"
MANIFEST_VALUES = {
    "game_version": "4.4",
    "snapshot_date": "2026-07-24",
    "sora_cli_version": "0.6.1",
}
NATIVE_HANDLER_FIELDS = [
    "id",
    "stable_key",
    "domain",
    "argument_schema_sha256",
    "determinism_note",
    "owner_note",
    "ir_insufficiency_reason",
    "removal_condition",
    "enabled",
]


def fields(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return [cell.value for cell in workbook.active[3][1:] if cell.value]
    finally:
        workbook.close()


def rows(path: Path) -> list[dict[str, Any]]:
    names = fields(path)
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        records = []
        for values in workbook.active.iter_rows(min_row=8, values_only=True):
            record = {name: values[index + 1] for index, name in enumerate(names)}
            if any(value is not None for value in record.values()):
                records.append(record)
        return records
    finally:
        workbook.close()


def normalized(record: dict[str, Any], names: list[str]) -> dict[str, str | None]:
    return {
        name: None if record.get(name) is None else str(record[name])
        for name in names
    }


def check_manifest() -> None:
    path = DATA / "ConfigManifest.xlsx"
    names = list(MANIFEST_VALUES)
    if fields(path) != names:
        raise ValueError("ConfigManifest.xlsx fields differ from the current schema")
    actual = rows(path)
    if len(actual) != 1 or normalized(actual[0], names) != MANIFEST_VALUES:
        raise ValueError("ConfigManifest.xlsx differs from current authored values")


def check_native_handlers() -> None:
    path = DATA / "NativeHandler.xlsx"
    template = TEMPLATES / "NativeHandler.xlsx"
    if fields(template) != NATIVE_HANDLER_FIELDS or rows(template):
        raise ValueError("generated NativeHandler.xlsx template is not current and empty")
    if fields(path) != NATIVE_HANDLER_FIELDS or rows(path):
        raise ValueError("NativeHandler.xlsx must match the current empty table schema")


def write_manifest() -> None:
    path = DATA / "ConfigManifest.xlsx"
    template = TEMPLATES / "ConfigManifest.xlsx"
    if fields(path) != fields(template):
        shutil.copyfile(template, path)
    if rows(path) == [MANIFEST_VALUES]:
        return
    workbook = load_workbook(path)
    sheet = workbook.active
    names = fields(path)
    if sheet.max_row >= 8:
        sheet.delete_rows(8, sheet.max_row - 7)
    for column, name in enumerate(names, start=2):
        sheet.cell(row=8, column=column, value=MANIFEST_VALUES[name])
    workbook.save(path)
    workbook.close()


def write_native_handlers() -> None:
    path = DATA / "NativeHandler.xlsx"
    template = TEMPLATES / "NativeHandler.xlsx"
    if fields(template) != NATIVE_HANDLER_FIELDS or rows(template):
        raise ValueError("generate the current NativeHandler.xlsx template before writing")
    if fields(path) != NATIVE_HANDLER_FIELDS or rows(path):
        shutil.copyfile(template, path)


def main() -> None:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--write", action="store_true")
    mode.add_argument("--check", action="store_true")
    arguments = parser.parse_args()
    if arguments.write:
        write_manifest()
        write_native_handlers()
    check_manifest()
    check_native_handlers()
    print("Current production workbook surfaces match schema and authored values.")


if __name__ == "__main__":
    main()
