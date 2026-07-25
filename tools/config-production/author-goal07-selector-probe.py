"""Author the first Goal 07 selector predicate through production Excel.

The row is intentionally behavior-neutral: every legal formation index passes.
Its purpose is to prevent SelectorPredicate from remaining a schema-only table
while the runtime and Sora lowerer silently ignore it.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "config" / "data" / "SelectorPredicate.xlsx"
SELECTOR_ID = 24_252
SEQUENCE = 1
PREDICATE = json.dumps(
    {
        "type": "FormationRange",
        "minimum_index": 0,
        "maximum_index": 31,
    },
    separators=(",", ":"),
)


def records(sheet) -> list[tuple[int, int, str]]:
    result = []
    for row in sheet.iter_rows(min_row=8, values_only=True):
        if row[1] is None:
            continue
        result.append((int(row[1]), int(row[2]), str(row[3])))
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--write", action="store_true")
    mode.add_argument("--check", action="store_true")
    args = parser.parse_args()

    workbook = load_workbook(WORKBOOK, read_only=False, data_only=False)
    sheet = workbook.active
    expected = (SELECTOR_ID, SEQUENCE, PREDICATE)
    current = records(sheet)
    matching = [row for row in current if row[:2] == expected[:2]]

    if args.check:
        if matching != [expected]:
            raise ValueError("Goal 07 selector predicate probe is missing or changed")
        print("Goal 07 selector predicate workbook probe matches.")
        return

    if matching and matching != [expected]:
        raise ValueError("Goal 07 selector predicate key is already owned by another row")
    if not matching:
        row = max(sheet.max_row + 1, 8)
        sheet.cell(row, 1, None)
        sheet.cell(row, 2, SELECTOR_ID)
        sheet.cell(row, 3, SEQUENCE)
        sheet.cell(row, 4, PREDICATE)
    workbook.save(WORKBOOK)
    print("Authored Goal 07 selector predicate workbook probe.")


if __name__ == "__main__":
    main()
