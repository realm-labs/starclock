"""Author Goal 07 action/break capability probes through production Excel."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "config" / "data" / "Operation.xlsx"
ROWS = (
    (
        24701,
        "goal07.probe.operation.force-break",
        "Battle",
        14001,
        None,
        "Fault",
        "EventSnapshot",
        "Rollback",
        {"type": "Break", "element": "Fire"},
    ),
    (
        24702,
        "goal07.probe.operation.delay-action",
        "Battle",
        24051,
        None,
        "Fault",
        "Dynamic",
        "Rollback",
        {"type": "DelayAction", "amount_expression_id": 24309},
    ),
    (
        24703,
        "goal07.probe.operation.grant-extra-turn",
        "Battle",
        24051,
        None,
        "Fault",
        "Dynamic",
        "Rollback",
        {"type": "GrantExtraTurn", "actor_selector_id": 24051},
    ),
)


def normalized(row) -> tuple:
    return (
        int(row[1]),
        str(row[2]),
        str(row[3]),
        None if row[4] is None else int(row[4]),
        None if row[5] is None else int(row[5]),
        str(row[6]),
        str(row[7]),
        str(row[8]),
        json.loads(str(row[9])),
    )


def records(sheet) -> list[tuple]:
    return [
        normalized(row)
        for row in sheet.iter_rows(min_row=8, values_only=True)
        if row[1] is not None
    ]


def main() -> None:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--write", action="store_true")
    mode.add_argument("--check", action="store_true")
    args = parser.parse_args()

    workbook = load_workbook(WORKBOOK, read_only=False, data_only=False)
    sheet = workbook.active
    existing = {row[0]: row for row in records(sheet)}
    if args.check:
        if [existing.get(row[0]) for row in ROWS] != list(ROWS):
            raise ValueError("Goal 07 action/break workbook probes are missing or changed")
        print("Goal 07 action/break workbook probes match.")
        return

    for expected in ROWS:
        current = existing.get(expected[0])
        if current is not None and current != expected:
            raise ValueError(f"Goal 07 operation probe key {expected[0]} is already owned")
        if current is None:
            row = max(sheet.max_row + 1, 8)
            sheet.cell(row, 1, None)
            for column, value in enumerate(expected[:-1], start=2):
                sheet.cell(row, column, value)
            sheet.cell(
                row,
                10,
                json.dumps(expected[-1], ensure_ascii=False, separators=(",", ":")),
            )
    workbook.save(WORKBOOK)
    print("Authored Goal 07 action/break workbook probes.")


if __name__ == "__main__":
    main()
