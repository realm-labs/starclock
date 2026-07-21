"""Author the six frozen Goal 01 V1B character forms into production Excel.

The prepared Version 4.4 reference pack is the only content input.  This
adapter owns IDs 20_000 through 29_999 and composes with Standard-v1 rows.

Run through the approved repository adapter:
  uv run --with openpyxl python tools/config-production/author-character-v1b.py --write
  uv run --with openpyxl python tools/config-production/author-character-v1b.py --check
"""

from __future__ import annotations

import argparse
import json
from decimal import Decimal, ROUND_HALF_EVEN
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from v1b_probe_promotion import (
    entry_rule_overrides,
    generate as generate_probe_rows,
    phase_program_overrides,
)


ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "config" / "data"
REFERENCE = ROOT / "content-reference" / "v4.4"
PARTITIONS = ROOT / "content-manifests" / "core-combat-v1" / "partitions.json"
REFERENCE_DIGEST = "0dca8ae581b4fa1e9fe8ce0c9e67ac6eb72c251deacbd4831751ce685e45ef5a"
V1B_MIN_ID = 20_000
V1B_MAX_ID = 29_999

CHARACTER_ELEMENTS = {
    "character.aglaea": "Lightning",
    "character.asta": "Fire",
    "character.clara": "Physical",
    "character.firefly": "Fire",
    "character.kafka": "Lightning",
    "character.silver-wolf-lv-999": "Imaginary",
}

# Exact primary/secondary coefficient positions for battle-facing damage
# envelopes. Rows omitted here carry setup/passive metadata but do not execute a
# direct damage hit merely because the prepared source graph mentions damage.
DIRECT_DAMAGE_PARAMETERS = {
    "character.aglaea.ability.meteoric-sunder.maze": {"All": 1},
    "character.aglaea.ability.slash-by-a-thousandfold-kiss.normal": {"Primary": 1, "Adjacent": 2},
    "character.aglaea.ability.thorned-nectar.normal": {"Primary": 1},
    "character.asta.ability.meteor-storm.bpskill": {"Primary": 1, "BounceDraw": 1},
    "character.asta.ability.miracle-flash.maze": {"All": 1},
    "character.asta.ability.spectrum-beam.normal": {"Primary": 1},
    "character.clara.ability.because-were-family.skillp01": {"Primary": 2, "Adjacent": 2},
    "character.clara.ability.i-want-to-help.normal": {"Primary": 1},
    "character.clara.ability.svarog-watches-over-you.bpskill": {"All": 1},
    "character.firefly.ability.fyrefly-type-iv-deathstar-overload.bpskill": {"Primary": 1, "Adjacent": 2},
    "character.firefly.ability.fyrefly-type-iv-pyrogenic-decimation.normal": {"Primary": 1},
    "character.firefly.ability.order-aerial-bombardment.bpskill": {"Primary": 1},
    "character.firefly.ability.order-flare-propulsion.normal": {"Primary": 1},
    "character.kafka.ability.caressing-moonlight.bpskill": {"Primary": 1, "Adjacent": 3},
    "character.kafka.ability.gentle-but-cruel.skillp01": {"Primary": 1},
    "character.kafka.ability.mercy-is-not-forgiveness.maze": {"All": 1},
    "character.kafka.ability.midnight-tumult.normal": {"Primary": 1},
    "character.kafka.ability.twilight-trill.ultra": {"All": 1},
    "character.silver-wolf-lv-999.ability.bonus-stage-αwolf-instant.normal": {"Primary": 1, "BounceDraw": 1, "All": 4},
    "character.silver-wolf-lv-999.ability.honkai-dmg-demo.elationdamage": {"Primary": 1, "BounceDraw": 1},
    "character.silver-wolf-lv-999.ability.one-punch.normal": {"Primary": 1},
    "character.silver-wolf-lv-999.ability.trigger-happy.bpskill": {"All": 1},
}

# The prepared source uses -1 both for an ordinary Basic ATK's one-point gain
# and for internal/enhanced Normal abilities that are Skill-Point neutral.  Do
# not infer resource behavior from the source kind alone.
SKILL_POINT_GAIN_ABILITIES = {
    "character.aglaea.ability.thorned-nectar.normal",
    "character.asta.ability.spectrum-beam.normal",
    "character.clara.ability.i-want-to-help.normal",
    "character.firefly.ability.order-flare-propulsion.normal",
    "character.kafka.ability.midnight-tumult.normal",
    "character.silver-wolf-lv-999.ability.one-punch.normal",
}

SILVER_WOLF_ENHANCED_BASIC = (
    "character.silver-wolf-lv-999.ability.bonus-stage-αwolf-instant.normal"
)

OWNED_TABLES = (
    "Ability",
    "AbilityHitPlanBinding",
    "AbilityLevelParameter",
    "AbilityPhase",
    "AbilityResourceDelta",
    "Character",
    "CharacterAbilityBinding",
    "CharacterResource",
    "CharacterStat",
    "ConditionExpression",
    "CountdownDefinition",
    "Effect",
    "EffectModifierBinding",
    "Eidolon",
    "EidolonPatch",
    "EventFilter",
    "HitPlan",
    "HitPlanHit",
    "LinkedUnitDefinition",
    "ModifierDefinition",
    "ModifierFilter",
    "ModifierStackingGroup",
    "Operation",
    "Program",
    "ProgramStep",
    "RuleDefinition",
    "RuleTrigger",
    "Selector",
    "StateSlot",
    "StateSlotReset",
    "TraceNode",
    "TracePatch",
    "ValueExpression",
)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def canonical_decimal(value: Decimal | str | int) -> str:
    decimal = Decimal(str(value))
    if decimal.as_tuple().exponent < -6:
        decimal = decimal.quantize(Decimal("0.000001"), rounding=ROUND_HALF_EVEN)
    text = format(decimal, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def workbook_rows(name: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(DATA / f"{name}.xlsx", read_only=True, data_only=False)
    sheet = workbook.active
    fields = [cell.value for cell in sheet[3][1:] if cell.value]
    rows = []
    for values in sheet.iter_rows(min_row=8, values_only=True):
        record = {field: values[index + 1] for index, field in enumerate(fields)}
        if any(value is not None for value in record.values()):
            rows.append(record)
    return fields, rows


def write_rows(name: str, records: list[dict[str, Any]]) -> None:
    path = DATA / f"{name}.xlsx"
    workbook = load_workbook(path)
    sheet = workbook.active
    fields = [cell.value for cell in sheet[3][1:] if cell.value]
    for record in records:
        unknown = set(record) - set(fields)
        if unknown:
            raise ValueError(f"{name} has unknown fields {sorted(unknown)}")
    if sheet.max_row >= 8:
        sheet.delete_rows(8, sheet.max_row - 7)
    for row_index, record in enumerate(records, start=8):
        for field_index, field in enumerate(fields, start=2):
            value = record.get(field)
            if value is not None:
                sheet.cell(row=row_index, column=field_index, value=value)
    workbook.save(path)


def normalized(value: Any) -> Any:
    if value == "":
        return None
    if value is None or isinstance(value, bool):
        return value
    return str(value)


def v1b_ids() -> list[str]:
    partitions = read_json(PARTITIONS)
    row = partitions["character_v1b"]
    if row["batch_id"] != "G01-P7-V1B" or len(row["ids"]) != 6:
        raise ValueError("frozen V1B partition changed")
    return row["ids"]


def sources() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    selected = set(v1b_ids())
    characters = [row for row in read_json(REFERENCE / "characters.json") if row["id"] in selected]
    abilities = [row for row in read_json(REFERENCE / "character-abilities.json") if row["character_id"] in selected]
    traces = [row for row in read_json(REFERENCE / "character-traces.json") if row["character_id"] in selected]
    eidolons = [row for row in read_json(REFERENCE / "character-eidolons.json") if row["character_id"] in selected]
    if (len(characters), len(abilities), len(traces), len(eidolons)) != (6, 45, 116, 36):
        raise ValueError("prepared V1B cardinality changed")
    return characters, abilities, traces, eidolons


def identity_map() -> dict[str, int]:
    _, rows = workbook_rows("ContentIdentity")
    return {str(row["stable_key"]): int(row["id"]) for row in rows}


def internal_maps(
    abilities: list[dict[str, Any]], traces: list[dict[str, Any]], eidolons: list[dict[str, Any]]
) -> dict[str, dict[str, int]]:
    ordered_abilities = sorted(abilities, key=lambda row: row["id"])
    damaging = [row for row in ordered_abilities if "damage" in row["mechanic_hints"]["operation_tags"]]
    return {
        "ability": {row["id"]: 20_001 + index for index, row in enumerate(ordered_abilities)},
        "hit_plan": {row["id"]: 21_001 + index for index, row in enumerate(damaging)},
        "trace": {row["id"]: 22_001 + index for index, row in enumerate(sorted(traces, key=lambda row: row["id"]))},
        "eidolon": {row["id"]: 23_001 + index for index, row in enumerate(sorted(eidolons, key=lambda row: row["id"]))},
    }


def identity(
    id_: int, stable_key: str, kind: str, name_en: str, name_zh_cn: str, summary: str
) -> dict[str, Any]:
    return {
        "id": id_,
        "stable_key": stable_key,
        "content_kind": kind,
        "name_en": name_en,
        "name_zh_cn": name_zh_cn,
        "summary_en": summary,
        "summary_zh_cn": "版本4.4准备数据中完整转录并绑定来源的战斗内容。",
        "game_version_introduced": "unresolved",
        "game_version_snapshot": "4.4",
        "release_state": "Released",
        "enabled": True,
        "coverage_state": "DataReady",
        "source_record_ids": "1",
    }


def ability_kind(row: dict[str, Any]) -> str:
    kind = row["kind"]
    stable_key = row["id"]
    if kind == "Normal":
        if any(token in stable_key for token in ("pyrogenic-decimation", "slash-by-a-thousandfold", "bonus-stage", "big-flipping", "funky-munch", "kaboom")):
            return "EnhancedBasic"
        return "Basic"
    if kind == "BPSkill":
        return "EnhancedSkill" if "deathstar-overload" in stable_key else "Skill"
    return {
        "Ultra": "Ultimate",
        "Passive": "Talent",
        "Maze": "Technique",
        "MazeNormal": "Entry",
        "ElationDamage": "Passive",
    }.get(kind, "Passive")


def ability_slot(kind: str) -> str:
    return {
        "Basic": "Basic",
        "Skill": "Skill",
        "Ultimate": "Ultimate",
        "Talent": "Talent",
        "Technique": "Technique",
        "EnhancedBasic": "Enhanced",
        "EnhancedSkill": "Enhanced",
        "Entry": "Passive",
        "Passive": "Passive",
    }[kind]


def invested_level_cap(source_kind: str, effective_cap: int) -> int:
    """Separate legal player investment from the prepared effective-level table."""
    if source_kind in ("Maze", "MazeNormal"):
        return 1
    if source_kind == "Normal":
        return min(6, effective_cap)
    return min(10, effective_cap)


def target_pattern(row: dict[str, Any]) -> str:
    if row["id"] == SILVER_WOLF_ENHANCED_BASIC:
        return "Bounce"
    return {
        "SingleEnemy": "SingleTarget",
        "Blast": "Blast",
        "AllEnemies": "Aoe",
        "RandomEnemy": "Bounce",
        "AllAllies": "Support",
        "SingleAlly": "Support",
        "Self": "Enhance",
        "Battlefield": "ContentDefined",
        "": "None",
    }.get(row["mechanic_hints"]["target_hint"], "ContentDefined")


def semantic_mask(kind: str, row: dict[str, Any]) -> int:
    tags = set(row["mechanic_hints"]["operation_tags"])
    mask = 1 if "damage" in tags else 0
    family = {
        "Basic": 1,
        "EnhancedBasic": 1,
        "Skill": 2,
        "EnhancedSkill": 2,
        "Ultimate": 3,
        "Technique": 0,
    }.get(kind)
    if family is not None:
        mask |= 1 << family
    if row["character_id"] == "character.aglaea" and ("summon" in tags or kind == "Talent"):
        mask |= 1 << 7
    if row["character_id"] == "character.silver-wolf-lv-999" and row["kind"] == "ElationDamage":
        mask |= 1 << 10
    if row["kind"] == "Assist":
        mask |= 1 << 11
    return mask


def hit_shape(row: dict[str, Any]) -> list[tuple[str, Decimal]]:
    pattern = target_pattern(row)
    toughness = [Decimal(str(value)) for value in row.get("display_toughness", [])]
    if pattern == "Bounce":
        count = next(
            (
                int(Decimal(str(value)))
                for value in row["levels"][0]["parameters"][1:]
                if Decimal(str(value)) == Decimal(str(value)).to_integral_value()
                and 1 < Decimal(str(value)) <= 100
            ),
            5,
        )
        share = Decimal(1) / Decimal(count)
        return [("Primary" if index == 0 else "BounceDraw", share) for index in range(count)]
    weighted = []
    for index, value in enumerate(toughness[:3]):
        if value > 0:
            weighted.append((("Primary", "All", "Adjacent")[index], value))
    if not weighted:
        return [("All" if pattern == "Aoe" else "Primary", Decimal(1))]
    total = sum(value for _, value in weighted)
    return [(group, value / total) for group, value in weighted]


def hit_phase_shapes(row: dict[str, Any]) -> list[tuple[int, list[tuple[str, Decimal]]]]:
    if row["id"] != SILVER_WOLF_ENHANCED_BASIC:
        return [(1, hit_shape(row))]
    # Sora caps one plan at 100 hits. The released ability already defines
    # explicit 34 + 34 + 32 resumable bounce stages followed by one final hit,
    # so each stage is represented as its own ordinary phase/plan.
    result = []
    for phase, count in enumerate((34, 34, 32), start=1):
        result.append((phase, [
            ("Primary" if phase == 1 and index == 0 else "BounceDraw", Decimal(1) / Decimal(count))
            for index in range(count)
        ]))
    result.append((4, [("All", Decimal(1))]))
    return result


def split_shares(shape: list[tuple[str, Decimal]]) -> list[tuple[str, str]]:
    scaled = []
    assigned = 0
    for index, (group, value) in enumerate(shape):
        if index == len(shape) - 1:
            units = 1_000_000 - assigned
        else:
            units = int(value * Decimal(1_000_000))
            assigned += units
        scaled.append((group, canonical_decimal(Decimal(units) / Decimal(1_000_000))))
    return scaled


def trace_kind(row: dict[str, Any]) -> str:
    if row["point_type"] == 1:
        return "MinorStat"
    if row["point_type"] == 2:
        return "BasicLevel" if row["max_level"] <= 6 else "AbilityLevel"
    if row["point_type"] == 3:
        return "MajorPassive"
    return "AbilityUnlock"


def minor_modifier_spec(property_type: str) -> tuple[str, str, str, str | None]:
    """Map exact prepared Trace stat rows into the closed modifier vocabulary."""
    direct = {
        "AttackAddedRatio": ("Atk", "PercentOfBase", "Stat", None),
        "HPAddedRatio": ("Hp", "PercentOfBase", "Stat", None),
        "DefenceAddedRatio": ("Def", "PercentOfBase", "Stat", None),
        "SpeedDelta": ("Spd", "Flat", "Stat", None),
        "CriticalChanceBase": ("CritRate", "BaseAdd", "Stat", None),
        "StatusResistanceBase": ("EffectResistance", "BaseAdd", "Stat", None),
        "StatusProbabilityBase": ("EffectHitRate", "BaseAdd", "Stat", None),
        "BreakDamageAddedRatioBase": ("BreakEffect", "BaseAdd", "Stat", None),
        "ElationDamageAddedRatioBase": ("Atk", "DamageBoost", "ElationDamage", None),
        "FireAddedRatio": ("Atk", "DamageBoost", "OrdinaryDamage", "Fire"),
        "PhysicalAddedRatio": ("Atk", "DamageBoost", "OrdinaryDamage", "Physical"),
        "ThunderAddedRatio": ("Atk", "DamageBoost", "OrdinaryDamage", "Lightning"),
    }
    try:
        return direct[property_type]
    except KeyError as error:
        raise ValueError(f"unsupported prepared minor-Trace property {property_type}") from error


def generated_rows() -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]], list[dict[str, Any]]]:
    characters, abilities, traces, eidolons = sources()
    frozen = identity_map()
    ids = internal_maps(abilities, traces, eidolons)
    rows: dict[str, list[dict[str, Any]]] = {name: [] for name in OWNED_TABLES}
    internals: list[dict[str, Any]] = []

    # Minor Traces compile as persistent, source-attributed modifier bindings.
    # One additive group is sufficient because the registry groups independently
    # by queried stat/stage/purpose/filter; authored Trace order remains explicit.
    rows["ModifierStackingGroup"].append({
        "id": 24_980, "stable_key": "v1b.trace-minor.additive", "aggregation": "Sum",
    })
    for selector_id, origin, label in (
        (24_990, "Owner", "Owner"),
        (24_991, "CurrentSubject", "Current Subject"),
    ):
        internals.append(identity(
            selector_id, f"selector.v1b.trace-minor.{origin.lower()}", "Selector",
            f"V1B Minor Trace {label} Selector", f"V1B小行迹{label}选择器",
            "Generic single-subject selector used by production minor-Trace modifiers.",
        ))
        rows["Selector"].append({
            "id": selector_id, "domain": "Battle", "origin": origin,
            "side_relationship": "SameSide", "life": "Alive", "presence": "Present",
            "reference_point": "CurrentState", "ordering": "StableId", "choice": "First",
            "minimum_count": 1, "maximum_count": 1, "allow_repeated_targets": False,
            "empty_pool_policy": "Fault",
        })

    for ability in sorted(abilities, key=lambda row: row["id"]):
        ability_id = ids["ability"][ability["id"]]
        kind = ability_kind(ability)
        damage = "damage" in ability["mechanic_hints"]["operation_tags"]
        internals.append(identity(
            ability_id, ability["id"], "Ability", ability["name_en"], ability["name_zh_cn"],
            "Complete prepared ability metadata, level parameters, resources and ordered hit structure.",
        ))
        rows["Ability"].append({
            "id": ability_id,
            "kind": kind,
            "target_pattern": target_pattern(ability),
            "retarget_policy": "RecomputeEachHit" if target_pattern(ability) == "Bounce" else "CancelRemaining",
            "level_cap": ability["max_level"],
            "cooldown_actions": max(0, int(ability.get("cooldown") or 0)),
            "semantic_tags_mask": semantic_mask(kind, ability),
        })
        rows["CharacterAbilityBinding"].append({
            "character_id": frozen[ability["character_id"]],
            "sequence": 0,
            "slot": ability_slot(kind),
            "ability_id": ability_id,
            "invested_level_cap": invested_level_cap(ability["kind"], ability["max_level"]),
        })
        for level in ability["levels"]:
            for parameter_index, value in enumerate(level["parameters"], start=1):
                rows["AbilityLevelParameter"].append({
                    "ability_id": ability_id,
                    "effective_level": level["level"],
                    "parameter_key": f"parameter.{parameter_index:02d}",
                    "value_decimal": canonical_decimal(value),
                })
        delta_sequence = 1
        skill_points = ability.get("skill_point_cost")
        if skill_points not in (None, "0", 0) and (
            Decimal(str(skill_points)) > 0 or ability["id"] in SKILL_POINT_GAIN_ABILITIES
        ):
            amount = Decimal(str(skill_points))
            rows["AbilityResourceDelta"].append({
                "ability_id": ability_id, "sequence": delta_sequence,
                "resource_kind": "SkillPoints", "delta_kind": "Spend" if amount > 0 else "Gain",
                "timing": "ActionStarted" if amount > 0 else "AbilityResolved",
                "amount_decimal": canonical_decimal(abs(amount)),
            })
            delta_sequence += 1
        energy = ability.get("energy_gain")
        if energy not in (None, "0", 0):
            rows["AbilityResourceDelta"].append({
                "ability_id": ability_id, "sequence": delta_sequence,
                "resource_kind": "Energy", "delta_kind": "Gain", "timing": "AbilityResolved",
                "amount_decimal": canonical_decimal(energy),
            })
        phase_shapes = hit_phase_shapes(ability) if damage else [(1, [])]
        for phase_sequence, _ in phase_shapes:
            rows["AbilityPhase"].append({
                "ability_id": ability_id, "sequence": phase_sequence,
                "kind": "Hits" if damage else "Resolved",
            })
        if damage:
            silver_bounce_ordinal = 0
            for phase_sequence, raw_shape in phase_shapes:
                plan_id = (
                    ids["hit_plan"][ability["id"]]
                    if phase_sequence == 1
                    else 21_988 + phase_sequence
                )
                suffix = "" if phase_sequence == 1 else f".phase-{phase_sequence}"
                internals.append(identity(
                    plan_id, f"program.hit-plan.{ability['id']}{suffix}", "Program",
                    f"{ability['name_en']} Hit Plan {phase_sequence}",
                    f"{ability['name_zh_cn']}命中计划{phase_sequence}",
                    "Ordered hit and Toughness-share structure transcribed from the prepared ability record.",
                ))
                shape = split_shares(raw_shape)
                phase_pattern = target_pattern(ability)
                rows["HitPlan"].append({
                    "id": plan_id, "target_pattern": phase_pattern,
                    "retarget_policy": "RecomputeEachHit" if phase_pattern == "Bounce" else "CancelRemaining",
                    "declared_hit_count": len(shape),
                })
                for sequence, (group, share) in enumerate(shape, start=1):
                    hit = {
                        "hit_plan_id": plan_id, "sequence": sequence, "target_group": group,
                        "damage_ratio_decimal": share, "toughness_ratio_decimal": share,
                        "crit_policy": "PerTarget",
                    }
                    parameter_map = DIRECT_DAMAGE_PARAMETERS.get(ability["id"])
                    parameter = parameter_map.get(group) if parameter_map else None
                    if parameter is not None:
                        operation_ratio = "1"
                        toughness_index = {"Primary": 0, "All": 1, "Adjacent": 2}.get(group, 0)
                        toughness = Decimal(str(ability["display_toughness"][toughness_index]))
                        if ability["id"] == SILVER_WOLF_ENHANCED_BASIC and phase_sequence <= 3:
                            # parameter.01 is the total coefficient of all 100
                            # bounces; the final hit independently uses parameter.04.
                            operation_ratio = "0.01"
                            silver_bounce_ordinal += 1
                            # Raw Toughness is integral. Preserve the prepared
                            # stage total exactly instead of flooring 0.3 on
                            # every bounce to zero.
                            toughness = Decimal(1) if silver_bounce_ordinal <= 30 else Decimal(0)
                        hit.update({
                            "damage_parameter_key_override": f"parameter.{parameter:02d}",
                            "damage_operation_ratio_decimal": operation_ratio,
                        })
                        if toughness > 0:
                            hit["toughness_amount_decimal"] = canonical_decimal(toughness)
                    rows["HitPlanHit"].append(hit)
                binding = {
                    "ability_id": ability_id, "phase_sequence": phase_sequence, "hit_plan_id": plan_id,
                }
                parameter_map = DIRECT_DAMAGE_PARAMETERS.get(ability["id"])
                if parameter_map:
                    first_parameter = next(iter(parameter_map.values()))
                    binding.update({
                        "damage_parameter_key": f"parameter.{first_parameter:02d}",
                        "damage_scaling_stat": "Atk",
                        "damage_class": "Elation" if ability["kind"] == "ElationDamage" else "Ordinary",
                        "element": CHARACTER_ELEMENTS[ability["character_id"]],
                    })
                rows["AbilityHitPlanBinding"].append(binding)

    by_character: dict[str, list[dict[str, Any]]] = {}
    for binding in rows["CharacterAbilityBinding"]:
        by_character.setdefault(str(binding["character_id"]), []).append(binding)
    for bindings in by_character.values():
        bindings.sort(key=lambda row: int(row["ability_id"]))
        for sequence, binding in enumerate(bindings, start=1):
            binding["sequence"] = sequence

    for character in sorted(characters, key=lambda row: row["id"]):
        character_id = frozen[character["id"]]
        rows["Character"].append({
            "id": character_id, "rarity": character["rarity"], "path": character["path"],
            "element": character["element"], "base_energy_decimal": character["max_energy"],
            "base_aggro_decimal": character["promotions"][0]["aggro"],
        })
        for promotion, stat in enumerate(character["promotions"]):
            first_level = 1 if promotion == 0 else promotion * 10 + 10
            for level in range(first_level, int(stat["max_level"]) + 1):
                offset = Decimal(level - 1)
                rows["CharacterStat"].append({
                    "character_id": character_id, "level": level, "promotion": promotion,
                    "hp_decimal": canonical_decimal(Decimal(stat["hp_base"]) + Decimal(stat["hp_per_level"]) * offset),
                    "atk_decimal": canonical_decimal(Decimal(stat["atk_base"]) + Decimal(stat["atk_per_level"]) * offset),
                    "def_decimal": canonical_decimal(Decimal(stat["def_base"]) + Decimal(stat["def_per_level"]) * offset),
                    "spd_decimal": canonical_decimal(stat["spd"]),
                })
        resources = {
            "character.aglaea": [("garmentmaker.speed-stacks", "6", "0")],
            "character.asta": [("charging", "5", "0")],
            "character.clara": [("enhanced-counter-charges", "2", "0")],
            "character.firefly": [("complete-combustion", "1", "0")],
            "character.kafka": [("follow-up-used", "1", "0")],
            "character.silver-wolf-lv-999": [("hidden-mmr", "300", "0"), ("enhanced-basic-count", "3", "0")],
        }[character["id"]]
        for sequence, (key, maximum, initial) in enumerate(resources, start=1):
            rows["CharacterResource"].append({
                "character_id": character_id, "sequence": sequence, "stable_key": key,
                "maximum_decimal": maximum, "initial_decimal": initial,
            })

    point_to_trace = {
        source_id: row["id"] for row in traces for source_id in row["source_point_ids"]
    }
    minor_index = 0
    for trace in sorted(traces, key=lambda row: row["id"]):
        trace_id = ids["trace"][trace["id"]]
        internals.append(identity(
            trace_id, trace["id"], "Trace", trace["name_en"], trace["name_zh_cn"],
            "Complete battle-relevant Trace identity, graph edge and prepared mechanic payload.",
        ))
        prerequisites = sorted({
            ids["trace"][point_to_trace[source]]
            for source in trace["prerequisites"]
            if source in point_to_trace
        })
        rows["TraceNode"].append({
            "id": trace_id, "character_id": frozen[trace["character_id"]],
            "kind": trace_kind(trace), "promotion_requirement": 0,
            "prerequisite_trace_ids": "|".join(str(value) for value in prerequisites) or None,
        })
        patch_sequence = 1
        unique_status_additions = {
            (addition["PropertyType"], addition["Value"])
            for addition in trace["status_additions"]
        }
        for property_type, value in sorted(unique_status_additions):
            minor_index += 1
            modifier_id = 24_700 + minor_index
            expression_id = 24_800 + minor_index
            stat, stage, purpose, element = minor_modifier_spec(property_type)
            internals.append(identity(
                modifier_id, f"modifier.{trace['id']}.{property_type}", "Modifier",
                f"{trace['name_en']} {property_type}", f"{trace['name_zh_cn']} {property_type}",
                "Exact prepared minor-Trace stat addition compiled as a persistent modifier.",
            ))
            value_kind = "Scalar" if property_type == "SpeedDelta" else "Ratio"
            rows["ValueExpression"].append({
                "id": expression_id,
                "stable_key": f"v1b.trace-minor.value.{minor_index:02d}",
                "result_kind": value_kind,
                "node": json.dumps(
                    {"type": f"{value_kind}Literal", "value_decimal": canonical_decimal(value)},
                    separators=(",", ":"),
                ),
            })
            rows["ModifierDefinition"].append({
                "id": modifier_id, "owner_selector_id": 24_990,
                "subject_selector_id": 24_991, "stat": stat, "formula_stage": stage,
                "formula_purpose": purpose, "value_expression_id": expression_id,
                "value_domain": value_kind,
                "stacking_group_id": 24_980, "priority": 0,
                "cap_formula_stage": stage, "snapshot_policy": "Dynamic",
                "duration_scope": "Battle",
            })
            if element is not None:
                rows["ModifierFilter"].append({
                    "modifier_id": modifier_id, "sequence": 1,
                    "filter": json.dumps(
                        {"type": "Element", "element": element}, separators=(",", ":")
                    ),
                })
            rows["TracePatch"].append({
                "trace_id": trace_id, "sequence": patch_sequence,
                "patch": json.dumps(
                    {"type": "AddModifier", "modifier_identity_id": modifier_id},
                    separators=(",", ":"),
                ),
            })
            patch_sequence += 1
        for source_skill_id in trace["level_up_skill_source_ids"]:
            matching = next((row for row in abilities if source_skill_id in row["source_skill_ids"]), None)
            if matching is not None and not trace["default_unlocked"]:
                rows["TracePatch"].append({
                    "trace_id": trace_id, "sequence": patch_sequence,
                    "patch": json.dumps({
                        "type": "AdjustAbilityLevel", "ability_id": ids["ability"][matching["id"]],
                        "bonus": 1, "cap_delta": 1,
                    }, separators=(",", ":")),
                })
                patch_sequence += 1

    ability_by_source = {
        source_id: row for row in abilities for source_id in row["source_skill_ids"]
    }
    for eidolon in sorted(eidolons, key=lambda row: row["id"]):
        eidolon_id = ids["eidolon"][eidolon["id"]]
        internals.append(identity(
            eidolon_id, eidolon["id"], "Eidolon", eidolon["name_en"], eidolon["name_zh_cn"],
            "Complete E1-E6 rank identity and exact prepared ability-level patches.",
        ))
        rows["Eidolon"].append({
            "id": eidolon_id, "character_id": frozen[eidolon["character_id"]], "rank": eidolon["rank"],
        })
        for sequence, addition in enumerate(eidolon["skill_level_additions"], start=1):
            ability = ability_by_source.get(addition["source_skill_id"])
            if ability is None:
                continue
            levels = int(addition["levels"])
            rows["EidolonPatch"].append({
                "eidolon_id": eidolon_id, "sequence": sequence,
                "patch": json.dumps({
                    "type": "AdjustAbilityLevel", "ability_id": ids["ability"][ability["id"]],
                    "bonus": levels, "cap_delta": levels,
                }, separators=(",", ":")),
            })

    stable_ids = dict(frozen)
    for family in ("ability", "trace", "eidolon"):
        stable_ids.update(ids[family])
    promoted, promoted_identities, resolved = generate_probe_rows(stable_ids)
    for name, promoted_table_rows in promoted.items():
        rows[name].extend(promoted_table_rows)
    internals.extend(promoted_identities)

    # Released Elation Skill: Pro-Gamer Move gains exactly 15 Hidden MMR.
    # This is a normal typed named-resource program; no character branch enters
    # the resolver.
    silver_owner_selector = 24_620
    silver_mmr_expression = 24_621
    silver_mmr_operation = 24_622
    silver_mmr_program = 24_623
    internals.extend([
        identity(
            silver_owner_selector,
            "selector.v1b.silver-wolf.owner",
            "Selector",
            "Silver Wolf Owner Selector",
            "银狼自身选择器",
            "Single owner selector for source-bound Silver Wolf resource programs.",
        ),
        identity(
            silver_mmr_program,
            "program.v1b.silver-wolf.pro-gamer-move",
            "Program",
            "Pro-Gamer Move Program",
            "高手玩家程序",
            "Exact released Elation Skill resource gain program.",
        ),
    ])
    rows["Selector"].append({
        "id": silver_owner_selector, "domain": "Battle", "origin": "Owner",
        "side_relationship": "SameSide", "life": "Alive", "presence": "Present",
        "reference_point": "CurrentState", "ordering": "StableId", "choice": "First",
        "minimum_count": 1, "maximum_count": 1, "allow_repeated_targets": False,
        "empty_pool_policy": "Fault",
    })
    rows["ValueExpression"].append({
        "id": silver_mmr_expression,
        "stable_key": "v1b.silver-wolf.expr.fifteen-hidden-mmr",
        "result_kind": "Scalar",
        "node": json.dumps({"type": "ScalarLiteral", "value_decimal": "15"}, separators=(",", ":")),
    })
    rows["Operation"].append({
        "id": silver_mmr_operation,
        "stable_key": "v1b.silver-wolf.operation.gain-hidden-mmr",
        "domain": "Battle", "target_selector_id": silver_owner_selector,
        "empty_target_policy": "Fault", "snapshot_boundary": "Dynamic",
        "fault_policy": "Rollback",
        "payload": json.dumps({
            "type": "ModifyResource", "resource_kind": "CharacterResource",
            "character_resource_key": "hidden-mmr", "update_kind": "Gain",
            "amount_expression_id": silver_mmr_expression,
            "scales_with_energy_regeneration": False, "rounding": "Floor",
        }, separators=(",", ":")),
    })
    rows["Program"].append({"id": silver_mmr_program, "domain": "Battle"})
    rows["ProgramStep"].append({
        "program_id": silver_mmr_program, "sequence": 1,
        "step": json.dumps({"type": "Operation", "operation_id": silver_mmr_operation}, separators=(",", ":")),
    })
    # Production lifecycle definitions replace the probe-only immediate teardown
    # with independently scheduled actors owned by the catalog.
    aglaea_program = resolved["aglaea-memosprite"]["6"]
    rows["ProgramStep"] = [
        row for row in rows["ProgramStep"]
        if int(row["program_id"]) != aglaea_program or int(row["sequence"]) == 1
    ]
    kafka_shock = resolved["kafka-dot"]["4"]
    for effect in rows["Effect"]:
        if int(effect["id"]) == kafka_shock:
            effect["dot_element"] = "Lightning"
    rows["EffectModifierBinding"].append({
        "effect_id": resolved["asta-modifier"]["6"],
        "sequence": 1,
        "modifier_id": resolved["asta-modifier"]["4"],
    })

    garmentmaker_action = 24_610
    firefly_countdown = 24_611
    internals.extend([
        identity(
            garmentmaker_action,
            "ability.v1b.aglaea.garmentmaker-action",
            "Ability",
            "Garmentmaker Action",
            "衣匠行动",
            "Catalog-owned independently scheduled Garmentmaker action entry point.",
        ),
        identity(
            firefly_countdown,
            "ability.v1b.firefly.complete-combustion-countdown",
            "Ability",
            "Complete Combustion Countdown",
            "完全燃烧倒计时",
            "Timeline-only action that ends Firefly's active transformation.",
        ),
    ])
    rows["Ability"].extend([
        {
            "id": garmentmaker_action,
            "kind": "Memosprite",
            "target_pattern": "SingleTarget",
            "retarget_policy": "CancelRemaining",
            "level_cap": 1,
            "cooldown_actions": 0,
            "semantic_tags_mask": 128,
        },
        {
            "id": firefly_countdown,
            "kind": "Countdown",
            "target_pattern": "Support",
            "retarget_policy": "Locked",
            "level_cap": 1,
            "cooldown_actions": 0,
            "semantic_tags_mask": 0,
        },
    ])
    rows["AbilityPhase"].extend([
        {"ability_id": garmentmaker_action, "sequence": 1, "kind": "Resolved"},
        {"ability_id": firefly_countdown, "sequence": 1, "kind": "Resolved"},
    ])
    rows["LinkedUnitDefinition"].append({
        "id": resolved["aglaea-memosprite"]["5"],
        "source_definition_identity_id": stable_ids[
            "character.aglaea.ability.rosy-fingered.skillp01"
        ],
        "kind": "Memosprite",
        "presence": "Linked",
        "ability_ids": str(garmentmaker_action),
        "action_ability_id": garmentmaker_action,
        "formation_index": 8,
        "initial_gauge_decimal": "10000",
        "hp_owner_ratio_decimal": "0.66",
        "hp_flat_decimal": "720",
        "atk_owner_ratio_decimal": "1",
        "atk_flat_decimal": "0",
        "def_owner_ratio_decimal": "1",
        "def_flat_decimal": "0",
        "spd_owner_ratio_decimal": "0",
        "spd_flat_decimal": "35",
        "owner_defeat_policy": "Depart",
        "owner_departure_policy": "Depart",
        "wave_policy": "Depart",
        "combatant_digest_sha256": "af709d10048b253540fe9439ce44beff539202c1336ebd7aee9a8e00c7b3d371",
    })
    rows["CountdownDefinition"].append({
        "code": resolved["firefly-damage"]["4"],
        "ability_id": firefly_countdown,
        "initial_gauge_decimal": "10000",
        "speed_decimal": "70",
        "owner_defeat_policy": "Depart",
        "owner_departure_policy": "Depart",
        "wave_policy": "Depart",
        "end_transformation": True,
    })
    for ability_id, program_id in phase_program_overrides(stable_ids, resolved).items():
        phase = next(
            row for row in rows["AbilityPhase"]
            if int(row["ability_id"]) == ability_id and int(row["sequence"]) == 1
        )
        phase["program_identity_id"] = program_id
    silver_elation_skill = stable_ids[
        "character.silver-wolf-lv-999.ability.pro-gamer-move.elationdamage"
    ]
    next(
        row for row in rows["AbilityPhase"]
        if int(row["ability_id"]) == silver_elation_skill and int(row["sequence"]) == 1
    )["program_identity_id"] = silver_mmr_program
    for ability_id, rule_id in entry_rule_overrides(stable_ids, resolved).items():
        ability = next(row for row in rows["Ability"] if int(row["id"]) == ability_id)
        ability["entry_rule_identity_id"] = rule_id

    for table_rows in rows.values():
        table_rows.sort(key=lambda row: tuple(str(value) for value in row.values()))
    return rows, internals, abilities + traces + eidolons


def owned_predicate(name: str) -> Callable[[dict[str, Any]], bool]:
    field = {
        "Ability": "id", "AbilityHitPlanBinding": "ability_id", "AbilityLevelParameter": "ability_id",
        "AbilityPhase": "ability_id", "AbilityResourceDelta": "ability_id", "Character": "id",
        "CharacterAbilityBinding": "ability_id", "CharacterResource": "character_id",
        "CharacterStat": "character_id", "Eidolon": "id", "EidolonPatch": "eidolon_id",
        "HitPlan": "id", "HitPlanHit": "hit_plan_id", "TraceNode": "id", "TracePatch": "trace_id",
        "ConditionExpression": "id", "Effect": "id",
        "EffectModifierBinding": "effect_id", "EventFilter": "id",
        "CountdownDefinition": "code", "LinkedUnitDefinition": "id",
        "ModifierDefinition": "id", "ModifierFilter": "modifier_id",
        "ModifierStackingGroup": "id", "Operation": "id",
        "Program": "id", "ProgramStep": "program_id", "RuleDefinition": "id",
        "RuleTrigger": "id", "Selector": "id", "StateSlot": "id",
        "StateSlotReset": "state_slot_id", "ValueExpression": "id",
    }[name]
    frozen_character_ids = set(identity_map()[key] for key in v1b_ids())
    if name in ("Character", "CharacterResource", "CharacterStat"):
        return lambda row: int(row[field]) in frozen_character_ids
    return lambda row: V1B_MIN_ID <= int(row[field]) <= V1B_MAX_ID


def merged_table(name: str, authored: list[dict[str, Any]]) -> list[dict[str, Any]]:
    _, existing = workbook_rows(name)
    owns = owned_predicate(name)
    owned_positions = [index for index, row in enumerate(existing) if owns(row)]
    insertion = min(owned_positions) if owned_positions else len(existing)
    retained = [dict(row) for row in existing if not owns(row)]
    retained[insertion:insertion] = authored
    return retained


def update_metadata(
    internals: list[dict[str, Any]], source_rows: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    _, identities = workbook_rows("ContentIdentity")
    selected = set(v1b_ids())
    retained = [dict(row) for row in identities if not (V1B_MIN_ID <= int(row["id"]) <= V1B_MAX_ID)]
    for row in retained:
        if str(row["stable_key"]) in selected:
            row["enabled"] = True
            row["coverage_state"] = "GoldenVerified"
            row["summary_en"] = str(row["summary_en"]).replace(
                " Catalog identity only; executable rows remain pending.",
                " Complete production statistics, abilities, Traces and E1-E6 rows are present.",
            )
            row["summary_zh_cn"] = str(row["summary_zh_cn"]).replace(
                " 当前仅为目录身份；可执行数据尚待转录。", " 已具备完整生产属性、技能、行迹与星魂数据。"
            )
    retained.extend(internals)
    retained.sort(key=lambda row: int(row["id"]))

    source_by_id = {row["id"]: row for row in source_rows}
    _, bindings = workbook_rows("ContentEvidenceBinding")
    selected_content_ids = {identity_map()[key] for key in selected}
    kept = [
        dict(row) for row in bindings
        if not (V1B_MIN_ID <= int(row["content_id"]) <= V1B_MAX_ID)
        and not (int(row["content_id"]) in selected_content_ids and int(row["sequence"]) >= 2)
    ]
    for record in internals:
        source = source_by_id.get(record["stable_key"])
        quality = source.get("quality", "ExactStructured") if source else "ExactStructured"
        mechanism = source.get("mechanism_quality", quality) if source else quality
        if mechanism == "ExactPreviousRelease":
            mechanism = "ExactPreviousReleaseText"
        kept.append({
            "content_id": record["id"], "sequence": 1,
            "fact_key": f"v1b.prepared:{record['stable_key']}",
            "source_record_id": 1, "evidence_record_id": 3,
            "quality": quality, "mechanism_quality": mechanism,
        })
    for stable_key in sorted(selected):
        kept.append({
            "content_id": identity_map()[stable_key], "sequence": 2,
            "fact_key": f"v1b.executable:{stable_key}",
            "source_record_id": 1, "evidence_record_id": 3,
            "quality": "ExactStructured", "mechanism_quality": "ExactStructured",
        })
    kept.sort(key=lambda row: (int(row["content_id"]), int(row["sequence"])))
    return retained, kept


def check_exact(name: str, expected: list[dict[str, Any]]) -> None:
    fields, actual = workbook_rows(name)
    project = lambda rows: [{field: normalized(row.get(field)) for field in fields} for row in rows]
    if project(actual) != project(expected):
        raise ValueError(f"{name}.xlsx differs from deterministic V1B authoring output")


def main() -> None:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--write", action="store_true")
    mode.add_argument("--check", action="store_true")
    args = parser.parse_args()
    pack = read_json(REFERENCE / "pack-index.json")
    if pack["pack_sha256"] != REFERENCE_DIGEST:
        raise ValueError("prepared reference pack digest changed")
    rows, internals, source_rows = generated_rows()
    expected = {name: merged_table(name, rows[name]) for name in OWNED_TABLES}
    identities, evidence = update_metadata(internals, source_rows)
    _, manifest_rows = workbook_rows("ConfigManifest")
    if len(manifest_rows) != 1:
        raise ValueError("production ConfigManifest must remain a singleton")
    if args.write:
        manifest_rows[0]["data_revision"] = "core-combat-v1-phase7-v1b"
        for name in OWNED_TABLES:
            write_rows(name, expected[name])
        write_rows("ContentIdentity", identities)
        write_rows("ContentEvidenceBinding", evidence)
        write_rows("ConfigManifest", manifest_rows)
        print("Authored six frozen V1B character forms into production workbooks.")
    else:
        for name in OWNED_TABLES:
            check_exact(name, expected[name])
        check_exact("ContentIdentity", identities)
        check_exact("ContentEvidenceBinding", evidence)
        check_exact("ConfigManifest", manifest_rows)
        print("Frozen V1B character workbooks match deterministic authoring output.")


if __name__ == "__main__":
    main()
