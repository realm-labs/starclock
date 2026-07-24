"""Author and verify the frozen Goal 01 Standard-v1 production workbook rows.

Run with the repository-pinned adapter:
  uv run --with openpyxl python tools/config-production/author-standard-v1.py --write
  uv run --with openpyxl python tools/config-production/author-standard-v1.py --check

The prepared reference pack and frozen goal manifest are the only content inputs.
The script owns the executable Standard-v1 tables and the additional internal
identities they require. It deliberately leaves character and Light Cone rows
for Phase 7.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "config" / "data"
REFERENCE = ROOT / "content-reference" / "v4.4"
MANIFEST = ROOT / "content-manifests" / "core-combat-v1" / "standard-v1.json"
REFERENCE_DIGEST = "0dca8ae581b4fa1e9fe8ce0c9e67ac6eb72c251deacbd4831751ce685e45ef5a"
GOAL_DIGEST = "e2188c7844d678253c98d569db017dbad7101541cf502aba4c2eb80c0435bf19"

OWNED_TABLES = (
    "Ability",
    "AbilityHitPlanBinding",
    "AbilityPhase",
    "ActivityDefinition",
    "ActivityEdge",
    "ActivityNode",
    "ActivitySection",
    "AiCandidate",
    "AiGraph",
    "AiState",
    "BattleBinding",
    "BattleParticipantSlot",
    "BattleResultProjection",
    "BattleResultProjectionField",
    "ConditionExpression",
    "Encounter",
    "EncounterWave",
    "EnemyAbility",
    "EnemyDebuffResistance",
    "EnemyLink",
    "EnemyPhase",
    "EnemyResistance",
    "EnemyStat",
    "EnemyTemplate",
    "EnemyToughnessLayer",
    "EnemyVariant",
    "EnemyVariantAbility",
    "EnemyWeakness",
    "HitPlan",
    "HitPlanHit",
    "ParticipantPolicy",
    "Selector",
    "StandardProfile",
    "StandardScenario",
    "WaveSlot",
)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def sha256_json(value: Any) -> str:
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def canonical_decimal(value: str) -> str:
    if "." not in value:
        return value
    return value.rstrip("0").rstrip(".")


def workbook_rows(name: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(DATA / f"{name}.xlsx", read_only=True, data_only=False)
    sheet = workbook.active
    fields = [cell.value for cell in sheet[3][1:] if cell.value]
    rows = []
    for values in sheet.iter_rows(min_row=8, values_only=True):
        record = {field: values[index + 1] for index, field in enumerate(fields)}
        if any(value is not None for value in record.values()):
            rows.append(record)
    return fields, rows


def normalized(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    return str(value)


def write_rows(name: str, records: list[dict[str, Any]]) -> None:
    path = DATA / f"{name}.xlsx"
    workbook = load_workbook(path)
    sheet = workbook.active
    fields = [cell.value for cell in sheet[3][1:] if cell.value]
    for record in records:
        unknown = set(record) - set(fields)
        if unknown:
            raise ValueError(f"{name} has unknown fields {sorted(unknown)}")
    if sheet.max_row >= 8:
        sheet.delete_rows(8, sheet.max_row - 7)
    for row_index, record in enumerate(records, start=8):
        for field_index, field in enumerate(fields, start=2):
            value = record.get(field)
            if value is not None:
                sheet.cell(row=row_index, column=field_index, value=value)
    workbook.save(path)


def source_data() -> tuple[dict[str, Any], list[dict[str, Any]], dict[str, Any], dict[str, Any], dict[str, Any]]:
    manifest = read_json(MANIFEST)
    templates = {row["id"]: row for row in read_json(REFERENCE / "enemy-templates.json")}
    variants = {row["id"]: row for row in read_json(REFERENCE / "enemy-variants.json")}
    abilities = {row["source_skill_id"]: row for row in read_json(REFERENCE / "enemy-abilities.json")}
    encounters = {row["id"]: row for row in read_json(REFERENCE / "encounters.json")}
    selected_variants = [variants[row["id"]] for row in manifest["enemies"]]
    selected_templates = [templates[row["enemy_id"]] for row in selected_variants]
    selected_abilities = [abilities[skill] for row in selected_variants for skill in row["source_skill_ids"]]
    if len(selected_variants) != 17 or len({row["id"] for row in selected_abilities}) != 63:
        raise ValueError("frozen Standard-v1 reference cardinality changed")
    return manifest, selected_variants, {row["id"]: row for row in selected_templates}, {row["id"]: row for row in selected_abilities}, encounters


def current_identity_rows() -> list[dict[str, Any]]:
    _, rows = workbook_rows("ContentIdentity")
    return rows


def transport_maps(
    identities: list[dict[str, Any]],
    variants: list[dict[str, Any]],
    templates: dict[str, Any],
    abilities: dict[str, Any],
) -> dict[str, dict[str, int]]:
    frozen = {str(row["stable_key"]): int(row["id"]) for row in identities if int(row["id"]) <= 283}
    template_ids = {key: 10_001 + index for index, key in enumerate(sorted(templates))}
    ability_ids = {key: 11_001 + index for index, key in enumerate(sorted(abilities))}
    hit_plan_ids = {key: 12_001 + index for index, key in enumerate(sorted(abilities))}
    graph_ids = {row["id"]: 13_001 + index for index, row in enumerate(sorted(variants, key=lambda item: item["id"]))}
    return {
        "frozen": frozen,
        "template": template_ids,
        "ability": ability_ids,
        "hit_plan": hit_plan_ids,
        "graph": graph_ids,
    }


def identity(
    id_: int,
    stable_key: str,
    kind: str,
    name_en: str,
    name_zh_cn: str,
    summary_en: str,
    summary_zh_cn: str,
    coverage: str = "DataReady",
) -> dict[str, Any]:
    return {
        "id": id_,
        "stable_key": stable_key,
        "content_kind": kind,
        "name_en": name_en,
        "name_zh_cn": name_zh_cn,
        "summary_en": summary_en,
        "summary_zh_cn": summary_zh_cn,
        "game_version_introduced": "unresolved",
        "game_version_snapshot": "4.4",
        "release_state": "Released",
        "enabled": True,
        "coverage_state": coverage,
        "source_record_ids": "1",
    }


def target_pattern(ability: dict[str, Any]) -> str:
    hint = ability["mechanic_hints"]["target_hint"]
    return {
        "SingleEnemy": "SingleTarget",
        "AllEnemies": "Aoe",
        "AllAllies": "Support",
        "Self": "Enhance",
        "Battlefield": "ContentDefined",
        "": "None",
    }.get(hint, "ContentDefined")


def has_damage(ability: dict[str, Any]) -> bool:
    return "damage" in ability["mechanic_hints"]["operation_tags"]


def rank(value: str) -> str:
    return {"Minion": "Minion", "MinionLv2": "Normal", "Elite": "Elite", "BigBoss": "Boss"}[value]


def json_cell(type_: str, **fields: Any) -> str:
    return json.dumps({"type": type_, **fields}, separators=(",", ":"))


def generated_rows() -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]], list[dict[str, Any]]]:
    manifest, variants, templates, abilities, encounters = source_data()
    current_identities = current_identity_rows()
    original_identities = [row for row in current_identities if int(row["id"]) <= 283]
    later_identities = [dict(row) for row in current_identities if int(row["id"]) >= 20_000]
    ids = transport_maps(original_identities, variants, templates, abilities)
    by_variant = {row["id"]: row for row in variants}
    ability_by_source = {row["source_skill_id"]: row for row in abilities.values()}
    rows: dict[str, list[dict[str, Any]]] = {name: [] for name in OWNED_TABLES}

    internal_identities: list[dict[str, Any]] = []
    for key in sorted(templates):
        row = templates[key]
        internal_identities.append(identity(
            ids["template"][key], key, "Enemy", row["name_en"], row["name_zh_cn"],
            "Prepared Version 4.4 enemy template used by the frozen Standard-v1 manifest.",
            "版本4.4准备数据中的敌人模板，用于已冻结的标准模式第一版清单。",
        ))
    for key in sorted(abilities):
        row = abilities[key]
        internal_identities.append(identity(
            ids["ability"][key], key, "Ability", row["name_en"], row["name_zh_cn"],
            "Prepared enemy ability with deterministic baseline target and hit-plan lowering.",
            "已准备的敌人技能，含确定性的基础目标与命中计划转录。",
        ))
        if has_damage(row):
            internal_identities.append(identity(
                ids["hit_plan"][key], f"program.hit-plan.{key}", "Program",
                f"{row['name_en']} Hit Plan", f"{row['name_zh_cn']}命中计划",
                "Deterministic hit structure derived from the prepared enemy ability record.",
                "根据已准备敌人技能记录派生的确定性命中结构。",
            ))
    for variant in sorted(variants, key=lambda item: item["id"]):
        template = templates[variant["enemy_id"]]
        internal_identities.append(identity(
            ids["graph"][variant["id"]], f"ai.standard-v1.{variant['id']}", "AiGraph",
            f"{template['name_en']} Standard AI", f"{template['name_zh_cn']}标准模式AI",
            "Finite first-legal AI graph preserving the prepared source-skill order.",
            "保留已准备来源技能顺序的有限首个合法AI图。",
        ))
    internal_identities.extend([
        identity(14_001, "selector.standard-v1.opposing", "Selector", "Standard Enemy Target", "标准敌方目标", "Formation-ordered opposing target selector.", "按站位排序的敌方目标选择器。"),
        identity(15_001, "activity.standard-v1", "Activity", "Standard Version 1 Activity", "标准模式第一版活动", "One-battle activity shared by the six frozen scenario bindings.", "六个冻结场景绑定共享的单战斗活动。"),
    ])

    external_keys = {row["id"] for row in manifest["enemies"]}
    external_keys.update(row["id"] for row in manifest["encounters"])
    external_keys.update(row["id"] for row in manifest["scenarios"])
    external_keys.add(manifest["profile"]["id"])
    updated_identities = []
    for record in original_identities:
        copied = dict(record)
        if str(record["stable_key"]) in external_keys:
            copied["enabled"] = True
            copied["coverage_state"] = "GoldenVerified"
            copied["summary_en"] = str(copied["summary_en"]).replace(" Catalog identity only; executable rows remain pending.", " Executable Standard-v1 rows and seeded golden evidence are present.")
            copied["summary_zh_cn"] = str(copied["summary_zh_cn"]).replace(" 当前仅为目录身份；可执行数据尚待转录。", " 已具备标准模式第一版可执行数据与种子黄金证据。")
        updated_identities.append(copied)
    updated_identities.extend(internal_identities)
    updated_identities.extend(later_identities)
    updated_identities.sort(key=lambda row: int(row["id"]))

    _, original_bindings = workbook_rows("ContentEvidenceBinding")
    evidence_bindings = [
        dict(row) for row in original_bindings
        if (int(row["content_id"]) <= 283 and int(row["sequence"]) == 1)
        or int(row["content_id"]) in {2, 8, 18, 27, 45, 68}
        or int(row["content_id"]) >= 20_000
    ]
    for record in internal_identities:
        evidence_bindings.append({
            "content_id": record["id"], "sequence": 1,
            "fact_key": f"standard-v1.executable:{record['stable_key']}",
            "source_record_id": 1, "evidence_record_id": 3,
            "quality": "ExactStructured", "mechanism_quality": "ExactStructured",
        })
    for key in sorted(external_keys):
        evidence_bindings.append({
            "content_id": ids["frozen"][key], "sequence": 2,
            "fact_key": f"standard-v1.executable:{key}",
            "source_record_id": 1, "evidence_record_id": 3,
            "quality": "ExactStructured", "mechanism_quality": "ExactStructured",
        })
    evidence_bindings.sort(key=lambda row: (int(row["content_id"]), int(row["sequence"])))

    rows["ConditionExpression"] = [{"id": 1, "stable_key": "standard-v1.condition.always", "node": json_cell("Constant", value=True)}]
    rows["Selector"] = [{
        "id": 14_001, "domain": "Battle", "origin": "Actor", "side_relationship": "OpposingSide",
        "life": "Alive", "presence": "Present", "reference_point": "CurrentState", "ordering": "Formation",
        "minimum_count": 1, "maximum_count": 1, "empty_pool_policy": "Fault", "choice": "First",
        "allow_repeated_targets": False,
    }]

    for ability_key in sorted(abilities):
        ability = abilities[ability_key]
        ability_id = ids["ability"][ability_key]
        pattern = target_pattern(ability)
        damage = has_damage(ability)
        rows["Ability"].append({
            "id": ability_id, "kind": "Passive" if ability["use_type"] == "Passive" else "Skill",
            "target_pattern": pattern, "retarget_policy": "CancelRemaining", "level_cap": 1,
            "cooldown_actions": ability["ai_cooldown"] or 0, "semantic_tags_mask": 5 if damage else 4,
        })
        rows["AbilityPhase"].append({"ability_id": ability_id, "sequence": 1, "kind": "Hits" if damage else "Resolved"})
        rows["EnemyAbility"].append({
            "id": ability_id, "telegraph": "Charge" if (ability["ai_initial_cooldown"] or 0) > 1 else "None",
            "cooldown_actions": ability["ai_cooldown"] or 0,
            "initial_cooldown_actions": ability["ai_initial_cooldown"] or 0,
            "charge_actions": 0, "ai_tag": ability["trigger_key"].lower(),
        })
        if damage:
            plan_id = ids["hit_plan"][ability_key]
            rows["HitPlan"].append({"id": plan_id, "target_pattern": pattern, "retarget_policy": "CancelRemaining", "declared_hit_count": 1})
            rows["HitPlanHit"].append({
                "hit_plan_id": plan_id, "sequence": 1,
                "target_group": "All" if pattern == "Aoe" else "Primary",
                "damage_ratio_decimal": "1", "toughness_ratio_decimal": "1", "crit_policy": "PerTarget",
            })
            rows["AbilityHitPlanBinding"].append({"ability_id": ability_id, "phase_sequence": 1, "hit_plan_id": plan_id})

    level_by_variant: dict[str, int] = {}
    for encounter_entry in manifest["encounters"]:
        encounter = encounters[encounter_entry["id"]]
        for wave in encounter["waves"]:
            for slot in wave["slots"]:
                level_by_variant[slot["enemy_variant_id"]] = max(level_by_variant.get(slot["enemy_variant_id"], 0), encounter["level"])

    state_id = 1
    candidate_id = 1
    phase_id = 1
    for variant in sorted(variants, key=lambda item: item["id"]):
        variant_id = ids["frozen"][variant["id"]]
        template = templates[variant["enemy_id"]]
        template_id = ids["template"][variant["enemy_id"]]
        graph_id = ids["graph"][variant["id"]]
        source_abilities = [ability_by_source[source] for source in variant["source_skill_ids"]]
        candidate_abilities = [ability for ability in source_abilities if ability["use_type"] != "Passive"]
        if not candidate_abilities:
            candidate_abilities = source_abilities[:1]
        fallback = ids["ability"][candidate_abilities[0]["id"]]
        rows["AiGraph"].append({"id": graph_id, "initial_state_id": state_id, "automatic_transition_budget": 8})
        rows["AiState"].append({
            "id": state_id, "stable_key": f"standard-v1.ai.state.{variant['id']}", "graph_id": graph_id,
            "mandatory_fallback_ability_id": fallback, "turn_counter_reset": False,
        })
        for sequence, ability in enumerate(candidate_abilities, start=1):
            ability_id = ids["ability"][ability["id"]]
            rows["AiCandidate"].append({
                "id": candidate_id, "stable_key": f"standard-v1.ai.candidate.{variant['id']}.{sequence:02d}",
                "state_id": state_id, "sequence": sequence, "ability_id": ability_id, "condition_id": 1,
                "target_selector_id": 14_001, "priority": sequence, "selection": "FirstLegal",
                "no_target_fallback": "UseFallbackAbility", "fallback_ability_id": fallback,
            })
            candidate_id += 1
        rows["EnemyTemplate"].append({"id": template_id, "rank": rank(template["rank"]), "base_aggro_decimal": "100", "default_ai_graph_id": graph_id})
        rows["EnemyVariant"].append({"id": variant_id, "template_id": template_id, "ai_graph_id": graph_id, "mechanically_distinct_key": variant["id"]})
        stats = template["base_stats"]
        multipliers = variant["stat_multipliers"]
        def scaled(field: str) -> str:
            value = Decimal(stats[field]) * Decimal(multipliers[field])
            return canonical_decimal(format(value, "f"))
        rows["EnemyStat"].append({
            "variant_id": variant_id, "level": level_by_variant[variant["id"]], "difficulty_key": "standard-v1",
            "hp_decimal": scaled("hp"), "atk_decimal": scaled("atk"), "def_decimal": scaled("def"),
            "spd_decimal": scaled("spd"), "effect_resistance_decimal": canonical_decimal(stats["effect_res"] or "0"),
            "crit_damage_decimal": canonical_decimal(stats["crit_damage"]),
        })
        for sequence, weakness in enumerate(sorted(variant["weaknesses"]), start=1):
            rows["EnemyWeakness"].append({"variant_id": variant_id, "sequence": sequence, "element": weakness})
        for resistance in variant["resistances"]:
            rows["EnemyResistance"].append({"variant_id": variant_id, "element": resistance["element"], "value_decimal": canonical_decimal(resistance["value"])})
        for resistance in variant["debuff_resistances"]:
            rows["EnemyDebuffResistance"].append({"variant_id": variant_id, "category_key": resistance["effect"], "value_decimal": canonical_decimal(resistance["value"])})
        for sequence in range(1, template["toughness_layers"] + 1):
            rows["EnemyToughnessLayer"].append({
                "variant_id": variant_id, "sequence": sequence, "layer_key": f"layer-{sequence}",
                "kind": "Ordinary" if sequence == 1 else "Sequential",
                "maximum_decimal": scaled("toughness"), "recovery_ratio_decimal": "1", "active_at_start": sequence == 1,
            })
        for sequence, ability in enumerate(source_abilities, start=1):
            rows["EnemyVariantAbility"].append({"variant_id": variant_id, "sequence": sequence, "ability_id": ids["ability"][ability["id"]]})
        phase_count = 2 if "cocolia-mother" in variant["id"] else 3 if "great-septimus" in variant["id"] else 0
        for sequence in range(1, phase_count + 1):
            rows["EnemyPhase"].append({
                "id": phase_id, "stable_key": f"standard-v1.phase.{variant['id']}.{sequence}",
                "variant_id": variant_id, "sequence": sequence, "entry_condition_id": 1, "exit_condition_id": 1,
                "replacement_priority": sequence, "ai_graph_id": graph_id, "targetable": True,
                "transition_model": "TransformSameUnit", "hp_carry": "CarryRatio", "action_gauge_carry": "CarryExact",
                "effect_carry": "Clear", "toughness_carry": "Reset", "summon_carry": "Clear",
            })
            phase_id += 1
        state_id += 1

    link_pairs = [
        ("enemy.aurumaton-gatekeeper.elite.variant.01", "enemy.entranced-ingenium-illumination-dragonfish.minionlv2.variant.01"),
        ("enemy.disciples-of-sanctus-medicus-shape-shifter.elite.variant.01", "enemy.mara-struck-soldier.minionlv2.variant.01"),
    ]
    for owner, linked in link_pairs:
        rows["EnemyLink"].append({
            "owner_variant_id": ids["frozen"][owner], "sequence": 1, "linked_variant_id": ids["frozen"][linked],
            "kind": "Summon", "maximum_simultaneous": 2, "overflow_policy": "ReplaceOldest",
            "owner_defeat_policy": "Despawn", "wave_persistence": "WaveOwned", "contributes_to_victory": False,
            "initial_action_gauge_decimal": "0", "formation_policy": "NextAvailable",
        })

    wave_id = 1
    for encounter_entry in manifest["encounters"]:
        encounter = encounters[encounter_entry["id"]]
        encounter_id = ids["frozen"][encounter_entry["id"]]
        rows["Encounter"].append({
            "id": encounter_id, "level": encounter["level"], "difficulty_key": f"hard-level-{encounter['hard_level_group']}",
            "environment_key": encounter["stage_type"].lower(), "wave_transition": "AfterAction",
            "initial_skill_points": 3, "maximum_skill_points": 5,
            "victory_policy": "DefeatRequiredHostiles", "loss_policy": "NoControllableAllies",
        })
        for wave in encounter["waves"]:
            rows["EncounterWave"].append({
                "id": wave_id, "stable_key": f"standard-v1.wave.{encounter_entry['id']}.{wave['order']}",
                "encounter_id": encounter_id, "sequence": wave["order"], "hp_carry": "CarryExact",
                "energy_carry": "CarryExact", "skill_point_carry": "CarryExact", "effect_carry": "Clear",
                "action_gauge_carry": "Reset",
            })
            for spawn_sequence, slot in enumerate(wave["slots"], start=1):
                variant_id = ids["frozen"][slot["enemy_variant_id"]]
                initial_phase = None
                if "cocolia-mother" in slot["enemy_variant_id"]:
                    initial_phase = next(row["id"] for row in rows["EnemyPhase"] if row["variant_id"] == variant_id and row["sequence"] == 1)
                if "great-septimus" in slot["enemy_variant_id"]:
                    initial_phase = next(row["id"] for row in rows["EnemyPhase"] if row["variant_id"] == variant_id and row["sequence"] == 1)
                rows["WaveSlot"].append({
                    "wave_id": wave_id, "spawn_sequence": spawn_sequence, "formation_index": spawn_sequence + 3,
                    "enemy_variant_id": variant_id, "level_override": encounter["level"],
                    "initial_phase_id": initial_phase, "required_for_victory": True,
                })
            wave_id += 1

    rows["ActivityDefinition"] = [{"id": 15_001, "entry_node_id": 1, "maximum_total_visits": 4}]
    rows["ActivitySection"] = [{"id": 1, "stable_key": "standard-v1.section", "activity_id": 15_001, "sequence": 1, "entry_node_id": 1}]
    rows["ActivityNode"] = [
        {"id": 1, "stable_key": "standard-v1.node.battle", "section_id": 1, "kind": "Battle", "maximum_visits": 1},
        {"id": 2, "stable_key": "standard-v1.node.complete", "section_id": 1, "kind": "Terminal", "maximum_visits": 1, "terminal_outcome": "Complete"},
        {"id": 3, "stable_key": "standard-v1.node.failed", "section_id": 1, "kind": "Terminal", "maximum_visits": 1, "terminal_outcome": "Failed"},
        {"id": 4, "stable_key": "standard-v1.node.faulted", "section_id": 1, "kind": "Terminal", "maximum_visits": 1, "terminal_outcome": "Faulted"},
    ]
    rows["ActivityEdge"] = [
        {"id": 1, "stable_key": "standard-v1.edge.won", "activity_id": 15_001, "source_node_id": 1, "target_node_id": 2, "condition": "BattleWon", "priority": 1, "maximum_traversals": 1},
        {"id": 2, "stable_key": "standard-v1.edge.lost", "activity_id": 15_001, "source_node_id": 1, "target_node_id": 3, "condition": "BattleLost", "priority": 2, "maximum_traversals": 1},
        {"id": 3, "stable_key": "standard-v1.edge.faulted", "activity_id": 15_001, "source_node_id": 1, "target_node_id": 4, "condition": "BattleFaulted", "priority": 3, "maximum_traversals": 1},
    ]
    rows["ParticipantPolicy"] = [{
        "id": 1, "stable_key": "standard-v1.participants", "activity_id": 15_001, "team_count": 1,
        "minimum_team_size": 1, "maximum_team_size": 4, "uniqueness_scope": "Team",
        "loadout_lock_scope": "Attempt", "allow_substitution": False,
    }]
    rows["BattleResultProjection"] = [{"id": 1, "stable_key": "standard-v1.projection"}]
    rows["BattleResultProjectionField"] = [
        {"projection_id": 1, "sequence": index, "field": json_cell(field)}
        for index, field in enumerate(("Outcome", "FinalStateHash", "EventDigest", "TerminalFault"), start=1)
    ]
    rows["StandardProfile"] = [{
        "id": ids["frozen"][manifest["profile"]["id"]], "activity_id": 15_001, "player_team_count": 1,
        "maximum_party_size": 4, "has_global_clock": False, "has_score": False, "has_seasonal_rules": False,
        "default_wave_transition": "AfterAction",
    }]
    character_ids = ids["frozen"]
    for binding_id, scenario in enumerate(manifest["scenarios"], start=1):
        encounter_id = ids["frozen"][scenario["encounter_id"]]
        lock = sha256_json(scenario["builds"])
        rows["BattleBinding"].append({
            "id": binding_id, "stable_key": f"standard-v1.binding.{scenario['id']}", "node_id": 1,
            "encounter_id": encounter_id, "participant_policy_id": 1, "projection_id": 1,
            "seed_stream_label": "standard-v1-battle", "participant_lock_sha256": lock,
            "battle_spec_policy_revision": "starclock.battle-spec.v1",
        })
        for formation, build in enumerate(scenario["builds"]):
            build_digest = sha256_json(build)
            spec_digest = sha256_json({"build": build, "policy": "standard-v1-precompiled-v1"})
            rows["BattleParticipantSlot"].append({
                "battle_binding_id": binding_id, "team_index": 0, "formation_index": formation,
                "character_id": character_ids[build["form_id"]], "resolved_spec_sha256": spec_digest,
                "build_digest_sha256": build_digest, "build_catalog_revision": "core-combat-v1-standard-lock-v1",
                "source_kind": "CompiledBuild",
            })
        rows["StandardScenario"].append({
            "id": ids["frozen"][scenario["id"]], "profile_id": ids["frozen"][manifest["profile"]["id"]],
            "activity_id": 15_001, "battle_binding_id": binding_id,
            "master_seed_hex": f"{int(scenario['seed']):016x}", "expected_outcome": "Won",
        })

    later_fields = {
        "Ability": "id",
        "AbilityHitPlanBinding": "ability_id",
        "AbilityPhase": "ability_id",
        "ConditionExpression": "id",
        "HitPlan": "id",
        "HitPlanHit": "hit_plan_id",
        "Selector": "id",
    }
    for name, field in later_fields.items():
        _, existing = workbook_rows(name)
        rows[name].extend(
            dict(row) for row in existing if int(row[field]) >= 20_000
        )
    for name in rows:
        rows[name].sort(key=lambda record: tuple(str(value) for value in record.values()))
    return rows, updated_identities, evidence_bindings


def expected_manifest_row() -> dict[str, Any]:
    return {
        "game_version": "4.4", "snapshot_date": "2026-07-17", "data_revision": "core-combat-v1-phase7-v1b",
        "required_rules_revision": "core-combat-rules-v1", "sora_cli_version": "0.3.0",
        "numeric_policy_revision": "fixed-i64-6dp-v1", "rng_algorithm_revision": "chacha8-rand-0.10.2-intmap-v1",
        "state_hash_revision": "sha256-v4", "replay_format_version": "replay-v1",
        "coverage_manifest_sha256": GOAL_DIGEST,
    }


def check_table(name: str, expected: list[dict[str, Any]]) -> None:
    fields, actual = workbook_rows(name)
    def project(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [{field: normalized(row.get(field)) for field in fields} for row in rows]
    if project(actual) != project(expected):
        raise ValueError(f"{name}.xlsx differs from deterministic Standard-v1 authoring output")


def main() -> None:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--write", action="store_true")
    mode.add_argument("--check", action="store_true")
    args = parser.parse_args()
    manifest = read_json(REFERENCE / "pack-index.json")
    if manifest["pack_sha256"] != REFERENCE_DIGEST:
        raise ValueError("prepared reference pack digest changed")
    rows, identities, evidence = generated_rows()
    if args.write:
        for name in OWNED_TABLES:
            write_rows(name, rows[name])
        write_rows("ContentIdentity", identities)
        write_rows("ContentEvidenceBinding", evidence)
        write_rows("ConfigManifest", [expected_manifest_row()])
        _, evidence_rows = workbook_rows("EvidenceRecord")
        for record in evidence_rows:
            if int(record["id"]) == 3:
                record["note"] = "Deterministically normalized Version 4.4 reference pack used by executable Standard-v1 rows."
        write_rows("EvidenceRecord", evidence_rows)
        print("Authored frozen Standard-v1 production workbooks.")
    else:
        for name in OWNED_TABLES:
            check_table(name, rows[name])
        check_table("ContentIdentity", identities)
        check_table("ContentEvidenceBinding", evidence)
        check_table("ConfigManifest", [expected_manifest_row()])
        print("Frozen Standard-v1 production workbooks match deterministic authoring output.")


if __name__ == "__main__":
    main()
