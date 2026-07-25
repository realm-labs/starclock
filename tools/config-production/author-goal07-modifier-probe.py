"""Author the Goal 07 comparator probe through production Excel.

The stacking group is intentionally unreferenced by enabled content. It proves
that an authored comparator survives Excel, Sora and domain lowering without
changing a released combat definition.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "config" / "data" / "ModifierStackingGroup.xlsx"
GROUP_ID = 970_001
STABLE_KEY = "goal07.probe.modifier.strongest-comparator"
AGGREGATION = "StrongestByComparator"
COMPARATOR_EXPRESSION_ID = 24_801
EXPECTED = (GROUP_ID, STABLE_KEY, AGGREGATION, COMPARATOR_EXPRESSION_ID)


def records(sheet) -> list[tuple[int, str, str, int | None]]:
    result = []
    for row in sheet.iter_rows(min_row=8, values_only=True):
        if row[1] is None:
            continue
        result.append(
            (
                int(row[1]),
                str(row[2]),
                str(row[3]),
                None if row[4] is None else int(row[4]),
            )
        )
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--write", action="store_true")
    mode.add_argument("--check", action="store_true")
    args = parser.parse_args()

    workbook = load_workbook(WORKBOOK, read_only=False, data_only=False)
    sheet = workbook.active
    matching = [row for row in records(sheet) if row[0] == GROUP_ID]

    if args.check:
        if matching != [EXPECTED]:
            raise ValueError("Goal 07 modifier comparator probe is missing or changed")
        print("Goal 07 modifier comparator workbook probe matches.")
        return

    if matching and matching != [EXPECTED]:
        raise ValueError("Goal 07 modifier comparator ID is already owned")
    if not matching:
        row = max(sheet.max_row + 1, 8)
        sheet.cell(row, 1, None)
        for column, value in enumerate(EXPECTED, start=2):
            sheet.cell(row, column, value)
    workbook.save(WORKBOOK)
    print("Authored Goal 07 modifier comparator workbook probe.")


if __name__ == "__main__":
    main()
