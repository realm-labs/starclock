"""Deterministic Goal 11 openpyxl authoring and structural verification."""

from __future__ import annotations

import hashlib
import json
import re
import zipfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKBOOKS = (
    "DivergentUniverse.xlsx",
    "DivergentUniverseBindings.xlsx",
    "DivergentUniverseReview.xlsx",
)
NORMALIZED_ROOT = Path("content-reference/divergent-universe-v1")
MANIFEST_ROOT = Path("content-manifests/divergent-universe-v1")
GENERATED_ROOT = Path("config/divergent-universe-generated")
FIXED_TIME = datetime(2000, 1, 1, tzinfo=timezone.utc)
EXCEL_CELL_LIMIT = 32_767
COMMON_FIELDS = {
    "id",
    "schema_revision",
    "kind",
    "name_en",
    "name_zh_cn",
    "summary_en",
    "summary_zh_cn",
    "ownership",
    "coverage_state",
    "evidence_quality",
    "source_refs",
    "tags",
    "source_id",
}
REFERENCE_SENTINELS = {"", "NotApplicable", "Unspecified"}
HEADER_FILL = PatternFill("solid", fgColor="17365D")
EVEN_FILL = PatternFill("solid", fgColor="EAF2F8")
ODD_FILL = PatternFill("solid", fgColor="FFFFFF")
POLICY_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="B7C9E2"))


def canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def schema(root: Path) -> dict[str, Any]:
    return json.loads(
        (root / GENERATED_ROOT / "schema.lock").read_text(encoding="utf-8")
    )["schema"]


def table_file_map(root: Path, tables: list[dict[str, Any]]) -> dict[str, str]:
    contracts = json.loads(
        (root / MANIFEST_ROOT / "normalized-schema.json").read_text(
            encoding="utf-8"
        )
    )
    result = {
        contract["record_kind"]: contract["file"]
        for contract in contracts["files"]
    }
    if len(result) != len(tables):
        raise ValueError("Sora table-to-normalized-file mapping is incomplete")
    authoring = json.loads(
        (root / MANIFEST_ROOT / "authoring-contract.json").read_text(
            encoding="utf-8"
        )
    )
    normalized_by_workbook = {
        workbook["file"]: set(workbook["normalized_files"])
        for workbook in authoring["workbooks"]
    }
    for table in tables:
        filename = result.get(table["name"])
        workbook = table["source"]["file"]
        if filename not in normalized_by_workbook[workbook]:
            raise ValueError(
                f"{table['name']}: normalized file violates workbook contract"
            )
    return result


def raw_rows(
    root: Path,
    tables: list[dict[str, Any]],
    files: dict[str, str],
) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = {}
    for table in tables:
        filename = files[table["name"]]
        value = json.loads(
            (root / NORMALIZED_ROOT / filename).read_text(encoding="utf-8")
        )
        if not isinstance(value, list):
            raise ValueError(f"{filename} must contain a JSON array")
        result[table["name"]] = value
    return result


def numeric_reference_maps(
    source_rows: dict[str, list[dict[str, Any]]],
) -> dict[str, dict[str, int]]:
    result: dict[str, dict[str, int]] = {}
    for table_name, rows in source_rows.items():
        values = {str(row["id"]): index for index, row in enumerate(rows, 1)}
        source_ids: dict[str, int] = {}
        duplicates: set[str] = set()
        for index, row in enumerate(rows, 1):
            source_id = row.get("source_id")
            if source_id is None:
                continue
            key = str(source_id)
            if key in source_ids:
                duplicates.add(key)
            else:
                source_ids[key] = index
        for key, index in source_ids.items():
            if key not in duplicates and key not in values:
                values[key] = index
        result[table_name] = values
    return result


def is_optional(field_type: Any) -> bool:
    return isinstance(field_type, dict) and "Optional" in field_type


def unwrapped(field_type: Any) -> Any:
    return field_type["Optional"] if is_optional(field_type) else field_type


def is_list(field_type: Any) -> bool:
    value = unwrapped(field_type)
    return isinstance(value, dict) and "List" in value


def scalar_type(field_type: Any) -> str | None:
    value = unwrapped(field_type)
    return value if isinstance(value, str) else None


def reference_target(field_type: Any) -> str | None:
    value = unwrapped(field_type)
    if isinstance(value, dict) and "List" in value:
        value = value["List"]
    if isinstance(value, dict) and "Ref" in value:
        return value["Ref"]["table"]
    return None


def compact_projection(value: Any) -> dict[str, Any]:
    encoded = canonical_json(value).encode("utf-8")
    result: dict[str, Any] = {
        "projection": "ExcelCellLimit",
        "canonical_json_bytes": len(encoded),
        "sha256": hashlib.sha256(encoded).hexdigest(),
    }
    if isinstance(value, (list, dict)):
        result["item_count"] = len(value)
    return result


def payload_value(row: dict[str, Any]) -> str:
    payload: dict[str, Any] = {}
    for name, value in row.items():
        if name in COMMON_FIELDS:
            continue
        encoded = canonical_json(value)
        payload[name] = (
            value if len(encoded) <= EXCEL_CELL_LIMIT else compact_projection(value)
        )
    result = canonical_json(payload)
    if len(result) > EXCEL_CELL_LIMIT:
        result = canonical_json(
            {
                "row_projection": compact_projection(payload),
                "field_projections": {
                    name: compact_projection(value)
                    for name, value in payload.items()
                    if len(canonical_json(value)) > 1_000
                },
            }
        )
    if len(result) > EXCEL_CELL_LIMIT:
        raise ValueError(f"payload projection exceeds Excel cell limit: {len(result)}")
    return result


def reference_keys(value: Any, is_sequence: bool) -> list[str]:
    if value is None:
        return []
    values = value if is_sequence and isinstance(value, list) else [value]
    keys: list[str] = []
    for item in values:
        if isinstance(item, dict) and "source_id" in item:
            item = item["source_id"]
        key = str(item)
        if key not in REFERENCE_SENTINELS:
            keys.append(key)
    return keys


def convert_row(
    table: dict[str, Any],
    row: dict[str, Any],
    numeric_id: int,
    numeric_ids: dict[str, dict[str, int]],
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for field in table["fields"]:
        name = field["name"]
        if name == "id":
            value: Any = numeric_id
        elif name == "stable_key":
            value = row["id"]
        elif name == "payload_json":
            value = payload_value(row)
        else:
            value = row.get(name)
        target = reference_target(field["ty"])
        if target is not None:
            converted = [
                numeric_ids[target][key]
                for key in reference_keys(value, is_list(field["ty"]))
                if key in numeric_ids[target]
            ]
            value = converted if is_list(field["ty"]) else (
                converted[0] if converted else None
            )
        if is_list(field["ty"]):
            value = None if not value else "|".join(str(item) for item in value)
        elif scalar_type(field["ty"]) == "String" and value is not None:
            if isinstance(value, (dict, list)):
                encoded = canonical_json(value)
                value = (
                    encoded
                    if len(encoded) <= EXCEL_CELL_LIMIT
                    else canonical_json(compact_projection(value))
                )
            elif not isinstance(value, str):
                value = str(value)
        elif scalar_type(field["ty"]) == "I32" and value is not None:
            value = int(value)
        if is_optional(field["ty"]) and value == "":
            value = None
        if value is None and not is_optional(field["ty"]):
            raise ValueError(f"{table['name']}.{name} is required")
        if isinstance(value, str) and len(value) > EXCEL_CELL_LIMIT:
            if not is_optional(field["ty"]):
                raise ValueError(
                    f"{table['name']}.{name} exceeds the Excel cell limit"
                )
            value = canonical_json(compact_projection(value))
        result[name] = value
    return result


def workbook_rows(root: Path) -> dict[str, list[dict[str, Any]]]:
    tables = schema(root)["tables"]
    files = table_file_map(root, tables)
    source_rows = raw_rows(root, tables, files)
    numeric_ids = numeric_reference_maps(source_rows)
    return {
        table["name"]: [
            convert_row(table, row, index, numeric_ids)
            for index, row in enumerate(source_rows[table["name"]], 1)
        ]
        for table in tables
    }


def field_columns(sheet: Any) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }


def write_rows(sheet: Any, rows: list[dict[str, Any]]) -> None:
    columns = field_columns(sheet)
    for offset, values in enumerate(rows, 8):
        unknown = sorted(set(values) - set(columns))
        if unknown:
            raise ValueError(f"{sheet.title}: unknown fields {unknown}")
        for field, value in values.items():
            sheet.cell(row=offset, column=columns[field], value=value)


def add_validations(
    sheet: Any,
    table: dict[str, Any],
    enums: dict[str, list[str]],
    last_row: int,
) -> None:
    columns = field_columns(sheet)
    for field in table["fields"]:
        field_type = unwrapped(field["ty"])
        values: list[str] | None = None
        if isinstance(field_type, dict) and "Enum" in field_type:
            values = enums[field_type["Enum"]]
        elif field_type == "Bool":
            values = ["TRUE", "FALSE"]
        if values:
            validation = DataValidation(
                type="list",
                formula1=f'"{",".join(values)}"',
                allow_blank=is_optional(field["ty"]),
            )
            sheet.add_data_validation(validation)
            letter = sheet.cell(row=3, column=columns[field["name"]]).column_letter
            validation.add(f"{letter}8:{letter}{last_row}")


def style_sheet(
    sheet: Any,
    table: dict[str, Any],
    rows: list[dict[str, Any]],
    enums: dict[str, list[str]],
) -> None:
    last_row = max(7 + len(rows), 7)
    last_column = sheet.max_column
    for cell in sheet[3]:
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    for row in sheet.iter_rows(min_row=8, max_row=last_row, max_col=last_column):
        fill = EVEN_FILL if row[0].row % 2 == 0 else ODD_FILL
        if any(isinstance(cell.value, str) and len(cell.value) > 500 for cell in row):
            sheet.row_dimensions[row[0].row].height = 72
        for cell in row:
            cell.fill = copy(fill)
            cell.border = copy(THIN_BORDER)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A8"
    last_letter = sheet.cell(row=3, column=last_column).column_letter
    sheet.auto_filter.ref = f"A3:{last_letter}{last_row}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[3].height = 30
    for column in range(1, last_column + 1):
        values = [
            str(sheet.cell(row=row, column=column).value or "")
            for row in range(2, min(last_row, 200) + 1)
        ]
        width = min(60, max(10, max(map(len, values), default=10) + 2))
        letter = sheet.cell(row=3, column=column).column_letter
        sheet.column_dimensions[letter].width = width
    if rows:
        add_validations(sheet, table, enums, last_row)
        evidence_column = field_columns(sheet).get("evidence_quality")
        if evidence_column:
            letter = sheet.cell(row=3, column=evidence_column).column_letter
            sheet.conditional_formatting.add(
                f"{letter}8:{letter}{last_row}",
                FormulaRule(
                    formula=[f'${letter}8="ProjectPolicy"'],
                    fill=copy(POLICY_FILL),
                ),
            )


def normalize_archive(path: Path) -> None:
    temporary = path.with_suffix(f"{path.suffix}.canonical")
    with zipfile.ZipFile(path, "r") as source:
        members = [(name, source.read(name)) for name in sorted(source.namelist())]
    with zipfile.ZipFile(
        temporary, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as target:
        for name, payload in members:
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2000-01-01T00:00:00Z\g<2>",
                    payload,
                )
            info = zipfile.ZipInfo(name, date_time=(2000, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0
            target.writestr(info, payload, compress_type=zipfile.ZIP_DEFLATED)
    temporary.replace(path)


def prepare_workbook(
    template: Path,
    target: Path,
    tables: list[dict[str, Any]],
    rows: dict[str, list[dict[str, Any]]],
    enums: dict[str, list[str]],
) -> dict[str, int]:
    workbook = load_workbook(template)
    workbook_tables = [
        table for table in tables if table["source"]["file"] == template.name
    ]
    expected_sheets = [table["source"]["sheet"] for table in workbook_tables]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"{template.name}: sheet order differs from schema")
    counts: dict[str, int] = {}
    for table in workbook_tables:
        sheet = workbook[table["source"]["sheet"]]
        values = rows[table["name"]]
        write_rows(sheet, values)
        style_sheet(sheet, table, values, enums)
        counts[table["name"]] = len(values)
    workbook.properties.creator = "Starclock Goal 11 openpyxl bootstrap"
    workbook.properties.lastModifiedBy = "Starclock Goal 11 openpyxl bootstrap"
    workbook.properties.created = FIXED_TIME
    workbook.properties.modified = FIXED_TIME
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.save(target)
    normalize_archive(target)
    return counts


def author(root: Path, output: Path) -> dict[str, int]:
    output.mkdir(parents=True, exist_ok=True)
    targets = [output / name for name in WORKBOOKS]
    existing = [target for target in targets if target.exists()]
    if existing:
        raise FileExistsError(
            "refusing to overwrite authored workbook(s): "
            + ", ".join(str(target) for target in existing)
        )
    schema_value = schema(root)
    tables = schema_value["tables"]
    rows = workbook_rows(root)
    enums = {enum["name"]: enum["values"] for enum in schema_value["enums"]}
    template_root = root / GENERATED_ROOT / "templates"
    counts: dict[str, int] = {}
    for target in targets:
        counts.update(
            prepare_workbook(
                template_root / target.name,
                target,
                tables,
                rows,
                enums,
            )
        )
    verify(root, output, counts)
    return counts


def verify(
    root: Path,
    directory: Path,
    expected_counts: dict[str, int] | None = None,
) -> dict[str, int]:
    schema_value = schema(root)
    tables = schema_value["tables"]
    expected_rows = workbook_rows(root)
    counts: dict[str, int] = {}
    template_root = root / GENERATED_ROOT / "templates"
    for name in WORKBOOKS:
        workbook = load_workbook(directory / name, data_only=False)
        template = load_workbook(template_root / name, data_only=False)
        workbook_tables = [
            table for table in tables if table["source"]["file"] == name
        ]
        expected_sheets = [table["source"]["sheet"] for table in workbook_tables]
        if workbook.sheetnames != expected_sheets:
            raise ValueError(f"{name}: missing or reordered sheet")
        for table in workbook_tables:
            sheet_name = table["source"]["sheet"]
            sheet = workbook[sheet_name]
            template_sheet = template[sheet_name]
            for row in range(1, 8):
                actual = [cell.value for cell in sheet[row]]
                expected = [cell.value for cell in template_sheet[row]]
                if actual != expected:
                    raise ValueError(
                        f"{name}/{sheet_name}: Sora metadata row {row} drifted"
                    )
            count = max(0, sheet.max_row - 7)
            counts[table["name"]] = count
            if count != len(expected_rows[table["name"]]):
                raise ValueError(f"{name}/{sheet_name}: row count differs")
            if (
                sheet.freeze_panes != "A8"
                or not sheet.auto_filter.ref
                or sheet.sheet_view.showGridLines
            ):
                raise ValueError(
                    f"{name}/{sheet_name}: authoring affordances missing"
                )
            columns = field_columns(sheet)
            for offset, expected_row in enumerate(
                expected_rows[table["name"]], start=8
            ):
                actual_row = {
                    field: sheet.cell(row=offset, column=column).value
                    for field, column in columns.items()
                }
                if actual_row != expected_row:
                    raise ValueError(f"{name}/{sheet_name}: row {offset} differs")
            for column in range(1, sheet.max_column + 1):
                letter = sheet.cell(row=3, column=column).column_letter
                width = sheet.column_dimensions[letter].width
                if width is None or not 10 <= width <= 60:
                    raise ValueError(
                        f"{name}/{sheet_name}/{letter}: invalid width {width}"
                    )
            for row in sheet.iter_rows(min_row=8):
                for cell in row:
                    if cell.data_type in (TYPE_FORMULA, TYPE_ERROR):
                        raise ValueError(
                            f"{name}/{sheet_name}/{cell.coordinate}: "
                            "formula or Excel error forbidden"
                        )
                    if cell.value is not None and not cell.alignment.wrap_text:
                        raise ValueError(
                            f"{name}/{sheet_name}/{cell.coordinate}: "
                            "data wrapping missing"
                        )
        template.close()
        workbook.close()
    if expected_counts is not None and counts != expected_counts:
        raise ValueError("workbook row counts changed after save/reload")
    return counts


def semantic_digest(directory: Path) -> str:
    payload: list[Any] = []
    for name in WORKBOOKS:
        workbook = load_workbook(directory / name, data_only=False)
        for sheet in workbook.worksheets:
            cells = [
                [[cell.value, cell.data_type] for cell in row]
                for row in sheet.iter_rows()
            ]
            widths = {
                key: value.width
                for key, value in sheet.column_dimensions.items()
                if value.width is not None
            }
            heights = {
                key: value.height
                for key, value in sheet.row_dimensions.items()
                if value.height is not None
            }
            validations = sorted(
                (
                    str(validation.sqref),
                    validation.type,
                    validation.formula1,
                    validation.allow_blank,
                )
                for validation in sheet.data_validations.dataValidation
            )
            payload.append(
                [
                    name,
                    sheet.title,
                    cells,
                    str(sheet.freeze_panes),
                    sheet.auto_filter.ref,
                    widths,
                    heights,
                    validations,
                ]
            )
        workbook.close()
    encoded = canonical_json(payload).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()
