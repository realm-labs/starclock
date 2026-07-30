"""Deterministic openpyxl authoring and QA for Goal 16 workbooks."""

from __future__ import annotations

import hashlib
import json
import re
import zipfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

if openpyxl.__version__ != "3.1.5":
    raise RuntimeError(f"openpyxl 3.1.5 required, got {openpyxl.__version__}")

FIXED_TIME = datetime(2000, 1, 1, tzinfo=timezone.utc)
TITLE_FILL = PatternFill("solid", fgColor="12372A")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
META_FILL = PatternFill("solid", fgColor="D9EAF7")
EVEN_FILL = PatternFill("solid", fgColor="EAF4EF")
ODD_FILL = PatternFill("solid", fgColor="FFFFFF")
QUALITY_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(name="Aptos Display", color="FFFFFF", bold=True)
HEADER_FONT = Font(name="Aptos", color="FFFFFF", bold=True)
META_FONT = Font(name="Aptos", color="17365D", italic=True, size=9)
BODY_FONT = Font(name="Aptos", size=10)
THIN_BORDER = Border(bottom=Side(style="thin", color="B7C9C1"))
COMMON_FIELDS = (
    "schema_revision",
    "kind",
    "name_en",
    "name_zh_cn",
    "summary_en",
    "summary_zh_cn",
    "profile_ids",
    "ownership",
    "coverage_state",
    "evidence_quality",
    "mechanism_quality",
    "manifest_record_ids",
    "source_refs",
    "tags",
)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def compact(value: Any) -> str:
    if isinstance(value, str):
        return value
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def sheet_name(file_name: str) -> str:
    parts = re.sub(r"\.json$", "", file_name).split("-")
    return "".join(part[:1].upper() + part[1:] for part in parts)[:31]


def normalized_files(root: Path) -> list[str]:
    schema = read_json(
        root
        / "content-manifests/galactic-baseballer-v1/normalized-schema.json"
    )
    return [entry["file"] for entry in schema["files"]]


def workbook_contract(root: Path) -> list[dict[str, Any]]:
    contract = read_json(
        root
        / "content-manifests/galactic-baseballer-v1/authoring-contract.json"
    )
    return contract["workbooks"]


def fields_for_rows(rows: list[dict[str, Any]]) -> list[str]:
    discovered = set().union(*(row.keys() for row in rows)) if rows else set()
    discovered.discard("id")
    fields = [field for field in COMMON_FIELDS if field in discovered]
    fields.extend(sorted(discovered - set(fields)))
    return fields


def authored_value(field: str, value: Any) -> str | None:
    if field == "source_refs":
        return compact([reference["source_id"] for reference in value])
    if value == "":
        return None
    return compact(value)


def authored_rows(
    root: Path,
    file_name: str,
) -> tuple[list[str], list[dict[str, Any]]]:
    source = (
        root / "content-reference/galactic-baseballer-v1" / file_name
    )
    rows = read_json(source)
    fields = fields_for_rows(rows)
    authored: list[dict[str, Any]] = []
    for private_id, row in enumerate(rows, start=1):
        values: dict[str, Any] = {
            "id": private_id,
            "stable_key": row["id"],
        }
        for field in fields:
            if field not in row:
                values[field] = None
            else:
                values[field] = authored_value(field, row[field])
        for field, value in values.items():
            if isinstance(value, str) and len(value) > 32767:
                raise ValueError(
                    f"{file_name}/{row['id']}/{field}: "
                    f"{len(value)} exceeds Excel cell limit"
                )
        authored.append(values)
    return fields, authored


def add_validation(
    sheet,
    field: str,
    values: list[str],
    maximum_row: int,
) -> None:
    columns = {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }
    column = columns.get(field)
    if column is None:
        return
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(values)}"',
        allow_blank=False,
    )
    validation.error = f"Choose one of: {', '.join(values)}"
    validation.errorTitle = f"Invalid {field}"
    validation.prompt = f"Goal 16 {field}"
    validation.promptTitle = "Starclock authoring contract"
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(
        f"{sheet.cell(row=8, column=column).coordinate}:"
        f"{sheet.cell(row=max(8, maximum_row), column=column).coordinate}"
    )


def style_sheet(sheet, row_count: int) -> None:
    maximum_row = max(7 + row_count, 7)
    maximum_column = sheet.max_column
    for row_number in range(1, 8):
        sheet.row_dimensions[row_number].height = 22
    sheet.row_dimensions[1].height = 26
    sheet.row_dimensions[3].height = 32
    for cell in sheet[1]:
        cell.fill = copy(TITLE_FILL)
        cell.font = copy(TITLE_FONT)
        cell.alignment = Alignment(vertical="center")
    for row_number in (2, 4, 5, 6, 7):
        for cell in sheet[row_number]:
            cell.fill = copy(META_FILL)
            cell.font = copy(META_FONT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in sheet[3]:
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    if row_count:
        for row in sheet.iter_rows(
            min_row=8,
            max_row=maximum_row,
            max_col=maximum_column,
        ):
            fill = EVEN_FILL if row[0].row % 2 == 0 else ODD_FILL
            sheet.row_dimensions[row[0].row].height = 30
            for cell in row:
                cell.fill = copy(fill)
                cell.border = copy(THIN_BORDER)
                cell.font = copy(BODY_FONT)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A8"
    last_column = sheet.cell(row=3, column=maximum_column).column_letter
    sheet.auto_filter.ref = f"A3:{last_column}{maximum_row}"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    for column in range(1, maximum_column + 1):
        samples = [
            str(sheet.cell(row=row, column=column).value or "")
            for row in range(1, min(maximum_row, 107) + 1)
        ]
        width = min(
            48,
            max(
                11,
                max((min(len(value), 46) for value in samples), default=11)
                + 2,
            ),
        )
        letter = sheet.cell(row=3, column=column).column_letter
        sheet.column_dimensions[letter].width = width
    add_validation(
        sheet,
        "ownership",
        ["Departure", "DemonKing", "SharedBase", "Shared"],
        maximum_row,
    )
    add_validation(
        sheet,
        "coverage_state",
        ["Cataloged", "Researched", "DataReady", "Blocked", "EvidenceOnly"],
        maximum_row,
    )
    add_validation(
        sheet,
        "evidence_quality",
        [
            "ExactStructured",
            "ExactPublicText",
            "Observed",
            "ApproximateFromReleasedText",
            "ProjectPolicy",
        ],
        maximum_row,
    )
    add_validation(
        sheet,
        "mechanism_quality",
        [
            "ExactProgram",
            "ExactRelationship",
            "ObservedBehavior",
            "IdentityCrossCheck",
            "PolicyBoundary",
            "ContextOnly",
        ],
        maximum_row,
    )
    columns = {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }
    quality_column = columns.get("evidence_quality")
    if quality_column is not None and row_count:
        letter = sheet.cell(row=8, column=quality_column).column_letter
        target = f"{letter}8:{letter}{maximum_row}"
        sheet.conditional_formatting.add(
            target,
            FormulaRule(
                formula=[
                    f'OR({letter}8="ProjectPolicy",'
                    f'{letter}8="ApproximateFromReleasedText")'
                ],
                fill=copy(QUALITY_FILL),
            ),
        )


def normalize_archive(file: Path) -> None:
    temporary = file.with_suffix(f"{file.suffix}.canonical")
    with zipfile.ZipFile(file, "r") as source:
        members = [(name, source.read(name)) for name in sorted(source.namelist())]
    with zipfile.ZipFile(
        temporary,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as target:
        for name, payload in members:
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:(?:created|modified)[^>]*>)"
                    rb"[^<]*(</dcterms:(?:created|modified)>)",
                    rb"\g<1>2000-01-01T00:00:00Z\g<2>",
                    payload,
                )
            info = zipfile.ZipInfo(name, date_time=(2000, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0
            target.writestr(
                info,
                payload,
                compress_type=zipfile.ZIP_DEFLATED,
                compresslevel=9,
            )
    temporary.replace(file)


def expected_layout(
    root: Path,
) -> dict[str, list[str]]:
    return {
        workbook["file"]: list(workbook["normalized_files"])
        for workbook in workbook_contract(root)
    }


def author(
    root: Path,
    output: Path,
    templates: Path,
) -> dict[str, int]:
    output.mkdir(parents=True, exist_ok=True)
    layout = expected_layout(root)
    targets = [output / name for name in layout]
    existing = [target for target in targets if target.exists()]
    if existing:
        raise FileExistsError(
            "refusing to overwrite authored workbook(s): "
            + ", ".join(map(str, existing))
        )
    counts: dict[str, int] = {}
    for workbook_name, files in layout.items():
        template = templates / workbook_name
        if not template.is_file():
            raise FileNotFoundError(f"Sora template missing: {template}")
        workbook = load_workbook(template, data_only=False)
        expected_sheets = [sheet_name(file_name) for file_name in files]
        if workbook.sheetnames != expected_sheets:
            raise ValueError(
                f"{workbook_name}: Sora template sheet partition drift"
            )
        for file_name in files:
            fields, rows = authored_rows(root, file_name)
            columns = ["id", "stable_key", *fields]
            sheet = workbook[sheet_name(file_name)]
            template_fields = [
                cell.value for cell in sheet[3] if cell.value is not None
            ]
            if template_fields != ["#field", *columns]:
                raise ValueError(
                    f"{workbook_name}/{sheet.title}: "
                    "Sora template field drift"
                )
            if sheet.max_row != 7:
                raise ValueError(
                    f"{workbook_name}/{sheet.title}: "
                    "Sora template contains authored rows"
                )
            for values in rows:
                sheet.append([values[column] for column in columns])
            style_sheet(sheet, len(rows))
            counts[f"{workbook_name}/{sheet.title}"] = len(rows)
        workbook.properties.creator = "Starclock Goal 16 openpyxl authoring"
        workbook.properties.lastModifiedBy = (
            "Starclock Goal 16 openpyxl authoring"
        )
        workbook.properties.created = FIXED_TIME
        workbook.properties.modified = FIXED_TIME
        workbook.calculation.fullCalcOnLoad = False
        workbook.calculation.forceFullCalc = False
        target = output / workbook_name
        workbook.save(target)
        workbook.close()
        normalize_archive(target)
    verify(root, output, templates, counts)
    return counts


def verify(
    root: Path,
    directory: Path,
    templates: Path,
    authored_counts: dict[str, int] | None = None,
) -> dict[str, int]:
    layout = expected_layout(root)
    observed: dict[str, int] = {}
    source_ids = {
        row["id"]
        for row in read_json(
            root / "content-reference/galactic-baseballer-v1/sources.json"
        )
    }
    for workbook_name, files in layout.items():
        template = load_workbook(
            templates / workbook_name,
            read_only=True,
            data_only=False,
        )
        workbook = load_workbook(
            directory / workbook_name,
            data_only=False,
        )
        expected_sheets = [sheet_name(file_name) for file_name in files]
        if (
            workbook.sheetnames != expected_sheets
            or template.sheetnames != expected_sheets
        ):
            raise ValueError(f"{workbook_name}: missing or reordered sheet")
        for file_name, sheet_name_value in zip(files, expected_sheets):
            sheet = workbook[sheet_name_value]
            template_sheet = template[sheet_name_value]
            fields, expected_rows = authored_rows(root, file_name)
            columns = ["id", "stable_key", *fields]
            for row_number in range(1, 8):
                observed_values = [
                    sheet.cell(row=row_number, column=column).value
                    for column in range(1, len(columns) + 2)
                ]
                template_values = [
                    template_sheet.cell(
                        row=row_number,
                        column=column,
                    ).value
                    for column in range(1, len(columns) + 2)
                ]
                if observed_values != template_values:
                    raise ValueError(
                        f"{workbook_name}/{sheet_name_value}: "
                        f"Sora metadata row {row_number} drift"
                    )
            key = f"{workbook_name}/{sheet_name_value}"
            count = len(expected_rows)
            observed[key] = count
            if sheet.max_row != 7 + count:
                raise ValueError(
                    f"{key}: expected {count} rows, got {sheet.max_row - 7}"
                )
            for offset, expected in enumerate(expected_rows, start=8):
                values = [
                    sheet.cell(row=offset, column=column).value
                    for column in range(1, len(columns) + 1)
                ]
                expected_values = [expected[column] for column in columns]
                if values != expected_values:
                    mismatches = [
                        (
                            columns[index],
                            expected_values[index],
                            values[index],
                        )
                        for index in range(len(columns))
                        if expected_values[index] != values[index]
                    ]
                    raise ValueError(
                        f"{key}/row {offset}: semantic drift "
                        f"{mismatches[:3]}"
                    )
                if "source_refs" in fields:
                    references = json.loads(
                        expected["source_refs"] or "[]"
                    )
                    if any(reference not in source_ids for reference in references):
                        raise ValueError(
                            f"{key}/row {offset}: unknown source reference"
                        )
            if sheet.freeze_panes != "A8" or not sheet.auto_filter.ref:
                raise ValueError(f"{key}: authoring affordances missing")
            for cell in sheet[3]:
                if cell.value is not None and not cell.alignment.wrap_text:
                    raise ValueError(
                        f"{key}/{cell.coordinate}: header wrap missing"
                    )
            for column in range(1, sheet.max_column + 1):
                letter = sheet.cell(row=3, column=column).column_letter
                width = sheet.column_dimensions[letter].width
                if width is None or not 11 <= width <= 48:
                    raise ValueError(f"{key}/{letter}: width {width} invalid")
            for row in sheet.iter_rows(min_row=8):
                for cell in row:
                    if cell.data_type in (TYPE_FORMULA, TYPE_ERROR):
                        raise ValueError(
                            f"{key}/{cell.coordinate}: formula/error forbidden"
                        )
                    if isinstance(cell.value, str) and len(cell.value) > 32767:
                        raise ValueError(
                            f"{key}/{cell.coordinate}: Excel text overflow"
                        )
        workbook.close()
        template.close()
    if authored_counts is not None and observed != authored_counts:
        raise ValueError("workbook counts changed after save/reload")
    return observed


def semantic_digest(directory: Path) -> str:
    result = hashlib.sha256()
    for workbook_name in sorted(path.name for path in directory.glob("*.xlsx")):
        workbook = load_workbook(
            directory / workbook_name,
            read_only=True,
            data_only=False,
        )
        result.update(f"{workbook_name}\n".encode())
        for sheet in workbook.worksheets:
            result.update(f"{sheet.title}\n".encode())
            for row in sheet.iter_rows(values_only=True):
                result.update((compact(list(row)) + "\n").encode("utf-8"))
        workbook.close()
    return result.hexdigest()


def byte_digests(directory: Path) -> dict[str, str]:
    return {
        path.name: hashlib.sha256(path.read_bytes()).hexdigest()
        for path in sorted(directory.glob("*.xlsx"))
    }
