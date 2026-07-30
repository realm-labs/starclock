"""Deterministic Goal 10 openpyxl authoring and structural verification."""

from __future__ import annotations

import hashlib
import json
import re
import zipfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKBOOKS = (
    "UnknowableDomain.xlsx",
    "UnknowableDomainBindings.xlsx",
    "UnknowableDomainReview.xlsx",
)
NORMALIZED_ROOT = Path("content-reference/unknowable-domain-v1")
FIXED_TIME = datetime(2000, 1, 1, tzinfo=timezone.utc)
HEADER_FILL = PatternFill("solid", fgColor="17365D")
EVEN_FILL = PatternFill("solid", fgColor="EAF2F8")
ODD_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="B7C9E2"))
POLICY_FILL = PatternFill("solid", fgColor="FFF2CC")

RENAMED_FIELDS = {
    ("UnknowableDomainArea", "default_alignment_id"): "default_alignment",
    ("UnknowableDomainArea", "extra_layer_key"): "extra_layer_id",
    ("UnknowableDomainWorkbenchFunction", "currency_key"): "currency_id",
    ("UnknowableDomainServiceOfferRule", "service_key"): "service_id",
    ("UnknowableDomainBossChoice", "pool_ids"): "pool_id",
}
REFERENCE_SENTINELS = {"", "NotApplicable", "Unspecified"}


def schema(root: Path) -> dict[str, Any]:
    return json.loads(
        (root / "config/unknowable-domain-generated/schema.lock").read_text(
            encoding="utf-8"
        )
    )["schema"]


def table_file_map(root: Path, tables: list[dict[str, Any]]) -> dict[str, str]:
    contract = json.loads(
        (
            root
            / "content-manifests/unknowable-domain-v1/authoring-contract.json"
        ).read_text(encoding="utf-8")
    )
    normalized_by_workbook = {
        workbook["file"]: workbook["normalized_files"]
        for workbook in contract["workbooks"]
    }
    generator = (
        root / "tools/unknowable-domain-reference/generate-sora-schema.mjs"
    ).read_text(encoding="utf-8")
    result = dict(
        re.findall(
            r'name:\s*"([^"]+)",\s*\n\s*sheet:\s*"[^"]+",\s*\n'
            r'\s*normalized:\s*"([^"]+)"',
            generator,
        )
    )
    if len(result) != len(tables):
        raise ValueError("Sora table-to-normalized-file mapping is incomplete")
    for table in tables:
        workbook = table["source"]["file"]
        if result[table["name"]] not in normalized_by_workbook[workbook]:
            raise ValueError(
                f"{table['name']}: normalized file violates workbook contract"
            )
    return result


def table_rows(
    root: Path,
    table_name: str,
    files: dict[str, str],
) -> list[dict[str, Any]]:
    value = json.loads(
        (root / NORMALIZED_ROOT / files[table_name]).read_text(encoding="utf-8")
    )
    if not isinstance(value, list):
        raise ValueError(f"{files[table_name]} must contain a JSON array")
    return value


def raw_rows(
    root: Path,
    tables: list[dict[str, Any]],
    files: dict[str, str],
) -> dict[str, list[dict[str, Any]]]:
    return {
        table["name"]: table_rows(root, table["name"], files) for table in tables
    }


def numeric_reference_maps(
    source_rows: dict[str, list[dict[str, Any]]],
) -> dict[str, dict[str, int]]:
    result: dict[str, dict[str, int]] = {}
    for table_name, rows in source_rows.items():
        values = {str(row["id"]): index for index, row in enumerate(rows, start=1)}
        source_ids: dict[str, int] = {}
        duplicates: set[str] = set()
        for index, row in enumerate(rows, start=1):
            if (source_id := row.get("source_id")) is None:
                continue
            key = str(source_id)
            if key in source_ids:
                duplicates.add(key)
            else:
                source_ids[key] = index
        for key, index in source_ids.items():
            if key not in duplicates and key not in values:
                values[key] = index
        result[table_name] = values
    return result


def workbook_rows(root: Path) -> dict[str, list[dict[str, Any]]]:
    tables = schema(root)["tables"]
    files = table_file_map(root, tables)
    source_rows = raw_rows(root, tables, files)
    numeric_ids = numeric_reference_maps(source_rows)
    return {
        table["name"]: [
            convert_row(table, row, index, numeric_ids)
            for index, row in enumerate(source_rows[table["name"]], start=1)
        ]
        for table in tables
    }


def convert_row(
    table: dict[str, Any],
    row: dict[str, Any],
    numeric_id: int,
    numeric_ids: dict[str, dict[str, int]],
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for field in table["fields"]:
        name = field["name"]
        if name == "id":
            value: Any = numeric_id
        elif name == "stable_key":
            value = row["id"]
        elif name == "source_refs":
            value = [record["source_id"] for record in row.get("source_refs", [])]
        elif name.endswith("_json"):
            value = row.get(name.removesuffix("_json"))
            value = canonical_json(value) if value is not None else None
        else:
            source_name = RENAMED_FIELDS.get((table["name"], name), name)
            value = row.get(source_name)
        reference = reference_target(field["ty"])
        if reference is not None:
            if value is None or (
                is_optional(field["ty"]) and str(value) in REFERENCE_SENTINELS
            ):
                value = None
            else:
                try:
                    value = numeric_ids[reference][str(value)]
                except KeyError as error:
                    raise ValueError(
                        f"{table['name']}.{name} references unknown "
                        f"{reference} key {value}"
                    ) from error
        if is_list(field["ty"]):
            value = None if not value else "|".join(str(item) for item in value)
        elif scalar_type(field["ty"]) == "String" and value is not None:
            if isinstance(value, (dict, list)):
                value = canonical_json(value)
            elif not isinstance(value, str):
                value = str(value)
        elif scalar_type(field["ty"]) == "I32" and value is not None:
            value = int(value)
        if is_optional(field["ty"]) and value == "":
            value = None
        if value is None and not is_optional(field["ty"]):
            raise ValueError(f"{table['name']}.{name} is required")
        if isinstance(value, str) and len(value) > 32767:
            raise ValueError(f"{table['name']}.{name} exceeds the Excel cell limit")
        result[name] = value
    return result


def canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def is_optional(field_type: Any) -> bool:
    return isinstance(field_type, dict) and "Optional" in field_type


def unwrapped(field_type: Any) -> Any:
    return field_type["Optional"] if is_optional(field_type) else field_type


def is_list(field_type: Any) -> bool:
    return isinstance(unwrapped(field_type), dict) and "List" in unwrapped(field_type)


def scalar_type(field_type: Any) -> str | None:
    value = unwrapped(field_type)
    return value if isinstance(value, str) else None


def reference_target(field_type: Any) -> str | None:
    value = unwrapped(field_type)
    if isinstance(value, dict) and "Ref" in value:
        return value["Ref"]["table"]
    return None


def field_columns(sheet: Any) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }


def write_rows(sheet: Any, rows: list[dict[str, Any]]) -> None:
    columns = field_columns(sheet)
    for offset, values in enumerate(rows, start=8):
        unknown = sorted(set(values) - set(columns))
        if unknown:
            raise ValueError(f"{sheet.title}: unknown fields {unknown}")
        for field, value in values.items():
            sheet.cell(row=offset, column=columns[field], value=value)


def style_sheet(
    sheet: Any,
    table: dict[str, Any],
    rows: list[dict[str, Any]],
    enums: dict[str, list[str]],
) -> None:
    maximum_row = max(7 + len(rows), 7)
    maximum_column = sheet.max_column
    for cell in sheet[3]:
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    for row in sheet.iter_rows(
        min_row=8,
        max_row=maximum_row,
        max_col=maximum_column,
    ):
        fill = EVEN_FILL if row[0].row % 2 == 0 else ODD_FILL
        if any(
            isinstance(cell.value, str) and len(cell.value) > 500
            for cell in row
        ):
            sheet.row_dimensions[row[0].row].height = 72
        for cell in row:
            cell.fill = copy(fill)
            cell.border = copy(THIN_BORDER)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A8"
    last_column = sheet.cell(row=3, column=maximum_column).column_letter
    sheet.auto_filter.ref = f"A3:{last_column}{maximum_row}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[3].height = 30
    for column in range(1, maximum_column + 1):
        values = [
            str(sheet.cell(row=row, column=column).value or "")
            for row in range(2, min(maximum_row, 200) + 1)
        ]
        width = min(
            60,
            max(10, max((len(value) for value in values), default=10) + 2),
        )
        letter = sheet.cell(row=3, column=column).column_letter
        sheet.column_dimensions[letter].width = width
    if rows:
        add_validations(sheet, table, enums, 8, maximum_row)
        evidence_column = field_columns(sheet).get("evidence_quality")
        if evidence_column:
            letter = sheet.cell(row=3, column=evidence_column).column_letter
            sheet.conditional_formatting.add(
                f"{letter}8:{letter}{maximum_row}",
                FormulaRule(
                    formula=[f'${letter}8="ProjectPolicy"'],
                    fill=copy(POLICY_FILL),
                ),
            )


def add_validations(
    sheet: Any,
    table: dict[str, Any],
    enums: dict[str, list[str]],
    first_row: int,
    last_row: int,
) -> None:
    columns = field_columns(sheet)
    for field in table["fields"]:
        field_type = unwrapped(field["ty"])
        values: list[str] | None = None
        if isinstance(field_type, dict) and "Enum" in field_type:
            values = enums[field_type["Enum"]]
        elif field_type == "Bool":
            values = ["TRUE", "FALSE"]
        if values:
            validation = DataValidation(
                type="list",
                formula1=f'"{",".join(values)}"',
                allow_blank=is_optional(field["ty"]),
            )
            sheet.add_data_validation(validation)
            letter = sheet.cell(row=3, column=columns[field["name"]]).column_letter
            validation.add(f"{letter}{first_row}:{letter}{last_row}")


def prepare_workbook(
    template: Path,
    target: Path,
    tables: list[dict[str, Any]],
    rows: dict[str, list[dict[str, Any]]],
    enums: dict[str, list[str]],
) -> dict[str, int]:
    workbook = load_workbook(template)
    workbook_tables = [
        table for table in tables if table["source"]["file"] == template.name
    ]
    expected_sheets = [table["source"]["sheet"] for table in workbook_tables]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"{template.name}: sheet order differs from schema")
    counts: dict[str, int] = {}
    for table in workbook_tables:
        sheet = workbook[table["source"]["sheet"]]
        table_rows_to_write = rows[table["name"]]
        write_rows(sheet, table_rows_to_write)
        style_sheet(sheet, table, table_rows_to_write, enums)
        counts[table["name"]] = len(table_rows_to_write)
    workbook.properties.creator = "Starclock Goal 10 openpyxl bootstrap"
    workbook.properties.lastModifiedBy = "Starclock Goal 10 openpyxl bootstrap"
    workbook.properties.created = FIXED_TIME
    workbook.properties.modified = FIXED_TIME
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.save(target)
    normalize_archive(target)
    return counts


def author(root: Path, output: Path) -> dict[str, int]:
    output.mkdir(parents=True, exist_ok=True)
    targets = [output / name for name in WORKBOOKS]
    existing = [target for target in targets if target.exists()]
    if existing:
        raise FileExistsError(
            "refusing to overwrite authored workbook(s): "
            + ", ".join(str(target) for target in existing)
        )
    schema_value = schema(root)
    tables = schema_value["tables"]
    rows = workbook_rows(root)
    enums = {enum["name"]: enum["values"] for enum in schema_value["enums"]}
    counts: dict[str, int] = {}
    template_root = root / "config/unknowable-domain-generated/templates"
    for target in targets:
        counts.update(
            prepare_workbook(
                template_root / target.name,
                target,
                tables,
                rows,
                enums,
            )
        )
    verify(root, output, counts)
    return counts


def verify(
    root: Path,
    directory: Path,
    expected_counts: dict[str, int] | None = None,
) -> dict[str, int]:
    schema_value = schema(root)
    tables = schema_value["tables"]
    expected_rows = workbook_rows(root)
    counts: dict[str, int] = {}
    template_root = root / "config/unknowable-domain-generated/templates"
    for name in WORKBOOKS:
        workbook = load_workbook(directory / name, data_only=False)
        template = load_workbook(template_root / name, data_only=False)
        workbook_tables = [
            table for table in tables if table["source"]["file"] == name
        ]
        expected_sheets = [table["source"]["sheet"] for table in workbook_tables]
        if workbook.sheetnames != expected_sheets:
            raise ValueError(f"{name}: missing or reordered sheet")
        for table in workbook_tables:
            sheet_name = table["source"]["sheet"]
            sheet = workbook[sheet_name]
            template_sheet = template[sheet_name]
            for row in range(1, 8):
                actual = [cell.value for cell in sheet[row]]
                expected = [cell.value for cell in template_sheet[row]]
                if actual != expected:
                    raise ValueError(
                        f"{name}/{sheet_name}: Sora metadata row {row} drifted"
                    )
            count = max(0, sheet.max_row - 7)
            counts[table["name"]] = count
            if count != len(expected_rows[table["name"]]):
                raise ValueError(f"{name}/{sheet_name}: row count differs")
            if sheet.freeze_panes != "A8" or not sheet.auto_filter.ref:
                raise ValueError(
                    f"{name}/{sheet_name}: authoring affordances missing"
                )
            columns = field_columns(sheet)
            for offset, expected_row in enumerate(
                expected_rows[table["name"]],
                start=8,
            ):
                actual_row = {
                    field: sheet.cell(row=offset, column=column).value
                    for field, column in columns.items()
                }
                if actual_row != expected_row:
                    raise ValueError(f"{name}/{sheet_name}: row {offset} differs")
            for column in range(1, sheet.max_column + 1):
                letter = sheet.cell(row=3, column=column).column_letter
                width = sheet.column_dimensions[letter].width
                if width is None or not 10 <= width <= 60:
                    raise ValueError(
                        f"{name}/{sheet_name}/{letter}: invalid width {width}"
                    )
            for row in sheet.iter_rows(min_row=8):
                has_long_value = any(
                    isinstance(cell.value, str) and len(cell.value) > 500
                    for cell in row
                )
                if has_long_value and sheet.row_dimensions[row[0].row].height != 72:
                    raise ValueError(
                        f"{name}/{sheet_name}/row {row[0].row}: "
                        "long-value height cap missing"
                    )
                for cell in row:
                    if cell.data_type in (TYPE_FORMULA, TYPE_ERROR):
                        raise ValueError(
                            f"{name}/{sheet_name}/{cell.coordinate}: "
                            "formula or Excel error forbidden"
                        )
                    if cell.value is not None and not cell.alignment.wrap_text:
                        raise ValueError(
                            f"{name}/{sheet_name}/{cell.coordinate}: "
                            "data wrapping missing"
                        )
        template.close()
        workbook.close()
    if expected_counts is not None and counts != expected_counts:
        raise ValueError("workbook row counts changed after save/reload")
    return counts


def normalize_archive(path: Path) -> None:
    """Canonicalize ZIP metadata so equivalent workbooks are byte-identical."""
    temporary = path.with_suffix(f"{path.suffix}.canonical")
    with zipfile.ZipFile(path, "r") as source:
        members = [(name, source.read(name)) for name in sorted(source.namelist())]
    with zipfile.ZipFile(
        temporary,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as target:
        for name, payload in members:
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2000-01-01T00:00:00Z\g<2>",
                    payload,
                )
            info = zipfile.ZipInfo(name, date_time=(2000, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0
            target.writestr(
                info,
                payload,
                compress_type=zipfile.ZIP_DEFLATED,
                compresslevel=9,
            )
    temporary.replace(path)


def semantic_digest(directory: Path) -> str:
    payload: list[Any] = []
    for name in WORKBOOKS:
        workbook = load_workbook(directory / name, data_only=False)
        for sheet in workbook.worksheets:
            cells = [
                [[cell.value, cell.data_type] for cell in row]
                for row in sheet.iter_rows()
            ]
            widths = {
                key: value.width
                for key, value in sheet.column_dimensions.items()
                if value.width is not None
            }
            heights = {
                key: value.height
                for key, value in sheet.row_dimensions.items()
                if value.height is not None
            }
            validations = sorted(
                (
                    str(validation.sqref),
                    validation.type,
                    validation.formula1,
                    validation.allow_blank,
                )
                for validation in sheet.data_validations.dataValidation
            )
            payload.append(
                [
                    name,
                    sheet.title,
                    cells,
                    str(sheet.freeze_panes),
                    sheet.auto_filter.ref,
                    widths,
                    heights,
                    validations,
                ]
            )
        workbook.close()
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()
