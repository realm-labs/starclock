#!/usr/bin/env python3

import argparse
import pathlib
import shutil

from openpyxl import load_workbook

WORKBOOKS = ["PureFiction.xlsx", "PureFictionBindings.xlsx", "PureFictionReview.xlsx"]

parser = argparse.ArgumentParser()
parser.add_argument("--source", type=pathlib.Path, required=True)
parser.add_argument("--output", type=pathlib.Path, required=True)
args = parser.parse_args()
args.output.mkdir(parents=True, exist_ok=False)
for name in WORKBOOKS:
    target = args.output / name
    shutil.copy2(args.source / name, target)
    workbook = load_workbook(target)
    for sheet in workbook.worksheets:
        sheet.print_area = f"A1:P{min(sheet.max_row, 10)}"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.sheet_view.zoomScale = 70
        for row in range(8, min(sheet.max_row, 10) + 1):
            sheet.row_dimensions[row].height = 42
    workbook.save(target)
print("Prepared three disposable workbooks for 37-sheet visual review.")
