#!/usr/bin/env python3

import argparse
import datetime
import json
import pathlib
import re
import shutil
import tempfile
import zipfile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path.cwd()
PACK = ROOT / "content-reference/pure-fiction-v1"
DEFAULT_DATA = ROOT / "config/pure-fiction/data"
TEMPLATES = ROOT / "config/pure-fiction-generated/templates"
GROUPS = {
    "PureFiction.xlsx": [
        ("profiles", "Profile"), ("seasons", "Season"), ("stages", "Stage"),
        ("nodes", "Node"), ("tierce-starward", "TierceStarward"),
        ("participant-policies", "ParticipantPolicy"),
        ("attempt-policies", "AttemptPolicy"),
    ],
    "PureFictionBindings.xlsx": [
        ("clocks", "Clock"), ("spawn-programs", "SpawnProgram"),
        ("score-programs", "ScoreProgram"), ("objectives", "Objective"),
        ("seasonal-mechanics", "SeasonalMechanic"),
        ("cacophonies", "Cacophony"), ("initial-resources", "InitialResource"),
        ("pool-proofs", "PoolProof"), ("themes", "Theme"),
        ("maze-buffs", "MazeBuff"), ("battle-events", "BattleEvent"),
        ("ability-programs", "AbilityProgram"), ("encounters", "Encounter"),
        ("waves", "Wave"), ("enemy-slots", "EnemySlot"),
        ("enemy-variants", "EnemyVariant"),
        ("enemy-templates", "EnemyTemplate"), ("enemy-skills", "EnemySkill"),
        ("enemy-character-configs", "EnemyCharacterConfig"),
        ("enemy-ai", "EnemyAI"), ("enemy-abilities", "EnemyAbility"),
        ("enemy-statuses", "EnemyStatus"), ("mechanic-rules", "MechanicRule"),
    ],
    "PureFictionReview.xlsx": [
        ("sources", "SourceRecord"), ("coverage", "ContentAudit"),
        ("coverage", "Coverage"), ("research-gaps", "ResearchGap"),
        ("reconciliation", "Reconciliation"),
        ("semantic-fixtures", "SemanticFixture"), ("pack-index", "PackFile"),
    ],
}
HEADERS = [
    "id", "stable_key", "row_order", "name_en", "name_zh_cn", "summary_en",
    "summary_zh_cn", "ownership", "coverage_state", "evidence_quality",
    "mechanism_quality", "manifest_record_ids", "source_record_ids",
    "payload_json", "runtime_executable",
]


def normalize_xlsx(path: pathlib.Path) -> None:
    with tempfile.TemporaryDirectory(prefix="starclock-g15-xlsx-") as temporary_dir:
        temporary = pathlib.Path(temporary_dir) / path.name
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
            temporary, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as target:
            for name in sorted(source.namelist()):
                info = zipfile.ZipInfo(name, date_time=(2000, 1, 1, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                info.external_attr = source.getinfo(name).external_attr
                payload = source.read(name)
                if name == "docProps/core.xml":
                    payload = re.sub(
                        rb"<dcterms:created[^>]*>[^<]*</dcterms:created>",
                        (b'<dcterms:created xsi:type="dcterms:W3CDTF">'
                         b'2000-01-01T00:00:00Z</dcterms:created>'),
                        payload,
                    )
                    payload = re.sub(
                        rb"<dcterms:modified[^>]*>[^<]*</dcterms:modified>",
                        (b'<dcterms:modified xsi:type="dcterms:W3CDTF">'
                         b'2000-01-01T00:00:00Z</dcterms:modified>'),
                        payload,
                    )
                target.writestr(info, payload)
        shutil.copyfile(temporary, path)


def envelope(record: dict, sheet_name: str) -> list:
    source_row = "name_en" not in record
    manifest_ids = record.get("manifest_record_ids")
    if manifest_ids is None:
        manifest_ids = [record["id"].removeprefix("source.")]
    source_ids = record.get("source_record_ids")
    if source_ids is None:
        source_ids = [record["id"]]
    stable_key = record["id"]
    if sheet_name == "ContentAudit":
        stable_key = f"audit.{stable_key}"
    return [
        stable_key,
        record.get("name_en", record["id"]),
        record.get("name_zh_cn", record["id"]),
        record.get("summary_en", "Manifest source evidence locator."),
        record.get("summary_zh_cn", "清单来源证据定位器。"),
        record.get("ownership", "EvidenceOnly" if source_row else "Shared"),
        record.get("coverage_state", "DataReady"),
        record.get("evidence_quality", "ProjectPolicy"),
        record.get("mechanism_quality", "Exact"),
        "|".join(manifest_ids),
        "|".join(source_ids),
        json.dumps(record, ensure_ascii=False, separators=(",", ":"), sort_keys=True),
        bool(record.get("runtime_executable", False)),
    ]


def create_workbook(filename: str, tables: list[tuple[str, str]], data_root: pathlib.Path) -> None:
    workbook = load_workbook(TEMPLATES / filename)
    workbook.properties.created = datetime.datetime(2000, 1, 1)
    workbook.properties.modified = datetime.datetime(2000, 1, 1)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for table_order, (file, sheet_name) in enumerate(tables, start=1):
        document = json.loads((PACK / f"{file}.json").read_text())
        sheet = workbook[sheet_name]
        sheet.freeze_panes = "B8"
        sheet.auto_filter.ref = f"B3:{get_column_letter(len(HEADERS) + 1)}3"
        for cell in sheet[3][1:]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row_order, record in enumerate(document["records"], start=1):
            values = envelope(record, sheet_name)
            sheet.append([None, table_order * 1_000_000 + row_order,
                          values[0], row_order, *values[1:]])
        widths = [12, 42, 12, 28, 24, 48, 48, 20, 16, 24, 24, 48, 48, 80, 18]
        for index, width in enumerate(widths, start=2):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.column_dimensions["A"].width = 3
        for row in sheet.iter_rows(min_row=8):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    data_root.mkdir(parents=True, exist_ok=True)
    output = data_root / filename
    workbook.save(output)
    normalize_xlsx(output)
    loaded = load_workbook(output, read_only=True, data_only=False)
    assert loaded.sheetnames == [sheet for _, sheet in tables]
    loaded.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=pathlib.Path, default=DEFAULT_DATA)
    args = parser.parse_args()
    for workbook_name, tables in GROUPS.items():
        create_workbook(workbook_name, tables, args.output.resolve())
    print("Pure Fiction workbooks: 3 files, 37 sheets.")


if __name__ == "__main__":
    main()
