#!/usr/bin/env python3

import json
import pathlib
import sys

from openpyxl import load_workbook

ROOT = pathlib.Path.cwd()
sys.path.insert(0, str(ROOT / "tools/pure-fiction-reference"))
from author_workbooks import GROUPS, HEADERS  # noqa: E402

data_root = ROOT / "config/pure-fiction/data"
sheet_count = 0
row_count = 0
for workbook_name, tables in GROUPS.items():
    workbook = load_workbook(data_root / workbook_name, read_only=False,
                             data_only=False)
    expected_sheets = [sheet for _, sheet in tables]
    assert workbook.sheetnames == expected_sheets
    for (file, sheet_name) in tables:
        sheet = workbook[sheet_name]
        document = json.loads((ROOT / "content-reference/pure-fiction-v1" /
                               f"{file}.json").read_text())
        assert sheet["A1"].value == "@table"
        assert sheet["C1"].value == "@mode"
        assert sheet["D1"].value == "map"
        assert [cell.value for cell in sheet[3][1:16]] == HEADERS
        authored = max(sheet.max_row - 7, 0)
        assert authored == len(document["records"]), (file, authored,
                                                       len(document["records"]))
        stable_keys = [sheet.cell(row=row, column=3).value
                       for row in range(8, sheet.max_row + 1)]
        expected_keys = [record["id"] for record in document["records"]]
        if sheet_name == "ContentAudit":
            expected_keys = [f"audit.{key}" for key in expected_keys]
        assert stable_keys == expected_keys
        assert len(stable_keys) == len(set(stable_keys))
        for row in sheet.iter_rows(min_row=8, min_col=2, max_col=16):
            assert not any(isinstance(cell.value, str) and
                           cell.value.startswith("=") for cell in row)
            assert row[-1].value is False
        assert sheet.freeze_panes == "B8"
        assert sheet.auto_filter.ref == "B3:P3"
        sheet_count += 1
        row_count += authored
    workbook.close()
assert sheet_count == 37
assert row_count == 6810
print(f"Pure Fiction workbooks verified: {sheet_count} sheets, "
      f"{row_count} rows, zero formulas/runtime rows.")
