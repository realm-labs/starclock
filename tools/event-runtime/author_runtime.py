#!/usr/bin/env python3
"""Author production Galactic Baseballer and Fate NIGHT runtime workbooks."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def source_ref(row: dict) -> str:
    refs = row.get("source_refs") or row.get("evidence_refs") or []
    return json.dumps(refs[0] if refs else {"locator": row["id"]},
                      ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def gb_rows(root: Path) -> dict[str, list[dict]]:
    base = root / "content-reference/galactic-baseballer-v1"
    load = lambda name: json.loads((base / f"{name}.json").read_text(encoding="utf-8"))
    profiles = load("profiles")
    stages = load("stages")
    stage_periods = load("stage-periods")
    weapons = load("weapons")
    accessories = load("accessories")
    shop_upgrades = load("shop-upgrades")
    strategies = load("adventure-strategies")
    progression = load("progression")
    recipes = load("synthesis-recipes")
    inputs = load("synthesis-inputs")
    scores = load("scoring-rules")
    approximations = load("approximations")
    profile_ids = {
        "galactic-baseballer.departure.v2_2": 1,
        "galactic-baseballer.demon-king.v3_3": 2,
    }
    authored_profiles = []
    for row in profiles:
        kind = "Departure" if "departure" in row["id"] else "DemonKing"
        authored_profiles.append({
            "id": profile_ids[row["id"]],
            "stable_key": row["id"],
            "kind": kind,
            "weapon_slots": 5,
            "unlocked_weapon_slots": 4,
            "accessory_slots": 6,
            "unlocked_accessory_slots": 4,
            "source_ref": source_ref(row),
        })
    trigger_exact = {
        row["parent_id"]: bool(row.get("runtime_executable", False))
        for row in load("weapon-triggers")
    }
    accessory_exact = {
        row["parent_id"]: bool(row.get("runtime_executable", False))
        for row in load("accessory-bindings")
    }
    equipment_rows = []
    equipment_ids = {}
    for row in [*weapons, *accessories]:
        numeric = int(row["source_numeric_id"])
        equipment_ids[row["id"]] = numeric
        if row in accessories:
            kind = "Accessory"
            exact = accessory_exact.get(row["id"], False)
        else:
            kind = "LegendaryWeapon" if row["tier"] != "Standard" else "StandardWeapon"
            exact = trigger_exact.get(row["id"], False)
        equipment_rows.append({
            "id": numeric,
            "stable_key": row["id"],
            "kind": kind,
            "maximum_level": row["maximum_level"],
            "profile_ids": ",".join(str(profile_ids[item]) for item in row["profile_ids"]),
            "runtime_binding_exact": exact,
            "source_ref": source_ref(row),
        })
    authored_shop_upgrades = []
    for index, row in enumerate(shop_upgrades, start=1):
        kind = row["shop_type"]
        authored_shop_upgrades.append({
            "id": index,
            "stable_key": row["id"],
            "profile_id": profile_ids[row["profile_ids"][0]],
            "source_numeric_id": int(row["source_numeric_id"]),
            "purchase_level": row["purchase_level"],
            "maximum_level": row["maximum_level"],
            "kind": kind,
            "currency_key": row["cost_currency_id"],
            "cost": row["cost"],
            "maze_buff_id": (
                int(row["maze_buff_id"]) if row.get("maze_buff_id") else None
            ),
            "maze_buff_parameters": (
                ",".join(row["maze_buff_parameters"])
                if row.get("maze_buff_parameters")
                else None
            ),
            "shop_parameter_values": ",".join(row["shop_parameter_values"]),
            "runtime_binding_exact": kind != "AddMazeBuff",
            "source_ref": source_ref(row),
        })
    recipe_ids = {row["id"]: index for index, row in enumerate(recipes, start=1)}
    authored_recipes = [{
        "id": recipe_ids[row["id"]],
        "profile_id": profile_ids[row["profile_ids"][0]],
        "stable_key": row["id"],
        "output_equipment_id": equipment_ids[row["output_weapon_id"]],
        "tier": row["tier"],
        "source_ref": source_ref(row),
    } for row in recipes]
    authored_inputs = []
    for index, row in enumerate(inputs, start=1):
        authored_inputs.append({
            "id": index,
            "recipe_id": recipe_ids[row["recipe_id"]],
            "input_order": row["input_order"],
            "kind": "Equipment",
            "equipment_id": equipment_ids[row["input_id"]],
            "required_level": row["required_level"],
            "consumed": row["consumed"],
            "source_ref": source_ref(row),
        })
    authored_stages = []
    stage_ids = {}
    for row in stages:
        thresholds = [item["minimum_score"] for item in row["rating_thresholds"]]
        authored_stages.append({
            "id": int(row["source_numeric_id"]),
            "profile_id": profile_ids[row["profile_ids"][0]],
            "difficulty": row["difficulty"],
            "weapon_selectable": row["weapon_selectable"],
            "initial_weapon_ids": (
                ",".join(
                    str(equipment_ids.get(item, int(item)))
                    for item in row["initial_weapon_ids"]
                )
                or None
            ),
            "rating_thresholds": ",".join(str(value) for value in thresholds),
            "source_ref": source_ref(row),
        })
        stage_ids[row["id"]] = int(row["source_numeric_id"])
    authored_stage_periods = []
    rank_names = {
        "PeriodFirst": "First",
        "PeriodSecond": "Second",
        "PeriodThird": "Third",
        "PeriodExtra": "Extra",
    }
    for row in stage_periods:
        parent = row.get("parent_stage_id")
        if parent not in stage_ids:
            continue
        authored_stage_periods.append({
            "id": int(row["source_numeric_id"]),
            "stage_id": stage_ids[parent],
            "rank": rank_names[row["period_rank"]],
            "encounter_id": int(row["shared_stage_config_id"]),
            "battle_event_id": int(row["battle_event_id"]),
            "wave_count": row["wave_count"],
            "countdown_by_wave": ",".join(str(value) for value in row["countdown_by_wave"]),
            "period_score": row["period_score"],
            "stage_score": row.get("stage_score"),
            "selection_weight": row["selection_weight"],
            "source_ref": source_ref(row),
        })
    authored_scores = [{
        "profile_id": profile_ids[row["profile_ids"][0]],
        "monster_base_score": row["monster_base_score"],
        "elite_scores": ",".join(str(value) for value in row["elite_score_vector"]),
        "monster_weights": ",".join(str(value) for value in row["monster_weight_vector"]),
        "score_cap": row["score_upper_limit"],
        "final_stage_extra_bonus": row["final_stage_extra_bonus"],
        "source_ref": source_ref(row),
    } for row in scores]
    policy_tests = {
        "galactic-baseballer.approximation.gb.policy.upgrade-candidate-weight": [
            "runtime::tests::reward_offer_uses_bounded_uniform_stable_options",
        ],
        "galactic-baseballer.approximation.gb.policy.no-legal-candidate": [
            "inventory::tests::no_legal_candidate_exposes_only_skip",
        ],
        "galactic-baseballer.approximation.gb.policy.simultaneous-synthesis-order": [
            "inventory::tests::eligible_synthesis_is_atomic_and_precedes_duplicate_upgrade",
        ],
        "galactic-baseballer.demon-king.approximation.shop-transaction-atomicity": [
            "progression::tests::purchase_commits_balance_level_and_effect_atomically",
            "progression::tests::rejected_purchase_leaves_state_byte_identical",
        ],
    }
    selected_policy_ids = list(policy_tests)
    authored_policies = []
    for index, policy_id in enumerate(selected_policy_ids, start=1):
        row = next(item for item in approximations if item["id"] == policy_id)
        authored_policies.append({
            "id": index,
            "stable_key": row["id"],
            "unavailable_fact": row["unavailable_fact"],
            "known_facts": " | ".join(row["known_released_facts"]),
            "selected_behavior": row["selected_policy"],
            "rejected_alternatives": ",".join(row["rejected_alternatives"]),
            "rationale": row["rationale"],
            "affected_tests": ",".join(policy_tests[policy_id]),
            "confidence": row["confidence"],
            "replacement_condition": row["replacement_condition"],
            "source_ref": source_ref(row),
        })
    authored_strategies = [{
        "id": int(row["source_numeric_id"]),
        "stable_key": row["id"],
        "profile_id": profile_ids[row["profile_ids"][0]],
        "kind": row["source_type"],
        "maximum_level": row["maximum_level"],
        "unlock_quest_id": int(row["unlock_quest_id"]) if row.get("unlock_quest_id") else None,
        "selectable_periods": (
            ",".join(row["selectable_period_ids"])
            if row.get("selectable_period_ids")
            else None
        ),
        "influence_scope": row["influence_scope"],
        "maze_buff_id": int(row["maze_buff_id"]),
        "maze_buff_parameters": (
            ",".join(row["maze_buff_parameters"])
            if row["maze_buff_parameters"]
            else None
        ),
        "ability_binding": row["program_summary"]["binding_key"],
        "runtime_binding_exact": False,
        "source_ref": source_ref(row),
    } for row in strategies]
    team_bonuses = [row for row in progression if row["kind"] == "TeamBonus"]
    authored_team_bonuses = [{
        "stage_id": stage_ids[row["stage_id"]],
        "profile_id": profile_ids[row["profile_ids"][0]],
        "maze_buff_id": int(row["source_maze_buff_id"]),
        "level": row["source_level"],
        "parameters": ",".join(row["parameter_values"]),
        "ability_binding": row["binding_key"],
        "runtime_binding_exact": False,
        "source_ref": source_ref(row),
    } for row in team_bonuses]
    authored_policies.extend([
        {
            "id": len(authored_policies) + 1,
            "stable_key": "policy.baseballer.adventure-strategy-runtime-binding",
            "unavailable_fact": "The 56 released Adventure Strategy ability programs have not been lowered into shared Combat operations.",
            "known_facts": "Strategy identity, type, unlock quest, selectable period, MazeBuff parameters and structural ability binding are exact released data.",
            "selected_behavior": "Expose exact strategy configuration with runtime bindings disabled; do not offer a strategy as an executable stage mutation.",
            "rejected_alternatives": "attach no-op effects,infer operations from localized prose,run a mode-local battle processor",
            "rationale": "Identity-only effects must not silently alter authoritative battle state.",
            "affected_tests": "event::tests::production_baseballer_profiles_and_synthesis_lower",
            "confidence": "High",
            "replacement_condition": "Enable each strategy only after its released ability program lowers to shared Combat operations and passes replay/hash fixtures.",
            "source_ref": source_ref(strategies[0]),
        },
        {
            "id": len(authored_policies) + 2,
            "stable_key": "policy.baseballer.team-bonus-runtime-binding",
            "unavailable_fact": "The seven released Demon King stage team-bonus programs have not been lowered into shared Combat operations.",
            "known_facts": "Each stage-to-MazeBuff relation, level, cumulative parameter vector and structural ability binding is exact released data.",
            "selected_behavior": "Expose exact team-bonus configuration with runtime bindings disabled and refuse to claim the bonus is active in BattleSpec.",
            "rejected_alternatives": "attach no-op modifiers,infer formulas from display text,drop stage-to-bonus identity",
            "rationale": "Failing closed preserves exact stage identity without inventing combat semantics.",
            "affected_tests": "event::tests::production_baseballer_profiles_and_synthesis_lower",
            "confidence": "High",
            "replacement_condition": "Enable each team bonus only after its released program lowers to shared Combat and semantic fixtures pass.",
            "source_ref": source_ref(team_bonuses[0]),
        },
    ])
    return {
        "Profiles": authored_profiles,
        "Policies": authored_policies,
        "Equipment": equipment_rows,
        "ShopUpgrades": authored_shop_upgrades,
        "Stages": authored_stages,
        "StagePeriods": authored_stage_periods,
        "Recipes": authored_recipes,
        "RecipeInputs": authored_inputs,
        "ScoreRules": authored_scores,
        "Strategies": authored_strategies,
        "TeamBonuses": authored_team_bonuses,
    }


def table_rows(root: Path, table: str) -> list[dict]:
    path = root / "config/fate-star-rail-night-generated/debug-json" / f"{table}.json"
    return json.loads(path.read_text(encoding="utf-8"))["table"]["rows"]


def typed(values: dict, key: str):
    value = values[key]
    if value == "Null":
        return None
    return next(iter(value.values()))


def reference_records(root: Path, file_name: str, family: str) -> list[dict]:
    path = root / "content-reference/fate-star-rail-night-v1" / file_name
    records = json.loads(path.read_text(encoding="utf-8"))["records"]
    return [row for row in records if row["family"] == family and row["enabled"]]


def fate_rows(root: Path) -> dict[str, list[dict]]:
    profiles = table_rows(root, "FsnProfiles")
    owners = reference_records(root, "participants.json", "FateRinOwner")
    decks = reference_records(root, "noble-phantasms.json", "FateRinDeck")
    recommendations = reference_records(
        root, "noble-phantasms.json", "FateRinDeckRecommend"
    )
    cards = reference_records(root, "noble-phantasms.json", "FateRinHouguConfig")
    boards = reference_records(root, "profile-graph.json", "FateRinCaseBoard")
    board_nodes = reference_records(root, "profile-graph.json", "FateRinCaseBoardInfo")
    story_fights = reference_records(root, "profile-graph.json", "FateRinStoryFight")
    challenge_fights = reference_records(
        root, "profile-graph.json", "FateRinChallengeFight"
    )
    map_fights = reference_records(root, "profile-graph.json", "FateRinHouguMapFight")
    profile_values = profiles[0]["values"]
    authored_profiles = [{
        "id": 1,
        "stable_key": typed(profile_values, "stable_key"),
        "source_ref": typed(profile_values, "source_refs_json"),
    }]
    authored_owners = [{
        "id": index,
        "owner": row["mechanic_payload"]["PHFMCACHFIJ"],
        "source_ref": source_ref(row),
    } for index, row in enumerate(owners, start=1)]
    authored_cards = []
    for row in cards:
        payload = row["mechanic_payload"]
        authored_cards.append({
            "id": int(payload["PHFMCACHFIJ"]),
            "stable_key": row["stable_id"],
            "owner": payload["GMPGDEINODK"],
            "magical_energy_cost": int(payload["NHALJPDONCP"]),
            "rarity": "Ssr" if payload["PMIEAEGJNMJ"] == "SSR" else payload["PMIEAEGJNMJ"],
            "ability_program": payload["GINFOPOAKHK"] or None,
            "runtime_binding_exact": False,
            "source_ref": source_ref(row),
        })
    authored_cards.sort(key=lambda row: row["id"])
    authored_decks = []
    for row in decks:
        payload = row["mechanic_payload"]
        authored_decks.append({
            "id": int(payload["PHFMCACHFIJ"]),
            "stable_key": row["stable_id"],
            "owner": payload["LOALOLNACOA"],
            "presentation_locator": int(payload["LIPCDDAPHNF"]),
            "action_locator": int(payload["ENKMNJDEMJE"]),
            "source_ref": source_ref(row),
        })
    authored_recommendations = []
    for index, row in enumerate(recommendations, start=1):
        payload = row["mechanic_payload"]
        authored_recommendations.append({
            "id": index,
            "owner": payload["LOALOLNACOA"],
            "kind": "Final" if payload.get("JGAKLKBOPEG") == "Final" else "Base",
            "owner_card_ids": ",".join(str(value) for value in payload["OFIGPIFELHJ"]),
            "neutral_card_ids": ",".join(str(value) for value in payload["NJBEMAEAEIL"]),
            "source_ref": source_ref(row),
        })
    authored_boards = [{
        "id": int(row["mechanic_payload"]["PHFMCACHFIJ"]),
        "stable_key": row["stable_id"],
        "source_ref": source_ref(row),
    } for row in boards]
    authored_board_nodes = []
    for index, row in enumerate(board_nodes):
        sequence = index % 3 + 1
        authored_board_nodes.append({
            "id": index + 1,
            "board_id": index // 3 + 1,
            "sequence": sequence,
            "kind": ("Choice" if sequence == 1 else "Battle" if sequence == 2 else "Completed"),
            "source_ref": source_ref(row),
        })
    authored_story_fights = [{
        "id": int(row["mechanic_payload"]["PHFMCACHFIJ"]),
        "battle_event_id": int(row["mechanic_payload"]["DOBKKDIECDO"]),
        "map_entrance_id": int(row["mechanic_payload"]["HNEIIAGADGO"]),
        "source_ref": source_ref(row),
    } for row in story_fights]
    authored_challenge_fights = [{
        "id": int(row["mechanic_payload"]["PHFMCACHFIJ"]),
        "battle_event_id": int(row["mechanic_payload"]["DOBKKDIECDO"]),
        "map_entrance_id": int(row["mechanic_payload"]["HNEIIAGADGO"]),
        "enemy_id": int(row["mechanic_payload"]["JFDHFPIIGCC"]),
        "buff_ids": ",".join(str(value) for value in row["mechanic_payload"]["FOHHOOKJPIM"]),
        "source_ref": source_ref(row),
    } for row in challenge_fights]
    authored_map_fights = []
    for row in map_fights:
        payload = row["mechanic_payload"]
        battle_events = [
            int(payload[key])
            for key in ("NHAINGEIMJA", "KAHNDIPJGHI", "HPJHKACDIMB")
            if key in payload
        ]
        authored_map_fights.append({
            "id": int(payload["PHFMCACHFIJ"]),
            "battle_event_ids": ",".join(str(value) for value in battle_events),
            "map_entrance_id": int(payload["HNEIIAGADGO"]),
            "reward_card_id": int(payload["OHFGNODANEP"]) if "OHFGNODANEP" in payload else None,
            "terminal": bool(payload.get("PKLFLANJCDG", False)),
            "enemy_id": int(payload["JFDHFPIIGCC"]),
            "relation": payload.get("EHAFJKIKKMC"),
            "source_ref": source_ref(row),
        })
    gaps = json.loads(
        (root / "content-reference/fate-star-rail-night-v1/research-gaps.json")
        .read_text(encoding="utf-8")
    )
    gap_records = [row for row in gaps.get("policies", []) if row]
    policies = [{
        "id": index,
        "stable_key": row["policy_id"],
        "unavailable_fact": row["unavailable_fact"],
        "known_facts": f"Released obligation {row['obligation_id']} remains identity-exact while operation semantics are unproven.",
        "selected_behavior": row["selected_policy"],
        "rejected_alternatives": ",".join(row["rejected_alternatives"]),
        "rationale": row["rationale"],
        "affected_tests": "event::tests::production_fate_card_surface_lowers_with_policies",
        "confidence": "Low",
        "replacement_condition": row.get("replacement_condition", "Replace when released exact runtime evidence is available."),
        "source_ref": json.dumps({
            "repository_or_url": "starclock",
            "revision_or_access_date": "current-tree",
            "game_version": "4.4",
            "path_or_page": "content-reference/fate-star-rail-night-v1/research-gaps.json",
            "locator": f"policy_id={row['policy_id']}",
            "evidence_quality": "ProjectPolicy",
        }, sort_keys=True, separators=(",", ":")),
    } for index, row in enumerate(gap_records, start=1)]
    policies.extend([
        {
            "id": len(policies) + 1,
            "stable_key": "policy.fate.case-board-node-grouping",
            "unavailable_fact": "Released Case Board rows do not expose exact edge selectors or node kinds.",
            "known_facts": "There are 18 released ordered CaseBoardInfo identities and six released board identities.",
            "selected_behavior": "Group the 18 released CaseBoardInfo rows into six stable three-node boards by authored row order: choice, battle, terminal.",
            "rejected_alternatives": "invent branching edges,discard the released board rows",
            "rationale": "A bounded linear board keeps every released identity reachable without claiming hidden adjacency.",
            "affected_tests": "event::tests::production_fate_card_surface_lowers_with_policies",
            "confidence": "Low",
            "replacement_condition": "Replace when a released edge selector exposes the exact Case Board adjacency and node-kind mapping.",
            "source_ref": authored_board_nodes[0]["source_ref"],
        },
        {
            "id": len(policies) + 2,
            "stable_key": "policy.fate.tactical-card-runtime-binding",
            "unavailable_fact": "The 107 released card ability programs and exact draw, discard, refill and end-turn ordering have not been lowered into the shared Combat aggregate.",
            "known_facts": "Released card identity, owner, magical-energy cost, rarity and ability-program path are exact; public released observations establish random hands, repeated energy-paid card plays and an explicit end-turn command while ordinary Ultimates remain interruptible.",
            "selected_behavior": "Expose exact card and deck configuration but keep every production runtime binding false and refuse to claim a playable tactical-card battle until shared Combat owns its commands, RNG, state hash and replay.",
            "rejected_alternatives": "run a second mode-local battle command processor,attach no-op card effects,infer operations from localized prose",
            "rationale": "Failing closed preserves one authoritative Combat aggregate and prevents identity-only cards from mutating battle state as invented effects.",
            "affected_tests": "event::tests::production_fate_card_surface_lowers_with_policies",
            "confidence": "High",
            "replacement_condition": "Replace card bindings individually after released ability programs and observed draw/turn fixtures pass shared Combat lowering, replay and hash tests.",
            "source_ref": authored_cards[0]["source_ref"],
        },
        {
            "id": len(policies) + 3,
            "stable_key": "policy.fate.custom-fight-battle-spec-mapping",
            "unavailable_fact": "The custom FateRin battle-event locators are not ordinary StageConfig encounter IDs and their exact BattleSpec assembly mapping is unavailable.",
            "known_facts": "Six story fights, four challenge fights and fifteen map fights publish exact battle-event, map-entrance, enemy and optional reward-card locators.",
            "selected_behavior": "Retain the custom fight locators as typed mode data and do not reinterpret the legacy 425001..425008 FateActivity stages as 4.4 Fate/Star Rail NIGHT encounters.",
            "rejected_alternatives": "bind legacy FateActivity StageConfig rows by ID order,treat battle-event IDs as EncounterId",
            "rationale": "The custom fight tables prove identity but not the shared BattleSpec assembly needed for an authoritative handoff.",
            "affected_tests": "event::tests::production_fate_card_surface_lowers_with_policies",
            "confidence": "Low",
            "replacement_condition": "Replace when released custom battle-event programs lower each fight locator to a validated BattleSpec and semantic fixture.",
            "source_ref": authored_story_fights[0]["source_ref"],
        },
    ])
    return {
        "Profiles": authored_profiles,
        "Owners": authored_owners,
        "Cards": authored_cards,
        "Decks": authored_decks,
        "DeckRecommendations": authored_recommendations,
        "Boards": authored_boards,
        "BoardNodes": authored_board_nodes,
        "StoryFights": authored_story_fights,
        "ChallengeFights": authored_challenge_fights,
        "MapFights": authored_map_fights,
        "Policies": policies,
    }


def columns(sheet) -> list[str]:
    return [str(cell.value) for cell in sheet[3] if cell.value not in (None, "#field")]


def style_sheet(sheet) -> None:
    for column in range(1, sheet.max_column + 1):
        values = (
            str(sheet.cell(row, column).value or "")
            for row in range(1, sheet.max_row + 1)
        )
        width = min(max(max(map(len, values), default=0) + 2, 10), 40)
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows(min_row=8):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:7"


def write(
    template: Path,
    output: Path,
    authored: dict[str, list[dict]],
    replace: bool,
) -> None:
    if output.exists() and not replace:
        raise FileExistsError(f"refusing to overwrite {output}")
    workbook = openpyxl.load_workbook(template)
    if workbook.sheetnames != list(authored):
        raise ValueError(f"template sheets differ: {workbook.sheetnames}")
    for name, rows in authored.items():
        sheet = workbook[name]
        if sheet.max_row >= 8:
            sheet.delete_rows(8, sheet.max_row - 7)
        headers = columns(sheet)
        for row in rows:
            sheet.append([None, *[row[column] for column in headers]])
        sheet.freeze_panes = "A8"
        sheet.sheet_view.showGridLines = False
        sheet.auto_filter.ref = f"A3:{sheet.cell(3, sheet.max_column).column_letter}{sheet.max_row}"
        style_sheet(sheet)
    output.parent.mkdir(parents=True, exist_ok=True)
    if replace:
        temporary = output.with_suffix(".authoring.tmp.xlsx")
        if temporary.exists():
            raise FileExistsError(f"refusing to overwrite stale temporary file {temporary}")
        workbook.save(temporary)
        temporary.replace(output)
    else:
        workbook.save(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--template-root", type=Path, required=True)
    parser.add_argument("--output-root", type=Path, required=True)
    parser.add_argument(
        "--replace",
        action="store_true",
        help="atomically replace known generated production workbooks",
    )
    args = parser.parse_args()
    root = args.root.resolve()
    write(
        args.template_root / "GalacticBaseballerRuntime.xlsx",
        args.output_root / "GalacticBaseballerRuntime.xlsx",
        gb_rows(root),
        args.replace,
    )
    write(
        args.template_root / "FateNightRuntime.xlsx",
        args.output_root / "FateNightRuntime.xlsx",
        fate_rows(root),
        args.replace,
    )
    print("Authored Galactic Baseballer and Fate NIGHT production workbooks")


if __name__ == "__main__":
    main()
