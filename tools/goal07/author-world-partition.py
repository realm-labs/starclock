"""Verify one Goal 07 world-structure partition in authoritative Excel and Sora."""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA

ROOT = Path(__file__).resolve().parents[2]
GOAL = "standard-universe-mechanics-complete-v1"
MANIFEST = ROOT / "content-manifests" / GOAL / "content-partitions.json"
REFERENCE = ROOT / "content-reference" / "standard-universe-v1"
DATA = ROOT / "config" / "data"
DEBUG = ROOT / "config" / "universe-generated" / "debug-json"

DOMAIN_KINDS = {
    "combat-primary": "CombatPrimary",
    "combat-secondary": "CombatSecondary",
    "occurrence": "Occurrence",
    "encounter": "Encounter",
    "respite": "Respite",
    "elite": "Elite",
    "boss": "Boss",
    "transaction": "Transaction",
    "adventure": "Adventure",
}


def rows(sheet: Any) -> list[dict[str, Any]]:
    fields = {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }
    output: list[dict[str, Any]] = []
    for cells in sheet.iter_rows(min_row=8):
        row = {field: cells[column - 1].value for field, column in fields.items()}
        if all(value is None for value in row.values()):
            continue
        if any(
            cells[column - 1].data_type in (TYPE_ERROR, TYPE_FORMULA)
            for column in fields.values()
        ):
            raise ValueError(f"{sheet.title}: formulas and errors are forbidden")
        output.append(row)
    return output


def keyed(values: list[dict[str, Any]], field: str) -> dict[Any, dict[str, Any]]:
    output: dict[Any, dict[str, Any]] = {}
    for row in values:
        if row[field] in output:
            raise ValueError(f"duplicate {field}={row[field]}")
        output[row[field]] = row
    return output


def unwrap(value: Any) -> Any:
    if value == "Null":
        return None
    if isinstance(value, list):
        return [unwrap(item) for item in value]
    if isinstance(value, dict):
        if len(value) == 1:
            tag, inner = next(iter(value.items()))
            if tag in {"String", "Integer", "Bool", "Decimal"}:
                return inner
            if tag == "List":
                return [unwrap(item) for item in inner]
        return {key: unwrap(inner) for key, inner in value.items()}
    return value


def sora_rows(table: str) -> list[dict[str, Any]]:
    payload = json.loads((DEBUG / f"{table}.json").read_text(encoding="utf-8"))
    return [
        {key: unwrap(value) for key, value in row["values"].items()}
        for row in payload["table"]["rows"]
    ]


def normalize(table: str, row: dict[str, Any]) -> dict[str, Any]:
    output = dict(row)
    if table == "UniverseDifficulty" and isinstance(
        output.get("recommended_elements"), str
    ):
        output["recommended_elements"] = [
            part for part in output["recommended_elements"].split("|") if part
        ]
    if table == "UniverseRoom" and isinstance(output.get("section_ids"), str):
        output["section_ids"] = [
            int(part) for part in output["section_ids"].split("|") if part
        ]
    if table == "UniverseReviewFixture":
        for field in ("input_stable_keys", "provenance_ids"):
            if isinstance(output.get(field), str):
                parts = [part for part in output[field].split("|") if part]
                output[field] = (
                    [int(part) for part in parts]
                    if field == "provenance_ids"
                    else parts
                )
    return output


def digest(table: str, values: list[dict[str, Any]]) -> str:
    values = [normalize(table, row) for row in values]
    ordered = sorted(
        values,
        key=lambda row: json.dumps(
            row, ensure_ascii=False, sort_keys=True, default=str
        ),
    )
    encoded = json.dumps(
        ordered,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )
    return hashlib.sha256(encoded.encode()).hexdigest()


def check_source_domain(reference: dict[str, Any], authored: dict[str, Any]) -> None:
    expected = {
        "stable_key": reference["id"],
        "source_type": int(reference["source_ids"][0]),
        "kind": DOMAIN_KINDS[reference["kind"]],
        "decision_policy": reference["decision_policy"],
        "terminal": reference["terminal"],
        "name_en": reference["name_en"],
        "name_zh_cn": reference["name_zh_cn"],
        "summary_en": reference["summary_en"],
        "summary_zh_cn": reference["summary_zh_cn"],
    }
    actual = {field: authored[field] for field in expected}
    if actual != expected:
        raise ValueError(f"{reference['id']}: Excel domain differs from public source")


def check_audits(
    partition: dict[str, Any], evidence: Any
) -> None:
    audits = keyed(rows(evidence["UniverseContentAudit"]), "content_stable_key")
    sources = keyed(rows(evidence["UniverseSourceRecord"]), "id")
    for key in partition["record_ids"]:
        audit = audits.get(key)
        provenance = [] if audit is None else str(audit["provenance_ids"]).split("|")
        if (
            audit is None
            or not audit["enabled"]
            or audit["mode_owner"] != "Standard"
            or audit["quality"] != "ExactStructured"
            or audit["mechanism_quality"] != "ExactStructured"
            or audit["coverage_state"] != "DataReady"
            or not provenance
            or any(int(item) not in sources for item in provenance)
        ):
            raise ValueError(f"{key}: exact enabled provenance audit is incomplete")


def build_domain_partition(
    partition: dict[str, Any],
    universe: Any,
    bindings: Any,
    evidence: Any,
) -> dict[str, list[dict[str, Any]]]:
    source = keyed(
        json.loads((REFERENCE / "domains.json").read_text(encoding="utf-8")),
        "id",
    )
    domains = keyed(rows(universe["UniverseDomain"]), "stable_key")
    selected_domains = [domains[key] for key in partition["record_ids"]]
    for row in selected_domains:
        check_source_domain(source[row["stable_key"]], row)

    domain_ids = {int(row["id"]) for row in selected_domains}
    selected_bindings = [
        row
        for row in rows(bindings["UniverseActivityDomainBinding"])
        if int(row["domain_id"]) in domain_ids
    ]
    if len(selected_bindings) != len(selected_domains):
        raise ValueError(
            f"{partition['id']}: Activity bindings do not exactly cover the domains"
        )
    by_domain = keyed(selected_bindings, "domain_id")
    for row in selected_domains:
        source_type = int(row["source_type"])
        binding = by_domain[int(row["id"])]
        expected_decision = (
            "BattleCommand"
            if row["decision_policy"] == "BattleHandoff"
            else "ExternalOutcome"
            if row["kind"] == "Adventure"
            else "RunCommand"
        )
        if int(binding["sequence"]) != source_type:
            raise ValueError(f"{row['stable_key']}: Activity sequence differs")
        if binding["decision_kind"] != expected_decision:
            raise ValueError(f"{row['stable_key']}: Activity decision differs")

    check_audits(partition, evidence)

    return {
        "UniverseDomain": selected_domains,
        "UniverseActivityDomainBinding": selected_bindings,
    }


def build_encounter_partition(
    partition: dict[str, Any],
    universe: Any,
    bindings: Any,
    evidence: Any,
) -> dict[str, list[dict[str, Any]]]:
    source = keyed(
        json.loads((REFERENCE / "encounter-pools.json").read_text(encoding="utf-8")),
        "id",
    )
    pools = keyed(rows(bindings["UniverseEncounterPool"]), "stable_key")
    rooms = keyed(rows(universe["UniverseRoom"]), "stable_key")
    groups = keyed(rows(bindings["UniverseEncounterGroup"]), "stable_key")
    selected_pools = [pools[key] for key in partition["record_ids"]]
    pool_ids = {int(row["id"]) for row in selected_pools}
    selected_groups = [
        row
        for row in rows(bindings["UniverseEncounterPoolGroup"])
        if int(row["pool_id"]) in pool_ids
    ]
    selected_fixed = [
        row
        for row in rows(bindings["UniverseEncounterPoolFixed"])
        if int(row["pool_id"]) in pool_ids
    ]
    groups_by_pool: dict[int, list[dict[str, Any]]] = {}
    fixed_by_pool: dict[int, list[dict[str, Any]]] = {}
    for row in selected_groups:
        groups_by_pool.setdefault(int(row["pool_id"]), []).append(row)
    for row in selected_fixed:
        fixed_by_pool.setdefault(int(row["pool_id"]), []).append(row)

    for authored in selected_pools:
        reference = source[authored["stable_key"]]
        expected = {
            "stable_key": reference["id"],
            "room_id": int(rooms[reference["room_id"]]["id"]),
            "domain_kind": DOMAIN_KINDS[reference["domain_kind"]],
            "map_entrance": reference["map_entrance"],
            "selection_policy": reference["selection_policy"],
            "source_primary_condition_key": reference[
                "source_primary_condition_key"
            ],
            "name_en": reference["name_en"],
            "name_zh_cn": reference["name_zh_cn"],
            "summary_en": reference["summary_en"],
            "summary_zh_cn": reference["summary_zh_cn"],
        }
        actual = {field: authored[field] for field in expected}
        if actual != expected:
            raise ValueError(
                f"{reference['id']}: Excel encounter pool differs from public source"
            )

        pool_id = int(authored["id"])
        authored_groups = sorted(
            groups_by_pool.get(pool_id, []), key=lambda row: int(row["sequence"])
        )
        expected_groups = [
            {
                "pool_id": pool_id,
                "sequence": sequence,
                "condition_key": item["condition_key"],
                "group_id": int(groups[item["group_id"]]["id"]),
                "weight_decimal": item["weight"],
            }
            for sequence, item in enumerate(reference["weighted_group_ids"], start=1)
        ]
        if authored_groups != expected_groups:
            raise ValueError(
                f"{reference['id']}: weighted encounter bindings differ from source"
            )
        authored_fixed = sorted(
            fixed_by_pool.get(pool_id, []), key=lambda row: int(row["sequence"])
        )
        expected_fixed = [
            {
                "pool_id": pool_id,
                "sequence": sequence,
                "condition_key": item["condition_key"],
                "source_content_id": item["source_content_id"],
            }
            for sequence, item in enumerate(
                reference["fixed_content_entries"], start=1
            )
        ]
        if authored_fixed != expected_fixed:
            raise ValueError(
                f"{reference['id']}: fixed encounter bindings differ from source"
            )

    fixtures = keyed(rows(evidence["UniverseReviewFixture"]), "stable_key")
    selected_fixtures = [fixtures[key] for key in partition["fixture_ids"]]
    check_audits(partition, evidence)
    return {
        "UniverseEncounterPool": selected_pools,
        "UniverseEncounterPoolGroup": selected_groups,
        "UniverseEncounterPoolFixed": selected_fixed,
        "UniverseReviewFixture": selected_fixtures,
    }


def build_map_partition(
    partition: dict[str, Any],
    universe: Any,
    evidence: Any,
) -> dict[str, list[dict[str, Any]]]:
    source = keyed(
        json.loads((REFERENCE / "maps.json").read_text(encoding="utf-8")),
        "id",
    )
    nodes = keyed(rows(universe["UniverseMapNode"]), "stable_key")
    selected_nodes = [nodes[key] for key in partition["record_ids"]]
    node_ids = {int(row["id"]) for row in selected_nodes}
    selected_edges = [
        row
        for row in rows(universe["UniverseMapEdge"])
        if int(row["source_node_id"]) in node_ids
    ]
    edges_by_source: dict[int, list[dict[str, Any]]] = {}
    for row in selected_edges:
        edges_by_source.setdefault(int(row["source_node_id"]), []).append(row)

    for authored in selected_nodes:
        reference = source[authored["stable_key"]]
        expected = {
            "stable_key": reference["id"],
            "source_map_id": int(reference["map_id"].rsplit(".", 1)[1]),
            "source_node_id": int(reference["node_id"]),
            "is_start": reference["start"],
            "position_x": int(reference["position_hint"]["x"]),
            "position_y": int(reference["position_hint"]["y"]),
        }
        actual = {field: authored[field] for field in expected}
        if actual != expected:
            raise ValueError(
                f"{reference['id']}: Excel map node differs from public source"
            )
        node_id = int(authored["id"])
        authored_edges = sorted(
            edges_by_source.get(node_id, []), key=lambda row: int(row["sequence"])
        )
        expected_edges = [
            {
                "source_node_id": node_id,
                "sequence": sequence,
                "target_node_id": int(nodes[target]["id"]),
            }
            for sequence, target in enumerate(reference["next_node_ids"], start=1)
        ]
        if authored_edges != expected_edges:
            raise ValueError(
                f"{reference['id']}: map edges differ from public source"
            )
    if any(int(row["target_node_id"]) not in node_ids for row in selected_edges):
        raise ValueError(f"{partition['id']}: map group has an external target")
    check_audits(partition, evidence)
    return {
        "UniverseMapNode": selected_nodes,
        "UniverseMapEdge": selected_edges,
    }


def build_room_partition(
    partition: dict[str, Any],
    universe: Any,
    bindings: Any,
    evidence: Any,
) -> dict[str, list[dict[str, Any]]]:
    source_rooms = keyed(
        json.loads((REFERENCE / "rooms.json").read_text(encoding="utf-8")),
        "id",
    )
    source_domains = keyed(
        json.loads((REFERENCE / "domains.json").read_text(encoding="utf-8")),
        "id",
    )
    source_groups = json.loads(
        (REFERENCE / "encounter-groups.json").read_text(encoding="utf-8")
    )
    group_key_by_source = {
        str(row["source_ids"][0]): row["id"] for row in source_groups
    }

    domains = keyed(rows(universe["UniverseDomain"]), "stable_key")
    rooms_by_key = keyed(rows(universe["UniverseRoom"]), "stable_key")
    groups = keyed(rows(bindings["UniverseEncounterGroup"]), "stable_key")
    selected_rooms = [rooms_by_key[key] for key in partition["record_ids"]]
    room_ids = {int(row["id"]) for row in selected_rooms}
    selected_content = [
        row
        for row in rows(universe["UniverseRoomContent"])
        if int(row["room_id"]) in room_ids
    ]
    content_by_room: dict[int, list[dict[str, Any]]] = {}
    for row in selected_content:
        content_by_room.setdefault(int(row["room_id"]), []).append(row)

    for authored in selected_rooms:
        reference = source_rooms[authored["stable_key"]]
        expected_room = {
            "stable_key": reference["id"],
            "domain_id": int(domains[reference["domain_id"]]["id"]),
            "source_room_id": str(reference["source_ids"][0]),
            "map_entrance": str(reference["map_entrance"]),
            "source_group_id": str(reference["source_group_id"]),
            "section_ids": "|".join(str(value) for value in reference["section_ids"]),
        }
        actual_room = {field: authored[field] for field in expected_room}
        if actual_room != expected_room:
            raise ValueError(
                f"{reference['id']}: Excel room differs from public source"
            )

        room_id = int(authored["id"])
        expected_content = []
        external = (
            source_domains[reference["domain_id"]]["decision_policy"]
            == "ExternalCommand"
        )
        for sequence, content in enumerate(reference["content_map"], start=1):
            source_content_id = str(content["content_source_id"])
            group_key = group_key_by_source.get(source_content_id)
            expected_content.append(
                {
                    "room_id": room_id,
                    "sequence": sequence,
                    "condition_key": str(content["group_id"]),
                    "source_content_id": source_content_id,
                    "kind": (
                        "EncounterGroup"
                        if group_key
                        else "ExternalDecision"
                        if external
                        else "FixedContent"
                    ),
                    "encounter_group_id": (
                        int(groups[group_key]["id"]) if group_key else None
                    ),
                }
            )
        authored_content = sorted(
            content_by_room.get(room_id, []),
            key=lambda row: int(row["sequence"]),
        )
        if authored_content != expected_content:
            raise ValueError(
                f"{reference['id']}: Excel room content differs from public source"
            )

    check_audits(partition, evidence)
    return {
        "UniverseRoom": selected_rooms,
        "UniverseRoomContent": selected_content,
    }


def build_world_partition(
    partition: dict[str, Any],
    universe: Any,
    bindings: Any,
    evidence: Any,
) -> dict[str, list[dict[str, Any]]]:
    source_worlds = keyed(
        json.loads((REFERENCE / "worlds.json").read_text(encoding="utf-8")),
        "id",
    )
    source_difficulties = keyed(
        json.loads(
            (REFERENCE / "world-difficulties.json").read_text(encoding="utf-8")
        ),
        "id",
    )
    worlds = keyed(rows(universe["UniverseWorld"]), "stable_key")
    difficulties = keyed(rows(universe["UniverseDifficulty"]), "stable_key")
    world_keys = [
        key for key in partition["record_ids"] if key in source_worlds
    ]
    difficulty_keys = [
        key for key in partition["record_ids"] if key in source_difficulties
    ]
    if len(world_keys) + len(difficulty_keys) != len(partition["record_ids"]):
        raise ValueError(f"{partition['id']}: unknown world/difficulty record")

    selected_worlds = [worlds[key] for key in world_keys]
    selected_difficulties = [difficulties[key] for key in difficulty_keys]
    difficulty_ids = {int(row["id"]) for row in selected_difficulties}
    selected_enemies = [
        row
        for row in rows(universe["UniverseDifficultyEnemy"])
        if int(row["difficulty_id"]) in difficulty_ids
    ]
    enemies_by_difficulty: dict[int, list[dict[str, Any]]] = {}
    for row in selected_enemies:
        enemies_by_difficulty.setdefault(int(row["difficulty_id"]), []).append(row)

    world_id_by_key = {
        row["stable_key"]: int(row["id"]) for row in selected_worlds
    }
    difficulty_id_by_key = {
        row["stable_key"]: int(row["id"]) for row in selected_difficulties
    }
    for authored in selected_worlds:
        reference = source_worlds[authored["stable_key"]]
        expected = {
            "stable_key": reference["id"],
            "profile_id": 1,
            "world_number": int(reference["world_id"]),
            "entry_rule_stable_key": reference["entry_rule_id"],
            "terminal_rule_stable_key": reference["terminal_rule_id"],
            "name_en": reference["name_en"],
            "name_zh_cn": reference["name_zh_cn"],
            "summary_en": reference["summary_en"],
            "summary_zh_cn": reference["summary_zh_cn"],
        }
        actual = {field: authored[field] for field in expected}
        if actual != expected:
            raise ValueError(
                f"{reference['id']}: Excel World differs from public source"
            )
        expected_difficulties = [
            difficulty_id_by_key[key] for key in reference["difficulty_ids"]
        ]
        actual_difficulties = [
            int(row["id"])
            for row in selected_difficulties
            if int(row["world_id"]) == int(authored["id"])
        ]
        if sorted(actual_difficulties) != sorted(expected_difficulties):
            raise ValueError(
                f"{reference['id']}: difficulty membership differs from source"
            )

    for authored in selected_difficulties:
        reference = source_difficulties[authored["stable_key"]]
        expected = {
            "stable_key": reference["id"],
            "world_id": world_id_by_key[reference["world_id"]],
            "source_area_id": str(reference["source_ids"][0]),
            "difficulty": int(reference["difficulty"]),
            "kind": reference["profile_kind"],
            "recommended_level": int(reference["recommended_level"]),
            "recommended_elements": "|".join(reference["recommended_elements"]),
            "score_curve_json": json.dumps(
                reference["score_curve"],
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ),
            "unlock_source_id": reference["unlock_source_id"] or None,
        }
        actual = {field: authored[field] for field in expected}
        if actual != expected:
            raise ValueError(
                f"{reference['id']}: Excel difficulty differs from public source"
            )

        difficulty_id = int(authored["id"])
        expected_enemies = []
        sequence = 1
        for role, values in (
            ("Boss", reference["boss_variant_ids"]),
            ("Elite", reference["elite_variant_ids"]),
        ):
            for value in values:
                expected_enemies.append(
                    {
                        "difficulty_id": difficulty_id,
                        "sequence": sequence,
                        "role": role,
                        "source_monster_id": str(value["source_monster_id"]),
                        "enemy_variant_stable_key": value["enemy_variant_id"],
                        "level": int(value["level"]),
                    }
                )
                sequence += 1
        actual_enemies = sorted(
            enemies_by_difficulty.get(difficulty_id, []),
            key=lambda row: int(row["sequence"]),
        )
        if actual_enemies != expected_enemies:
            raise ValueError(
                f"{reference['id']}: difficulty enemies differ from public source"
            )

    check_audits(partition, evidence)
    return {
        "UniverseWorld": selected_worlds,
        "UniverseDifficulty": selected_difficulties,
        "UniverseDifficultyEnemy": selected_enemies,
    }


def build(partition_id: str) -> dict[str, Any]:
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    partition = next(
        (value for value in manifest["partitions"] if value["id"] == partition_id),
        None,
    )
    if (
        partition is None
        or partition["mechanic_family"]
        != "enemies-encounters-worlds-difficulty-carry"
        or partition["lane"]
        not in {
            "domain-graph",
            "encounter-selection",
            "topology-map",
            "room-content",
            "world-difficulty",
        }
    ):
        raise ValueError(f"{partition_id}: unsupported world-structure partition")

    universe = load_workbook(DATA / "Universe.xlsx", read_only=True, data_only=False)
    bindings = load_workbook(
        DATA / "UniverseBindings.xlsx", read_only=True, data_only=False
    )
    evidence = load_workbook(
        DATA / "UniverseEvidence.xlsx", read_only=True, data_only=False
    )
    if partition["lane"] == "domain-graph":
        selected = build_domain_partition(partition, universe, bindings, evidence)
    elif partition["lane"] == "encounter-selection":
        selected = build_encounter_partition(
            partition, universe, bindings, evidence
        )
    elif partition["lane"] == "topology-map":
        selected = build_map_partition(partition, universe, evidence)
    elif partition["lane"] == "room-content":
        selected = build_room_partition(partition, universe, bindings, evidence)
    else:
        selected = build_world_partition(partition, universe, bindings, evidence)
    domain_ids = {
        int(row["id"]) for row in selected.get("UniverseDomain", [])
    }
    pool_ids = {
        int(row["id"]) for row in selected.get("UniverseEncounterPool", [])
    }
    node_ids = {
        int(row["id"]) for row in selected.get("UniverseMapNode", [])
    }
    room_ids = {
        int(row["id"]) for row in selected.get("UniverseRoom", [])
    }
    difficulty_ids = {
        int(row["id"]) for row in selected.get("UniverseDifficulty", [])
    }
    for table, selected_rows in selected.items():
        exported = sora_rows(table)
        if table == "UniverseDomain":
            keys = {row["stable_key"] for row in selected_rows}
            exported = [row for row in exported if row.get("stable_key") in keys]
        elif table == "UniverseActivityDomainBinding":
            exported = [
                row for row in exported if int(row.get("domain_id", -1)) in domain_ids
            ]
        elif table == "UniverseEncounterPool":
            keys = {row["stable_key"] for row in selected_rows}
            exported = [row for row in exported if row.get("stable_key") in keys]
        elif table in {"UniverseEncounterPoolGroup", "UniverseEncounterPoolFixed"}:
            exported = [
                row for row in exported if int(row.get("pool_id", -1)) in pool_ids
            ]
        elif table == "UniverseReviewFixture":
            keys = {row["stable_key"] for row in selected_rows}
            exported = [row for row in exported if row.get("stable_key") in keys]
        elif table == "UniverseMapNode":
            keys = {row["stable_key"] for row in selected_rows}
            exported = [row for row in exported if row.get("stable_key") in keys]
        elif table == "UniverseMapEdge":
            exported = [
                row
                for row in exported
                if int(row.get("source_node_id", -1)) in node_ids
            ]
        elif table == "UniverseRoom":
            keys = {row["stable_key"] for row in selected_rows}
            exported = [row for row in exported if row.get("stable_key") in keys]
        elif table == "UniverseRoomContent":
            exported = [
                row
                for row in exported
                if int(row.get("room_id", -1)) in room_ids
            ]
        elif table in {"UniverseWorld", "UniverseDifficulty"}:
            keys = {row["stable_key"] for row in selected_rows}
            exported = [row for row in exported if row.get("stable_key") in keys]
        elif table == "UniverseDifficultyEnemy":
            exported = [
                row
                for row in exported
                if int(row.get("difficulty_id", -1)) in difficulty_ids
            ]
        if digest(table, selected_rows) != digest(table, exported):
            raise ValueError(f"{partition_id}: Sora {table} rows differ from Excel")

    return {
        "schema_revision": "starclock.goal07-world-partition-golden.v1",
        "partition_id": partition_id,
        "lane": partition["lane"],
        "record_ids": partition["record_ids"],
        "rule_ids": partition["rule_ids"],
        "fixture_ids": partition["fixture_ids"],
        "tables": {
            table: {"rows": len(values), "semantic_sha256": digest(table, values)}
            for table, values in selected.items()
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--partition", required=True)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--check", action="store_true")
    mode.add_argument("--write", action="store_true")
    args = parser.parse_args()
    target = ROOT / "evidence" / GOAL / "goldens" / f"{args.partition}.json"
    encoded = json.dumps(build(args.partition), ensure_ascii=False, indent=2) + "\n"
    if args.write:
        target.parent.mkdir(parents=True, exist_ok=True)
        with target.open("w", encoding="utf-8", newline="\n") as stream:
            stream.write(encoded)
        print(f"Wrote {target.relative_to(ROOT)}.")
    elif not target.is_file() or target.read_text(encoding="utf-8") != encoded:
        raise ValueError(f"{args.partition}: world partition golden drifted")
    else:
        print(f"Goal 07 world partition {args.partition} matches Excel and Sora.")


if __name__ == "__main__":
    main()
