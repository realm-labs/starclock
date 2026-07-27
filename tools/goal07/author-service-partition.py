"""Verify one Goal 07 service partition in authoritative Excel and Sora."""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA

ROOT = Path(__file__).resolve().parents[2]
GOAL = "standard-universe-mechanics-complete-v1"
MANIFEST = ROOT / "content-manifests" / GOAL / "content-partitions.json"
DATA = ROOT / "config" / "data"
DEBUG = ROOT / "config" / "universe-generated" / "debug-json"


def rows(sheet: Any) -> list[dict[str, Any]]:
    fields = {str(cell.value): cell.column for cell in sheet[3] if cell.value not in (None, "#field")}
    output: list[dict[str, Any]] = []
    for cells in sheet.iter_rows(min_row=8):
        row = {field: cells[column - 1].value for field, column in fields.items()}
        if all(value is None for value in row.values()):
            continue
        if any(cells[column - 1].data_type in (TYPE_ERROR, TYPE_FORMULA) for column in fields.values()):
            raise ValueError(f"{sheet.title}: formulas and errors are forbidden")
        output.append(row)
    return output


def keyed(values: list[dict[str, Any]], field: str) -> dict[Any, dict[str, Any]]:
    output: dict[Any, dict[str, Any]] = {}
    for row in values:
        if row[field] in output:
            raise ValueError(f"duplicate {field}={row[field]}")
        output[row[field]] = row
    return output


def unwrap(value: Any) -> Any:
    if value == "Null":
        return None
    if isinstance(value, list):
        return [unwrap(item) for item in value]
    if isinstance(value, dict):
        if len(value) == 1:
            tag, inner = next(iter(value.items()))
            if tag in {"String", "Integer", "Bool", "Decimal"}:
                return inner
            if tag == "List":
                return [unwrap(item) for item in inner]
        return {key: unwrap(inner) for key, inner in value.items()}
    return value


def sora_rows(table: str) -> list[dict[str, Any]]:
    payload = json.loads((DEBUG / f"{table}.json").read_text(encoding="utf-8"))
    return [{key: unwrap(value) for key, value in row["values"].items()} for row in payload["table"]["rows"]]


def normalize(table: str, row: dict[str, Any]) -> dict[str, Any]:
    output = dict(row)
    if table == "UniverseMechanicRule" and isinstance(output.get("mechanic_tags"), str):
        output["mechanic_tags"] = [part for part in output["mechanic_tags"].split("|") if part]
    if table == "UniverseReviewFixture":
        for field in ("input_stable_keys", "provenance_ids"):
            if isinstance(output.get(field), str):
                parts = [part for part in output[field].split("|") if part]
                output[field] = [int(part) for part in parts] if field == "provenance_ids" else parts
    return output


def digest(table: str, values: list[dict[str, Any]]) -> str:
    normalized = [normalize(table, row) for row in values]
    normalized.sort(key=lambda row: json.dumps(row, ensure_ascii=False, sort_keys=True, default=str))
    encoded = json.dumps(normalized, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(encoded.encode()).hexdigest()


def build(partition_id: str) -> dict[str, Any]:
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    partition = next((value for value in manifest["partitions"] if value["id"] == partition_id), None)
    if partition is None or partition["mechanic_family"] != "services-shops-roster-interactables":
        raise ValueError(f"{partition_id}: not a service partition")
    universe = load_workbook(DATA / "Universe.xlsx", read_only=True, data_only=False)
    bindings = load_workbook(DATA / "UniverseBindings.xlsx", read_only=True, data_only=False)
    evidence = load_workbook(DATA / "UniverseEvidence.xlsx", read_only=True, data_only=False)
    services = keyed(rows(universe["UniverseService"]), "stable_key")
    selected_services = [services[key] for key in partition["record_ids"]]
    service_ids = {int(row["id"]) for row in selected_services}
    parameters = [
        row for row in rows(universe["UniverseServiceParameter"])
        if int(row["service_id"]) in service_ids
    ]
    sequences: dict[int, list[int]] = {}
    for row in parameters:
        sequences.setdefault(int(row["service_id"]), []).append(int(row["sequence"]))
    if any(sorted(value) != list(range(1, len(value) + 1)) for value in sequences.values()):
        raise ValueError(f"{partition_id}: service parameters are not contiguous")
    rules = keyed(rows(bindings["UniverseMechanicRule"]), "stable_key")
    selected_rules = [rules[key] for key in partition["rule_ids"]]
    fixtures = keyed(rows(evidence["UniverseReviewFixture"]), "stable_key")
    selected_fixtures = [fixtures[key] for key in partition["fixture_ids"]]
    audits = keyed(rows(evidence["UniverseContentAudit"]), "content_stable_key")
    sources = keyed(rows(evidence["UniverseSourceRecord"]), "id")
    for key in partition["record_ids"]:
        audit = audits.get(key)
        provenance = [] if audit is None else str(audit["provenance_ids"]).split("|")
        if audit is None or not audit["enabled"] or not provenance or any(int(item) not in sources for item in provenance):
            raise ValueError(f"{key}: enabled provenance audit is incomplete")
    selected = {
        "UniverseService": selected_services,
        "UniverseServiceParameter": parameters,
        "UniverseMechanicRule": selected_rules,
        "UniverseReviewFixture": selected_fixtures,
    }
    for table, selected_rows in selected.items():
        exported = sora_rows(table)
        if table == "UniverseServiceParameter":
            exported = [row for row in exported if int(row.get("service_id", -1)) in service_ids]
        else:
            keys = {row["stable_key"] for row in selected_rows}
            exported = [row for row in exported if row.get("stable_key") in keys]
        if digest(table, selected_rows) != digest(table, exported):
            raise ValueError(f"{partition_id}: Sora {table} rows differ from Excel")
    return {
        "schema_revision": "starclock.goal07-service-partition-golden.v2",
        "partition_id": partition_id,
        "record_ids": partition["record_ids"],
        "rule_ids": partition["rule_ids"],
        "fixture_ids": partition["fixture_ids"],
        "tables": {
            table: {"rows": len(values), "semantic_sha256": digest(table, values)}
            for table, values in selected.items()
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--partition", required=True)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--check", action="store_true")
    mode.add_argument("--write", action="store_true")
    args = parser.parse_args()
    target = ROOT / "evidence" / GOAL / "goldens" / f"{args.partition}.json"
    encoded = json.dumps(build(args.partition), ensure_ascii=False, indent=2) + "\n"
    if args.write:
        target.parent.mkdir(parents=True, exist_ok=True)
        with target.open("w", encoding="utf-8", newline="\n") as stream:
            stream.write(encoded)
        print(f"Wrote {target.relative_to(ROOT)}.")
    elif not target.is_file() or target.read_text(encoding="utf-8") != encoded:
        raise ValueError(f"{args.partition}: service partition golden drifted")
    else:
        print(f"Goal 07 service partition {args.partition} matches Excel and Sora.")


if __name__ == "__main__":
    main()
