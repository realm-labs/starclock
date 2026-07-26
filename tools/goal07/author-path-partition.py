"""Verify one Goal 07 path partition in authoritative Excel and Sora.

The complete Universe workbooks remain reproducibly authored by
tools/universe-reference/author_workbooks.py. This focused openpyxl gate proves
that one frozen path partition owns complete workbook rows and that Sora's
committed production export preserves their identities and relationships.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA


ROOT = Path(__file__).resolve().parents[2]
PARTITION_MANIFEST = (
    ROOT
    / "content-manifests"
    / "standard-universe-mechanics-complete-v1"
    / "content-partitions.json"
)
DATA = ROOT / "config" / "data"
DEBUG = ROOT / "config" / "universe-generated" / "debug-json"
BUNDLE = ROOT / "config" / "universe-generated" / "config.sora"
REFERENCE_LEVELS = (
    ROOT / "content-reference" / "standard-universe-v1" / "blessing-levels.json"
)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def fields(sheet: Any) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }


def rows(sheet: Any) -> list[dict[str, Any]]:
    columns = fields(sheet)
    result: list[dict[str, Any]] = []
    for cells in sheet.iter_rows(min_row=8):
        values = {
            field: cells[column - 1].value
            for field, column in columns.items()
        }
        if all(value is None for value in values.values()):
            continue
        for column in columns.values():
            cell = cells[column - 1]
            if cell.data_type in (TYPE_ERROR, TYPE_FORMULA):
                raise ValueError(
                    f"{sheet.title}/{cell.coordinate}: formulas and errors are forbidden"
                )
        result.append(values)
    return result


def keyed(items: list[dict[str, Any]], key: str) -> dict[Any, dict[str, Any]]:
    result: dict[Any, dict[str, Any]] = {}
    for item in items:
        value = item[key]
        if value in result:
            raise ValueError(f"duplicate {key}={value}")
        result[value] = item
    return result


def split_ids(value: Any) -> list[int]:
    if value in (None, ""):
        return []
    if isinstance(value, (int, float)):
        return [int(value)]
    return [int(part) for part in str(value).split("|") if part]


def unwrap(value: Any) -> Any:
    if value == "Null":
        return None
    if isinstance(value, list):
        return [unwrap(item) for item in value]
    if isinstance(value, dict):
        if len(value) == 1:
            tag, inner = next(iter(value.items()))
            if tag in {"String", "Integer", "Bool", "Decimal"}:
                return inner
            if tag == "List":
                return [unwrap(item) for item in inner]
        return {key: unwrap(inner) for key, inner in value.items()}
    return value


def sora_rows(table: str) -> list[dict[str, Any]]:
    payload = json.loads((DEBUG / f"{table}.json").read_text(encoding="utf-8"))
    return [
        {key: unwrap(value) for key, value in row["values"].items()}
        for row in payload["table"]["rows"]
    ]


def canonical_digest(items: list[dict[str, Any]]) -> str:
    encoded = json.dumps(
        items,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def normalized_for_sora(
    table: str, items: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    list_fields = {
        "UniverseBlessing": {"pool_tags", "mechanic_tags"},
        "UniverseMechanicRule": {"mechanic_tags"},
        "UniverseReviewFixture": {"input_stable_keys", "provenance_ids"},
    }.get(table, set())
    result: list[dict[str, Any]] = []
    for item in items:
        normalized = dict(item)
        for field in list_fields:
            value = normalized.get(field)
            if isinstance(value, str):
                parts = [part for part in value.split("|") if part]
                normalized[field] = (
                    [int(part) for part in parts]
                    if field == "provenance_ids"
                    else parts
                )
        result.append(normalized)
    return result


def partition(partition_id: str) -> dict[str, Any]:
    manifest = json.loads(PARTITION_MANIFEST.read_text(encoding="utf-8"))
    match = next(
        (item for item in manifest["partitions"] if item["id"] == partition_id),
        None,
    )
    if match is None or not str(match["mechanic_family"]).startswith("path-"):
        raise ValueError(f"{partition_id}: not a path partition")
    return match


def build_golden(partition_id: str) -> dict[str, Any]:
    assigned = partition(partition_id)
    workbook = load_workbook(DATA / "Universe.xlsx", read_only=True, data_only=False)
    bindings = load_workbook(
        DATA / "UniverseBindings.xlsx", read_only=True, data_only=False
    )
    evidence = load_workbook(
        DATA / "UniverseEvidence.xlsx", read_only=True, data_only=False
    )

    record_ids = set(assigned["record_ids"])
    assigned_path_keys = {
        item for item in record_ids if item.startswith("universe.path.")
    }
    assigned_blessing_keys = {
        item
        for item in record_ids
        if item.startswith("universe.blessing.") and ".level." not in item
    }
    level_keys = {item for item in record_ids if ".level." in item}
    if len(assigned_path_keys) > 1 or not level_keys:
        raise ValueError(f"{partition_id}: path/blessing/level assignment is incomplete")

    paths = keyed(rows(workbook["UniversePath"]), "stable_key")
    path_keys = assigned_path_keys or {
        f"universe.path.{str(assigned['mechanic_family']).removeprefix('path-')}"
    }
    selected_paths = [paths[key] for key in sorted(path_keys)]
    blessings = keyed(rows(workbook["UniverseBlessing"]), "stable_key")
    levels = keyed(rows(workbook["UniverseBlessingLevel"]), "stable_key")
    selected_levels = [levels[key] for key in sorted(level_keys)]
    blessing_by_id = {int(row["id"]): row for row in blessings.values()}
    blessing_keys = assigned_blessing_keys | {
        blessing_by_id[int(level["blessing_id"])]["stable_key"]
        for level in selected_levels
    }
    selected_blessings = [blessings[key] for key in sorted(blessing_keys)]
    blessing_ids = {int(row["id"]) for row in selected_blessings}
    level_ids = {int(row["id"]) for row in selected_levels}
    path_links = [
        row
        for row in rows(workbook["UniversePathBlessing"])
        if row["blessing_stable_key"] in blessing_keys
    ]
    parameters = [
        row
        for row in rows(workbook["UniverseBlessingParameter"])
        if int(row["blessing_level_id"]) in level_ids
    ]
    if {int(row["blessing_id"]) for row in selected_levels} - blessing_ids:
        raise ValueError(f"{partition_id}: assigned level has no assigned blessing")
    if {row["blessing_stable_key"] for row in path_links} != blessing_keys:
        raise ValueError(f"{partition_id}: path-to-blessing links are incomplete")
    parameter_level_ids = {int(row["blessing_level_id"]) for row in parameters}
    missing_parameter_levels = level_ids - parameter_level_ids
    if missing_parameter_levels:
        reference_levels = keyed(
            json.loads(REFERENCE_LEVELS.read_text(encoding="utf-8")),
            "id",
        )
        selected_by_id = {int(row["id"]): row for row in selected_levels}
        for level_id in sorted(missing_parameter_levels):
            stable_key = selected_by_id[level_id]["stable_key"]
            reference = reference_levels.get(stable_key)
            if reference is None or reference.get("parameter_values") != []:
                raise ValueError(
                    f"{partition_id}: {stable_key} has no exact parameter rows"
                )
    for blessing in selected_blessings:
        if (
            blessing["stable_key"] in assigned_blessing_keys
            and blessing["rule_stable_key"] not in assigned["rule_ids"]
        ):
            raise ValueError(f"{blessing['stable_key']}: assigned rule link differs")
    for level in selected_levels:
        if level["rule_stable_key"] not in assigned["rule_ids"]:
            raise ValueError(f"{level['stable_key']}: assigned level rule link differs")

    rules = keyed(rows(bindings["UniverseMechanicRule"]), "stable_key")
    selected_rules = [rules[key] for key in assigned["rule_ids"]]
    for rule in selected_rules:
        if rule["source_record_stable_key"] not in record_ids:
            raise ValueError(f"{rule['stable_key']}: assigned source link differs")
        if not rule["source_file"]:
            raise ValueError(f"{rule['stable_key']}: source-file provenance is missing")

    fixtures = keyed(rows(evidence["UniverseReviewFixture"]), "stable_key")
    selected_fixtures = [fixtures[key] for key in assigned["fixture_ids"]]
    audits = keyed(rows(evidence["UniverseContentAudit"]), "content_stable_key")
    sources = keyed(rows(evidence["UniverseSourceRecord"]), "id")
    for stable_key in assigned["record_ids"]:
        audit = audits.get(stable_key)
        if audit is None or not audit["enabled"]:
            raise ValueError(f"{stable_key}: enabled content audit is missing")
        provenance = split_ids(audit["provenance_ids"])
        if not provenance or any(source not in sources for source in provenance):
            raise ValueError(f"{stable_key}: provenance does not resolve")
    for fixture in selected_fixtures:
        provenance = split_ids(fixture["provenance_ids"])
        if not provenance or any(source not in sources for source in provenance):
            raise ValueError(f"{fixture['stable_key']}: provenance does not resolve")

    selected_by_table = {
        "UniversePath": selected_paths,
        "UniversePathBlessing": path_links,
        "UniverseBlessing": selected_blessings,
        "UniverseBlessingLevel": selected_levels,
        "UniverseBlessingParameter": parameters,
        "UniverseMechanicRule": selected_rules,
        "UniverseReviewFixture": selected_fixtures,
    }
    for table, selected in selected_by_table.items():
        exported = sora_rows(table)
        if table == "UniversePathBlessing":
            exported = [
                row
                for row in exported
                if row.get("blessing_stable_key") in blessing_keys
            ]
        elif table == "UniverseBlessingParameter":
            exported = [
                row
                for row in exported
                if int(row.get("blessing_level_id", -1)) in level_ids
            ]
        else:
            stable_keys = {item["stable_key"] for item in selected}
            exported = [
                row for row in exported if row.get("stable_key") in stable_keys
            ]
        if canonical_digest(normalized_for_sora(table, selected)) != canonical_digest(
            normalized_for_sora(table, exported)
        ):
            raise ValueError(f"{partition_id}: Sora {table} rows differ from Excel")

    return {
        "schema_revision": "starclock.goal07-path-partition-golden.v1",
        "partition_id": partition_id,
        "record_ids": assigned["record_ids"],
        "rule_ids": assigned["rule_ids"],
        "fixture_ids": assigned["fixture_ids"],
        "workbooks": {
            name: sha256(DATA / name)
            for name in (
                "Universe.xlsx",
                "UniverseBindings.xlsx",
                "UniverseEvidence.xlsx",
            )
        },
        "sora_bundle_sha256": sha256(BUNDLE),
        "tables": {
            name: {
                "rows": len(table_rows),
                "semantic_sha256": canonical_digest(table_rows),
            }
            for name, table_rows in selected_by_table.items()
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--partition", required=True)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--check", action="store_true")
    mode.add_argument("--write-golden", type=Path)
    args = parser.parse_args()

    golden = build_golden(args.partition)
    target = (
        ROOT
        / "evidence"
        / "standard-universe-mechanics-complete-v1"
        / "goldens"
        / f"{args.partition}.json"
    )
    if args.write_golden is not None:
        target = args.write_golden if args.write_golden.is_absolute() else ROOT / args.write_golden
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(
            json.dumps(golden, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        print(f"Wrote {args.partition} openpyxl/Sora golden to {target}.")
        return
    if not target.is_file():
        raise ValueError(f"{args.partition}: committed golden is missing")
    if json.loads(target.read_text(encoding="utf-8")) != golden:
        raise ValueError(f"{args.partition}: openpyxl/Sora golden drifted")
    print(f"{args.partition} authoritative Excel and Sora rows verified.")


if __name__ == "__main__":
    main()
