"""Author and verify Goal 07 enemy partitions in the core production workbooks.

S01 owns the Abundant Ebon Deer (Complete), S02 owns the Automaton Direwolf
(Complete), S03 owns the Automaton Grizzly (Complete), S04 owns the Blaze Out
of Space, S05 owns Cloud Knight Lieutenant: Yanqing (Complete), S06 owns
Cocolia (Complete), S07 owns Gepard (Complete), S08 owns Ice Out of Space,
S09 owns Memory Zone Meme "Something Unto Death" (Complete), S10 owns
Stellaron Hunter: Kafka (Complete), S11 owns Svarog (Complete), and S12 owns
the first frozen ordinary-enemy batch. S13 owns the second frozen
ordinary-enemy batch. Each partition receives an isolated
10,000-ID range so authoring and verification never consume rows owned by
another partition.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA


ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "config" / "data"
DEBUG = ROOT / "config" / "generated" / "debug-json"
PARTITIONS = (
    ROOT
    / "content-manifests"
    / "standard-universe-mechanics-complete-v1"
    / "content-partitions.json"
)
PARTITION_CONFIG = {
    "G07-P5-M15-S01": {
        "base": 980_000,
        "variant": "enemy.abundant-ebon-deer-complete.littleboss.variant.01",
        "source_record_id": 3,
        "evidence_record_id": 4,
    },
    "G07-P5-M15-S02": {
        "base": 990_000,
        "variant": "enemy.automaton-direwolf-complete.elite.variant.01",
        "source_record_id": 4,
        "evidence_record_id": 5,
    },
    "G07-P5-M15-S03": {
        "base": 1_000_000,
        "variant": "enemy.automaton-grizzly-complete.elite.variant.01",
        "source_record_id": 5,
        "evidence_record_id": 6,
    },
    "G07-P5-M15-S04": {
        "base": 1_010_000,
        "variant": "enemy.blaze-out-of-space.elite.variant.01",
        "source_record_id": 6,
        "evidence_record_id": 7,
    },
    "G07-P5-M15-S05": {
        "base": 1_020_000,
        "variant": "enemy.cloud-knight-lieutenant-yanqing-complete.littleboss.variant.01",
        "source_record_id": 7,
        "evidence_record_id": 8,
    },
    "G07-P5-M15-S06": {
        "base": 1_030_000,
        "variant": "enemy.cocolia-complete.littleboss.variant.01",
        "source_record_id": 8,
        "evidence_record_id": 9,
    },
    "G07-P5-M15-S07": {
        "base": 1_040_000,
        "variant": "enemy.gepard-complete.littleboss.variant.01",
        "source_record_id": 9,
        "evidence_record_id": 10,
    },
    "G07-P5-M15-S08": {
        "base": 1_050_000,
        "variant": "enemy.ice-out-of-space.elite.variant.01",
        "source_record_id": 10,
        "evidence_record_id": 11,
    },
    "G07-P5-M15-S09": {
        "base": 1_060_000,
        "variant": (
            "enemy.memory-zone-meme-something-unto-death-complete."
            "littleboss.variant.01"
        ),
        "source_record_id": 11,
        "evidence_record_id": 12,
    },
    "G07-P5-M15-S10": {
        "base": 1_070_000,
        "variant": "enemy.stellaron-hunter-kafka-complete.littleboss.variant.01",
        "source_record_id": 12,
        "evidence_record_id": 13,
    },
    "G07-P5-M15-S11": {
        "base": 1_080_000,
        "variant": "enemy.svarog-complete.littleboss.variant.01",
        "source_record_id": 13,
        "evidence_record_id": 14,
    },
    "G07-P5-M15-S12": {
        "base": 1_090_000,
        "variant": "enemy.abundance-sprite-golden-hound.minionlv2.variant.01",
        "variants": [
            "enemy.abundance-sprite-golden-hound.minionlv2.variant.01",
            "enemy.abundance-sprite-malefic-ape-bug.elite.variant.01",
            "enemy.abundance-sprite-malefic-ape.elite.variant.01",
            "enemy.abundance-sprite-wooden-lupus.minionlv2.variant.01",
            "enemy.antibaryon.minion.variant.01",
            "enemy.aurumaton-gatekeeper-bug.elite.variant.01",
            "enemy.aurumaton-gatekeeper.elite.variant.01",
            "enemy.aurumaton-spectral-envoy.elite.variant.01",
            "enemy.automaton-beetle.minionlv2.variant.01",
            "enemy.automaton-direwolf.elite.variant.01",
            "enemy.automaton-grizzly.elite.variant.01",
            "enemy.automaton-hound.minionlv2.variant.01",
        ],
        "source_record_id": 14,
        "evidence_record_id": 15,
    },
    "G07-P5-M15-S13": {
        "base": 1_100_000,
        "variant": "enemy.automaton-spider.minionlv2.variant.01",
        "variants": [
            "enemy.automaton-spider.minionlv2.variant.01",
            "enemy.baryon.minion.variant.01",
            "enemy.cloud-knights-patroller.minionlv2.variant.01",
            "enemy.decaying-shadow.elite.variant.01",
            (
                "enemy.disciples-of-sanctus-medicus-ballistarius."
                "minionlv2.variant.01"
            ),
            (
                "enemy.disciples-of-sanctus-medicus-internal-alchemist."
                "minionlv2.variant.01"
            ),
            (
                "enemy.disciples-of-sanctus-medicus-shape-shifter-bug."
                "elite.variant.01"
            ),
            (
                "enemy.disciples-of-sanctus-medicus-shape-shifter."
                "elite.variant.01"
            ),
            (
                "enemy.dreamjolt-troupes-beyond-overcooked-bug."
                "elite.variant.01"
            ),
            "enemy.dreamjolt-troupes-beyond-overcooked.elite.variant.01",
            "enemy.dreamjolt-troupes-birdskull.minion.variant.01",
            "enemy.dreamjolt-troupes-bubble-hound.minionlv2.variant.01",
        ],
        "source_record_id": 15,
        "evidence_record_id": 16,
    },
    "G07-P5-M15-S14": {
        "base": 1_110_000,
        "variant": "enemy.dreamjolt-troupes-mr-domescreen.minionlv2.variant.01",
        "variants": [
            "enemy.dreamjolt-troupes-mr-domescreen.minionlv2.variant.01",
            "enemy.dreamjolt-troupes-spring-loader.minion.variant.01",
            "enemy.dreamjolt-troupes-sweet-gorilla-bug.elite.variant.01",
            "enemy.dreamjolt-troupes-sweet-gorilla.elite.variant.01",
            "enemy.dreamjolt-troupes-winder-goon.minionlv2.variant.01",
            "enemy.entranced-ingenium-golden-cloud-toad.minion.variant.01",
            (
                "enemy.entranced-ingenium-illumination-dragonfish."
                "minionlv2.variant.01"
            ),
            "enemy.entranced-ingenium-obedient-dracolion.minion.variant.01",
            "enemy.everwinter-shadewalker.minionlv2.variant.01",
            "enemy.flamespawn.minion.variant.01",
            "enemy.frigid-prowler.elite.variant.01",
            "enemy.frostspawn.minion.variant.01",
        ],
        "source_record_id": 16,
        "evidence_record_id": 17,
    },
}
PARTITION = "G07-P5-M15-S01"
VARIANT_KEY = PARTITION_CONFIG[PARTITION]["variant"]
VARIANT_KEYS = [VARIANT_KEY]
BASE = PARTITION_CONFIG[PARTITION]["base"]
SOURCE_RECORD_ID = PARTITION_CONFIG[PARTITION]["source_record_id"]
EVIDENCE_RECORD_ID = PARTITION_CONFIG[PARTITION]["evidence_record_id"]


def anchor_path(partition: str) -> Path:
    return (
        ROOT
        / "evidence"
        / "standard-universe-mechanics-complete-v1"
        / "sources"
        / f"{partition}-numeric-anchors.json"
    )


def json_cell(type_: str, **fields: Any) -> str:
    return json.dumps({"type": type_, **fields}, separators=(",", ":"))


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_text(value: str) -> str:
    return sha256_bytes(value.encode("utf-8"))


def fields(sheet: Any) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }


def sheet_rows(sheet: Any) -> list[dict[str, Any]]:
    columns = fields(sheet)
    result: list[dict[str, Any]] = []
    for cells in sheet.iter_rows(min_row=8):
        row = {name: cells[column - 1].value for name, column in columns.items()}
        if all(value is None for value in row.values()):
            continue
        for column in columns.values():
            cell = cells[column - 1]
            if cell.data_type in (TYPE_ERROR, TYPE_FORMULA):
                raise ValueError(
                    f"{sheet.title}/{cell.coordinate}: formulas and errors are forbidden"
                )
        result.append(row)
    return result


def workbook_rows(table: str) -> list[dict[str, Any]]:
    workbook = load_workbook(DATA / f"{table}.xlsx", read_only=True, data_only=False)
    return sheet_rows(workbook.active)


def write_rows(table: str, rows: list[dict[str, Any]]) -> None:
    path = DATA / f"{table}.xlsx"
    workbook = load_workbook(path)
    sheet = workbook.active
    columns = fields(sheet)
    for row in rows:
        unknown = set(row) - set(columns)
        if unknown:
            raise ValueError(f"{table}: unknown fields {sorted(unknown)}")
    if sheet.max_row >= 8:
        sheet.delete_rows(8, sheet.max_row - 7)
    for row_index, row in enumerate(rows, start=8):
        for name, column in columns.items():
            value = row.get(name)
            if value is not None:
                sheet.cell(row=row_index, column=column, value=value)
    workbook.save(path)


def normalize(value: Any) -> Any:
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith(("{", "[")):
            try:
                return normalize(json.loads(stripped))
            except json.JSONDecodeError:
                pass
        return value
    if isinstance(value, list):
        return [normalize(item) for item in value]
    if isinstance(value, dict):
        if set(value) == {"Object"}:
            return normalize(value["Object"])
        return {key: normalize(inner) for key, inner in sorted(value.items())}
    return str(value)


def normalize_field(field: str, value: Any) -> Any:
    if field in {"ability_ids", "source_record_ids"}:
        if isinstance(value, str):
            value = [part for part in value.split("|") if part]
        if isinstance(value, list):
            return [normalize(item) for item in value]
    return normalize(value)


def semantic_digest(rows: list[dict[str, Any]]) -> str:
    projected = [
        {
            key: normalize_field(key, value)
            for key, value in sorted(row.items())
            if value is not None
        }
        for row in rows
    ]
    projected.sort(
        key=lambda row: json.dumps(
            row, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        )
    )
    return sha256_text(
        json.dumps(
            projected, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        )
    )


def unwrap(value: Any) -> Any:
    if value == "Null":
        return None
    if isinstance(value, list):
        return [unwrap(item) for item in value]
    if isinstance(value, dict):
        if len(value) == 1:
            tag, inner = next(iter(value.items()))
            if tag in {"String", "Integer", "Bool", "Decimal"}:
                return inner
            if tag == "List":
                return [unwrap(item) for item in inner]
        return {key: unwrap(inner) for key, inner in value.items()}
    return value


def sora_rows(table: str) -> list[dict[str, Any]]:
    payload = json.loads((DEBUG / f"{table}.json").read_text(encoding="utf-8"))
    return [
        {key: unwrap(value) for key, value in row["values"].items()}
        for row in payload["table"]["rows"]
    ]


def identity(
    id_: int,
    stable_key: str,
    kind: str,
    name_en: str,
    name_zh_cn: str,
    summary: str,
    sources: str = "1",
) -> dict[str, Any]:
    return {
        "id": id_,
        "stable_key": stable_key,
        "content_kind": kind,
        "name_en": name_en,
        "name_zh_cn": name_zh_cn,
        "summary_en": summary,
        "summary_zh_cn": "Goal 07 S01 来源绑定的丰饶玄鹿可执行定义。",
        "game_version_introduced": "1.2",
        "game_version_snapshot": "4.4",
        "release_state": "Released",
        "enabled": True,
        "coverage_state": "GoldenVerified",
        "source_record_ids": sources,
    }


def selector(
    id_: int,
    origin: str,
    side: str,
    life: str = "Alive",
    presence: str = "Present",
    minimum: int = 1,
    maximum: int = 1,
    empty: str = "Fault",
    choice: str = "First",
) -> dict[str, Any]:
    return {
        "id": id_,
        "domain": "Battle",
        "origin": origin,
        "side_relationship": side,
        "life": life,
        "presence": presence,
        "reference_point": "CurrentState",
        "ordering": "Formation",
        "minimum_count": minimum,
        "maximum_count": maximum,
        "empty_pool_policy": empty,
        "choice": choice,
        "allow_repeated_targets": False,
    }


def operation(
    id_: int,
    key: str,
    payload: str,
    target: int | None = None,
    empty: str = "Fault",
) -> dict[str, Any]:
    return {
        "id": id_,
        "stable_key": f"goal07.enemy.s01.operation.{key}",
        "domain": "Battle",
        "target_selector_id": target,
        "empty_target_policy": empty,
        "snapshot_boundary": "Dynamic",
        "fault_policy": "Rollback",
        "payload": payload,
    }


def owned_rows_s01() -> dict[str, list[dict[str, Any]]]:
    anchor = json.loads(anchor_path(PARTITION).read_text(encoding="utf-8"))
    manifest = json.loads(PARTITIONS.read_text(encoding="utf-8"))
    assigned = next(item for item in manifest["partitions"] if item["id"] == PARTITION)
    if assigned["enemy_variant_ids"] != [VARIANT_KEY]:
        raise ValueError("S01 frozen enemy assignment changed")

    variant = BASE + 1
    template = BASE + 2
    graphs = [BASE + 10, BASE + 11, BASE + 12]
    boss_abilities = {skill: BASE + 100 + skill for skill in range(1, 15)}
    branch_abilities = {
        "wintry": BASE + 121,
        "maple": BASE + 122,
        "glorious": BASE + 123,
        "lavish": BASE + 124,
    }
    linked = {
        "wintry-1": BASE + 201,
        "wintry-2": BASE + 202,
        "maple-1": BASE + 203,
        "maple-2": BASE + 204,
        "glorious-1": BASE + 205,
        "lavish-1": BASE + 206,
        "lavish-2": BASE + 207,
    }
    selectors = {
        "actor": BASE + 401,
        "owner": BASE + 402,
        "opposing-single": BASE + 403,
        "opposing-all": BASE + 404,
        "same-all": BASE + 405,
        "branches": BASE + 406,
        "lavish-branches": BASE + 407,
        "current-subject": BASE + 408,
        "maple-branches": BASE + 409,
        "primary-target": BASE + 410,
    }
    effects = {
        "wind-shear": BASE + 501,
        "shock": BASE + 502,
        "synwood-renewal": BASE + 503,
        "hardy-leaf": BASE + 504,
        "engender": BASE + 505,
        "vigor-overflow": BASE + 506,
        "maple-counter": BASE + 507,
        "outrage": BASE + 508,
    }
    modifiers = {
        "hardy-leaf-defense": BASE + 521,
        "engender-attack": BASE + 523,
        "vigor-damage": BASE + 524,
    }
    modifier_groups = {
        "hardy-leaf-defense": BASE + 522,
        "engender-attack": BASE + 525,
        "vigor-damage": BASE + 526,
    }
    programs: dict[str, int] = {}
    next_program = BASE + 301
    next_operation = BASE + 1_001
    next_expression = BASE + 1_101
    rows: dict[str, list[dict[str, Any]]] = {}

    def add(table: str, row: dict[str, Any]) -> None:
        rows.setdefault(table, []).append(row)

    identities = [
        identity(
            variant,
            VARIANT_KEY,
            "EnemyVariant",
            "Abundant Ebon Deer (Complete)",
            "丰饶玄鹿（完整）",
            "Goal 07 exact three-phase boss variant with reviewed Universe level rows.",
            "1|3",
        ),
        identity(
            template,
            "enemy.abundant-ebon-deer-complete.littleboss",
            "Enemy",
            "Abundant Ebon Deer (Complete) Template",
            "丰饶玄鹿（完整）模板",
            "Version 4.4 boss template retained from source monster 2024011.",
        ),
    ]
    for sequence, graph in enumerate(graphs, start=1):
        identities.append(
            identity(
                graph,
                f"ai.goal07.abundant-ebon-deer-complete.phase-{sequence}",
                "AiGraph",
                f"Abundant Ebon Deer Phase {sequence} AI",
                f"丰饶玄鹿第{sequence}阶段AI",
                "Finite source-ordered phase action graph.",
            )
        )
    for skill, ability in boss_abilities.items():
        identities.append(
            identity(
                ability,
                f"enemy.abundant-ebon-deer-complete.littleboss.ability.skill{skill:02d}",
                "Ability",
                f"Abundant Ebon Deer Skill {skill:02d}",
                f"丰饶玄鹿技能{skill:02d}",
                f"Executable transcription of source skill 2024011{skill:02d}.",
            )
        )
    for name, ability in branch_abilities.items():
        identities.append(
            identity(
                ability,
                f"enemy.abundant-ebon-deer-complete.branch.ability.{name}",
                "Ability",
                f"Abundant Ebon Deer {name.title()} Branch Action",
                f"丰饶玄鹿{name}枝条行动",
                "Executable linked-branch action used by the S01 boss.",
            )
        )
    for name, unit in linked.items():
        identities.append(
            identity(
                unit,
                f"unit.goal07.abundant-ebon-deer-complete.branch.{name}",
                "CharacterForm",
                f"Abundant Ebon Deer Branch {name}",
                f"丰饶玄鹿枝条{name}",
                "Internal linked-unit slot preserving one authored branch summon.",
            )
        )
    for name, selector_id in selectors.items():
        identities.append(
            identity(
                selector_id,
                f"selector.goal07.abundant-ebon-deer-complete.{name}",
                "Selector",
                f"Abundant Ebon Deer {name} Selector",
                f"丰饶玄鹿{name}选择器",
                "S01 battle selector.",
            )
        )
    for name, effect in effects.items():
        identities.append(
            identity(
                effect,
                f"effect.goal07.abundant-ebon-deer-complete.{name}",
                "Effect",
                f"Abundant Ebon Deer {name} Effect",
                f"丰饶玄鹿{name}效果",
                "S01 executable enemy effect.",
            )
        )
    for name, modifier in modifiers.items():
        display = {
            "hardy-leaf-defense": ("Hardy Leaf Defense", "蕉覆防御", "Exact +200% base DEF."),
            "engender-attack": ("Engender Attack", "繁生攻击", "Exact +30% base ATK for two turns."),
            "vigor-damage": (
                "Vigor Overflow Damage",
                "生机充盈伤害",
                "Exact +15% damage per enemy-granted buff occurrence.",
            ),
        }[name]
        identities.append(
            identity(
                modifier,
                f"modifier.goal07.abundant-ebon-deer-complete.{name}",
                "Modifier",
                display[0],
                display[1],
                display[2],
            )
        )

    add(
        "Selector",
        selector(selectors["actor"], "Actor", "SameSide"),
    )
    add(
        "Selector",
        selector(selectors["owner"], "Owner", "SameSide"),
    )
    add(
        "Selector",
        selector(selectors["current-subject"], "CurrentSubject", "SameSide"),
    )
    add(
        "Selector",
        selector(selectors["opposing-single"], "Actor", "OpposingSide"),
    )
    add(
        "Selector",
        selector(selectors["primary-target"], "PrimaryTarget", "OpposingSide"),
    )
    add(
        "Selector",
        selector(
            selectors["opposing-all"],
            "Actor",
            "OpposingSide",
            minimum=1,
            maximum=4,
            choice="All",
        ),
    )
    add(
        "Selector",
        selector(
            selectors["same-all"],
            "Actor",
            "SameSide",
            minimum=1,
            maximum=16,
            choice="All",
        ),
    )
    add(
        "Selector",
        selector(
            selectors["branches"],
            "Actor",
            "SameSide",
            presence="Linked",
            minimum=0,
            maximum=16,
            empty="NoOp",
            choice="All",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["branches"],
            "sequence": 1,
            "predicate": json_cell(
                "OwnedBy", owner_selector_id=selectors["actor"]
            ),
        },
    )
    add(
        "Selector",
        selector(
            selectors["maple-branches"],
            "Actor",
            "SameSide",
            presence="Linked",
            minimum=0,
            maximum=2,
            empty="NoOp",
            choice="All",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["maple-branches"],
            "sequence": 1,
            "predicate": json_cell(
                "OwnedBy", owner_selector_id=selectors["actor"]
            ),
        },
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["maple-branches"],
            "sequence": 2,
            "predicate": json_cell(
                "FormationRange", minimum_index=10, maximum_index=11
            ),
        },
    )
    add(
        "Selector",
        selector(
            selectors["lavish-branches"],
            "Actor",
            "SameSide",
            presence="Linked",
            minimum=0,
            maximum=2,
            empty="NoOp",
            choice="All",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["lavish-branches"],
            "sequence": 1,
            "predicate": json_cell(
                "OwnedBy", owner_selector_id=selectors["actor"]
            ),
        },
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["lavish-branches"],
            "sequence": 2,
            "predicate": json_cell(
                "FormationRange", minimum_index=13, maximum_index=14
            ),
        },
    )

    expression_ids: dict[str, int] = {}

    def expr(name: str, kind: str, node: str) -> int:
        nonlocal next_expression
        id_ = next_expression
        next_expression += 1
        expression_ids[name] = id_
        add(
            "ValueExpression",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s01.expression.{name}",
                "result_kind": kind,
                "node": node,
            },
        )
        return id_

    actor_atk = expr(
        "actor-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    owner_atk = expr(
        "owner-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["owner"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    actor_hp_heal = expr(
        "actor-hp-heal",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Hp",
            formula_purpose="Healing",
        ),
    )
    owner_hp_heal = expr(
        "owner-hp-heal",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["owner"],
            stat="Hp",
            formula_purpose="Healing",
        ),
    )
    ratios: dict[str, int] = {}
    for name, value in {
        "ratio-3-5": "3.5",
        "ratio-2-8": "2.8",
        "ratio-2-7": "2.7",
        "ratio-1-5": "1.5",
        "ratio-1-2": "1.2",
        "ratio-0-5": "0.5",
        "ratio-0-25": "0.25",
        "ratio-0-3": "0.3",
        "ratio-0-15": "0.15",
        "ratio-0-1": "0.1",
        "ratio-0-04": "0.04",
        "ratio-1": "1",
        "ratio-2": "2",
    }.items():
        ratios[name] = expr(
            name, "Scalar", json_cell("ScalarLiteral", value_decimal=value)
        )
    duration_two = expr(
        "duration-two", "Integer", json_cell("IntegerLiteral", value=2)
    )
    duration_one = expr(
        "duration-one", "Integer", json_cell("IntegerLiteral", value=1)
    )
    branch_count = expr(
        "branch-count",
        "Integer",
        json_cell("SelectorCount", selector_id=selectors["branches"]),
    )
    branch_count_scalar = expr(
        "branch-count-scalar",
        "Scalar",
        json_cell(
            "Convert",
            operand_expression_id=branch_count,
            target_kind="Scalar",
            rounding="NearestTiesAway",
        ),
    )
    vigor_stacks = expr(
        "vigor-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["current-subject"],
            effect_id=effects["vigor-overflow"],
        ),
    )
    vigor_stacks_scalar = expr(
        "vigor-stacks-scalar",
        "Scalar",
        json_cell(
            "Convert",
            operand_expression_id=vigor_stacks,
            target_kind="Scalar",
            rounding="NearestTiesAway",
        ),
    )

    def multiply(name: str, left: int, right: int) -> int:
        return expr(
            name,
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedMultiply",
                left_expression_id=left,
                right_expression_id=right,
                rounding="NearestTiesAway",
            ),
        )

    def add_expr(name: str, left: int, right: int) -> int:
        return expr(
            name,
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedAdd",
                left_expression_id=left,
                right_expression_id=right,
                rounding="NearestTiesAway",
            ),
        )

    damage_amounts = {
        "wavering": multiply("wavering-damage", actor_atk, ratios["ratio-3-5"]),
        "caress": multiply("caress-damage", actor_atk, ratios["ratio-2-8"]),
        "wintry": multiply("wintry-damage", actor_atk, ratios["ratio-1-2"]),
        "maple": multiply("maple-damage", actor_atk, ratios["ratio-1-5"]),
        "maple-counter": multiply(
            "maple-counter-damage", owner_atk, ratios["ratio-1-5"]
        ),
    }
    vigor_damage_boost = multiply(
        "vigor-damage-boost",
        vigor_stacks_scalar,
        ratios["ratio-0-15"],
    )
    gore_bonus = multiply(
        "gore-branch-bonus",
        branch_count_scalar,
        ratios["ratio-0-5"],
    )
    gore_ratio = add_expr("gore-ratio", ratios["ratio-2-7"], gore_bonus)
    damage_amounts["gore"] = multiply("gore-damage", actor_atk, gore_ratio)
    heal_amounts = {
        "everlife": multiply("everlife-heal", actor_hp_heal, ratios["ratio-0-1"]),
        "renewal": multiply("renewal-heal", owner_hp_heal, ratios["ratio-0-04"]),
        "last-spring": multiply(
            "last-spring-heal", actor_hp_heal, ratios["ratio-0-25"]
        ),
    }
    dot_amounts = {
        "wind-shear": multiply(
            "wind-shear-damage", actor_atk, ratios["ratio-0-5"]
        ),
        "shock": multiply("shock-damage", actor_atk, ratios["ratio-0-5"]),
    }

    def program(name: str, operations: list[dict[str, Any]]) -> int:
        nonlocal next_program
        id_ = next_program
        next_program += 1
        programs[name] = id_
        identities.append(
            identity(
                id_,
                f"program.goal07.abundant-ebon-deer-complete.{name}",
                "Program",
                f"Abundant Ebon Deer {name} Program",
                f"丰饶玄鹿{name}程序",
                "Ordered Rule IR program for the S01 enemy.",
            )
        )
        add("Program", {"id": id_, "domain": "Battle"})
        for sequence, row in enumerate(operations, start=1):
            add(
                "ProgramStep",
                {
                    "program_id": id_,
                    "sequence": sequence,
                    "step": json_cell("Operation", operation_id=row["id"]),
                },
            )
            add("Operation", row)
        return id_

    def damage_op(
        name: str,
        amount: int,
        element: str,
        all_targets: bool,
        target_selector: int | None = None,
    ) -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        return operation(
            id_,
            name,
            json_cell(
                "Damage",
                amount_expression_id=amount,
                damage_class="Ordinary",
                element=element,
                can_crit=True,
            ),
            target_selector
            if target_selector is not None
            else selectors["opposing-all" if all_targets else "opposing-single"],
        )

    def heal_op(name: str, amount: int, target: int) -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        return operation(
            id_,
            name,
            json_cell("Heal", amount_expression_id=amount),
            target,
        )

    def apply_effect_op(
        name: str,
        effect: int,
        target: int,
        resistible: bool,
        stacks_expression_id: int | None = None,
    ) -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        return operation(
            id_,
            name,
            json_cell(
                "ApplyEffect",
                effect_id=effect,
                stacks_expression_id=stacks_expression_id,
                chance_policy="Resistible" if resistible else "Guaranteed",
                base_chance_expression_id=ratios["ratio-1"] if resistible else None,
                rng_purpose_key="effect-application" if resistible else None,
            ),
            target,
        )

    def summon_op(name: str, unit: int) -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        return operation(
            id_,
            name,
            json_cell(
                "Summon",
                unit_definition_identity_id=unit,
                owner_selector_id=selectors["actor"],
            ),
            None,
        )

    def grant_extra_turn_op(name: str) -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        return operation(
            id_,
            name,
            json_cell(
                "GrantExtraTurn",
                actor_selector_id=selectors["current-subject"],
            ),
            None,
        )

    def despawn_lavish_branches_op(name: str) -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        return operation(
            id_,
            name,
            json_cell("Despawn"),
            selectors["lavish-branches"],
            "NoOp",
        )

    ability_programs: dict[int, int] = {}
    ability_programs[boss_abilities[5]] = program(
        "wavering-bleat",
        [
            damage_op(
                "wavering-bleat-damage",
                damage_amounts["wavering"],
                "Lightning",
                False,
                selectors["primary-target"],
            )
        ],
    )
    ability_programs[boss_abilities[6]] = program(
        "caress-of-wind",
        [damage_op("caress-of-wind-damage", damage_amounts["caress"], "Wind", True)],
    )
    ability_programs[boss_abilities[8]] = program(
        "flamboyant-gore",
        [
            damage_op("flamboyant-gore-damage", damage_amounts["gore"], "Lightning", True),
            despawn_lavish_branches_op("gore-despawn-lavish-branches"),
        ],
    )
    ability_programs[boss_abilities[9]] = program(
        "everlife",
        [
            heal_op("everlife-heal", heal_amounts["everlife"], selectors["actor"]),
            apply_effect_op(
                "everlife-synwood-renewal",
                effects["synwood-renewal"],
                selectors["actor"],
                False,
            ),
            apply_effect_op(
                "everlife-vigor-owner",
                effects["vigor-overflow"],
                selectors["actor"],
                False,
            ),
            apply_effect_op(
                "everlife-vigor-branches",
                effects["vigor-overflow"],
                selectors["branches"],
                False,
            ),
        ],
    )
    ability_programs[boss_abilities[7]] = program(
        "hardy-leaf-sheath",
        [
            apply_effect_op(
                "hardy-leaf-sheath",
                effects["hardy-leaf"],
                selectors["actor"],
                False,
            ),
            apply_effect_op(
                "hardy-leaf-vigor-owner",
                effects["vigor-overflow"],
                selectors["actor"],
                False,
            ),
            apply_effect_op(
                "hardy-leaf-vigor-branches",
                effects["vigor-overflow"],
                selectors["branches"],
                False,
            ),
        ],
    )
    for skill, names in {
        10: ["wintry-1", "wintry-2"],
        11: ["maple-1", "maple-2"],
        12: ["maple-1", "maple-2", "wintry-1", "glorious-1"],
        13: ["lavish-1", "lavish-2", "maple-1", "wintry-1"],
    }.items():
        summon_operations = [
            summon_op(f"summon-skill-{skill:02d}-{name}", linked[name])
            for name in names
        ]
        if any(name.startswith("maple-") for name in names):
            summon_operations.append(
                apply_effect_op(
                    f"summon-skill-{skill:02d}-maple-counter",
                    effects["maple-counter"],
                    selectors["maple-branches"],
                    False,
                )
            )
        if skill == 13:
            summon_operations.extend(
                [
                    damage_op(
                        "entwined-vines-immediate-gore",
                        damage_amounts["gore"],
                        "Lightning",
                        True,
                    ),
                    despawn_lavish_branches_op(
                        "entwined-vines-despawn-lavish-branches"
                    ),
                ]
            )
        ability_programs[boss_abilities[skill]] = program(
            f"summon-skill-{skill:02d}",
            summon_operations,
        )
    ability_programs[branch_abilities["wintry"]] = program(
        "wintry-branch-action",
        [
            damage_op(
                "wintry-branch-damage",
                damage_amounts["wintry"],
                "Wind",
                False,
                selectors["primary-target"],
            ),
            apply_effect_op(
                "wintry-branch-wind-shear",
                effects["wind-shear"],
                selectors["primary-target"],
                True,
            ),
            apply_effect_op(
                "wintry-branch-outrage",
                effects["outrage"],
                selectors["primary-target"],
                True,
            ),
        ],
    )
    ability_programs[branch_abilities["maple"]] = program(
        "maple-branch-action",
        [
            apply_effect_op(
                "maple-engender",
                effects["engender"],
                selectors["owner"],
                False,
            ),
            apply_effect_op(
                "maple-vigor-owner",
                effects["vigor-overflow"],
                selectors["owner"],
                False,
            ),
            apply_effect_op(
                "maple-vigor-branches",
                effects["vigor-overflow"],
                selectors["branches"],
                False,
            ),
        ],
    )
    ability_programs[branch_abilities["glorious"]] = program(
        "glorious-branch-action",
        [heal_op("glorious-last-spring", heal_amounts["last-spring"], selectors["same-all"])],
    )

    renewal_program = program(
        "synwood-renewal-turn-heal",
        [heal_op("synwood-renewal-turn-heal", heal_amounts["renewal"], selectors["owner"])],
    )
    maple_counter_program = program(
        "maple-counter",
        [
            damage_op(
                "maple-counter-damage",
                damage_amounts["maple-counter"],
                "Lightning",
                False,
                selectors["actor"],
            ),
            apply_effect_op(
                "maple-counter-shock",
                effects["shock"],
                selectors["actor"],
                True,
            ),
        ],
    )
    phase_three_entry_program = program(
        "phase-three-extra-actions",
        [
            grant_extra_turn_op("phase-three-extra-action-one"),
            grant_extra_turn_op("phase-three-extra-action-two"),
        ],
    )

    target_patterns = {
        5: "SingleTarget",
        6: "Aoe",
        8: "Aoe",
        7: "Enhance",
        9: "Enhance",
        10: "Enhance",
        11: "Enhance",
        12: "Enhance",
        13: "Enhance",
    }
    for skill, ability in boss_abilities.items():
        kind = "Passive" if skill == 14 else "Skill"
        pattern = target_patterns.get(skill, "None")
        add(
            "Ability",
            {
                "id": ability,
                "kind": kind,
                "target_pattern": pattern,
                "retarget_policy": "CancelRemaining",
                "level_cap": 1,
                "cooldown_actions": 1,
                "semantic_tags_mask": 5 if skill in {5, 6, 8, 14} else 4,
            },
        )
        add(
            "AbilityPhase",
            {
                "ability_id": ability,
                "sequence": 1,
                "kind": "Resolved",
                "program_identity_id": ability_programs.get(ability),
            },
        )
        add(
            "EnemyAbility",
            {
                "id": ability,
                "telegraph": "Charge" if skill == 8 else "None",
                "cooldown_actions": 1,
                "initial_cooldown_actions": 0,
                "charge_actions": 0,
                "ai_tag": f"skill{skill:02d}",
            },
        )
    for name, ability in branch_abilities.items():
        add(
            "Ability",
            {
                "id": ability,
                "kind": "Summon",
                "target_pattern": (
                    "Support" if name == "glorious" else
                    "None" if name == "lavish" else
                    "SingleTarget"
                ),
                "retarget_policy": "CancelRemaining",
                "level_cap": 1,
                "cooldown_actions": 0,
                "semantic_tags_mask": 5 if name in {"wintry", "maple"} else 4,
            },
        )
        add(
            "AbilityPhase",
            {
                "ability_id": ability,
                "sequence": 1,
                "kind": "Resolved",
                "program_identity_id": ability_programs.get(ability),
            },
        )
        add(
            "EnemyAbility",
            {
                "id": ability,
                "telegraph": "None",
                "cooldown_actions": 0,
                "initial_cooldown_actions": 0,
                "charge_actions": 0,
                "ai_tag": f"branch-{name}",
            },
        )

    for name, effect in effects.items():
        is_dot = name in {"wind-shear", "shock"}
        permanent = name in {"vigor-overflow", "maple-counter"}
        add(
            "Effect",
            {
                "id": effect,
                "category": (
                    "Dot" if is_dot else
                    "Control" if name == "outrage" else
                    "NeutralState" if name == "maple-counter" else
                    "Buff"
                ),
                "dispel_category": (
                    "DispellableDebuff" if is_dot else
                    "CleanseableControl" if name == "outrage" else
                    "NonDispellable" if permanent else
                    "DispellableBuff"
                ),
                "stack_limit": (
                    5 if name == "wind-shear" else
                    100 if name == "vigor-overflow" else
                    1
                ),
                "duration_expression_id": (
                    None if permanent else
                    duration_two if name in {"wind-shear", "shock", "engender", "outrage"} else
                    duration_one
                ),
                "duration_clock": "Permanent" if permanent else "TargetTurnEnd",
                "tick_phase": "TurnStart" if is_dot else "None",
                "stack_policy": (
                    "RefreshAndAddStacks"
                    if name in {"wind-shear", "vigor-overflow"}
                    else "Refresh"
                ),
                "magnitude_comparator_expression_id": dot_amounts.get(name),
                "dot_element": (
                    "Wind" if name == "wind-shear" else
                    "Lightning" if name == "shock" else None
                ),
                "snapshot_policy": "OnApplication",
                "teardown_policy": "RemoveWithOwner",
                "application_priority": 0,
            },
        )

    add(
        "EffectTag",
        {
            "effect_id": effects["hardy-leaf"],
            "sequence": 1,
            "tag": "prevents-toughness-reduction",
        },
    )
    add(
        "EffectTag",
        {
            "effect_id": effects["outrage"],
            "sequence": 1,
            "tag": "forced-basic-attack-random-ally",
        },
    )

    for name, group in modifier_groups.items():
        add(
            "ModifierStackingGroup",
            {
                "id": group,
                "stable_key": f"goal07.enemy.s01.{name}",
                "aggregation": "Sum",
            },
        )
    modifier_rows = {
        "hardy-leaf-defense": {
            "effect": "hardy-leaf",
            "stat": "Def",
            "stage": "PercentOfBase",
            "value": ratios["ratio-2"],
            "snapshot": "OnApplication",
        },
        "engender-attack": {
            "effect": "engender",
            "stat": "Atk",
            "stage": "PercentOfBase",
            "value": ratios["ratio-0-3"],
            "snapshot": "OnApplication",
        },
        "vigor-damage": {
            "effect": "vigor-overflow",
            "stat": "Atk",
            "stage": "DamageBoost",
            "value": vigor_damage_boost,
            "snapshot": "Dynamic",
        },
    }
    for name, modifier in modifiers.items():
        definition = modifier_rows[name]
        add(
            "ModifierDefinition",
            {
                "id": modifier,
                "source_effect_id": effects[definition["effect"]],
                "owner_selector_id": selectors["owner"],
                "subject_selector_id": selectors["current-subject"],
                "stat": definition["stat"],
                "formula_stage": definition["stage"],
                "formula_purpose": "Stat",
                "value_expression_id": definition["value"],
                "value_domain": "Ratio",
                "stacking_group_id": modifier_groups[name],
                "priority": 0,
                "cap_formula_stage": definition["stage"],
                "snapshot_policy": definition["snapshot"],
                "duration_scope": "Turn",
            },
        )
        add(
            "EffectModifierBinding",
            {
                "effect_id": effects[definition["effect"]],
                "sequence": 1,
                "modifier_id": modifier,
            },
        )

    regen_rule = BASE + 541
    regen_filter = BASE + 542
    regen_trigger = BASE + 543
    counter_rule = BASE + 544
    counter_filter = BASE + 545
    counter_trigger = BASE + 546
    identities.append(
        identity(
            regen_rule,
            "rule.goal07.abundant-ebon-deer-complete.synwood-renewal",
            "Rule",
            "Synwood Renewal Turn Heal",
            "古木逢春回合治疗",
            "Effect-owned turn-start heal rule.",
        )
    )
    add(
        "RuleDefinition",
        {
            "id": regen_rule,
            "domain": "Battle",
            "source_definition_identity_id": effects["synwood-renewal"],
            "source_class": "Effect",
            "source_digest_sha256": (
                "316d3542e27f1a899ceed57eab31bdbcbd60ec6dd4168f638f00d75b80d0590a"
            ),
        },
    )
    add(
        "EventFilter",
        {
            "id": regen_filter,
            "stable_key": "goal07.enemy.s01.filter.synwood-renewal-owner-turn",
            "owner_selector_id": selectors["owner"],
            "cause_ancestry": "Any",
        },
    )
    add(
        "RuleTrigger",
        {
            "id": regen_trigger,
            "stable_key": "goal07.enemy.s01.trigger.synwood-renewal-owner-turn",
            "rule_id": regen_rule,
            "sequence": 1,
            "event": json_cell("Turn", point="Started"),
            "phase": "AfterEvent",
            "filter_id": regen_filter,
            "condition_id": BASE + 551,
            "once_scope": "Event",
            "priority": 0,
            "program_id": renewal_program,
        },
    )
    add(
        "EffectRuleBinding",
        {
            "effect_id": effects["synwood-renewal"],
            "sequence": 1,
            "rule_id": regen_rule,
        },
    )
    identities.append(
        identity(
            counter_rule,
            "rule.goal07.abundant-ebon-deer-complete.maple-retaliation",
            "Rule",
            "Maple Retaliation",
            "缃叶反击",
            "Damage-targeted immediate Leaf Hinging counter rule.",
        )
    )
    add(
        "RuleDefinition",
        {
            "id": counter_rule,
            "domain": "Battle",
            "source_definition_identity_id": effects["maple-counter"],
            "source_class": "Effect",
            "source_digest_sha256": (
                "a43312b59cc4ea5ef3c70e0c65594b5cbd80beb15c78d39cc73a33b66568fe70"
            ),
        },
    )
    add(
        "EventFilter",
        {
            "id": counter_filter,
            "stable_key": "goal07.enemy.s01.filter.maple-retaliation",
            "target_selector_id": selectors["owner"],
            "damage_class": "Ordinary",
            "cause_ancestry": "Any",
        },
    )
    add(
        "RuleTrigger",
        {
            "id": counter_trigger,
            "stable_key": "goal07.enemy.s01.trigger.maple-retaliation",
            "rule_id": counter_rule,
            "sequence": 1,
            "event": json_cell("Damage", point="Applied"),
            "phase": "AfterEvent",
            "filter_id": counter_filter,
            "condition_id": BASE + 551,
            "once_scope": "Event",
            "priority": 0,
            "program_id": maple_counter_program,
        },
    )
    add(
        "EffectRuleBinding",
        {
            "effect_id": effects["maple-counter"],
            "sequence": 1,
            "rule_id": counter_rule,
        },
    )

    add(
        "ConditionExpression",
        {
            "id": BASE + 551,
            "stable_key": "goal07.enemy.s01.condition.always",
            "node": json_cell("Constant", value=True),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": BASE + 552,
            "stable_key": "goal07.enemy.s01.condition.no-branches",
            "node": json_cell(
                "SelectorCardinality",
                selector_id=selectors["branches"],
                minimum_count=0,
                maximum_count=0,
            ),
        },
    )
    phase_sequences = [
        [10, 11, 5, 5, 6],
        [12, 9, 5, 5, 6],
        [13, 7, 5, 5, 6],
    ]
    next_state = BASE + 701
    next_candidate = BASE + 801
    next_transition = BASE + 901
    phase_initial_states: list[int] = []
    for phase_index, skills in enumerate(phase_sequences):
        state_ids = list(range(next_state, next_state + len(skills)))
        next_state += len(skills)
        phase_initial_states.append(state_ids[0])
        add(
            "AiGraph",
            {
                "id": graphs[phase_index],
                "initial_state_id": state_ids[0],
                "automatic_transition_budget": 8,
            },
        )
        for offset, (state_id, skill) in enumerate(zip(state_ids, skills)):
            main_ability = boss_abilities[skill]
            add(
                "AiState",
                {
                    "id": state_id,
                    "stable_key": (
                        f"goal07.enemy.s01.ai.phase-{phase_index + 1}.state-{offset + 1}"
                    ),
                    "graph_id": graphs[phase_index],
                    "mandatory_fallback_ability_id": boss_abilities[5],
                    "turn_counter_reset": offset == 0,
                },
            )
            sequence = 1
            condition = (
                BASE + 552 if skill in {10, 11, 12, 13} else BASE + 551
            )
            target = (
                selectors["opposing-single"] if skill == 5 else
                selectors["opposing-all"] if skill in {6, 8} else
                selectors["actor"]
            )
            add(
                "AiCandidate",
                {
                    "id": next_candidate,
                    "stable_key": (
                        f"goal07.enemy.s01.ai.phase-{phase_index + 1}."
                        f"state-{offset + 1}.main"
                    ),
                    "state_id": state_id,
                    "sequence": sequence,
                    "ability_id": main_ability,
                    "condition_id": condition,
                    "target_selector_id": target,
                    "priority": 10,
                    "selection": "FirstLegal",
                    "no_target_fallback": "UseFallbackAbility",
                    "fallback_ability_id": boss_abilities[5],
                },
            )
            next_candidate += 1
            add(
                "AiTransition",
                {
                    "id": next_transition,
                    "stable_key": (
                        f"goal07.enemy.s01.ai.phase-{phase_index + 1}."
                        f"transition-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "target_state_id": state_ids[(offset + 1) % len(state_ids)],
                    "condition_id": BASE + 551,
                    "priority": 0,
                    "timing": "AfterAction",
                },
            )
            next_transition += 1

    add(
        "EnemyTemplate",
        {
            "id": template,
            "rank": "Boss",
            "base_aggro_decimal": "100",
            "default_ai_graph_id": graphs[0],
        },
    )
    add(
        "EnemyVariant",
        {
            "id": variant,
            "template_id": template,
            "ai_graph_id": graphs[0],
            "mechanically_distinct_key": VARIANT_KEY,
        },
    )
    for level in anchor["levels"]:
        add(
            "EnemyStat",
            {
                "variant_id": variant,
                "level": level["authored_level"],
                "difficulty_key": "standard-universe-v1",
                "hp_decimal": level["base_hp"],
                "atk_decimal": level["base_atk"],
                "def_decimal": level["base_def"],
                "spd_decimal": level["base_spd"],
                "effect_hit_rate_decimal": level["effect_hit_rate"],
                "effect_resistance_decimal": level["effect_resistance"],
                "crit_damage_decimal": "0.2",
            },
        )
    for sequence, weakness in enumerate(["Fire", "Ice", "Quantum"], start=1):
        add(
            "EnemyWeakness",
            {"variant_id": variant, "sequence": sequence, "element": weakness},
        )
    for element in ["Physical", "Lightning", "Wind", "Imaginary"]:
        add(
            "EnemyResistance",
            {"variant_id": variant, "element": element, "value_decimal": "0.2"},
        )
    for category, value in [("Frozen", "0.5"), ("Imprisonment", "0.5")]:
        add(
            "EnemyDebuffResistance",
            {
                "variant_id": variant,
                "category_key": category,
                "value_decimal": value,
            },
        )
    add(
        "EnemyToughnessLayer",
        {
            "variant_id": variant,
            "sequence": 1,
            "layer_key": "ordinary",
            "kind": "Ordinary",
            "maximum_decimal": "420",
            "recovery_ratio_decimal": "1",
            "active_at_start": True,
        },
    )
    for sequence, skill in enumerate(range(1, 15), start=1):
        add(
            "EnemyVariantAbility",
            {
                "variant_id": variant,
                "sequence": sequence,
                "ability_id": boss_abilities[skill],
            },
        )
    for sequence, graph in enumerate(graphs, start=1):
        add(
            "EnemyPhase",
            {
                "id": BASE + 600 + sequence,
                "stable_key": f"goal07.enemy.s01.phase-{sequence}",
                "variant_id": variant,
                "sequence": sequence,
                "entry_condition_id": BASE + 551,
                "exit_condition_id": BASE + 551,
                "replacement_priority": sequence,
                "ai_graph_id": graph,
                "targetable": True,
                "transition_model": "TransformSameUnit",
                "entry_program_id": (
                    phase_three_entry_program if sequence == 3 else None
                ),
                "hp_carry": "Reset",
                "action_gauge_carry": "Reset",
                "effect_carry": "Clear",
                "toughness_carry": "Reset",
                "summon_carry": "Clear",
            },
        )

    branch_type = {
        "wintry": (branch_abilities["wintry"], "0.173333", "120"),
        "maple": (branch_abilities["maple"], "0.266667", "144"),
        "glorious": (branch_abilities["glorious"], "0.266667", "120"),
        "lavish": (branch_abilities["lavish"], "0.4", "83"),
    }
    for formation, (name, unit) in enumerate(linked.items(), start=8):
        kind = name.split("-")[0]
        action, hp_ratio, speed = branch_type[kind]
        add(
            "LinkedUnitDefinition",
            {
                "id": unit,
                "source_definition_identity_id": unit,
                "kind": "Summon",
                "presence": "Linked",
                "ability_ids": str(action),
                "action_ability_id": None if kind == "lavish" else action,
                "formation_index": formation,
                "initial_gauge_decimal": "10000",
                "hp_owner_ratio_decimal": hp_ratio,
                "hp_flat_decimal": "0",
                "atk_owner_ratio_decimal": "1",
                "atk_flat_decimal": "0",
                "def_owner_ratio_decimal": "1",
                "def_flat_decimal": "0",
                "spd_owner_ratio_decimal": "0",
                "spd_flat_decimal": speed,
                "owner_defeat_policy": "Depart",
                "owner_departure_policy": "Depart",
                "wave_policy": "Depart",
                "combatant_digest_sha256": sha256_text(
                    f"goal07-s01-linked-{name}-v1"
                ),
            },
        )

    anchor_digest = sha256_bytes(anchor_path(PARTITION).read_bytes())
    add(
        "SourceRecord",
        {
            "id": 3,
            "stable_key": "source.hsr-wiki.abundant-ebon-deer-complete.2026-07-28",
            "category": "CommunityMaintained",
            "publisher": "Honkai: Star Rail Wiki",
            "url": anchor["source"]["url"],
            "accessed_on": anchor["source"]["accessed_on"],
            "applicable_game_version": anchor["source"]["game_version"],
            "confidence": "SecondaryVersionSensitiveCrossCheck",
            "evidence_sha256": anchor_digest,
            "usage_note": (
                "Exact public per-level HP, ATK, DEF, SPD, EHR and Effect RES "
                "transcribed into committed Goal 07 evidence."
            ),
        },
    )
    add(
        "EvidenceRecord",
        {
            "id": 4,
            "stable_key": "evidence.goal07.enemy.s01.numeric-anchors",
            "kind": "SourcePayload",
            "source_record_id": 3,
            "sha256": anchor_digest,
            "note": "Committed exact public per-level numeric anchors for Goal 07 S01.",
        },
    )
    for item in identities:
        add("ContentIdentity", item)
        add(
            "ContentEvidenceBinding",
            {
                "content_id": item["id"],
                "sequence": 1,
                "fact_key": f"goal07.s01.executable:{item['stable_key']}",
                "source_record_id": 1,
                "evidence_record_id": 3,
                "quality": "ExactStructured",
                "mechanism_quality": "ExactStructured",
            },
        )
    add(
        "ContentEvidenceBinding",
        {
            "content_id": variant,
            "sequence": 2,
            "fact_key": "goal07.s01.public-level-stats",
            "source_record_id": 3,
            "evidence_record_id": 4,
            "quality": "ExactStructured",
            "mechanism_quality": "ExactStructured",
        },
    )

    for table_rows in rows.values():
        table_rows.sort(
            key=lambda row: json.dumps(
                row, ensure_ascii=False, sort_keys=True, default=str
            )
        )
    return rows


def owned_rows_s02() -> dict[str, list[dict[str, Any]]]:
    anchor = json.loads(anchor_path(PARTITION).read_text(encoding="utf-8"))
    manifest = json.loads(PARTITIONS.read_text(encoding="utf-8"))
    assigned = next(item for item in manifest["partitions"] if item["id"] == PARTITION)
    if assigned["enemy_variant_ids"] != [VARIANT_KEY]:
        raise ValueError("S02 frozen enemy assignment changed")

    variant = BASE + 1
    template = BASE + 2
    graphs = [BASE + 10, BASE + 11, BASE + 12]
    ability_names = {
        201: ("felling-order", "Felling Order", "砍伐指令"),
        202: ("lock-on-target", "Lock On Target", "目标锁定"),
        203: ("targeting-order", "Targeting Order", "瞄准指令"),
        204: ("teamwork-order", "Teamwork Order", "协力指令"),
        205: ("disintegration-order", "Disintegration Order", "解体指令"),
        206: ("dismantle", "Dismantle", "解体拆除"),
        207: ("phase-transition-helper", "Phase Transition Helper", "阶段转换辅助"),
        208: ("combat-speed-up", "Combat Speed-Up", "作战加速"),
    }
    abilities = {source: BASE + 100 + source - 200 for source in ability_names}
    selectors = {
        "actor": BASE + 401,
        "primary-target": BASE + 402,
        "opposing-single": BASE + 403,
        "opposing-two-random": BASE + 404,
        "locked-targets": BASE + 405,
        "current-subject": BASE + 406,
        "other-opponents": BASE + 407,
        "owner": BASE + 408,
    }
    effects = {
        "felling-lock": BASE + 501,
        "bleed": BASE + 502,
        "teamwork-order": BASE + 503,
        "combat-speed-up": BASE + 504,
        "targeting-speed-up": BASE + 505,
    }
    modifiers = {
        "combat-speed-up": BASE + 521,
        "targeting-speed-up": BASE + 522,
    }
    modifier_groups = {
        "combat-speed-up": BASE + 531,
        "targeting-speed-up": BASE + 532,
    }
    condition_always = BASE + 551
    condition_unshielded = BASE + 552
    rows: dict[str, list[dict[str, Any]]] = {}
    identities: list[dict[str, Any]] = []
    next_program = BASE + 301
    next_operation = BASE + 1_001
    next_expression = BASE + 1_101

    def add(table: str, row: dict[str, Any]) -> None:
        rows.setdefault(table, []).append(row)

    def identity_s02(
        id_: int,
        stable_key: str,
        kind: str,
        name_en: str,
        name_zh_cn: str,
        summary: str,
        sources: str = "1",
    ) -> dict[str, Any]:
        row = identity(
            id_,
            stable_key,
            kind,
            name_en,
            name_zh_cn,
            summary,
            sources,
        )
        row["summary_zh_cn"] = "Goal 07 S02 来源绑定的完整形态自动机兵·齿狼可执行定义。"
        row["game_version_introduced"] = "1.0"
        return row

    identities.extend(
        [
            identity_s02(
                variant,
                VARIANT_KEY,
                "EnemyVariant",
                "Automaton Direwolf (Complete)",
                "自动机兵·齿狼（完整）",
                "Exact World 2 level-27 three-phase boss variant.",
                "1|4",
            ),
            identity_s02(
                template,
                "enemy.automaton-direwolf-complete.elite",
                "Enemy",
                "Automaton Direwolf (Complete) Template",
                "自动机兵·齿狼（完整）模板",
                "Version 4.4 elite template retained from source monster 1013022.",
            ),
        ]
    )
    for sequence, graph in enumerate(graphs, start=1):
        identities.append(
            identity_s02(
                graph,
                f"ai.goal07.automaton-direwolf-complete.phase-{sequence}",
                "AiGraph",
                f"Automaton Direwolf Phase {sequence} AI",
                f"自动机兵·齿狼第{sequence}阶段AI",
                "Finite source-ordered phase action graph.",
            )
        )
    for source, (key, name_en, name_zh_cn) in ability_names.items():
        identities.append(
            identity_s02(
                abilities[source],
                f"enemy.automaton-direwolf-complete.elite.ability.{key}",
                "Ability",
                name_en,
                name_zh_cn,
                f"Executable transcription of source skill 101302{source - 200:02d}.",
            )
        )
    for name, selector_id in selectors.items():
        identities.append(
            identity_s02(
                selector_id,
                f"selector.goal07.automaton-direwolf-complete.{name}",
                "Selector",
                f"Automaton Direwolf {name} Selector",
                f"自动机兵·齿狼{name}选择器",
                "S02 battle selector.",
            )
        )
    for name, effect in effects.items():
        identities.append(
            identity_s02(
                effect,
                f"effect.goal07.automaton-direwolf-complete.{name}",
                "Effect",
                f"Automaton Direwolf {name} Effect",
                f"自动机兵·齿狼{name}效果",
                "S02 executable enemy effect.",
            )
        )
    for name, modifier in modifiers.items():
        identities.append(
            identity_s02(
                modifier,
                f"modifier.goal07.automaton-direwolf-complete.{name}",
                "Modifier",
                f"Automaton Direwolf {name} Modifier",
                f"自动机兵·齿狼{name}调整器",
                "Exact phase-three speed modifier.",
            )
        )

    add("Selector", selector(selectors["actor"], "Actor", "SameSide"))
    add("Selector", selector(selectors["owner"], "Owner", "SameSide"))
    add(
        "Selector",
        selector(selectors["current-subject"], "CurrentSubject", "OpposingSide"),
    )
    add(
        "Selector",
        selector(selectors["primary-target"], "PrimaryTarget", "OpposingSide"),
    )
    add(
        "Selector",
        selector(selectors["opposing-single"], "Actor", "OpposingSide"),
    )
    add(
        "Selector",
        {
            **selector(
                selectors["opposing-two-random"],
                "Actor",
                "OpposingSide",
                minimum=1,
                maximum=2,
                choice="RngUniform",
            ),
            "rng_purpose_key": "behavior-choice",
        },
    )
    add(
        "Selector",
        selector(
            selectors["locked-targets"],
            "Actor",
            "OpposingSide",
            minimum=1,
            maximum=2,
            choice="All",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["locked-targets"],
            "sequence": 1,
            "predicate": json_cell("HasEffect", effect_id=effects["felling-lock"]),
        },
    )
    add(
        "Selector",
        {
            **selector(
                selectors["other-opponents"],
                "Actor",
                "OpposingSide",
                minimum=0,
                maximum=1,
                empty="NoOp",
                choice="RngUniform",
            ),
            "rng_purpose_key": "damage-target",
        },
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["other-opponents"],
            "sequence": 1,
            "predicate": json_cell(
                "Excludes", excluded_selector_id=selectors["current-subject"]
            ),
        },
    )

    expression_ids: dict[str, int] = {}

    def expr(name: str, kind: str, node: str) -> int:
        nonlocal next_expression
        id_ = next_expression
        next_expression += 1
        expression_ids[name] = id_
        add(
            "ValueExpression",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s02.expression.{name}",
                "result_kind": kind,
                "node": node,
            },
        )
        return id_

    def multiply(name: str, left: int, right: int) -> int:
        return expr(
            name,
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedMultiply",
                left_expression_id=left,
                right_expression_id=right,
                rounding="NearestTiesAway",
            ),
        )

    actor_atk = expr(
        "actor-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    current_shield = expr(
        "current-subject-shield",
        "Scalar",
        json_cell(
            "QueryShield",
            subject_selector_id=selectors["current-subject"],
            observation="Current",
        ),
    )
    zero = expr("zero", "Scalar", json_cell("ScalarLiteral", value_decimal="0"))
    one = expr("one", "Scalar", json_cell("ScalarLiteral", value_decimal="1"))
    ratio_08 = expr(
        "ratio-0-8", "Scalar", json_cell("ScalarLiteral", value_decimal="0.8")
    )
    ratio_005 = expr(
        "ratio-0-05", "Scalar", json_cell("ScalarLiteral", value_decimal="0.05")
    )
    ratio_25 = expr(
        "ratio-2-5", "Scalar", json_cell("ScalarLiteral", value_decimal="2.5")
    )
    ratio_06 = expr(
        "ratio-0-6", "Scalar", json_cell("ScalarLiteral", value_decimal="0.6")
    )
    ratio_02 = expr(
        "ratio-0-2", "Scalar", json_cell("ScalarLiteral", value_decimal="0.2")
    )
    duration_two = expr(
        "duration-two", "Integer", json_cell("IntegerLiteral", value=2)
    )
    targeting_stacks = expr(
        "targeting-speed-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["current-subject"],
            effect_id=effects["targeting-speed-up"],
        ),
    )
    targeting_stacks_scalar = expr(
        "targeting-speed-stacks-scalar",
        "Scalar",
        json_cell(
            "Convert",
            operand_expression_id=targeting_stacks,
            target_kind="Scalar",
            rounding="NearestTiesAway",
        ),
    )
    felling_damage = multiply("felling-hit-damage", actor_atk, ratio_08)
    bleed_damage = multiply("bleed-damage", actor_atk, ratio_005)
    heavy_damage = multiply("heavy-damage", actor_atk, ratio_25)
    targeting_speed_value = multiply(
        "targeting-speed-value", targeting_stacks_scalar, ratio_02
    )
    add(
        "ConditionExpression",
        {
            "id": condition_always,
            "stable_key": "goal07.enemy.s02.condition.always",
            "node": json_cell("Constant", value=True),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": condition_unshielded,
            "stable_key": "goal07.enemy.s02.condition.current-subject-unshielded",
            "node": json_cell(
                "Compare",
                left_expression_id=current_shield,
                comparison="Equal",
                right_expression_id=zero,
            ),
        },
    )

    def operation_s02(
        id_: int,
        name: str,
        payload: str,
        target: int | None = None,
        empty: str = "Fault",
    ) -> dict[str, Any]:
        row = operation(id_, name, payload, target, empty)
        row["stable_key"] = f"goal07.enemy.s02.operation.{name}"
        return row

    def damage_op(name: str, amount: int, target: int) -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        return operation_s02(
            id_,
            name,
            json_cell(
                "Damage",
                amount_expression_id=amount,
                damage_class="Ordinary",
                element="Physical",
                can_crit=True,
            ),
            target,
            "NoOp" if target == selectors["other-opponents"] else "Fault",
        )

    def effect_op(
        name: str,
        effect: int,
        target: int,
        *,
        resistible: bool = False,
        remove: bool = False,
        empty: str = "Fault",
    ) -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        payload = (
            json_cell("RemoveEffect", effect_id=effect)
            if remove
            else json_cell(
                "ApplyEffect",
                effect_id=effect,
                stacks_expression_id=None,
                chance_policy="Resistible" if resistible else "Guaranteed",
                base_chance_expression_id=one if resistible else None,
                rng_purpose_key="effect-application" if resistible else None,
            )
        )
        return operation_s02(id_, name, payload, target, empty)

    programs: dict[str, int] = {}

    def program(name: str, steps: list[dict[str, Any] | str]) -> int:
        nonlocal next_program
        id_ = next_program
        next_program += 1
        programs[name] = id_
        identities.append(
            identity_s02(
                id_,
                f"program.goal07.automaton-direwolf-complete.{name}",
                "Program",
                f"Automaton Direwolf {name} Program",
                f"自动机兵·齿狼{name}程序",
                "Ordered Rule IR program for the S02 enemy.",
            )
        )
        add("Program", {"id": id_, "domain": "Battle"})
        for sequence, step in enumerate(steps, start=1):
            if isinstance(step, dict):
                add("Operation", step)
                encoded = json_cell("Operation", operation_id=step["id"])
            else:
                encoded = step
            add(
                "ProgramStep",
                {"program_id": id_, "sequence": sequence, "step": encoded},
            )
        return id_

    dismantle_followup = program(
        "unshielded-dismantle-followup",
        [
            damage_op(
                "unshielded-dismantle-followup-damage",
                heavy_damage,
                selectors["other-opponents"],
            )
        ],
    )
    conditional_dismantle = program(
        "conditional-unshielded-dismantle",
        [
            json_cell(
                "If",
                condition_id=condition_unshielded,
                then_program_id=dismantle_followup,
                else_program_id=None,
            )
        ],
    )
    felling_steps: list[dict[str, Any] | str] = []
    for hit in range(1, 11):
        felling_steps.append(
            damage_op(
                f"felling-order-hit-{hit:02d}",
                felling_damage,
                selectors["locked-targets"],
            )
        )
        felling_steps.append(
            effect_op(
                f"felling-order-bleed-{hit:02d}",
                effects["bleed"],
                selectors["locked-targets"],
                resistible=True,
            )
        )
    felling_steps.append(
        json_cell(
            "ForEach",
            selector_id=selectors["locked-targets"],
            body_program_id=conditional_dismantle,
            maximum_iterations=2,
        )
    )
    felling_steps.extend(
        [
            effect_op(
                "felling-order-clear-lock",
                effects["felling-lock"],
                selectors["locked-targets"],
                remove=True,
            ),
            effect_op(
                "felling-order-clear-teamwork",
                effects["teamwork-order"],
                selectors["actor"],
                remove=True,
                empty="NoOp",
            ),
        ]
    )
    ability_programs = {
        abilities[201]: program("felling-order", felling_steps),
        abilities[202]: program(
            "lock-on-target",
            [
                effect_op(
                    "lock-on-target-mark",
                    effects["felling-lock"],
                    selectors["primary-target"],
                )
            ],
        ),
        abilities[203]: program(
            "targeting-order",
            [
                effect_op(
                    "targeting-order-mark-two",
                    effects["felling-lock"],
                    selectors["opposing-two-random"],
                ),
                effect_op(
                    "targeting-order-speed-up",
                    effects["targeting-speed-up"],
                    selectors["actor"],
                ),
            ],
        ),
        abilities[204]: program(
            "teamwork-order",
            [
                effect_op(
                    "teamwork-order-mark-two",
                    effects["felling-lock"],
                    selectors["opposing-two-random"],
                ),
                effect_op(
                    "teamwork-order-coordinate-grizzly",
                    effects["teamwork-order"],
                    selectors["actor"],
                ),
            ],
        ),
        abilities[205]: program(
            "disintegration-order",
            [
                damage_op(
                    "disintegration-order-damage",
                    heavy_damage,
                    selectors["primary-target"],
                )
            ],
        ),
        abilities[206]: program(
            "dismantle",
            [
                damage_op(
                    "dismantle-damage",
                    heavy_damage,
                    selectors["primary-target"],
                )
            ],
        ),
    }
    phase_three_entry = program(
        "phase-three-combat-speed-up",
        [
            effect_op(
                "phase-three-combat-speed-up",
                effects["combat-speed-up"],
                selectors["actor"],
            )
        ],
    )

    target_patterns = {
        201: "SingleTarget",
        202: "SingleTarget",
        203: "Aoe",
        204: "Aoe",
        205: "SingleTarget",
        206: "SingleTarget",
        207: "None",
        208: "None",
    }
    for source, ability in abilities.items():
        add(
            "Ability",
            {
                "id": ability,
                "kind": "Passive" if source in {207, 208} else "Skill",
                "target_pattern": target_patterns[source],
                "retarget_policy": "CancelRemaining",
                "level_cap": 1,
                "cooldown_actions": 1,
                "semantic_tags_mask": 5 if source in {201, 205, 206} else 4,
            },
        )
        add(
            "AbilityPhase",
            {
                "ability_id": ability,
                "sequence": 1,
                "kind": "Resolved",
                "program_identity_id": ability_programs.get(ability),
            },
        )
        add(
            "EnemyAbility",
            {
                "id": ability,
                "telegraph": "LockOn" if source in {202, 203, 204} else "None",
                "cooldown_actions": 1,
                "initial_cooldown_actions": 0,
                "charge_actions": 0,
                "ai_tag": ability_names[source][0],
            },
        )

    effect_rows = {
        "felling-lock": {
            "category": "Mark",
            "dispel": "NonDispellable",
            "stack_limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": None,
            "dot": None,
        },
        "bleed": {
            "category": "Dot",
            "dispel": "DispellableDebuff",
            "stack_limit": 1,
            "duration": duration_two,
            "clock": "TargetTurnEnd",
            "tick": "TurnStart",
            "policy": "Refresh",
            "magnitude": bleed_damage,
            "dot": "Physical",
        },
        "teamwork-order": {
            "category": "NeutralState",
            "dispel": "NonDispellable",
            "stack_limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": None,
            "dot": None,
        },
        "combat-speed-up": {
            "category": "Buff",
            "dispel": "NonDispellable",
            "stack_limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": ratio_06,
            "dot": None,
        },
        "targeting-speed-up": {
            "category": "Buff",
            "dispel": "NonDispellable",
            "stack_limit": 2,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "RefreshAndAddStacks",
            "magnitude": targeting_speed_value,
            "dot": None,
        },
    }
    for name, effect in effects.items():
        definition = effect_rows[name]
        add(
            "Effect",
            {
                "id": effect,
                "category": definition["category"],
                "dispel_category": definition["dispel"],
                "stack_limit": definition["stack_limit"],
                "duration_expression_id": definition["duration"],
                "duration_clock": definition["clock"],
                "tick_phase": definition["tick"],
                "stack_policy": definition["policy"],
                "magnitude_comparator_expression_id": definition["magnitude"],
                "dot_element": definition["dot"],
                "snapshot_policy": "OnApplication",
                "teardown_policy": "RemoveWithOwner",
                "application_priority": 0,
            },
        )
    for effect_name, tag in [
        ("felling-lock", "direwolf-felling-target"),
        ("teamwork-order", "coordinates-automaton-grizzly-purge-order"),
    ]:
        add(
            "EffectTag",
            {"effect_id": effects[effect_name], "sequence": 1, "tag": tag},
        )
    for name, group in modifier_groups.items():
        add(
            "ModifierStackingGroup",
            {
                "id": group,
                "stable_key": f"goal07.enemy.s02.{name}",
                "aggregation": "Sum",
            },
        )
    for name, modifier in modifiers.items():
        add(
            "ModifierDefinition",
            {
                "id": modifier,
                "source_effect_id": effects[name],
                "owner_selector_id": selectors["owner"],
                "subject_selector_id": selectors["current-subject"],
                "stat": "Spd",
                "formula_stage": "PercentOfBase",
                "formula_purpose": "Stat",
                "value_expression_id": (
                    ratio_06 if name == "combat-speed-up" else targeting_speed_value
                ),
                "value_domain": "Ratio",
                "stacking_group_id": modifier_groups[name],
                "priority": 0,
                "cap_formula_stage": "PercentOfBase",
                "snapshot_policy": (
                    "OnApplication" if name == "combat-speed-up" else "Dynamic"
                ),
                "duration_scope": "Turn",
            },
        )
        add(
            "EffectModifierBinding",
            {
                "effect_id": effects[name],
                "sequence": 1,
                "modifier_id": modifier,
            },
        )

    phase_sequences = [
        [202, 201, 205],
        [204, 201, 205],
        [203, 201, 205],
    ]
    target_selectors = {
        201: selectors["locked-targets"],
        202: selectors["opposing-single"],
        203: selectors["opposing-two-random"],
        204: selectors["opposing-two-random"],
        205: selectors["opposing-single"],
    }
    next_state = BASE + 701
    next_candidate = BASE + 801
    next_transition = BASE + 901
    for phase_index, sequence_sources in enumerate(phase_sequences):
        state_ids = list(range(next_state, next_state + len(sequence_sources)))
        next_state += len(sequence_sources)
        add(
            "AiGraph",
            {
                "id": graphs[phase_index],
                "initial_state_id": state_ids[0],
                "automatic_transition_budget": 8,
            },
        )
        for offset, (state_id, source) in enumerate(
            zip(state_ids, sequence_sources)
        ):
            add(
                "AiState",
                {
                    "id": state_id,
                    "stable_key": (
                        f"goal07.enemy.s02.ai.phase-{phase_index + 1}."
                        f"state-{offset + 1}"
                    ),
                    "graph_id": graphs[phase_index],
                    "mandatory_fallback_ability_id": abilities[205],
                    "turn_counter_reset": offset == 0,
                },
            )
            add(
                "AiCandidate",
                {
                    "id": next_candidate,
                    "stable_key": (
                        f"goal07.enemy.s02.ai.phase-{phase_index + 1}."
                        f"state-{offset + 1}.main"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "ability_id": abilities[source],
                    "condition_id": condition_always,
                    "target_selector_id": target_selectors[source],
                    "priority": 10,
                    "selection": "FirstLegal",
                    "no_target_fallback": "UseFallbackAbility",
                    "fallback_ability_id": abilities[205],
                },
            )
            next_candidate += 1
            add(
                "AiTransition",
                {
                    "id": next_transition,
                    "stable_key": (
                        f"goal07.enemy.s02.ai.phase-{phase_index + 1}."
                        f"transition-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "target_state_id": state_ids[(offset + 1) % len(state_ids)],
                    "condition_id": condition_always,
                    "priority": 0,
                    "timing": "AfterAction",
                },
            )
            next_transition += 1

    add(
        "EnemyTemplate",
        {
            "id": template,
            "rank": "Elite",
            "base_aggro_decimal": "100",
            "default_ai_graph_id": graphs[0],
        },
    )
    add(
        "EnemyVariant",
        {
            "id": variant,
            "template_id": template,
            "ai_graph_id": graphs[0],
            "mechanically_distinct_key": VARIANT_KEY,
        },
    )
    for level in anchor["levels"]:
        add(
            "EnemyStat",
            {
                "variant_id": variant,
                "level": level["authored_level"],
                "difficulty_key": "standard-universe-v1",
                "hp_decimal": level["base_hp"],
                "atk_decimal": level["base_atk"],
                "def_decimal": level["base_def"],
                "spd_decimal": level["base_spd"],
                "effect_hit_rate_decimal": level["effect_hit_rate"],
                "effect_resistance_decimal": level["effect_resistance"],
                "crit_damage_decimal": "0.2",
            },
        )
    for sequence, weakness in enumerate(
        ["Ice", "Imaginary", "Lightning"], start=1
    ):
        add(
            "EnemyWeakness",
            {"variant_id": variant, "sequence": sequence, "element": weakness},
        )
    for element in ["Fire", "Physical", "Quantum", "Wind"]:
        add(
            "EnemyResistance",
            {"variant_id": variant, "element": element, "value_decimal": "0.2"},
        )
    for category in ["STAT_Confine", "STAT_CTRL_Frozen", "STAT_Entangle"]:
        add(
            "EnemyDebuffResistance",
            {
                "variant_id": variant,
                "category_key": category,
                "value_decimal": "0.5",
            },
        )
    add(
        "EnemyToughnessLayer",
        {
            "variant_id": variant,
            "sequence": 1,
            "layer_key": "ordinary",
            "kind": "Ordinary",
            "maximum_decimal": "300",
            "recovery_ratio_decimal": "1",
            "active_at_start": True,
        },
    )
    for sequence, source in enumerate(ability_names, start=1):
        add(
            "EnemyVariantAbility",
            {
                "variant_id": variant,
                "sequence": sequence,
                "ability_id": abilities[source],
            },
        )
    for sequence, graph in enumerate(graphs, start=1):
        add(
            "EnemyPhase",
            {
                "id": BASE + 600 + sequence,
                "stable_key": f"goal07.enemy.s02.phase-{sequence}",
                "variant_id": variant,
                "sequence": sequence,
                "entry_condition_id": condition_always,
                "exit_condition_id": condition_always,
                "replacement_priority": sequence,
                "ai_graph_id": graph,
                "targetable": True,
                "transition_model": "TransformSameUnit",
                "entry_program_id": phase_three_entry if sequence == 3 else None,
                "hp_carry": "Reset",
                "action_gauge_carry": "Reset",
                "effect_carry": "Clear",
                "toughness_carry": "Reset",
                "summon_carry": "Clear",
            },
        )

    anchor_digest = sha256_bytes(anchor_path(PARTITION).read_bytes())
    add(
        "SourceRecord",
        {
            "id": SOURCE_RECORD_ID,
            "stable_key": "source.hsr-wiki.automaton-direwolf-complete.2026-07-29",
            "category": "CommunityMaintained",
            "publisher": "Honkai: Star Rail Wiki",
            "url": anchor["source"]["url"],
            "accessed_on": anchor["source"]["accessed_on"],
            "applicable_game_version": anchor["source"]["game_version"],
            "confidence": "SecondaryVersionSensitiveCrossCheck",
            "evidence_sha256": anchor_digest,
            "usage_note": (
                "Exact public World 2 level-27 HP, ATK, DEF, SPD, EHR and "
                "Effect RES transcribed into committed Goal 07 evidence."
            ),
        },
    )
    add(
        "EvidenceRecord",
        {
            "id": EVIDENCE_RECORD_ID,
            "stable_key": "evidence.goal07.enemy.s02.numeric-anchors",
            "kind": "SourcePayload",
            "source_record_id": SOURCE_RECORD_ID,
            "sha256": anchor_digest,
            "note": "Committed exact public per-level numeric anchors for Goal 07 S02.",
        },
    )
    for item in identities:
        add("ContentIdentity", item)
        add(
            "ContentEvidenceBinding",
            {
                "content_id": item["id"],
                "sequence": 1,
                "fact_key": f"goal07.s02.executable:{item['stable_key']}",
                "source_record_id": 1,
                "evidence_record_id": 3,
                "quality": "ExactStructured",
                "mechanism_quality": "ExactStructured",
            },
        )
    add(
        "ContentEvidenceBinding",
        {
            "content_id": variant,
            "sequence": 2,
            "fact_key": "goal07.s02.public-level-stats",
            "source_record_id": SOURCE_RECORD_ID,
            "evidence_record_id": EVIDENCE_RECORD_ID,
            "quality": "ExactStructured",
            "mechanism_quality": "ExactStructured",
        },
    )
    for table_rows in rows.values():
        table_rows.sort(
            key=lambda row: json.dumps(
                row, ensure_ascii=False, sort_keys=True, default=str
            )
        )
    return rows


def owned_rows_s03() -> dict[str, list[dict[str, Any]]]:
    anchor = json.loads(anchor_path(PARTITION).read_text(encoding="utf-8"))
    manifest = json.loads(PARTITIONS.read_text(encoding="utf-8"))
    assigned = next(item for item in manifest["partitions"] if item["id"] == PARTITION)
    if assigned["enemy_variant_ids"] != [VARIANT_KEY]:
        raise ValueError("S03 frozen enemy assignment changed")

    variant = BASE + 1
    template = BASE + 2
    graphs = [BASE + 10, BASE + 11]
    linked_spider = BASE + 201
    ability_names = {
        201: ("purge-order", "Purge Order", "清除指令"),
        202: ("destruction-order", "Destruction Order", "毁灭指令"),
        203: ("detonation-order", "Detonation Order", "引爆指令"),
        204: ("enrage-order", "Enrage Order", "激怒指令"),
        205: ("overcombust-order", "Overcombust Order", "过载指令"),
        206: ("obliteration-order", "Obliteration Order", "歼灭指令"),
    }
    abilities = {source: BASE + 100 + source - 200 for source in ability_names}
    enrage_phase_two = BASE + 109
    spider_self_explode = BASE + 110
    selectors = {
        "actor": BASE + 401,
        "owner": BASE + 402,
        "primary-target": BASE + 403,
        "opposing-single": BASE + 404,
        "opposing-all": BASE + 405,
        "current-subject": BASE + 406,
        "coordinated-allies": BASE + 407,
    }
    effects = {
        "taunt": BASE + 501,
        "overcombust": BASE + 502,
        "obliteration": BASE + 503,
    }
    modifier = BASE + 521
    modifier_group = BASE + 531
    condition_always = BASE + 551
    condition_coordinated = BASE + 552
    rows: dict[str, list[dict[str, Any]]] = {}
    identities: list[dict[str, Any]] = []
    next_program = BASE + 301
    next_operation = BASE + 1_001
    next_expression = BASE + 1_101

    def add(table: str, row: dict[str, Any]) -> None:
        rows.setdefault(table, []).append(row)

    def identity_s03(
        id_: int,
        stable_key: str,
        kind: str,
        name_en: str,
        name_zh_cn: str,
        summary: str,
        sources: str = "1",
    ) -> dict[str, Any]:
        row = identity(
            id_,
            stable_key,
            kind,
            name_en,
            name_zh_cn,
            summary,
            sources,
        )
        row["summary_zh_cn"] = "Goal 07 S03 来源绑定的完整形态自动机兵·灰熊可执行定义。"
        row["game_version_introduced"] = "1.0"
        return row

    identities.extend(
        [
            identity_s03(
                variant,
                VARIANT_KEY,
                "EnemyVariant",
                "Automaton Grizzly (Complete)",
                "自动机兵·灰熊（完整）",
                "Exact World 2 level-27 two-phase boss variant.",
                "1|5",
            ),
            identity_s03(
                template,
                "enemy.automaton-grizzly-complete.elite",
                "Enemy",
                "Automaton Grizzly (Complete) Template",
                "自动机兵·灰熊（完整）模板",
                "Version 4.4 elite template retained from source monster 1013012.",
            ),
            identity_s03(
                linked_spider,
                "unit.goal07.automaton-grizzly-complete.automaton-spider",
                "CharacterForm",
                "Automaton Grizzly Summoned Spider",
                "自动机兵·灰熊召唤的蜘蛛",
                "Executable linked Automaton Spider summoned by Detonation Order.",
            ),
        ]
    )
    for sequence, graph in enumerate(graphs, start=1):
        identities.append(
            identity_s03(
                graph,
                f"ai.goal07.automaton-grizzly-complete.phase-{sequence}",
                "AiGraph",
                f"Automaton Grizzly Phase {sequence} AI",
                f"自动机兵·灰熊第{sequence}阶段AI",
                "Finite source-ordered phase action graph.",
            )
        )
    for source, (key, name_en, name_zh_cn) in ability_names.items():
        identities.append(
            identity_s03(
                abilities[source],
                f"enemy.automaton-grizzly-complete.elite.ability.{key}",
                "Ability",
                name_en,
                name_zh_cn,
                f"Executable transcription of source skill 101301{source - 200:02d}.",
            )
        )
    identities.append(
        identity_s03(
            enrage_phase_two,
            "enemy.automaton-grizzly-complete.elite.ability.enrage-order-phase-2",
            "Ability",
            "Enrage Order (Phase 2)",
            "激怒指令（第二阶段）",
            "Phase-two guaranteed application of source skill 101301204.",
        )
    )
    identities.append(
        identity_s03(
            spider_self_explode,
            "unit.goal07.automaton-grizzly-complete.automaton-spider.self-explode",
            "Ability",
            "Automaton Spider Self-Explosion",
            "自动机兵·蜘蛛自爆",
            "Linked Summon action preserving source skill 101202103.",
        )
    )
    for name, selector_id in selectors.items():
        identities.append(
            identity_s03(
                selector_id,
                f"selector.goal07.automaton-grizzly-complete.{name}",
                "Selector",
                f"Automaton Grizzly {name} Selector",
                f"自动机兵·灰熊{name}选择器",
                "S03 battle selector.",
            )
        )
    for name, effect in effects.items():
        identities.append(
            identity_s03(
                effect,
                f"effect.goal07.automaton-grizzly-complete.{name}",
                "Effect",
                f"Automaton Grizzly {name} Effect",
                f"自动机兵·灰熊{name}效果",
                "S03 executable enemy effect.",
            )
        )
    identities.append(
        identity_s03(
            modifier,
            "modifier.goal07.automaton-grizzly-complete.obliteration",
            "Modifier",
            "Automaton Grizzly Obliteration Modifier",
            "自动机兵·灰熊歼灭调整器",
            "Stack-scaled ordinary damage boost applied by Overcombust Order.",
        )
    )

    add("Selector", selector(selectors["actor"], "Actor", "SameSide"))
    add("Selector", selector(selectors["owner"], "Owner", "SameSide"))
    add(
        "Selector",
        selector(selectors["primary-target"], "PrimaryTarget", "OpposingSide"),
    )
    add(
        "Selector",
        selector(selectors["opposing-single"], "Actor", "OpposingSide"),
    )
    add(
        "Selector",
        selector(
            selectors["opposing-all"],
            "Actor",
            "OpposingSide",
            minimum=1,
            maximum=8,
            choice="All",
        ),
    )
    add(
        "Selector",
        selector(selectors["current-subject"], "CurrentSubject", "OpposingSide"),
    )
    add(
        "Selector",
        selector(
            selectors["coordinated-allies"],
            "Actor",
            "SameSide",
            minimum=0,
            maximum=8,
            empty="NoOp",
            choice="All",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["coordinated-allies"],
            "sequence": 1,
            "predicate": json_cell("HasEffect", effect_id=990_503),
        },
    )

    def expr(name: str, kind: str, node: str) -> int:
        nonlocal next_expression
        id_ = next_expression
        next_expression += 1
        add(
            "ValueExpression",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s03.expression.{name}",
                "result_kind": kind,
                "node": node,
            },
        )
        return id_

    def multiply(name: str, left: int, right: int) -> int:
        return expr(
            name,
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedMultiply",
                left_expression_id=left,
                right_expression_id=right,
                rounding="NearestTiesAway",
            ),
        )

    actor_atk = expr(
        "actor-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    ratio_four = expr(
        "ratio-4", "Scalar", json_cell("ScalarLiteral", value_decimal="4")
    )
    ratio_three = expr(
        "ratio-3", "Scalar", json_cell("ScalarLiteral", value_decimal="3")
    )
    ratio_half = expr(
        "ratio-0-5", "Scalar", json_cell("ScalarLiteral", value_decimal="0.5")
    )
    ratio_one = expr(
        "ratio-1", "Scalar", json_cell("ScalarLiteral", value_decimal="1")
    )
    ratio_five = expr(
        "ratio-5", "Scalar", json_cell("ScalarLiteral", value_decimal="5")
    )
    duration_two = expr(
        "duration-two", "Integer", json_cell("IntegerLiteral", value=2)
    )
    coordinated_count = expr(
        "coordinated-allies-count",
        "Integer",
        json_cell("SelectorCount", selector_id=selectors["coordinated-allies"]),
    )
    integer_zero = expr(
        "integer-zero", "Integer", json_cell("IntegerLiteral", value=0)
    )
    obliteration_stacks = expr(
        "obliteration-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["current-subject"],
            effect_id=effects["obliteration"],
        ),
    )
    obliteration_stacks_scalar = expr(
        "obliteration-stacks-scalar",
        "Scalar",
        json_cell(
            "Convert",
            operand_expression_id=obliteration_stacks,
            target_kind="Scalar",
            rounding="NearestTiesAway",
        ),
    )
    purge_damage = multiply("purge-damage", actor_atk, ratio_four)
    destruction_damage = multiply("destruction-damage", actor_atk, ratio_three)
    spider_damage = multiply("spider-self-explode-damage", actor_atk, ratio_five)
    obliteration_value = multiply(
        "obliteration-damage-boost", obliteration_stacks_scalar, ratio_half
    )
    add(
        "ConditionExpression",
        {
            "id": condition_always,
            "stable_key": "goal07.enemy.s03.condition.always",
            "node": json_cell("Constant", value=True),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": condition_coordinated,
            "stable_key": "goal07.enemy.s03.condition.direwolf-coordinated",
            "node": json_cell(
                "Compare",
                left_expression_id=coordinated_count,
                comparison="Greater",
                right_expression_id=integer_zero,
            ),
        },
    )

    def operation_s03(
        id_: int,
        name: str,
        payload: str,
        target: int | None = None,
        empty: str = "Fault",
    ) -> dict[str, Any]:
        row = operation(id_, name, payload, target, empty)
        row["stable_key"] = f"goal07.enemy.s03.operation.{name}"
        return row

    def damage_op(
        name: str,
        amount: int,
        target: int,
        element: str = "Physical",
    ) -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        return operation_s03(
            id_,
            name,
            json_cell(
                "Damage",
                amount_expression_id=amount,
                damage_class="Ordinary",
                element=element,
                can_crit=True,
            ),
            target,
        )

    def effect_op(
        name: str,
        effect: int,
        target: int,
        *,
        chance: int | None = None,
        remove: bool = False,
        empty: str = "Fault",
    ) -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        payload = (
            json_cell("RemoveEffect", effect_id=effect)
            if remove
            else json_cell(
                "ApplyEffect",
                effect_id=effect,
                stacks_expression_id=None,
                chance_policy="Resistible" if chance is not None else "Guaranteed",
                base_chance_expression_id=chance,
                rng_purpose_key="effect-application" if chance is not None else None,
            )
        )
        return operation_s03(id_, name, payload, target, empty)

    def summon_op() -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        return operation_s03(
            id_,
            "detonation-order-summon-spider",
            json_cell(
                "Summon",
                unit_definition_identity_id=linked_spider,
                owner_selector_id=selectors["actor"],
            ),
        )

    def despawn_op() -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        return operation_s03(
            id_,
            "automaton-spider-despawn-after-self-explode",
            json_cell("Despawn"),
            selectors["actor"],
        )

    programs: dict[str, int] = {}

    def program(name: str, steps: list[dict[str, Any]]) -> int:
        nonlocal next_program
        id_ = next_program
        next_program += 1
        programs[name] = id_
        identities.append(
            identity_s03(
                id_,
                f"program.goal07.automaton-grizzly-complete.{name}",
                "Program",
                f"Automaton Grizzly {name} Program",
                f"自动机兵·灰熊{name}程序",
                "Ordered Rule IR program for the S03 enemy.",
            )
        )
        add("Program", {"id": id_, "domain": "Battle"})
        for sequence, step in enumerate(steps, start=1):
            add("Operation", step)
            add(
                "ProgramStep",
                {
                    "program_id": id_,
                    "sequence": sequence,
                    "step": json_cell("Operation", operation_id=step["id"]),
                },
            )
        return id_

    ability_programs = {
        abilities[201]: program(
            "purge-order",
            [
                damage_op("purge-order-damage", purge_damage, selectors["opposing-all"]),
                effect_op(
                    "purge-order-clear-charge",
                    effects["overcombust"],
                    selectors["actor"],
                    remove=True,
                    empty="NoOp",
                ),
                effect_op(
                    "purge-order-clear-direwolf-coordination",
                    990_503,
                    selectors["coordinated-allies"],
                    remove=True,
                    empty="NoOp",
                ),
            ],
        ),
        abilities[202]: program(
            "destruction-order",
            [
                damage_op(
                    "destruction-order-damage",
                    destruction_damage,
                    selectors["primary-target"],
                )
            ],
        ),
        abilities[203]: program("detonation-order", [summon_op()]),
        abilities[204]: program(
            "enrage-order-phase-1",
            [
                effect_op(
                    "enrage-order-phase-1-taunt",
                    effects["taunt"],
                    selectors["opposing-all"],
                    chance=ratio_half,
                )
            ],
        ),
        enrage_phase_two: program(
            "enrage-order-phase-2",
            [
                effect_op(
                    "enrage-order-phase-2-taunt",
                    effects["taunt"],
                    selectors["opposing-all"],
                    chance=ratio_one,
                )
            ],
        ),
        abilities[205]: program(
            "overcombust-order",
            [
                effect_op(
                    "overcombust-order-charge",
                    effects["overcombust"],
                    selectors["actor"],
                ),
                effect_op(
                    "overcombust-order-obliteration-stack",
                    effects["obliteration"],
                    selectors["actor"],
                ),
            ],
        ),
    }
    spider_program = program(
        "automaton-spider-self-explode",
        [
            damage_op(
                "automaton-spider-self-explode-damage",
                spider_damage,
                selectors["primary-target"],
                "Fire",
            ),
            despawn_op(),
        ],
    )
    target_patterns = {
        201: "Aoe",
        202: "SingleTarget",
        203: "None",
        204: "Aoe",
        205: "None",
        206: "None",
    }
    for source, ability in abilities.items():
        add(
            "Ability",
            {
                "id": ability,
                "kind": "Passive" if source == 206 else "Skill",
                "target_pattern": target_patterns[source],
                "retarget_policy": "CancelRemaining",
                "level_cap": 1,
                "cooldown_actions": 1,
                "semantic_tags_mask": 5 if source in {201, 202} else 4,
            },
        )
        add(
            "AbilityPhase",
            {
                "ability_id": ability,
                "sequence": 1,
                "kind": "Resolved",
                "program_identity_id": ability_programs.get(ability),
            },
        )
        add(
            "EnemyAbility",
            {
                "id": ability,
                "telegraph": "Charge" if source == 205 else "None",
                "cooldown_actions": 1,
                "initial_cooldown_actions": 0,
                "charge_actions": 1 if source == 205 else 0,
                "ai_tag": ability_names[source][0],
            },
        )
    add(
        "Ability",
        {
            "id": enrage_phase_two,
            "kind": "Skill",
            "target_pattern": "Aoe",
            "retarget_policy": "CancelRemaining",
            "level_cap": 1,
            "cooldown_actions": 1,
            "semantic_tags_mask": 4,
        },
    )
    add(
        "AbilityPhase",
        {
            "ability_id": enrage_phase_two,
            "sequence": 1,
            "kind": "Resolved",
            "program_identity_id": ability_programs[enrage_phase_two],
        },
    )
    add(
        "EnemyAbility",
        {
            "id": enrage_phase_two,
            "telegraph": "None",
            "cooldown_actions": 1,
            "initial_cooldown_actions": 0,
            "charge_actions": 0,
            "ai_tag": "enrage-order-phase-2",
        },
    )
    add(
        "Ability",
        {
            "id": spider_self_explode,
            "kind": "Summon",
            "target_pattern": "SingleTarget",
            "retarget_policy": "CancelRemaining",
            "level_cap": 1,
            "cooldown_actions": 1,
            "semantic_tags_mask": 5,
        },
    )
    add(
        "AbilityPhase",
        {
            "ability_id": spider_self_explode,
            "sequence": 1,
            "kind": "Resolved",
            "program_identity_id": spider_program,
        },
    )

    for name, definition in {
        "taunt": {
            "category": "Control",
            "dispel": "CleanseableControl",
            "stack_limit": 1,
            "duration": duration_two,
            "clock": "TargetTurnEnd",
            "policy": "Refresh",
            "magnitude": None,
        },
        "overcombust": {
            "category": "NeutralState",
            "dispel": "NonDispellable",
            "stack_limit": 1,
            "duration": None,
            "clock": "Permanent",
            "policy": "Replace",
            "magnitude": None,
        },
        "obliteration": {
            "category": "Buff",
            "dispel": "NonDispellable",
            "stack_limit": 100,
            "duration": None,
            "clock": "Permanent",
            "policy": "RefreshAndAddStacks",
            "magnitude": obliteration_value,
        },
    }.items():
        add(
            "Effect",
            {
                "id": effects[name],
                "category": definition["category"],
                "dispel_category": definition["dispel"],
                "stack_limit": definition["stack_limit"],
                "duration_expression_id": definition["duration"],
                "duration_clock": definition["clock"],
                "tick_phase": "None",
                "stack_policy": definition["policy"],
                "magnitude_comparator_expression_id": definition["magnitude"],
                "snapshot_policy": "OnApplication",
                "teardown_policy": "RemoveWithOwner",
                "application_priority": 0,
            },
        )
    for effect_name, tag in [
        ("taunt", "forced-basic-attack-applier"),
        ("overcombust", "charging-next-action-purge-order"),
    ]:
        add(
            "EffectTag",
            {"effect_id": effects[effect_name], "sequence": 1, "tag": tag},
        )
    add(
        "ModifierStackingGroup",
        {
            "id": modifier_group,
            "stable_key": "goal07.enemy.s03.obliteration",
            "aggregation": "Sum",
        },
    )
    add(
        "ModifierDefinition",
        {
            "id": modifier,
            "source_effect_id": effects["obliteration"],
            "owner_selector_id": selectors["owner"],
            "subject_selector_id": selectors["current-subject"],
            "stat": "Atk",
            "formula_stage": "DamageBoost",
            "formula_purpose": "OrdinaryDamage",
            "value_expression_id": obliteration_value,
            "value_domain": "Ratio",
            "stacking_group_id": modifier_group,
            "priority": 0,
            "cap_formula_stage": "DamageBoost",
            "snapshot_policy": "Dynamic",
            "duration_scope": "Turn",
        },
    )
    add(
        "EffectModifierBinding",
        {
            "effect_id": effects["obliteration"],
            "sequence": 1,
            "modifier_id": modifier,
        },
    )

    phase_sequences = [
        [abilities[202], abilities[203], abilities[204], abilities[205], abilities[201]],
        [abilities[202], abilities[203], enrage_phase_two, abilities[205], abilities[201]],
    ]
    target_selectors = {
        abilities[201]: selectors["opposing-all"],
        abilities[202]: selectors["opposing-single"],
        abilities[203]: selectors["actor"],
        abilities[204]: selectors["opposing-all"],
        enrage_phase_two: selectors["opposing-all"],
        abilities[205]: selectors["actor"],
    }
    next_state = BASE + 701
    next_candidate = BASE + 801
    next_transition = BASE + 901
    for phase_index, sequence_abilities in enumerate(phase_sequences):
        state_ids = list(range(next_state, next_state + len(sequence_abilities)))
        next_state += len(sequence_abilities)
        add(
            "AiGraph",
            {
                "id": graphs[phase_index],
                "initial_state_id": state_ids[0],
                "automatic_transition_budget": 8,
            },
        )
        for offset, (state_id, ability) in enumerate(zip(state_ids, sequence_abilities)):
            add(
                "AiState",
                {
                    "id": state_id,
                    "stable_key": (
                        f"goal07.enemy.s03.ai.phase-{phase_index + 1}."
                        f"state-{offset + 1}"
                    ),
                    "graph_id": graphs[phase_index],
                    "mandatory_fallback_ability_id": abilities[202],
                    "turn_counter_reset": offset == 0,
                },
            )
            if phase_index == 1 and offset == 0:
                add(
                    "AiCandidate",
                    {
                        "id": next_candidate,
                        "stable_key": "goal07.enemy.s03.ai.phase-2.coordinated-purge",
                        "state_id": state_id,
                        "sequence": 1,
                        "ability_id": abilities[201],
                        "condition_id": condition_coordinated,
                        "target_selector_id": selectors["opposing-all"],
                        "priority": 0,
                        "selection": "FirstLegal",
                        "no_target_fallback": "UseFallbackAbility",
                        "fallback_ability_id": abilities[202],
                    },
                )
                next_candidate += 1
                candidate_sequence = 2
            else:
                candidate_sequence = 1
            add(
                "AiCandidate",
                {
                    "id": next_candidate,
                    "stable_key": (
                        f"goal07.enemy.s03.ai.phase-{phase_index + 1}."
                        f"state-{offset + 1}.main"
                    ),
                    "state_id": state_id,
                    "sequence": candidate_sequence,
                    "ability_id": ability,
                    "condition_id": condition_always,
                    "target_selector_id": target_selectors[ability],
                    "priority": 10,
                    "selection": "FirstLegal",
                    "no_target_fallback": "UseFallbackAbility",
                    "fallback_ability_id": abilities[202],
                },
            )
            next_candidate += 1
            add(
                "AiTransition",
                {
                    "id": next_transition,
                    "stable_key": (
                        f"goal07.enemy.s03.ai.phase-{phase_index + 1}."
                        f"transition-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "target_state_id": state_ids[(offset + 1) % len(state_ids)],
                    "condition_id": condition_always,
                    "priority": 0,
                    "timing": "AfterAction",
                },
            )
            next_transition += 1

    add(
        "LinkedUnitDefinition",
        {
            "id": linked_spider,
            "source_definition_identity_id": 10_003,
            "kind": "Summon",
            "presence": "Linked",
            "ability_ids": str(spider_self_explode),
            "action_ability_id": spider_self_explode,
            "formation_index": 8,
            "initial_gauge_decimal": "10000",
            "hp_owner_ratio_decimal": "0.15",
            "hp_flat_decimal": "0",
            "atk_owner_ratio_decimal": "1",
            "atk_flat_decimal": "0",
            "def_owner_ratio_decimal": "1",
            "def_flat_decimal": "0",
            "spd_owner_ratio_decimal": "0",
            "spd_flat_decimal": "83",
            "owner_defeat_policy": "Depart",
            "owner_departure_policy": "Depart",
            "wave_policy": "Depart",
            "combatant_digest_sha256": sha256_text("goal07-s03-linked-spider-v1"),
        },
    )
    add(
        "EnemyTemplate",
        {
            "id": template,
            "rank": "Elite",
            "base_aggro_decimal": "100",
            "default_ai_graph_id": graphs[0],
        },
    )
    add(
        "EnemyVariant",
        {
            "id": variant,
            "template_id": template,
            "ai_graph_id": graphs[0],
            "mechanically_distinct_key": VARIANT_KEY,
        },
    )
    for level in anchor["levels"]:
        add(
            "EnemyStat",
            {
                "variant_id": variant,
                "level": level["authored_level"],
                "difficulty_key": "standard-universe-v1",
                "hp_decimal": level["base_hp"],
                "atk_decimal": level["base_atk"],
                "def_decimal": level["base_def"],
                "spd_decimal": level["base_spd"],
                "effect_hit_rate_decimal": level["effect_hit_rate"],
                "effect_resistance_decimal": level["effect_resistance"],
                "crit_damage_decimal": "0.2",
            },
        )
    for sequence, weakness in enumerate(["Fire", "Ice", "Lightning"], start=1):
        add(
            "EnemyWeakness",
            {"variant_id": variant, "sequence": sequence, "element": weakness},
        )
    for element in ["Imaginary", "Physical", "Quantum", "Wind"]:
        add(
            "EnemyResistance",
            {"variant_id": variant, "element": element, "value_decimal": "0.2"},
        )
    for category in ["STAT_Confine", "STAT_CTRL_Frozen", "STAT_Entangle"]:
        add(
            "EnemyDebuffResistance",
            {
                "variant_id": variant,
                "category_key": category,
                "value_decimal": "0.5",
            },
        )
    add(
        "EnemyToughnessLayer",
        {
            "variant_id": variant,
            "sequence": 1,
            "layer_key": "ordinary",
            "kind": "Ordinary",
            "maximum_decimal": "480",
            "recovery_ratio_decimal": "1",
            "active_at_start": True,
        },
    )
    variant_abilities = list(abilities.values()) + [enrage_phase_two]
    for sequence, ability in enumerate(variant_abilities, start=1):
        add(
            "EnemyVariantAbility",
            {"variant_id": variant, "sequence": sequence, "ability_id": ability},
        )
    for sequence, graph in enumerate(graphs, start=1):
        add(
            "EnemyPhase",
            {
                "id": BASE + 600 + sequence,
                "stable_key": f"goal07.enemy.s03.phase-{sequence}",
                "variant_id": variant,
                "sequence": sequence,
                "entry_condition_id": condition_always,
                "exit_condition_id": condition_always,
                "replacement_priority": sequence,
                "ai_graph_id": graph,
                "targetable": True,
                "transition_model": "TransformSameUnit",
                "hp_carry": "Reset",
                "action_gauge_carry": "Reset",
                "effect_carry": "Clear",
                "toughness_carry": "Reset",
                "summon_carry": "Clear",
            },
        )

    anchor_digest = sha256_bytes(anchor_path(PARTITION).read_bytes())
    add(
        "SourceRecord",
        {
            "id": SOURCE_RECORD_ID,
            "stable_key": "source.hsr-wiki.automaton-grizzly-complete.2026-07-29",
            "category": "CommunityMaintained",
            "publisher": "Honkai: Star Rail Wiki",
            "url": anchor["source"]["url"],
            "accessed_on": anchor["source"]["accessed_on"],
            "applicable_game_version": anchor["source"]["game_version"],
            "confidence": "SecondaryVersionSensitiveCrossCheck",
            "evidence_sha256": anchor_digest,
            "usage_note": (
                "Exact public World 2 level-27 HP, ATK, DEF, SPD, EHR and "
                "Effect RES transcribed into committed Goal 07 evidence."
            ),
        },
    )
    add(
        "EvidenceRecord",
        {
            "id": EVIDENCE_RECORD_ID,
            "stable_key": "evidence.goal07.enemy.s03.numeric-anchors",
            "kind": "SourcePayload",
            "source_record_id": SOURCE_RECORD_ID,
            "sha256": anchor_digest,
            "note": "Committed exact public per-level numeric anchors for Goal 07 S03.",
        },
    )
    for item in identities:
        add("ContentIdentity", item)
        add(
            "ContentEvidenceBinding",
            {
                "content_id": item["id"],
                "sequence": 1,
                "fact_key": f"goal07.s03.executable:{item['stable_key']}",
                "source_record_id": 1,
                "evidence_record_id": 3,
                "quality": "ExactStructured",
                "mechanism_quality": "ExactStructured",
            },
        )
    add(
        "ContentEvidenceBinding",
        {
            "content_id": variant,
            "sequence": 2,
            "fact_key": "goal07.s03.public-level-stats",
            "source_record_id": SOURCE_RECORD_ID,
            "evidence_record_id": EVIDENCE_RECORD_ID,
            "quality": "ExactStructured",
            "mechanism_quality": "ExactStructured",
        },
    )
    for table_rows in rows.values():
        table_rows.sort(
            key=lambda row: json.dumps(
                row, ensure_ascii=False, sort_keys=True, default=str
            )
        )
    return rows


def owned_rows_s04() -> dict[str, list[dict[str, Any]]]:
    anchor = json.loads(anchor_path(PARTITION).read_text(encoding="utf-8"))
    manifest = json.loads(PARTITIONS.read_text(encoding="utf-8"))
    assigned = next(item for item in manifest["partitions"] if item["id"] == PARTITION)
    if assigned["enemy_variant_ids"] != [VARIANT_KEY]:
        raise ValueError("S04 frozen enemy assignment changed")

    variant = BASE + 1
    template = BASE + 2
    graph = BASE + 10
    abilities = {
        "bellowing-inferno": BASE + 101,
        "blazing-absorption": BASE + 102,
        "rain-of-purifying-flames": BASE + 103,
        "molten-fusion": BASE + 104,
    }
    selectors = {
        "actor": BASE + 401,
        "owner": BASE + 402,
        "current-subject": BASE + 403,
        "primary-target": BASE + 404,
        "opposing-random": BASE + 405,
    }
    effects = {
        "enkindle": BASE + 501,
        "spontaneous-combustion": BASE + 502,
        "molten-fusion": BASE + 503,
        "reset-sequence": BASE + 504,
        "needs-absorption": BASE + 505,
    }
    modifier = BASE + 521
    modifier_group = BASE + 531
    reset_rule = BASE + 541
    reset_filter = BASE + 542
    reset_trigger = BASE + 543
    conditions = {
        "always": BASE + 551,
        "reset-sequence": BASE + 552,
        "needs-absorption": BASE + 553,
    }
    rows: dict[str, list[dict[str, Any]]] = {}
    identities: list[dict[str, Any]] = []
    next_program = BASE + 301
    next_operation = BASE + 1_001
    next_expression = BASE + 1_101

    def add(table: str, row: dict[str, Any]) -> None:
        rows.setdefault(table, []).append(row)

    def identity_s04(
        id_: int,
        stable_key: str,
        kind: str,
        name_en: str,
        name_zh_cn: str,
        summary: str,
        sources: str = "1",
    ) -> dict[str, Any]:
        row = identity(
            id_,
            stable_key,
            kind,
            name_en,
            name_zh_cn,
            summary,
            sources,
        )
        row["summary_zh_cn"] = "Goal 07 S04 来源绑定的外宇宙之炎可执行定义。"
        row["game_version_introduced"] = "1.0"
        return row

    identities.extend(
        [
            identity_s04(
                variant,
                VARIANT_KEY,
                "EnemyVariant",
                "Blaze Out of Space",
                "外宇宙之炎",
                "Exact materialization variant used by the frozen universe bindings.",
                "1|6",
            ),
            identity_s04(
                template,
                "enemy.blaze-out-of-space.elite",
                "Enemy",
                "Blaze Out of Space Template",
                "外宇宙之炎模板",
                "Version 4.4 elite template retained from source monster 8003020.",
            ),
            identity_s04(
                graph,
                "ai.goal07.blaze-out-of-space.phase-1",
                "AiGraph",
                "Blaze Out of Space AI",
                "外宇宙之炎AI",
                "Finite source-ordered setup and spontaneous-combustion action graph.",
            ),
        ]
    )
    ability_metadata = {
        "bellowing-inferno": (
            "Bellowing Inferno",
            "咆哮烈焰",
            "Executable transcription of source skill 800302001.",
        ),
        "blazing-absorption": (
            "Blazing Absorption",
            "炽焰吸收",
            "Executable transcription of source skill 800302002.",
        ),
        "rain-of-purifying-flames": (
            "Rain of Purifying Flames",
            "净火之雨",
            "Executable five-hit transcription of source skill 800302003.",
        ),
        "molten-fusion": (
            "Molten Fusion",
            "熔火聚变",
            "Executable transcription of source skill 800302004.",
        ),
    }
    for key, ability_id in abilities.items():
        name_en, name_zh_cn, summary = ability_metadata[key]
        identities.append(
            identity_s04(
                ability_id,
                f"enemy.blaze-out-of-space.elite.ability.{key}",
                "Ability",
                name_en,
                name_zh_cn,
                summary,
            )
        )
    for name, selector_id in selectors.items():
        identities.append(
            identity_s04(
                selector_id,
                f"selector.goal07.blaze-out-of-space.{name}",
                "Selector",
                f"Blaze Out of Space {name} Selector",
                f"外宇宙之炎{name}选择器",
                "S04 battle selector.",
            )
        )
    effect_metadata = {
        "enkindle": ("Enkindle", "焚化", "Stacking two-turn Fire damage-over-time."),
        "spontaneous-combustion": (
            "Spontaneous Combustion",
            "自燃",
            "Non-dispellable state enabling the phase-two action cycle.",
        ),
        "molten-fusion": (
            "Molten Fusion",
            "熔火聚变",
            "Three-stack two-turn ATK increase.",
        ),
        "reset-sequence": (
            "Combustion Reset Sequence",
            "自燃重置序列",
            "Internal Rule IR marker that restarts the source setup sequence.",
        ),
        "needs-absorption": (
            "Combustion Needs Absorption",
            "自燃待吸收",
            "Internal Rule IR marker between Bellowing Inferno and Blazing Absorption.",
        ),
    }
    for key, effect_id in effects.items():
        name_en, name_zh_cn, summary = effect_metadata[key]
        identities.append(
            identity_s04(
                effect_id,
                f"effect.goal07.blaze-out-of-space.{key}",
                "Effect",
                name_en,
                name_zh_cn,
                summary,
            )
        )
    identities.extend(
        [
            identity_s04(
                modifier,
                "modifier.goal07.blaze-out-of-space.molten-fusion",
                "Modifier",
                "Blaze Out of Space Molten Fusion Modifier",
                "外宇宙之炎熔火聚变调整器",
                "Stack-scaled ATK increase sourced by Molten Fusion.",
            ),
            identity_s04(
                reset_rule,
                "rule.goal07.blaze-out-of-space.weakness-break-reset",
                "Rule",
                "Blaze Out of Space Weakness Break Reset",
                "外宇宙之炎弱点击破重置",
                "Effect-owned WeaknessBroken reset for the combustion state and ATK buff.",
            ),
        ]
    )

    add("Selector", selector(selectors["actor"], "Actor", "SameSide"))
    add("Selector", selector(selectors["owner"], "Owner", "SameSide"))
    add(
        "Selector",
        selector(
            selectors["current-subject"],
            "CurrentSubject",
            "OpposingSide",
        ),
    )
    add(
        "Selector",
        selector(selectors["primary-target"], "PrimaryTarget", "OpposingSide"),
    )
    random_selector = selector(
        selectors["opposing-random"],
        "Actor",
        "OpposingSide",
        choice="RngUniform",
    )
    random_selector["rng_purpose_key"] = "damage-target"
    add("Selector", random_selector)

    def expr(name: str, kind: str, node: str) -> int:
        nonlocal next_expression
        id_ = next_expression
        next_expression += 1
        add(
            "ValueExpression",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s04.expression.{name}",
                "result_kind": kind,
                "node": node,
            },
        )
        return id_

    def multiply(name: str, left: int, right: int) -> int:
        return expr(
            name,
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedMultiply",
                left_expression_id=left,
                right_expression_id=right,
                rounding="NearestTiesAway",
            ),
        )

    actor_atk = expr(
        "actor-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    ratio_2_5 = expr(
        "ratio-2-5", "Scalar", json_cell("ScalarLiteral", value_decimal="2.5")
    )
    ratio_1 = expr(
        "ratio-1", "Scalar", json_cell("ScalarLiteral", value_decimal="1")
    )
    ratio_0_5 = expr(
        "ratio-0-5", "Scalar", json_cell("ScalarLiteral", value_decimal="0.5")
    )
    ratio_0_3 = expr(
        "ratio-0-3", "Scalar", json_cell("ScalarLiteral", value_decimal="0.3")
    )
    duration_2 = expr(
        "duration-two", "Integer", json_cell("IntegerLiteral", value=2)
    )
    integer_zero = expr(
        "integer-zero", "Integer", json_cell("IntegerLiteral", value=0)
    )
    reset_stacks = expr(
        "reset-sequence-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["actor"],
            effect_id=effects["reset-sequence"],
        ),
    )
    needs_absorption_stacks = expr(
        "needs-absorption-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["actor"],
            effect_id=effects["needs-absorption"],
        ),
    )
    molten_stacks = expr(
        "molten-fusion-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["owner"],
            effect_id=effects["molten-fusion"],
        ),
    )
    molten_stacks_scalar = expr(
        "molten-fusion-stacks-scalar",
        "Scalar",
        json_cell(
            "Convert",
            operand_expression_id=molten_stacks,
            target_kind="Scalar",
            rounding="NearestTiesAway",
        ),
    )
    bellow_damage = multiply("bellowing-inferno-damage", actor_atk, ratio_2_5)
    rain_hit_damage = multiply("rain-of-purifying-flames-hit", actor_atk, ratio_1)
    enkindle_dot = multiply("enkindle-dot-per-stack", actor_atk, ratio_0_5)
    molten_value = multiply(
        "molten-fusion-atk-ratio", molten_stacks_scalar, ratio_0_3
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["always"],
            "stable_key": "goal07.enemy.s04.condition.always",
            "node": json_cell("Constant", value=True),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["reset-sequence"],
            "stable_key": "goal07.enemy.s04.condition.reset-sequence",
            "node": json_cell(
                "Compare",
                left_expression_id=reset_stacks,
                comparison="Greater",
                right_expression_id=integer_zero,
            ),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["needs-absorption"],
            "stable_key": "goal07.enemy.s04.condition.needs-absorption",
            "node": json_cell(
                "Compare",
                left_expression_id=needs_absorption_stacks,
                comparison="Greater",
                right_expression_id=integer_zero,
            ),
        },
    )

    def operation_s04(
        id_: int,
        name: str,
        payload: str,
        target: int | None = None,
        empty: str = "Fault",
    ) -> dict[str, Any]:
        row = operation(id_, name, payload, target, empty)
        row["stable_key"] = f"goal07.enemy.s04.operation.{name}"
        return row

    def damage_op(name: str, amount: int, target: int) -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        return operation_s04(
            id_,
            name,
            json_cell(
                "Damage",
                amount_expression_id=amount,
                damage_class="Ordinary",
                element="Fire",
                can_crit=True,
            ),
            target,
        )

    def effect_op(
        name: str,
        effect: int,
        target: int,
        *,
        resistible: bool = False,
        remove: bool = False,
        empty: str = "Fault",
    ) -> dict[str, Any]:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        payload = (
            json_cell("RemoveEffect", effect_id=effect)
            if remove
            else json_cell(
                "ApplyEffect",
                effect_id=effect,
                stacks_expression_id=None,
                chance_policy="Resistible" if resistible else "Guaranteed",
                base_chance_expression_id=ratio_1 if resistible else None,
                rng_purpose_key="effect-application" if resistible else None,
            )
        )
        return operation_s04(id_, name, payload, target, empty)

    programs: dict[str, int] = {}

    def program(name: str, steps: list[dict[str, Any] | str]) -> int:
        nonlocal next_program
        id_ = next_program
        next_program += 1
        programs[name] = id_
        identities.append(
            identity_s04(
                id_,
                f"program.goal07.blaze-out-of-space.{name}",
                "Program",
                f"Blaze Out of Space {name} Program",
                f"外宇宙之炎{name}程序",
                "Ordered Rule IR program for the S04 enemy.",
            )
        )
        add("Program", {"id": id_, "domain": "Battle"})
        for sequence, step in enumerate(steps, start=1):
            if isinstance(step, dict):
                add("Operation", step)
                encoded = json_cell("Operation", operation_id=step["id"])
            else:
                encoded = step
            add(
                "ProgramStep",
                {"program_id": id_, "sequence": sequence, "step": encoded},
            )
        return id_

    rain_hit_program = program(
        "rain-of-purifying-flames-hit",
        [
            damage_op(
                "rain-of-purifying-flames-hit-damage",
                rain_hit_damage,
                selectors["current-subject"],
            ),
            effect_op(
                "rain-of-purifying-flames-hit-enkindle",
                effects["enkindle"],
                selectors["current-subject"],
                resistible=True,
            ),
        ],
    )
    ability_programs = {
        abilities["bellowing-inferno"]: program(
            "bellowing-inferno",
            [
                damage_op(
                    "bellowing-inferno-damage",
                    bellow_damage,
                    selectors["primary-target"],
                ),
                effect_op(
                    "bellowing-inferno-enkindle",
                    effects["enkindle"],
                    selectors["primary-target"],
                    resistible=True,
                ),
                effect_op(
                    "bellowing-inferno-clear-reset",
                    effects["reset-sequence"],
                    selectors["actor"],
                    remove=True,
                    empty="NoOp",
                ),
                effect_op(
                    "bellowing-inferno-mark-needs-absorption",
                    effects["needs-absorption"],
                    selectors["actor"],
                ),
            ],
        ),
        abilities["blazing-absorption"]: program(
            "blazing-absorption",
            [
                effect_op(
                    "blazing-absorption-enter-spontaneous-combustion",
                    effects["spontaneous-combustion"],
                    selectors["actor"],
                ),
                effect_op(
                    "blazing-absorption-clear-needs-absorption",
                    effects["needs-absorption"],
                    selectors["actor"],
                    remove=True,
                    empty="NoOp",
                ),
            ],
        ),
        abilities["rain-of-purifying-flames"]: program(
            "rain-of-purifying-flames",
            [
                json_cell(
                    "ForEach",
                    selector_id=selectors["opposing-random"],
                    body_program_id=rain_hit_program,
                    maximum_iterations=1,
                )
                for _ in range(5)
            ],
        ),
        abilities["molten-fusion"]: program(
            "molten-fusion",
            [
                effect_op(
                    "molten-fusion-add-stack",
                    effects["molten-fusion"],
                    selectors["actor"],
                )
            ],
        ),
    }
    reset_program = program(
        "weakness-break-reset",
        [
            effect_op(
                "weakness-break-clear-spontaneous-combustion",
                effects["spontaneous-combustion"],
                selectors["owner"],
                remove=True,
                empty="NoOp",
            ),
            effect_op(
                "weakness-break-clear-molten-fusion",
                effects["molten-fusion"],
                selectors["owner"],
                remove=True,
                empty="NoOp",
            ),
            effect_op(
                "weakness-break-mark-reset-sequence",
                effects["reset-sequence"],
                selectors["owner"],
            ),
        ],
    )

    target_patterns = {
        "bellowing-inferno": "SingleTarget",
        "blazing-absorption": "None",
        "rain-of-purifying-flames": "Bounce",
        "molten-fusion": "None",
    }
    ai_tags = {
        "bellowing-inferno": "skill01-bellowing-inferno",
        "blazing-absorption": "skill03-blazing-absorption",
        "rain-of-purifying-flames": "skill05-rain-of-purifying-flames",
        "molten-fusion": "skill07-molten-fusion",
    }
    for key, ability_id in abilities.items():
        add(
            "Ability",
            {
                "id": ability_id,
                "kind": "Skill",
                "target_pattern": target_patterns[key],
                "retarget_policy": "CancelRemaining",
                "level_cap": 1,
                "cooldown_actions": 1,
                "semantic_tags_mask": (
                    5
                    if key in {"bellowing-inferno", "rain-of-purifying-flames"}
                    else 4
                ),
            },
        )
        add(
            "AbilityPhase",
            {
                "ability_id": ability_id,
                "sequence": 1,
                "kind": "Resolved",
                "program_identity_id": ability_programs[ability_id],
            },
        )
        add(
            "EnemyAbility",
            {
                "id": ability_id,
                "telegraph": "None",
                "cooldown_actions": 1,
                "initial_cooldown_actions": 0,
                "charge_actions": 0,
                "ai_tag": ai_tags[key],
            },
        )

    effect_definitions = {
        "enkindle": {
            "category": "Dot",
            "dispel": "DispellableDebuff",
            "limit": 5,
            "duration": duration_2,
            "clock": "TargetTurnEnd",
            "tick": "TurnStart",
            "policy": "RefreshAndAddStacks",
            "magnitude": enkindle_dot,
            "dot_element": "Fire",
        },
        "spontaneous-combustion": {
            "category": "NeutralState",
            "dispel": "NonDispellable",
            "limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": None,
            "dot_element": None,
        },
        "molten-fusion": {
            "category": "Buff",
            "dispel": "DispellableBuff",
            "limit": 3,
            "duration": duration_2,
            "clock": "OwnerTurnEnd",
            "tick": "None",
            "policy": "RefreshAndAddStacks",
            "magnitude": molten_value,
            "dot_element": None,
        },
        "reset-sequence": {
            "category": "NeutralState",
            "dispel": "NonDispellable",
            "limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": None,
            "dot_element": None,
        },
        "needs-absorption": {
            "category": "NeutralState",
            "dispel": "NonDispellable",
            "limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": None,
            "dot_element": None,
        },
    }
    for key, effect_id in effects.items():
        definition = effect_definitions[key]
        add(
            "Effect",
            {
                "id": effect_id,
                "category": definition["category"],
                "dispel_category": definition["dispel"],
                "stack_limit": definition["limit"],
                "duration_expression_id": definition["duration"],
                "duration_clock": definition["clock"],
                "tick_phase": definition["tick"],
                "stack_policy": definition["policy"],
                "magnitude_comparator_expression_id": definition["magnitude"],
                "dot_element": definition["dot_element"],
                "snapshot_policy": "OnApplication",
                "teardown_policy": "RemoveWithOwner",
                "application_priority": 0,
            },
        )
    for effect_key, tags in {
        "enkindle": ["burn", "enkindle"],
        "spontaneous-combustion": ["spontaneous-combustion"],
        "molten-fusion": ["attack-up", "remove-on-weakness-break"],
        "reset-sequence": ["internal-ai-reset"],
        "needs-absorption": ["internal-ai-setup"],
    }.items():
        for sequence, tag in enumerate(tags, start=1):
            add(
                "EffectTag",
                {
                    "effect_id": effects[effect_key],
                    "sequence": sequence,
                    "tag": tag,
                },
            )

    add(
        "ModifierStackingGroup",
        {
            "id": modifier_group,
            "stable_key": "goal07.enemy.s04.molten-fusion",
            "aggregation": "Sum",
        },
    )
    add(
        "ModifierDefinition",
        {
            "id": modifier,
            "source_effect_id": effects["molten-fusion"],
            "owner_selector_id": selectors["owner"],
            "subject_selector_id": selectors["owner"],
            "stat": "Atk",
            "formula_stage": "PercentOfBase",
            "formula_purpose": "Stat",
            "value_expression_id": molten_value,
            "value_domain": "Ratio",
            "stacking_group_id": modifier_group,
            "priority": 0,
            "cap_formula_stage": "PercentOfBase",
            "snapshot_policy": "Dynamic",
            "duration_scope": "Turn",
        },
    )
    add(
        "EffectModifierBinding",
        {
            "effect_id": effects["molten-fusion"],
            "sequence": 1,
            "modifier_id": modifier,
        },
    )

    add(
        "RuleDefinition",
        {
            "id": reset_rule,
            "domain": "Battle",
            "source_definition_identity_id": effects["spontaneous-combustion"],
            "source_class": "Effect",
            "source_digest_sha256": sha256_text(
                "goal07-s04-spontaneous-combustion-weakness-break-reset-v1"
            ),
        },
    )
    add(
        "EventFilter",
        {
            "id": reset_filter,
            "stable_key": "goal07.enemy.s04.filter.weakness-break-owner",
            "target_selector_id": selectors["owner"],
            "cause_ancestry": "Any",
        },
    )
    add(
        "RuleTrigger",
        {
            "id": reset_trigger,
            "stable_key": "goal07.enemy.s04.trigger.weakness-break-reset",
            "rule_id": reset_rule,
            "sequence": 1,
            "event": json_cell("WeaknessBroken"),
            "phase": "AfterEvent",
            "filter_id": reset_filter,
            "condition_id": conditions["always"],
            "once_scope": "Event",
            "priority": 0,
            "program_id": reset_program,
        },
    )
    add(
        "EffectRuleBinding",
        {
            "effect_id": effects["spontaneous-combustion"],
            "sequence": 1,
            "rule_id": reset_rule,
        },
    )

    state_ids = {
        "initial-bellow": BASE + 701,
        "initial-absorption": BASE + 702,
        "rain-one": BASE + 703,
        "molten": BASE + 704,
        "rain-two": BASE + 705,
        "reset-absorption": BASE + 706,
    }
    add(
        "AiGraph",
        {
            "id": graph,
            "initial_state_id": state_ids["initial-bellow"],
            "automatic_transition_budget": 8,
        },
    )
    normal_abilities = {
        "initial-bellow": abilities["bellowing-inferno"],
        "initial-absorption": abilities["blazing-absorption"],
        "rain-one": abilities["rain-of-purifying-flames"],
        "molten": abilities["molten-fusion"],
        "rain-two": abilities["rain-of-purifying-flames"],
        "reset-absorption": abilities["blazing-absorption"],
    }
    target_selectors = {
        abilities["bellowing-inferno"]: selectors["opposing-random"],
        abilities["blazing-absorption"]: selectors["actor"],
        abilities["rain-of-purifying-flames"]: selectors["opposing-random"],
        abilities["molten-fusion"]: selectors["actor"],
    }
    for state_name, state_id in state_ids.items():
        add(
            "AiState",
            {
                "id": state_id,
                "stable_key": f"goal07.enemy.s04.ai.{state_name}",
                "graph_id": graph,
                "mandatory_fallback_ability_id": abilities["bellowing-inferno"],
                "turn_counter_reset": state_name == "initial-bellow",
            },
        )
    next_candidate = BASE + 801
    phase_two_states = {"rain-one", "molten", "rain-two"}
    for state_name, state_id in state_ids.items():
        candidate_sequence = 1
        if state_name in phase_two_states:
            add(
                "AiCandidate",
                {
                    "id": next_candidate,
                    "stable_key": f"goal07.enemy.s04.ai.{state_name}.reset-bellow",
                    "state_id": state_id,
                    "sequence": candidate_sequence,
                    "ability_id": abilities["bellowing-inferno"],
                    "condition_id": conditions["reset-sequence"],
                    "target_selector_id": selectors["opposing-random"],
                    "priority": 0,
                    "selection": "FirstLegal",
                    "no_target_fallback": "UseFallbackAbility",
                    "fallback_ability_id": abilities["bellowing-inferno"],
                },
            )
            next_candidate += 1
            candidate_sequence += 1
        normal_ability = normal_abilities[state_name]
        add(
            "AiCandidate",
            {
                "id": next_candidate,
                "stable_key": f"goal07.enemy.s04.ai.{state_name}.main",
                "state_id": state_id,
                "sequence": candidate_sequence,
                "ability_id": normal_ability,
                "condition_id": conditions["always"],
                "target_selector_id": target_selectors[normal_ability],
                "priority": 10,
                "selection": "FirstLegal",
                "no_target_fallback": "UseFallbackAbility",
                "fallback_ability_id": abilities["bellowing-inferno"],
            },
        )
        next_candidate += 1

    normal_transitions = {
        "initial-bellow": "initial-absorption",
        "initial-absorption": "rain-one",
        "rain-one": "molten",
        "molten": "rain-two",
        "rain-two": "rain-one",
        "reset-absorption": "rain-one",
    }
    next_transition = BASE + 901
    for state_name, target_name in normal_transitions.items():
        state_id = state_ids[state_name]
        transition_sequence = 1
        if state_name in phase_two_states:
            add(
                "AiTransition",
                {
                    "id": next_transition,
                    "stable_key": (
                        f"goal07.enemy.s04.ai.{state_name}.to-reset-absorption"
                    ),
                    "state_id": state_id,
                    "sequence": transition_sequence,
                    "target_state_id": state_ids["reset-absorption"],
                    "condition_id": conditions["needs-absorption"],
                    "priority": 0,
                    "timing": "AfterAction",
                },
            )
            next_transition += 1
            transition_sequence += 1
        add(
            "AiTransition",
            {
                "id": next_transition,
                "stable_key": f"goal07.enemy.s04.ai.{state_name}.normal-transition",
                "state_id": state_id,
                "sequence": transition_sequence,
                "target_state_id": state_ids[target_name],
                "condition_id": conditions["always"],
                "priority": 10,
                "timing": "AfterAction",
            },
        )
        next_transition += 1

    add(
        "EnemyTemplate",
        {
            "id": template,
            "rank": "Elite",
            "base_aggro_decimal": "100",
            "default_ai_graph_id": graph,
        },
    )
    add(
        "EnemyVariant",
        {
            "id": variant,
            "template_id": template,
            "ai_graph_id": graph,
            "mechanically_distinct_key": VARIANT_KEY,
        },
    )
    for level in anchor["levels"]:
        add(
            "EnemyStat",
            {
                "variant_id": variant,
                "level": level["authored_level"],
                "difficulty_key": "standard-universe-v1",
                "hp_decimal": level["base_hp"],
                "atk_decimal": level["base_atk"],
                "def_decimal": level["base_def"],
                "spd_decimal": level["base_spd"],
                "effect_hit_rate_decimal": level["effect_hit_rate"],
                "effect_resistance_decimal": level["effect_resistance"],
                "crit_damage_decimal": "0.2",
            },
        )
    for sequence, weakness in enumerate(
        ["Physical", "Ice", "Quantum"], start=1
    ):
        add(
            "EnemyWeakness",
            {"variant_id": variant, "sequence": sequence, "element": weakness},
        )
    for element, value in [
        ("Fire", "0.4"),
        ("Lightning", "0.2"),
        ("Wind", "0.2"),
        ("Imaginary", "0.2"),
    ]:
        add(
            "EnemyResistance",
            {"variant_id": variant, "element": element, "value_decimal": value},
        )
    add(
        "EnemyDebuffResistance",
        {
            "variant_id": variant,
            "category_key": "STAT_DOT_Burn",
            "value_decimal": "1",
        },
    )
    add(
        "EnemyToughnessLayer",
        {
            "variant_id": variant,
            "sequence": 1,
            "layer_key": "ordinary",
            "kind": "Ordinary",
            "maximum_decimal": "300",
            "recovery_ratio_decimal": "1",
            "active_at_start": True,
        },
    )
    for sequence, ability_id in enumerate(abilities.values(), start=1):
        add(
            "EnemyVariantAbility",
            {
                "variant_id": variant,
                "sequence": sequence,
                "ability_id": ability_id,
            },
        )
    add(
        "EnemyPhase",
        {
            "id": BASE + 601,
            "stable_key": "goal07.enemy.s04.phase-1",
            "variant_id": variant,
            "sequence": 1,
            "entry_condition_id": conditions["always"],
            "exit_condition_id": conditions["always"],
            "replacement_priority": 1,
            "ai_graph_id": graph,
            "targetable": True,
            "transition_model": "TransformSameUnit",
            "hp_carry": "Reset",
            "action_gauge_carry": "Reset",
            "effect_carry": "Clear",
            "toughness_carry": "Reset",
            "summon_carry": "Clear",
        },
    )

    anchor_digest = sha256_bytes(anchor_path(PARTITION).read_bytes())
    add(
        "SourceRecord",
        {
            "id": SOURCE_RECORD_ID,
            "stable_key": "source.hsr-db.blaze-out-of-space.2026-07-29",
            "category": "CommunityMaintained",
            "publisher": anchor["source"]["publisher"],
            "url": anchor["source"]["url"],
            "accessed_on": anchor["source"]["accessed_on"],
            "applicable_game_version": anchor["source"]["game_version"],
            "confidence": "SecondaryVersionSensitiveCrossCheck",
            "evidence_sha256": anchor_digest,
            "usage_note": (
                "Public level values and the exact retained hard-level curve inputs "
                "are committed as Goal 07 evidence."
            ),
        },
    )
    add(
        "EvidenceRecord",
        {
            "id": EVIDENCE_RECORD_ID,
            "stable_key": "evidence.goal07.enemy.s04.numeric-anchors",
            "kind": "SourcePayload",
            "source_record_id": SOURCE_RECORD_ID,
            "sha256": anchor_digest,
            "note": "Committed exact public level-curve numeric anchors for Goal 07 S04.",
        },
    )
    for item in identities:
        add("ContentIdentity", item)
        add(
            "ContentEvidenceBinding",
            {
                "content_id": item["id"],
                "sequence": 1,
                "fact_key": f"goal07.s04.executable:{item['stable_key']}",
                "source_record_id": 1,
                "evidence_record_id": 3,
                "quality": "ExactStructured",
                "mechanism_quality": "ExactStructured",
            },
        )
    add(
        "ContentEvidenceBinding",
        {
            "content_id": variant,
            "sequence": 2,
            "fact_key": "goal07.s04.public-level-curve",
            "source_record_id": SOURCE_RECORD_ID,
            "evidence_record_id": EVIDENCE_RECORD_ID,
            "quality": "ExactStructured",
            "mechanism_quality": "ExactStructured",
        },
    )
    for table_rows in rows.values():
        table_rows.sort(
            key=lambda row: json.dumps(
                row, ensure_ascii=False, sort_keys=True, default=str
            )
        )
    return rows


def owned_rows_s05() -> dict[str, list[dict[str, Any]]]:
    anchor = json.loads(anchor_path(PARTITION).read_text(encoding="utf-8"))
    manifest = json.loads(PARTITIONS.read_text(encoding="utf-8"))
    assigned = next(item for item in manifest["partitions"] if item["id"] == PARTITION)
    if assigned["enemy_variant_ids"] != [VARIANT_KEY]:
        raise ValueError("S05 frozen enemy assignment changed")

    variant = BASE + 1
    template = BASE + 2
    graphs = [BASE + 10, BASE + 11, BASE + 12]
    abilities = {
        "rapturous-wind": BASE + 101,
        "swallow-return": BASE + 102,
        "ironthorn": BASE + 103,
        "ballad-formation-breaker": BASE + 104,
        "qi-advance": BASE + 105,
        "cascading-laceration": BASE + 106,
        "qi-converge": BASE + 107,
        "aethereal-dreamflux": BASE + 108,
        "swallow-return-ordeal": BASE + 109,
        "jadecarve-strike": BASE + 121,
    }
    linked = {
        "sword-1": BASE + 201,
        "sword-2": BASE + 202,
        "sword-4": BASE + 204,
        "sword-5": BASE + 205,
    }
    selectors = {
        "actor": BASE + 401,
        "owner": BASE + 402,
        "applier": BASE + 403,
        "current-subject": BASE + 404,
        "primary-target": BASE + 405,
        "opposing-random": BASE + 406,
        "opposing-all": BASE + 407,
        "primary-adjacent": BASE + 408,
        "locked-target": BASE + 409,
        "actor-summons": BASE + 410,
        "owner-summons": BASE + 411,
        "pair-2-4-random": BASE + 412,
        "pair-1-5-random": BASE + 413,
        "ordeal-summons": BASE + 414,
        "ordinary-summons": BASE + 415,
        "ordinary-random": BASE + 416,
        "event-target": BASE + 417,
        "pair-2-4-all": BASE + 418,
        "ordeal-pair-2-4": BASE + 419,
        "ordeal-pair-1-5": BASE + 420,
    }
    effects = {
        "chilling-light": BASE + 501,
        "sword-formation": BASE + 502,
        "formation-core-wind": BASE + 503,
        "formation-core-lightning": BASE + 504,
        "formation-core-imaginary": BASE + 505,
        "ordeal": BASE + 506,
        "freeze": BASE + 507,
        "qi-lock": BASE + 508,
        "qi-advance": BASE + 509,
    }
    modifiers = {
        "chilling-light": BASE + 521,
        "sword-formation": BASE + 522,
    }
    modifier_groups = {
        "chilling-light": BASE + 531,
        "sword-formation": BASE + 532,
    }
    rules = {
        "core-wind-break": BASE + 541,
        "core-lightning-break": BASE + 542,
        "core-imaginary-break": BASE + 543,
        "formation-collapse": BASE + 544,
        "ordeal-freeze": BASE + 545,
    }
    filters = {
        "core-wind-break": BASE + 551,
        "core-lightning-break": BASE + 552,
        "core-imaginary-break": BASE + 553,
        "formation-collapse": BASE + 554,
        "ordeal-freeze": BASE + 555,
    }
    conditions = {
        "always": BASE + 561,
        "core-wind": BASE + 562,
        "core-lightning": BASE + 563,
        "core-imaginary": BASE + 564,
        "no-swords": BASE + 565,
    }
    rows: dict[str, list[dict[str, Any]]] = {}
    identities: list[dict[str, Any]] = []
    next_program = BASE + 301
    next_operation = BASE + 1_001
    next_expression = BASE + 1_101

    def add(table: str, row: dict[str, Any]) -> None:
        rows.setdefault(table, []).append(row)

    def identity_s05(
        id_: int,
        stable_key: str,
        kind: str,
        name_en: str,
        name_zh_cn: str,
        summary: str,
        sources: str = "1",
    ) -> dict[str, Any]:
        row = identity(
            id_,
            stable_key,
            kind,
            name_en,
            name_zh_cn,
            summary,
            sources,
        )
        row["summary_zh_cn"] = "Goal 07 S05 来源绑定的云骑骁卫·彦卿（完整）可执行定义。"
        row["game_version_introduced"] = "1.5"
        return row

    identities.extend(
        [
            identity_s05(
                variant,
                VARIANT_KEY,
                "EnemyVariant",
                "Cloud Knight Lieutenant: Yanqing (Complete)",
                "云骑骁卫·彦卿（完整）",
                "Exact materialization variant used by frozen World 8 bindings.",
                "1|7",
            ),
            identity_s05(
                template,
                "enemy.cloud-knight-lieutenant-yanqing-complete.littleboss",
                "Enemy",
                "Cloud Knight Lieutenant: Yanqing (Complete) Template",
                "云骑骁卫·彦卿（完整）模板",
                "Version 4.4 boss template retained from source monster 2004021.",
            ),
        ]
    )
    for phase, graph in enumerate(graphs, start=1):
        identities.append(
            identity_s05(
                graph,
                f"ai.goal07.cloud-knight-lieutenant-yanqing-complete.phase-{phase}",
                "AiGraph",
                f"Yanqing Complete Phase {phase} AI",
                f"彦卿完整形态{phase}阶段AI",
                "Finite source-ordered boss action graph.",
            )
        )
    ability_metadata = {
        "rapturous-wind": ("Rapturous Wind", "快雨燕相逐", "200% Ice single-target strike."),
        "swallow-return": ("Swallow Return", "遥击三尺水", "Summons four Flying Swords and creates Sword Formation."),
        "ironthorn": ("Ironthorn", "铁马冰河入梦来", "350% Ice strike with one-turn Freeze."),
        "ballad-formation-breaker": (
            "Ballad, Formation Breaker",
            "剑气吟",
            "300% primary and 200% adjacent Ice damage.",
        ),
        "qi-advance": ("Qi Advance", "蓄势", "Charges Cascading Laceration."),
        "cascading-laceration": (
            "Cascading Laceration",
            "破阵",
            "400% Ice damage to all opponents.",
        ),
        "qi-converge": ("Qi Converge", "凝滞", "Locks the future Aethereal target."),
        "aethereal-dreamflux": (
            "Aethereal Dreamflux",
            "空梦",
            "300% Ice damage plus 300% for each surviving Flying Sword.",
        ),
        "swallow-return-ordeal": (
            "Swallow Return — Ordeal",
            "遥击三尺水·罹厄",
            "Phase-three Sword Formation with two Ordeal swords.",
        ),
        "jadecarve-strike": (
            "Jadecarve Strike",
            "琢玉",
            "Flying Sword 240% Ice single-target strike.",
        ),
    }
    for key, ability_id in abilities.items():
        name_en, name_zh_cn, summary = ability_metadata[key]
        identities.append(
            identity_s05(
                ability_id,
                f"enemy.cloud-knight-lieutenant-yanqing-complete.ability.{key}",
                "Ability",
                name_en,
                name_zh_cn,
                summary,
            )
        )
    for key, linked_id in linked.items():
        identities.append(
            identity_s05(
                linked_id,
                f"unit.goal07.cloud-knight-lieutenant-yanqing-complete.{key}",
                "CharacterForm",
                f"Yanqing Flying Sword {key[-1]}",
                f"彦卿飞剑{key[-1]}",
                "Owner-scaled targetable Flying Sword summon.",
            )
        )
    for key, selector_id_ in selectors.items():
        identities.append(
            identity_s05(
                selector_id_,
                f"selector.goal07.cloud-knight-lieutenant-yanqing-complete.{key}",
                "Selector",
                f"Yanqing {key} Selector",
                f"彦卿{key}选择器",
                "S05 battle selector.",
            )
        )
    effect_metadata = {
        "chilling-light": ("Chilling Light", "寒光", "Permanent 10% damage stack."),
        "sword-formation": (
            "Sword Formation",
            "剑阵",
            "Protects toughness and grants 60 flat SPD.",
        ),
        "formation-core-wind": ("Formation Core — Wind", "阵眼·风", "True Wind weakness."),
        "formation-core-lightning": (
            "Formation Core — Lightning",
            "阵眼·雷",
            "True Lightning weakness.",
        ),
        "formation-core-imaginary": (
            "Formation Core — Imaginary",
            "阵眼·虚数",
            "True Imaginary weakness.",
        ),
        "ordeal": ("Ordeal", "罹厄", "Empowered sword toughness and attack Freeze."),
        "freeze": ("Yanqing Freeze", "彦卿冻结", "One-turn control with delayed Ice damage."),
        "qi-lock": ("Qi Converge Lock", "凝滞锁定", "Aethereal target marker."),
        "qi-advance": ("Qi Advance Charge", "蓄势", "Cascading Laceration charge marker."),
    }
    for key, effect_id in effects.items():
        name_en, name_zh_cn, summary = effect_metadata[key]
        identities.append(
            identity_s05(
                effect_id,
                f"effect.goal07.cloud-knight-lieutenant-yanqing-complete.{key}",
                "Effect",
                name_en,
                name_zh_cn,
                summary,
            )
        )
    for key, modifier_id in modifiers.items():
        identities.append(
            identity_s05(
                modifier_id,
                f"modifier.goal07.cloud-knight-lieutenant-yanqing-complete.{key}",
                "Modifier",
                f"Yanqing {key} Modifier",
                f"彦卿{key}调整器",
                "Effect-owned S05 stat modifier.",
            )
        )
    for key, rule_id in rules.items():
        identities.append(
            identity_s05(
                rule_id,
                f"rule.goal07.cloud-knight-lieutenant-yanqing-complete.{key}",
                "Rule",
                f"Yanqing {key} Rule",
                f"彦卿{key}规则",
                "Effect-owned S05 Rule IR lifecycle rule.",
            )
        )

    add("Selector", selector(selectors["actor"], "Actor", "SameSide"))
    add("Selector", selector(selectors["owner"], "Owner", "SameSide"))
    add("Selector", selector(selectors["applier"], "Applier", "SameSide"))
    add(
        "Selector",
        selector(selectors["current-subject"], "CurrentSubject", "AnySide"),
    )
    add(
        "Selector",
        selector(selectors["primary-target"], "PrimaryTarget", "OpposingSide"),
    )
    random_opponent = selector(
        selectors["opposing-random"],
        "Actor",
        "OpposingSide",
        choice="RngUniform",
    )
    random_opponent["rng_purpose_key"] = "damage-target"
    add("Selector", random_opponent)
    add(
        "Selector",
        selector(
            selectors["opposing-all"],
            "Actor",
            "OpposingSide",
            minimum=1,
            maximum=8,
            choice="All",
        ),
    )
    add(
        "Selector",
        selector(
            selectors["primary-adjacent"],
            "Actor",
            "OpposingSide",
            minimum=1,
            maximum=3,
            choice="PrimaryPlusAdjacent",
        ),
    )
    add(
        "Selector",
        selector(selectors["locked-target"], "Actor", "OpposingSide"),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["locked-target"],
            "sequence": 1,
            "predicate": json_cell("HasEffect", effect_id=effects["qi-lock"]),
        },
    )
    for selector_name, origin in [
        ("actor-summons", "Actor"),
        ("owner-summons", "Owner"),
    ]:
        add(
            "Selector",
            selector(
                selectors[selector_name],
                origin,
                "SameSide",
                minimum=0,
                maximum=4,
                empty="NoOp",
                choice="All",
            ),
        )
        add(
            "SelectorPredicate",
            {
                "selector_id": selectors[selector_name],
                "sequence": 1,
                "predicate": json_cell(
                    "OwnedBy",
                    owner_selector_id=(
                        selectors["actor"] if origin == "Actor" else selectors["owner"]
                    ),
                ),
            },
        )
    for selector_name, lower, upper in [
        ("pair-2-4-random", 2, 4),
        ("pair-1-5-random", 1, 5),
    ]:
        selected = selector(
            selectors[selector_name],
            "Actor",
            "SameSide",
            minimum=1,
            maximum=1,
            choice="RngUniform",
        )
        selected["rng_purpose_key"] = "behavior-choice"
        add("Selector", selected)
        add(
            "SelectorPredicate",
            {
                "selector_id": selectors[selector_name],
                "sequence": 1,
                "predicate": json_cell(
                    "OwnedBy", owner_selector_id=selectors["actor"]
                ),
            },
        )
        add(
            "SelectorPredicate",
            {
                "selector_id": selectors[selector_name],
                "sequence": 2,
                "predicate": json_cell(
                    "FormationRange",
                    minimum_index=lower,
                    maximum_index=upper,
                ),
            },
        )
        if selector_name == "pair-1-5-random":
            add(
                "SelectorPredicate",
                {
                    "selector_id": selectors[selector_name],
                    "sequence": 3,
                    "predicate": json_cell(
                        "Excludes", excluded_selector_id=selectors["pair-2-4-all"]
                    ),
                },
            )
    add(
        "Selector",
        selector(
            selectors["pair-2-4-all"],
            "Actor",
            "SameSide",
            minimum=0,
            maximum=2,
            empty="NoOp",
            choice="All",
        ),
    )
    for sequence, predicate in enumerate(
        [
            json_cell("OwnedBy", owner_selector_id=selectors["actor"]),
            json_cell("FormationRange", minimum_index=2, maximum_index=4),
        ],
        start=1,
    ):
        add(
            "SelectorPredicate",
            {
                "selector_id": selectors["pair-2-4-all"],
                "sequence": sequence,
                "predicate": predicate,
            },
        )
    for selector_name, effect_id, excluded in [
        ("ordeal-summons", effects["ordeal"], None),
        ("ordinary-summons", None, selectors["ordeal-summons"]),
    ]:
        add(
            "Selector",
            selector(
                selectors[selector_name],
                "Actor",
                "SameSide",
                minimum=0,
                maximum=4,
                empty="NoOp",
                choice="All",
            ),
        )
        add(
            "SelectorPredicate",
            {
                "selector_id": selectors[selector_name],
                "sequence": 1,
                "predicate": json_cell(
                    "OwnedBy", owner_selector_id=selectors["actor"]
                ),
            },
        )
        add(
            "SelectorPredicate",
            {
                "selector_id": selectors[selector_name],
                "sequence": 2,
                "predicate": (
                    json_cell("HasEffect", effect_id=effect_id)
                    if effect_id is not None
                    else json_cell("Excludes", excluded_selector_id=excluded)
                ),
            },
        )
    for selector_name, lower, upper in [
        ("ordeal-pair-2-4", 2, 4),
        ("ordeal-pair-1-5", 1, 5),
    ]:
        add(
            "Selector",
            selector(
                selectors[selector_name],
                "Actor",
                "SameSide",
                minimum=1,
                maximum=1,
            ),
        )
        predicates = [
            json_cell("OwnedBy", owner_selector_id=selectors["actor"]),
            json_cell("FormationRange", minimum_index=lower, maximum_index=upper),
            json_cell("HasEffect", effect_id=effects["ordeal"]),
        ]
        if selector_name == "ordeal-pair-1-5":
            predicates.insert(
                2,
                json_cell(
                    "Excludes", excluded_selector_id=selectors["pair-2-4-all"]
                ),
            )
        for sequence, predicate in enumerate(predicates, start=1):
            add(
                "SelectorPredicate",
                {
                    "selector_id": selectors[selector_name],
                    "sequence": sequence,
                    "predicate": predicate,
                },
            )
    ordinary_random = selector(
        selectors["ordinary-random"],
        "Actor",
        "SameSide",
        minimum=0,
        maximum=1,
        empty="NoOp",
        choice="RngUniform",
    )
    ordinary_random["rng_purpose_key"] = "behavior-choice"
    add("Selector", ordinary_random)
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["ordinary-random"],
            "sequence": 1,
            "predicate": json_cell(
                "OwnedBy", owner_selector_id=selectors["actor"]
            ),
        },
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["ordinary-random"],
            "sequence": 2,
            "predicate": json_cell(
                "Excludes", excluded_selector_id=selectors["ordeal-summons"]
            ),
        },
    )
    add(
        "Selector",
        selector(
            selectors["event-target"],
            "CurrentSubject",
            "AnySide",
            life="Any",
            presence="Any",
        ),
    )

    def expr(name: str, kind: str, node: str) -> int:
        nonlocal next_expression
        id_ = next_expression
        next_expression += 1
        add(
            "ValueExpression",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s05.expression.{name}",
                "result_kind": kind,
                "node": node,
            },
        )
        return id_

    def multiply(name: str, left: int, right: int) -> int:
        return expr(
            name,
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedMultiply",
                left_expression_id=left,
                right_expression_id=right,
                rounding="NearestTiesAway",
            ),
        )

    actor_atk = expr(
        "actor-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    applier_atk = expr(
        "applier-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["applier"],
            stat="Atk",
            formula_purpose="Dot",
        ),
    )
    owner_max_hp = expr(
        "owner-max-hp",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["owner"],
            stat="Hp",
            formula_purpose="OrdinaryDamage",
        ),
    )
    ratios: dict[str, int] = {}
    for name, value in [
        ("zero", "0"),
        ("one-tenth", "0.1"),
        ("one-fifth", "0.2"),
        ("one-half", "0.5"),
        ("one", "1"),
        ("one-point-two", "1.2"),
        ("two", "2"),
        ("two-point-four", "2.4"),
        ("three", "3"),
        ("three-point-five", "3.5"),
        ("four", "4"),
        ("thirty", "30"),
        ("sixty", "60"),
    ]:
        ratios[name] = expr(
            f"scalar-{name}",
            "Scalar",
            json_cell("ScalarLiteral", value_decimal=value),
        )
    integer_one = expr(
        "integer-one", "Integer", json_cell("IntegerLiteral", value=1)
    )
    duration_one = expr(
        "duration-one", "Integer", json_cell("IntegerLiteral", value=1)
    )
    chilling_stacks = expr(
        "chilling-light-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["owner"],
            effect_id=effects["chilling-light"],
        ),
    )
    chilling_scalar = expr(
        "chilling-light-stacks-scalar",
        "Scalar",
        json_cell(
            "Convert",
            operand_expression_id=chilling_stacks,
            target_kind="Scalar",
            rounding="NearestTiesAway",
        ),
    )
    chilling_value = multiply(
        "chilling-light-damage-boost", chilling_scalar, ratios["one-tenth"]
    )
    damage = {
        "rapturous": multiply("rapturous-wind-damage", actor_atk, ratios["two"]),
        "ironthorn": multiply("ironthorn-damage", actor_atk, ratios["three-point-five"]),
        "ballad-primary": multiply(
            "ballad-primary-damage", actor_atk, ratios["three"]
        ),
        "ballad-adjacent": multiply(
            "ballad-adjacent-damage", actor_atk, ratios["two"]
        ),
        "cascading": multiply(
            "cascading-laceration-damage", actor_atk, ratios["four"]
        ),
        "aethereal": multiply(
            "aethereal-dreamflux-damage", actor_atk, ratios["three"]
        ),
        "jadecarve": multiply(
            "jadecarve-strike-damage", actor_atk, ratios["two-point-four"]
        ),
        "freeze": multiply(
            "freeze-delayed-damage", applier_atk, ratios["one-point-two"]
        ),
        "formation-collapse": multiply(
            "formation-collapse-hp", owner_max_hp, ratios["one-fifth"]
        ),
    }

    add(
        "ConditionExpression",
        {
            "id": conditions["always"],
            "stable_key": "goal07.enemy.s05.condition.always",
            "node": json_cell("Constant", value=True),
        },
    )
    for key, effect_key in [
        ("core-wind", "formation-core-wind"),
        ("core-lightning", "formation-core-lightning"),
        ("core-imaginary", "formation-core-imaginary"),
    ]:
        add(
            "ConditionExpression",
            {
                "id": conditions[key],
                "stable_key": f"goal07.enemy.s05.condition.{key}",
                "node": json_cell(
                    "EffectExists",
                    selector_id=selectors["current-subject"],
                    effect_id=effects[effect_key],
                ),
            },
        )
    add(
        "ConditionExpression",
        {
            "id": conditions["no-swords"],
            "stable_key": "goal07.enemy.s05.condition.no-flying-swords",
            "node": json_cell(
                "SelectorCardinality",
                selector_id=selectors["owner-summons"],
                minimum_count=0,
                maximum_count=0,
            ),
        },
    )

    def new_operation(
        name: str,
        payload: str,
        target: int | None = None,
        empty: str = "Fault",
    ) -> int:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        row = operation(id_, name, payload, target, empty)
        row["stable_key"] = f"goal07.enemy.s05.operation.{name}"
        add("Operation", row)
        return id_

    def operation_step(operation_id: int) -> str:
        return json_cell("Operation", operation_id=operation_id)

    def make_program(name: str, steps: list[str]) -> int:
        nonlocal next_program
        id_ = next_program
        next_program += 1
        identities.append(
            identity_s05(
                id_,
                f"program.goal07.cloud-knight-lieutenant-yanqing-complete.{name}",
                "Program",
                f"Yanqing {name} Program",
                f"彦卿{name}程序",
                "Ordered Rule IR program for the S05 enemy.",
            )
        )
        add("Program", {"id": id_, "domain": "Battle"})
        for sequence, step in enumerate(steps, start=1):
            add(
                "ProgramStep",
                {"program_id": id_, "sequence": sequence, "step": step},
            )
        return id_

    def damage_op(name: str, amount: int, target: int) -> int:
        return new_operation(
            name,
            json_cell(
                "Damage",
                amount_expression_id=amount,
                damage_class="Ordinary",
                element="Ice",
                can_crit=True,
            ),
            target,
        )

    def apply_effect_op(
        name: str,
        effect_id: int,
        target: int,
        *,
        chance: int | None = None,
        stacks: int | None = None,
        empty: str = "Fault",
    ) -> int:
        return new_operation(
            name,
            json_cell(
                "ApplyEffect",
                effect_id=effect_id,
                stacks_expression_id=stacks,
                chance_policy="Resistible" if chance is not None else "Guaranteed",
                base_chance_expression_id=chance,
                rng_purpose_key="effect-application" if chance is not None else None,
            ),
            target,
            empty,
        )

    def remove_effect_op(
        name: str, effect_id: int, target: int, empty: str = "NoOp"
    ) -> int:
        return new_operation(
            name,
            json_cell("RemoveEffect", effect_id=effect_id),
            target,
            empty,
        )

    add_weakness_programs: dict[str, int] = {}
    for element in ["Wind", "Lightning", "Imaginary"]:
        operation_id = new_operation(
            f"formation-core-add-{element.lower()}-weakness",
            json_cell("AddWeakness", element=element),
            selectors["current-subject"],
        )
        add_weakness_programs[element] = make_program(
            f"formation-core-add-{element.lower()}-weakness",
            [operation_step(operation_id)],
        )
    random_core = new_operation(
        "formation-core-random-true-weakness",
        json_cell(
            "ApplyRandomEffect",
            effect_ids=[
                effects["formation-core-wind"],
                effects["formation-core-lightning"],
                effects["formation-core-imaginary"],
            ],
            stacks_expression_id=integer_one,
            choice_rng_purpose_key="weakness-element",
            chance_policy="Guaranteed",
            base_chance_expression_id=None,
            chance_rng_purpose_key=None,
        ),
        selectors["current-subject"],
    )
    core_body = make_program(
        "formation-core-random-true-weakness",
        [
            operation_step(random_core),
            json_cell(
                "If",
                condition_id=conditions["core-wind"],
                then_program_id=add_weakness_programs["Wind"],
                else_program_id=None,
            ),
            json_cell(
                "If",
                condition_id=conditions["core-lightning"],
                then_program_id=add_weakness_programs["Lightning"],
                else_program_id=None,
            ),
            json_cell(
                "If",
                condition_id=conditions["core-imaginary"],
                then_program_id=add_weakness_programs["Imaginary"],
                else_program_id=None,
            ),
        ],
    )

    def chilling_steps(prefix: str) -> list[str]:
        return [
            operation_step(
                apply_effect_op(
                    f"{prefix}-chilling-light-yanqing",
                    effects["chilling-light"],
                    selectors["actor"],
                )
            ),
            operation_step(
                apply_effect_op(
                    f"{prefix}-chilling-light-flying-swords",
                    effects["chilling-light"],
                    selectors["actor-summons"],
                    empty="NoOp",
                )
            ),
        ]

    summon_steps: list[str] = []
    for key, linked_id in linked.items():
        summon_steps.append(
            operation_step(
                new_operation(
                    f"swallow-return-summon-{key}",
                    json_cell(
                        "Summon",
                        unit_definition_identity_id=linked_id,
                        owner_selector_id=selectors["actor"],
                    ),
                )
            )
        )
    summon_steps.append(
        operation_step(
            apply_effect_op(
                "swallow-return-sword-formation",
                effects["sword-formation"],
                selectors["actor"],
            )
        )
    )
    summon_steps.append(
        json_cell(
            "ForEach",
            selector_id=selectors["actor-summons"],
            body_program_id=core_body,
            maximum_iterations=4,
        )
    )
    create_ordinary_toughness = new_operation(
        "swallow-return-create-flying-sword-toughness",
        json_cell(
            "CreateToughnessLayer",
            layer_key="formation-core",
            maximum_expression_id=ratios["thirty"],
        ),
        selectors["actor-summons"],
    )
    normal_summon_program = make_program(
        "swallow-return",
        summon_steps + [operation_step(create_ordinary_toughness)],
    )

    ordeal_steps = list(summon_steps)
    for name in ["pair-2-4-random", "pair-1-5-random"]:
        ordeal_steps.append(
            operation_step(
                apply_effect_op(
                    f"swallow-return-ordeal-mark-{name}",
                    effects["ordeal"],
                    selectors[name],
                )
            )
        )
    create_ordeal_toughness = new_operation(
        "swallow-return-create-ordeal-toughness",
        json_cell(
            "CreateToughnessLayer",
            layer_key="formation-core",
            maximum_expression_id=ratios["sixty"],
        ),
        selectors["ordeal-summons"],
        "NoOp",
    )
    create_non_ordeal_toughness = new_operation(
        "swallow-return-create-non-ordeal-toughness",
        json_cell(
            "CreateToughnessLayer",
            layer_key="formation-core",
            maximum_expression_id=ratios["thirty"],
        ),
        selectors["ordinary-summons"],
        "NoOp",
    )
    advance_first_ordeal = new_operation(
        "swallow-return-advance-first-ordeal",
        json_cell("AdvanceAction", amount_expression_id=ratios["one"]),
        selectors["ordeal-pair-2-4"],
    )
    advance_second_ordeal = new_operation(
        "swallow-return-advance-second-ordeal",
        json_cell("AdvanceAction", amount_expression_id=ratios["one-half"]),
        selectors["ordeal-pair-1-5"],
    )
    delay_ordinary = new_operation(
        "swallow-return-delay-one-ordinary-sword",
        json_cell("DelayAction", amount_expression_id=ratios["one-half"]),
        selectors["ordinary-random"],
        "NoOp",
    )
    ordeal_summon_program = make_program(
        "swallow-return-ordeal",
        ordeal_steps
        + [
            operation_step(create_ordeal_toughness),
            operation_step(create_non_ordeal_toughness),
            operation_step(advance_first_ordeal),
            operation_step(advance_second_ordeal),
            operation_step(delay_ordinary),
        ],
    )

    ability_programs = {
        abilities["rapturous-wind"]: make_program(
            "rapturous-wind",
            [
                operation_step(
                    damage_op(
                        "rapturous-wind-damage",
                        damage["rapturous"],
                        selectors["primary-target"],
                    )
                ),
                *chilling_steps("rapturous-wind"),
            ],
        ),
        abilities["swallow-return"]: normal_summon_program,
        abilities["ironthorn"]: make_program(
            "ironthorn",
            [
                operation_step(
                    damage_op(
                        "ironthorn-damage",
                        damage["ironthorn"],
                        selectors["primary-target"],
                    )
                ),
                operation_step(
                    apply_effect_op(
                        "ironthorn-freeze",
                        effects["freeze"],
                        selectors["primary-target"],
                        chance=ratios["one"],
                    )
                ),
                *chilling_steps("ironthorn"),
            ],
        ),
        abilities["ballad-formation-breaker"]: make_program(
            "ballad-formation-breaker",
            [
                operation_step(
                    damage_op(
                        "ballad-primary-damage",
                        damage["ballad-primary"],
                        selectors["primary-target"],
                    )
                ),
                operation_step(
                    damage_op(
                        "ballad-adjacent-damage",
                        damage["ballad-adjacent"],
                        selectors["primary-adjacent"],
                    )
                ),
                *chilling_steps("ballad-formation-breaker"),
            ],
        ),
        abilities["qi-advance"]: make_program(
            "qi-advance",
            [
                operation_step(
                    apply_effect_op(
                        "qi-advance-charge",
                        effects["qi-advance"],
                        selectors["actor"],
                    )
                )
            ],
        ),
        abilities["cascading-laceration"]: make_program(
            "cascading-laceration",
            [
                operation_step(
                    damage_op(
                        "cascading-laceration-damage",
                        damage["cascading"],
                        selectors["opposing-all"],
                    )
                ),
                operation_step(
                    remove_effect_op(
                        "cascading-laceration-clear-charge",
                        effects["qi-advance"],
                        selectors["actor"],
                    )
                ),
                *chilling_steps("cascading-laceration"),
            ],
        ),
        abilities["qi-converge"]: make_program(
            "qi-converge",
            [
                operation_step(
                    apply_effect_op(
                        "qi-converge-lock",
                        effects["qi-lock"],
                        selectors["primary-target"],
                    )
                )
            ],
        ),
        abilities["swallow-return-ordeal"]: ordeal_summon_program,
    }
    aethereal_sword_damage = new_operation(
        "aethereal-surviving-sword-damage",
        json_cell(
            "Damage",
            amount_expression_id=damage["aethereal"],
            damage_class="Ordinary",
            element="Ice",
            can_crit=True,
        ),
        selectors["primary-target"],
    )
    aethereal_despawn = new_operation(
        "aethereal-despawn-surviving-sword",
        json_cell("Despawn"),
        selectors["current-subject"],
    )
    aethereal_body = make_program(
        "aethereal-surviving-sword",
        [operation_step(aethereal_sword_damage), operation_step(aethereal_despawn)],
    )
    ability_programs[abilities["aethereal-dreamflux"]] = make_program(
        "aethereal-dreamflux",
        [
            operation_step(
                damage_op(
                    "aethereal-base-damage",
                    damage["aethereal"],
                    selectors["primary-target"],
                )
            ),
            json_cell(
                "ForEach",
                selector_id=selectors["actor-summons"],
                body_program_id=aethereal_body,
                maximum_iterations=4,
            ),
            operation_step(
                remove_effect_op(
                    "aethereal-clear-sword-formation",
                    effects["sword-formation"],
                    selectors["actor"],
                )
            ),
            operation_step(
                remove_effect_op(
                    "aethereal-clear-target-lock",
                    effects["qi-lock"],
                    selectors["primary-target"],
                )
            ),
            *chilling_steps("aethereal-dreamflux"),
        ],
    )
    ability_programs[abilities["jadecarve-strike"]] = make_program(
        "jadecarve-strike",
        [
            operation_step(
                damage_op(
                    "jadecarve-strike-damage",
                    damage["jadecarve"],
                    selectors["primary-target"],
                )
            )
        ],
    )

    target_patterns = {
        "rapturous-wind": "SingleTarget",
        "swallow-return": "None",
        "ironthorn": "SingleTarget",
        "ballad-formation-breaker": "Blast",
        "qi-advance": "None",
        "cascading-laceration": "Aoe",
        "qi-converge": "SingleTarget",
        "aethereal-dreamflux": "SingleTarget",
        "swallow-return-ordeal": "None",
        "jadecarve-strike": "SingleTarget",
    }
    for key, ability_id in abilities.items():
        add(
            "Ability",
            {
                "id": ability_id,
                "kind": "Summon" if key == "jadecarve-strike" else "Skill",
                "target_pattern": target_patterns[key],
                "retarget_policy": "CancelRemaining",
                "level_cap": 1,
                "cooldown_actions": 1,
                "semantic_tags_mask": (
                    5 if target_patterns[key] not in {"None"} else 4
                ),
            },
        )
        add(
            "AbilityPhase",
            {
                "ability_id": ability_id,
                "sequence": 1,
                "kind": "Resolved",
                "program_identity_id": ability_programs[ability_id],
            },
        )
        if key != "jadecarve-strike":
            add(
                "EnemyAbility",
                {
                    "id": ability_id,
                    "telegraph": "Charge" if key == "qi-advance" else "None",
                    "cooldown_actions": 1,
                    "initial_cooldown_actions": 0,
                    "charge_actions": 1 if key == "qi-advance" else 0,
                    "ai_tag": key,
                },
            )

    effect_definitions = {
        "chilling-light": {
            "category": "Buff",
            "dispel": "NonDispellable",
            "limit": 1000,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "RefreshAndAddStacks",
            "magnitude": chilling_value,
            "dot": None,
        },
        "sword-formation": {
            "category": "NeutralState",
            "dispel": "NonDispellable",
            "limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": None,
            "dot": None,
        },
        "formation-core-wind": {
            "category": "Mark",
            "dispel": "NonDispellable",
            "limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": None,
            "dot": None,
        },
        "formation-core-lightning": {
            "category": "Mark",
            "dispel": "NonDispellable",
            "limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": None,
            "dot": None,
        },
        "formation-core-imaginary": {
            "category": "Mark",
            "dispel": "NonDispellable",
            "limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": None,
            "dot": None,
        },
        "ordeal": {
            "category": "NeutralState",
            "dispel": "NonDispellable",
            "limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": None,
            "dot": None,
        },
        "freeze": {
            "category": "Control",
            "dispel": "CleanseableControl",
            "limit": 1,
            "duration": duration_one,
            "clock": "TargetTurnStart",
            "tick": "TurnStart",
            "policy": "Refresh",
            "magnitude": damage["freeze"],
            "dot": "Ice",
        },
        "qi-lock": {
            "category": "Mark",
            "dispel": "NonDispellable",
            "limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": None,
            "dot": None,
        },
        "qi-advance": {
            "category": "NeutralState",
            "dispel": "NonDispellable",
            "limit": 1,
            "duration": None,
            "clock": "Permanent",
            "tick": "None",
            "policy": "Replace",
            "magnitude": None,
            "dot": None,
        },
    }
    for key, effect_id in effects.items():
        definition = effect_definitions[key]
        add(
            "Effect",
            {
                "id": effect_id,
                "category": definition["category"],
                "dispel_category": definition["dispel"],
                "stack_limit": definition["limit"],
                "duration_expression_id": definition["duration"],
                "duration_clock": definition["clock"],
                "tick_phase": definition["tick"],
                "stack_policy": definition["policy"],
                "magnitude_comparator_expression_id": definition["magnitude"],
                "dot_element": definition["dot"],
                "snapshot_policy": "OnApplication",
                "teardown_policy": "RemoveWithOwner",
                "application_priority": 0,
            },
        )
    for effect_key, tags in {
        "chilling-light": ["chilling-light"],
        "sword-formation": ["sword-formation", "prevents-toughness-reduction"],
        "formation-core-wind": ["formation-core", "true-weakness-wind"],
        "formation-core-lightning": ["formation-core", "true-weakness-lightning"],
        "formation-core-imaginary": ["formation-core", "true-weakness-imaginary"],
        "ordeal": ["ordeal"],
        "freeze": ["freeze", "blocks-normal-action"],
        "qi-lock": ["qi-converge-lock"],
        "qi-advance": ["qi-advance-charge"],
    }.items():
        for sequence, tag in enumerate(tags, start=1):
            add(
                "EffectTag",
                {"effect_id": effects[effect_key], "sequence": sequence, "tag": tag},
            )
    for key, group_id in modifier_groups.items():
        add(
            "ModifierStackingGroup",
            {
                "id": group_id,
                "stable_key": f"goal07.enemy.s05.{key}",
                "aggregation": "Sum",
            },
        )
    add(
        "ModifierDefinition",
        {
            "id": modifiers["chilling-light"],
            "source_effect_id": effects["chilling-light"],
            "owner_selector_id": selectors["owner"],
            "subject_selector_id": selectors["owner"],
            "stat": "Atk",
            "formula_stage": "DamageBoost",
            "formula_purpose": "OrdinaryDamage",
            "value_expression_id": chilling_value,
            "value_domain": "Ratio",
            "stacking_group_id": modifier_groups["chilling-light"],
            "priority": 0,
            "cap_formula_stage": "DamageBoost",
            "snapshot_policy": "Dynamic",
            "duration_scope": "Turn",
        },
    )
    add(
        "ModifierDefinition",
        {
            "id": modifiers["sword-formation"],
            "source_effect_id": effects["sword-formation"],
            "owner_selector_id": selectors["owner"],
            "subject_selector_id": selectors["owner"],
            "stat": "Spd",
            "formula_stage": "Flat",
            "formula_purpose": "Stat",
            "value_expression_id": ratios["sixty"],
            "value_domain": "Scalar",
            "stacking_group_id": modifier_groups["sword-formation"],
            "priority": 0,
            "cap_formula_stage": "Flat",
            "snapshot_policy": "OnApplication",
            "duration_scope": "Turn",
        },
    )
    for sequence, key in enumerate(["chilling-light", "sword-formation"], start=1):
        add(
            "EffectModifierBinding",
            {
                "effect_id": effects[key],
                "sequence": 1,
                "modifier_id": modifiers[key],
            },
        )

    core_break_programs: dict[str, int] = {}
    for element in ["wind", "lightning", "imaginary"]:
        defeat = new_operation(
            f"formation-core-{element}-break-defeat-sword",
            json_cell(
                "ConsumeHp",
                amount_expression_id=owner_max_hp,
                floor_expression_id=ratios["zero"],
            ),
            selectors["owner"],
        )
        core_break_programs[element] = make_program(
            f"formation-core-{element}-break", [operation_step(defeat)]
        )
    collapse_program = make_program(
        "sword-formation-collapse",
        [
            operation_step(
                remove_effect_op(
                    "sword-formation-collapse-remove",
                    effects["sword-formation"],
                    selectors["owner"],
                )
            ),
            operation_step(
                new_operation(
                    "sword-formation-collapse-consume-hp",
                    json_cell(
                        "ConsumeHp",
                        amount_expression_id=damage["formation-collapse"],
                        floor_expression_id=ratios["zero"],
                    ),
                    selectors["owner"],
                )
            ),
        ],
    )
    ordeal_freeze_program = make_program(
        "ordeal-freeze",
        [
            operation_step(
                apply_effect_op(
                    "ordeal-freeze-primary-target",
                    effects["freeze"],
                    selectors["primary-target"],
                    chance=ratios["one"],
                )
            )
        ],
    )
    rule_specs = [
        (
            "core-wind-break",
            effects["formation-core-wind"],
            json_cell("WeaknessBroken"),
            selectors["owner"],
            None,
            core_break_programs["wind"],
            "Battle",
        ),
        (
            "core-lightning-break",
            effects["formation-core-lightning"],
            json_cell("WeaknessBroken"),
            selectors["owner"],
            None,
            core_break_programs["lightning"],
            "Battle",
        ),
        (
            "core-imaginary-break",
            effects["formation-core-imaginary"],
            json_cell("WeaknessBroken"),
            selectors["owner"],
            None,
            core_break_programs["imaginary"],
            "Battle",
        ),
        (
            "formation-collapse",
            effects["sword-formation"],
            json_cell("Unit", point="Defeated"),
            selectors["event-target"],
            None,
            collapse_program,
            "Battle",
        ),
        (
            "ordeal-freeze",
            effects["ordeal"],
            json_cell("Action", point="Resolved"),
            None,
            "Summon",
            ordeal_freeze_program,
            "Action",
        ),
    ]
    for key, source_effect, event, target_selector, action_kind, program_id, once in rule_specs:
        add(
            "RuleDefinition",
            {
                "id": rules[key],
                "domain": "Battle",
                "source_definition_identity_id": source_effect,
                "source_class": "Effect",
                "source_digest_sha256": sha256_text(f"goal07-s05-{key}-v1"),
            },
        )
        event_filter = {
            "id": filters[key],
            "stable_key": f"goal07.enemy.s05.filter.{key}",
            "cause_ancestry": "Any",
        }
        if target_selector is not None:
            event_filter["target_selector_id"] = target_selector
        if action_kind is not None:
            event_filter["actor_selector_id"] = selectors["owner"]
            event_filter["action_kind"] = action_kind
        add("EventFilter", event_filter)
        add(
            "RuleTrigger",
            {
                "id": BASE + 570 + len(rows.get("RuleTrigger", [])),
                "stable_key": f"goal07.enemy.s05.trigger.{key}",
                "rule_id": rules[key],
                "sequence": 1,
                "event": event,
                "phase": "AfterEvent",
                "filter_id": filters[key],
                "condition_id": (
                    conditions["no-swords"]
                    if key == "formation-collapse"
                    else conditions["always"]
                ),
                "once_scope": once,
                "priority": 0,
                "program_id": program_id,
            },
        )
        add(
            "EffectRuleBinding",
            {"effect_id": source_effect, "sequence": 1, "rule_id": rules[key]},
        )

    phase_sequences = [
        [
            abilities["rapturous-wind"],
            abilities["swallow-return"],
            abilities["ironthorn"],
            abilities["ballad-formation-breaker"],
        ],
        [
            abilities["swallow-return"],
            abilities["rapturous-wind"],
            abilities["ballad-formation-breaker"],
            abilities["qi-advance"],
            abilities["cascading-laceration"],
            abilities["ironthorn"],
        ],
        [
            abilities["swallow-return-ordeal"],
            abilities["rapturous-wind"],
            abilities["qi-converge"],
            abilities["ballad-formation-breaker"],
            abilities["aethereal-dreamflux"],
            abilities["ironthorn"],
        ],
    ]
    target_selectors = {
        abilities["rapturous-wind"]: selectors["opposing-random"],
        abilities["swallow-return"]: selectors["actor"],
        abilities["ironthorn"]: selectors["opposing-random"],
        abilities["ballad-formation-breaker"]: selectors["opposing-random"],
        abilities["qi-advance"]: selectors["actor"],
        abilities["cascading-laceration"]: selectors["opposing-all"],
        abilities["qi-converge"]: selectors["opposing-random"],
        abilities["aethereal-dreamflux"]: selectors["locked-target"],
        abilities["swallow-return-ordeal"]: selectors["actor"],
    }
    next_state = BASE + 701
    next_candidate = BASE + 801
    next_transition = BASE + 901
    for phase_index, sequence in enumerate(phase_sequences):
        state_ids = list(range(next_state, next_state + len(sequence)))
        next_state += len(sequence)
        add(
            "AiGraph",
            {
                "id": graphs[phase_index],
                "initial_state_id": state_ids[0],
                "automatic_transition_budget": 8,
            },
        )
        for offset, (state_id, ability_id) in enumerate(zip(state_ids, sequence)):
            add(
                "AiState",
                {
                    "id": state_id,
                    "stable_key": (
                        f"goal07.enemy.s05.ai.phase-{phase_index + 1}."
                        f"state-{offset + 1}"
                    ),
                    "graph_id": graphs[phase_index],
                    "mandatory_fallback_ability_id": abilities["rapturous-wind"],
                    "turn_counter_reset": offset == 0,
                },
            )
            add(
                "AiCandidate",
                {
                    "id": next_candidate,
                    "stable_key": (
                        f"goal07.enemy.s05.ai.phase-{phase_index + 1}."
                        f"candidate-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "ability_id": ability_id,
                    "condition_id": conditions["always"],
                    "target_selector_id": target_selectors[ability_id],
                    "priority": 0,
                    "selection": "FirstLegal",
                    "no_target_fallback": "UseFallbackAbility",
                    "fallback_ability_id": abilities["rapturous-wind"],
                },
            )
            next_candidate += 1
            add(
                "AiTransition",
                {
                    "id": next_transition,
                    "stable_key": (
                        f"goal07.enemy.s05.ai.phase-{phase_index + 1}."
                        f"transition-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "target_state_id": state_ids[(offset + 1) % len(state_ids)],
                    "condition_id": conditions["always"],
                    "priority": 0,
                    "timing": "AfterAction",
                },
            )
            next_transition += 1

    for key, linked_id in linked.items():
        add(
            "LinkedUnitDefinition",
            {
                "id": linked_id,
                "source_definition_identity_id": linked_id,
                "kind": "Summon",
                "presence": "Present",
                "ability_ids": str(abilities["jadecarve-strike"]),
                "action_ability_id": abilities["jadecarve-strike"],
                "formation_index": int(key[-1]),
                "initial_gauge_decimal": "10000",
                "hp_owner_ratio_decimal": "0.36",
                "hp_flat_decimal": "0",
                "atk_owner_ratio_decimal": "1",
                "atk_flat_decimal": "0",
                "def_owner_ratio_decimal": "1",
                "def_flat_decimal": "0",
                "spd_owner_ratio_decimal": "0",
                "spd_flat_decimal": "100",
                "owner_defeat_policy": "Depart",
                "owner_departure_policy": "Depart",
                "wave_policy": "Depart",
                "combatant_digest_sha256": sha256_text(
                    f"goal07-s05-linked-{key}-v1"
                ),
            },
        )

    add(
        "EnemyTemplate",
        {
            "id": template,
            "rank": "Boss",
            "base_aggro_decimal": "100",
            "default_ai_graph_id": graphs[0],
        },
    )
    add(
        "EnemyVariant",
        {
            "id": variant,
            "template_id": template,
            "ai_graph_id": graphs[0],
            "mechanically_distinct_key": VARIANT_KEY,
        },
    )
    for level in anchor["levels"]:
        add(
            "EnemyStat",
            {
                "variant_id": variant,
                "level": level["authored_level"],
                "difficulty_key": "standard-universe-v1",
                "hp_decimal": level["base_hp"],
                "atk_decimal": level["base_atk"],
                "def_decimal": level["base_def"],
                "spd_decimal": level["base_spd"],
                "effect_hit_rate_decimal": level["effect_hit_rate"],
                "effect_resistance_decimal": level["effect_resistance"],
                "crit_damage_decimal": "0.2",
            },
        )
    for sequence, weakness in enumerate(["Imaginary", "Lightning", "Wind"], start=1):
        add(
            "EnemyWeakness",
            {"variant_id": variant, "sequence": sequence, "element": weakness},
        )
    for element, value in [
        ("Fire", "0.2"),
        ("Ice", "0.4"),
        ("Physical", "0.2"),
        ("Quantum", "0.2"),
    ]:
        add(
            "EnemyResistance",
            {"variant_id": variant, "element": element, "value_decimal": value},
        )
    for category in ["STAT_CTRL_Confine", "STAT_CTRL_Frozen", "STAT_CTRL_Entangle"]:
        add(
            "EnemyDebuffResistance",
            {
                "variant_id": variant,
                "category_key": category,
                "value_decimal": "0.75",
            },
        )
    add(
        "EnemyToughnessLayer",
        {
            "variant_id": variant,
            "sequence": 1,
            "layer_key": "ordinary",
            "kind": "Ordinary",
            "maximum_decimal": "120",
            "recovery_ratio_decimal": "1",
            "active_at_start": True,
        },
    )
    for sequence, ability_id in enumerate(
        [value for key, value in abilities.items() if key != "jadecarve-strike"],
        start=1,
    ):
        add(
            "EnemyVariantAbility",
            {
                "variant_id": variant,
                "sequence": sequence,
                "ability_id": ability_id,
            },
        )
    for sequence, graph in enumerate(graphs, start=1):
        add(
            "EnemyPhase",
            {
                "id": BASE + 600 + sequence,
                "stable_key": f"goal07.enemy.s05.phase-{sequence}",
                "variant_id": variant,
                "sequence": sequence,
                "entry_condition_id": conditions["always"],
                "exit_condition_id": conditions["always"],
                "replacement_priority": sequence,
                "ai_graph_id": graph,
                "targetable": True,
                "transition_model": "TransformSameUnit",
                "hp_carry": "Reset",
                "action_gauge_carry": "Reset",
                "effect_carry": "Clear",
                "toughness_carry": "Reset",
                "summon_carry": "Clear",
            },
        )

    anchor_digest = sha256_bytes(anchor_path(PARTITION).read_bytes())
    add(
        "SourceRecord",
        {
            "id": SOURCE_RECORD_ID,
            "stable_key": "source.hsr-wiki.cloud-knight-lieutenant-yanqing-complete.2026-07-29",
            "category": "CommunityMaintained",
            "publisher": anchor["source"]["publisher"],
            "url": anchor["source"]["url"],
            "accessed_on": anchor["source"]["accessed_on"],
            "applicable_game_version": anchor["source"]["game_version"],
            "confidence": "SecondaryVersionSensitiveCrossCheck",
            "evidence_sha256": anchor_digest,
            "usage_note": (
                "Exact public per-level values are committed with retained "
                "structured AI and ability source hashes."
            ),
        },
    )
    add(
        "EvidenceRecord",
        {
            "id": EVIDENCE_RECORD_ID,
            "stable_key": "evidence.goal07.enemy.s05.numeric-anchors",
            "kind": "SourcePayload",
            "source_record_id": SOURCE_RECORD_ID,
            "sha256": anchor_digest,
            "note": "Committed exact public per-level numeric anchors for Goal 07 S05.",
        },
    )
    for item in identities:
        add("ContentIdentity", item)
        add(
            "ContentEvidenceBinding",
            {
                "content_id": item["id"],
                "sequence": 1,
                "fact_key": f"goal07.s05.executable:{item['stable_key']}",
                "source_record_id": 1,
                "evidence_record_id": 3,
                "quality": "ExactStructured",
                "mechanism_quality": "ExactStructured",
            },
        )
    add(
        "ContentEvidenceBinding",
        {
            "content_id": variant,
            "sequence": 2,
            "fact_key": "goal07.s05.public-level-stats",
            "source_record_id": SOURCE_RECORD_ID,
            "evidence_record_id": EVIDENCE_RECORD_ID,
            "quality": "ExactStructured",
            "mechanism_quality": "ExactStructured",
        },
    )
    for table_rows in rows.values():
        table_rows.sort(
            key=lambda row: json.dumps(
                row, ensure_ascii=False, sort_keys=True, default=str
            )
        )
    return rows


def owned_rows_s06() -> dict[str, list[dict[str, Any]]]:
    anchor = json.loads(anchor_path(PARTITION).read_text(encoding="utf-8"))
    manifest = json.loads(PARTITIONS.read_text(encoding="utf-8"))
    assigned = next(item for item in manifest["partitions"] if item["id"] == PARTITION)
    if assigned["enemy_variant_ids"] != [VARIANT_KEY]:
        raise ValueError("S06 frozen enemy assignment changed")

    variant = BASE + 1
    template = BASE + 2
    graphs = [BASE + 10, BASE + 11, BASE + 12]
    abilities = {
        "chill": BASE + 101,
        "omen": BASE + 102,
        "hoarfrost": BASE + 104,
        "wrath": BASE + 105,
        "punishment": BASE + 103,
        "reverberating": BASE + 106,
        "omen-bronya": BASE + 108,
        "doomsday": BASE + 109,
        "icy-wind": BASE + 121,
        "bronya-cycle": BASE + 122,
    }
    linked = {
        "ice-edge-left": BASE + 201,
        "ice-edge-right": BASE + 202,
        "bronya": BASE + 203,
    }
    selectors = {
        "actor": BASE + 401,
        "owner": BASE + 402,
        "applier": BASE + 403,
        "current-subject": BASE + 404,
        "primary-target": BASE + 405,
        "opposing-random": BASE + 406,
        "opposing-all": BASE + 407,
        "adjacent": BASE + 408,
        "frozen-primary": BASE + 409,
        "frozen-all": BASE + 410,
        "adjacent-frozen": BASE + 411,
        "nonfrozen-random": BASE + 412,
        "actor-summons": BASE + 413,
        "event-target": BASE + 414,
    }
    effects = {
        "freeze": BASE + 501,
        "charging": BASE + 502,
        "intensifying-cold": BASE + 503,
        "redeployment": BASE + 504,
    }
    modifiers = {
        "charging": BASE + 521,
        "intensifying-cold": BASE + 522,
        "redeployment": BASE + 523,
    }
    modifier_groups = {
        "charging": BASE + 531,
        "intensifying-cold": BASE + 532,
        "redeployment": BASE + 533,
    }
    rules = {"freeze-applied": BASE + 541, "freeze-removed": BASE + 542}
    filters = {"freeze-applied": BASE + 551, "freeze-removed": BASE + 552}
    conditions = {"always": BASE + 561, "adjacent-frozen": BASE + 562}
    rows: dict[str, list[dict[str, Any]]] = {}
    identities: list[dict[str, Any]] = []
    next_program = BASE + 301
    next_operation = BASE + 1_001
    next_expression = BASE + 1_101

    def add(table: str, row: dict[str, Any]) -> None:
        rows.setdefault(table, []).append(row)

    def identity_s06(
        id_: int,
        stable_key: str,
        kind: str,
        name_en: str,
        name_zh_cn: str,
        summary: str,
        sources: str = "1",
    ) -> dict[str, Any]:
        row = identity(id_, stable_key, kind, name_en, name_zh_cn, summary, sources)
        row["summary_zh_cn"] = "Goal 07 S06 来源绑定的可可利亚（完整）可执行定义。"
        row["game_version_introduced"] = "1.0"
        return row

    identities.extend(
        [
            identity_s06(
                variant,
                VARIANT_KEY,
                "EnemyVariant",
                "Cocolia (Complete)",
                "可可利亚（完整）",
                "Exact materialization variant used by frozen World 6 bindings.",
                "1|8",
            ),
            identity_s06(
                template,
                "enemy.cocolia-complete.littleboss",
                "Enemy",
                "Cocolia (Complete) Template",
                "可可利亚（完整）模板",
                "Version 4.4 boss template retained from source monster 1004011.",
            ),
        ]
    )
    for phase, graph in enumerate(graphs, start=1):
        identities.append(
            identity_s06(
                graph,
                f"ai.goal07.cocolia-complete.phase-{phase}",
                "AiGraph",
                f"Cocolia Complete Phase {phase} AI",
                f"可可利亚完整形态{phase}阶段AI",
                "Finite source-ordered boss action graph.",
            )
        )
    ability_metadata = {
        "chill": ("Chill of Bone-Piercing Coagulation", "刺骨凝血的寒芒", "225% Ice single-target strike."),
        "omen": ("Omen of Everlasting Freeze", "漫长冰期的预兆", "Summons two Ice Edges."),
        "hoarfrost": ("Hoarfrost of Eternal Isolation", "永囚于此的白霜", "150% Ice strike with Freeze."),
        "wrath": ("Wrath of Winterland Saints", "雪国圣徒的烈怒", "Charges the next Punishment with 30% damage."),
        "punishment": ("Punishment of Endless Winter", "无尽长冬的绝罚", "375% Ice damage to all opponents."),
        "reverberating": ("Reverberating Ice", "碎冰震荡", "Shatters a Frozen target and adjacent units."),
        "omen-bronya": ("Omen of Everlasting Freeze — Reinforced", "漫长冰期的预兆·增援", "Summons Ice Edges and Bronya."),
        "doomsday": ("Doomsday Ice Cascade", "末日冰瀑", "Phase-three immediate Ice cascade and Freeze."),
        "icy-wind": ("Icy Wind", "冰风", "Ice Edge 150% Ice damage to all opponents."),
        "bronya-cycle": ("Bronya Complete Combat Cycle", "布洛妮娅完整作战循环", "Suppressive Fire followed by Combat Redeployment."),
    }
    for key, ability_id in abilities.items():
        name_en, name_zh_cn, summary = ability_metadata[key]
        identities.append(
            identity_s06(
                ability_id,
                f"enemy.cocolia-complete.ability.{key}",
                "Ability",
                name_en,
                name_zh_cn,
                summary,
            )
        )
    for key, linked_id in linked.items():
        name = "Bronya (Complete)" if key == "bronya" else "Ice Edge"
        identities.append(
            identity_s06(
                linked_id,
                f"unit.goal07.cocolia-complete.{key}",
                "CharacterForm",
                name,
                "布洛妮娅（完整）" if key == "bronya" else "冰锋",
                "Owner-scaled targetable Cocolia summon.",
            )
        )
    for key, selector_id_ in selectors.items():
        identities.append(
            identity_s06(
                selector_id_,
                f"selector.goal07.cocolia-complete.{key}",
                "Selector",
                f"Cocolia {key} Selector",
                f"可可利亚{key}选择器",
                "S06 battle selector.",
            )
        )
    for key, effect_id in effects.items():
        identities.append(
            identity_s06(
                effect_id,
                f"effect.goal07.cocolia-complete.{key}",
                "Effect",
                f"Cocolia {key}",
                f"可可利亚{key}",
                "Executable Cocolia battle effect.",
            )
        )
    for key, modifier_id in modifiers.items():
        identities.append(
            identity_s06(
                modifier_id,
                f"modifier.goal07.cocolia-complete.{key}",
                "Modifier",
                f"Cocolia {key} Modifier",
                f"可可利亚{key}调整器",
                "Effect-owned S06 stat modifier.",
            )
        )
    for key, rule_id in rules.items():
        identities.append(
            identity_s06(
                rule_id,
                f"rule.goal07.cocolia-complete.{key}",
                "Rule",
                f"Cocolia {key} Rule",
                f"可可利亚{key}规则",
                "Freeze lifecycle rule for Intensifying Cold.",
            )
        )

    add("Selector", selector(selectors["actor"], "Actor", "SameSide"))
    add("Selector", selector(selectors["owner"], "Owner", "SameSide"))
    add("Selector", selector(selectors["applier"], "Applier", "SameSide"))
    add("Selector", selector(selectors["current-subject"], "CurrentSubject", "AnySide"))
    add("Selector", selector(selectors["primary-target"], "PrimaryTarget", "OpposingSide"))
    opposing_random = selector(
        selectors["opposing-random"], "Actor", "OpposingSide", choice="RngUniform"
    )
    opposing_random["rng_purpose_key"] = "damage-target"
    add("Selector", opposing_random)
    add(
        "Selector",
        selector(
            selectors["opposing-all"],
            "Actor",
            "OpposingSide",
            minimum=1,
            maximum=8,
            choice="All",
        ),
    )
    for name, frozen, excludes_primary in [
        ("adjacent", False, True),
        ("adjacent-frozen", True, True),
    ]:
        add(
            "Selector",
            selector(
                selectors[name],
                "Actor",
                "OpposingSide",
                minimum=0,
                maximum=2,
                empty="NoOp",
                choice="PrimaryPlusAdjacent",
            ),
        )
        sequence = 1
        if excludes_primary:
            add(
                "SelectorPredicate",
                {
                    "selector_id": selectors[name],
                    "sequence": sequence,
                    "predicate": json_cell(
                        "Excludes", excluded_selector_id=selectors["primary-target"]
                    ),
                },
            )
            sequence += 1
        if frozen:
            add(
                "SelectorPredicate",
                {
                    "selector_id": selectors[name],
                    "sequence": sequence,
                    "predicate": json_cell("HasEffect", effect_id=effects["freeze"]),
                },
            )
    add("Selector", selector(selectors["frozen-primary"], "Actor", "OpposingSide"))
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["frozen-primary"],
            "sequence": 1,
            "predicate": json_cell("HasEffect", effect_id=effects["freeze"]),
        },
    )
    add(
        "Selector",
        selector(
            selectors["frozen-all"],
            "Actor",
            "OpposingSide",
            minimum=0,
            maximum=8,
            empty="NoOp",
            choice="All",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["frozen-all"],
            "sequence": 1,
            "predicate": json_cell("HasEffect", effect_id=effects["freeze"]),
        },
    )
    nonfrozen_random = selector(
        selectors["nonfrozen-random"],
        "Actor",
        "OpposingSide",
        minimum=0,
        maximum=1,
        empty="NoOp",
        choice="RngUniform",
    )
    nonfrozen_random["rng_purpose_key"] = "damage-target"
    add("Selector", nonfrozen_random)
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["nonfrozen-random"],
            "sequence": 1,
            "predicate": json_cell(
                "Excludes", excluded_selector_id=selectors["frozen-all"]
            ),
        },
    )
    add(
        "Selector",
        selector(
            selectors["actor-summons"],
            "Actor",
            "SameSide",
            minimum=0,
            maximum=3,
            empty="NoOp",
            choice="All",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["actor-summons"],
            "sequence": 1,
            "predicate": json_cell(
                "OwnedBy", owner_selector_id=selectors["actor"]
            ),
        },
    )
    add(
        "Selector",
        selector(
            selectors["event-target"],
            "CurrentSubject",
            "AnySide",
            life="Any",
            presence="Any",
        ),
    )

    def expr(name: str, kind: str, node: str) -> int:
        nonlocal next_expression
        id_ = next_expression
        next_expression += 1
        add(
            "ValueExpression",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s06.expression.{name}",
                "result_kind": kind,
                "node": node,
            },
        )
        return id_

    def multiply(name: str, left: int, right: int) -> int:
        return expr(
            name,
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedMultiply",
                left_expression_id=left,
                right_expression_id=right,
                rounding="NearestTiesAway",
            ),
        )

    actor_atk = expr(
        "actor-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    applier_atk = expr(
        "applier-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["applier"],
            stat="Atk",
            formula_purpose="Dot",
        ),
    )
    ratios: dict[str, int] = {}
    for name, value in [
        ("negative-one", "-1"),
        ("one-tenth", "0.1"),
        ("one-fifth", "0.2"),
        ("three-tenths", "0.3"),
        ("one-half", "0.5"),
        ("one", "1"),
        ("one-point-one-two-five", "1.125"),
        ("one-point-five", "1.5"),
        ("two-point-two-five", "2.25"),
        ("three-point-seven-five", "3.75"),
    ]:
        ratios[name] = expr(
            f"scalar-{name}",
            "Scalar",
            json_cell("ScalarLiteral", value_decimal=value),
        )
    integer_one = expr("integer-one", "Integer", json_cell("IntegerLiteral", value=1))
    integer_negative_one = expr(
        "integer-negative-one", "Integer", json_cell("IntegerLiteral", value=-1)
    )
    intensifying_stacks = expr(
        "intensifying-cold-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["owner"],
            effect_id=effects["intensifying-cold"],
        ),
    )
    intensifying_scalar = expr(
        "intensifying-cold-stacks-scalar",
        "Scalar",
        json_cell(
            "Convert",
            operand_expression_id=intensifying_stacks,
            target_kind="Scalar",
            rounding="NearestTiesAway",
        ),
    )
    intensifying_value = multiply(
        "intensifying-cold-damage-boost",
        intensifying_scalar,
        ratios["one-tenth"],
    )
    damage = {
        "chill": multiply("chill-damage", actor_atk, ratios["two-point-two-five"]),
        "hoarfrost": multiply(
            "hoarfrost-damage", actor_atk, ratios["one-point-five"]
        ),
        "freeze": multiply(
            "freeze-delayed-damage", applier_atk, ratios["one-point-one-two-five"]
        ),
        "punishment": multiply(
            "punishment-damage", actor_atk, ratios["three-point-seven-five"]
        ),
        "reverberating-primary": multiply(
            "reverberating-primary-damage",
            actor_atk,
            ratios["three-point-seven-five"],
        ),
        "reverberating-adjacent": multiply(
            "reverberating-adjacent-damage",
            actor_atk,
            ratios["one-point-five"],
        ),
        "icy-wind": multiply("icy-wind-damage", actor_atk, ratios["one-point-five"]),
        "bronya": multiply(
            "bronya-suppressive-fire-damage",
            actor_atk,
            ratios["three-point-seven-five"],
        ),
        "doomsday": multiply(
            "doomsday-ice-cascade-damage", actor_atk, ratios["one-point-five"]
        ),
    }

    add(
        "ConditionExpression",
        {
            "id": conditions["always"],
            "stable_key": "goal07.enemy.s06.condition.always",
            "node": json_cell("Constant", value=True),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["adjacent-frozen"],
            "stable_key": "goal07.enemy.s06.condition.adjacent-frozen",
            "node": json_cell(
                "SelectorCardinality",
                selector_id=selectors["adjacent-frozen"],
                minimum_count=1,
                maximum_count=2,
            ),
        },
    )

    def new_operation(
        name: str,
        payload: str,
        target: int | None = None,
        empty: str = "Fault",
    ) -> int:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        row = operation(id_, name, payload, target, empty)
        row["stable_key"] = f"goal07.enemy.s06.operation.{name}"
        add("Operation", row)
        return id_

    def operation_step(operation_id: int) -> str:
        return json_cell("Operation", operation_id=operation_id)

    def make_program(name: str, steps: list[str]) -> int:
        nonlocal next_program
        id_ = next_program
        next_program += 1
        identities.append(
            identity_s06(
                id_,
                f"program.goal07.cocolia-complete.{name}",
                "Program",
                f"Cocolia {name} Program",
                f"可可利亚{name}程序",
                "Ordered Rule IR program for the S06 enemy.",
            )
        )
        add("Program", {"id": id_, "domain": "Battle"})
        for sequence, step in enumerate(steps, start=1):
            add("ProgramStep", {"program_id": id_, "sequence": sequence, "step": step})
        return id_

    def damage_op(
        name: str, amount: int, target: int, element: str = "Ice", empty: str = "Fault"
    ) -> int:
        return new_operation(
            name,
            json_cell(
                "Damage",
                amount_expression_id=amount,
                damage_class="Ordinary",
                element=element,
                can_crit=True,
            ),
            target,
            empty,
        )

    def apply_effect_op(
        name: str,
        effect_id: int,
        target: int,
        *,
        chance: int | None = None,
        stacks: int | None = None,
        empty: str = "Fault",
    ) -> int:
        return new_operation(
            name,
            json_cell(
                "ApplyEffect",
                effect_id=effect_id,
                stacks_expression_id=stacks,
                chance_policy="Resistible" if chance is not None else "Guaranteed",
                base_chance_expression_id=chance,
                rng_purpose_key="effect-application" if chance is not None else None,
            ),
            target,
            empty,
        )

    def remove_effect_op(name: str, effect_id: int, target: int) -> int:
        return new_operation(
            name, json_cell("RemoveEffect", effect_id=effect_id), target, "NoOp"
        )

    def summon_steps(include_bronya: bool) -> list[str]:
        keys = ["ice-edge-left", "ice-edge-right"]
        if include_bronya:
            keys.append("bronya")
        return [
            operation_step(
                new_operation(
                    f"summon-{key}{'-reinforced' if include_bronya else ''}",
                    json_cell(
                        "Summon",
                        unit_definition_identity_id=linked[key],
                        owner_selector_id=selectors["actor"],
                    ),
                )
            )
            for key in keys
        ]

    repeat_primary = make_program(
        "reverberating-repeat-primary",
        [
            operation_step(
                damage_op(
                    "reverberating-repeat-primary-damage",
                    damage["reverberating-primary"],
                    selectors["primary-target"],
                )
            )
        ],
    )
    ability_programs = {
        abilities["chill"]: make_program(
            "chill",
            [
                operation_step(
                    damage_op("chill-damage", damage["chill"], selectors["primary-target"])
                )
            ],
        ),
        abilities["omen"]: make_program("omen", summon_steps(False)),
        abilities["omen-bronya"]: make_program(
            "omen-reinforced", summon_steps(True)
        ),
        abilities["hoarfrost"]: make_program(
            "hoarfrost",
            [
                operation_step(
                    damage_op(
                        "hoarfrost-damage",
                        damage["hoarfrost"],
                        selectors["primary-target"],
                    )
                ),
                operation_step(
                    apply_effect_op(
                        "hoarfrost-freeze",
                        effects["freeze"],
                        selectors["primary-target"],
                        chance=ratios["one"],
                    )
                ),
            ],
        ),
        abilities["wrath"]: make_program(
            "wrath",
            [
                operation_step(
                    apply_effect_op(
                        "wrath-charging", effects["charging"], selectors["actor"]
                    )
                )
            ],
        ),
        abilities["punishment"]: make_program(
            "punishment",
            [
                operation_step(
                    damage_op(
                        "punishment-damage",
                        damage["punishment"],
                        selectors["opposing-all"],
                    )
                ),
                operation_step(
                    remove_effect_op(
                        "punishment-clear-charging",
                        effects["charging"],
                        selectors["actor"],
                    )
                ),
            ],
        ),
        abilities["reverberating"]: make_program(
            "reverberating",
            [
                operation_step(
                    damage_op(
                        "reverberating-primary-damage",
                        damage["reverberating-primary"],
                        selectors["primary-target"],
                    )
                ),
                operation_step(
                    damage_op(
                        "reverberating-adjacent-damage",
                        damage["reverberating-adjacent"],
                        selectors["adjacent"],
                        empty="NoOp",
                    )
                ),
                json_cell(
                    "If",
                    condition_id=conditions["adjacent-frozen"],
                    then_program_id=repeat_primary,
                    else_program_id=None,
                ),
                operation_step(
                    remove_effect_op(
                        "reverberating-remove-primary-freeze",
                        effects["freeze"],
                        selectors["primary-target"],
                    )
                ),
            ],
        ),
        abilities["doomsday"]: make_program(
            "doomsday",
            [
                operation_step(
                    damage_op(
                        "doomsday-damage",
                        damage["doomsday"],
                        selectors["opposing-all"],
                    )
                ),
                operation_step(
                    apply_effect_op(
                        "doomsday-freeze",
                        effects["freeze"],
                        selectors["nonfrozen-random"],
                        chance=ratios["one"],
                        empty="NoOp",
                    )
                ),
            ],
        ),
        abilities["icy-wind"]: make_program(
            "icy-wind",
            [
                operation_step(
                    damage_op(
                        "icy-wind-damage",
                        damage["icy-wind"],
                        selectors["opposing-all"],
                    )
                )
            ],
        ),
        abilities["bronya-cycle"]: make_program(
            "bronya-cycle",
            [
                operation_step(
                    damage_op(
                        "bronya-suppressive-fire",
                        damage["bronya"],
                        selectors["primary-target"],
                        element="Wind",
                    )
                ),
                operation_step(
                    new_operation(
                        "bronya-delay-target",
                        json_cell(
                            "DelayAction", amount_expression_id=ratios["one-half"]
                        ),
                        selectors["primary-target"],
                    )
                ),
                operation_step(
                    new_operation(
                        "bronya-redeploy-owner",
                        json_cell("AdvanceAction", amount_expression_id=ratios["one"]),
                        selectors["owner"],
                    )
                ),
                operation_step(
                    apply_effect_op(
                        "bronya-redeployment-damage-boost",
                        effects["redeployment"],
                        selectors["owner"],
                    )
                ),
            ],
        ),
    }

    target_patterns = {
        "chill": "SingleTarget",
        "omen": "None",
        "hoarfrost": "SingleTarget",
        "wrath": "None",
        "punishment": "Aoe",
        "reverberating": "Blast",
        "omen-bronya": "None",
        "doomsday": "Aoe",
        "icy-wind": "Aoe",
        "bronya-cycle": "SingleTarget",
    }
    for key, ability_id in abilities.items():
        linked_ability = key in {"icy-wind", "bronya-cycle"}
        add(
            "Ability",
            {
                "id": ability_id,
                "kind": "Summon" if linked_ability else "Skill",
                "target_pattern": target_patterns[key],
                "retarget_policy": "CancelRemaining",
                "level_cap": 1,
                "cooldown_actions": 1,
                "semantic_tags_mask": 5 if target_patterns[key] != "None" else 4,
            },
        )
        add(
            "AbilityPhase",
            {
                "ability_id": ability_id,
                "sequence": 1,
                "kind": "Resolved",
                "program_identity_id": ability_programs[ability_id],
            },
        )
        if not linked_ability:
            add(
                "EnemyAbility",
                {
                    "id": ability_id,
                    "telegraph": "Charge" if key == "wrath" else "None",
                    "cooldown_actions": 1,
                    "initial_cooldown_actions": 0,
                    "charge_actions": 1 if key == "wrath" else 0,
                    "ai_tag": key,
                },
            )

    effect_definitions = {
        "freeze": (
            "Control",
            "CleanseableControl",
            1,
            integer_one,
            "TargetTurnStart",
            "TurnStart",
            "Refresh",
            damage["freeze"],
            "Ice",
        ),
        "charging": (
            "Buff",
            "NonDispellable",
            1,
            integer_one,
            "OwnerTurnEnd",
            "None",
            "Refresh",
            None,
            None,
        ),
        "intensifying-cold": (
            "Buff",
            "NonDispellable",
            8,
            None,
            "Permanent",
            "None",
            "RefreshAndAddStacks",
            intensifying_value,
            None,
        ),
        "redeployment": (
            "Buff",
            "DispellableBuff",
            1,
            integer_one,
            "OwnerTurnEnd",
            "None",
            "Refresh",
            None,
            None,
        ),
    }
    for key, effect_id in effects.items():
        category, dispel, limit, duration, clock, tick, policy, magnitude, dot = (
            effect_definitions[key]
        )
        add(
            "Effect",
            {
                "id": effect_id,
                "category": category,
                "dispel_category": dispel,
                "stack_limit": limit,
                "duration_expression_id": duration,
                "duration_clock": clock,
                "tick_phase": tick,
                "stack_policy": policy,
                "magnitude_comparator_expression_id": magnitude,
                "dot_element": dot,
                "snapshot_policy": "OnApplication",
                "teardown_policy": "RemoveWithOwner",
                "application_priority": 0,
            },
        )
    for key, tags in {
        "freeze": ["freeze", "blocks-normal-action"],
        "charging": ["charging", "punishment-damage-boost"],
        "intensifying-cold": ["intensifying-cold"],
        "redeployment": ["combat-redeployment"],
    }.items():
        for sequence, tag in enumerate(tags, start=1):
            add("EffectTag", {"effect_id": effects[key], "sequence": sequence, "tag": tag})
    for key, group_id in modifier_groups.items():
        add(
            "ModifierStackingGroup",
            {
                "id": group_id,
                "stable_key": f"goal07.enemy.s06.{key}",
                "aggregation": "Sum",
            },
        )
    for key, value in [
        ("charging", ratios["three-tenths"]),
        ("intensifying-cold", intensifying_value),
        ("redeployment", ratios["one-fifth"]),
    ]:
        add(
            "ModifierDefinition",
            {
                "id": modifiers[key],
                "source_effect_id": effects[key],
                "owner_selector_id": selectors["owner"],
                "subject_selector_id": selectors["owner"],
                "stat": "Atk",
                "formula_stage": "DamageBoost",
                "formula_purpose": "OrdinaryDamage",
                "value_expression_id": value,
                "value_domain": "Ratio",
                "stacking_group_id": modifier_groups[key],
                "priority": 0,
                "cap_formula_stage": "DamageBoost",
                "snapshot_policy": "Dynamic",
                "duration_scope": "Turn",
            },
        )
        add(
            "EffectModifierBinding",
            {"effect_id": effects[key], "sequence": 1, "modifier_id": modifiers[key]},
        )

    intensify_program = make_program(
        "freeze-applied-intensify",
        [
            operation_step(
                apply_effect_op(
                    "freeze-applied-intensify-owner",
                    effects["intensifying-cold"],
                    selectors["owner"],
                    stacks=integer_one,
                )
            ),
            operation_step(
                apply_effect_op(
                    "freeze-applied-intensify-summons",
                    effects["intensifying-cold"],
                    selectors["actor-summons"],
                    stacks=integer_one,
                    empty="NoOp",
                )
            ),
        ],
    )
    weaken_program = make_program(
        "freeze-removed-weaken",
        [
            operation_step(
                new_operation(
                    "freeze-removed-weaken-owner",
                    json_cell(
                        "ModifyEffect",
                        effect_id=effects["intensifying-cold"],
                        stack_delta_expression_id=integer_negative_one,
                    ),
                    selectors["owner"],
                    "NoOp",
                )
            ),
            operation_step(
                new_operation(
                    "freeze-removed-weaken-summons",
                    json_cell(
                        "ModifyEffect",
                        effect_id=effects["intensifying-cold"],
                        stack_delta_expression_id=integer_negative_one,
                    ),
                    selectors["actor-summons"],
                    "NoOp",
                )
            ),
        ],
    )
    for key, point, program_id in [
        ("freeze-applied", "Applied", intensify_program),
        ("freeze-removed", "Removed", weaken_program),
    ]:
        add(
            "RuleDefinition",
            {
                "id": rules[key],
                "domain": "Battle",
                "source_definition_identity_id": effects["freeze"],
                "source_class": "Effect",
                "source_digest_sha256": sha256_text(f"goal07-s06-{key}-v1"),
            },
        )
        add(
            "EventFilter",
            {
                "id": filters[key],
                "stable_key": f"goal07.enemy.s06.filter.{key}",
                "source_definition_identity_id": effects["freeze"],
                "source_class": "Effect",
                "owner_selector_id": selectors["owner"],
                "target_selector_id": selectors["event-target"],
                "cause_ancestry": "Any",
            },
        )
        add(
            "RuleTrigger",
            {
                "id": BASE + 570 + len(rows.get("RuleTrigger", [])),
                "stable_key": f"goal07.enemy.s06.trigger.{key}",
                "rule_id": rules[key],
                "sequence": 1,
                "event": json_cell("Effect", point=point),
                "phase": "AfterEvent",
                "filter_id": filters[key],
                "condition_id": conditions["always"],
                "once_scope": "Battle",
                "priority": 0,
                "program_id": program_id,
            },
        )
        add(
            "EffectRuleBinding",
            {
                "effect_id": effects["freeze"],
                "sequence": 1 if key == "freeze-applied" else 2,
                "rule_id": rules[key],
            },
        )

    phase_sequences = [
        [abilities["chill"], abilities["omen"], abilities["hoarfrost"]],
        [
            abilities["omen-bronya"],
            abilities["chill"],
            abilities["hoarfrost"],
            abilities["reverberating"],
            abilities["wrath"],
            abilities["punishment"],
        ],
        [
            abilities["doomsday"],
            abilities["omen"],
            abilities["hoarfrost"],
            abilities["reverberating"],
            abilities["wrath"],
            abilities["punishment"],
        ],
    ]
    target_selectors = {
        abilities["chill"]: selectors["opposing-random"],
        abilities["omen"]: selectors["actor"],
        abilities["omen-bronya"]: selectors["actor"],
        abilities["hoarfrost"]: selectors["opposing-random"],
        abilities["reverberating"]: selectors["frozen-primary"],
        abilities["wrath"]: selectors["actor"],
        abilities["punishment"]: selectors["opposing-all"],
        abilities["doomsday"]: selectors["opposing-all"],
    }
    next_state = BASE + 701
    next_candidate = BASE + 801
    next_transition = BASE + 901
    for phase_index, sequence in enumerate(phase_sequences):
        state_ids = list(range(next_state, next_state + len(sequence)))
        next_state += len(sequence)
        add(
            "AiGraph",
            {
                "id": graphs[phase_index],
                "initial_state_id": state_ids[0],
                "automatic_transition_budget": 8,
            },
        )
        for offset, (state_id, ability_id) in enumerate(zip(state_ids, sequence)):
            add(
                "AiState",
                {
                    "id": state_id,
                    "stable_key": (
                        f"goal07.enemy.s06.ai.phase-{phase_index + 1}."
                        f"state-{offset + 1}"
                    ),
                    "graph_id": graphs[phase_index],
                    "mandatory_fallback_ability_id": abilities["chill"],
                    "turn_counter_reset": offset == 0,
                },
            )
            add(
                "AiCandidate",
                {
                    "id": next_candidate,
                    "stable_key": (
                        f"goal07.enemy.s06.ai.phase-{phase_index + 1}."
                        f"candidate-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "ability_id": ability_id,
                    "condition_id": conditions["always"],
                    "target_selector_id": target_selectors[ability_id],
                    "priority": 0,
                    "selection": "FirstLegal",
                    "no_target_fallback": "UseFallbackAbility",
                    "fallback_ability_id": abilities["chill"],
                },
            )
            next_candidate += 1
            add(
                "AiTransition",
                {
                    "id": next_transition,
                    "stable_key": (
                        f"goal07.enemy.s06.ai.phase-{phase_index + 1}."
                        f"transition-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "target_state_id": state_ids[(offset + 1) % len(state_ids)],
                    "condition_id": conditions["always"],
                    "priority": 0,
                    "timing": "AfterAction",
                },
            )
            next_transition += 1

    linked_specs = {
        "ice-edge-left": (1, "0.09375", "144", abilities["icy-wind"]),
        "ice-edge-right": (5, "0.09375", "144", abilities["icy-wind"]),
        "bronya": (4, "0.75", "143", abilities["bronya-cycle"]),
    }
    for key, linked_id in linked.items():
        formation, hp_ratio, spd, action_ability = linked_specs[key]
        add(
            "LinkedUnitDefinition",
            {
                "id": linked_id,
                "source_definition_identity_id": linked_id,
                "kind": "Summon",
                "presence": "Present",
                "ability_ids": str(action_ability),
                "action_ability_id": action_ability,
                "formation_index": formation,
                "initial_gauge_decimal": "10000",
                "hp_owner_ratio_decimal": hp_ratio,
                "hp_flat_decimal": "0",
                "atk_owner_ratio_decimal": "1",
                "atk_flat_decimal": "0",
                "def_owner_ratio_decimal": "1",
                "def_flat_decimal": "0",
                "spd_owner_ratio_decimal": "0",
                "spd_flat_decimal": spd,
                "owner_defeat_policy": "Depart",
                "owner_departure_policy": "Depart",
                "wave_policy": "Depart",
                "combatant_digest_sha256": sha256_text(
                    f"goal07-s06-linked-{key}-v1"
                ),
            },
        )

    add(
        "EnemyTemplate",
        {
            "id": template,
            "rank": "Boss",
            "base_aggro_decimal": "100",
            "default_ai_graph_id": graphs[0],
        },
    )
    add(
        "EnemyVariant",
        {
            "id": variant,
            "template_id": template,
            "ai_graph_id": graphs[0],
            "mechanically_distinct_key": VARIANT_KEY,
        },
    )
    for level in anchor["levels"]:
        add(
            "EnemyStat",
            {
                "variant_id": variant,
                "level": level["authored_level"],
                "difficulty_key": "standard-universe-v1",
                "hp_decimal": level["base_hp"],
                "atk_decimal": level["base_atk"],
                "def_decimal": level["base_def"],
                "spd_decimal": level["base_spd"],
                "effect_hit_rate_decimal": level["effect_hit_rate"],
                "effect_resistance_decimal": level["effect_resistance"],
                "crit_damage_decimal": "0.2",
            },
        )
    for sequence, weakness in enumerate(["Fire", "Lightning", "Quantum"], start=1):
        add(
            "EnemyWeakness",
            {"variant_id": variant, "sequence": sequence, "element": weakness},
        )
    for element, value in [
        ("Ice", "0.6"),
        ("Imaginary", "0.2"),
        ("Physical", "0.4"),
        ("Wind", "0.4"),
    ]:
        add(
            "EnemyResistance",
            {"variant_id": variant, "element": element, "value_decimal": value},
        )
    for category in ["STAT_CTRL_Confine", "STAT_CTRL_Frozen"]:
        add(
            "EnemyDebuffResistance",
            {
                "variant_id": variant,
                "category_key": category,
                "value_decimal": "1",
            },
        )
    add(
        "EnemyToughnessLayer",
        {
            "variant_id": variant,
            "sequence": 1,
            "layer_key": "ordinary",
            "kind": "Ordinary",
            "maximum_decimal": "120",
            "recovery_ratio_decimal": "1",
            "active_at_start": True,
        },
    )
    for sequence, ability_id in enumerate(
        [
            abilities["chill"],
            abilities["omen"],
            abilities["punishment"],
            abilities["hoarfrost"],
            abilities["wrath"],
            abilities["reverberating"],
            abilities["omen-bronya"],
            abilities["doomsday"],
        ],
        start=1,
    ):
        add(
            "EnemyVariantAbility",
            {
                "variant_id": variant,
                "sequence": sequence,
                "ability_id": ability_id,
            },
        )
    for sequence, graph in enumerate(graphs, start=1):
        add(
            "EnemyPhase",
            {
                "id": BASE + 600 + sequence,
                "stable_key": f"goal07.enemy.s06.phase-{sequence}",
                "variant_id": variant,
                "sequence": sequence,
                "entry_condition_id": conditions["always"],
                "exit_condition_id": conditions["always"],
                "replacement_priority": sequence,
                "ai_graph_id": graph,
                "targetable": True,
                "transition_model": "TransformSameUnit",
                "hp_carry": "Reset",
                "action_gauge_carry": "Reset",
                "effect_carry": "Clear",
                "toughness_carry": "Reset",
                "summon_carry": "Clear",
            },
        )

    anchor_digest = sha256_bytes(anchor_path(PARTITION).read_bytes())
    add(
        "SourceRecord",
        {
            "id": SOURCE_RECORD_ID,
            "stable_key": "source.hsr-wiki.cocolia-complete.2026-07-29",
            "category": "CommunityMaintained",
            "publisher": anchor["source"]["publisher"],
            "url": anchor["source"]["url"],
            "accessed_on": anchor["source"]["accessed_on"],
            "applicable_game_version": anchor["source"]["game_version"],
            "confidence": "SecondaryVersionSensitiveCrossCheck",
            "evidence_sha256": anchor_digest,
            "usage_note": (
                "Exact public per-level values are committed with retained "
                "structured AI and ability source hashes."
            ),
        },
    )
    add(
        "EvidenceRecord",
        {
            "id": EVIDENCE_RECORD_ID,
            "stable_key": "evidence.goal07.enemy.s06.numeric-anchors",
            "kind": "SourcePayload",
            "source_record_id": SOURCE_RECORD_ID,
            "sha256": anchor_digest,
            "note": "Committed exact public per-level numeric anchors for Goal 07 S06.",
        },
    )
    for item in identities:
        add("ContentIdentity", item)
        add(
            "ContentEvidenceBinding",
            {
                "content_id": item["id"],
                "sequence": 1,
                "fact_key": f"goal07.s06.executable:{item['stable_key']}",
                "source_record_id": 1,
                "evidence_record_id": 3,
                "quality": "ExactStructured",
                "mechanism_quality": "ExactStructured",
            },
        )
    add(
        "ContentEvidenceBinding",
        {
            "content_id": variant,
            "sequence": 2,
            "fact_key": "goal07.s06.public-level-stats",
            "source_record_id": SOURCE_RECORD_ID,
            "evidence_record_id": EVIDENCE_RECORD_ID,
            "quality": "ExactStructured",
            "mechanism_quality": "ExactStructured",
        },
    )
    for table_rows in rows.values():
        table_rows.sort(
            key=lambda row: json.dumps(
                row, ensure_ascii=False, sort_keys=True, default=str
            )
        )
    return rows


def owned_rows_s07() -> dict[str, list[dict[str, Any]]]:
    anchor = json.loads(anchor_path(PARTITION).read_text(encoding="utf-8"))
    manifest = json.loads(PARTITIONS.read_text(encoding="utf-8"))
    assigned = next(item for item in manifest["partitions"] if item["id"] == PARTITION)
    if assigned["enemy_variant_ids"] != [VARIANT_KEY]:
        raise ValueError("S07 frozen enemy assignment changed")

    variant = BASE + 1
    template = BASE + 2
    graphs = [BASE + 10, BASE + 11, BASE + 12]
    abilities = {
        "fist": BASE + 101,
        "smite": BASE + 102,
        "garrison": BASE + 103,
        "frigid": BASE + 104,
        "besiege": BASE + 105,
        "frigid-rapid": BASE + 106,
        "besiege-rapid": BASE + 107,
        "counter": BASE + 108,
        "smite-rapid": BASE + 112,
        "besiege-end": BASE + 113,
        "soldier-charge": BASE + 121,
        "cannoneer-barrage": BASE + 122,
        "lieutenant-pierce": BASE + 123,
        "coordinated-strike": BASE + 124,
    }
    linked = {
        "phase-1-soldier": BASE + 201,
        "phase-1-cannoneer": BASE + 202,
        "phase-2-cannoneer-left": BASE + 203,
        "phase-2-cannoneer-right": BASE + 204,
        "phase-3-lieutenant-left": BASE + 205,
        "phase-3-lieutenant-right": BASE + 206,
    }
    selectors = {
        "actor": BASE + 401,
        "owner": BASE + 402,
        "current-subject": BASE + 403,
        "primary-target": BASE + 404,
        "opposing-random": BASE + 405,
        "opposing-all": BASE + 406,
        "same-all": BASE + 407,
        "actor-summons": BASE + 408,
        "adjacent": BASE + 409,
    }
    effects = {
        "frigid-escalation": BASE + 501,
        "collective-shield": BASE + 502,
        "counter": BASE + 503,
        "lock-on": BASE + 504,
        "def-down": BASE + 505,
    }
    modifiers = {
        "frigid-escalation": BASE + 521,
        "def-down": BASE + 522,
    }
    modifier_groups = {
        "frigid-escalation": BASE + 531,
        "def-down": BASE + 532,
    }
    rules = {"counter": BASE + 541}
    filters = {"counter": BASE + 551}
    conditions = {
        "always": BASE + 561,
        "shielded": BASE + 562,
        "unshielded": BASE + 563,
    }
    rows: dict[str, list[dict[str, Any]]] = {}
    identities: list[dict[str, Any]] = []
    next_program = BASE + 301
    next_operation = BASE + 1_001
    next_expression = BASE + 1_101

    def add(table: str, row: dict[str, Any]) -> None:
        rows.setdefault(table, []).append(row)

    def identity_s07(
        id_: int,
        stable_key: str,
        kind: str,
        name_en: str,
        name_zh_cn: str,
        summary: str,
        sources: str = "1",
    ) -> dict[str, Any]:
        row = identity(id_, stable_key, kind, name_en, name_zh_cn, summary, sources)
        row["summary_zh_cn"] = "Goal 07 S07 来源绑定的杰帕德（完整）可执行定义。"
        row["game_version_introduced"] = "1.0"
        return row

    identities.extend(
        [
            identity_s07(
                variant,
                VARIANT_KEY,
                "EnemyVariant",
                "Gepard (Complete)",
                "杰帕德（完整）",
                "Exact materialization variant used by frozen World 3 bindings.",
                "1|9",
            ),
            identity_s07(
                template,
                "enemy.gepard-complete.littleboss",
                "Enemy",
                "Gepard (Complete) Template",
                "杰帕德（完整）模板",
                "Version 4.4 boss template retained from source monster 1004022.",
            ),
        ]
    )
    for phase, graph in enumerate(graphs, start=1):
        identities.append(
            identity_s07(
                graph,
                f"ai.goal07.gepard-complete.phase-{phase}",
                "AiGraph",
                f"Gepard Complete Phase {phase} AI",
                f"杰帕德完整形态{phase}阶段AI",
                "Finite source-ordered boss action graph.",
            )
        )
    ability_metadata = {
        "fist": ("Fist of Conviction", "一意之拳", "300% Ice single-target strike."),
        "smite": ("Smite of Frost", "霜之惩击", "350% Ice single-target strike."),
        "garrison": (
            "Garrison Aura Field",
            "戍卫气场",
            "Collective allied shield and Toughness protection.",
        ),
        "frigid": (
            "Frigid Waterfall",
            "极寒瀑流",
            "380% Ice area strike with stacking 12% damage escalation.",
        ),
        "besiege": (
            "Besiege",
            "围攻",
            "Locks one opponent and orders all surviving summons to strike.",
        ),
        "frigid-rapid": (
            "Frigid Waterfall — Rapid",
            "极寒瀑流·连动",
            "Extra-action opener for the unshielded phase rotation.",
        ),
        "besiege-rapid": (
            "Besiege — Rapid",
            "围攻·连动",
            "Extra-action opener for phase two and three.",
        ),
        "counter": (
            "Tit for Tat",
            "以牙还牙",
            "One-turn, one-trigger Smite of Frost counter state.",
        ),
        "smite-rapid": (
            "Smite of Frost — Rapid",
            "霜之惩击·连动",
            "Extra-action opener for the shielded phase rotation.",
        ),
        "besiege-end": (
            "Besiege — Rotation End",
            "围攻·轮转终段",
            "No-extra-action Besiege at the end of the unshielded rotation.",
        ),
        "soldier-charge": (
            "Tireless Charge",
            "不倦冲锋",
            "300% Physical strike with a resistible 50% DEF reduction.",
        ),
        "cannoneer-barrage": (
            "Barrage",
            "炮击",
            "130% Physical primary and 100% adjacent damage.",
        ),
        "lieutenant-pierce": (
            "Pierce",
            "穿刺",
            "420% Physical single-target strike.",
        ),
        "coordinated-strike": (
            "Besiege Coordinated Strike",
            "围攻协同打击",
            "Forced 300% Physical strike shared by Gepard-owned summons.",
        ),
    }
    for key, ability_id in abilities.items():
        name_en, name_zh_cn, summary = ability_metadata[key]
        identities.append(
            identity_s07(
                ability_id,
                f"enemy.gepard-complete.ability.{key}",
                "Ability",
                name_en,
                name_zh_cn,
                summary,
            )
        )
    linked_metadata = {
        "phase-1-soldier": ("Silvermane Soldier", "银鬃近卫"),
        "phase-1-cannoneer": ("Silvermane Cannoneer", "银鬃炮手"),
        "phase-2-cannoneer-left": ("Silvermane Cannoneer", "银鬃炮手"),
        "phase-2-cannoneer-right": ("Silvermane Cannoneer", "银鬃炮手"),
        "phase-3-lieutenant-left": (
            "Silvermane Lieutenant (Complete)",
            "银鬃尉官（完整）",
        ),
        "phase-3-lieutenant-right": (
            "Silvermane Lieutenant (Complete)",
            "银鬃尉官（完整）",
        ),
    }
    for key, linked_id in linked.items():
        name_en, name_zh_cn = linked_metadata[key]
        identities.append(
            identity_s07(
                linked_id,
                f"unit.goal07.gepard-complete.{key}",
                "CharacterForm",
                name_en,
                name_zh_cn,
                "Owner-scaled targetable Gepard phase summon.",
            )
        )
    for key, selector_id in selectors.items():
        identities.append(
            identity_s07(
                selector_id,
                f"selector.goal07.gepard-complete.{key}",
                "Selector",
                f"Gepard {key} Selector",
                f"杰帕德{key}选择器",
                "S07 battle selector.",
            )
        )
    for key, effect_id in effects.items():
        identities.append(
            identity_s07(
                effect_id,
                f"effect.goal07.gepard-complete.{key}",
                "Effect",
                f"Gepard {key} Effect",
                f"杰帕德{key}效果",
                "Executable Gepard battle effect.",
            )
        )
    for key, modifier_id in modifiers.items():
        identities.append(
            identity_s07(
                modifier_id,
                f"modifier.goal07.gepard-complete.{key}",
                "Modifier",
                f"Gepard {key} Modifier",
                f"杰帕德{key}调整器",
                "Effect-owned S07 stat modifier.",
            )
        )
    identities.append(
        identity_s07(
            rules["counter"],
            "rule.goal07.gepard-complete.tit-for-tat",
            "Rule",
            "Gepard Tit for Tat Counter",
            "杰帕德以牙还牙反击",
            "Damage-targeted, single-trigger queued Smite counter.",
        )
    )

    add("Selector", selector(selectors["actor"], "Actor", "SameSide"))
    add("Selector", selector(selectors["owner"], "Owner", "SameSide"))
    add(
        "Selector",
        selector(selectors["current-subject"], "CurrentSubject", "AnySide"),
    )
    add(
        "Selector",
        selector(selectors["primary-target"], "PrimaryTarget", "OpposingSide"),
    )
    opposing_random = selector(
        selectors["opposing-random"], "Actor", "OpposingSide", choice="RngUniform"
    )
    opposing_random["rng_purpose_key"] = "damage-target"
    add("Selector", opposing_random)
    add(
        "Selector",
        selector(
            selectors["opposing-all"],
            "Actor",
            "OpposingSide",
            minimum=1,
            maximum=8,
            choice="All",
        ),
    )
    add(
        "Selector",
        selector(
            selectors["same-all"],
            "Actor",
            "SameSide",
            minimum=1,
            maximum=8,
            choice="All",
        ),
    )
    add(
        "Selector",
        selector(
            selectors["actor-summons"],
            "Actor",
            "SameSide",
            presence="Linked",
            minimum=0,
            maximum=6,
            empty="NoOp",
            choice="All",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["actor-summons"],
            "sequence": 1,
            "predicate": json_cell(
                "OwnedBy", owner_selector_id=selectors["actor"]
            ),
        },
    )
    add(
        "Selector",
        selector(
            selectors["adjacent"],
            "Actor",
            "OpposingSide",
            minimum=0,
            maximum=2,
            empty="NoOp",
            choice="PrimaryPlusAdjacent",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["adjacent"],
            "sequence": 1,
            "predicate": json_cell(
                "Excludes", excluded_selector_id=selectors["primary-target"]
            ),
        },
    )

    def expr(name: str, kind: str, node: str) -> int:
        nonlocal next_expression
        id_ = next_expression
        next_expression += 1
        add(
            "ValueExpression",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s07.expression.{name}",
                "result_kind": kind,
                "node": node,
            },
        )
        return id_

    def multiply(name: str, left: int, right: int) -> int:
        return expr(
            name,
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedMultiply",
                left_expression_id=left,
                right_expression_id=right,
                rounding="NearestTiesAway",
            ),
        )

    actor_atk = expr(
        "actor-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    actor_hp = expr(
        "actor-hp",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Hp",
            formula_purpose="Shield",
        ),
    )
    actor_shield = expr(
        "actor-shield",
        "Scalar",
        json_cell(
            "QueryShield",
            subject_selector_id=selectors["actor"],
            observation="Current",
        ),
    )
    ratios: dict[str, int] = {}
    for name, value in [
        ("negative-half", "-0.5"),
        ("zero", "0"),
        ("twelve-percent", "0.12"),
        ("two-fifths", "0.4"),
        ("one", "1"),
        ("one-point-three", "1.3"),
        ("three", "3"),
        ("three-point-five", "3.5"),
        ("three-point-eight", "3.8"),
        ("four-point-two", "4.2"),
    ]:
        ratios[name] = expr(
            f"scalar-{name}",
            "Scalar",
            json_cell("ScalarLiteral", value_decimal=value),
        )
    integer_one = expr("integer-one", "Integer", json_cell("IntegerLiteral", value=1))
    escalation_stacks = expr(
        "frigid-escalation-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["owner"],
            effect_id=effects["frigid-escalation"],
        ),
    )
    escalation_scalar = expr(
        "frigid-escalation-stacks-scalar",
        "Scalar",
        json_cell(
            "Convert",
            operand_expression_id=escalation_stacks,
            target_kind="Scalar",
            rounding="NearestTiesAway",
        ),
    )
    escalation_value = multiply(
        "frigid-escalation-damage-boost",
        escalation_scalar,
        ratios["twelve-percent"],
    )
    damage = {
        "fist": multiply("fist-damage", actor_atk, ratios["three"]),
        "smite": multiply("smite-damage", actor_atk, ratios["three-point-five"]),
        "frigid": multiply(
            "frigid-waterfall-damage", actor_atk, ratios["three-point-eight"]
        ),
        "soldier": multiply(
            "soldier-charge-damage", actor_atk, ratios["three"]
        ),
        "cannoneer-primary": multiply(
            "cannoneer-primary-damage", actor_atk, ratios["one-point-three"]
        ),
        "cannoneer-adjacent": multiply(
            "cannoneer-adjacent-damage", actor_atk, ratios["one"]
        ),
        "lieutenant": multiply(
            "lieutenant-pierce-damage", actor_atk, ratios["four-point-two"]
        ),
        "coordinated": multiply(
            "coordinated-strike-damage", actor_atk, ratios["three"]
        ),
    }
    shield_amount = multiply(
        "collective-shield-allied-slice", actor_hp, ratios["two-fifths"]
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["always"],
            "stable_key": "goal07.enemy.s07.condition.always",
            "node": json_cell("Constant", value=True),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["shielded"],
            "stable_key": "goal07.enemy.s07.condition.shielded",
            "node": json_cell(
                "Compare",
                left_expression_id=actor_shield,
                comparison="Greater",
                right_expression_id=ratios["zero"],
            ),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["unshielded"],
            "stable_key": "goal07.enemy.s07.condition.unshielded",
            "node": json_cell(
                "Compare",
                left_expression_id=actor_shield,
                comparison="Equal",
                right_expression_id=ratios["zero"],
            ),
        },
    )

    def new_operation(
        name: str,
        payload: str,
        target: int | None = None,
        empty: str = "Fault",
    ) -> int:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        row = operation(id_, name, payload, target, empty)
        row["stable_key"] = f"goal07.enemy.s07.operation.{name}"
        add("Operation", row)
        return id_

    def operation_step(operation_id: int) -> str:
        return json_cell("Operation", operation_id=operation_id)

    def make_program(name: str, steps: list[str]) -> int:
        nonlocal next_program
        id_ = next_program
        next_program += 1
        identities.append(
            identity_s07(
                id_,
                f"program.goal07.gepard-complete.{name}",
                "Program",
                f"Gepard {name} Program",
                f"杰帕德{name}程序",
                "Ordered Rule IR program for the S07 enemy.",
            )
        )
        add("Program", {"id": id_, "domain": "Battle"})
        for sequence, step in enumerate(steps, start=1):
            add("ProgramStep", {"program_id": id_, "sequence": sequence, "step": step})
        return id_

    def damage_op(
        name: str,
        amount: int,
        target: int,
        *,
        element: str = "Ice",
        empty: str = "Fault",
    ) -> int:
        return new_operation(
            name,
            json_cell(
                "Damage",
                amount_expression_id=amount,
                damage_class="Ordinary",
                element=element,
                can_crit=True,
            ),
            target,
            empty,
        )

    def apply_effect_op(
        name: str,
        effect_id: int,
        target: int,
        *,
        chance: int | None = None,
        stacks: int | None = None,
        empty: str = "Fault",
    ) -> int:
        return new_operation(
            name,
            json_cell(
                "ApplyEffect",
                effect_id=effect_id,
                stacks_expression_id=stacks,
                chance_policy="Resistible" if chance is not None else "Guaranteed",
                base_chance_expression_id=chance,
                rng_purpose_key="effect-application" if chance is not None else None,
            ),
            target,
            empty,
        )

    def extra_turn_op(name: str) -> int:
        return new_operation(
            name,
            json_cell("GrantExtraTurn", actor_selector_id=selectors["actor"]),
        )

    def queue_action_op(
        name: str,
        ability_id: int,
        actor_selector_id: int,
        target_selector_id: int,
        boundary: str,
        *,
        empty: str = "Fault",
    ) -> int:
        return new_operation(
            name,
            json_cell(
                "QueueAction",
                ability_id=ability_id,
                actor_selector_id=actor_selector_id,
                priority=100,
                forced_use=True,
                reaction_boundary=boundary,
                owner_policy="Actor",
                payment_policy="Suppressed",
                payment_resource_key=None,
            ),
            target_selector_id,
            empty,
        )

    def summon_op(name: str, linked_id: int) -> int:
        return new_operation(
            name,
            json_cell(
                "Summon",
                unit_definition_identity_id=linked_id,
                owner_selector_id=selectors["actor"],
            ),
        )

    def besiege_steps(prefix: str, grant_extra: bool) -> list[str]:
        steps = [
            operation_step(
                apply_effect_op(
                    f"{prefix}-lock-on",
                    effects["lock-on"],
                    selectors["primary-target"],
                )
            ),
            operation_step(
                queue_action_op(
                    f"{prefix}-coordinated-strikes",
                    abilities["coordinated-strike"],
                    selectors["actor-summons"],
                    selectors["primary-target"],
                    "AfterAction",
                    empty="NoOp",
                )
            ),
        ]
        if grant_extra:
            steps.append(operation_step(extra_turn_op(f"{prefix}-extra-turn")))
        return steps

    ability_programs: dict[int, int] = {}
    ability_programs[abilities["fist"]] = make_program(
        "fist-of-conviction",
        [
            operation_step(
                damage_op(
                    "fist-of-conviction-damage",
                    damage["fist"],
                    selectors["primary-target"],
                )
            )
        ],
    )
    ability_programs[abilities["smite"]] = make_program(
        "smite-of-frost",
        [
            operation_step(
                damage_op(
                    "smite-of-frost-damage",
                    damage["smite"],
                    selectors["primary-target"],
                )
            )
        ],
    )
    ability_programs[abilities["smite-rapid"]] = make_program(
        "smite-of-frost-rapid",
        [
            operation_step(
                damage_op(
                    "smite-of-frost-rapid-damage",
                    damage["smite"],
                    selectors["primary-target"],
                )
            ),
            operation_step(extra_turn_op("smite-of-frost-rapid-extra-turn")),
        ],
    )
    frigid_steps = [
        operation_step(
            damage_op(
                "frigid-waterfall-damage",
                damage["frigid"],
                selectors["opposing-all"],
            )
        ),
        operation_step(
            apply_effect_op(
                "frigid-waterfall-escalation",
                effects["frigid-escalation"],
                selectors["actor"],
                stacks=integer_one,
            )
        ),
    ]
    ability_programs[abilities["frigid"]] = make_program(
        "frigid-waterfall", frigid_steps
    )
    ability_programs[abilities["frigid-rapid"]] = make_program(
        "frigid-waterfall-rapid",
        [
            operation_step(
                damage_op(
                    "frigid-waterfall-rapid-damage",
                    damage["frigid"],
                    selectors["opposing-all"],
                )
            ),
            operation_step(
                apply_effect_op(
                    "frigid-waterfall-rapid-escalation",
                    effects["frigid-escalation"],
                    selectors["actor"],
                    stacks=integer_one,
                )
            ),
            operation_step(extra_turn_op("frigid-waterfall-rapid-extra-turn")),
        ],
    )
    ability_programs[abilities["garrison"]] = make_program(
        "garrison-aura-field",
        [
            operation_step(
                new_operation(
                    "garrison-aura-field-shield",
                    json_cell(
                        "Shield",
                        amount_expression_id=shield_amount,
                        effect_id=effects["collective-shield"],
                    ),
                    selectors["same-all"],
                )
            ),
            operation_step(extra_turn_op("garrison-aura-field-extra-turn")),
        ],
    )
    ability_programs[abilities["besiege"]] = make_program(
        "besiege", besiege_steps("besiege", False)
    )
    ability_programs[abilities["besiege-rapid"]] = make_program(
        "besiege-rapid", besiege_steps("besiege-rapid", True)
    )
    ability_programs[abilities["besiege-end"]] = make_program(
        "besiege-rotation-end", besiege_steps("besiege-rotation-end", False)
    )
    ability_programs[abilities["counter"]] = make_program(
        "tit-for-tat",
        [
            operation_step(
                apply_effect_op(
                    "tit-for-tat-counter",
                    effects["counter"],
                    selectors["actor"],
                )
            )
        ],
    )
    ability_programs[abilities["soldier-charge"]] = make_program(
        "tireless-charge",
        [
            operation_step(
                damage_op(
                    "tireless-charge-damage",
                    damage["soldier"],
                    selectors["primary-target"],
                    element="Physical",
                )
            ),
            operation_step(
                apply_effect_op(
                    "tireless-charge-def-down",
                    effects["def-down"],
                    selectors["primary-target"],
                    chance=ratios["one"],
                )
            ),
        ],
    )
    ability_programs[abilities["cannoneer-barrage"]] = make_program(
        "cannoneer-barrage",
        [
            operation_step(
                damage_op(
                    "cannoneer-barrage-primary",
                    damage["cannoneer-primary"],
                    selectors["primary-target"],
                    element="Physical",
                )
            ),
            operation_step(
                damage_op(
                    "cannoneer-barrage-adjacent",
                    damage["cannoneer-adjacent"],
                    selectors["adjacent"],
                    element="Physical",
                    empty="NoOp",
                )
            ),
        ],
    )
    ability_programs[abilities["lieutenant-pierce"]] = make_program(
        "lieutenant-pierce",
        [
            operation_step(
                damage_op(
                    "lieutenant-pierce-damage",
                    damage["lieutenant"],
                    selectors["primary-target"],
                    element="Physical",
                )
            )
        ],
    )
    ability_programs[abilities["coordinated-strike"]] = make_program(
        "besiege-coordinated-strike",
        [
            operation_step(
                damage_op(
                    "besiege-coordinated-strike-damage",
                    damage["coordinated"],
                    selectors["primary-target"],
                    element="Physical",
                )
            )
        ],
    )
    phase_entry_programs = [
        make_program(
            "phase-1-support",
            [
                operation_step(
                    summon_op(
                        "phase-1-summon-soldier", linked["phase-1-soldier"]
                    )
                ),
                operation_step(
                    summon_op(
                        "phase-1-summon-cannoneer", linked["phase-1-cannoneer"]
                    )
                ),
            ],
        ),
        make_program(
            "phase-2-support",
            [
                operation_step(
                    summon_op(
                        "phase-2-summon-cannoneer-left",
                        linked["phase-2-cannoneer-left"],
                    )
                ),
                operation_step(
                    summon_op(
                        "phase-2-summon-cannoneer-right",
                        linked["phase-2-cannoneer-right"],
                    )
                ),
            ],
        ),
        make_program(
            "phase-3-support",
            [
                operation_step(
                    summon_op(
                        "phase-3-summon-lieutenant-left",
                        linked["phase-3-lieutenant-left"],
                    )
                ),
                operation_step(
                    summon_op(
                        "phase-3-summon-lieutenant-right",
                        linked["phase-3-lieutenant-right"],
                    )
                ),
            ],
        ),
    ]
    counter_program = make_program(
        "tit-for-tat-counter",
        [
            operation_step(
                queue_action_op(
                    "tit-for-tat-queue-smite",
                    abilities["smite"],
                    selectors["owner"],
                    selectors["actor"],
                    "AfterHit",
                )
            ),
            operation_step(
                new_operation(
                    "tit-for-tat-consume-counter",
                    json_cell("RemoveEffect", effect_id=effects["counter"]),
                    selectors["owner"],
                    "NoOp",
                )
            ),
        ],
    )

    target_patterns = {
        "fist": "SingleTarget",
        "smite": "SingleTarget",
        "garrison": "None",
        "frigid": "Aoe",
        "besiege": "SingleTarget",
        "frigid-rapid": "Aoe",
        "besiege-rapid": "SingleTarget",
        "counter": "None",
        "smite-rapid": "SingleTarget",
        "besiege-end": "SingleTarget",
        "soldier-charge": "SingleTarget",
        "cannoneer-barrage": "Blast",
        "lieutenant-pierce": "SingleTarget",
        "coordinated-strike": "SingleTarget",
    }
    linked_ability_keys = {
        "soldier-charge",
        "cannoneer-barrage",
        "lieutenant-pierce",
        "coordinated-strike",
    }
    for key, ability_id in abilities.items():
        is_linked = key in linked_ability_keys
        add(
            "Ability",
            {
                "id": ability_id,
                "kind": "Summon" if is_linked else "Skill",
                "target_pattern": target_patterns[key],
                "retarget_policy": "CancelRemaining",
                "level_cap": 1,
                "cooldown_actions": 1,
                "semantic_tags_mask": (
                    2_053
                    if key == "smite"
                    else 5 if target_patterns[key] != "None" else 4
                ),
            },
        )
        add(
            "AbilityPhase",
            {
                "ability_id": ability_id,
                "sequence": 1,
                "kind": "Resolved",
                "program_identity_id": ability_programs[ability_id],
            },
        )
        if not is_linked:
            add(
                "EnemyAbility",
                {
                    "id": ability_id,
                    "telegraph": "None",
                    "cooldown_actions": 1,
                    "initial_cooldown_actions": 0,
                    "charge_actions": 0,
                    "ai_tag": key,
                },
            )

    effect_definitions = {
        "frigid-escalation": (
            "Buff",
            "NonDispellable",
            100,
            None,
            "Permanent",
            "RefreshAndAddStacks",
        ),
        "collective-shield": (
            "Shield",
            "NonDispellable",
            1,
            None,
            "Permanent",
            "Replace",
        ),
        "counter": (
            "NeutralState",
            "NonDispellable",
            1,
            integer_one,
            "TargetTurnEnd",
            "Refresh",
        ),
        "lock-on": (
            "Mark",
            "NonDispellable",
            1,
            integer_one,
            "TargetTurnEnd",
            "Refresh",
        ),
        "def-down": (
            "Debuff",
            "DispellableDebuff",
            1,
            integer_one,
            "TargetTurnEnd",
            "Refresh",
        ),
    }
    for key, effect_id in effects.items():
        category, dispel, limit, duration, clock, policy = effect_definitions[key]
        add(
            "Effect",
            {
                "id": effect_id,
                "category": category,
                "dispel_category": dispel,
                "stack_limit": limit,
                "duration_expression_id": duration,
                "duration_clock": clock,
                "tick_phase": "None",
                "stack_policy": policy,
                "magnitude_comparator_expression_id": (
                    escalation_value if key == "frigid-escalation" else None
                ),
                "snapshot_policy": (
                    "RecomputeOnStackChange"
                    if key == "frigid-escalation"
                    else "OnApplication"
                ),
                "teardown_policy": "RemoveWithOwner",
                "application_priority": 0,
            },
        )
    for key, tags in {
        "frigid-escalation": ["frigid-waterfall-escalation"],
        "collective-shield": [
            "collective-shield",
            "prevents-toughness-reduction",
        ],
        "counter": ["counter", "single-trigger"],
        "lock-on": ["lock-on", "besiege-target"],
        "def-down": ["def-down"],
    }.items():
        for sequence, tag in enumerate(tags, start=1):
            add(
                "EffectTag",
                {"effect_id": effects[key], "sequence": sequence, "tag": tag},
            )
    for key, group_id in modifier_groups.items():
        add(
            "ModifierStackingGroup",
            {
                "id": group_id,
                "stable_key": f"goal07.enemy.s07.{key}",
                "aggregation": "Sum",
            },
        )
    add(
        "ModifierDefinition",
        {
            "id": modifiers["frigid-escalation"],
            "source_effect_id": effects["frigid-escalation"],
            "owner_selector_id": selectors["owner"],
            "subject_selector_id": selectors["owner"],
            "stat": "Atk",
            "formula_stage": "DamageBoost",
            "formula_purpose": "OrdinaryDamage",
            "value_expression_id": escalation_value,
            "value_domain": "Ratio",
            "stacking_group_id": modifier_groups["frigid-escalation"],
            "priority": 0,
            "cap_formula_stage": "DamageBoost",
            "snapshot_policy": "Dynamic",
            "duration_scope": "Turn",
        },
    )
    add(
        "ModifierDefinition",
        {
            "id": modifiers["def-down"],
            "source_effect_id": effects["def-down"],
            "owner_selector_id": selectors["owner"],
            "subject_selector_id": selectors["owner"],
            "stat": "Def",
            "formula_stage": "PercentOfBase",
            "formula_purpose": "Stat",
            "value_expression_id": ratios["negative-half"],
            "value_domain": "Ratio",
            "stacking_group_id": modifier_groups["def-down"],
            "priority": 0,
            "cap_formula_stage": "PercentOfBase",
            "snapshot_policy": "OnApplication",
            "duration_scope": "Turn",
        },
    )
    for sequence, key in enumerate(
        ["frigid-escalation", "def-down"], start=1
    ):
        add(
            "EffectModifierBinding",
            {
                "effect_id": effects[key],
                "sequence": 1,
                "modifier_id": modifiers[key],
            },
        )

    add(
        "RuleDefinition",
        {
            "id": rules["counter"],
            "domain": "Battle",
            "source_definition_identity_id": effects["counter"],
            "source_class": "Effect",
            "source_digest_sha256": sha256_text("goal07-s07-tit-for-tat-v1"),
        },
    )
    add(
        "EventFilter",
        {
            "id": filters["counter"],
            "stable_key": "goal07.enemy.s07.filter.tit-for-tat",
            "target_selector_id": selectors["owner"],
            "damage_class": "Ordinary",
            "cause_ancestry": "Any",
        },
    )
    add(
        "RuleTrigger",
        {
            "id": BASE + 571,
            "stable_key": "goal07.enemy.s07.trigger.tit-for-tat",
            "rule_id": rules["counter"],
            "sequence": 1,
            "event": json_cell("Damage", point="Applied"),
            "phase": "AfterEvent",
            "filter_id": filters["counter"],
            "condition_id": conditions["always"],
            "once_scope": "Turn",
            "priority": 100,
            "program_id": counter_program,
        },
    )
    add(
        "EffectRuleBinding",
        {
            "effect_id": effects["counter"],
            "sequence": 1,
            "rule_id": rules["counter"],
        },
    )

    phase_sequences = [
        [
            abilities["besiege"],
            abilities["frigid"],
            abilities["counter"],
            abilities["fist"],
        ],
        [
            abilities["besiege-rapid"],
            abilities["frigid"],
            abilities["garrison"],
            abilities["frigid"],
            abilities["smite-rapid"],
            abilities["frigid"],
            abilities["frigid-rapid"],
            abilities["counter"],
            abilities["smite-rapid"],
            abilities["besiege-end"],
        ],
        [
            abilities["besiege-rapid"],
            abilities["frigid"],
            abilities["garrison"],
            abilities["frigid"],
            abilities["smite-rapid"],
            abilities["frigid"],
            abilities["frigid-rapid"],
            abilities["counter"],
            abilities["smite-rapid"],
            abilities["besiege-end"],
        ],
    ]
    target_selectors = {
        abilities["fist"]: selectors["opposing-random"],
        abilities["smite"]: selectors["opposing-random"],
        abilities["garrison"]: selectors["actor"],
        abilities["frigid"]: selectors["opposing-all"],
        abilities["besiege"]: selectors["opposing-random"],
        abilities["frigid-rapid"]: selectors["opposing-all"],
        abilities["besiege-rapid"]: selectors["opposing-random"],
        abilities["counter"]: selectors["actor"],
        abilities["smite-rapid"]: selectors["opposing-random"],
        abilities["besiege-end"]: selectors["opposing-random"],
    }
    next_state = BASE + 701
    next_candidate = BASE + 801
    next_transition = BASE + 901
    for phase_index, sequence in enumerate(phase_sequences):
        state_ids = list(range(next_state, next_state + len(sequence)))
        next_state += len(sequence)
        add(
            "AiGraph",
            {
                "id": graphs[phase_index],
                "initial_state_id": state_ids[0],
                "automatic_transition_budget": 8,
            },
        )
        for offset, (state_id, ability_id) in enumerate(zip(state_ids, sequence)):
            add(
                "AiState",
                {
                    "id": state_id,
                    "stable_key": (
                        f"goal07.enemy.s07.ai.phase-{phase_index + 1}."
                        f"state-{offset + 1}"
                    ),
                    "graph_id": graphs[phase_index],
                    "mandatory_fallback_ability_id": abilities["fist"],
                    "turn_counter_reset": offset == 0,
                },
            )
            add(
                "AiCandidate",
                {
                    "id": next_candidate,
                    "stable_key": (
                        f"goal07.enemy.s07.ai.phase-{phase_index + 1}."
                        f"candidate-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "ability_id": ability_id,
                    "condition_id": conditions["always"],
                    "target_selector_id": target_selectors[ability_id],
                    "priority": 0,
                    "selection": "FirstLegal",
                    "no_target_fallback": "UseFallbackAbility",
                    "fallback_ability_id": abilities["fist"],
                },
            )
            next_candidate += 1
            transitions: list[tuple[int, int]] = []
            if phase_index > 0 and offset == 5:
                transitions = [
                    (state_ids[4], conditions["shielded"]),
                    (state_ids[6], conditions["unshielded"]),
                ]
            elif phase_index > 0 and offset == 9:
                transitions = [(state_ids[2], conditions["always"])]
            else:
                transitions = [
                    (state_ids[(offset + 1) % len(state_ids)], conditions["always"])
                ]
            for sequence_index, (target_state, condition_id) in enumerate(
                transitions, start=1
            ):
                add(
                    "AiTransition",
                    {
                        "id": next_transition,
                        "stable_key": (
                            f"goal07.enemy.s07.ai.phase-{phase_index + 1}."
                            f"transition-{offset + 1}-{sequence_index}"
                        ),
                        "state_id": state_id,
                        "sequence": sequence_index,
                        "target_state_id": target_state,
                        "condition_id": condition_id,
                        "priority": 1 if sequence_index == 1 else 0,
                        "timing": "AfterAction",
                    },
                )
                next_transition += 1

    linked_specs = {
        "phase-1-soldier": (
            1,
            "0.307692",
            "0.923077",
            "83",
            abilities["soldier-charge"],
        ),
        "phase-1-cannoneer": (
            5,
            "0.323077",
            "1.333333",
            "100",
            abilities["cannoneer-barrage"],
        ),
        "phase-2-cannoneer-left": (
            1,
            "0.323077",
            "1.333333",
            "100",
            abilities["cannoneer-barrage"],
        ),
        "phase-2-cannoneer-right": (
            5,
            "0.323077",
            "1.333333",
            "100",
            abilities["cannoneer-barrage"],
        ),
        "phase-3-lieutenant-left": (
            1,
            "0.692308",
            "1.333333",
            "144",
            abilities["lieutenant-pierce"],
        ),
        "phase-3-lieutenant-right": (
            5,
            "0.692308",
            "1.333333",
            "144",
            abilities["lieutenant-pierce"],
        ),
    }
    for key, linked_id in linked.items():
        formation, hp_ratio, atk_ratio, spd, action_ability = linked_specs[key]
        add(
            "LinkedUnitDefinition",
            {
                "id": linked_id,
                "source_definition_identity_id": linked_id,
                "kind": "Summon",
                "presence": "Present",
                "ability_ids": f"{action_ability}|{abilities['coordinated-strike']}",
                "action_ability_id": action_ability,
                "formation_index": formation,
                "initial_gauge_decimal": "10000",
                "hp_owner_ratio_decimal": hp_ratio,
                "hp_flat_decimal": "0",
                "atk_owner_ratio_decimal": atk_ratio,
                "atk_flat_decimal": "0",
                "def_owner_ratio_decimal": "1",
                "def_flat_decimal": "0",
                "spd_owner_ratio_decimal": "0",
                "spd_flat_decimal": spd,
                "owner_defeat_policy": "Depart",
                "owner_departure_policy": "Depart",
                "wave_policy": "Depart",
                "combatant_digest_sha256": sha256_text(
                    f"goal07-s07-linked-{key}-v1"
                ),
            },
        )

    add(
        "EnemyTemplate",
        {
            "id": template,
            "rank": "Boss",
            "base_aggro_decimal": "100",
            "default_ai_graph_id": graphs[0],
        },
    )
    add(
        "EnemyVariant",
        {
            "id": variant,
            "template_id": template,
            "ai_graph_id": graphs[0],
            "mechanically_distinct_key": VARIANT_KEY,
        },
    )
    for level in anchor["levels"]:
        add(
            "EnemyStat",
            {
                "variant_id": variant,
                "level": level["authored_level"],
                "difficulty_key": "standard-universe-v1",
                "hp_decimal": level["base_hp"],
                "atk_decimal": level["base_atk"],
                "def_decimal": level["base_def"],
                "spd_decimal": level["base_spd"],
                "effect_hit_rate_decimal": level["effect_hit_rate"],
                "effect_resistance_decimal": level["effect_resistance"],
                "crit_damage_decimal": "0.2",
            },
        )
    for sequence, weakness in enumerate(
        ["Imaginary", "Lightning", "Physical"], start=1
    ):
        add(
            "EnemyWeakness",
            {"variant_id": variant, "sequence": sequence, "element": weakness},
        )
    for element, value in [
        ("Fire", "0.2"),
        ("Ice", "0.4"),
        ("Quantum", "0.2"),
        ("Wind", "0.2"),
    ]:
        add(
            "EnemyResistance",
            {"variant_id": variant, "element": element, "value_decimal": value},
        )
    for category in [
        "STAT_CTRL_Confine",
        "STAT_CTRL_Frozen",
        "STAT_CTRL_Entangle",
    ]:
        add(
            "EnemyDebuffResistance",
            {
                "variant_id": variant,
                "category_key": category,
                "value_decimal": "0.75",
            },
        )
    add(
        "EnemyToughnessLayer",
        {
            "variant_id": variant,
            "sequence": 1,
            "layer_key": "ordinary",
            "kind": "Ordinary",
            "maximum_decimal": "100",
            "recovery_ratio_decimal": "1",
            "active_at_start": True,
        },
    )
    for sequence, ability_id in enumerate(
        [
            abilities["fist"],
            abilities["smite"],
            abilities["garrison"],
            abilities["frigid"],
            abilities["besiege"],
            abilities["frigid-rapid"],
            abilities["besiege-rapid"],
            abilities["counter"],
            abilities["smite-rapid"],
            abilities["besiege-end"],
        ],
        start=1,
    ):
        add(
            "EnemyVariantAbility",
            {
                "variant_id": variant,
                "sequence": sequence,
                "ability_id": ability_id,
            },
        )
    for sequence, graph in enumerate(graphs, start=1):
        add(
            "EnemyPhase",
            {
                "id": BASE + 600 + sequence,
                "stable_key": f"goal07.enemy.s07.phase-{sequence}",
                "variant_id": variant,
                "sequence": sequence,
                "entry_condition_id": conditions["always"],
                "exit_condition_id": conditions["always"],
                "replacement_priority": sequence,
                "ai_graph_id": graph,
                "targetable": True,
                "transition_model": "TransformSameUnit",
                "entry_program_id": phase_entry_programs[sequence - 1],
                "hp_carry": "Reset",
                "action_gauge_carry": "Reset",
                "effect_carry": "Clear",
                "toughness_carry": "Reset",
                "summon_carry": "Clear",
            },
        )

    anchor_digest = sha256_bytes(anchor_path(PARTITION).read_bytes())
    add(
        "SourceRecord",
        {
            "id": SOURCE_RECORD_ID,
            "stable_key": "source.hsr-wiki.gepard-complete.2026-07-29",
            "category": "CommunityMaintained",
            "publisher": anchor["source"]["publisher"],
            "url": anchor["source"]["url"],
            "accessed_on": anchor["source"]["accessed_on"],
            "applicable_game_version": anchor["source"]["game_version"],
            "confidence": "SecondaryVersionSensitiveCrossCheck",
            "evidence_sha256": anchor_digest,
            "usage_note": (
                "Exact public per-level values are committed with retained "
                "structured AI and ability source hashes."
            ),
        },
    )
    add(
        "EvidenceRecord",
        {
            "id": EVIDENCE_RECORD_ID,
            "stable_key": "evidence.goal07.enemy.s07.numeric-anchors",
            "kind": "SourcePayload",
            "source_record_id": SOURCE_RECORD_ID,
            "sha256": anchor_digest,
            "note": "Committed exact public per-level numeric anchors for Goal 07 S07.",
        },
    )
    for item in identities:
        add("ContentIdentity", item)
        add(
            "ContentEvidenceBinding",
            {
                "content_id": item["id"],
                "sequence": 1,
                "fact_key": f"goal07.s07.executable:{item['stable_key']}",
                "source_record_id": 1,
                "evidence_record_id": 3,
                "quality": "ExactStructured",
                "mechanism_quality": "ExactStructured",
            },
        )
    add(
        "ContentEvidenceBinding",
        {
            "content_id": variant,
            "sequence": 2,
            "fact_key": "goal07.s07.public-level-stats",
            "source_record_id": SOURCE_RECORD_ID,
            "evidence_record_id": EVIDENCE_RECORD_ID,
            "quality": "ExactStructured",
            "mechanism_quality": "ExactStructured",
        },
    )
    for table_rows in rows.values():
        table_rows.sort(
            key=lambda row: json.dumps(
                row, ensure_ascii=False, sort_keys=True, default=str
            )
        )
    return rows


def owned_rows_s08() -> dict[str, list[dict[str, Any]]]:
    anchor = json.loads(anchor_path(PARTITION).read_text(encoding="utf-8"))
    manifest = json.loads(PARTITIONS.read_text(encoding="utf-8"))
    assigned = next(item for item in manifest["partitions"] if item["id"] == PARTITION)
    if assigned["enemy_variant_ids"] != [VARIANT_KEY]:
        raise ValueError("S08 frozen enemy assignment changed")

    variant, template, graph = BASE + 1, BASE + 2, BASE + 10
    abilities = {
        "lament": BASE + 101,
        "absorption": BASE + 102,
        "rain": BASE + 103,
        "enhanced-lament": BASE + 104,
    }
    selectors = {
        "actor": BASE + 401,
        "owner": BASE + 402,
        "primary": BASE + 403,
        "all-opposing": BASE + 404,
        "frozen-primary": BASE + 405,
    }
    effects = {
        "freeze": BASE + 501,
        "freezing-point": BASE + 502,
        "reset": BASE + 503,
    }
    conditions = {"always": BASE + 551, "reset": BASE + 552}
    reset_rule, reset_filter, reset_trigger = BASE + 541, BASE + 542, BASE + 543
    rows: dict[str, list[dict[str, Any]]] = {}
    identities: list[dict[str, Any]] = []
    next_program, next_operation, next_expression = BASE + 301, BASE + 1_001, BASE + 1_101

    def add(table: str, row: dict[str, Any]) -> None:
        rows.setdefault(table, []).append(row)

    def ident(
        id_: int,
        key: str,
        kind: str,
        name_en: str,
        name_zh_cn: str,
        summary: str,
        sources: str = "1",
    ) -> dict[str, Any]:
        row = identity(id_, key, kind, name_en, name_zh_cn, summary, sources)
        row["summary_zh_cn"] = "Goal 07 S08 来源绑定的外宇宙之冰可执行定义。"
        row["game_version_introduced"] = "1.0"
        return row

    identities.extend(
        [
            ident(
                variant,
                VARIANT_KEY,
                "EnemyVariant",
                "Ice Out of Space",
                "外宇宙之冰",
                "Exact materialization variant used by frozen universe bindings.",
                "1|10",
            ),
            ident(
                template,
                "enemy.ice-out-of-space.elite",
                "Enemy",
                "Ice Out of Space Template",
                "外宇宙之冰模板",
                "Version 4.4 elite template retained from source monster 8003010.",
            ),
            ident(
                graph,
                "ai.goal07.ice-out-of-space.phase-1",
                "AiGraph",
                "Ice Out of Space AI",
                "外宇宙之冰AI",
                "Source-ordered setup, enhanced attack and break-reset graph.",
            ),
        ]
    )
    ability_metadata = {
        "lament": ("Chilling Lament", "叹息之寒", "800301001"),
        "absorption": ("Frosty Absorption", "汲取霜晶", "800301002"),
        "rain": ("Everwinter Rain", "永冬之雨", "800301003"),
        "enhanced-lament": ("Chilling Lament (Enhanced)", "叹息之寒·强化", "800301004"),
    }
    for key, id_ in abilities.items():
        name_en, name_zh_cn, source = ability_metadata[key]
        identities.append(
            ident(
                id_,
                f"enemy.ice-out-of-space.elite.ability.{key}",
                "Ability",
                name_en,
                name_zh_cn,
                f"Executable transcription of source skill {source}.",
            )
        )
    for key, id_ in selectors.items():
        identities.append(
            ident(
                id_,
                f"selector.goal07.ice-out-of-space.{key}",
                "Selector",
                f"Ice Out of Space {key} Selector",
                f"外宇宙之冰{key}选择器",
                "S08 battle selector.",
            )
        )
    for key, id_, name_en, name_zh_cn, summary in [
        ("freeze", effects["freeze"], "Ice Out of Space Freeze", "外宇宙之冰冻结", "One-turn control with delayed Ice damage."),
        ("freezing-point", effects["freezing-point"], "Freezing Point", "冰点", "Enhanced mode removed only by Weakness Break."),
        ("reset", effects["reset"], "Freezing Point Reset", "冰点重置", "Internal AI reset marker."),
    ]:
        identities.append(
            ident(
                id_,
                f"effect.goal07.ice-out-of-space.{key}",
                "Effect",
                name_en,
                name_zh_cn,
                summary,
            )
        )
    identities.append(
        ident(
            reset_rule,
            "rule.goal07.ice-out-of-space.weakness-break-reset",
            "Rule",
            "Ice Out of Space Weakness Break Reset",
            "外宇宙之冰弱点击破重置",
            "WeaknessBroken removes Freezing Point and restarts setup.",
        )
    )

    add("Selector", selector(selectors["actor"], "Actor", "SameSide"))
    add("Selector", selector(selectors["owner"], "Owner", "SameSide"))
    add("Selector", selector(selectors["primary"], "PrimaryTarget", "OpposingSide"))
    add(
        "Selector",
        selector(
            selectors["all-opposing"],
            "Actor",
            "OpposingSide",
            minimum=1,
            maximum=8,
            choice="All",
        ),
    )
    add(
        "Selector",
        selector(
            selectors["frozen-primary"],
            "PrimaryTarget",
            "OpposingSide",
            minimum=0,
            maximum=1,
            empty="NoOp",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["frozen-primary"],
            "sequence": 1,
            "predicate": json_cell("HasEffect", effect_id=effects["freeze"]),
        },
    )

    def expression(name: str, kind: str, node: str) -> int:
        nonlocal next_expression
        id_ = next_expression
        next_expression += 1
        add(
            "ValueExpression",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s08.expression.{name}",
                "result_kind": kind,
                "node": node,
            },
        )
        return id_

    def product(name: str, left: int, right: int) -> int:
        return expression(
            name,
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedMultiply",
                left_expression_id=left,
                right_expression_id=right,
                rounding="NearestTiesAway",
            ),
        )

    actor_atk = expression(
        "actor-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    ratios = {
        "two-point-five": expression("ratio-2-5", "Scalar", json_cell("ScalarLiteral", value_decimal="2.5")),
        "two": expression("ratio-2", "Scalar", json_cell("ScalarLiteral", value_decimal="2")),
        "one": expression("ratio-1", "Scalar", json_cell("ScalarLiteral", value_decimal="1")),
        "six-tenths": expression("ratio-0-6", "Scalar", json_cell("ScalarLiteral", value_decimal="0.6")),
    }
    integer_one = expression("integer-one", "Integer", json_cell("IntegerLiteral", value=1))
    integer_zero = expression("integer-zero", "Integer", json_cell("IntegerLiteral", value=0))
    reset_stacks = expression(
        "reset-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["actor"],
            effect_id=effects["reset"],
        ),
    )
    damage = {
        "lament": product("lament-damage", actor_atk, ratios["two-point-five"]),
        "extra": product("lament-extra-damage", actor_atk, ratios["one"]),
        "rain": product("rain-damage", actor_atk, ratios["two"]),
        "freeze": product("freeze-delayed-damage", actor_atk, ratios["one"]),
    }
    add(
        "ConditionExpression",
        {
            "id": conditions["always"],
            "stable_key": "goal07.enemy.s08.condition.always",
            "node": json_cell("Constant", value=True),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["reset"],
            "stable_key": "goal07.enemy.s08.condition.reset",
            "node": json_cell(
                "Compare",
                left_expression_id=reset_stacks,
                comparison="Greater",
                right_expression_id=integer_zero,
            ),
        },
    )

    def new_operation(
        name: str,
        payload: str,
        target: int | None = None,
        empty: str = "Fault",
    ) -> int:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        row = operation(id_, name, payload, target, empty)
        row["stable_key"] = f"goal07.enemy.s08.operation.{name}"
        add("Operation", row)
        return id_

    def op_step(id_: int) -> str:
        return json_cell("Operation", operation_id=id_)

    def damage_op(name: str, amount: int, target: int, empty: str = "Fault") -> int:
        return new_operation(
            name,
            json_cell(
                "Damage",
                amount_expression_id=amount,
                damage_class="Ordinary",
                element="Ice",
                can_crit=True,
            ),
            target,
            empty,
        )

    def apply_op(
        name: str,
        effect: int,
        target: int,
        chance: int | None = None,
    ) -> int:
        return new_operation(
            name,
            json_cell(
                "ApplyEffect",
                effect_id=effect,
                stacks_expression_id=None,
                chance_policy="Resistible" if chance is not None else "Guaranteed",
                base_chance_expression_id=chance,
                rng_purpose_key="effect-application" if chance is not None else None,
            ),
            target,
        )

    def remove_op(name: str, effect: int, target: int) -> int:
        return new_operation(
            name,
            json_cell("RemoveEffect", effect_id=effect),
            target,
            "NoOp",
        )

    def make_program(name: str, steps: list[str]) -> int:
        nonlocal next_program
        id_ = next_program
        next_program += 1
        identities.append(
            ident(
                id_,
                f"program.goal07.ice-out-of-space.{name}",
                "Program",
                f"Ice Out of Space {name} Program",
                f"外宇宙之冰{name}程序",
                "Ordered Rule IR program for the S08 enemy.",
            )
        )
        add("Program", {"id": id_, "domain": "Battle"})
        for sequence, step in enumerate(steps, start=1):
            add("ProgramStep", {"program_id": id_, "sequence": sequence, "step": step})
        return id_

    ability_programs = {
        abilities["lament"]: make_program(
            "lament",
            [op_step(damage_op("lament-damage", damage["lament"], selectors["primary"]))],
        ),
        abilities["absorption"]: make_program(
            "absorption",
            [
                op_step(apply_op("absorption-freezing-point", effects["freezing-point"], selectors["actor"])),
                op_step(remove_op("absorption-clear-reset", effects["reset"], selectors["actor"])),
            ],
        ),
        abilities["rain"]: make_program(
            "rain",
            [
                op_step(damage_op("rain-damage", damage["rain"], selectors["all-opposing"])),
                op_step(apply_op("rain-freeze", effects["freeze"], selectors["all-opposing"], ratios["six-tenths"])),
            ],
        ),
        abilities["enhanced-lament"]: make_program(
            "enhanced-lament",
            [
                op_step(damage_op("enhanced-lament-damage", damage["lament"], selectors["primary"])),
                op_step(damage_op("enhanced-lament-frozen-extra", damage["extra"], selectors["frozen-primary"], "NoOp")),
            ],
        ),
    }
    reset_program = make_program(
        "weakness-break-reset",
        [
            op_step(remove_op("weakness-break-clear-freezing-point", effects["freezing-point"], selectors["owner"])),
            op_step(apply_op("weakness-break-mark-reset", effects["reset"], selectors["owner"])),
        ],
    )

    patterns = {
        "lament": "SingleTarget",
        "absorption": "None",
        "rain": "Aoe",
        "enhanced-lament": "SingleTarget",
    }
    for key, id_ in abilities.items():
        add(
            "Ability",
            {
                "id": id_,
                "kind": "Skill",
                "target_pattern": patterns[key],
                "retarget_policy": "CancelRemaining",
                "level_cap": 1,
                "cooldown_actions": 1,
                "semantic_tags_mask": 5 if patterns[key] != "None" else 4,
            },
        )
        add(
            "AbilityPhase",
            {
                "ability_id": id_,
                "sequence": 1,
                "kind": "Resolved",
                "program_identity_id": ability_programs[id_],
            },
        )
        add(
            "EnemyAbility",
            {
                "id": id_,
                "telegraph": "None",
                "cooldown_actions": 1,
                "initial_cooldown_actions": 0,
                "charge_actions": 0,
                "ai_tag": key,
            },
        )

    effect_rows = {
        "freeze": ("Control", "CleanseableControl", 1, integer_one, "TargetTurnStart", "TurnStart", "Refresh", damage["freeze"], "Ice"),
        "freezing-point": ("NeutralState", "NonDispellable", 1, None, "Permanent", "None", "Replace", None, None),
        "reset": ("NeutralState", "NonDispellable", 1, None, "Permanent", "None", "Replace", None, None),
    }
    for key, id_ in effects.items():
        category, dispel, limit, duration, clock, tick, policy, magnitude, dot = effect_rows[key]
        add(
            "Effect",
            {
                "id": id_,
                "category": category,
                "dispel_category": dispel,
                "stack_limit": limit,
                "duration_expression_id": duration,
                "duration_clock": clock,
                "tick_phase": tick,
                "stack_policy": policy,
                "magnitude_comparator_expression_id": magnitude,
                "dot_element": dot,
                "snapshot_policy": "OnApplication",
                "teardown_policy": "RemoveWithOwner",
                "application_priority": 0,
            },
        )
    for key, tags in {
        "freeze": ["freeze", "blocks-normal-action"],
        "freezing-point": ["freezing-point", "remove-on-weakness-break"],
        "reset": ["internal-ai-reset"],
    }.items():
        for sequence, tag in enumerate(tags, start=1):
            add("EffectTag", {"effect_id": effects[key], "sequence": sequence, "tag": tag})

    add(
        "RuleDefinition",
        {
            "id": reset_rule,
            "domain": "Battle",
            "source_definition_identity_id": effects["freezing-point"],
            "source_class": "Effect",
            "source_digest_sha256": sha256_text("goal07-s08-freezing-point-break-reset-v1"),
        },
    )
    add(
        "EventFilter",
        {
            "id": reset_filter,
            "stable_key": "goal07.enemy.s08.filter.weakness-break-owner",
            "target_selector_id": selectors["owner"],
            "cause_ancestry": "Any",
        },
    )
    add(
        "RuleTrigger",
        {
            "id": reset_trigger,
            "stable_key": "goal07.enemy.s08.trigger.weakness-break-reset",
            "rule_id": reset_rule,
            "sequence": 1,
            "event": json_cell("WeaknessBroken"),
            "phase": "AfterEvent",
            "filter_id": reset_filter,
            "condition_id": conditions["always"],
            "once_scope": "Event",
            "priority": 0,
            "program_id": reset_program,
        },
    )
    add(
        "EffectRuleBinding",
        {"effect_id": effects["freezing-point"], "sequence": 1, "rule_id": reset_rule},
    )

    states = {
        "initial-lament": BASE + 701,
        "initial-absorption": BASE + 702,
        "rain": BASE + 703,
        "enhanced-lament": BASE + 704,
        "reset-absorption": BASE + 705,
    }
    add(
        "AiGraph",
        {"id": graph, "initial_state_id": states["initial-lament"], "automatic_transition_budget": 8},
    )
    state_abilities = {
        "initial-lament": abilities["lament"],
        "initial-absorption": abilities["absorption"],
        "rain": abilities["rain"],
        "enhanced-lament": abilities["enhanced-lament"],
        "reset-absorption": abilities["absorption"],
    }
    targets = {
        abilities["lament"]: selectors["primary"],
        abilities["absorption"]: selectors["actor"],
        abilities["rain"]: selectors["all-opposing"],
        abilities["enhanced-lament"]: selectors["primary"],
    }
    next_candidate = BASE + 801
    for name, state_id in states.items():
        add(
            "AiState",
            {
                "id": state_id,
                "stable_key": f"goal07.enemy.s08.ai.{name}",
                "graph_id": graph,
                "mandatory_fallback_ability_id": abilities["lament"],
                "turn_counter_reset": name == "initial-lament",
            },
        )
        sequence = 1
        if name in {"rain", "enhanced-lament"}:
            add(
                "AiCandidate",
                {
                    "id": next_candidate,
                    "stable_key": f"goal07.enemy.s08.ai.{name}.reset-lament",
                    "state_id": state_id,
                    "sequence": sequence,
                    "ability_id": abilities["lament"],
                    "condition_id": conditions["reset"],
                    "target_selector_id": selectors["primary"],
                    "priority": 0,
                    "selection": "FirstLegal",
                    "no_target_fallback": "UseFallbackAbility",
                    "fallback_ability_id": abilities["lament"],
                },
            )
            next_candidate += 1
            sequence += 1
        ability_id = state_abilities[name]
        add(
            "AiCandidate",
            {
                "id": next_candidate,
                "stable_key": f"goal07.enemy.s08.ai.{name}.main",
                "state_id": state_id,
                "sequence": sequence,
                "ability_id": ability_id,
                "condition_id": conditions["always"],
                "target_selector_id": targets[ability_id],
                "priority": 10,
                "selection": "FirstLegal",
                "no_target_fallback": "UseFallbackAbility",
                "fallback_ability_id": abilities["lament"],
            },
        )
        next_candidate += 1
    normal_transitions = {
        "initial-lament": "initial-absorption",
        "initial-absorption": "rain",
        "rain": "enhanced-lament",
        "enhanced-lament": "rain",
        "reset-absorption": "rain",
    }
    next_transition = BASE + 901
    for name, target in normal_transitions.items():
        sequence = 1
        if name in {"rain", "enhanced-lament"}:
            add(
                "AiTransition",
                {
                    "id": next_transition,
                    "stable_key": f"goal07.enemy.s08.ai.{name}.to-reset-absorption",
                    "state_id": states[name],
                    "sequence": sequence,
                    "target_state_id": states["reset-absorption"],
                    "condition_id": conditions["reset"],
                    "priority": 0,
                    "timing": "AfterAction",
                },
            )
            next_transition += 1
            sequence += 1
        add(
            "AiTransition",
            {
                "id": next_transition,
                "stable_key": f"goal07.enemy.s08.ai.{name}.normal-transition",
                "state_id": states[name],
                "sequence": sequence,
                "target_state_id": states[target],
                "condition_id": conditions["always"],
                "priority": 10,
                "timing": "AfterAction",
            },
        )
        next_transition += 1

    add("EnemyTemplate", {"id": template, "rank": "Elite", "base_aggro_decimal": "100", "default_ai_graph_id": graph})
    add("EnemyVariant", {"id": variant, "template_id": template, "ai_graph_id": graph, "mechanically_distinct_key": VARIANT_KEY})
    for level in anchor["levels"]:
        add(
            "EnemyStat",
            {
                "variant_id": variant,
                "level": level["authored_level"],
                "difficulty_key": "standard-universe-v1",
                "hp_decimal": level["base_hp"],
                "atk_decimal": level["base_atk"],
                "def_decimal": level["base_def"],
                "spd_decimal": level["base_spd"],
                "effect_hit_rate_decimal": level["effect_hit_rate"],
                "effect_resistance_decimal": level["effect_resistance"],
                "crit_damage_decimal": "0.2",
            },
        )
    for sequence, weakness in enumerate(["Fire", "Wind", "Quantum"], start=1):
        add("EnemyWeakness", {"variant_id": variant, "sequence": sequence, "element": weakness})
    for element, value in [("Ice", "0.4"), ("Physical", "0.2"), ("Lightning", "0.2"), ("Imaginary", "0.2")]:
        add("EnemyResistance", {"variant_id": variant, "element": element, "value_decimal": value})
    add("EnemyDebuffResistance", {"variant_id": variant, "category_key": "STAT_CTRL_Frozen", "value_decimal": "1"})
    add(
        "EnemyToughnessLayer",
        {
            "variant_id": variant,
            "sequence": 1,
            "layer_key": "ordinary",
            "kind": "Ordinary",
            "maximum_decimal": "300",
            "recovery_ratio_decimal": "1",
            "active_at_start": True,
        },
    )
    for sequence, ability_id in enumerate(abilities.values(), start=1):
        add("EnemyVariantAbility", {"variant_id": variant, "sequence": sequence, "ability_id": ability_id})
    add(
        "EnemyPhase",
        {
            "id": BASE + 601,
            "stable_key": "goal07.enemy.s08.phase-1",
            "variant_id": variant,
            "sequence": 1,
            "entry_condition_id": conditions["always"],
            "exit_condition_id": conditions["always"],
            "replacement_priority": 1,
            "ai_graph_id": graph,
            "targetable": True,
            "transition_model": "TransformSameUnit",
            "hp_carry": "Reset",
            "action_gauge_carry": "Reset",
            "effect_carry": "Clear",
            "toughness_carry": "Reset",
            "summon_carry": "Clear",
        },
    )

    anchor_digest = sha256_bytes(anchor_path(PARTITION).read_bytes())
    add(
        "SourceRecord",
        {
            "id": SOURCE_RECORD_ID,
            "stable_key": "source.hsr-wiki.ice-out-of-space.2026-07-29",
            "category": "CommunityMaintained",
            "publisher": anchor["source"]["publisher"],
            "url": anchor["source"]["url"],
            "accessed_on": anchor["source"]["accessed_on"],
            "applicable_game_version": anchor["source"]["game_version"],
            "confidence": "SecondaryVersionSensitiveCrossCheck",
            "evidence_sha256": anchor_digest,
            "usage_note": "Exact public per-level values and mechanics are committed as Goal 07 evidence.",
        },
    )
    add(
        "EvidenceRecord",
        {
            "id": EVIDENCE_RECORD_ID,
            "stable_key": "evidence.goal07.enemy.s08.numeric-anchors",
            "kind": "SourcePayload",
            "source_record_id": SOURCE_RECORD_ID,
            "sha256": anchor_digest,
            "note": "Committed exact public per-level numeric anchors for Goal 07 S08.",
        },
    )
    for item in identities:
        add("ContentIdentity", item)
        add(
            "ContentEvidenceBinding",
            {
                "content_id": item["id"],
                "sequence": 1,
                "fact_key": f"goal07.s08.executable:{item['stable_key']}",
                "source_record_id": 1,
                "evidence_record_id": 3,
                "quality": "ExactStructured",
                "mechanism_quality": "ExactStructured",
            },
        )
    add(
        "ContentEvidenceBinding",
        {
            "content_id": variant,
            "sequence": 2,
            "fact_key": "goal07.s08.public-level-stats",
            "source_record_id": SOURCE_RECORD_ID,
            "evidence_record_id": EVIDENCE_RECORD_ID,
            "quality": "ExactStructured",
            "mechanism_quality": "ExactStructured",
        },
    )
    for table_rows in rows.values():
        table_rows.sort(key=lambda row: json.dumps(row, ensure_ascii=False, sort_keys=True, default=str))
    return rows


def owned_rows_s09() -> dict[str, list[dict[str, Any]]]:
    anchor = json.loads(anchor_path(PARTITION).read_text(encoding="utf-8"))
    manifest = json.loads(PARTITIONS.read_text(encoding="utf-8"))
    assigned = next(item for item in manifest["partitions"] if item["id"] == PARTITION)
    if assigned["enemy_variant_ids"] != [VARIANT_KEY]:
        raise ValueError("S09 frozen enemy assignment changed")

    variant, template = BASE + 1, BASE + 2
    graphs = [BASE + 10, BASE + 11, BASE + 12]
    abilities = {
        "funereal-kiss": BASE + 101,
        "sunken-rain": BASE + 102,
        "harrowing-obituary": BASE + 103,
        "fading-radiance": BASE + 104,
        "losing-eventide-light": BASE + 105,
        "go-into-that-good-night": BASE + 106,
        "watery-dissolution": BASE + 107,
        "capture-insert": BASE + 108,
        "sad-true-lover": BASE + 109,
    }
    tombs = {f"formation-{index}": BASE + 201 + index for index in range(4)}
    selectors = {
        "actor": BASE + 401,
        "owner": BASE + 402,
        "primary": BASE + 403,
        "all-opposing": BASE + 404,
        "random-two": BASE + 405,
        "current-subject": BASE + 406,
        "dreaming-opponents": BASE + 407,
    }
    for index in range(4):
        selectors[f"primary-{index}"] = BASE + 410 + index
        selectors[f"current-{index}"] = BASE + 420 + index
        selectors[f"release-{index}"] = BASE + 430 + index
        selectors[f"actor-tomb-{index}"] = BASE + 440 + index
        selectors[f"owner-tomb-{index}"] = BASE + 450 + index
    effects = {
        "obituary-permanent": BASE + 501,
        "obituary-burst": BASE + 502,
        "sunset": BASE + 503,
        "nightfall": BASE + 504,
        "morbid-dream": BASE + 505,
        "tomb-bars": BASE + 506,
    }
    for index in range(4):
        effects[f"tomb-marker-{index}"] = BASE + 510 + index
    modifiers = {
        "obituary-permanent": BASE + 521,
        "obituary-burst": BASE + 522,
    }
    modifier_groups = {
        "obituary-permanent": BASE + 531,
        "obituary-burst": BASE + 532,
    }
    rules = {
        "sunset-actions": BASE + 541,
        "sunset-break": BASE + 542,
        "nightfall-actions": BASE + 543,
        "nightfall-break": BASE + 544,
        "tomb-bars": BASE + 545,
    }
    for index in range(4):
        rules[f"tomb-release-{index}"] = BASE + 546 + index
    filters = {key: rule_id + 100 for key, rule_id in rules.items()}
    conditions = {
        "always": BASE + 551,
        "sunset-ready": BASE + 552,
        "nightfall-second": BASE + 553,
        "tomb-last": BASE + 554,
        "tomb-more": BASE + 555,
        "no-dreams": BASE + 556,
        "sunset-not-ready": BASE + 557,
        "nightfall-not-second": BASE + 558,
    }
    for index in range(4):
        conditions[f"primary-{index}"] = BASE + 560 + index
        conditions[f"current-{index}"] = BASE + 570 + index

    rows: dict[str, list[dict[str, Any]]] = {}
    identities: list[dict[str, Any]] = []
    next_program = BASE + 301
    next_operation = BASE + 1_001
    next_expression = BASE + 1_101
    next_trigger = BASE + 671

    def add(table: str, row: dict[str, Any]) -> None:
        rows.setdefault(table, []).append(row)

    def ident(
        id_: int,
        key: str,
        kind: str,
        name_en: str,
        name_zh_cn: str,
        summary: str,
        sources: str = "1",
    ) -> dict[str, Any]:
        row = identity(id_, key, kind, name_en, name_zh_cn, summary, sources)
        row["summary_zh_cn"] = "Goal 07 S09 来源绑定的忆域迷因「何物朝向死亡」可执行定义。"
        row["game_version_introduced"] = "2.0"
        return row

    identities.extend(
        [
            ident(
                variant,
                VARIANT_KEY,
                "EnemyVariant",
                'Memory Zone Meme "Something Unto Death" (Complete)',
                "忆域迷因「何物朝向死亡」（完整）",
                "Exact three-phase World 9 boss materialization.",
                "1|11",
            ),
            ident(
                template,
                "enemy.memory-zone-meme-something-unto-death-complete.littleboss",
                "Enemy",
                'Memory Zone Meme "Something Unto Death" Template',
                "忆域迷因「何物朝向死亡」（完整）模板",
                "Version 4.4 boss template retained from source monster 3014022.",
            ),
        ]
    )
    for sequence, graph_id in enumerate(graphs, start=1):
        identities.append(
            ident(
                graph_id,
                f"ai.goal07.memory-zone-meme-something-unto-death.phase-{sequence}",
                "AiGraph",
                f"Something Unto Death Phase {sequence} AI",
                f"何物朝向死亡第{sequence}阶段AI",
                "Finite source-phase action graph with queued capture reactions.",
            )
        )
    ability_metadata = {
        "funereal-kiss": ("Funereal Kiss", "葬仪之吻", "301402201"),
        "sunken-rain": ("Sunken Rain", "沦没黑雨", "301402202"),
        "harrowing-obituary": ("Harrowing Obituary", "彻心悼词", "301402203"),
        "fading-radiance": ("Fading Radiance", "光明消歇", "301402204"),
        "losing-eventide-light": (
            "Losing Eventide Light",
            "已在黄昏失色",
            "301402205",
        ),
        "go-into-that-good-night": (
            "Go Into That Good Night",
            "步入良夜",
            "301402206",
        ),
        "watery-dissolution": ("Watery Dissolution", "如水逝于水中", "301402207"),
        "capture-insert": ("Morbid Dream Insert", "梦死插入", "301402208"),
        "sad-true-lover": (
            "Sad True Lover Never Find My Grave",
            "无从凭吊的荒场",
            "301402209",
        ),
    }
    for key, ability_id in abilities.items():
        name_en, name_zh_cn, source_skill = ability_metadata[key]
        identities.append(
            ident(
                ability_id,
                (
                    "enemy.memory-zone-meme-something-unto-death-complete."
                    f"littleboss.ability.{key}"
                ),
                "Ability",
                name_en,
                name_zh_cn,
                f"Executable transcription of source skill {source_skill}.",
            )
        )
    for index, tomb_id in enumerate(tombs.values()):
        identities.append(
            ident(
                tomb_id,
                f"unit.goal07.something-unto-death.sombrous-sepulcher-{index + 1}",
                "CharacterForm",
                f"Sombrous Sepulcher {index + 1}",
                f"阴影墓碑{index + 1}",
                "Formation-linked five-bar captivity unit from source monster 3012030.",
            )
        )
    for key, selector_id in selectors.items():
        identities.append(
            ident(
                selector_id,
                f"selector.goal07.something-unto-death.{key}",
                "Selector",
                f"Something Unto Death {key} Selector",
                f"何物朝向死亡{key}选择器",
                "S09 battle selector.",
            )
        )
    effect_names = {
        "obituary-permanent": ("Harrowing Obituary Growth", "彻心悼词成长"),
        "obituary-burst": ("Harrowing Obituary Burst", "彻心悼词爆发"),
        "sunset": ("Sunset", "落日"),
        "nightfall": ("Nightfall", "夜幕"),
        "morbid-dream": ("Morbid Dream", "梦死"),
        "tomb-bars": ("Sombrous Sepulcher Bars", "阴影墓碑生命段"),
    }
    for index in range(4):
        effect_names[f"tomb-marker-{index}"] = (
            f"Sombrous Sepulcher Formation {index + 1}",
            f"阴影墓碑站位{index + 1}",
        )
    for key, effect_id in effects.items():
        name_en, name_zh_cn = effect_names[key]
        identities.append(
            ident(
                effect_id,
                f"effect.goal07.something-unto-death.{key}",
                "Effect",
                name_en,
                name_zh_cn,
                "Effect-backed S09 capture, reset, or tomb lifecycle state.",
            )
        )
    for key, modifier_id in modifiers.items():
        identities.append(
            ident(
                modifier_id,
                f"modifier.goal07.something-unto-death.{key}",
                "Modifier",
                f"Something Unto Death {key} Modifier",
                f"何物朝向死亡{key}调整器",
                "Effect-owned public damage increase.",
            )
        )
    for key, rule_id in rules.items():
        identities.append(
            ident(
                rule_id,
                f"rule.goal07.something-unto-death.{key}",
                "Rule",
                f"Something Unto Death {key} Rule",
                f"何物朝向死亡{key}规则",
                "Shared Rule IR lifecycle rule for S09.",
            )
        )

    add("Selector", selector(selectors["actor"], "Actor", "SameSide"))
    add("Selector", selector(selectors["owner"], "Owner", "SameSide"))
    add("Selector", selector(selectors["primary"], "PrimaryTarget", "OpposingSide"))
    add(
        "Selector",
        selector(
            selectors["all-opposing"],
            "Actor",
            "OpposingSide",
            minimum=1,
            maximum=8,
            choice="All",
        ),
    )
    random_two = selector(
        selectors["random-two"],
        "Actor",
        "OpposingSide",
        minimum=1,
        maximum=2,
        choice="RngUniform",
    )
    random_two["rng_purpose_key"] = "behavior-choice"
    add("Selector", random_two)
    add(
        "Selector",
        selector(
            selectors["current-subject"],
            "CurrentSubject",
            "AnySide",
            life="Any",
            presence="Any",
        ),
    )
    add(
        "Selector",
        selector(
            selectors["dreaming-opponents"],
            "Actor",
            "OpposingSide",
            life="Any",
            presence="Any",
            minimum=0,
            maximum=4,
            empty="NoOp",
            choice="All",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["dreaming-opponents"],
            "sequence": 1,
            "predicate": json_cell(
                "HasEffect", effect_id=effects["morbid-dream"]
            ),
        },
    )
    for index in range(4):
        selector_specs = [
            (f"primary-{index}", "PrimaryTarget", "OpposingSide", "Alive", "Present"),
            (f"current-{index}", "CurrentSubject", "AnySide", "Alive", "Present"),
            (f"release-{index}", "Owner", "OpposingSide", "Any", "Any"),
        ]
        for key, origin, side, life, presence in selector_specs:
            add(
                "Selector",
                selector(
                    selectors[key],
                    origin,
                    side,
                    life=life,
                    presence=presence,
                    minimum=0,
                    maximum=1,
                    empty="NoOp",
                ),
            )
            add(
                "SelectorPredicate",
                {
                    "selector_id": selectors[key],
                    "sequence": 1,
                    "predicate": json_cell(
                        "FormationRange",
                        minimum_index=index,
                        maximum_index=index,
                    ),
                },
            )
        for prefix, origin, owner_selector in [
            ("actor-tomb", "Actor", selectors["actor"]),
            ("owner-tomb", "Owner", selectors["owner"]),
        ]:
            key = f"{prefix}-{index}"
            add(
                "Selector",
                selector(
                    selectors[key],
                    origin,
                    "SameSide",
                    life="Any",
                    presence="Any",
                    minimum=0,
                    maximum=1,
                    empty="NoOp",
                ),
            )
            for sequence, predicate in enumerate(
                [
                    json_cell("OwnedBy", owner_selector_id=owner_selector),
                    json_cell(
                        "FormationRange",
                        minimum_index=8 + index,
                        maximum_index=8 + index,
                    ),
                ],
                start=1,
            ):
                add(
                    "SelectorPredicate",
                    {
                        "selector_id": selectors[key],
                        "sequence": sequence,
                        "predicate": predicate,
                    },
                )

    def expr(name: str, kind: str, node: str) -> int:
        nonlocal next_expression
        id_ = next_expression
        next_expression += 1
        add(
            "ValueExpression",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s09.expression.{name}",
                "result_kind": kind,
                "node": node,
            },
        )
        return id_

    def binary(name: str, kind: str, operator: str, left: int, right: int) -> int:
        return expr(
            name,
            kind,
            json_cell(
                "CheckedBinary",
                operator=operator,
                left_expression_id=left,
                right_expression_id=right,
                rounding="NearestTiesAway",
            ),
        )

    actor_atk = expr(
        "actor-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    ratios = {
        name: expr(
            f"scalar-{name}",
            "Scalar",
            json_cell("ScalarLiteral", value_decimal=value),
        )
        for name, value in [
            ("zero", "0"),
            ("twelve-percent", "0.12"),
            ("one-half", "0.5"),
            ("one", "1"),
            ("two-point-five", "2.5"),
            ("four", "4"),
            ("energy-cap", "1000000"),
        ]
    }
    integers = {
        name: expr(
            f"integer-{name}",
            "Integer",
            json_cell("IntegerLiteral", value=value),
        )
        for name, value in [
            ("negative-one", -1),
            ("one", 1),
            ("two", 2),
            ("five", 5),
        ]
    }
    actor_obituary_stacks = expr(
        "actor-obituary-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["actor"],
            effect_id=effects["obituary-permanent"],
        ),
    )
    owner_obituary_stacks = expr(
        "owner-obituary-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["owner"],
            effect_id=effects["obituary-permanent"],
        ),
    )
    initial_bars_actor = binary(
        "initial-bars-actor",
        "Integer",
        "CheckedAdd",
        integers["five"],
        actor_obituary_stacks,
    )
    initial_bars_owner = binary(
        "initial-bars-owner",
        "Integer",
        "CheckedAdd",
        integers["five"],
        owner_obituary_stacks,
    )
    sunset_stacks = expr(
        "sunset-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["owner"],
            effect_id=effects["sunset"],
        ),
    )
    nightfall_stacks = expr(
        "nightfall-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["owner"],
            effect_id=effects["nightfall"],
        ),
    )
    tomb_bar_stacks = expr(
        "tomb-bar-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["owner"],
            effect_id=effects["tomb-bars"],
        ),
    )
    damage = {
        "funereal-kiss": binary(
            "funereal-kiss-damage",
            "Scalar",
            "CheckedMultiply",
            actor_atk,
            ratios["four"],
        ),
        "sunken-rain": binary(
            "sunken-rain-damage",
            "Scalar",
            "CheckedMultiply",
            actor_atk,
            ratios["two-point-five"],
        ),
    }
    release_heals: dict[int, int] = {}
    for index in range(4):
        maximum_hp = expr(
            f"release-{index}-maximum-hp",
            "Scalar",
            json_cell(
                "QueryStat",
                subject_selector_id=selectors[f"release-{index}"],
                stat="Hp",
                formula_purpose="Healing",
            ),
        )
        release_heals[index] = binary(
            f"release-{index}-half-hp",
            "Scalar",
            "CheckedMultiply",
            maximum_hp,
            ratios["one-half"],
        )
    tomb_max_hp = expr(
        "tomb-maximum-hp",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["owner"],
            stat="Hp",
            formula_purpose="OrdinaryDamage",
        ),
    )

    add(
        "ConditionExpression",
        {
            "id": conditions["always"],
            "stable_key": "goal07.enemy.s09.condition.always",
            "node": json_cell("Constant", value=True),
        },
    )
    for key, left, comparison, right in [
        ("sunset-ready", sunset_stacks, "GreaterOrEqual", integers["two"]),
        ("nightfall-second", nightfall_stacks, "GreaterOrEqual", integers["two"]),
        ("tomb-last", tomb_bar_stacks, "LessOrEqual", integers["one"]),
        ("tomb-more", tomb_bar_stacks, "Greater", integers["one"]),
    ]:
        add(
            "ConditionExpression",
            {
                "id": conditions[key],
                "stable_key": f"goal07.enemy.s09.condition.{key}",
                "node": json_cell(
                    "Compare",
                    left_expression_id=left,
                    comparison=comparison,
                    right_expression_id=right,
                ),
            },
        )
    for key, inner in [
        ("sunset-not-ready", "sunset-ready"),
        ("nightfall-not-second", "nightfall-second"),
    ]:
        add(
            "ConditionExpression",
            {
                "id": conditions[key],
                "stable_key": f"goal07.enemy.s09.condition.{key}",
                "node": json_cell(
                    "Not",
                    condition_id=conditions[inner],
                ),
            },
        )
    add(
        "ConditionExpression",
        {
            "id": conditions["no-dreams"],
            "stable_key": "goal07.enemy.s09.condition.no-dreams",
            "node": json_cell(
                "SelectorCardinality",
                selector_id=selectors["dreaming-opponents"],
                minimum_count=0,
                maximum_count=0,
            ),
        },
    )
    for index in range(4):
        for prefix in ["primary", "current"]:
            add(
                "ConditionExpression",
                {
                    "id": conditions[f"{prefix}-{index}"],
                    "stable_key": (
                        f"goal07.enemy.s09.condition.{prefix}-formation-{index}"
                    ),
                    "node": json_cell(
                        "SelectorCardinality",
                        selector_id=selectors[f"{prefix}-{index}"],
                        minimum_count=1,
                        maximum_count=1,
                    ),
                },
            )

    def new_operation(
        name: str,
        payload: str,
        target: int | None = None,
        empty: str = "Fault",
    ) -> int:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        row = operation(id_, name, payload, target, empty)
        row["stable_key"] = f"goal07.enemy.s09.operation.{name}"
        add("Operation", row)
        return id_

    def op_step(operation_id: int) -> str:
        return json_cell("Operation", operation_id=operation_id)

    def make_program(name: str, steps: list[str]) -> int:
        nonlocal next_program
        id_ = next_program
        next_program += 1
        identities.append(
            ident(
                id_,
                f"program.goal07.something-unto-death.{name}",
                "Program",
                f"Something Unto Death {name} Program",
                f"何物朝向死亡{name}程序",
                "Ordered Rule IR program for the S09 enemy.",
            )
        )
        add("Program", {"id": id_, "domain": "Battle"})
        for sequence, step in enumerate(steps, start=1):
            add(
                "ProgramStep",
                {"program_id": id_, "sequence": sequence, "step": step},
            )
        return id_

    def apply_effect_op(
        name: str,
        effect_id: int,
        target: int,
        *,
        stacks: int | None = None,
        empty: str = "Fault",
    ) -> int:
        return new_operation(
            name,
            json_cell(
                "ApplyEffect",
                effect_id=effect_id,
                stacks_expression_id=stacks,
                chance_policy="Guaranteed",
                base_chance_expression_id=None,
                rng_purpose_key=None,
            ),
            target,
            empty,
        )

    def remove_effect_op(name: str, effect_id: int, target: int) -> int:
        return new_operation(
            name,
            json_cell("RemoveEffect", effect_id=effect_id),
            target,
            "NoOp",
        )

    def modify_effect_op(
        name: str,
        effect_id: int,
        delta: int,
        target: int,
    ) -> int:
        return new_operation(
            name,
            json_cell(
                "ModifyEffect",
                effect_id=effect_id,
                stack_delta_expression_id=delta,
            ),
            target,
            "NoOp",
        )

    def summon_op(name: str, tomb_id: int, owner_selector: int) -> int:
        return new_operation(
            name,
            json_cell(
                "Summon",
                unit_definition_identity_id=tomb_id,
                owner_selector_id=owner_selector,
            ),
        )

    def capture_program(
        name: str,
        formation_prefix: str,
        tomb_prefix: str,
        summon_owner_selector: int,
        initial_bars: int,
    ) -> tuple[int, list[int]]:
        formation_programs: list[int] = []
        for index in range(4):
            target = selectors[f"{formation_prefix}-{index}"]
            tomb_target = selectors[f"{tomb_prefix}-{index}"]
            formation_programs.append(
                make_program(
                    f"{name}-formation-{index}",
                    [
                        op_step(
                            apply_effect_op(
                                f"{name}-{index}-morbid-dream",
                                effects["morbid-dream"],
                                target,
                                empty="NoOp",
                            )
                        ),
                        op_step(
                            new_operation(
                                f"{name}-{index}-untargetable",
                                json_cell("ChangePresence", presence="Untargetable"),
                                target,
                                "NoOp",
                            )
                        ),
                        op_step(
                            summon_op(
                                f"{name}-{index}-summon-tomb",
                                tombs[f"formation-{index}"],
                                summon_owner_selector,
                            )
                        ),
                        op_step(
                            apply_effect_op(
                                f"{name}-{index}-tomb-bars",
                                effects["tomb-bars"],
                                tomb_target,
                                stacks=initial_bars,
                                empty="NoOp",
                            )
                        ),
                        op_step(
                            apply_effect_op(
                                f"{name}-{index}-tomb-marker",
                                effects[f"tomb-marker-{index}"],
                                tomb_target,
                                empty="NoOp",
                            )
                        ),
                    ],
                )
            )
        root = make_program(
            name,
            [
                json_cell(
                    "If",
                    condition_id=conditions[f"{formation_prefix}-{index}"],
                    then_program_id=formation_programs[index],
                    else_program_id=None,
                )
                for index in range(4)
            ],
        )
        return root, formation_programs

    capture_primary, _capture_primary_formations = capture_program(
        "capture-primary",
        "primary",
        "actor-tomb",
        selectors["actor"],
        initial_bars_actor,
    )
    capture_current_actor, _capture_current_actor_formations = capture_program(
        "capture-current-actor",
        "current",
        "actor-tomb",
        selectors["actor"],
        initial_bars_actor,
    )
    capture_current_owner, capture_current_owner_formations = capture_program(
        "capture-current-owner",
        "current",
        "owner-tomb",
        selectors["owner"],
        initial_bars_owner,
    )

    funereal_program = make_program(
        "funereal-kiss",
        [
            op_step(
                new_operation(
                    "funereal-kiss-damage",
                    json_cell(
                        "Damage",
                        amount_expression_id=damage["funereal-kiss"],
                        damage_class="Ordinary",
                        element="Physical",
                        can_crit=True,
                    ),
                    selectors["primary"],
                )
            )
        ],
    )
    rain_program = make_program(
        "sunken-rain",
        [
            op_step(
                new_operation(
                    "sunken-rain-damage",
                    json_cell(
                        "Damage",
                        amount_expression_id=damage["sunken-rain"],
                        damage_class="Ordinary",
                        element="Physical",
                        can_crit=True,
                    ),
                    selectors["all-opposing"],
                )
            )
        ],
    )
    obituary_program = make_program(
        "harrowing-obituary",
        [
            op_step(
                apply_effect_op(
                    "harrowing-obituary-permanent-stack",
                    effects["obituary-permanent"],
                    selectors["actor"],
                    stacks=integers["one"],
                )
            ),
            op_step(
                apply_effect_op(
                    "harrowing-obituary-burst",
                    effects["obituary-burst"],
                    selectors["actor"],
                )
            ),
        ],
    )
    fading_program = make_program(
        "fading-radiance",
        [
            op_step(
                apply_effect_op(
                    "fading-radiance-sunset",
                    effects["sunset"],
                    selectors["actor"],
                    stacks=integers["one"],
                )
            )
        ],
    )
    good_night_program = make_program(
        "go-into-that-good-night",
        [
            op_step(
                new_operation(
                    "go-into-that-good-night-fill-energy",
                    json_cell(
                        "ModifyResource",
                        resource_kind="Energy",
                        character_resource_key=None,
                        update_kind="Set",
                        amount_expression_id=ratios["energy-cap"],
                        scales_with_energy_regeneration=False,
                        rounding="Floor",
                    ),
                    selectors["all-opposing"],
                )
            ),
            op_step(
                apply_effect_op(
                    "go-into-that-good-night-nightfall",
                    effects["nightfall"],
                    selectors["actor"],
                    stacks=integers["one"],
                )
            ),
        ],
    )
    watery_fallback = make_program(
        "watery-dissolution-fallback",
        [
            json_cell(
                "ForEach",
                selector_id=selectors["random-two"],
                body_program_id=capture_current_actor,
                maximum_iterations=2,
            )
        ],
    )
    watery_program = make_program(
        "watery-dissolution",
        [
            json_cell(
                "If",
                condition_id=conditions["no-dreams"],
                then_program_id=watery_fallback,
                else_program_id=None,
            )
        ],
    )
    inert_program = make_program(
        "capture-lifecycle-passive",
        [
            op_step(
                remove_effect_op(
                    "capture-lifecycle-passive-noop",
                    effects["morbid-dream"],
                    selectors["actor"],
                )
            )
        ],
    )
    ability_programs = {
        abilities["funereal-kiss"]: funereal_program,
        abilities["sunken-rain"]: rain_program,
        abilities["harrowing-obituary"]: obituary_program,
        abilities["fading-radiance"]: fading_program,
        abilities["losing-eventide-light"]: capture_primary,
        abilities["go-into-that-good-night"]: good_night_program,
        abilities["watery-dissolution"]: watery_program,
        abilities["capture-insert"]: inert_program,
        abilities["sad-true-lover"]: inert_program,
    }
    target_patterns = {
        "funereal-kiss": ("SingleTarget", 5),
        "sunken-rain": ("Aoe", 5),
        "harrowing-obituary": ("None", 4),
        "fading-radiance": ("None", 4),
        "losing-eventide-light": ("SingleTarget", 4),
        "go-into-that-good-night": ("Aoe", 4),
        "watery-dissolution": ("Aoe", 4),
        "capture-insert": ("None", 4),
        "sad-true-lover": ("None", 4),
    }
    for key, ability_id in abilities.items():
        target_pattern, tags = target_patterns[key]
        if key in {"losing-eventide-light", "watery-dissolution"}:
            tags = 2_053
        add(
            "Ability",
            {
                "id": ability_id,
                "kind": "Passive" if key == "sad-true-lover" else "Skill",
                "target_pattern": target_pattern,
                "retarget_policy": "CancelRemaining",
                "level_cap": 1,
                "cooldown_actions": 1,
                "semantic_tags_mask": tags,
            },
        )
        add(
            "AbilityPhase",
            {
                "ability_id": ability_id,
                "sequence": 1,
                "kind": "Resolved",
                "program_identity_id": ability_programs[ability_id],
            },
        )
        add(
            "EnemyAbility",
            {
                "id": ability_id,
                "telegraph": "None",
                "cooldown_actions": 1,
                "initial_cooldown_actions": 0,
                "charge_actions": 0,
                "ai_tag": key,
            },
        )

    effect_specs = {
        "obituary-permanent": (
            "Buff",
            "NonDispellable",
            99,
            None,
            "Permanent",
            "RefreshAndAddStacks",
        ),
        "obituary-burst": (
            "Buff",
            "NonDispellable",
            1,
            integers["one"],
            "OwnerTurnEnd",
            "Refresh",
        ),
        "sunset": (
            "NeutralState",
            "NonDispellable",
            3,
            None,
            "Permanent",
            "RefreshAndAddStacks",
        ),
        "nightfall": (
            "NeutralState",
            "NonDispellable",
            3,
            None,
            "Permanent",
            "RefreshAndAddStacks",
        ),
        "morbid-dream": (
            "Control",
            "NonDispellable",
            1,
            None,
            "Permanent",
            "Replace",
        ),
        "tomb-bars": (
            "NeutralState",
            "NonDispellable",
            99,
            None,
            "Permanent",
            "Replace",
        ),
    }
    for index in range(4):
        effect_specs[f"tomb-marker-{index}"] = (
            "Mark",
            "NonDispellable",
            1,
            None,
            "Permanent",
            "Replace",
        )
    for key, effect_id in effects.items():
        category, dispel, limit, duration, clock, policy = effect_specs[key]
        add(
            "Effect",
            {
                "id": effect_id,
                "category": category,
                "dispel_category": dispel,
                "stack_limit": limit,
                "duration_expression_id": duration,
                "duration_clock": clock,
                "tick_phase": "None",
                "stack_policy": policy,
                "snapshot_policy": "OnApplication",
                "teardown_policy": "RemoveWithOwner",
                "application_priority": 0,
            },
        )
    effect_tags = {
        "obituary-permanent": ["harrowing-obituary", "permanent-damage-growth"],
        "obituary-burst": ["harrowing-obituary", "remove-on-weakness-break"],
        "sunset": ["sunset", "remove-on-weakness-break"],
        "nightfall": [
            "nightfall",
            "weakness-protection",
            "prevents-toughness-reduction",
            "remove-on-weakness-break",
        ],
        "morbid-dream": ["morbid-dream", "blocks-normal-action"],
        "tomb-bars": ["sombrous-sepulcher", "fixed-hit-bars"],
    }
    for index in range(4):
        effect_tags[f"tomb-marker-{index}"] = [
            "sombrous-sepulcher",
            f"captured-formation-{index}",
        ]
    for key, tags in effect_tags.items():
        for sequence, tag in enumerate(tags, start=1):
            add(
                "EffectTag",
                {"effect_id": effects[key], "sequence": sequence, "tag": tag},
            )

    for key, value_expression in [
        ("obituary-permanent", ratios["twelve-percent"]),
        ("obituary-burst", ratios["one-half"]),
    ]:
        add(
            "ModifierStackingGroup",
            {
                "id": modifier_groups[key],
                "stable_key": f"goal07.enemy.s09.modifier-group.{key}",
                "aggregation": "Sum",
                "comparator_expression_id": None,
            },
        )
        add(
            "ModifierDefinition",
            {
                "id": modifiers[key],
                "source_effect_id": effects[key],
                "owner_selector_id": selectors["owner"],
                "subject_selector_id": selectors["owner"],
                "stat": "Atk",
                "formula_stage": "DamageBoost",
                "formula_purpose": "OrdinaryDamage",
                "value_expression_id": value_expression,
                "value_domain": "Ratio",
                "stacking_group_id": modifier_groups[key],
                "priority": 0,
                "cap_formula_stage": "DamageBoost",
                "snapshot_policy": "Dynamic",
                "duration_scope": "Turn",
            },
        )
        add(
            "EffectModifierBinding",
            {
                "effect_id": effects[key],
                "sequence": 1,
                "modifier_id": modifiers[key],
            },
        )

    queue_losing = new_operation(
        "sunset-queue-losing-eventide-light",
        json_cell(
            "QueueAction",
            ability_id=abilities["losing-eventide-light"],
            actor_selector_id=selectors["owner"],
            priority=100,
            forced_use=True,
            reaction_boundary="AfterAction",
            owner_policy="Actor",
            payment_policy="Suppressed",
            payment_resource_key=None,
        ),
        selectors["current-subject"],
    )
    sunset_continue = make_program(
        "sunset-continue",
        [
            op_step(
                modify_effect_op(
                    "sunset-increment",
                    effects["sunset"],
                    integers["one"],
                    selectors["owner"],
                )
            )
        ],
    )
    sunset_release = make_program(
        "sunset-release",
        [
            op_step(queue_losing),
            op_step(
                remove_effect_op(
                    "sunset-release-clear",
                    effects["sunset"],
                    selectors["owner"],
                )
            ),
        ],
    )
    sunset_action_program = make_program(
        "sunset-action-count",
        [
            json_cell(
                "If",
                condition_id=conditions["sunset-ready"],
                then_program_id=sunset_release,
                else_program_id=sunset_continue,
            )
        ],
    )
    queue_watery = new_operation(
        "nightfall-queue-watery-dissolution",
        json_cell(
            "QueueAction",
            ability_id=abilities["watery-dissolution"],
            actor_selector_id=selectors["owner"],
            priority=100,
            forced_use=True,
            reaction_boundary="AfterAction",
            owner_policy="Actor",
            payment_policy="Suppressed",
            payment_resource_key=None,
        ),
        selectors["all-opposing"],
    )
    nightfall_continue = make_program(
        "nightfall-first-actor",
        [
            op_step(
                modify_effect_op(
                    "nightfall-increment",
                    effects["nightfall"],
                    integers["one"],
                    selectors["owner"],
                )
            ),
        ],
    )
    nightfall_release = make_program(
        "nightfall-second-actor",
        [
            op_step(queue_watery),
            op_step(
                remove_effect_op(
                    "nightfall-release-clear",
                    effects["nightfall"],
                    selectors["owner"],
                )
            ),
        ],
    )
    nightfall_action_program = make_program(
        "nightfall-action-count",
        [
            json_cell(
                "If",
                condition_id=conditions["nightfall-second"],
                then_program_id=nightfall_release,
                else_program_id=nightfall_continue,
            )
        ],
    )
    reset_sunset_program = make_program(
        "weakness-break-clear-sunset",
        [
            op_step(
                remove_effect_op(
                    "weakness-break-clear-sunset",
                    effects["sunset"],
                    selectors["owner"],
                )
            ),
            op_step(
                remove_effect_op(
                    "weakness-break-clear-obituary-burst-from-sunset",
                    effects["obituary-burst"],
                    selectors["owner"],
                )
            ),
        ],
    )
    reset_nightfall_program = make_program(
        "weakness-break-clear-nightfall",
        [
            op_step(
                remove_effect_op(
                    "weakness-break-clear-nightfall",
                    effects["nightfall"],
                    selectors["owner"],
                )
            ),
            op_step(
                remove_effect_op(
                    "weakness-break-clear-obituary-burst-from-nightfall",
                    effects["obituary-burst"],
                    selectors["owner"],
                )
            ),
        ],
    )
    defeat_tomb_program = make_program(
        "tomb-last-bar",
        [
            op_step(
                new_operation(
                    "tomb-last-bar-consume-hp",
                    json_cell(
                        "ConsumeHp",
                        amount_expression_id=tomb_max_hp,
                        floor_expression_id=ratios["zero"],
                    ),
                    selectors["owner"],
                )
            )
        ],
    )
    decrement_tomb_program = make_program(
        "tomb-decrement-bar",
        [
            op_step(
                modify_effect_op(
                    "tomb-decrement-bar",
                    effects["tomb-bars"],
                    integers["negative-one"],
                    selectors["owner"],
                )
            )
        ],
    )
    release_programs: dict[int, int] = {}
    for index in range(4):
        target = selectors[f"release-{index}"]
        release_programs[index] = make_program(
            f"tomb-release-formation-{index}",
            [
                op_step(
                    new_operation(
                        f"tomb-release-{index}-present",
                        json_cell("ChangePresence", presence="Present"),
                        target,
                        "NoOp",
                    )
                ),
                op_step(
                    remove_effect_op(
                        f"tomb-release-{index}-clear-dream",
                        effects["morbid-dream"],
                        target,
                    )
                ),
                op_step(
                    new_operation(
                        f"tomb-release-{index}-heal",
                        json_cell(
                            "Heal",
                            amount_expression_id=release_heals[index],
                        ),
                        target,
                        "NoOp",
                    )
                ),
                op_step(
                    new_operation(
                        f"tomb-release-{index}-energy",
                        json_cell(
                            "ModifyResource",
                            resource_kind="Energy",
                            character_resource_key=None,
                            update_kind="Set",
                            amount_expression_id=ratios["energy-cap"],
                            scales_with_energy_regeneration=False,
                            rounding="Floor",
                        ),
                        target,
                        "NoOp",
                    )
                ),
            ],
        )

    rule_specs = [
        (
            "sunset-actions",
            effects["sunset"],
            json_cell("Action", point="Resolved"),
            {"actor_selector_id": selectors["current-subject"]},
            [
                (conditions["sunset-ready"], sunset_release, "Event"),
                (conditions["sunset-not-ready"], sunset_continue, "Event"),
            ],
        ),
        (
            "sunset-break",
            effects["sunset"],
            json_cell("WeaknessBroken"),
            {"target_selector_id": selectors["owner"]},
            [(conditions["always"], reset_sunset_program, "Event")],
        ),
        (
            "nightfall-actions",
            effects["nightfall"],
            json_cell("Action", point="Resolved"),
            {"actor_selector_id": selectors["current-subject"]},
            [
                *[
                    (
                        conditions[f"current-{index}"],
                        capture_current_owner_formations[index],
                        "Event",
                    )
                    for index in range(4)
                ],
                (
                    conditions["nightfall-not-second"],
                    nightfall_continue,
                    "Event",
                ),
                (
                    conditions["nightfall-second"],
                    nightfall_release,
                    "Event",
                ),
            ],
        ),
        (
            "nightfall-break",
            effects["nightfall"],
            json_cell("WeaknessBroken"),
            {"target_selector_id": selectors["owner"]},
            [(conditions["always"], reset_nightfall_program, "Event")],
        ),
        (
            "tomb-bars",
            effects["tomb-bars"],
            json_cell("Damage", point="Applied"),
            {"target_selector_id": selectors["owner"]},
            [
                (conditions["tomb-last"], defeat_tomb_program, "Event"),
                (conditions["tomb-more"], decrement_tomb_program, "Event"),
            ],
        ),
    ]
    for index in range(4):
        rule_specs.append(
            (
                f"tomb-release-{index}",
                effects[f"tomb-marker-{index}"],
                json_cell("Unit", point="Defeated"),
                {"target_selector_id": selectors["owner"]},
                [(conditions["always"], release_programs[index], "Event")],
            )
        )
    effect_rule_sequences: dict[int, int] = {}
    for key, source_effect, event, filter_fields, triggers in rule_specs:
        add(
            "RuleDefinition",
            {
                "id": rules[key],
                "domain": "Battle",
                "source_definition_identity_id": source_effect,
                "source_class": "Effect",
                "source_digest_sha256": sha256_text(f"goal07-s09-{key}-v1"),
            },
        )
        add(
            "EventFilter",
            {
                "id": filters[key],
                "stable_key": f"goal07.enemy.s09.filter.{key}",
                **filter_fields,
                "cause_ancestry": "Any",
            },
        )
        for sequence, (condition_id, program_id, once_scope) in enumerate(
            triggers, start=1
        ):
            add(
                "RuleTrigger",
                {
                    "id": next_trigger,
                    "stable_key": (
                        f"goal07.enemy.s09.trigger.{key}-{sequence}"
                    ),
                    "rule_id": rules[key],
                    "sequence": sequence,
                    "event": event,
                    "phase": "AfterEvent",
                    "filter_id": filters[key],
                    "condition_id": condition_id,
                    "once_scope": once_scope,
                    "priority": 0,
                    "program_id": program_id,
                },
            )
            next_trigger += 1
        binding_sequence = effect_rule_sequences.get(source_effect, 0) + 1
        effect_rule_sequences[source_effect] = binding_sequence
        add(
            "EffectRuleBinding",
            {
                "effect_id": source_effect,
                "sequence": binding_sequence,
                "rule_id": rules[key],
            },
        )
    add(
        "RuleTrigger",
        {
            "id": next_trigger,
            "stable_key": "goal07.enemy.s09.trigger.tomb-bars-weakness-last",
            "rule_id": rules["tomb-bars"],
            "sequence": 3,
            "event": json_cell("WeaknessBroken"),
            "phase": "AfterEvent",
            "filter_id": filters["tomb-bars"],
            "condition_id": conditions["tomb-last"],
            "once_scope": "Event",
            "priority": 0,
            "program_id": defeat_tomb_program,
        },
    )
    next_trigger += 1
    add(
        "RuleTrigger",
        {
            "id": next_trigger,
            "stable_key": "goal07.enemy.s09.trigger.tomb-bars-weakness-more",
            "rule_id": rules["tomb-bars"],
            "sequence": 4,
            "event": json_cell("WeaknessBroken"),
            "phase": "AfterEvent",
            "filter_id": filters["tomb-bars"],
            "condition_id": conditions["tomb-more"],
            "once_scope": "Event",
            "priority": 0,
            "program_id": decrement_tomb_program,
        },
    )

    phase_sequences = [
        [
            abilities["funereal-kiss"],
            abilities["fading-radiance"],
            abilities["harrowing-obituary"],
            abilities["sunken-rain"],
        ],
        [
            abilities["funereal-kiss"],
            abilities["go-into-that-good-night"],
            abilities["harrowing-obituary"],
            abilities["sunken-rain"],
        ],
        [
            abilities["fading-radiance"],
            abilities["funereal-kiss"],
            abilities["go-into-that-good-night"],
            abilities["harrowing-obituary"],
            abilities["sunken-rain"],
        ],
    ]
    ai_targets = {
        abilities["funereal-kiss"]: selectors["primary"],
        abilities["fading-radiance"]: selectors["actor"],
        abilities["harrowing-obituary"]: selectors["actor"],
        abilities["sunken-rain"]: selectors["all-opposing"],
        abilities["go-into-that-good-night"]: selectors["all-opposing"],
    }
    next_candidate = BASE + 801
    next_transition = BASE + 901
    for phase_index, sequence_abilities in enumerate(phase_sequences):
        state_ids = [
            BASE + 701 + phase_index * 20 + offset
            for offset in range(len(sequence_abilities))
        ]
        add(
            "AiGraph",
            {
                "id": graphs[phase_index],
                "initial_state_id": state_ids[0],
                "automatic_transition_budget": 8,
            },
        )
        for offset, (state_id, ability_id) in enumerate(
            zip(state_ids, sequence_abilities, strict=True)
        ):
            add(
                "AiState",
                {
                    "id": state_id,
                    "stable_key": (
                        f"goal07.enemy.s09.ai.phase-{phase_index + 1}."
                        f"state-{offset + 1}"
                    ),
                    "graph_id": graphs[phase_index],
                    "mandatory_fallback_ability_id": abilities["funereal-kiss"],
                    "turn_counter_reset": offset == 0,
                },
            )
            add(
                "AiCandidate",
                {
                    "id": next_candidate,
                    "stable_key": (
                        f"goal07.enemy.s09.ai.phase-{phase_index + 1}."
                        f"candidate-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "ability_id": ability_id,
                    "condition_id": conditions["always"],
                    "target_selector_id": ai_targets[ability_id],
                    "priority": 0,
                    "selection": "FirstLegal",
                    "no_target_fallback": "UseFallbackAbility",
                    "fallback_ability_id": abilities["funereal-kiss"],
                },
            )
            next_candidate += 1
            add(
                "AiTransition",
                {
                    "id": next_transition,
                    "stable_key": (
                        f"goal07.enemy.s09.ai.phase-{phase_index + 1}."
                        f"transition-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "target_state_id": state_ids[(offset + 1) % len(state_ids)],
                    "condition_id": conditions["always"],
                    "priority": 0,
                    "timing": "AfterAction",
                },
            )
            next_transition += 1

    for index, tomb_id in enumerate(tombs.values()):
        add(
            "LinkedUnitDefinition",
            {
                "id": tomb_id,
                "source_definition_identity_id": tomb_id,
                "kind": "Summon",
                "presence": "Present",
                "ability_ids": str(abilities["capture-insert"]),
                "action_ability_id": None,
                "formation_index": 8 + index,
                "initial_gauge_decimal": "10000",
                "hp_owner_ratio_decimal": "10",
                "hp_flat_decimal": "0",
                "atk_owner_ratio_decimal": "1",
                "atk_flat_decimal": "0",
                "def_owner_ratio_decimal": "1",
                "def_flat_decimal": "0",
                "spd_owner_ratio_decimal": "0",
                "spd_flat_decimal": "200",
                "owner_defeat_policy": "Depart",
                "owner_departure_policy": "Depart",
                "wave_policy": "Depart",
                "combatant_digest_sha256": sha256_text(
                    f"goal07-s09-sombrous-sepulcher-{index}-v1"
                ),
            },
        )

    add(
        "EnemyTemplate",
        {
            "id": template,
            "rank": "Boss",
            "base_aggro_decimal": "100",
            "default_ai_graph_id": graphs[0],
        },
    )
    add(
        "EnemyVariant",
        {
            "id": variant,
            "template_id": template,
            "ai_graph_id": graphs[0],
            "mechanically_distinct_key": VARIANT_KEY,
        },
    )
    for level in anchor["levels"]:
        add(
            "EnemyStat",
            {
                "variant_id": variant,
                "level": level["authored_level"],
                "difficulty_key": "standard-universe-v1",
                "hp_decimal": level["base_hp"],
                "atk_decimal": level["base_atk"],
                "def_decimal": level["base_def"],
                "spd_decimal": level["base_spd"],
                "effect_hit_rate_decimal": level["effect_hit_rate"],
                "effect_resistance_decimal": level["effect_resistance"],
                "crit_damage_decimal": "0.2",
            },
        )
    for sequence, weakness in enumerate(["Fire", "Wind", "Imaginary"], start=1):
        add(
            "EnemyWeakness",
            {"variant_id": variant, "sequence": sequence, "element": weakness},
        )
    for element in ["Physical", "Ice", "Lightning", "Quantum"]:
        add(
            "EnemyResistance",
            {"variant_id": variant, "element": element, "value_decimal": "0.2"},
        )
    for category in ["STAT_CTRL_Frozen", "STAT_Entangle"]:
        add(
            "EnemyDebuffResistance",
            {
                "variant_id": variant,
                "category_key": category,
                "value_decimal": "0.5",
            },
        )
    add(
        "EnemyToughnessLayer",
        {
            "variant_id": variant,
            "sequence": 1,
            "layer_key": "ordinary",
            "kind": "Ordinary",
            "maximum_decimal": "720",
            "recovery_ratio_decimal": "1",
            "active_at_start": True,
        },
    )
    for sequence, ability_id in enumerate(abilities.values(), start=1):
        add(
            "EnemyVariantAbility",
            {
                "variant_id": variant,
                "sequence": sequence,
                "ability_id": ability_id,
            },
        )
    for sequence, graph_id in enumerate(graphs, start=1):
        add(
            "EnemyPhase",
            {
                "id": BASE + 600 + sequence,
                "stable_key": f"goal07.enemy.s09.phase-{sequence}",
                "variant_id": variant,
                "sequence": sequence,
                "entry_condition_id": conditions["always"],
                "exit_condition_id": conditions["always"],
                "replacement_priority": sequence,
                "ai_graph_id": graph_id,
                "targetable": True,
                "transition_model": "TransformSameUnit",
                "hp_carry": "Reset",
                "action_gauge_carry": "Reset",
                "effect_carry": "Clear",
                "toughness_carry": "Reset",
                "summon_carry": "Clear",
            },
        )

    anchor_digest = sha256_bytes(anchor_path(PARTITION).read_bytes())
    add(
        "SourceRecord",
        {
            "id": SOURCE_RECORD_ID,
            "stable_key": "source.hsr-wiki.something-unto-death-complete.2026-07-29",
            "category": "CommunityMaintained",
            "publisher": anchor["source"]["publisher"],
            "url": anchor["source"]["url"],
            "accessed_on": anchor["source"]["accessed_on"],
            "applicable_game_version": anchor["source"]["game_version"],
            "confidence": "SecondaryVersionSensitiveCrossCheck",
            "evidence_sha256": anchor_digest,
            "usage_note": (
                "Exact public World 9 level values and capture mechanics are "
                "committed as Goal 07 evidence."
            ),
        },
    )
    add(
        "EvidenceRecord",
        {
            "id": EVIDENCE_RECORD_ID,
            "stable_key": "evidence.goal07.enemy.s09.numeric-anchors",
            "kind": "SourcePayload",
            "source_record_id": SOURCE_RECORD_ID,
            "sha256": anchor_digest,
            "note": "Committed exact public per-level numeric anchors for Goal 07 S09.",
        },
    )
    for item in identities:
        add("ContentIdentity", item)
        add(
            "ContentEvidenceBinding",
            {
                "content_id": item["id"],
                "sequence": 1,
                "fact_key": f"goal07.s09.executable:{item['stable_key']}",
                "source_record_id": 1,
                "evidence_record_id": 3,
                "quality": "ExactStructured",
                "mechanism_quality": "ExactStructured",
            },
        )
    add(
        "ContentEvidenceBinding",
        {
            "content_id": variant,
            "sequence": 2,
            "fact_key": "goal07.s09.public-level-stats",
            "source_record_id": SOURCE_RECORD_ID,
            "evidence_record_id": EVIDENCE_RECORD_ID,
            "quality": "ExactStructured",
            "mechanism_quality": "ExactStructured",
        },
    )
    for table_rows in rows.values():
        table_rows.sort(
            key=lambda row: json.dumps(
                row, ensure_ascii=False, sort_keys=True, default=str
            )
        )
    return rows


def owned_rows_s10() -> dict[str, list[dict[str, Any]]]:
    anchor = json.loads(anchor_path(PARTITION).read_text(encoding="utf-8"))
    manifest = json.loads(PARTITIONS.read_text(encoding="utf-8"))
    assigned = next(item for item in manifest["partitions"] if item["id"] == PARTITION)
    if assigned["enemy_variant_ids"] != [VARIANT_KEY]:
        raise ValueError("S10 frozen enemy assignment changed")

    variant, template = BASE + 1, BASE + 2
    graphs = [BASE + 10, BASE + 11, BASE + 12]
    abilities = {
        "midnight": BASE + 101,
        "caressing": BASE + 102,
        "mockery": BASE + 103,
        "spirit": BASE + 104,
        "psychological": BASE + 105,
        "seething": BASE + 106,
        "tremble": BASE + 107,
        "cruelty": BASE + 108,
        "griever": BASE + 109,
        "revelation": BASE + 110,
    }
    selectors = {
        "actor": BASE + 401,
        "owner": BASE + 402,
        "primary": BASE + 403,
        "adjacent": BASE + 404,
        "all-opposing": BASE + 405,
        "random-two": BASE + 406,
        "suggested": BASE + 407,
        "shocked-primary": BASE + 408,
        "shocked-other": BASE + 409,
        "current-shocked": BASE + 410,
        "actor-other": BASE + 411,
    }
    effects = {
        "shock": BASE + 501,
        "dominated": BASE + 502,
        "suggestion": BASE + 503,
        "cruelty-listener": BASE + 504,
        "griever": BASE + 505,
    }
    modifier, modifier_group = BASE + 521, BASE + 531
    cruelty_rule, cruelty_filter, cruelty_trigger = BASE + 541, BASE + 641, BASE + 671
    conditions = {"always": BASE + 551, "primary-shocked": BASE + 552}
    rows: dict[str, list[dict[str, Any]]] = {}
    identities: list[dict[str, Any]] = []
    next_program, next_operation, next_expression = BASE + 301, BASE + 1_001, BASE + 1_101

    def add(table: str, row: dict[str, Any]) -> None:
        rows.setdefault(table, []).append(row)

    def ident(
        id_: int,
        key: str,
        kind: str,
        name_en: str,
        name_zh_cn: str,
        summary: str,
        sources: str = "1",
    ) -> dict[str, Any]:
        row = identity(id_, key, kind, name_en, name_zh_cn, summary, sources)
        row["summary_zh_cn"] = "Goal 07 S10 来源绑定的「星核猎手」卡芙卡可执行定义。"
        row["game_version_introduced"] = "1.0"
        return row

    identities.extend(
        [
            ident(
                variant,
                VARIANT_KEY,
                "EnemyVariant",
                "Stellaron Hunter: Kafka (Complete)",
                "「星核猎手」卡芙卡（完整）",
                "Exact three-phase World 5 boss materialization.",
                "1|12",
            ),
            ident(
                template,
                "enemy.stellaron-hunter-kafka-complete.littleboss",
                "Enemy",
                "Stellaron Hunter: Kafka (Complete) Template",
                "「星核猎手」卡芙卡（完整）模板",
                "Version 4.4 boss template retained from source monster 2004011.",
            ),
        ]
    )
    for phase, graph in enumerate(graphs, start=1):
        identities.append(
            ident(
                graph,
                f"ai.goal07.stellaron-hunter-kafka-complete.phase-{phase}",
                "AiGraph",
                f"Kafka Complete Phase {phase} AI",
                f"卡芙卡完整形态第{phase}阶段AI",
                "Finite source-phase rotation with explicit control and Shock behavior.",
            )
        )
    ability_metadata = {
        "midnight": ("Midnight Tumult", "夜间喧嚣不止", "200401101"),
        "caressing": ("Caressing Moonlight", "月光摩挲连绵", "200401102"),
        "mockery": ("Silent and Sharp Mockery", "缄默厉声嘲笑", "200401103"),
        "spirit": ("Spirit Whisper", "言灵", "200401104"),
        "psychological": ("Psychological Suggestion", "心理暗示", "200401106"),
        "seething": ("Seething Whisper of the Fallen", "亡者切齿呢喃", "200401105"),
        "tremble": ("Tremble", "颤慄", "200401107"),
        "cruelty": ("Cruelty", "残酷", "200401108"),
        "griever": ("Griever", "吊唁", "200401109"),
        "revelation": ("Revelation", "启示", "derived-psychological-suggestion"),
    }
    for key, id_ in abilities.items():
        name_en, name_zh_cn, source = ability_metadata[key]
        identities.append(
            ident(
                id_,
                f"enemy.stellaron-hunter-kafka-complete.littleboss.ability.{key}",
                "Ability",
                name_en,
                name_zh_cn,
                f"Executable transcription of source skill {source}.",
            )
        )
    for key, id_ in selectors.items():
        identities.append(
            ident(
                id_,
                f"selector.goal07.stellaron-hunter-kafka-complete.{key}",
                "Selector",
                f"Kafka Complete {key} Selector",
                f"卡芙卡完整形态{key}选择器",
                "S10 battle selector.",
            )
        )
    effect_metadata = {
        "shock": ("Kafka Shock", "卡芙卡触电", "Two-turn 300% Lightning damage-over-time state."),
        "dominated": ("Dominated", "支配", "Forces the target to Basic ATK a random ally."),
        "suggestion": ("Psychological Suggestion", "心理暗示", "Marks targets for Revelation."),
        "cruelty-listener": ("Cruelty Listener", "残酷监听", "Once-per-turn shocked-target follow-up listener."),
        "griever": ("Griever Growth", "吊唁成长", "Stacking two-percent Kafka damage growth."),
    }
    for key, id_ in effects.items():
        name_en, name_zh_cn, summary = effect_metadata[key]
        identities.append(
            ident(
                id_,
                f"effect.goal07.stellaron-hunter-kafka-complete.{key}",
                "Effect",
                name_en,
                name_zh_cn,
                summary,
            )
        )
    identities.extend(
        [
            ident(
                modifier,
                "modifier.goal07.stellaron-hunter-kafka-complete.griever",
                "Modifier",
                "Kafka Griever Damage Growth",
                "卡芙卡吊唁增伤",
                "Two-percent dynamic damage boost per Griever stack.",
            ),
            ident(
                modifier_group,
                "modifier-group.goal07.stellaron-hunter-kafka-complete.griever",
                "Modifier",
                "Kafka Griever Stacking",
                "卡芙卡吊唁叠加组",
                "Additive stacking group for Griever.",
            ),
            ident(
                cruelty_rule,
                "rule.goal07.stellaron-hunter-kafka-complete.cruelty",
                "Rule",
                "Kafka Cruelty Follow-up",
                "卡芙卡残酷追击",
                "Once per turn an eligible non-Kafka attacker causes a Lightning follow-up.",
            ),
        ]
    )

    add("Selector", selector(selectors["actor"], "Actor", "SameSide"))
    add("Selector", selector(selectors["owner"], "Owner", "SameSide"))
    add("Selector", selector(selectors["primary"], "PrimaryTarget", "OpposingSide"))
    add(
        "Selector",
        selector(
            selectors["adjacent"],
            "Actor",
            "OpposingSide",
            minimum=0,
            maximum=2,
            empty="NoOp",
            choice="PrimaryPlusAdjacent",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["adjacent"],
            "sequence": 1,
            "predicate": json_cell("Excludes", excluded_selector_id=selectors["primary"]),
        },
    )
    add(
        "Selector",
        selector(
            selectors["all-opposing"],
            "Actor",
            "OpposingSide",
            minimum=1,
            maximum=8,
            choice="All",
        ),
    )
    random_two = selector(
        selectors["random-two"],
        "Actor",
        "OpposingSide",
        minimum=1,
        maximum=2,
        choice="RngUniform",
    )
    random_two["rng_purpose_key"] = "behavior-choice"
    add("Selector", random_two)
    for name, origin, side, minimum, maximum, empty, choice in [
        ("suggested", "Actor", "OpposingSide", 0, 8, "NoOp", "All"),
        ("shocked-primary", "PrimaryTarget", "OpposingSide", 0, 1, "NoOp", "First"),
        ("shocked-other", "Actor", "OpposingSide", 0, 8, "NoOp", "All"),
        ("current-shocked", "CurrentSubject", "OpposingSide", 0, 1, "NoOp", "First"),
        ("actor-other", "Actor", "AnySide", 0, 1, "NoOp", "First"),
    ]:
        add(
            "Selector",
            selector(
                selectors[name],
                origin,
                side,
                minimum=minimum,
                maximum=maximum,
                empty=empty,
                choice=choice,
            ),
        )
    for name, effect in [
        ("suggested", effects["suggestion"]),
        ("shocked-primary", effects["shock"]),
        ("shocked-other", effects["shock"]),
        ("current-shocked", effects["shock"]),
    ]:
        add(
            "SelectorPredicate",
            {
                "selector_id": selectors[name],
                "sequence": 1,
                "predicate": json_cell("HasEffect", effect_id=effect),
            },
        )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["shocked-other"],
            "sequence": 2,
            "predicate": json_cell("Excludes", excluded_selector_id=selectors["primary"]),
        },
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["actor-other"],
            "sequence": 1,
            "predicate": json_cell("Excludes", excluded_selector_id=selectors["owner"]),
        },
    )

    def expression(name: str, kind: str, node: str) -> int:
        nonlocal next_expression
        id_ = next_expression
        next_expression += 1
        add(
            "ValueExpression",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s10.expression.{name}",
                "result_kind": kind,
                "node": node,
            },
        )
        return id_

    def product(name: str, left: int, right: int) -> int:
        return expression(
            name,
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedMultiply",
                left_expression_id=left,
                right_expression_id=right,
                rounding="NearestTiesAway",
            ),
        )

    actor_atk = expression(
        "actor-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    ratios = {
        "three": expression("ratio-3", "Scalar", json_cell("ScalarLiteral", value_decimal="3")),
        "two-point-five": expression("ratio-2-5", "Scalar", json_cell("ScalarLiteral", value_decimal="2.5")),
        "two": expression("ratio-2", "Scalar", json_cell("ScalarLiteral", value_decimal="2")),
        "one": expression("ratio-1", "Scalar", json_cell("ScalarLiteral", value_decimal="1")),
        "one-point-two": expression("ratio-1-2", "Scalar", json_cell("ScalarLiteral", value_decimal="1.2")),
        "zero-point-zero-two": expression("ratio-0-02", "Scalar", json_cell("ScalarLiteral", value_decimal="0.02")),
    }
    integers = {
        "one": expression("integer-one", "Integer", json_cell("IntegerLiteral", value=1)),
        "two": expression("integer-two", "Integer", json_cell("IntegerLiteral", value=2)),
    }
    damage = {
        "midnight": product("midnight-damage", actor_atk, ratios["two-point-five"]),
        "caressing-primary": product("caressing-primary-damage", actor_atk, ratios["three"]),
        "caressing-adjacent": product("caressing-adjacent-damage", actor_atk, ratios["two"]),
        "mockery": product("mockery-damage", actor_atk, ratios["two-point-five"]),
        "seething-primary": product("seething-primary-damage", actor_atk, ratios["three"]),
        "seething-other": product("seething-other-damage", actor_atk, ratios["two-point-five"]),
        "cruelty": product("cruelty-damage", actor_atk, ratios["one"]),
        "shock": product("shock-damage", actor_atk, ratios["three"]),
    }
    add(
        "ConditionExpression",
        {
            "id": conditions["always"],
            "stable_key": "goal07.enemy.s10.condition.always",
            "node": json_cell("Constant", value=True),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["primary-shocked"],
            "stable_key": "goal07.enemy.s10.condition.primary-shocked",
            "node": json_cell(
                "SelectorCardinality",
                selector_id=selectors["shocked-primary"],
                minimum_count=1,
                maximum_count=1,
            ),
        },
    )

    def new_operation(
        name: str,
        payload: str,
        target: int | None = None,
        empty: str = "Fault",
    ) -> int:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        row = operation(id_, name, payload, target, empty)
        row["stable_key"] = f"goal07.enemy.s10.operation.{name}"
        add("Operation", row)
        return id_

    def op_step(id_: int) -> str:
        return json_cell("Operation", operation_id=id_)

    def damage_op(name: str, amount: int, target: int, empty: str = "Fault") -> int:
        return new_operation(
            name,
            json_cell(
                "Damage",
                amount_expression_id=amount,
                damage_class="Ordinary",
                element="Lightning",
                can_crit=True,
            ),
            target,
            empty,
        )

    def apply_op(
        name: str,
        effect: int,
        target: int,
        chance: int | None = None,
        stacks: int | None = None,
        empty: str = "Fault",
    ) -> int:
        return new_operation(
            name,
            json_cell(
                "ApplyEffect",
                effect_id=effect,
                stacks_expression_id=stacks,
                chance_policy="Resistible" if chance is not None else "Guaranteed",
                base_chance_expression_id=chance,
                rng_purpose_key="effect-application" if chance is not None else None,
            ),
            target,
            empty,
        )

    def remove_op(name: str, effect: int, target: int) -> int:
        return new_operation(
            name,
            json_cell("RemoveEffect", effect_id=effect),
            target,
            "NoOp",
        )

    def detonate_op(name: str, target: int) -> int:
        return new_operation(
            name,
            json_cell(
                "DetonateDot",
                fraction_expression_id=ratios["one"],
                required_effect_tag=None,
            ),
            target,
            "NoOp",
        )

    def make_program(name: str, steps: list[str]) -> int:
        nonlocal next_program
        id_ = next_program
        next_program += 1
        identities.append(
            ident(
                id_,
                f"program.goal07.stellaron-hunter-kafka-complete.{name}",
                "Program",
                f"Kafka Complete {name} Program",
                f"卡芙卡完整形态{name}程序",
                "Ordered shared Rule IR program for the S10 enemy.",
            )
        )
        add("Program", {"id": id_, "domain": "Battle"})
        for sequence, step in enumerate(steps, start=1):
            add("ProgramStep", {"program_id": id_, "sequence": sequence, "step": step})
        return id_

    griever_stack = lambda name: apply_op(
        name,
        effects["griever"],
        selectors["actor"],
        stacks=integers["one"],
    )
    spread_shock = make_program(
        "caressing-spread-shock",
        [
            op_step(
                apply_op(
                    "caressing-apply-adjacent-shock",
                    effects["shock"],
                    selectors["adjacent"],
                    ratios["one-point-two"],
                    empty="NoOp",
                )
            ),
            op_step(griever_stack("caressing-griever-growth")),
        ],
    )
    ability_programs = {
        abilities["midnight"]: make_program(
            "midnight-tumult",
            [
                op_step(damage_op("midnight-damage", damage["midnight"], selectors["primary"])),
                op_step(detonate_op("midnight-tremble", selectors["shocked-primary"])),
                op_step(
                    apply_op(
                        "midnight-shock",
                        effects["shock"],
                        selectors["primary"],
                        ratios["one-point-two"],
                    )
                ),
                op_step(griever_stack("midnight-griever-growth")),
            ],
        ),
        abilities["caressing"]: make_program(
            "caressing-moonlight",
            [
                op_step(
                    damage_op(
                        "caressing-primary-damage",
                        damage["caressing-primary"],
                        selectors["primary"],
                    )
                ),
                op_step(
                    damage_op(
                        "caressing-adjacent-damage",
                        damage["caressing-adjacent"],
                        selectors["adjacent"],
                        "NoOp",
                    )
                ),
                op_step(detonate_op("caressing-tremble", selectors["shocked-primary"])),
                json_cell(
                    "If",
                    condition_id=conditions["primary-shocked"],
                    then_program_id=spread_shock,
                    else_program_id=None,
                ),
            ],
        ),
        abilities["mockery"]: make_program(
            "silent-and-sharp-mockery",
            [
                op_step(
                    damage_op(
                        "mockery-damage",
                        damage["mockery"],
                        selectors["all-opposing"],
                    )
                ),
                op_step(detonate_op("mockery-tremble", selectors["shocked-other"])),
                op_step(detonate_op("mockery-primary-tremble", selectors["shocked-primary"])),
            ],
        ),
        abilities["spirit"]: make_program(
            "spirit-whisper",
            [
                op_step(
                    apply_op(
                        "spirit-dominate",
                        effects["dominated"],
                        selectors["primary"],
                        ratios["one-point-two"],
                    )
                ),
                op_step(
                    new_operation(
                        "spirit-advance",
                        json_cell("AdvanceAction", amount_expression_id=ratios["one"]),
                        selectors["primary"],
                    )
                ),
            ],
        ),
        abilities["seething"]: make_program(
            "seething-whisper-of-the-fallen",
            [
                op_step(
                    damage_op(
                        "seething-primary-damage",
                        damage["seething-primary"],
                        selectors["primary"],
                    )
                ),
                op_step(detonate_op("seething-tremble", selectors["shocked-primary"])),
                op_step(
                    apply_op(
                        "seething-shock",
                        effects["shock"],
                        selectors["primary"],
                        ratios["one-point-two"],
                    )
                ),
                op_step(griever_stack("seething-griever-growth")),
                op_step(
                    damage_op(
                        "seething-other-shocked-damage",
                        damage["seething-other"],
                        selectors["shocked-other"],
                        "NoOp",
                    )
                ),
            ],
        ),
    }
    revelation_program = make_program(
        "revelation",
        [
            op_step(
                apply_op(
                    "revelation-dominate",
                    effects["dominated"],
                    selectors["suggested"],
                    ratios["one-point-two"],
                    empty="NoOp",
                )
            ),
            op_step(
                new_operation(
                    "revelation-advance",
                    json_cell("AdvanceAction", amount_expression_id=ratios["one"]),
                    selectors["suggested"],
                    "NoOp",
                )
            ),
            op_step(remove_op("revelation-clear-suggestion", effects["suggestion"], selectors["suggested"])),
        ],
    )
    ability_programs[abilities["revelation"]] = revelation_program
    queue_revelation = new_operation(
        "psychological-queue-revelation",
        json_cell(
            "QueueAction",
            ability_id=abilities["revelation"],
            actor_selector_id=selectors["actor"],
            priority=100,
            forced_use=True,
            reaction_boundary="AfterAction",
            owner_policy="Actor",
            payment_policy="Suppressed",
            payment_resource_key=None,
        ),
        selectors["suggested"],
        "NoOp",
    )
    ability_programs[abilities["psychological"]] = make_program(
        "psychological-suggestion",
        [
            op_step(
                apply_op(
                    "psychological-mark",
                    effects["suggestion"],
                    selectors["random-two"],
                )
            ),
            op_step(queue_revelation),
        ],
    )

    patterns = {
        "midnight": "SingleTarget",
        "caressing": "Blast",
        "mockery": "Aoe",
        "spirit": "SingleTarget",
        "psychological": "Aoe",
        "seething": "Aoe",
        "tremble": "None",
        "cruelty": "None",
        "griever": "None",
        "revelation": "Aoe",
    }
    passive = {"tremble", "cruelty", "griever"}
    for key, id_ in abilities.items():
        add(
            "Ability",
            {
                "id": id_,
                "kind": "Passive" if key in passive else "Skill",
                "target_pattern": patterns[key],
                "retarget_policy": "CancelRemaining",
                "level_cap": 1,
                "cooldown_actions": 1 if key not in passive else 0,
                "semantic_tags_mask": (
                    2053 if key == "revelation" else 5 if patterns[key] != "None" else 4
                ),
            },
        )
        add(
            "AbilityPhase",
            {
                "ability_id": id_,
                "sequence": 1,
                "kind": "Resolved",
                "program_identity_id": ability_programs.get(id_),
            },
        )
        add(
            "EnemyAbility",
            {
                "id": id_,
                "telegraph": "Charge" if key == "psychological" else "None",
                "cooldown_actions": 1 if key not in passive else 0,
                "initial_cooldown_actions": 0,
                "charge_actions": 0,
                "ai_tag": key,
            },
        )

    effect_rows = {
        "shock": ("Dot", "DispellableDebuff", 1, integers["two"], "TargetTurnEnd", "TurnStart", "Refresh", damage["shock"], "Lightning"),
        "dominated": ("Control", "CleanseableControl", 1, integers["one"], "TargetTurnEnd", "None", "Refresh", None, None),
        "suggestion": ("Debuff", "NonDispellable", 1, None, "Permanent", "None", "Replace", None, None),
        "cruelty-listener": ("NeutralState", "NonDispellable", 1, None, "Permanent", "None", "Replace", None, None),
        "griever": ("Buff", "NonDispellable", 100, None, "Permanent", "None", "RefreshAndAddStacks", None, None),
    }
    for key, id_ in effects.items():
        category, dispel, limit, duration, clock, tick, policy, magnitude, dot = effect_rows[key]
        add(
            "Effect",
            {
                "id": id_,
                "category": category,
                "dispel_category": dispel,
                "stack_limit": limit,
                "duration_expression_id": duration,
                "duration_clock": clock,
                "tick_phase": tick,
                "stack_policy": policy,
                "magnitude_comparator_expression_id": magnitude,
                "dot_element": dot,
                "snapshot_policy": "OnApplication",
                "teardown_policy": "RemoveWithOwner",
                "application_priority": 0,
            },
        )
    for key, tags in {
        "shock": ["shock", "kafka-shock"],
        "dominated": ["dominated", "forced-basic-attack-random-ally"],
        "suggestion": ["psychological-suggestion"],
        "cruelty-listener": ["cruelty-listener"],
        "griever": ["griever-damage-growth"],
    }.items():
        for sequence, tag in enumerate(tags, start=1):
            add("EffectTag", {"effect_id": effects[key], "sequence": sequence, "tag": tag})

    add(
        "ModifierStackingGroup",
        {
            "id": modifier_group,
            "stable_key": "goal07.enemy.s10.modifier-group.griever",
            "aggregation": "Sum",
            "comparator_expression_id": None,
        },
    )
    add(
        "ModifierDefinition",
        {
            "id": modifier,
            "source_effect_id": effects["griever"],
            "owner_selector_id": selectors["owner"],
            "subject_selector_id": selectors["owner"],
            "stat": "Atk",
            "formula_stage": "DamageBoost",
            "formula_purpose": "OrdinaryDamage",
            "value_expression_id": ratios["zero-point-zero-two"],
            "value_domain": "Ratio",
            "stacking_group_id": modifier_group,
            "priority": 0,
            "cap_formula_stage": "DamageBoost",
            "snapshot_policy": "Dynamic",
            "duration_scope": "Turn",
        },
    )
    add(
        "EffectModifierBinding",
        {"effect_id": effects["griever"], "sequence": 1, "modifier_id": modifier},
    )

    cruelty_program = make_program(
        "cruelty-follow-up",
        [
            op_step(
                damage_op(
                    "cruelty-follow-up-damage",
                    damage["cruelty"],
                    selectors["current-shocked"],
                    "NoOp",
                )
            )
        ],
    )
    add(
        "RuleDefinition",
        {
            "id": cruelty_rule,
            "domain": "Battle",
            "source_definition_identity_id": effects["cruelty-listener"],
            "source_class": "Effect",
            "source_digest_sha256": sha256_text("goal07-s10-kafka-cruelty-v1"),
        },
    )
    add(
        "EventFilter",
        {
            "id": cruelty_filter,
            "stable_key": "goal07.enemy.s10.filter.cruelty",
            "actor_selector_id": selectors["actor-other"],
            "target_selector_id": selectors["current-shocked"],
            "cause_ancestry": "RootCommand",
        },
    )
    add(
        "RuleTrigger",
        {
            "id": cruelty_trigger,
            "stable_key": "goal07.enemy.s10.trigger.cruelty",
            "rule_id": cruelty_rule,
            "sequence": 1,
            "event": json_cell("Damage", point="Applied"),
            "phase": "AfterEvent",
            "filter_id": cruelty_filter,
            "condition_id": conditions["always"],
            "once_scope": "Turn",
            "priority": 0,
            "program_id": cruelty_program,
        },
    )
    add(
        "EffectRuleBinding",
        {
            "effect_id": effects["cruelty-listener"],
            "sequence": 1,
            "rule_id": cruelty_rule,
        },
    )

    phase_entry_programs = [
        make_program(
            f"phase-{phase}-entry",
            [
                op_step(
                    apply_op(
                        f"phase-{phase}-install-cruelty",
                        effects["cruelty-listener"],
                        selectors["actor"],
                    )
                )
            ],
        )
        for phase in range(1, 4)
    ]
    phase_sequences = [
        ["midnight", "caressing", "mockery", "spirit"],
        ["midnight", "caressing", "mockery", "psychological"],
        ["caressing", "mockery", "seething", "psychological"],
    ]
    targets = {
        "midnight": selectors["primary"],
        "caressing": selectors["primary"],
        "mockery": selectors["all-opposing"],
        "spirit": selectors["primary"],
        "psychological": selectors["random-two"],
        "seething": selectors["primary"],
    }
    next_state, next_candidate, next_transition = BASE + 701, BASE + 801, BASE + 901
    for phase_index, sequence_keys in enumerate(phase_sequences):
        state_ids = list(range(next_state, next_state + len(sequence_keys)))
        next_state += len(sequence_keys)
        add(
            "AiGraph",
            {
                "id": graphs[phase_index],
                "initial_state_id": state_ids[0],
                "automatic_transition_budget": 8,
            },
        )
        for offset, (state_id, ability_key) in enumerate(zip(state_ids, sequence_keys)):
            ability_id = abilities[ability_key]
            add(
                "AiState",
                {
                    "id": state_id,
                    "stable_key": (
                        f"goal07.enemy.s10.ai.phase-{phase_index + 1}."
                        f"step-{offset + 1}-{ability_key}"
                    ),
                    "graph_id": graphs[phase_index],
                    "mandatory_fallback_ability_id": abilities["mockery"],
                    "turn_counter_reset": offset == 0,
                },
            )
            add(
                "AiCandidate",
                {
                    "id": next_candidate,
                    "stable_key": (
                        f"goal07.enemy.s10.ai.phase-{phase_index + 1}."
                        f"candidate-{offset + 1}-{ability_key}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "ability_id": ability_id,
                    "condition_id": conditions["always"],
                    "target_selector_id": targets[ability_key],
                    "priority": 0,
                    "selection": "FirstLegal",
                    "no_target_fallback": "UseFallbackAbility",
                    "fallback_ability_id": abilities["mockery"],
                },
            )
            next_candidate += 1
            add(
                "AiTransition",
                {
                    "id": next_transition,
                    "stable_key": (
                        f"goal07.enemy.s10.ai.phase-{phase_index + 1}."
                        f"transition-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "target_state_id": state_ids[(offset + 1) % len(state_ids)],
                    "condition_id": conditions["always"],
                    "priority": 0,
                    "timing": "AfterAction",
                },
            )
            next_transition += 1

    add(
        "EnemyTemplate",
        {
            "id": template,
            "rank": "Boss",
            "base_aggro_decimal": "100",
            "default_ai_graph_id": graphs[0],
        },
    )
    add(
        "EnemyVariant",
        {
            "id": variant,
            "template_id": template,
            "ai_graph_id": graphs[0],
            "mechanically_distinct_key": VARIANT_KEY,
        },
    )
    for level in anchor["levels"]:
        add(
            "EnemyStat",
            {
                "variant_id": variant,
                "level": level["authored_level"],
                "difficulty_key": "standard-universe-v1",
                "hp_decimal": level["base_hp"],
                "atk_decimal": level["base_atk"],
                "def_decimal": level["base_def"],
                "spd_decimal": level["base_spd"],
                "effect_hit_rate_decimal": level["effect_hit_rate"],
                "effect_resistance_decimal": level["effect_resistance"],
                "crit_damage_decimal": "0.2",
            },
        )
    for sequence, weakness in enumerate(["Physical", "Wind", "Imaginary"], start=1):
        add(
            "EnemyWeakness",
            {"variant_id": variant, "sequence": sequence, "element": weakness},
        )
    for element, value in [
        ("Fire", "0.2"),
        ("Ice", "0.2"),
        ("Lightning", "0.4"),
        ("Quantum", "0.2"),
    ]:
        add(
            "EnemyResistance",
            {"variant_id": variant, "element": element, "value_decimal": value},
        )
    add(
        "EnemyDebuffResistance",
        {
            "variant_id": variant,
            "category_key": "STAT_CTRL_Frozen",
            "value_decimal": "0.75",
        },
    )
    add(
        "EnemyToughnessLayer",
        {
            "variant_id": variant,
            "sequence": 1,
            "layer_key": "ordinary",
            "kind": "Ordinary",
            "maximum_decimal": "450",
            "recovery_ratio_decimal": "1",
            "active_at_start": True,
        },
    )
    for sequence, ability_id in enumerate(abilities.values(), start=1):
        add(
            "EnemyVariantAbility",
            {"variant_id": variant, "sequence": sequence, "ability_id": ability_id},
        )
    for phase, graph in enumerate(graphs, start=1):
        add(
            "EnemyPhase",
            {
                "id": BASE + 600 + phase,
                "stable_key": f"goal07.enemy.s10.phase-{phase}",
                "variant_id": variant,
                "sequence": phase,
                "entry_condition_id": conditions["always"],
                "exit_condition_id": conditions["always"],
                "replacement_priority": phase,
                "ai_graph_id": graph,
                "entry_program_id": phase_entry_programs[phase - 1],
                "targetable": True,
                "transition_model": "TransformSameUnit",
                "hp_carry": "Reset",
                "action_gauge_carry": "Reset",
                "effect_carry": "Clear",
                "toughness_carry": "Reset",
                "summon_carry": "Clear",
            },
        )

    anchor_digest = sha256_bytes(anchor_path(PARTITION).read_bytes())
    add(
        "SourceRecord",
        {
            "id": SOURCE_RECORD_ID,
            "stable_key": "source.hsr-wiki.stellaron-hunter-kafka-complete.2026-07-29",
            "category": "CommunityMaintained",
            "publisher": anchor["source"]["publisher"],
            "url": anchor["source"]["url"],
            "accessed_on": anchor["source"]["accessed_on"],
            "applicable_game_version": anchor["source"]["game_version"],
            "confidence": "SecondaryVersionSensitiveCrossCheck",
            "evidence_sha256": anchor_digest,
            "usage_note": "Exact World 5 levels and public Kafka mechanics are committed as Goal 07 evidence.",
        },
    )
    add(
        "EvidenceRecord",
        {
            "id": EVIDENCE_RECORD_ID,
            "stable_key": "evidence.goal07.enemy.s10.numeric-anchors",
            "kind": "SourcePayload",
            "source_record_id": SOURCE_RECORD_ID,
            "sha256": anchor_digest,
            "note": "Committed exact public per-level numeric anchors for Goal 07 S10.",
        },
    )
    for item in identities:
        add("ContentIdentity", item)
        add(
            "ContentEvidenceBinding",
            {
                "content_id": item["id"],
                "sequence": 1,
                "fact_key": f"goal07.s10.executable:{item['stable_key']}",
                "source_record_id": 1,
                "evidence_record_id": 3,
                "quality": "ExactStructured",
                "mechanism_quality": "ExactStructured",
            },
        )
    add(
        "ContentEvidenceBinding",
        {
            "content_id": variant,
            "sequence": 2,
            "fact_key": "goal07.s10.public-level-stats",
            "source_record_id": SOURCE_RECORD_ID,
            "evidence_record_id": EVIDENCE_RECORD_ID,
            "quality": "ExactStructured",
            "mechanism_quality": "ExactStructured",
        },
    )
    for table_rows in rows.values():
        table_rows.sort(
            key=lambda row: json.dumps(
                row, ensure_ascii=False, sort_keys=True, default=str
            )
        )
    return rows


def owned_rows_s11() -> dict[str, list[dict[str, Any]]]:
    anchor = json.loads(anchor_path(PARTITION).read_text(encoding="utf-8"))
    manifest = json.loads(PARTITIONS.read_text(encoding="utf-8"))
    assigned = next(item for item in manifest["partitions"] if item["id"] == PARTITION)
    if assigned["enemy_variant_ids"] != [VARIANT_KEY]:
        raise ValueError("S11 frozen enemy assignment changed")

    variant, template = BASE + 1, BASE + 2
    graphs = [BASE + 10, BASE + 11, BASE + 12]
    abilities = {
        "banishing": BASE + 101,
        "burning": BASE + 102,
        "bombardment": BASE + 103,
        "power": BASE + 104,
        "emergency": BASE + 105,
        "tactical": BASE + 106,
        "boost": BASE + 107,
        "support-cycle": BASE + 120,
        "beetle-strike": BASE + 121,
        "hound-strike": BASE + 122,
        "direwolf-strike": BASE + 123,
        "arm-cycle": BASE + 124,
        "oppressive": BASE + 125,
        "disabling": BASE + 126,
        "overload": BASE + 127,
        "controlled": BASE + 128,
    }
    linked = {
        "phase-1-support-1": BASE + 201,
        "phase-1-support-2": BASE + 202,
        "phase-1-support-3": BASE + 203,
        "phase-1-support-4": BASE + 204,
        "phase-2-direwolf-left": BASE + 205,
        "phase-2-direwolf-right": BASE + 206,
        "phase-3-arm": BASE + 207,
    }
    selectors = {
        "actor": BASE + 401,
        "owner": BASE + 402,
        "primary": BASE + 403,
        "random-opponent": BASE + 404,
        "all-opposing": BASE + 405,
        "same-all": BASE + 406,
        "restrained": BASE + 407,
        "actor-summons": BASE + 408,
    }
    effects = {
        "def-down": BASE + 501,
        "power-amplification": BASE + 502,
        "restrain": BASE + 503,
        "arm-turns": BASE + 504,
        "overload": BASE + 505,
        "beetle-role": BASE + 506,
        "hound-role": BASE + 507,
    }
    modifiers = {
        "def-down": BASE + 521,
        "power-amplification": BASE + 522,
    }
    modifier_groups = {
        "def-down": BASE + 531,
        "power-amplification": BASE + 532,
    }
    conditions = {
        "always": BASE + 551,
        "restrained": BASE + 552,
        "arm-ready": BASE + 553,
        "overloaded": BASE + 554,
        "beetle-role": BASE + 555,
        "hound-role": BASE + 556,
        "support-role-assigned": BASE + 557,
        "arm-turn-ready": BASE + 558,
        "arm-hp-low": BASE + 559,
    }
    rows: dict[str, list[dict[str, Any]]] = {}
    identities: list[dict[str, Any]] = []
    next_program = BASE + 301
    next_operation = BASE + 1_001
    next_expression = BASE + 1_101

    def add(table: str, row: dict[str, Any]) -> None:
        rows.setdefault(table, []).append(row)

    def ident(
        id_: int,
        key: str,
        kind: str,
        name_en: str,
        name_zh_cn: str,
        summary: str,
        sources: str = "1",
    ) -> dict[str, Any]:
        row = identity(id_, key, kind, name_en, name_zh_cn, summary, sources)
        row["summary_zh_cn"] = "Goal 07 S11 来源绑定的史瓦罗（完整）可执行定义。"
        row["game_version_introduced"] = "1.0"
        return row

    identities.extend(
        [
            ident(
                variant,
                VARIANT_KEY,
                "EnemyVariant",
                "Svarog (Complete)",
                "史瓦罗（完整）",
                "Exact three-phase World 4 boss materialization.",
                "1|13",
            ),
            ident(
                template,
                "enemy.svarog-complete.littleboss",
                "Enemy",
                "Svarog (Complete) Template",
                "史瓦罗（完整）模板",
                "Version 4.4 boss template retained from source monster 1014011.",
            ),
        ]
    )
    for phase, graph in enumerate(graphs, start=1):
        identities.append(
            ident(
                graph,
                f"ai.goal07.svarog-complete.phase-{phase}",
                "AiGraph",
                f"Svarog Complete Phase {phase} AI",
                f"史瓦罗完整形态第{phase}阶段AI",
                "Finite phase rotation with explicit support deployment.",
            )
        )
    ability_metadata = {
        "banishing": ("Banishing Punch", "驱逐拳击", "101401101"),
        "burning": ("Burning Beam", "延烧光束", "101401102"),
        "bombardment": ("Oversaturated Bombardment", "过饱和轰炸", "101401104"),
        "power": ("Power Amplification", "输出增幅", "101401112"),
        "emergency": ("Emergency Support", "应急支援", "101401113"),
        "tactical": ("Tactical Support", "战术支援", "101401114"),
        "boost": ("Boost Deployment", "增效部署", "101401107"),
        "support-cycle": (
            "Emergency Support Random Action",
            "应急支援随机行动",
            "derived-101203001-or-101201101",
        ),
        "beetle-strike": ("Unstable Forcefield", "不稳定力场", "101203001"),
        "hound-strike": ("Vertical Strike", "垂直打击", "101201101"),
        "direwolf-strike": ("Disintegration Order", "解体指令", "101302003"),
        "arm-cycle": (
            "Auxiliary Robot Arm Action Cycle",
            "辅助机械臂行动循环",
            "derived-101204104-through-101204107",
        ),
        "oppressive": ("Oppressive Embrace", "巨掌之间", "101204104"),
        "disabling": ("Disabling Field", "压制力场", "101204105"),
        "overload": ("Overload Warning", "过载警告", "101204107"),
        "controlled": ("Controlled Blasting", "控场起爆", "101204106"),
    }
    for key, id_ in abilities.items():
        name_en, name_zh_cn, source = ability_metadata[key]
        identities.append(
            ident(
                id_,
                f"enemy.svarog-complete.littleboss.ability.{key}",
                "Ability",
                name_en,
                name_zh_cn,
                f"Executable transcription of source skill {source}.",
            )
        )
    linked_metadata = {
        "phase-1-support-1": ("Random Automaton Support", "随机自动机兵支援"),
        "phase-1-support-2": ("Random Automaton Support", "随机自动机兵支援"),
        "phase-1-support-3": ("Random Automaton Support", "随机自动机兵支援"),
        "phase-1-support-4": ("Random Automaton Support", "随机自动机兵支援"),
        "phase-2-direwolf-left": ("Automaton Direwolf", "自动机兵「齿狼」"),
        "phase-2-direwolf-right": ("Automaton Direwolf", "自动机兵「齿狼」"),
        "phase-3-arm": ("Auxiliary Robot Arm Unit", "辅助机械臂单元"),
    }
    for key, id_ in linked.items():
        name_en, name_zh_cn = linked_metadata[key]
        identities.append(
            ident(
                id_,
                f"unit.goal07.svarog-complete.{key}",
                "CharacterForm",
                name_en,
                name_zh_cn,
                "Owner-scaled targetable support deployed by Svarog.",
            )
        )
    for key, id_ in selectors.items():
        identities.append(
            ident(
                id_,
                f"selector.goal07.svarog-complete.{key}",
                "Selector",
                f"Svarog Complete {key} Selector",
                f"史瓦罗完整形态{key}选择器",
                "S11 battle selector.",
            )
        )
    effect_metadata = {
        "def-down": (
            "Bombardment DEF Reduction",
            "轰炸防御降低",
            "Stackable three-turn 20% DEF reduction.",
        ),
        "power-amplification": (
            "Power Amplification",
            "输出增幅",
            "Permanent stackable 15% allied damage increase.",
        ),
        "restrain": (
            "Restrained",
            "拘束",
            "Blocks ordinary actions while the controlling arm remains.",
        ),
        "arm-turns": (
            "Auxiliary Arm Turn Clock",
            "辅助机械臂行动计数",
            "Counts three linked-unit turns before overload.",
        ),
        "overload": (
            "Auxiliary Arm Overload",
            "辅助机械臂过载",
            "Queues Controlled Blasting for the arm's next action.",
        ),
        "beetle-role": (
            "Emergency Support Beetle Role",
            "应急支援甲虫职责",
            "Randomly fixes one support slot as an Automaton Beetle.",
        ),
        "hound-role": (
            "Emergency Support Hound Role",
            "应急支援战犬职责",
            "Randomly fixes one support slot as an Automaton Hound.",
        ),
    }
    for key, id_ in effects.items():
        name_en, name_zh_cn, summary = effect_metadata[key]
        identities.append(
            ident(
                id_,
                f"effect.goal07.svarog-complete.{key}",
                "Effect",
                name_en,
                name_zh_cn,
                summary,
            )
        )
    for key, id_ in modifiers.items():
        identities.append(
            ident(
                id_,
                f"modifier.goal07.svarog-complete.{key}",
                "Modifier",
                f"Svarog Complete {key} Modifier",
                f"史瓦罗完整形态{key}调整器",
                "Effect-owned S11 stat modifier.",
            )
        )
    for key, id_ in modifier_groups.items():
        identities.append(
            ident(
                id_,
                f"modifier-group.goal07.svarog-complete.{key}",
                "Modifier",
                f"Svarog Complete {key} Stacking",
                f"史瓦罗完整形态{key}叠加组",
                "Additive S11 modifier stacking group.",
            )
        )

    add("Selector", selector(selectors["actor"], "Actor", "SameSide"))
    add("Selector", selector(selectors["owner"], "Owner", "SameSide"))
    add("Selector", selector(selectors["primary"], "PrimaryTarget", "OpposingSide"))
    random_opponent = selector(
        selectors["random-opponent"],
        "Actor",
        "OpposingSide",
        choice="RngUniform",
    )
    random_opponent["rng_purpose_key"] = "damage-target"
    add("Selector", random_opponent)
    add(
        "Selector",
        selector(
            selectors["all-opposing"],
            "Actor",
            "OpposingSide",
            minimum=1,
            maximum=8,
            choice="All",
        ),
    )
    add(
        "Selector",
        selector(
            selectors["same-all"],
            "Actor",
            "SameSide",
            minimum=1,
            maximum=8,
            choice="All",
        ),
    )
    add(
        "Selector",
        selector(
            selectors["restrained"],
            "Actor",
            "OpposingSide",
            minimum=0,
            maximum=1,
            empty="NoOp",
            choice="First",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["restrained"],
            "sequence": 1,
            "predicate": json_cell("HasEffect", effect_id=effects["restrain"]),
        },
    )
    add(
        "Selector",
        selector(
            selectors["actor-summons"],
            "Actor",
            "SameSide",
            presence="Linked",
            minimum=0,
            maximum=8,
            empty="NoOp",
            choice="All",
        ),
    )
    add(
        "SelectorPredicate",
        {
            "selector_id": selectors["actor-summons"],
            "sequence": 1,
            "predicate": json_cell(
                "OwnedBy",
                owner_selector_id=selectors["actor"],
            ),
        },
    )

    def expression(name: str, kind: str, node: str) -> int:
        nonlocal next_expression
        id_ = next_expression
        next_expression += 1
        add(
            "ValueExpression",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s11.expression.{name}",
                "result_kind": kind,
                "node": node,
            },
        )
        return id_

    def product(name: str, left: int, right: int) -> int:
        return expression(
            name,
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedMultiply",
                left_expression_id=left,
                right_expression_id=right,
                rounding="NearestTiesAway",
            ),
        )

    actor_atk = expression(
        "actor-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    actor_hp = expression(
        "actor-current-hp",
        "Scalar",
        json_cell(
            "QueryHp",
            subject_selector_id=selectors["actor"],
        ),
    )
    actor_maximum_hp = (
        expression(
            "actor-maximum-hp",
            "Scalar",
            json_cell(
                "QueryStat",
                subject_selector_id=selectors["actor"],
                stat="Hp",
                formula_purpose="Stat",
            ),
        )
        if PARTITION in {"G07-P5-M15-S13", "G07-P5-M15-S14"}
        else None
    )
    primary_hp = expression(
        "primary-maximum-hp",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["primary"],
            stat="Hp",
            formula_purpose="OrdinaryDamage",
        ),
    )
    restrained_hp = expression(
        "restrained-maximum-hp",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["restrained"],
            stat="Hp",
            formula_purpose="OrdinaryDamage",
        ),
    )
    ratios = {}
    for key, value in [
        ("negative-point-two", "-0.2"),
        ("point-one-five", "0.15"),
        ("point-two", "0.2"),
        ("point-five", "0.5"),
        ("one", "1"),
        ("two-point-five", "2.5"),
        ("three", "3"),
        ("fifteen", "15"),
    ]:
        ratios[key] = expression(
            f"ratio-{key}",
            "Scalar",
            json_cell("ScalarLiteral", value_decimal=value),
        )
    integers = {}
    for key, value in [("one", 1), ("two", 2), ("three", 3)]:
        integers[key] = expression(
            f"integer-{key}",
            "Integer",
            json_cell("IntegerLiteral", value=value),
        )
    arm_turn_stacks = expression(
        "arm-turn-stacks",
        "Integer",
        json_cell(
            "QueryEffectStacks",
            subject_selector_id=selectors["actor"],
            effect_id=effects["arm-turns"],
        ),
    )
    arm_half_hp = product(
        "arm-half-maximum-hp",
        actor_maximum_hp,
        ratios["point-five"],
    )
    damage = {
        "banishing": product("banishing-damage", actor_atk, ratios["three"]),
        "burning": product("burning-damage", actor_atk, ratios["three"]),
        "bombardment": product(
            "bombardment-hit-damage",
            actor_atk,
            ratios["point-one-five"],
        ),
        "beetle": product("beetle-damage", actor_atk, ratios["three"]),
        "hound": product("hound-damage", actor_atk, ratios["two-point-five"]),
        "direwolf": product(
            "direwolf-damage",
            actor_atk,
            ratios["two-point-five"],
        ),
        "controlled": product(
            "controlled-blasting-damage",
            actor_atk,
            ratios["fifteen"],
        ),
        "oppressive": product(
            "oppressive-embrace-damage",
            primary_hp,
            ratios["point-five"],
        ),
        "disabling": product(
            "disabling-field-damage",
            restrained_hp,
            ratios["point-five"],
        ),
    }
    add(
        "ConditionExpression",
        {
            "id": conditions["always"],
            "stable_key": "goal07.enemy.s11.condition.always",
            "node": json_cell("Constant", value=True),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["restrained"],
            "stable_key": "goal07.enemy.s11.condition.restrained-target-exists",
            "node": json_cell(
                "SelectorCardinality",
                selector_id=selectors["restrained"],
                minimum_count=1,
                maximum_count=1,
            ),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["arm-turn-ready"],
            "stable_key": "goal07.enemy.s11.condition.arm-turn-overload-ready",
            "node": json_cell(
                "Compare",
                left_expression_id=arm_turn_stacks,
                comparison="GreaterOrEqual",
                right_expression_id=integers["two"],
            ),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["arm-hp-low"],
            "stable_key": "goal07.enemy.s11.condition.arm-hp-overload-ready",
            "node": json_cell(
                "Compare",
                left_expression_id=actor_hp,
                comparison="Less",
                right_expression_id=arm_half_hp,
            ),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["arm-ready"],
            "stable_key": "goal07.enemy.s11.condition.arm-overload-ready",
            "node": json_cell(
                "Any",
                condition_ids=[
                    conditions["arm-turn-ready"],
                    conditions["arm-hp-low"],
                ],
            ),
        },
    )
    add(
        "ConditionExpression",
        {
            "id": conditions["overloaded"],
            "stable_key": "goal07.enemy.s11.condition.arm-overloaded",
            "node": json_cell(
                "EffectExists",
                selector_id=selectors["actor"],
                effect_id=effects["overload"],
            ),
        },
    )
    for key in ["beetle-role", "hound-role"]:
        add(
            "ConditionExpression",
            {
                "id": conditions[key],
                "stable_key": f"goal07.enemy.s11.condition.{key}",
                "node": json_cell(
                    "EffectExists",
                    selector_id=selectors["actor"],
                    effect_id=effects[key],
                ),
            },
        )
    add(
        "ConditionExpression",
        {
            "id": conditions["support-role-assigned"],
            "stable_key": "goal07.enemy.s11.condition.support-role-assigned",
            "node": json_cell(
                "Any",
                condition_ids=[
                    conditions["beetle-role"],
                    conditions["hound-role"],
                ],
            ),
        },
    )

    def new_operation(
        name: str,
        payload: str,
        target: int | None = None,
        empty: str = "Fault",
    ) -> int:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        row = operation(id_, name, payload, target, empty)
        row["stable_key"] = f"goal07.enemy.s11.operation.{name}"
        add("Operation", row)
        return id_

    def op_step(id_: int) -> str:
        return json_cell("Operation", operation_id=id_)

    def damage_op(
        name: str,
        amount: int,
        target: int,
        empty: str = "Fault",
    ) -> int:
        return new_operation(
            name,
            json_cell(
                "Damage",
                amount_expression_id=amount,
                damage_class="Ordinary",
                element="Physical",
                can_crit=True,
            ),
            target,
            empty,
        )

    def apply_op(
        name: str,
        effect: int,
        target: int,
        chance: int | None = None,
        stacks: int | None = None,
        empty: str = "Fault",
    ) -> int:
        return new_operation(
            name,
            json_cell(
                "ApplyEffect",
                effect_id=effect,
                stacks_expression_id=stacks,
                chance_policy="Resistible" if chance is not None else "Guaranteed",
                base_chance_expression_id=chance,
                rng_purpose_key="effect-application" if chance is not None else None,
            ),
            target,
            empty,
        )

    def remove_op(name: str, effect: int, target: int) -> int:
        return new_operation(
            name,
            json_cell("RemoveEffect", effect_id=effect),
            target,
            "NoOp",
        )

    def summon_op(name: str, linked_id: int) -> int:
        return new_operation(
            name,
            json_cell(
                "Summon",
                unit_definition_identity_id=linked_id,
                owner_selector_id=selectors["actor"],
            ),
        )

    def make_program(name: str, steps: list[str]) -> int:
        nonlocal next_program
        id_ = next_program
        next_program += 1
        identities.append(
            ident(
                id_,
                f"program.goal07.svarog-complete.{name}",
                "Program",
                f"Svarog Complete {name} Program",
                f"史瓦罗完整形态{name}程序",
                "Ordered shared Rule IR program for the S11 enemy.",
            )
        )
        add("Program", {"id": id_, "domain": "Battle"})
        for sequence, step in enumerate(steps, start=1):
            add(
                "ProgramStep",
                {"program_id": id_, "sequence": sequence, "step": step},
            )
        return id_

    oppressive_program = make_program(
        "oppressive-embrace",
        [
            op_step(
                damage_op(
                    "oppressive-embrace-damage",
                    damage["oppressive"],
                    selectors["primary"],
                )
            ),
            op_step(
                apply_op(
                    "oppressive-embrace-restrain",
                    effects["restrain"],
                    selectors["primary"],
                )
            ),
            op_step(
                apply_op(
                    "oppressive-embrace-turn-count",
                    effects["arm-turns"],
                    selectors["actor"],
                    stacks=integers["one"],
                )
            ),
        ],
    )
    disabling_program = make_program(
        "disabling-field",
        [
            op_step(
                damage_op(
                    "disabling-field-damage",
                    damage["disabling"],
                    selectors["restrained"],
                    "NoOp",
                )
            ),
            op_step(
                apply_op(
                    "disabling-field-turn-count",
                    effects["arm-turns"],
                    selectors["actor"],
                    stacks=integers["one"],
                )
            ),
        ],
    )
    overload_program = make_program(
        "overload-warning",
        [
            op_step(
                remove_op(
                    "overload-release-restrained",
                    effects["restrain"],
                    selectors["all-opposing"],
                )
            ),
            op_step(
                apply_op(
                    "overload-marker",
                    effects["overload"],
                    selectors["actor"],
                )
            ),
            op_step(
                remove_op(
                    "overload-clear-turn-count",
                    effects["arm-turns"],
                    selectors["actor"],
                )
            ),
            op_step(
                new_operation(
                    "overload-advance-svarog",
                    json_cell(
                        "AdvanceAction",
                        amount_expression_id=ratios["one"],
                    ),
                    selectors["owner"],
                )
            ),
        ],
    )
    controlled_program = make_program(
        "controlled-blasting",
        [
            op_step(
                damage_op(
                    "controlled-blasting-damage",
                    damage["controlled"],
                    selectors["primary"],
                )
            ),
            op_step(
                new_operation(
                    "controlled-blasting-despawn-arm",
                    json_cell("Despawn"),
                    selectors["actor"],
                )
            ),
        ],
    )
    ordinary_arm_program = make_program(
        "auxiliary-arm-ordinary-choice",
        [
            json_cell(
                "If",
                condition_id=conditions["restrained"],
                then_program_id=disabling_program,
                else_program_id=oppressive_program,
            )
        ],
    )
    ready_arm_program = make_program(
        "auxiliary-arm-overload-choice",
        [
            json_cell(
                "If",
                condition_id=conditions["arm-ready"],
                then_program_id=overload_program,
                else_program_id=ordinary_arm_program,
            )
        ],
    )
    arm_cycle_program = make_program(
        "auxiliary-arm-cycle",
        [
            json_cell(
                "If",
                condition_id=conditions["overloaded"],
                then_program_id=controlled_program,
                else_program_id=ready_arm_program,
            )
        ],
    )
    beetle_program = make_program(
        "unstable-forcefield",
        [
            op_step(
                damage_op(
                    "unstable-forcefield-damage",
                    damage["beetle"],
                    selectors["primary"],
                )
            )
        ],
    )
    hound_program = make_program(
        "vertical-strike",
        [
            op_step(
                damage_op(
                    "vertical-strike-damage",
                    damage["hound"],
                    selectors["primary"],
                )
            )
        ],
    )
    assigned_support_program = make_program(
        "emergency-support-assigned-action",
        [
            json_cell(
                "If",
                condition_id=conditions["beetle-role"],
                then_program_id=beetle_program,
                else_program_id=hound_program,
            )
        ],
    )
    random_support_role = new_operation(
        "emergency-support-random-role",
        json_cell(
            "ApplyRandomEffect",
            effect_ids=[
                effects["beetle-role"],
                effects["hound-role"],
            ],
            stacks_expression_id=integers["one"],
            choice_rng_purpose_key="behavior-choice",
            chance_policy="Guaranteed",
            base_chance_expression_id=None,
            chance_rng_purpose_key=None,
        ),
        selectors["actor"],
    )
    choose_support_program = make_program(
        "emergency-support-choose-role",
        [
            op_step(random_support_role),
            json_cell(
                "If",
                condition_id=conditions["beetle-role"],
                then_program_id=beetle_program,
                else_program_id=hound_program,
            ),
        ],
    )
    support_cycle_program = make_program(
        "emergency-support-action-cycle",
        [
            json_cell(
                "If",
                condition_id=conditions["support-role-assigned"],
                then_program_id=assigned_support_program,
                else_program_id=choose_support_program,
            )
        ],
    )

    ability_programs = {
        abilities["banishing"]: make_program(
            "banishing-punch",
            [
                op_step(
                    damage_op(
                        "banishing-punch-damage",
                        damage["banishing"],
                        selectors["primary"],
                    )
                )
            ],
        ),
        abilities["burning"]: make_program(
            "burning-beam",
            [
                op_step(
                    damage_op(
                        "burning-beam-damage",
                        damage["burning"],
                        selectors["primary"],
                    )
                ),
                op_step(
                    new_operation(
                        "burning-beam-delay",
                        json_cell(
                            "DelayAction",
                            amount_expression_id=ratios["point-five"],
                        ),
                        selectors["primary"],
                    )
                ),
            ],
        ),
        abilities["power"]: make_program(
            "power-amplification",
            [
                op_step(
                    apply_op(
                        "power-amplification-allies",
                        effects["power-amplification"],
                        selectors["same-all"],
                        stacks=integers["one"],
                    )
                )
            ],
        ),
        abilities["emergency"]: make_program(
            "emergency-support",
            [
                op_step(
                    summon_op(
                        f"emergency-support-{key}",
                        linked[key],
                    )
                )
                for key in [
                    "phase-1-support-1",
                    "phase-1-support-2",
                    "phase-1-support-3",
                    "phase-1-support-4",
                ]
            ],
        ),
        abilities["tactical"]: make_program(
            "tactical-support",
            [
                op_step(
                    summon_op(
                        f"tactical-support-{key}",
                        linked[key],
                    )
                )
                for key in [
                    "phase-2-direwolf-left",
                    "phase-2-direwolf-right",
                ]
            ],
        ),
        abilities["boost"]: make_program(
            "boost-deployment",
            [
                op_step(
                    summon_op(
                        "boost-deployment-auxiliary-arm",
                        linked["phase-3-arm"],
                    )
                )
            ],
        ),
        abilities["support-cycle"]: support_cycle_program,
        abilities["beetle-strike"]: beetle_program,
        abilities["hound-strike"]: hound_program,
        abilities["direwolf-strike"]: make_program(
            "disintegration-order",
            [
                op_step(
                    damage_op(
                        "disintegration-order-damage",
                        damage["direwolf"],
                        selectors["primary"],
                    )
                )
            ],
        ),
        abilities["arm-cycle"]: arm_cycle_program,
        abilities["oppressive"]: oppressive_program,
        abilities["disabling"]: disabling_program,
        abilities["overload"]: overload_program,
        abilities["controlled"]: controlled_program,
    }
    bombardment_steps = [
        op_step(
            damage_op(
                f"oversaturated-bombardment-hit-{hit}",
                damage["bombardment"],
                selectors["all-opposing"],
            )
        )
        for hit in range(1, 13)
    ]
    bombardment_steps.append(
        op_step(
            apply_op(
                "oversaturated-bombardment-def-down",
                effects["def-down"],
                selectors["all-opposing"],
                chance=ratios["one"],
                stacks=integers["one"],
            )
        )
    )
    ability_programs[abilities["bombardment"]] = make_program(
        "oversaturated-bombardment",
        bombardment_steps,
    )

    linked_ability_keys = {
        "support-cycle",
        "beetle-strike",
        "hound-strike",
        "direwolf-strike",
        "arm-cycle",
        "oppressive",
        "disabling",
        "overload",
        "controlled",
    }
    target_patterns = {
        "banishing": "SingleTarget",
        "burning": "SingleTarget",
        "bombardment": "Aoe",
        "power": "None",
        "emergency": "None",
        "tactical": "None",
        "boost": "None",
        "support-cycle": "SingleTarget",
        "beetle-strike": "SingleTarget",
        "hound-strike": "SingleTarget",
        "direwolf-strike": "SingleTarget",
        "arm-cycle": "SingleTarget",
        "oppressive": "SingleTarget",
        "disabling": "SingleTarget",
        "overload": "None",
        "controlled": "SingleTarget",
    }
    for key, id_ in abilities.items():
        linked_ability = key in linked_ability_keys
        add(
            "Ability",
            {
                "id": id_,
                "kind": "Summon" if linked_ability else "Skill",
                "target_pattern": target_patterns[key],
                "retarget_policy": "CancelRemaining",
                "level_cap": 1,
                "cooldown_actions": 1,
                "semantic_tags_mask": (
                    5 if target_patterns[key] != "None" else 4
                ),
            },
        )
        add(
            "AbilityPhase",
            {
                "ability_id": id_,
                "sequence": 1,
                "kind": "Resolved",
                "program_identity_id": ability_programs[id_],
            },
        )
        if not linked_ability:
            add(
                "EnemyAbility",
                {
                    "id": id_,
                    "telegraph": (
                        "Charge" if key == "bombardment" else "None"
                    ),
                    "cooldown_actions": 1,
                    "initial_cooldown_actions": 0,
                    "charge_actions": 0,
                    "ai_tag": key,
                },
            )

    effect_rows = {
        "def-down": (
            "Debuff",
            "DispellableDebuff",
            100,
            integers["three"],
            "TargetTurnEnd",
            "RefreshAndAddStacks",
        ),
        "power-amplification": (
            "Buff",
            "NonDispellable",
            100,
            None,
            "Permanent",
            "RefreshAndAddStacks",
        ),
        "restrain": (
            "Control",
            "NonDispellable",
            1,
            None,
            "Permanent",
            "Replace",
        ),
        "arm-turns": (
            "NeutralState",
            "NonDispellable",
            3,
            None,
            "Permanent",
            "RefreshAndAddStacks",
        ),
        "overload": (
            "NeutralState",
            "NonDispellable",
            1,
            None,
            "Permanent",
            "Replace",
        ),
        "beetle-role": (
            "NeutralState",
            "NonDispellable",
            1,
            None,
            "Permanent",
            "Replace",
        ),
        "hound-role": (
            "NeutralState",
            "NonDispellable",
            1,
            None,
            "Permanent",
            "Replace",
        ),
    }
    for key, id_ in effects.items():
        category, dispel, limit, duration, clock, policy = effect_rows[key]
        add(
            "Effect",
            {
                "id": id_,
                "category": category,
                "dispel_category": dispel,
                "stack_limit": limit,
                "duration_expression_id": duration,
                "duration_clock": clock,
                "tick_phase": "None",
                "stack_policy": policy,
                "snapshot_policy": "OnApplication",
                "teardown_policy": "RemoveWithOwner",
                "application_priority": 0,
            },
        )
    effect_tags = {
        "def-down": ["def-down", "oversaturated-bombardment"],
        "power-amplification": ["power-amplification", "stackable-damage-boost"],
        "restrain": ["restrained", "blocks-normal-action", "remove-with-applier"],
        "arm-turns": ["auxiliary-arm-turn-clock"],
        "overload": ["overload-warning", "controlled-blasting-next-action"],
        "beetle-role": ["emergency-support-role", "automaton-beetle"],
        "hound-role": ["emergency-support-role", "automaton-hound"],
    }
    for key, tags in effect_tags.items():
        for sequence, tag in enumerate(tags, start=1):
            add(
                "EffectTag",
                {"effect_id": effects[key], "sequence": sequence, "tag": tag},
            )
    for key, id_ in modifier_groups.items():
        add(
            "ModifierStackingGroup",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s11.modifier-group.{key}",
                "aggregation": "Sum",
                "comparator_expression_id": None,
            },
        )
    modifier_specs = {
        "def-down": (
            "Def",
            "PercentOfBase",
            "Stat",
            ratios["negative-point-two"],
            "OnApplication",
        ),
        "power-amplification": (
            "Atk",
            "DamageBoost",
            "OrdinaryDamage",
            ratios["point-one-five"],
            "Dynamic",
        ),
    }
    for key, id_ in modifiers.items():
        stat, stage, purpose, value, snapshot = modifier_specs[key]
        add(
            "ModifierDefinition",
            {
                "id": id_,
                "source_effect_id": effects[key],
                "owner_selector_id": selectors["owner"],
                "subject_selector_id": selectors["owner"],
                "stat": stat,
                "formula_stage": stage,
                "formula_purpose": purpose,
                "value_expression_id": value,
                "value_domain": "Ratio",
                "stacking_group_id": modifier_groups[key],
                "priority": 0,
                "cap_formula_stage": stage,
                "snapshot_policy": snapshot,
                "duration_scope": "Turn",
            },
        )
        add(
            "EffectModifierBinding",
            {
                "effect_id": effects[key],
                "sequence": 1,
                "modifier_id": id_,
            },
        )

    phase_sequences = [
        ["emergency", "banishing", "burning", "power"],
        ["tactical", "bombardment", "banishing", "burning", "power"],
        ["boost", "bombardment", "banishing", "burning"],
    ]
    target_selectors = {
        "emergency": selectors["actor"],
        "banishing": selectors["random-opponent"],
        "burning": selectors["random-opponent"],
        "power": selectors["actor"],
        "tactical": selectors["actor"],
        "bombardment": selectors["all-opposing"],
        "boost": selectors["actor"],
    }
    next_state = BASE + 701
    next_candidate = BASE + 801
    next_transition = BASE + 901
    for phase_index, sequence_keys in enumerate(phase_sequences):
        state_ids = list(range(next_state, next_state + len(sequence_keys)))
        next_state += len(sequence_keys)
        add(
            "AiGraph",
            {
                "id": graphs[phase_index],
                "initial_state_id": state_ids[0],
                "automatic_transition_budget": 8,
            },
        )
        for offset, (state_id, ability_key) in enumerate(
            zip(state_ids, sequence_keys)
        ):
            ability_id = abilities[ability_key]
            add(
                "AiState",
                {
                    "id": state_id,
                    "stable_key": (
                        f"goal07.enemy.s11.ai.phase-{phase_index + 1}."
                        f"step-{offset + 1}-{ability_key}"
                    ),
                    "graph_id": graphs[phase_index],
                    "mandatory_fallback_ability_id": abilities["banishing"],
                    "turn_counter_reset": offset == 0,
                },
            )
            add(
                "AiCandidate",
                {
                    "id": next_candidate,
                    "stable_key": (
                        f"goal07.enemy.s11.ai.phase-{phase_index + 1}."
                        f"candidate-{offset + 1}-{ability_key}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "ability_id": ability_id,
                    "condition_id": conditions["always"],
                    "target_selector_id": target_selectors[ability_key],
                    "priority": 0,
                    "selection": "FirstLegal",
                    "no_target_fallback": "UseFallbackAbility",
                    "fallback_ability_id": abilities["banishing"],
                },
            )
            next_candidate += 1
            add(
                "AiTransition",
                {
                    "id": next_transition,
                    "stable_key": (
                        f"goal07.enemy.s11.ai.phase-{phase_index + 1}."
                        f"transition-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "target_state_id": state_ids[
                        (offset + 1) % len(state_ids)
                    ],
                    "condition_id": conditions["always"],
                    "priority": 0,
                    "timing": "AfterAction",
                },
            )
            next_transition += 1

    linked_specs = {
        "phase-1-support-1": (
            1,
            "0.093333",
            "1",
            "100",
            abilities["support-cycle"],
        ),
        "phase-1-support-2": (
            2,
            "0.093333",
            "1",
            "100",
            abilities["support-cycle"],
        ),
        "phase-1-support-3": (
            4,
            "0.093333",
            "1",
            "100",
            abilities["support-cycle"],
        ),
        "phase-1-support-4": (
            5,
            "0.093333",
            "1",
            "100",
            abilities["support-cycle"],
        ),
        "phase-2-direwolf-left": (
            1,
            "0.266667",
            "0.833333",
            "144",
            abilities["direwolf-strike"],
        ),
        "phase-2-direwolf-right": (
            5,
            "0.266667",
            "0.833333",
            "144",
            abilities["direwolf-strike"],
        ),
        "phase-3-arm": (
            5,
            "0.233333",
            "1",
            "120",
            abilities["arm-cycle"],
        ),
    }
    for key, id_ in linked.items():
        formation, hp_ratio, atk_ratio, spd, action_ability = linked_specs[key]
        ability_ids = [action_ability]
        if key.startswith("phase-1-support"):
            ability_ids.extend(
                [
                    abilities["beetle-strike"],
                    abilities["hound-strike"],
                ]
            )
        elif key == "phase-3-arm":
            ability_ids.extend(
                [
                    abilities["oppressive"],
                    abilities["disabling"],
                    abilities["overload"],
                    abilities["controlled"],
                ]
            )
        add(
            "LinkedUnitDefinition",
            {
                "id": id_,
                "source_definition_identity_id": id_,
                "kind": "Summon",
                "presence": "Present",
                "ability_ids": "|".join(str(value) for value in ability_ids),
                "action_ability_id": action_ability,
                "formation_index": formation,
                "initial_gauge_decimal": "10000",
                "hp_owner_ratio_decimal": hp_ratio,
                "hp_flat_decimal": "0",
                "atk_owner_ratio_decimal": atk_ratio,
                "atk_flat_decimal": "0",
                "def_owner_ratio_decimal": "1",
                "def_flat_decimal": "0",
                "spd_owner_ratio_decimal": "0",
                "spd_flat_decimal": spd,
                "owner_defeat_policy": "Depart",
                "owner_departure_policy": "Depart",
                "wave_policy": "Depart",
                "combatant_digest_sha256": sha256_text(
                    f"goal07-s11-linked-{key}-v1"
                ),
            },
        )

    add(
        "EnemyTemplate",
        {
            "id": template,
            "rank": "Boss",
            "base_aggro_decimal": "100",
            "default_ai_graph_id": graphs[0],
        },
    )
    add(
        "EnemyVariant",
        {
            "id": variant,
            "template_id": template,
            "ai_graph_id": graphs[0],
            "mechanically_distinct_key": VARIANT_KEY,
        },
    )
    for level in anchor["levels"]:
        add(
            "EnemyStat",
            {
                "variant_id": variant,
                "level": level["authored_level"],
                "difficulty_key": "standard-universe-v1",
                "hp_decimal": level["base_hp"],
                "atk_decimal": level["base_atk"],
                "def_decimal": level["base_def"],
                "spd_decimal": level["base_spd"],
                "effect_hit_rate_decimal": level["effect_hit_rate"],
                "effect_resistance_decimal": level["effect_resistance"],
                "crit_damage_decimal": "0.2",
            },
        )
    for sequence, weakness in enumerate(["Fire", "Lightning", "Wind"], start=1):
        add(
            "EnemyWeakness",
            {"variant_id": variant, "sequence": sequence, "element": weakness},
        )
    for element in ["Ice", "Imaginary", "Physical", "Quantum"]:
        add(
            "EnemyResistance",
            {
                "variant_id": variant,
                "element": element,
                "value_decimal": "0.2",
            },
        )
    for category in [
        "STAT_CTRL_Confine",
        "STAT_CTRL_Frozen",
        "STAT_CTRL_Entangle",
    ]:
        add(
            "EnemyDebuffResistance",
            {
                "variant_id": variant,
                "category_key": category,
                "value_decimal": "0.75",
            },
        )
    add(
        "EnemyToughnessLayer",
        {
            "variant_id": variant,
            "sequence": 1,
            "layer_key": "ordinary",
            "kind": "Ordinary",
            "maximum_decimal": "360",
            "recovery_ratio_decimal": "1",
            "active_at_start": True,
        },
    )
    for sequence, ability_key in enumerate(
        [
            "banishing",
            "burning",
            "bombardment",
            "power",
            "emergency",
            "tactical",
            "boost",
        ],
        start=1,
    ):
        add(
            "EnemyVariantAbility",
            {
                "variant_id": variant,
                "sequence": sequence,
                "ability_id": abilities[ability_key],
            },
        )
    for phase, graph in enumerate(graphs, start=1):
        add(
            "EnemyPhase",
            {
                "id": BASE + 600 + phase,
                "stable_key": f"goal07.enemy.s11.phase-{phase}",
                "variant_id": variant,
                "sequence": phase,
                "entry_condition_id": conditions["always"],
                "exit_condition_id": conditions["always"],
                "replacement_priority": phase,
                "ai_graph_id": graph,
                "targetable": True,
                "transition_model": "TransformSameUnit",
                "hp_carry": "Reset",
                "action_gauge_carry": "Reset",
                "effect_carry": "Clear",
                "toughness_carry": "Reset",
                "summon_carry": "Clear",
            },
        )

    anchor_digest = sha256_bytes(anchor_path(PARTITION).read_bytes())
    add(
        "SourceRecord",
        {
            "id": SOURCE_RECORD_ID,
            "stable_key": "source.hsr-wiki.svarog-complete.2026-07-29",
            "category": "CommunityMaintained",
            "publisher": anchor["source"]["publisher"],
            "url": anchor["source"]["url"],
            "accessed_on": anchor["source"]["accessed_on"],
            "applicable_game_version": anchor["source"]["game_version"],
            "confidence": "SecondaryVersionSensitiveCrossCheck",
            "evidence_sha256": anchor_digest,
            "usage_note": (
                "Exact World 4 levels and public Svarog mechanics are "
                "committed as Goal 07 evidence."
            ),
        },
    )
    add(
        "EvidenceRecord",
        {
            "id": EVIDENCE_RECORD_ID,
            "stable_key": "evidence.goal07.enemy.s11.numeric-anchors",
            "kind": "SourcePayload",
            "source_record_id": SOURCE_RECORD_ID,
            "sha256": anchor_digest,
            "note": (
                "Committed exact public per-level numeric anchors for "
                "Goal 07 S11."
            ),
        },
    )
    for item in identities:
        add("ContentIdentity", item)
        add(
            "ContentEvidenceBinding",
            {
                "content_id": item["id"],
                "sequence": 1,
                "fact_key": f"goal07.s11.executable:{item['stable_key']}",
                "source_record_id": 1,
                "evidence_record_id": 3,
                "quality": "ExactStructured",
                "mechanism_quality": "ExactStructured",
            },
        )
    add(
        "ContentEvidenceBinding",
        {
            "content_id": variant,
            "sequence": 2,
            "fact_key": "goal07.s11.public-level-stats",
            "source_record_id": SOURCE_RECORD_ID,
            "evidence_record_id": EVIDENCE_RECORD_ID,
            "quality": "ExactStructured",
            "mechanism_quality": "ExactStructured",
        },
    )
    for table_rows in rows.values():
        table_rows.sort(
            key=lambda row: json.dumps(
                row,
                ensure_ascii=False,
                sort_keys=True,
                default=str,
            )
        )
    return rows


def enemy_specs_s13() -> list[dict[str, Any]]:
    return [
        {
            "key": VARIANT_KEYS[0],
            "template": "enemy.automaton-spider.minionlv2",
            "name": "Automaton Spider",
            "zh": "自动机兵•蜘蛛",
            "rank": "Normal",
            "toughness": "60",
            "weaknesses": ["Lightning", "Wind"],
            "resistances": [
                ("Fire", "0.2"),
                ("Ice", "0.2"),
                ("Imaginary", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("ram", "Ram", "冲撞", "101202001", "2", "Physical", "primary", True, None),
                ("timed-module", "Timed Module", "定时模块", "101202002", None, None, "actor", True, "timed"),
                (
                    "self-exploding-module",
                    "Self-Exploding Module",
                    "自爆模块",
                    "101202003",
                    "5",
                    "Fire",
                    "blast",
                    True,
                    "self-explode",
                ),
                (
                    "chains-of-destruction",
                    "Chains of Destruction",
                    "毁灭连锁",
                    "101202004",
                    None,
                    None,
                    "actor",
                    False,
                    "chains",
                ),
            ],
            "cycle": ["ram", "timed-module", "self-exploding-module"],
        },
        {
            "key": VARIANT_KEYS[1],
            "template": "enemy.baryon.minion",
            "name": "Baryon",
            "zh": "重子",
            "rank": "Minion",
            "toughness": "30",
            "weaknesses": ["Ice", "Wind"],
            "resistances": [
                ("Fire", "0.2"),
                ("Imaginary", "0.2"),
                ("Lightning", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("obliterate", "Obliterate", "湮灭", "801101001", "2.5", "Quantum", "primary", True, None),
            ],
            "cycle": ["obliterate"],
        },
        {
            "key": VARIANT_KEYS[2],
            "template": "enemy.cloud-knights-patroller.minionlv2",
            "name": "Cloud Knights Patroller",
            "zh": "云骑巡防士卒",
            "rank": "Normal",
            "toughness": "60",
            "weaknesses": ["Fire", "Imaginary", "Wind"],
            "resistances": [
                ("Ice", "0.2"),
                ("Lightning", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("cloud-edge", "Cloud Edge", "云刃", "200201001", "2.5", "Wind", "primary", True, "cloud-edge"),
                ("patrol-sweep", "Patrol Sweep", "巡防扫击", "200201004", "3.5", "Physical", "primary", True, None),
            ],
            "cycle": ["cloud-edge", "patrol-sweep"],
        },
        {
            "key": VARIANT_KEYS[3],
            "template": "enemy.decaying-shadow.elite",
            "name": "Decaying Shadow",
            "zh": "蚕食者之影",
            "rank": "Elite",
            "toughness": "300",
            "weaknesses": ["Fire", "Lightning", "Wind"],
            "resistances": [
                ("Ice", "0.2"),
                ("Imaginary", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("gilded-spikes", "Fleeting Gilded Spikes", "流年金刺", "800304002", "3", "Imaginary", "primary", True, None),
                ("binding", "Binding of the Golden Age", "黄金年代的绸缪", "800304004", None, None, "actor", True, "binding"),
                ("punishment", "Fleeting Punishment", "流年谴罚", "800304003", "5", "Imaginary", "primary", True, None),
                ("liberation", "Liberation of the Golden Age", "黄金年代的解放", "800304005", "2.25", "Imaginary", "all-opposing", True, "liberation"),
            ],
            "cycle": ["gilded-spikes", "binding", "punishment", "liberation"],
        },
        {
            "key": VARIANT_KEYS[4],
            "template": "enemy.disciples-of-sanctus-medicus-ballistarius.minionlv2",
            "name": "Disciples of Sanctus Medicus: Ballistarius",
            "zh": "「药王秘传」器元士",
            "rank": "Normal",
            "toughness": "90",
            "weaknesses": ["Ice", "Physical", "Wind"],
            "resistances": [
                ("Fire", "0.2"),
                ("Imaginary", "0.2"),
                ("Lightning", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("mounted-crossbow", "Mounted Crossbow", "犀弩", "202206001", None, "Lightning", "random", True, "bounce-6"),
            ],
            "cycle": ["mounted-crossbow"],
        },
        {
            "key": VARIANT_KEYS[5],
            "template": "enemy.disciples-of-sanctus-medicus-internal-alchemist.minionlv2",
            "name": "Disciples of Sanctus Medicus: Internal Alchemist",
            "zh": "「药王秘传」内丹士",
            "rank": "Normal",
            "toughness": "60",
            "weaknesses": ["Imaginary", "Physical"],
            "resistances": [
                ("Fire", "0.2"),
                ("Ice", "0.2"),
                ("Lightning", "0.4"),
                ("Quantum", "0.2"),
                ("Wind", "0.2"),
            ],
            "abilities": [
                ("seeding-core", "Seeding Core", "种丹", "202202001", None, None, "primary", True, "seed-core"),
                ("thundering-reign", "Thundering Reign", "役雷", "202202002", "2.5", "Lightning", "primary", True, None),
                ("core-rupture", "Core Rupture", "丹毁", "202202003", "4.5", "Lightning", "blast", False, "core-rupture"),
            ],
            "cycle": ["seeding-core", "thundering-reign"],
        },
        {
            "key": VARIANT_KEYS[6],
            "template": "enemy.disciples-of-sanctus-medicus-shape-shifter-bug.elite",
            "name": "Disciples of Sanctus Medicus: Shape Shifter (Bug)",
            "zh": "「药王秘传」炼形者（错误）",
            "rank": "Elite",
            "toughness": "360",
            "weaknesses": ["Ice", "Imaginary", "Wind"],
            "resistances": [
                ("Fire", "0.2"),
                ("Lightning", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("thunder-shock", "Thunder-Shock", "震击", "202301101", "3.5", "Lightning", "primary", True, "drain-03"),
                ("punishing-bolts", "Punishing Bolts", "霆罚", "202301102", "3", "Lightning", "primary", True, "drain-03"),
                ("sympathicus", "Sympathicus", "交感", "202301105", None, None, "actor", True, "mara-summon"),
                ("thudding-calamity", "Thudding Calamity", "劫雷", "202301103", "4.5", "Lightning", "primary", True, "drain-03"),
            ],
            "cycle": ["sympathicus", "thunder-shock", "punishing-bolts", "thudding-calamity"],
        },
        {
            "key": VARIANT_KEYS[7],
            "template": "enemy.disciples-of-sanctus-medicus-shape-shifter.elite",
            "name": "Disciples of Sanctus Medicus: Shape Shifter",
            "zh": "「药王秘传」炼形者",
            "rank": "Elite",
            "toughness": "360",
            "weaknesses": ["Ice", "Imaginary", "Wind"],
            "resistances": [
                ("Fire", "0.2"),
                ("Lightning", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("thunder-shock", "Thunder-Shock", "震击", "202301001", "3.5", "Lightning", "primary", True, "drain-025"),
                ("punishing-bolts", "Punishing Bolts", "霆罚", "202301002", "3", "Lightning", "primary", True, "drain-025"),
                ("mara-summon", "Mara-Summon", "召遣", "202301004", None, None, "actor", True, "mara-summon"),
                ("thudding-calamity", "Thudding Calamity", "劫雷", "202301003", "5", "Lightning", "primary", True, "drain-025"),
            ],
            "cycle": ["mara-summon", "thunder-shock", "punishing-bolts", "thudding-calamity"],
        },
        {
            "key": VARIANT_KEYS[8],
            "template": "enemy.dreamjolt-troupes-beyond-overcooked-bug.elite",
            "name": "Dreamjolt Troupe's Beyond Overcooked (Bug)",
            "zh": "惊梦剧团的十七分熟（错误）",
            "rank": "Elite",
            "toughness": "720",
            "weaknesses": ["Fire", "Lightning", "Wind"],
            "resistances": [
                ("Ice", "0.2"),
                ("Imaginary", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("ladle-swirl", "Ladle Swirl", "大力颠勺", "300301201", "3.8", "Fire", "primary", True, None),
                ("tornado-toss", "Tornado Toss", "旋风翻锅", "300301202", "3", "Fire", "primary", True, None),
                ("stove-detonation", "Stove Detonation", "引爆炉灶", "300301203", None, None, "actor", True, "stove"),
                ("flaming-stir-fry", "Flaming Stir-Fry", "烈焰爆炒", "300301204", "5.8", "Fire", "all-opposing", True, None),
                ("kitchen-mishap", "Kitchen Mishap", "厨房事故", "300301205", "12", "Fire", "all-allies", True, None),
            ],
            "cycle": ["ladle-swirl", "tornado-toss", "stove-detonation", "flaming-stir-fry", "kitchen-mishap"],
        },
        {
            "key": VARIANT_KEYS[9],
            "template": "enemy.dreamjolt-troupes-beyond-overcooked.elite",
            "name": "Dreamjolt Troupe's Beyond Overcooked",
            "zh": "惊梦剧团的十七分熟",
            "rank": "Elite",
            "toughness": "480",
            "weaknesses": ["Fire", "Lightning", "Wind"],
            "resistances": [
                ("Ice", "0.2"),
                ("Imaginary", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("ladle-swirl", "Ladle Swirl", "大力颠勺", "300301001", "3.2", "Fire", "primary", True, None),
                ("tornado-toss", "Tornado Toss", "旋风翻锅", "300301002", "2.8", "Fire", "primary", True, None),
                ("ignite-stove", "Ignite Stove", "点燃炉灶", "300301003", None, None, "actor", True, "stove"),
                ("flaming-stir-fry", "Flaming Stir-Fry", "烈焰爆炒", "300301004", "6.2", "Fire", "all-opposing", True, None),
                ("kitchen-mishap", "Kitchen Mishap", "厨房事故", "300301005", "6", "Fire", "all-allies", True, None),
            ],
            "cycle": ["ladle-swirl", "tornado-toss", "ignite-stove", "flaming-stir-fry", "kitchen-mishap"],
        },
        {
            "key": VARIANT_KEYS[10],
            "template": "enemy.dreamjolt-troupes-birdskull.minion",
            "name": "Dreamjolt Troupe's Birdskull",
            "zh": "惊梦剧团的舞鸫假面",
            "rank": "Minion",
            "toughness": "30",
            "weaknesses": ["Imaginary", "Lightning"],
            "resistances": [
                ("Fire", "0.2"),
                ("Ice", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
                ("Wind", "0.2"),
            ],
            "abilities": [
                ("dance-invite", "Dance Invite", "邀舞", "300102001", "2.5", "Physical", "primary", True, "bleed"),
            ],
            "cycle": ["dance-invite"],
        },
        {
            "key": VARIANT_KEYS[11],
            "template": "enemy.dreamjolt-troupes-bubble-hound.minionlv2",
            "name": "Dreamjolt Troupe's Bubble Hound",
            "zh": "惊梦剧团的气泡锂犬",
            "rank": "Normal",
            "toughness": "90",
            "weaknesses": ["Lightning", "Physical", "Wind"],
            "resistances": [
                ("Fire", "0.2"),
                ("Ice", "0.2"),
                ("Imaginary", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("drink-invite", "Drink Invite", "邀饮", "300203001", "2.5", "Physical", "primary", True, None),
                ("libation-of-sweetness", "Libation of Sweetness", "泼洒甜蜜", "300203002", None, None, "actor", True, "sweetness"),
                ("hearty-revelry", "Hearty Revelry", "酣畅", "300203003", None, None, "actor", False, "sweetness"),
            ],
            "cycle": ["drink-invite", "libation-of-sweetness"],
        },
    ]


def enemy_specs_s14() -> list[dict[str, Any]]:
    return [
        {
            "key": VARIANT_KEYS[0],
            "template": "enemy.dreamjolt-troupes-mr-domescreen.minionlv2",
            "name": "Dreamjolt Troupe's Mr. Domescreen",
            "zh": "惊梦剧团的圆幕先生",
            "rank": "Normal",
            "toughness": "90",
            "weaknesses": ["Ice", "Imaginary", "Lightning"],
            "resistances": [("Fire", "0.2"), ("Physical", "0.2"), ("Quantum", "0.2"), ("Wind", "0.2")],
            "abilities": [
                ("surprise-variety-show", "Surprise Variety Show", "惊喜综艺", "300201001", None, None, "primary", True, "channel"),
                ("startling-broadcast", "Startling Broadcast", "惊吓播报", "300201002", "1.6", "Quantum", "all-opposing", True, "channel"),
                ("channel-switch", "Channel Switch", "频道切换", "300201003", None, None, "actor", True, "channel"),
                ("channel-crossover", "Channel Crossover", "串台", "300201004", None, None, "actor", False, "channel"),
            ],
            "cycle": ["surprise-variety-show", "startling-broadcast", "channel-switch"],
        },
        {
            "key": VARIANT_KEYS[1],
            "template": "enemy.dreamjolt-troupes-spring-loader.minion",
            "name": "Dreamjolt Troupe's Spring Loader",
            "zh": "惊梦剧团的弹簧荷官",
            "rank": "Minion",
            "toughness": "30",
            "weaknesses": ["Physical", "Wind"],
            "resistances": [("Fire", "0.2"), ("Ice", "0.2"), ("Imaginary", "0.2"), ("Lightning", "0.2"), ("Quantum", "0.2")],
            "abilities": [
                ("throw-die", "Throw Die", "掷点", "300101001", "2.5", "Physical", "primary", True, "die-debuff"),
            ],
            "cycle": ["throw-die"],
        },
        {
            "key": VARIANT_KEYS[2],
            "template": "enemy.dreamjolt-troupes-sweet-gorilla-bug.elite",
            "name": "Dreamjolt Troupe's Sweet Gorilla (Bug)",
            "zh": "惊梦剧团的甜猿泰山（错误）",
            "rank": "Elite",
            "toughness": "360",
            "weaknesses": ["Fire", "Imaginary", "Physical"],
            "resistances": [("Ice", "0.2"), ("Lightning", "0.2"), ("Quantum", "0.2"), ("Wind", "0.2")],
            "abilities": [
                ("limitless-free-drinks", "Limitless Free Drinks", "不限量赠饮", "300302201", "3.8", "Ice", "primary", True, "drink-shield"),
                ("frenzied-free-drinks", "Frenzied Free Drinks", "疯狂畅饮", "300302202", "3", "Ice", "primary", True, "drink-shield"),
                ("soulglad-party", "SoulGlad Party", "苏乐达派对", "300302203", None, None, "actor", True, "soulglad"),
                ("celebration-fountain", "Celebration Fountain", "酬宾喷泉", "300302204", None, None, "actor", True, "fountain"),
            ],
            "cycle": ["limitless-free-drinks", "frenzied-free-drinks", "soulglad-party", "celebration-fountain"],
        },
        {
            "key": VARIANT_KEYS[3],
            "template": "enemy.dreamjolt-troupes-sweet-gorilla.elite",
            "name": "Dreamjolt Troupe's Sweet Gorilla",
            "zh": "惊梦剧团的甜猿泰山",
            "rank": "Elite",
            "toughness": "300",
            "weaknesses": ["Fire", "Imaginary", "Physical"],
            "resistances": [("Ice", "0.2"), ("Lightning", "0.2"), ("Quantum", "0.2"), ("Wind", "0.2")],
            "abilities": [
                ("limited-free-drinks", "Limited Free Drinks", "限量赠饮", "300302001", "3.8", "Ice", "primary", True, "drink-shield"),
                ("unlimited-free-drinks", "Unlimited Free Drinks", "无限畅饮", "300302002", "3.6", "Ice", "primary", True, "drink-shield"),
                ("soulglad-party", "SoulGlad Party", "苏乐达派对", "300302003", None, None, "actor", True, "soulglad"),
                ("celebration-fountain", "Celebration Fountain", "酬宾喷泉", "300302004", None, None, "actor", True, "fountain"),
            ],
            "cycle": ["limited-free-drinks", "unlimited-free-drinks", "soulglad-party", "celebration-fountain"],
        },
        {
            "key": VARIANT_KEYS[4],
            "template": "enemy.dreamjolt-troupes-winder-goon.minionlv2",
            "name": "Dreamjolt Troupe's Winder Goon",
            "zh": "惊梦剧团的发条暴徒",
            "rank": "Normal",
            "toughness": "90",
            "weaknesses": ["Lightning", "Quantum"],
            "resistances": [("Fire", "0.2"), ("Ice", "0.2"), ("Imaginary", "0.2"), ("Physical", "0.2"), ("Wind", "0.2")],
            "abilities": [
                ("set-alarm", "Set Alarm", "预定闹钟", "300204001", None, None, "actor", True, "alarm"),
                ("wake-up-alarm", "Wake-Up Alarm", "起床铃声", "300204002", None, None, "actor", True, "alarm"),
                ("wake-up-punch", "Wake-Up Punch", "清醒铁拳", "300204003", "2.5", "Physical", "primary", True, None),
            ],
            "cycle": ["set-alarm", "wake-up-alarm", "wake-up-punch"],
        },
        {
            "key": VARIANT_KEYS[5],
            "template": "enemy.entranced-ingenium-golden-cloud-toad.minion",
            "name": "Entranced Ingenium: Golden Cloud Toad",
            "zh": "入魔机巧 • 浓云金蟾",
            "rank": "Minion",
            "toughness": "30",
            "weaknesses": ["Lightning", "Quantum"],
            "resistances": [("Fire", "0.2"), ("Ice", "0.2"), ("Imaginary", "0.2"), ("Physical", "0.2"), ("Wind", "0.2")],
            "abilities": [
                ("dense-fog", "Dense Fog", "霏雾", "201102001", "2", "Imaginary", "primary", True, "fog"),
            ],
            "cycle": ["dense-fog"],
        },
        {
            "key": VARIANT_KEYS[6],
            "template": "enemy.entranced-ingenium-illumination-dragonfish.minionlv2",
            "name": "Entranced Ingenium: Illumination Dragonfish",
            "zh": "入魔机巧 • 灯昼龙鱼",
            "rank": "Normal",
            "toughness": "60",
            "weaknesses": ["Imaginary", "Lightning", "Physical"],
            "resistances": [("Fire", "0.2"), ("Ice", "0.2"), ("Quantum", "0.2"), ("Wind", "0.2")],
            "abilities": [
                ("nightly-luminosity", "Nightly Luminosity", "宵明", "201201001", "1.5", "Fire", "all-opposing", True, "def-down"),
                ("candle-flame", "Candle Flame", "烛花", "201201002", None, "Fire", "actor", False, "candle-flame"),
            ],
            "cycle": ["nightly-luminosity"],
        },
        {
            "key": VARIANT_KEYS[7],
            "template": "enemy.entranced-ingenium-obedient-dracolion.minion",
            "name": "Entranced Ingenium: Obedient Dracolion",
            "zh": "入魔机巧 • 率从狻猊",
            "rank": "Minion",
            "toughness": "30",
            "weaknesses": ["Imaginary", "Wind"],
            "resistances": [("Fire", "0.2"), ("Ice", "0.2"), ("Lightning", "0.2"), ("Physical", "0.2"), ("Quantum", "0.2")],
            "abilities": [
                ("lions-roar", "Lion's Roar", "狮子吼", "201101001", "2", "Fire", "primary", True, None),
            ],
            "cycle": ["lions-roar"],
        },
        {
            "key": VARIANT_KEYS[8],
            "template": "enemy.everwinter-shadewalker.minionlv2",
            "name": "Everwinter Shadewalker",
            "zh": "永冬灾影",
            "rank": "Normal",
            "toughness": "60",
            "weaknesses": ["Fire", "Physical", "Quantum"],
            "resistances": [("Ice", "0.4"), ("Imaginary", "0.2"), ("Lightning", "0.2"), ("Wind", "0.2")],
            "abilities": [
                ("frost-crush", "Frost Crush", "霜冻重击", "102201001", "3", "Ice", "primary", True, "delay-50"),
            ],
            "cycle": ["frost-crush"],
        },
        {
            "key": VARIANT_KEYS[9],
            "template": "enemy.flamespawn.minion",
            "name": "Flamespawn",
            "zh": "炎华造物",
            "rank": "Minion",
            "toughness": "30",
            "weaknesses": ["Ice", "Physical"],
            "resistances": [("Fire", "0.4"), ("Imaginary", "0.2"), ("Lightning", "0.2"), ("Quantum", "0.2"), ("Wind", "0.2")],
            "abilities": [
                ("distract", "Distract", "扰乱", "800101001", "2.5", "Fire", "primary", True, None),
            ],
            "cycle": ["distract"],
        },
        {
            "key": VARIANT_KEYS[10],
            "template": "enemy.frigid-prowler.elite",
            "name": "Frigid Prowler",
            "zh": "深寒徘徊者",
            "rank": "Elite",
            "toughness": "300",
            "weaknesses": ["Fire", "Lightning", "Quantum"],
            "resistances": [("Ice", "0.4"), ("Imaginary", "0.2"), ("Physical", "0.2"), ("Wind", "0.2")],
            "abilities": [
                ("ice-wheel-fist", "Ice Wheel Fist", "冰轮铁拳", "102302001", "3", "Ice", "primary", True, None),
                ("summon-otherling", "Summon Otherling", "召来衍生物", "102302005", None, None, "actor", True, "otherling"),
                ("devour-otherling", "Devour Otherling", "吞噬衍生物", "102302007", None, None, "actor", True, "devour"),
                ("ice-wheel-crusher", "Ice Wheel Crusher", "冰轮碎碾", "102302009", "5", "Ice", "primary", True, "deep-freeze"),
                ("frozen-storm", "Frozen Storm", "冻结风暴", "102302006", "1", "Ice", "all-opposing", True, "deep-freeze"),
            ],
            "cycle": ["summon-otherling", "ice-wheel-fist", "devour-otherling", "ice-wheel-crusher", "frozen-storm"],
        },
        {
            "key": VARIANT_KEYS[11],
            "template": "enemy.frostspawn.minion",
            "name": "Frostspawn",
            "zh": "霜晶造物",
            "rank": "Minion",
            "toughness": "30",
            "weaknesses": ["Fire", "Wind"],
            "resistances": [("Ice", "0.4"), ("Imaginary", "0.2"), ("Lightning", "0.2"), ("Physical", "0.2"), ("Quantum", "0.2")],
            "abilities": [
                ("distract", "Distract", "扰乱", "800102001", "2.5", "Ice", "primary", True, None),
            ],
            "cycle": ["distract"],
        },
    ]


def owned_rows_ordinary() -> dict[str, list[dict[str, Any]]]:
    anchor = json.loads(anchor_path(PARTITION).read_text(encoding="utf-8"))
    manifest = json.loads(PARTITIONS.read_text(encoding="utf-8"))
    assigned = next(item for item in manifest["partitions"] if item["id"] == PARTITION)
    if assigned["enemy_variant_ids"] != VARIANT_KEYS:
        raise ValueError("S12 frozen enemy assignment changed")
    if [item["enemy_variant_id"] for item in anchor["variants"]] != VARIANT_KEYS:
        raise ValueError("S12 numeric anchor variants changed")

    enemy_specs = (
        enemy_specs_s14()
        if PARTITION == "G07-P5-M15-S14"
        else enemy_specs_s13()
        if PARTITION == "G07-P5-M15-S13"
        else [
        {
            "key": VARIANT_KEYS[0],
            "template": "enemy.abundance-sprite-golden-hound.minionlv2",
            "name": "Abundance Sprite: Golden Hound",
            "zh": "丰饶灵兽·娄金",
            "rank": "Normal",
            "toughness": "90",
            "weaknesses": ["Ice", "Lightning", "Quantum"],
            "resistances": [
                ("Fire", "0.2"),
                ("Imaginary", "0.2"),
                ("Physical", "0.2"),
                ("Wind", "0.2"),
            ],
            "abilities": [
                ("horn-gore", "Horn Gore", "角牴", "202204001", "2.5", "Wind", "primary", True, None),
                ("rebound-roar", "Rebound Roar", "嗥吠", "202204002", None, None, "actor", False, "rebound"),
            ],
            "cycle": ["horn-gore"],
        },
        {
            "key": VARIANT_KEYS[1],
            "template": "enemy.abundance-sprite-malefic-ape-bug.elite",
            "name": "Abundance Sprite: Malefic Ape (Bug)",
            "zh": "丰饶灵兽·长右（错误）",
            "rank": "Elite",
            "toughness": "360",
            "weaknesses": ["Fire", "Ice", "Wind"],
            "resistances": [
                ("Imaginary", "0.2"),
                ("Lightning", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("wrathful-roar", "Wrathful Roar", "怒啸", "202302102", None, None, "random", True, "monitor"),
                ("simian-strike", "Simian Strike", "猱击", "202302101", "3", "Quantum", "primary", True, None),
                ("simian-howl", "Simian Howl", "猿啼", "202302104", "2.5", "Quantum", "all-opposing", True, None),
            ],
            "cycle": ["wrathful-roar", "simian-strike", "simian-strike", "simian-howl"],
        },
        {
            "key": VARIANT_KEYS[2],
            "template": "enemy.abundance-sprite-malefic-ape.elite",
            "name": "Abundance Sprite: Malefic Ape",
            "zh": "丰饶灵兽·长右",
            "rank": "Elite",
            "toughness": "360",
            "weaknesses": ["Fire", "Ice", "Wind"],
            "resistances": [
                ("Imaginary", "0.2"),
                ("Lightning", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("wrathful-roar", "Wrathful Roar", "怒啸", "202302004", None, None, "random", True, "monitor"),
                ("simian-strike", "Simian Strike", "猱击", "202302001", "4.75", "Quantum", "primary", True, None),
                ("fiery-majesty", "Fiery Majesty", "奋威", "202302003", None, None, "actor", True, "fiery"),
            ],
            "cycle": ["wrathful-roar", "simian-strike", "fiery-majesty", "simian-strike"],
        },
        {
            "key": VARIANT_KEYS[3],
            "template": "enemy.abundance-sprite-wooden-lupus.minionlv2",
            "name": "Abundance Sprite: Wooden Lupus",
            "zh": "丰饶灵兽·奎木",
            "rank": "Normal",
            "toughness": "90",
            "weaknesses": ["Fire", "Imaginary", "Wind"],
            "resistances": [
                ("Ice", "0.2"),
                ("Lightning", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("rallying-howl", "Rallying Howl", "啸聚", "202205002", None, None, "actor", True, "rally"),
                ("wolf-predation", "Wolf Predation", "狼吞", "202205001", "2", "Lightning", "primary", True, "bleed"),
            ],
            "cycle": ["rallying-howl", "wolf-predation"],
        },
        {
            "key": VARIANT_KEYS[4],
            "template": "enemy.antibaryon.minion",
            "name": "Antibaryon",
            "zh": "反重子",
            "rank": "Minion",
            "toughness": "30",
            "weaknesses": ["Physical", "Quantum"],
            "resistances": [
                ("Fire", "0.2"),
                ("Ice", "0.2"),
                ("Imaginary", "0.2"),
                ("Lightning", "0.2"),
                ("Wind", "0.2"),
            ],
            "abilities": [
                ("obliterate", "Obliterate", "湮灭", "801102001", "2.5", "Imaginary", "primary", True, None),
            ],
            "cycle": ["obliterate"],
        },
        {
            "key": VARIANT_KEYS[5],
            "template": "enemy.aurumaton-gatekeeper-bug.elite",
            "name": "Aurumaton Gatekeeper (Bug)",
            "zh": "金人司阍（错误）",
            "rank": "Elite",
            "toughness": "300",
            "weaknesses": ["Lightning", "Quantum", "Wind"],
            "resistances": [
                ("Fire", "0.2"),
                ("Ice", "0.2"),
                ("Imaginary", "0.4"),
                ("Physical", "0.2"),
            ],
            "abilities": [
                ("penal-code", "Penal Code", "惩戒律", "201301101", None, None, "actor", True, "sanction-bug"),
                ("dread", "Dread", "惩戒", "201301102", "3", "Imaginary", "all-opposing", True, "weaken"),
                ("track-down", "Track Down", "追缉", "201301104", "6", "Imaginary", "primary", True, "imprison-60"),
            ],
            "cycle": ["penal-code", "dread", "track-down"],
        },
        {
            "key": VARIANT_KEYS[6],
            "template": "enemy.aurumaton-gatekeeper.elite",
            "name": "Aurumaton Gatekeeper",
            "zh": "金人司阍",
            "rank": "Elite",
            "toughness": "300",
            "weaknesses": ["Lightning", "Quantum", "Wind"],
            "resistances": [
                ("Fire", "0.2"),
                ("Ice", "0.2"),
                ("Imaginary", "0.4"),
                ("Physical", "0.2"),
            ],
            "abilities": [
                ("dread", "Dread", "惩戒", "201301002", "1.5", "Imaginary", "all-opposing", True, "weaken"),
                ("farewell-etiquette", "Farewell Etiquette", "大逆不道", "201301001", None, None, "actor", True, "sanction"),
                ("restraint", "Restraint", "拘束", "201301003", "3", "Imaginary", "primary", True, None),
                ("enchainment", "Enchainment", "锁禁", "201301004", "5", "Imaginary", "primary", True, "imprison-50"),
            ],
            "cycle": ["dread", "dread", "dread", "farewell-etiquette", "restraint", "enchainment"],
        },
        {
            "key": VARIANT_KEYS[7],
            "template": "enemy.aurumaton-spectral-envoy.elite",
            "name": "Aurumaton Spectral Envoy",
            "zh": "金人勾魂使",
            "rank": "Elite",
            "toughness": "300",
            "weaknesses": ["Imaginary", "Lightning", "Physical"],
            "resistances": [
                ("Fire", "0.2"),
                ("Ice", "0.2"),
                ("Quantum", "0.2"),
                ("Wind", "0.2"),
            ],
            "abilities": [
                ("adjudicate", "Adjudicate", "判罚", "201302001", "3", "Physical", "primary", True, None),
                ("subdue", "Subdue", "镇魂", "201302002", "3.6", "Physical", "primary", True, "reverberation-70"),
                ("revert-yin-and-yang", "Revert Yin and Yang", "倒转阴阳", "201302003", "3", "Physical", "all-opposing", True, "reverberation-35"),
                ("heavens-fall", "Heaven's Fall", "天坠", "201302004", "4", "Physical", "primary", True, None),
            ],
            "cycle": ["adjudicate", "subdue", "revert-yin-and-yang", "heavens-fall"],
        },
        {
            "key": VARIANT_KEYS[8],
            "template": "enemy.automaton-beetle.minionlv2",
            "name": "Automaton Beetle",
            "zh": "自动机兵·甲虫",
            "rank": "Normal",
            "toughness": "60",
            "weaknesses": ["Imaginary", "Lightning", "Wind"],
            "resistances": [
                ("Fire", "0.2"),
                ("Ice", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
            ],
            "abilities": [
                ("unstable-forcefield", "Unstable Forcefield", "不稳定力场", "101203001", "3", "Physical", "primary", True, "bleed"),
            ],
            "cycle": ["unstable-forcefield"],
        },
        {
            "key": VARIANT_KEYS[9],
            "template": "enemy.automaton-direwolf.elite",
            "name": "Automaton Direwolf",
            "zh": "自动机兵·齿狼",
            "rank": "Elite",
            "toughness": "300",
            "weaknesses": ["Ice", "Imaginary", "Lightning"],
            "resistances": [
                ("Fire", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
                ("Wind", "0.2"),
            ],
            "abilities": [
                ("lock-on-target", "Lock On Target", "目标锁定", "101302002", None, None, "primary", True, "lock"),
                ("felling-order", "Felling Order", "砍伐指令", "101302001", "0.5", "Physical", "primary", True, "bleed"),
                ("disintegration-order", "Disintegration Order", "解体指令", "101302003", "2.5", "Physical", "primary", True, None),
                ("dismantle", "Dismantle", "解体拆除", "101302004", "2.5", "Physical", "primary", True, None),
            ],
            "cycle": ["lock-on-target", "felling-order", "disintegration-order", "dismantle"],
        },
        {
            "key": VARIANT_KEYS[10],
            "template": "enemy.automaton-grizzly.elite",
            "name": "Automaton Grizzly",
            "zh": "自动机兵·灰熊",
            "rank": "Elite",
            "toughness": "420",
            "weaknesses": ["Fire", "Ice", "Lightning"],
            "resistances": [
                ("Imaginary", "0.2"),
                ("Physical", "0.2"),
                ("Quantum", "0.2"),
                ("Wind", "0.2"),
            ],
            "abilities": [
                ("destruction-order", "Destruction Order", "破坏指令", "101301009", "3", "Physical", "primary", True, None),
                ("detonation-order", "Detonation Order", "爆裂指令", "101301001", None, None, "actor", True, "rally"),
                ("enrage-order", "Enrage Order", "愤怒指令", "101301002", None, None, "all-opposing", True, "weaken"),
                ("overcombust-order", "Overcombust Order", "过载指令", "101301003", None, None, "actor", True, "overcombust"),
                ("purge-order", "Purge Order", "剿灭指令", "101301004", "4", "Physical", "all-opposing", True, None),
            ],
            "cycle": ["destruction-order", "detonation-order", "enrage-order", "overcombust-order", "purge-order"],
        },
        {
            "key": VARIANT_KEYS[11],
            "template": "enemy.automaton-hound.minionlv2",
            "name": "Automaton Hound",
            "zh": "自动机兵·战犬",
            "rank": "Normal",
            "toughness": "60",
            "weaknesses": ["Lightning", "Physical"],
            "resistances": [
                ("Fire", "0.2"),
                ("Ice", "0.2"),
                ("Imaginary", "0.2"),
                ("Quantum", "0.2"),
                ("Wind", "0.2"),
            ],
            "abilities": [
                ("vertical-strike", "Vertical Strike", "垂直打击", "101201001", "2.5", "Physical", "primary", True, None),
                ("self-healing-module", "Self-Healing Module", "自愈模块", "101201002", None, None, "all-allies", True, "heal"),
            ],
            "cycle": ["vertical-strike", "self-healing-module"],
        },
    ]
    )
    if PARTITION == "G07-P5-M15-S12":
        variant_row_ids = [
            95 if index == 0 else 96 if index == 6 else BASE + 1 + index
            for index in range(len(enemy_specs))
        ]
        template_row_ids = [
            10_001 if index == 0 else 10_002 if index == 6 else BASE + 21 + index
            for index in range(len(enemy_specs))
        ]
    elif PARTITION == "G07-P5-M15-S13":
        variant_row_ids = [
            97 if index == 0 else 99 if index == 4 else 100 if index == 7 else BASE + 1 + index
            for index in range(len(enemy_specs))
        ]
        template_row_ids = [
            10_003
            if index == 0
            else 10_005
            if index == 4
            else 10_006
            if index == 7
            else BASE + 21 + index
            for index in range(len(enemy_specs))
        ]
    elif PARTITION == "G07-P5-M15-S14":
        variant_row_ids = [
            102
            if index == 6
            else 103
            if index == 9
            else 104
            if index == 11
            else BASE + 1 + index
            for index in range(len(enemy_specs))
        ]
        template_row_ids = [
            10_008
            if index == 6
            else 10_009
            if index == 9
            else 10_010
            if index == 11
            else BASE + 21 + index
            for index in range(len(enemy_specs))
        ]
    else:
        variant_row_ids = [BASE + 1 + index for index in range(len(enemy_specs))]
        template_row_ids = [BASE + 21 + index for index in range(len(enemy_specs))]

    rows: dict[str, list[dict[str, Any]]] = {}
    identities: list[dict[str, Any]] = []
    selectors = {
        "actor": BASE + 401,
        "owner": BASE + 402,
        "current-subject": BASE + 403,
        "primary": BASE + 404,
        "random": BASE + 405,
        "all-opposing": BASE + 406,
        "all-allies": BASE + 407,
        "adjacent-allies": BASE + 408,
        "blast": BASE + 409,
    }
    if PARTITION == "G07-P5-M15-S12":
        selectors.pop("blast")
    effects = (
        {
            "channel": BASE + 5_001,
            "die-debuff": BASE + 5_002,
            "drink-shield": BASE + 5_003,
            "soulglad": BASE + 5_004,
            "fountain": BASE + 5_005,
            "alarm": BASE + 5_006,
            "fog": BASE + 5_007,
            "def-down": BASE + 5_008,
            "candle-flame": BASE + 5_009,
            "deep-freeze": BASE + 5_010,
            "devour": BASE + 5_011,
        }
        if PARTITION == "G07-P5-M15-S14"
        else
        {
            "chains": BASE + 5_001,
            "timed": BASE + 5_002,
            "cloud-edge": BASE + 5_003,
            "binding": BASE + 5_004,
            "seed-core": BASE + 5_005,
            "draining": BASE + 5_006,
            "stove": BASE + 5_007,
            "bleed": BASE + 5_008,
            "sweetness": BASE + 5_009,
            "rebirth": BASE + 5_010,
        }
        if PARTITION == "G07-P5-M15-S13"
        else {
            "rebound": BASE + 5_001,
            "monitor": BASE + 5_002,
            "fiery": BASE + 5_003,
            "rally": BASE + 5_004,
            "bleed": BASE + 5_005,
            "weaken": BASE + 5_006,
            "sanction": BASE + 5_007,
            "sanction-bug": BASE + 5_008,
            "imprisonment": BASE + 5_009,
            "reverberation": BASE + 5_010,
            "lock": BASE + 5_011,
            "overcombust": BASE + 5_012,
        }
    )
    condition_always = BASE + 4_801
    next_program = BASE + 1_001
    next_operation = BASE + 2_001
    next_expression = BASE + 3_001
    next_state = BASE + 4_101
    next_candidate = BASE + 4_301
    next_transition = BASE + 4_501

    def add(table: str, row: dict[str, Any]) -> None:
        rows.setdefault(table, []).append(row)

    def ident(
        id_: int,
        key: str,
        kind: str,
        name_en: str,
        name_zh_cn: str,
        summary: str,
        sources: str = "1",
    ) -> dict[str, Any]:
        row = identity(id_, key, kind, name_en, name_zh_cn, summary, sources)
        row["summary_zh_cn"] = f"Goal 07 {PARTITION[-3:]} 来源绑定的普通敌人批次可执行定义。"
        row["game_version_introduced"] = "1.0"
        return row

    selector_metadata = {
        "actor": ("Actor", "SameSide", "First", 1, 1, "Fault"),
        "owner": ("Owner", "SameSide", "First", 1, 1, "Fault"),
        "current-subject": ("CurrentSubject", "AnySide", "First", 1, 1, "Fault"),
        "primary": ("PrimaryTarget", "OpposingSide", "First", 1, 1, "Fault"),
        "random": ("Actor", "OpposingSide", "RngUniform", 1, 1, "Fault"),
        "all-opposing": ("Actor", "OpposingSide", "All", 1, 8, "Fault"),
        "all-allies": ("Actor", "SameSide", "All", 1, 8, "Fault"),
        "adjacent-allies": (
            "CurrentSubject",
            "SameSide",
            "PrimaryPlusAdjacent",
            0,
            2,
            "NoOp",
        ),
        "blast": (
            "PrimaryTarget",
            "OpposingSide",
            "PrimaryPlusAdjacent",
            1,
            3,
            "Fault",
        ),
    }
    for key, id_ in selectors.items():
        origin, side, choice, minimum, maximum, empty = selector_metadata[key]
        identities.append(
            ident(
                id_,
                f"selector.goal07.enemy-s12.{key}",
                "Selector",
                f"S12 {key} Selector",
                f"S12 {key}选择器",
                "Shared selector for the S12 ordinary-enemy batch.",
            )
        )
        row = selector(
            id_,
            origin,
            side,
            minimum=minimum,
            maximum=maximum,
            empty=empty,
            choice=choice,
        )
        if key == "random":
            row["rng_purpose_key"] = "behavior-choice"
        add("Selector", row)

    def expression(name: str, kind: str, node: str) -> int:
        nonlocal next_expression
        id_ = next_expression
        next_expression += 1
        add(
            "ValueExpression",
            {
                "id": id_,
                "stable_key": f"goal07.enemy.s12.expression.{name}",
                "result_kind": kind,
                "node": node,
            },
        )
        return id_

    actor_atk = expression(
        "actor-atk",
        "Scalar",
        json_cell(
            "QueryStat",
            subject_selector_id=selectors["actor"],
            stat="Atk",
            formula_purpose="OrdinaryDamage",
        ),
    )
    actor_maximum_hp = (
        expression(
            "actor-maximum-hp",
            "Scalar",
            json_cell(
                "QueryStat",
                subject_selector_id=selectors["actor"],
                stat="Hp",
                formula_purpose="Stat",
            ),
        )
        if PARTITION in {"G07-P5-M15-S13", "G07-P5-M15-S14"}
        else None
    )
    ratio_values = {
        value
        for spec in enemy_specs
        for _, _, _, _, value, _, _, _, _ in spec["abilities"]
        if value is not None
    } | {
        "0.25",
        "0.35",
        "0.5",
        "0.6",
        "0.7",
        "1",
    } | (
        {"0.025", "0.03", "0.075", "0.1", "1.2"}
        if PARTITION == "G07-P5-M15-S13"
        else {"0.4"}
        if PARTITION == "G07-P5-M15-S14"
        else set()
    )
    ratios = {
        value: expression(
            f"ratio-{value.replace('.', '-')}",
            "Scalar",
            json_cell("ScalarLiteral", value_decimal=value),
        )
        for value in sorted(ratio_values, key=float)
    }
    integer_two = expression(
        "integer-two", "Integer", json_cell("IntegerLiteral", value=2)
    )
    damage_amounts = {
        value: expression(
            f"atk-times-{value.replace('.', '-')}",
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedMultiply",
                left_expression_id=actor_atk,
                right_expression_id=ratios[value],
                rounding="NearestTiesAway",
            ),
        )
        for value in sorted(ratio_values, key=float)
    }
    maximum_hp_amounts = {
        value: expression(
            f"maximum-hp-times-{value.replace('.', '-')}",
            "Scalar",
            json_cell(
                "CheckedBinary",
                operator="CheckedMultiply",
                left_expression_id=actor_maximum_hp,
                right_expression_id=ratios[value],
                rounding="NearestTiesAway",
            ),
        )
        for value in (
            ("0.025", "0.03", "0.1", "1.2")
            if PARTITION == "G07-P5-M15-S13"
            else ("0.4",)
            if PARTITION == "G07-P5-M15-S14"
            else ()
        )
    }

    add(
        "ConditionExpression",
        {
            "id": condition_always,
            "stable_key": "goal07.enemy.s12.condition.always",
            "node": json_cell("Constant", value=True),
        },
    )

    effect_metadata = (
        {
            "channel": ("Channel State", "频道状态", "NeutralState", "NonDispellable", 1, integer_two, "OwnerTurnEnd", "Refresh", None, None),
            "die-debuff": ("Bad Roll", "坏点数", "Debuff", "DispellableDebuff", 1, integer_two, "TargetTurnEnd", "Refresh", None, None),
            "drink-shield": ("Free Drinks Shield", "赠饮护盾", "Buff", "DispellableBuff", 2, integer_two, "OwnerTurnEnd", "RefreshAndAddStacks", None, None),
            "soulglad": ("SoulGlad Party", "苏乐达派对", "Buff", "NonDispellable", 1, integer_two, "OwnerTurnEnd", "Refresh", None, None),
            "fountain": ("Celebration Fountain", "酬宾喷泉", "NeutralState", "NonDispellable", 1, integer_two, "OwnerTurnEnd", "Refresh", None, None),
            "alarm": ("Alarm Set", "闹钟预定", "Buff", "NonDispellable", 1, integer_two, "OwnerTurnEnd", "Refresh", None, None),
            "fog": ("Dense Fog", "霏雾", "Debuff", "DispellableDebuff", 1, integer_two, "TargetTurnEnd", "Refresh", None, None),
            "def-down": ("Defense Down", "防御降低", "Debuff", "DispellableDebuff", 1, integer_two, "TargetTurnEnd", "Refresh", None, None),
            "candle-flame": ("Candle Flame", "烛花", "NeutralState", "NonDispellable", 1, None, "Permanent", "Replace", None, None),
            "deep-freeze": ("Deep Freeze", "深度冻结", "Control", "CleanseableControl", 3, integer_two, "TargetTurnEnd", "RefreshAndAddStacks", None, None),
            "devour": ("Devour Otherling", "吞噬衍生物", "Buff", "DispellableBuff", 1, integer_two, "OwnerTurnEnd", "Refresh", None, None),
        }
        if PARTITION == "G07-P5-M15-S14"
        else
        {
            "chains": ("Chains of Destruction", "毁灭连锁", "NeutralState", "NonDispellable", 1, None, "Permanent", "Replace", None, None),
            "timed": ("Timed Module", "定时模块", "Buff", "NonDispellable", 1, integer_two, "OwnerTurnEnd", "Refresh", None, None),
            "cloud-edge": ("Cloud Edge", "云刃", "Buff", "DispellableBuff", 2, integer_two, "OwnerTurnEnd", "RefreshAndAddStacks", None, None),
            "binding": ("Binding of the Golden Age", "黄金年代的绸缪", "Buff", "NonDispellable", 1, integer_two, "OwnerTurnEnd", "Refresh", None, None),
            "seed-core": ("Seeding Core", "种丹", "Mark", "DispellableDebuff", 1, integer_two, "TargetTurnEnd", "Refresh", None, None),
            "draining": ("Vigor Draining", "采气", "Buff", "NonDispellable", 1, integer_two, "OwnerTurnEnd", "Refresh", None, None),
            "stove": ("Inferno Stove", "猛火炉灶", "Buff", "NonDispellable", 6, integer_two, "OwnerTurnEnd", "RefreshAndAddStacks", None, None),
            "bleed": ("Dreamjolt Bleed", "惊梦裂伤", "Dot", "DispellableDebuff", 1, integer_two, "TargetTurnStart", "Refresh", damage_amounts["0.5"], "Physical"),
            "sweetness": ("Libation of Sweetness", "泼洒甜蜜", "Buff", "DispellableBuff", 4, integer_two, "OwnerTurnEnd", "RefreshAndAddStacks", None, None),
            "rebirth": ("Rebirth", "复生", "Buff", "NonDispellable", 1, None, "Permanent", "Replace", None, None),
        }
        if PARTITION == "G07-P5-M15-S13"
        else {
            "rebound": ("Rebound Roar", "嗥吠", "NeutralState", "NonDispellable", 1, None, "Permanent", "Replace", None, None),
            "monitor": ("Monitor", "监视", "Mark", "DispellableDebuff", 1, integer_two, "TargetTurnEnd", "Refresh", None, None),
            "fiery": ("Fiery Majesty", "奋威", "Buff", "DispellableBuff", 3, integer_two, "OwnerTurnEnd", "RefreshAndAddStacks", None, None),
            "rally": ("Rallying Howl", "啸聚", "Buff", "DispellableBuff", 1, integer_two, "OwnerTurnEnd", "Refresh", None, None),
            "bleed": ("Machine Bleed", "机械裂伤", "Dot", "DispellableDebuff", 1, integer_two, "TargetTurnStart", "Refresh", damage_amounts["0.5"], "Physical"),
            "weaken": ("Weaken", "虚弱", "Debuff", "DispellableDebuff", 1, integer_two, "TargetTurnEnd", "Refresh", None, None),
            "sanction": ("Sanction Mode", "惩戒模式", "Buff", "NonDispellable", 1, integer_two, "OwnerTurnEnd", "Refresh", None, None),
            "sanction-bug": ("Sanction Mode: Punisher", "惩戒模式·惩罚者", "Buff", "NonDispellable", 1, integer_two, "OwnerTurnEnd", "Refresh", None, None),
            "imprisonment": ("Imprisonment", "禁锢", "Control", "CleanseableControl", 1, integer_two, "TargetTurnStart", "Refresh", None, None),
            "reverberation": ("Reverberation", "震荡", "Debuff", "DispellableDebuff", 1, integer_two, "TargetTurnEnd", "Refresh", None, None),
            "lock": ("Direwolf Lock", "齿狼锁定", "Mark", "DispellableDebuff", 1, integer_two, "TargetTurnEnd", "Refresh", None, None),
            "overcombust": ("Overcombust", "过载", "Buff", "DispellableBuff", 1, integer_two, "OwnerTurnEnd", "Refresh", None, None),
        }
    )
    for key, id_ in effects.items():
        (
            name_en,
            name_zh_cn,
            category,
            dispel,
            limit,
            duration,
            clock,
            policy,
            magnitude,
            dot_element,
        ) = effect_metadata[key]
        identities.append(
            ident(
                id_,
                f"effect.goal07.enemy-s12.{key}",
                "Effect",
                name_en,
                name_zh_cn,
                "Shared executable status for the S12 ordinary-enemy batch.",
            )
        )
        add(
            "Effect",
            {
                "id": id_,
                "category": category,
                "dispel_category": dispel,
                "stack_limit": limit,
                "duration_expression_id": duration,
                "duration_clock": clock,
                "tick_phase": "TurnStart" if category == "Dot" else "None",
                "stack_policy": policy,
                "magnitude_comparator_expression_id": magnitude,
                "dot_element": dot_element,
                "snapshot_policy": "OnApplication",
                "teardown_policy": "RemoveWithOwner",
                "application_priority": 0,
            },
        )
    effect_tags = (
        {
            "channel": ["channel-state"],
            "die-debuff": ["bad-roll"],
            "drink-shield": ["shield"],
            "soulglad": ["soulglad-party"],
            "fountain": ["celebration-fountain"],
            "alarm": ["alarm-set"],
            "fog": ["dense-fog"],
            "def-down": ["defense-down"],
            "candle-flame": ["candle-flame"],
            "deep-freeze": ["deep-freeze", "speed-down"],
            "devour": ["attack-up", "lock-on-target"],
        }
        if PARTITION == "G07-P5-M15-S14"
        else
        {
            "chains": ["chains-of-destruction"],
            "timed": ["timed-module"],
            "cloud-edge": ["damage-up"],
            "binding": ["golden-age-binding"],
            "seed-core": ["seeding-core"],
            "draining": ["vigor-draining"],
            "stove": ["inferno-stove"],
            "bleed": ["bleed"],
            "sweetness": ["sweetness"],
            "rebirth": ["rebirth"],
        }
        if PARTITION == "G07-P5-M15-S13"
        else {
            "rebound": ["rebound-roar"],
            "monitor": ["monitor"],
            "fiery": ["damage-up"],
            "rally": ["rallying-howl"],
            "bleed": ["bleed"],
            "weaken": ["weaken"],
            "sanction": ["sanction-mode", "toughness-protected"],
            "sanction-bug": ["sanction-mode", "toughness-protected"],
            "imprisonment": ["imprisonment", "blocks-normal-action"],
            "reverberation": ["reverberation"],
            "lock": ["lock-on-target"],
            "overcombust": ["overcombust"],
        }
    )
    for key, tags in effect_tags.items():
        for sequence, tag in enumerate(tags, start=1):
            add(
                "EffectTag",
                {"effect_id": effects[key], "sequence": sequence, "tag": tag},
            )

    def new_operation(
        name: str,
        payload: str,
        target: int | None = None,
        empty: str = "Fault",
    ) -> int:
        nonlocal next_operation
        id_ = next_operation
        next_operation += 1
        row = operation(id_, name, payload, target, empty)
        row["stable_key"] = f"goal07.enemy.s12.operation.{name}"
        add("Operation", row)
        return id_

    def op_step(id_: int) -> str:
        return json_cell("Operation", operation_id=id_)

    def make_program(name: str, steps: list[str]) -> int:
        nonlocal next_program
        id_ = next_program
        next_program += 1
        identities.append(
            ident(
                id_,
                f"program.goal07.enemy-s12.{name}",
                "Program",
                f"S12 {name} Program",
                f"S12 {name}程序",
                "Ordered Rule IR program for the S12 ordinary-enemy batch.",
            )
        )
        add("Program", {"id": id_, "domain": "Battle"})
        for sequence, step in enumerate(steps, start=1):
            add(
                "ProgramStep",
                {"program_id": id_, "sequence": sequence, "step": step},
            )
        return id_

    linked_units = [BASE + 351, BASE + 352]
    linked_ability = BASE + 381
    is_s13 = PARTITION == "G07-P5-M15-S13"
    is_s14 = PARTITION == "G07-P5-M15-S14"
    linked_slug = (
        "everwinter-shadewalker"
        if is_s14
        else "mara-struck-soldier"
        if is_s13
        else "illumination-dragonfish"
    )
    linked_name = (
        "Everwinter Shadewalker"
        if is_s14
        else "Mara-Struck Soldier"
        if is_s13
        else "Illumination Dragonfish"
    )
    linked_name_zh = (
        "永冬灾影" if is_s14 else "魔阴身士卒" if is_s13 else "入魔机巧·灯昼龙鱼"
    )
    linked_skill = (
        "Frost Crush" if is_s14 else "Thunder Strike" if is_s13 else "Candle Flame"
    )
    linked_skill_zh = "霜冻重击" if is_s14 else "雷霆突刺" if is_s13 else "烛焰"
    linked_element = "Ice" if is_s14 else "Lightning" if is_s13 else "Fire"
    linked_ability_key = (
        f"enemy.goal07.s12.{linked_slug}.ability.basic-attack"
        if is_s13 or is_s14
        else "enemy.goal07.s12.illumination-dragonfish.ability.candle-flame"
    )
    linked_program_slug = (
        f"{linked_slug}-basic-attack"
        if is_s13 or is_s14
        else "illumination-dragonfish-candle-flame"
    )
    linked_operation_slug = (
        f"{linked_slug}-basic-attack-damage"
        if is_s13 or is_s14
        else "illumination-dragonfish-candle-flame-damage"
    )
    linked_summary = (
        "Formation-linked Everwinter Shadewalker summoned by Frigid Prowler."
        if is_s14
        else "Formation-linked Mara-Struck Soldier summoned by Shape Shifter."
        if is_s13
        else "Formation-linked Gatekeeper sanction summon."
    )
    identities.append(
        ident(
            linked_ability,
            linked_ability_key,
            "Ability",
            linked_skill,
            linked_skill_zh,
            (
                f"Executable linked-unit attack used by {linked_name}."
                if is_s13 or is_s14
                else "Executable linked-unit attack used by Aurumaton Gatekeeper."
            ),
        )
    )
    linked_program = make_program(
        linked_program_slug,
        [
            op_step(
                new_operation(
                    linked_operation_slug,
                    json_cell(
                        "Damage",
                        amount_expression_id=damage_amounts["1"],
                        damage_class="Ordinary",
                        element=linked_element,
                        can_crit=True,
                    ),
                    selectors["primary"],
                )
            )
        ],
    )
    add(
        "Ability",
        {
            "id": linked_ability,
            "kind": "Summon",
            "target_pattern": "SingleTarget",
            "retarget_policy": "CancelRemaining",
            "level_cap": 1,
            "cooldown_actions": 1,
            "semantic_tags_mask": 5,
        },
    )
    add(
        "AbilityPhase",
        {
            "ability_id": linked_ability,
            "sequence": 1,
            "kind": "Resolved",
            "program_identity_id": linked_program,
        },
    )
    for index, unit_id in enumerate(linked_units, start=1):
        identities.append(
            ident(
                unit_id,
                (
                    f"unit.goal07.s12.{linked_slug}-{index}"
                    if is_s13 or is_s14
                    else (
                        "unit.goal07.s12.aurumaton-gatekeeper."
                        f"{linked_slug}-{index}"
                    )
                ),
                "CharacterForm",
                f"{linked_name} {index}",
                f"{linked_name_zh}{index}",
                linked_summary,
            )
        )
        add(
            "LinkedUnitDefinition",
            {
                "id": unit_id,
                "source_definition_identity_id": unit_id,
                "kind": "Summon",
                "presence": "Present",
                "ability_ids": str(linked_ability),
                "action_ability_id": linked_ability,
                "formation_index": 2 if index == 1 else 4,
                "initial_gauge_decimal": "10000",
                "hp_owner_ratio_decimal": "0.2",
                "hp_flat_decimal": "0",
                "atk_owner_ratio_decimal": "0.5",
                "atk_flat_decimal": "0",
                "def_owner_ratio_decimal": "1",
                "def_flat_decimal": "0",
                "spd_owner_ratio_decimal": "0",
                "spd_flat_decimal": "120",
                "owner_defeat_policy": "Depart",
                "owner_departure_policy": "Depart",
                "wave_policy": "Depart",
                "combatant_digest_sha256": sha256_text(
                    f"goal07-s12-{linked_slug}-{index}-v1"
                ),
            },
        )

    ability_ids: dict[tuple[int, str], int] = {}
    ability_programs: dict[int, int] = {}
    ability_targets: dict[int, int] = {}
    for enemy_index, spec in enumerate(enemy_specs):
        for ability_index, ability in enumerate(spec["abilities"], start=1):
            key, name_en, name_zh_cn, source_skill, ratio, element, target, active, mechanic = ability
            ability_id = BASE + 101 + enemy_index * 20 + ability_index
            ability_ids[(enemy_index, key)] = ability_id
            identities.append(
                ident(
                    ability_id,
                    f"{spec['template']}.ability.{key}",
                    "Ability",
                    name_en,
                    name_zh_cn,
                    f"Executable transcription of source skill {source_skill}.",
                )
            )
            target_selector = selectors[target]
            ability_targets[ability_id] = target_selector
            steps: list[str] = []
            if ratio is not None:
                steps.append(
                    op_step(
                        new_operation(
                            f"{enemy_index + 1}-{key}-damage",
                            json_cell(
                                "Damage",
                                amount_expression_id=damage_amounts[ratio],
                                damage_class="Ordinary",
                                element=element,
                                can_crit=True,
                            ),
                            target_selector,
                        )
                    )
                )
            if mechanic == "bounce-6":
                for hit in range(1, 7):
                    steps.append(
                        op_step(
                            new_operation(
                                f"{enemy_index + 1}-{key}-bounce-{hit}",
                                json_cell(
                                    "Damage",
                                    amount_expression_id=damage_amounts["1"],
                                    damage_class="Ordinary",
                                    element=element,
                                    can_crit=True,
                                ),
                                target_selector,
                            )
                        )
                    )
            effect_key = (
                {
                    "channel": "channel",
                    "die-debuff": "die-debuff",
                    "drink-shield": "drink-shield",
                    "soulglad": "soulglad",
                    "fountain": "fountain",
                    "alarm": "alarm",
                    "fog": "fog",
                    "def-down": "def-down",
                    "candle-flame": "candle-flame",
                    "deep-freeze": "deep-freeze",
                    "devour": "devour",
                }
                if is_s14
                else
                {
                    "chains": "chains",
                    "timed": "timed",
                    "cloud-edge": "cloud-edge",
                    "binding": "binding",
                    "seed-core": "seed-core",
                    "mara-summon": "draining",
                    "stove": "stove",
                    "bleed": "bleed",
                    "sweetness": "sweetness",
                }
                if is_s13
                else {
                    "monitor": "monitor",
                    "fiery": "fiery",
                    "rally": "rally",
                    "bleed": "bleed",
                    "weaken": "weaken",
                    "sanction": "sanction",
                    "sanction-bug": "sanction-bug",
                    "imprison-50": "imprisonment",
                    "imprison-60": "imprisonment",
                    "reverberation-70": "reverberation",
                    "reverberation-35": "reverberation",
                    "lock": "lock",
                    "overcombust": "overcombust",
                }
            ).get(mechanic)
            if effect_key is not None:
                chance = {
                    "imprison-50": "1",
                    "imprison-60": "1",
                    "reverberation-70": "0.7",
                    "reverberation-35": "0.35",
                }.get(mechanic)
                steps.append(
                    op_step(
                        new_operation(
                            f"{enemy_index + 1}-{key}-effect",
                            json_cell(
                                "ApplyEffect",
                                effect_id=effects[effect_key],
                                stacks_expression_id=None,
                                chance_policy=(
                                    "Resistible"
                                    if chance is not None
                                    else "Guaranteed"
                                ),
                                base_chance_expression_id=(
                                    ratios[chance] if chance is not None else None
                                ),
                                rng_purpose_key=(
                                    "effect-application"
                                    if chance is not None
                                    else None
                                ),
                            ),
                            (
                                selectors["actor"]
                                if mechanic
                                in {
                                    "chains",
                                    "timed",
                                    "cloud-edge",
                                    "binding",
                                    "mara-summon",
                                    "stove",
                                    "sweetness",
                                    "channel",
                                    "drink-shield",
                                    "soulglad",
                                    "fountain",
                                    "alarm",
                                    "candle-flame",
                                    "devour",
                                }
                                else target_selector
                            ),
                        )
                    )
                )
            if mechanic in {"imprison-50", "imprison-60", "delay-50"}:
                delay_ratio = "0.6" if mechanic == "imprison-60" else "0.5"
                steps.append(
                    op_step(
                        new_operation(
                            f"{enemy_index + 1}-{key}-delay",
                            json_cell(
                                "DelayAction",
                                amount_expression_id=ratios[delay_ratio],
                            ),
                            target_selector,
                        )
                    )
                )
            if mechanic == "heal":
                steps.append(
                    op_step(
                        new_operation(
                            f"{enemy_index + 1}-{key}-heal",
                            json_cell(
                                "Heal",
                                amount_expression_id=damage_amounts["0.25"],
                            ),
                            target_selector,
                        )
                    )
                )
            if mechanic in {"drain-025", "drain-03"}:
                heal_ratio = "0.025" if mechanic == "drain-025" else "0.03"
                steps.append(
                    op_step(
                        new_operation(
                            f"{enemy_index + 1}-{key}-vigor-drain",
                            json_cell(
                                "Heal",
                                amount_expression_id=maximum_hp_amounts[heal_ratio],
                            ),
                            selectors["actor"],
                        )
                    )
                )
            if mechanic == "mara-summon":
                steps.append(
                    op_step(
                        new_operation(
                            f"{enemy_index + 1}-{key}-hp-cost",
                            json_cell(
                                "Damage",
                                amount_expression_id=maximum_hp_amounts["0.1"],
                                damage_class="Ordinary",
                                element="Physical",
                                can_crit=False,
                            ),
                            selectors["actor"],
                        )
                    )
                )
            if mechanic in {"sanction", "mara-summon", "otherling"}:
                for unit_id in linked_units:
                    steps.append(
                        op_step(
                            new_operation(
                                f"{enemy_index + 1}-{key}-summon-{unit_id}",
                                json_cell(
                                    "Summon",
                                    unit_definition_identity_id=unit_id,
                                    owner_selector_id=selectors["actor"],
                                ),
                            )
                        )
                    )
            program = make_program(f"{enemy_index + 1}-{key}", steps)
            ability_programs[ability_id] = program
            target_pattern = (
                "Aoe"
                if target in {"all-opposing", "all-allies"}
                else "Blast"
                if target == "blast"
                else "None"
                if target == "actor"
                else "SingleTarget"
            )
            add(
                "Ability",
                {
                    "id": ability_id,
                    "kind": "Passive" if not active else "Skill",
                    "target_pattern": target_pattern,
                    "retarget_policy": "CancelRemaining",
                    "level_cap": 1,
                    "cooldown_actions": 1,
                    "semantic_tags_mask": 5 if ratio is not None else 4,
                },
            )
            add(
                "AbilityPhase",
                {
                    "ability_id": ability_id,
                    "sequence": 1,
                    "kind": "Resolved",
                    "program_identity_id": program,
                },
            )
            add(
                "EnemyAbility",
                {
                    "id": ability_id,
                    "telegraph": "None",
                    "cooldown_actions": 1,
                    "initial_cooldown_actions": 0,
                    "charge_actions": 0,
                    "ai_tag": f"s12-{enemy_index + 1}-{key}",
                },
            )

    defeat_effect = (
        effects["candle-flame"]
        if is_s14
        else effects["chains"]
        if is_s13
        else effects["rebound"]
    )
    defeat_slug = (
        "illumination-dragonfish-candle-flame"
        if is_s14
        else "automaton-spider-chains"
        if is_s13
        else "golden-hound-rebound"
    )
    defeat_name = (
        "Illumination Dragonfish Candle Flame Rule"
        if is_s14
        else "Automaton Spider Chains Rule"
        if is_s13
        else "Golden Hound Rebound Rule"
    )
    defeat_name_zh = (
        "灯昼龙鱼烛花规则"
        if is_s14
        else "自动机兵蜘蛛毁灭连锁规则"
        if is_s13
        else "娄金嗥吠规则"
    )
    defeat_summary = (
        "Deals 40% of the Illumination Dragonfish's maximum HP as Fire "
        "damage to all same-side units after it is defeated."
        if is_s14
        else "Deals 120% of the Automaton Spider's maximum HP as Fire damage "
        "to adjacent same-side units after it is defeated."
        if is_s13
        else "Advances adjacent living allies after the Golden Hound is defeated."
    )
    rebound_entry = make_program(
        f"{defeat_slug}-entry",
        [
            op_step(
                new_operation(
                    f"{defeat_slug}-entry",
                    json_cell(
                        "ApplyEffect",
                        effect_id=defeat_effect,
                        stacks_expression_id=None,
                        chance_policy="Guaranteed",
                        base_chance_expression_id=None,
                        rng_purpose_key=None,
                    ),
                    selectors["actor"],
                )
            )
        ],
    )
    rebound_program = make_program(
        f"{defeat_slug}-trigger",
        [
            op_step(
                new_operation(
                    (
                        f"{defeat_slug}-adjacent-allies"
                        if is_s13 or is_s14
                        else "golden-hound-rebound-advance-adjacent-allies"
                    ),
                    (
                        json_cell(
                            "Damage",
                            amount_expression_id=maximum_hp_amounts[
                                "0.4" if is_s14 else "1.2"
                            ],
                            damage_class="Ordinary",
                            element="Fire",
                            can_crit=False,
                        )
                        if is_s13 or is_s14
                        else json_cell(
                            "AdvanceAction",
                            amount_expression_id=ratios["1"],
                        )
                    ),
                    selectors["all-allies"] if is_s14 else selectors["adjacent-allies"],
                    "NoOp",
                )
            )
        ],
    )
    rebound_rule = BASE + 5_101
    rebound_filter = BASE + 5_102
    rebound_trigger = BASE + 5_103
    identities.append(
        ident(
            rebound_rule,
            f"rule.goal07.enemy-s12.{defeat_slug}",
            "Rule",
            defeat_name,
            defeat_name_zh,
            defeat_summary,
        )
    )
    add(
        "RuleDefinition",
        {
            "id": rebound_rule,
            "domain": "Battle",
            "source_definition_identity_id": defeat_effect,
            "source_class": "Effect",
            "source_digest_sha256": sha256_text(
                f"goal07-s12-{defeat_slug}-v1"
            ),
        },
    )
    add(
        "EventFilter",
        {
            "id": rebound_filter,
            "stable_key": (
                f"goal07.enemy.s12.filter.{defeat_slug}-defeated"
                if is_s13 or is_s14
                else "goal07.enemy.s12.filter.golden-hound-defeated"
            ),
            "target_selector_id": selectors["owner"],
            "cause_ancestry": "Any",
        },
    )
    add(
        "RuleTrigger",
        {
            "id": rebound_trigger,
            "stable_key": f"goal07.enemy.s12.trigger.{defeat_slug}",
            "rule_id": rebound_rule,
            "sequence": 1,
            "event": json_cell("Unit", point="Defeated"),
            "phase": "AfterDefeatSettlement",
            "filter_id": rebound_filter,
            "condition_id": condition_always,
            "once_scope": "Battle",
            "priority": 0,
            "program_id": rebound_program,
        },
    )
    add(
        "EffectRuleBinding",
        {
            "effect_id": defeat_effect,
            "sequence": 1,
            "rule_id": rebound_rule,
        },
    )

    anchor_by_variant = {
        item["enemy_variant_id"]: item for item in anchor["variants"]
    }
    for enemy_index, spec in enumerate(enemy_specs):
        variant = variant_row_ids[enemy_index]
        template = template_row_ids[enemy_index]
        graph = BASE + 41 + enemy_index
        identities.extend(
            [
                ident(
                    variant,
                    spec["key"],
                    "EnemyVariant",
                    spec["name"],
                    spec["zh"],
                    "Production materialization variant for frozen Standard Universe bindings.",
                    f"1|{SOURCE_RECORD_ID}",
                ),
                ident(
                    template,
                    spec["template"],
                    "Enemy",
                    f"{spec['name']} Template",
                    f"{spec['zh']}模板",
                    "Version 4.4 source-bound enemy template.",
                ),
                ident(
                    graph,
                    f"ai.goal07.enemy-s12.{enemy_index + 1}",
                    "AiGraph",
                    f"{spec['name']} AI",
                    f"{spec['zh']}AI",
                    "Finite source-ordered ordinary-enemy action graph.",
                ),
            ]
        )
        active_ability_ids = [
            ability_ids[(enemy_index, key)] for key in spec["cycle"]
        ]
        state_ids = list(range(next_state, next_state + len(active_ability_ids)))
        next_state += len(active_ability_ids)
        add(
            "AiGraph",
            {
                "id": graph,
                "initial_state_id": state_ids[0],
                "automatic_transition_budget": 8,
            },
        )
        for offset, (state_id, ability_id) in enumerate(
            zip(state_ids, active_ability_ids)
        ):
            add(
                "AiState",
                {
                    "id": state_id,
                    "stable_key": (
                        f"goal07.enemy.s12.ai.{enemy_index + 1}.step-{offset + 1}"
                    ),
                    "graph_id": graph,
                    "mandatory_fallback_ability_id": active_ability_ids[0],
                    "turn_counter_reset": offset == 0,
                },
            )
            add(
                "AiCandidate",
                {
                    "id": next_candidate,
                    "stable_key": (
                        f"goal07.enemy.s12.ai.{enemy_index + 1}."
                        f"candidate-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "ability_id": ability_id,
                    "condition_id": condition_always,
                    "target_selector_id": ability_targets[ability_id],
                    "priority": 0,
                    "selection": "FirstLegal",
                    "no_target_fallback": "UseFallbackAbility",
                    "fallback_ability_id": active_ability_ids[0],
                },
            )
            next_candidate += 1
            add(
                "AiTransition",
                {
                    "id": next_transition,
                    "stable_key": (
                        f"goal07.enemy.s12.ai.{enemy_index + 1}."
                        f"transition-{offset + 1}"
                    ),
                    "state_id": state_id,
                    "sequence": 1,
                    "target_state_id": state_ids[
                        (offset + 1) % len(state_ids)
                    ],
                    "condition_id": condition_always,
                    "priority": 0,
                    "timing": "AfterAction",
                },
            )
            next_transition += 1

        add(
            "EnemyTemplate",
            {
                "id": template,
                "rank": spec["rank"],
                "base_aggro_decimal": "100",
                "default_ai_graph_id": graph,
            },
        )
        add(
            "EnemyVariant",
            {
                "id": variant,
                "template_id": template,
                "ai_graph_id": graph,
                "mechanically_distinct_key": spec["key"],
            },
        )
        for level in anchor_by_variant[spec["key"]]["levels"]:
            add(
                "EnemyStat",
                {
                    "variant_id": variant,
                    "level": level["authored_level"],
                    "difficulty_key": "standard-universe-v1",
                    "hp_decimal": level["base_hp"],
                    "atk_decimal": level["base_atk"],
                    "def_decimal": level["base_def"],
                    "spd_decimal": level["base_spd"],
                    "effect_hit_rate_decimal": level["effect_hit_rate"],
                    "effect_resistance_decimal": level["effect_resistance"],
                    "crit_damage_decimal": "0.2",
                },
            )
        for sequence, weakness in enumerate(spec["weaknesses"], start=1):
            add(
                "EnemyWeakness",
                {
                    "variant_id": variant,
                    "sequence": sequence,
                    "element": weakness,
                },
            )
        for element, value in spec["resistances"]:
            add(
                "EnemyResistance",
                {
                    "variant_id": variant,
                    "element": element,
                    "value_decimal": value,
                },
            )
        add(
            "EnemyToughnessLayer",
            {
                "variant_id": variant,
                "sequence": 1,
                "layer_key": "ordinary",
                "kind": "Ordinary",
                "maximum_decimal": spec["toughness"],
                "recovery_ratio_decimal": "1",
                "active_at_start": True,
            },
        )
        for sequence, ability in enumerate(spec["abilities"], start=1):
            add(
                "EnemyVariantAbility",
                {
                    "variant_id": variant,
                    "sequence": sequence,
                    "ability_id": ability_ids[(enemy_index, ability[0])],
                },
            )
        add(
            "EnemyPhase",
            {
                "id": BASE + 4_001 + enemy_index,
                "stable_key": f"goal07.enemy.s12.phase-{enemy_index + 1}",
                "variant_id": variant,
                "sequence": 1,
                "entry_condition_id": condition_always,
                "exit_condition_id": condition_always,
                "replacement_priority": 1,
                "ai_graph_id": graph,
                "entry_program_id": (
                    rebound_entry
                    if enemy_index == (6 if is_s14 else 0)
                    else None
                ),
                "targetable": True,
                "transition_model": "TransformSameUnit",
                "hp_carry": "Reset",
                "action_gauge_carry": "Reset",
                "effect_carry": "Clear",
                "toughness_carry": "Reset",
                "summon_carry": "Clear",
            },
        )

    anchor_digest = sha256_bytes(anchor_path(PARTITION).read_bytes())
    add(
        "SourceRecord",
        {
            "id": SOURCE_RECORD_ID,
            "stable_key": "source.goal07.enemy-s12.2026-07-29",
            "category": "CommunityMaintained",
            "publisher": anchor["source"]["publisher"],
            "url": anchor["source"]["url"],
            "accessed_on": anchor["source"]["accessed_on"],
            "applicable_game_version": anchor["source"]["game_version"],
            "confidence": "SecondaryVersionSensitiveCrossCheck",
            "evidence_sha256": anchor_digest,
            "usage_note": (
                "Per-variant exact or approved derived level values and "
                "version 4.4 mechanics are committed as Goal 07 evidence."
            ),
        },
    )
    add(
        "EvidenceRecord",
        {
            "id": EVIDENCE_RECORD_ID,
            "stable_key": "evidence.goal07.enemy.s12.numeric-anchors",
            "kind": "SourcePayload",
            "source_record_id": SOURCE_RECORD_ID,
            "sha256": anchor_digest,
            "note": "Committed per-variant numeric anchors for Goal 07 S12.",
        },
    )
    for item in identities:
        add("ContentIdentity", item)
        add(
            "ContentEvidenceBinding",
            {
                "content_id": item["id"],
                "sequence": 1,
                "fact_key": f"goal07.s12.executable:{item['stable_key']}",
                "source_record_id": 1,
                "evidence_record_id": 3,
                "quality": "ExactStructured",
                "mechanism_quality": "ExactStructured",
            },
        )
    for enemy_index, spec in enumerate(enemy_specs):
        add(
            "ContentEvidenceBinding",
            {
                "content_id": variant_row_ids[enemy_index],
                "sequence": 2,
                "fact_key": f"goal07.s12.public-level-stats:{spec['key']}",
                "source_record_id": SOURCE_RECORD_ID,
                "evidence_record_id": EVIDENCE_RECORD_ID,
                "quality": (
                    "ExactStructured"
                    if anchor_by_variant[spec["key"]]["accuracy_disposition"]
                    == "ExactPublic"
                    else "ApproximateFromReleasedText"
                ),
                "mechanism_quality": "ExactStructured",
            },
        )
    if is_s13 or is_s14:
        batch = "s14" if is_s14 else "s13"
        batch_upper = batch.upper()
        for table_rows in rows.values():
            for row in table_rows:
                for field, value in row.items():
                    if isinstance(value, str):
                        row[field] = value.replace("s12", batch).replace(
                            "S12", batch_upper
                        )
    for table_rows in rows.values():
        table_rows.sort(
            key=lambda row: json.dumps(
                row,
                ensure_ascii=False,
                sort_keys=True,
                default=str,
            )
        )
    return rows


def owns_partition_range(value: Any) -> bool:
    return BASE <= int(value) < BASE + 10_000


def owns_partition_identity(value: Any) -> bool:
    id_ = int(value)
    return owns_partition_range(id_) or (
        PARTITION == "G07-P5-M15-S12" and id_ in {95, 96, 10_001, 10_002}
    ) or (
        PARTITION == "G07-P5-M15-S13"
        and id_ in {97, 99, 100, 10_003, 10_005, 10_006}
    ) or (
        PARTITION == "G07-P5-M15-S14"
        and id_ in {102, 103, 104, 10_008, 10_009, 10_010}
    )


def owns_partition_variant(value: Any) -> bool:
    id_ = int(value)
    return owns_partition_range(id_) or (
        PARTITION == "G07-P5-M15-S12" and id_ in {95, 96}
    ) or (
        PARTITION == "G07-P5-M15-S13" and id_ in {97, 99, 100}
    ) or (
        PARTITION == "G07-P5-M15-S14" and id_ in {102, 103, 104}
    )


def owns_partition_template(value: Any) -> bool:
    id_ = int(value)
    return owns_partition_range(id_) or (
        PARTITION == "G07-P5-M15-S12" and id_ in {10_001, 10_002}
    ) or (
        PARTITION == "G07-P5-M15-S13" and id_ in {10_003, 10_005, 10_006}
    ) or (
        PARTITION == "G07-P5-M15-S14" and id_ in {10_008, 10_009, 10_010}
    )


OWNERSHIP: dict[str, Callable[[dict[str, Any]], bool]] = {
    "Ability": lambda row: owns_partition_range(row["id"]),
    "AbilityPhase": lambda row: owns_partition_range(row["ability_id"]),
    "AiCandidate": lambda row: owns_partition_range(row["id"]),
    "AiGraph": lambda row: owns_partition_range(row["id"]),
    "AiState": lambda row: owns_partition_range(row["id"]),
    "AiTransition": lambda row: owns_partition_range(row["id"]),
    "ConditionExpression": lambda row: owns_partition_range(row["id"]),
    "ContentEvidenceBinding": lambda row: owns_partition_identity(row["content_id"]),
    "ContentIdentity": lambda row: owns_partition_identity(row["id"]),
    "Effect": lambda row: owns_partition_range(row["id"]),
    "EffectModifierBinding": lambda row: owns_partition_range(row["effect_id"]),
    "EffectRuleBinding": lambda row: owns_partition_range(row["effect_id"]),
    "EffectTag": lambda row: owns_partition_range(row["effect_id"]),
    "EnemyAbility": lambda row: owns_partition_range(row["id"]),
    "EnemyDebuffResistance": lambda row: owns_partition_variant(row["variant_id"]),
    "EnemyPhase": lambda row: owns_partition_range(row["id"]),
    "EnemyResistance": lambda row: owns_partition_variant(row["variant_id"]),
    "EnemyStat": lambda row: owns_partition_variant(row["variant_id"]),
    "EnemyTemplate": lambda row: owns_partition_template(row["id"]),
    "EnemyToughnessLayer": lambda row: owns_partition_variant(row["variant_id"]),
    "EnemyVariant": lambda row: owns_partition_variant(row["id"]),
    "EnemyVariantAbility": lambda row: owns_partition_variant(row["variant_id"]),
    "EnemyWeakness": lambda row: owns_partition_variant(row["variant_id"]),
    "EventFilter": lambda row: owns_partition_range(row["id"]),
    "LinkedUnitDefinition": lambda row: owns_partition_range(row["id"]),
    "ModifierDefinition": lambda row: owns_partition_range(row["id"]),
    "ModifierStackingGroup": lambda row: owns_partition_range(row["id"]),
    "Operation": lambda row: owns_partition_range(row["id"]),
    "Program": lambda row: owns_partition_range(row["id"]),
    "ProgramStep": lambda row: owns_partition_range(row["program_id"]),
    "RuleDefinition": lambda row: owns_partition_range(row["id"]),
    "RuleTrigger": lambda row: owns_partition_range(row["id"]),
    "Selector": lambda row: owns_partition_range(row["id"]),
    "SelectorPredicate": lambda row: owns_partition_range(row["selector_id"]),
    "ValueExpression": lambda row: owns_partition_range(row["id"]),
    "SourceRecord": lambda row: int(row["id"]) == SOURCE_RECORD_ID,
    "EvidenceRecord": lambda row: int(row["id"]) == EVIDENCE_RECORD_ID,
}


def selected(table: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    owns = OWNERSHIP[table]
    return [row for row in rows if owns(row)]


def merged(table: str, expected: list[dict[str, Any]]) -> list[dict[str, Any]]:
    owns = OWNERSHIP[table]
    current = [row for row in workbook_rows(table) if not owns(row)]
    current.extend(expected)
    return current


def build_golden(expected: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    tables: dict[str, Any] = {}
    for table, rows in sorted(expected.items()):
        excel = selected(table, workbook_rows(table))
        if semantic_digest(excel) != semantic_digest(rows):
            raise ValueError(f"{table}: {PARTITION} Excel rows differ")
        exported = selected(table, sora_rows(table))
        if semantic_digest(exported) != semantic_digest(rows):
            raise ValueError(f"{table}: {PARTITION} Sora rows differ")
        tables[table] = {
            "rows": len(rows),
            "semantic_sha256": semantic_digest(rows),
        }
    return {
        "schema_revision": "starclock.goal07-enemy-partition-golden.v1",
        "partition_id": PARTITION,
        "enemy_variant_ids": VARIANT_KEYS,
        "tables": tables,
    }


def main() -> None:
    global BASE, EVIDENCE_RECORD_ID, PARTITION, SOURCE_RECORD_ID, VARIANT_KEY, VARIANT_KEYS
    parser = argparse.ArgumentParser()
    parser.add_argument("--partition", required=True)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--write", action="store_true")
    mode.add_argument("--write-golden", action="store_true")
    mode.add_argument("--check", action="store_true")
    args = parser.parse_args()
    if args.partition not in PARTITION_CONFIG:
        raise ValueError(f"{args.partition}: enemy partition authoring is not implemented")
    PARTITION = args.partition
    partition_config = PARTITION_CONFIG[PARTITION]
    BASE = partition_config["base"]
    VARIANT_KEY = partition_config["variant"]
    VARIANT_KEYS = partition_config.get("variants", [VARIANT_KEY])
    SOURCE_RECORD_ID = partition_config["source_record_id"]
    EVIDENCE_RECORD_ID = partition_config["evidence_record_id"]
    expected = {
        "G07-P5-M15-S01": owned_rows_s01,
        "G07-P5-M15-S02": owned_rows_s02,
        "G07-P5-M15-S03": owned_rows_s03,
        "G07-P5-M15-S04": owned_rows_s04,
        "G07-P5-M15-S05": owned_rows_s05,
        "G07-P5-M15-S06": owned_rows_s06,
        "G07-P5-M15-S07": owned_rows_s07,
        "G07-P5-M15-S08": owned_rows_s08,
        "G07-P5-M15-S09": owned_rows_s09,
        "G07-P5-M15-S10": owned_rows_s10,
        "G07-P5-M15-S11": owned_rows_s11,
        "G07-P5-M15-S12": owned_rows_ordinary,
        "G07-P5-M15-S13": owned_rows_ordinary,
        "G07-P5-M15-S14": owned_rows_ordinary,
    }[PARTITION]()
    golden_path = (
        ROOT
        / "evidence"
        / "standard-universe-mechanics-complete-v1"
        / "goldens"
        / f"{PARTITION}.json"
    )
    if args.write:
        for table, rows in expected.items():
            write_rows(table, merged(table, rows))
        print(f"Authored Goal 07 enemy partition {PARTITION}.")
        return
    golden = build_golden(expected)
    encoded = json.dumps(golden, ensure_ascii=False, indent=2) + "\n"
    if args.write_golden:
        golden_path.parent.mkdir(parents=True, exist_ok=True)
        golden_path.write_text(encoded, encoding="utf-8", newline="\n")
        print(f"Wrote {golden_path.relative_to(ROOT)}.")
        return
    if not golden_path.is_file() or golden_path.read_text(encoding="utf-8") != encoded:
        raise ValueError(f"{PARTITION}: enemy partition golden drifted")
    print(f"Goal 07 enemy partition {PARTITION} matches Excel and Sora.")


if __name__ == "__main__":
    main()
