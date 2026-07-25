"""Verify one Goal 07 Ability Tree partition in authoritative Excel and Sora.

The complete Universe workbooks remain reproducibly authored by
tools/universe-reference/author_workbooks.py. This focused openpyxl gate proves
that one frozen Goal 07 partition owns complete workbook rows and that Sora's
committed debug output preserves their identities and relationships.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA


ROOT = Path(__file__).resolve().parents[2]
PARTITION_MANIFEST = (
    ROOT
    / "content-manifests"
    / "standard-universe-mechanics-complete-v1"
    / "content-partitions.json"
)
DATA = ROOT / "config" / "data"
DEBUG = ROOT / "config" / "universe-generated" / "debug-json"
BUNDLE = ROOT / "config" / "universe-generated" / "config.sora"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def fields(sheet: Any) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }


def rows(sheet: Any) -> list[dict[str, Any]]:
    columns = fields(sheet)
    result: list[dict[str, Any]] = []
    for cells in sheet.iter_rows(min_row=8):
        values = {
            field: cells[column - 1].value
            for field, column in columns.items()
        }
        if all(value is None for value in values.values()):
            continue
        for column in columns.values():
            cell = cells[column - 1]
            if cell.data_type in (TYPE_ERROR, TYPE_FORMULA):
                raise ValueError(f"{sheet.title}/{cell.coordinate}: formulas and errors are forbidden")
        result.append(values)
    return result


def keyed(items: list[dict[str, Any]], key: str) -> dict[Any, dict[str, Any]]:
    result: dict[Any, dict[str, Any]] = {}
    for item in items:
        value = item[key]
        if value in result:
            raise ValueError(f"duplicate {key}={value}")
        result[value] = item
    return result


def split_ids(value: Any) -> list[int]:
    if value in (None, ""):
        return []
    if isinstance(value, (int, float)):
        return [int(value)]
    return [int(part) for part in str(value).split("|") if part]


def unwrap(value: Any) -> Any:
    if value == "Null":
        return None
    if isinstance(value, list):
        return [unwrap(item) for item in value]
    if isinstance(value, dict):
        if len(value) == 1:
            tag, inner = next(iter(value.items()))
            if tag in {"String", "Integer", "Bool", "Decimal"}:
                return inner
            if tag == "List":
                return [unwrap(item) for item in inner]
        return {key: unwrap(inner) for key, inner in value.items()}
    return value


def sora_rows(table: str) -> list[dict[str, Any]]:
    payload = json.loads((DEBUG / f"{table}.json").read_text(encoding="utf-8"))
    return [
        {key: unwrap(value) for key, value in row["values"].items()}
        for row in payload["table"]["rows"]
    ]


def canonical_digest(items: list[dict[str, Any]]) -> str:
    encoded = json.dumps(
        items,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def partition(partition_id: str) -> dict[str, Any]:
    manifest = json.loads(PARTITION_MANIFEST.read_text(encoding="utf-8"))
    match = next(
        (item for item in manifest["partitions"] if item["id"] == partition_id),
        None,
    )
    if match is None or match["mechanic_family"] != "shared-activity-and-ability-tree":
        raise ValueError(f"{partition_id}: not an Ability Tree partition")
    return match


def build_golden(partition_id: str) -> dict[str, Any]:
    assigned = partition(partition_id)
    workbook = load_workbook(DATA / "Universe.xlsx", read_only=True, data_only=False)
    bindings = load_workbook(
        DATA / "UniverseBindings.xlsx", read_only=True, data_only=False
    )
    evidence = load_workbook(
        DATA / "UniverseEvidence.xlsx", read_only=True, data_only=False
    )

    nodes = keyed(rows(workbook["UniverseAbilityTreeNode"]), "stable_key")
    selected_nodes = [nodes[stable_key] for stable_key in assigned["record_ids"]]
    node_ids = {int(node["id"]) for node in selected_nodes}
    effects = [
        row
        for row in rows(workbook["UniverseAbilityTreeEffect"])
        if int(row["node_id"]) in node_ids
    ]
    costs = [
        row
        for row in rows(workbook["UniverseAbilityTreeCost"])
        if int(row["node_id"]) in node_ids
    ]
    parameters = [
        row
        for row in rows(workbook["UniverseAbilityTreeParameter"])
        if int(row["node_id"]) in node_ids
    ]
    edges = [
        row
        for row in rows(workbook["UniverseAbilityTreeEdge"])
        if int(row["node_id"]) in node_ids
    ]
    effect_node_ids = {int(effect["node_id"]) for effect in effects}
    if effect_node_ids != node_ids:
        raise ValueError(f"{partition_id}: one or more assigned nodes have no typed effect")

    rules = keyed(rows(bindings["UniverseMechanicRule"]), "stable_key")
    selected_rules = [rules[stable_key] for stable_key in assigned["rule_ids"]]
    for node in selected_nodes:
        if node["rule_stable_key"] not in assigned["rule_ids"]:
            raise ValueError(f"{node['stable_key']}: assigned rule link differs")
    for rule in selected_rules:
        if rule["source_record_stable_key"] not in assigned["record_ids"]:
            raise ValueError(f"{rule['stable_key']}: assigned source link differs")

    fixtures = keyed(rows(evidence["UniverseReviewFixture"]), "stable_key")
    selected_fixtures = [fixtures[stable_key] for stable_key in assigned["fixture_ids"]]
    audits = keyed(rows(evidence["UniverseContentAudit"]), "content_stable_key")
    sources = keyed(rows(evidence["UniverseSourceRecord"]), "id")
    for stable_key in assigned["record_ids"]:
        audit = audits.get(stable_key)
        if audit is None or not audit["enabled"]:
            raise ValueError(f"{stable_key}: enabled content audit is missing")
        provenance = split_ids(audit["provenance_ids"])
        if not provenance or any(source not in sources for source in provenance):
            raise ValueError(f"{stable_key}: provenance does not resolve")
    if any(not rule["source_file"] for rule in selected_rules):
        raise ValueError(f"{partition_id}: one or more rules lack source-file provenance")
    for fixture in selected_fixtures:
        provenance = split_ids(fixture["provenance_ids"])
        if not provenance or any(source not in sources for source in provenance):
            raise ValueError(f"{fixture['stable_key']}: provenance does not resolve")

    sora_nodes = keyed(sora_rows("UniverseAbilityTreeNode"), "stable_key")
    sora_rules = keyed(sora_rows("UniverseMechanicRule"), "stable_key")
    sora_fixtures = keyed(sora_rows("UniverseReviewFixture"), "stable_key")
    for node in selected_nodes:
        exported = sora_nodes.get(node["stable_key"])
        if exported is None or int(exported["id"]) != int(node["id"]):
            raise ValueError(f"{node['stable_key']}: Sora node differs")
        if exported["rule_stable_key"] != node["rule_stable_key"]:
            raise ValueError(f"{node['stable_key']}: Sora rule reference differs")
    for rule in selected_rules:
        exported = sora_rules.get(rule["stable_key"])
        if exported is None or exported["source_record_stable_key"] != rule[
            "source_record_stable_key"
        ]:
            raise ValueError(f"{rule['stable_key']}: Sora rule differs")
    for fixture in selected_fixtures:
        if fixture["stable_key"] not in sora_fixtures:
            raise ValueError(f"{fixture['stable_key']}: Sora fixture is missing")

    table_rows = {
        "UniverseAbilityTreeNode": selected_nodes,
        "UniverseAbilityTreeEdge": edges,
        "UniverseAbilityTreeCost": costs,
        "UniverseAbilityTreeEffect": effects,
        "UniverseAbilityTreeParameter": parameters,
        "UniverseMechanicRule": selected_rules,
        "UniverseReviewFixture": selected_fixtures,
    }
    return {
        "schema_revision": "starclock.goal07-ability-tree-partition-golden.v1",
        "partition_id": partition_id,
        "record_ids": assigned["record_ids"],
        "rule_ids": assigned["rule_ids"],
        "fixture_ids": assigned["fixture_ids"],
        "workbooks": {
            name: sha256(DATA / name)
            for name in (
                "Universe.xlsx",
                "UniverseBindings.xlsx",
                "UniverseEvidence.xlsx",
            )
        },
        "sora_bundle_sha256": sha256(BUNDLE),
        "tables": {
            name: {
                "rows": len(table_rows),
                "semantic_sha256": canonical_digest(table_rows),
            }
            for name, table_rows in table_rows.items()
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--partition", required=True)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--check", action="store_true")
    mode.add_argument("--write-golden", type=Path)
    args = parser.parse_args()

    golden = build_golden(args.partition)
    if args.write_golden is not None:
        target = args.write_golden
        if not target.is_absolute():
            target = ROOT / target
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(
            json.dumps(golden, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        print(f"Wrote {args.partition} openpyxl/Sora golden to {target}.")
        return

    target = (
        ROOT
        / "evidence"
        / "standard-universe-mechanics-complete-v1"
        / "goldens"
        / f"{args.partition}.json"
    )
    if not target.is_file():
        raise ValueError(f"{args.partition}: committed golden is missing")
    expected = json.loads(target.read_text(encoding="utf-8"))
    if expected != golden:
        raise ValueError(f"{args.partition}: openpyxl/Sora golden drifted")
    print(f"{args.partition} authoritative Excel and Sora rows verified.")


if __name__ == "__main__":
    main()
