"""Verify one Goal 07 Curio partition in authoritative Excel and Sora."""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA


ROOT = Path(__file__).resolve().parents[2]
MANIFEST = (
    ROOT
    / "content-manifests"
    / "standard-universe-mechanics-complete-v1"
    / "content-partitions.json"
)
DATA = ROOT / "config" / "data"
DEBUG = ROOT / "config" / "universe-generated" / "debug-json"
BUNDLE = ROOT / "config" / "universe-generated" / "config.sora"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def sheet_rows(sheet: Any) -> list[dict[str, Any]]:
    fields = {
        str(cell.value): cell.column
        for cell in sheet[3]
        if cell.value not in (None, "#field")
    }
    result: list[dict[str, Any]] = []
    for cells in sheet.iter_rows(min_row=8):
        row = {field: cells[column - 1].value for field, column in fields.items()}
        if all(value is None for value in row.values()):
            continue
        for column in fields.values():
            cell = cells[column - 1]
            if cell.data_type in (TYPE_ERROR, TYPE_FORMULA):
                raise ValueError(
                    f"{sheet.title}/{cell.coordinate}: formulas and errors are forbidden"
                )
        result.append(row)
    return result


def keyed(rows: list[dict[str, Any]], key: str) -> dict[Any, dict[str, Any]]:
    result: dict[Any, dict[str, Any]] = {}
    for row in rows:
        value = row[key]
        if value in result:
            raise ValueError(f"duplicate {key}={value}")
        result[value] = row
    return result


def split_ids(value: Any) -> list[int]:
    if value in (None, ""):
        return []
    if isinstance(value, (int, float)):
        return [int(value)]
    return [int(part) for part in str(value).split("|") if part]


def unwrap(value: Any) -> Any:
    if value == "Null":
        return None
    if isinstance(value, list):
        return [unwrap(item) for item in value]
    if isinstance(value, dict):
        if len(value) == 1:
            tag, inner = next(iter(value.items()))
            if tag in {"String", "Integer", "Bool", "Decimal"}:
                return inner
            if tag == "List":
                return [unwrap(item) for item in inner]
        return {key: unwrap(inner) for key, inner in value.items()}
    return value


def sora_rows(table: str) -> list[dict[str, Any]]:
    payload = json.loads((DEBUG / f"{table}.json").read_text(encoding="utf-8"))
    return [
        {key: unwrap(value) for key, value in row["values"].items()}
        for row in payload["table"]["rows"]
    ]


def normalize(table: str, row: dict[str, Any]) -> dict[str, Any]:
    result = dict(row)
    list_fields = {
        "UniverseCurio": {"tags", "pool_tags"},
        "UniverseMechanicRule": {"mechanic_tags"},
        "UniverseReviewFixture": {"input_stable_keys", "provenance_ids"},
    }.get(table, set())
    for field in list_fields:
        value = result.get(field)
        if isinstance(value, str):
            parts = [part for part in value.split("|") if part]
            result[field] = (
                [int(part) for part in parts]
                if field == "provenance_ids"
                else parts
            )
    return result


def digest(rows: list[dict[str, Any]], table: str = "") -> str:
    values = [normalize(table, row) for row in rows]
    values.sort(
        key=lambda row: json.dumps(
            row, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str
        )
    )
    encoded = json.dumps(
        values,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def assigned_partition(partition_id: str) -> dict[str, Any]:
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    match = next(
        (item for item in manifest["partitions"] if item["id"] == partition_id),
        None,
    )
    if match is None or not str(match["mechanic_family"]).startswith("curio-"):
        raise ValueError(f"{partition_id}: not a Curio partition")
    return match


def build_golden(partition_id: str) -> dict[str, Any]:
    assigned = assigned_partition(partition_id)
    universe = load_workbook(DATA / "Universe.xlsx", read_only=True, data_only=False)
    bindings = load_workbook(
        DATA / "UniverseBindings.xlsx", read_only=True, data_only=False
    )
    evidence = load_workbook(
        DATA / "UniverseEvidence.xlsx", read_only=True, data_only=False
    )
    record_ids = set(assigned["record_ids"])
    curio_keys = {key for key in record_ids if ".state." not in key}
    state_keys = {key for key in record_ids if ".state." in key}
    curios = keyed(sheet_rows(universe["UniverseCurio"]), "stable_key")
    states = keyed(sheet_rows(universe["UniverseCurioState"]), "stable_key")
    selected_curios = [curios[key] for key in sorted(curio_keys)]
    selected_states = [states[key] for key in sorted(state_keys)]
    curio_ids = {int(row["id"]) for row in selected_curios}
    all_curio_ids = {int(row["id"]) for row in curios.values()}
    state_ids = {int(row["id"]) for row in selected_states}
    if any(int(row["curio_id"]) not in all_curio_ids for row in selected_states):
        raise ValueError(f"{partition_id}: Curio state ownership differs")
    for row in selected_curios:
        initial = states.get(row["initial_state_stable_key"])
        if initial is None or int(initial["curio_id"]) != int(row["id"]):
            raise ValueError(f"{partition_id}: initial state links differ")
    parameters = [
        row
        for row in sheet_rows(universe["UniverseCurioParameter"])
        if int(row["curio_state_id"]) in state_ids
    ]
    sequence_by_state: dict[int, list[int]] = {}
    for row in parameters:
        sequence_by_state.setdefault(int(row["curio_state_id"]), []).append(
            int(row["sequence"])
        )
    if any(
        sorted(sequence) != list(range(1, len(sequence) + 1))
        for sequence in sequence_by_state.values()
    ):
        raise ValueError(f"{partition_id}: Curio parameters are not contiguous")

    rules = keyed(sheet_rows(bindings["UniverseMechanicRule"]), "stable_key")
    selected_rules = [rules[key] for key in assigned["rule_ids"]]
    for rule in selected_rules:
        if rule["source_record_stable_key"] not in record_ids or not rule["source_file"]:
            raise ValueError(f"{rule['stable_key']}: source/provenance link differs")

    fixtures = keyed(sheet_rows(evidence["UniverseReviewFixture"]), "stable_key")
    selected_fixtures = [fixtures[key] for key in assigned["fixture_ids"]]
    audits = keyed(sheet_rows(evidence["UniverseContentAudit"]), "content_stable_key")
    sources = keyed(sheet_rows(evidence["UniverseSourceRecord"]), "id")
    for stable_key in assigned["record_ids"]:
        audit = audits.get(stable_key)
        provenance = split_ids(audit["provenance_ids"]) if audit else []
        if (
            audit is None
            or not audit["enabled"]
            or not provenance
            or any(source not in sources for source in provenance)
        ):
            raise ValueError(f"{stable_key}: enabled provenance audit is incomplete")
    for fixture in selected_fixtures:
        provenance = split_ids(fixture["provenance_ids"])
        if not provenance or any(source not in sources for source in provenance):
            raise ValueError(f"{fixture['stable_key']}: provenance does not resolve")

    selected = {
        "UniverseCurio": selected_curios,
        "UniverseCurioState": selected_states,
        "UniverseCurioParameter": parameters,
        "UniverseMechanicRule": selected_rules,
        "UniverseReviewFixture": selected_fixtures,
    }
    for table, rows in selected.items():
        exported = sora_rows(table)
        if table == "UniverseCurioParameter":
            exported = [
                row
                for row in exported
                if int(row.get("curio_state_id", -1)) in state_ids
            ]
        else:
            keys = {row["stable_key"] for row in rows}
            exported = [row for row in exported if row.get("stable_key") in keys]
        if digest(rows, table) != digest(exported, table):
            raise ValueError(f"{partition_id}: Sora {table} rows differ from Excel")

    return {
        "schema_revision": "starclock.goal07-curio-partition-golden.v1",
        "partition_id": partition_id,
        "record_ids": assigned["record_ids"],
        "rule_ids": assigned["rule_ids"],
        "fixture_ids": assigned["fixture_ids"],
        "workbooks": {
            name: sha256(DATA / name)
            for name in (
                "Universe.xlsx",
                "UniverseBindings.xlsx",
                "UniverseEvidence.xlsx",
            )
        },
        "sora_bundle_sha256": sha256(BUNDLE),
        "tables": {
            table: {"rows": len(rows), "semantic_sha256": digest(rows, table)}
            for table, rows in selected.items()
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--partition", required=True)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--check", action="store_true")
    mode.add_argument("--write", action="store_true")
    args = parser.parse_args()
    golden = build_golden(args.partition)
    target = (
        ROOT
        / "evidence"
        / "standard-universe-mechanics-complete-v1"
        / "goldens"
        / f"{args.partition}.json"
    )
    encoded = json.dumps(golden, ensure_ascii=False, indent=2) + "\n"
    if args.write:
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(encoded, encoding="utf-8")
        print(f"Wrote {target.relative_to(ROOT)}.")
    else:
        if not target.is_file() or target.read_text(encoding="utf-8") != encoded:
            raise ValueError(f"{args.partition}: Curio partition golden drifted")
        print(f"Goal 07 Curio partition {args.partition} matches Excel and Sora.")


if __name__ == "__main__":
    main()
